# apps/reports/views.py
"""
Views للتقارير المركزية
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Q, Sum, Count, Avg, F, Max, Min
from django.utils.translation import gettext_lazy as _
from django.db.models.functions import TruncMonth, TruncYear, Coalesce
from datetime import date, timedelta, datetime
from decimal import Decimal
import io

# استيراد النماذج
from apps.core.models import (
    Item, ItemCategory, BusinessPartner, Warehouse, Branch,
    PriceList, UnitOfMeasure
)
from apps.accounting.models.account_models import Account
from apps.accounting.models.journal_models import JournalEntry, JournalEntryLine
from apps.accounting.models.voucher_models import PaymentVoucher, ReceiptVoucher
from apps.accounting.models.balance_models import AccountBalance
from apps.inventory.models import ItemStock, StockMovement, StockIn, StockOut

# استيراد النماذج
from .forms import (
    ItemsReportFilterForm, PartnersReportFilterForm, WarehousesReportFilterForm,
    ChartOfAccountsFilterForm, TrialBalanceFilterForm, GeneralLedgerFilterForm,
    JournalEntriesFilterForm, ReceiptsPaymentsFilterForm,
    IncomeStatementFilterForm, BalanceSheetFilterForm,
    StockReportFilterForm, StockMovementFilterForm, StockInOutFilterForm
)

# استيراد مكتبات التصدير
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===========================
# Helper Functions
# ===========================

def create_excel_workbook(title, headers, data, column_widths=None):
    """إنشاء workbook Excel مع تنسيق موحد"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel title limit

    # تنسيق العنوان
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # كتابة العناوين
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # كتابة البيانات
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="right")

    # تعيين عرض الأعمدة
    if column_widths:
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
    else:
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20

    return wb


def get_pdf_response(html_content, filename):
    """تحويل HTML إلى PDF"""
    try:
        from weasyprint import HTML
        pdf_file = HTML(string=html_content).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        messages.error(request, 'مكتبة WeasyPrint غير مثبتة')
        return redirect('reports:dashboard')


# ===========================
# Dashboard
# ===========================

@login_required
def reports_dashboard(request):
    """لوحة تحكم التقارير"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('core:dashboard')

    # تقارير Core Module
    core_reports = [
        {
            'title': 'تقرير الأصناف',
            'description': 'تقرير شامل بجميع الأصناف والمنتجات',
            'icon': 'fas fa-box',
            'color': 'primary',
            'url': reverse('reports:items_report'),
            'features': [
                'التصنيف حسب الفئات',
                'تفاصيل الأسعار والتكلفة',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'تقرير الشركاء',
            'description': 'تقرير العملاء والموردين',
            'icon': 'fas fa-handshake',
            'color': 'success',
            'url': reverse('reports:partners_report'),
            'features': [
                'العملاء والموردين',
                'بيانات الاتصال',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'تقرير المخازن',
            'description': 'تقرير شامل بجميع المخازن',
            'icon': 'fas fa-warehouse',
            'color': 'info',
            'url': reverse('reports:warehouses_report'),
            'features': [
                'المخازن حسب الفروع',
                'السعات والأرصدة',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'تقرير قوائم الأسعار',
            'description': 'تقرير قوائم الأسعار والخصومات',
            'icon': 'fas fa-tags',
            'color': 'warning',
            'url': reverse('reports:pricelists_report'),
            'features': [
                'الأسعار والخصومات',
                'الفترات الزمنية',
                'تصدير Excel/PDF',
            ]
        },
    ]

    # تقارير Accounting Module
    accounting_reports = [
        {
            'title': 'دليل الحسابات',
            'description': 'تقرير شجرة الحسابات',
            'icon': 'fas fa-sitemap',
            'color': 'primary',
            'url': reverse('reports:chart_of_accounts_report'),
            'features': [
                'شجرة الحسابات الكاملة',
                'الأرصدة الافتتاحية',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'ميزان المراجعة',
            'description': 'ميزان المراجعة لفترة محددة',
            'icon': 'fas fa-balance-scale',
            'color': 'success',
            'url': reverse('reports:trial_balance_report'),
            'features': [
                'الأرصدة المدينة والدائنة',
                'الفترة الزمنية',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'الأستاذ العام',
            'description': 'تقرير حركة الحسابات',
            'icon': 'fas fa-book',
            'color': 'info',
            'url': reverse('reports:general_ledger_report'),
            'features': [
                'حركة الحسابات التفصيلية',
                'الأرصدة الجارية',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'قيود اليومية',
            'description': 'تقرير قيود اليومية',
            'icon': 'fas fa-file-invoice',
            'color': 'warning',
            'url': reverse('reports:journal_entries_report'),
            'features': [
                'القيود اليدوية والتلقائية',
                'حالة الترحيل',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'المقبوضات والمدفوعات',
            'description': 'تقرير سندات القبض والصرف',
            'icon': 'fas fa-money-bill-wave',
            'color': 'danger',
            'url': reverse('reports:receipts_payments_report'),
            'features': [
                'سندات القبض والصرف',
                'طرق الدفع',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'قائمة الدخل',
            'description': 'قائمة الأرباح والخسائر',
            'icon': 'fas fa-chart-line',
            'color': 'primary',
            'url': reverse('reports:income_statement_report'),
            'features': [
                'الإيرادات والمصروفات',
                'صافي الربح/الخسارة',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'المركز المالي',
            'description': 'الميزانية العمومية',
            'icon': 'fas fa-chart-pie',
            'color': 'success',
            'url': reverse('reports:balance_sheet_report'),
            'features': [
                'الأصول والخصوم',
                'حقوق الملكية',
                'تصدير Excel/PDF',
            ]
        },
    ]

    # تقارير Inventory Module
    inventory_reports = [
        {
            'title': 'تقرير الجرد',
            'description': 'تقرير أرصدة المخزون',
            'icon': 'fas fa-clipboard-list',
            'color': 'primary',
            'url': reverse('reports:stock_report'),
            'features': [
                'الأرصدة الحالية',
                'حسب المخازن',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'حركة المخزون',
            'description': 'تقرير حركة المخزون التفصيلية',
            'icon': 'fas fa-exchange-alt',
            'color': 'info',
            'url': reverse('reports:stock_movement_report'),
            'features': [
                'جميع الحركات',
                'الإدخال والإخراج',
                'تصدير Excel/PDF',
            ]
        },
        {
            'title': 'الإدخالات والإخراجات',
            'description': 'تقرير عمليات الإدخال والإخراج',
            'icon': 'fas fa-arrows-alt-h',
            'color': 'success',
            'url': reverse('reports:stock_in_out_report'),
            'features': [
                'إدخالات المخزون',
                'إخراجات المخزون',
                'تصدير Excel/PDF',
            ]
        },
    ]

    context = {
        'title': _('التقارير'),
        'core_reports': core_reports,
        'accounting_reports': accounting_reports,
        'inventory_reports': inventory_reports,
        'breadcrumbs': [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('التقارير'), 'url': ''},
        ]
    }

    return render(request, 'reports/dashboard.html', context)


# ===========================
# Core Module Reports
# ===========================

@login_required
def items_report(request):
    """تقرير الأصناف"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = ItemsReportFilterForm(request.GET or None, company=request.current_company)

    items = Item.objects.filter(
        company=request.current_company
    ).select_related('category', 'unit', 'created_by')

    if form.is_valid():
        if form.cleaned_data.get('category'):
            items = items.filter(category=form.cleaned_data['category'])

        if form.cleaned_data.get('item_type'):
            items = items.filter(item_type=form.cleaned_data['item_type'])

        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            items = items.filter(is_active=is_active)

        if form.cleaned_data.get('date_from'):
            items = items.filter(created_at__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            items = items.filter(created_at__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    stats = {
        'total_count': items.count(),
        'active_count': items.filter(is_active=True).count(),
        'inactive_count': items.filter(is_active=False).count(),
    }

    # حسب الفئة
    by_category = items.values(
        'category__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')

    # حسب النوع
    by_type = items.values('item_type').annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'title': _('تقرير الأصناف'),
        'form': form,
        'items': items.order_by('code')[:500],
        'stats': stats,
        'by_category': by_category,
        'by_type': by_type,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('تقرير الأصناف'), 'url': ''},
        ]
    }

    return render(request, 'reports/core/items_report.html', context)


@login_required
def export_items_excel(request):
    """تصدير تقرير الأصناف إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # استخدام نفس الفلاتر
    form = ItemsReportFilterForm(request.GET or None, company=request.current_company)

    items = Item.objects.filter(
        company=request.current_company
    ).select_related('category', 'unit')

    if form.is_valid():
        if form.cleaned_data.get('category'):
            items = items.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('item_type'):
            items = items.filter(item_type=form.cleaned_data['item_type'])
        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            items = items.filter(is_active=is_active)

    # إعداد البيانات
    headers = ['الكود', 'الاسم', 'الفئة', 'النوع', 'الوحدة', 'التكلفة', 'السعر', 'الحالة']
    data = []

    for item in items:
        data.append([
            item.code,
            item.name,
            item.category.name if item.category else '-',
            item.get_item_type_display(),
            item.unit.name if item.unit else '-',
            float(item.cost_price) if item.cost_price else 0,
            float(item.selling_price) if item.selling_price else 0,
            'نشط' if item.is_active else 'غير نشط',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('تقرير الأصناف', headers, data,
                                column_widths=[15, 30, 20, 15, 15, 15, 15, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="items_report.xlsx"'

    return response


@login_required
def export_items_pdf(request):
    """تصدير تقرير الأصناف إلى PDF"""
    # سيتم تنفيذه لاحقاً
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:items_report')


@login_required
def partners_report(request):
    """تقرير الشركاء (العملاء والموردين)"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = PartnersReportFilterForm(request.GET or None)

    partners = BusinessPartner.objects.filter(
        company=request.current_company
    )

    if form.is_valid():
        if form.cleaned_data.get('partner_type'):
            partners = partners.filter(partner_type=form.cleaned_data['partner_type'])

        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            partners = partners.filter(is_active=is_active)

    # الإحصائيات
    stats = {
        'total_count': partners.count(),
        'customers_count': partners.filter(partner_type__in=['customer', 'both']).count(),
        'suppliers_count': partners.filter(partner_type__in=['supplier', 'both']).count(),
        'active_count': partners.filter(is_active=True).count(),
    }

    # حسب النوع
    by_type = partners.values('partner_type').annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'title': _('تقرير الشركاء'),
        'form': form,
        'partners': partners.order_by('name')[:500],
        'stats': stats,
        'by_type': by_type,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('تقرير الشركاء'), 'url': ''},
        ]
    }

    return render(request, 'reports/core/partners_report.html', context)


@login_required
def export_partners_excel(request):
    """تصدير تقرير الشركاء إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    form = PartnersReportFilterForm(request.GET or None)

    partners = BusinessPartner.objects.filter(
        company=request.current_company
    )

    if form.is_valid():
        if form.cleaned_data.get('partner_type'):
            partners = partners.filter(partner_type=form.cleaned_data['partner_type'])
        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            partners = partners.filter(is_active=is_active)

    # إعداد البيانات
    headers = ['الكود', 'الاسم', 'النوع', 'الهاتف', 'البريد الإلكتروني', 'العنوان', 'الحالة']
    data = []

    for partner in partners:
        data.append([
            partner.code,
            partner.name,
            partner.get_partner_type_display(),
            partner.phone or '-',
            partner.email or '-',
            partner.address or '-',
            'نشط' if partner.is_active else 'غير نشط',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('تقرير الشركاء', headers, data,
                                column_widths=[15, 25, 15, 15, 25, 30, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="partners_report.xlsx"'

    return response


@login_required
def export_partners_pdf(request):
    """تصدير تقرير الشركاء إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:partners_report')


@login_required
def warehouses_report(request):
    """تقرير المخازن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = WarehousesReportFilterForm(request.GET or None, company=request.current_company)

    warehouses = Warehouse.objects.filter(
        company=request.current_company
    ).select_related('branch')

    if form.is_valid():
        if form.cleaned_data.get('branch'):
            warehouses = warehouses.filter(branch=form.cleaned_data['branch'])

        if form.cleaned_data.get('warehouse_type'):
            warehouses = warehouses.filter(warehouse_type=form.cleaned_data['warehouse_type'])

        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            warehouses = warehouses.filter(is_active=is_active)

    # الإحصائيات
    stats = {
        'total_count': warehouses.count(),
        'active_count': warehouses.filter(is_active=True).count(),
        'inactive_count': warehouses.filter(is_active=False).count(),
    }

    # حسب الفرع
    by_branch = warehouses.values(
        'branch__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'title': _('تقرير المخازن'),
        'form': form,
        'warehouses': warehouses.order_by('branch', 'name'),
        'stats': stats,
        'by_branch': by_branch,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('تقرير المخازن'), 'url': ''},
        ]
    }

    return render(request, 'reports/core/warehouses_report.html', context)


@login_required
def export_warehouses_excel(request):
    """تصدير تقرير المخازن إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    form = WarehousesReportFilterForm(request.GET or None, company=request.current_company)

    warehouses = Warehouse.objects.filter(
        company=request.current_company
    ).select_related('branch')

    if form.is_valid():
        if form.cleaned_data.get('branch'):
            warehouses = warehouses.filter(branch=form.cleaned_data['branch'])
        if form.cleaned_data.get('warehouse_type'):
            warehouses = warehouses.filter(warehouse_type=form.cleaned_data['warehouse_type'])
        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            warehouses = warehouses.filter(is_active=is_active)

    # إعداد البيانات
    headers = ['الكود', 'الاسم', 'الفرع', 'النوع', 'العنوان', 'الحالة']
    data = []

    for warehouse in warehouses:
        data.append([
            warehouse.code,
            warehouse.name,
            warehouse.branch.name if warehouse.branch else '-',
            warehouse.get_warehouse_type_display() if hasattr(warehouse, 'warehouse_type') else '-',
            warehouse.address or '-',
            'نشط' if warehouse.is_active else 'غير نشط',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('تقرير المخازن', headers, data,
                                column_widths=[15, 25, 20, 15, 30, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="warehouses_report.xlsx"'

    return response


@login_required
def export_warehouses_pdf(request):
    """تصدير تقرير المخازن إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:warehouses_report')


@login_required
def pricelists_report(request):
    """تقرير قوائم الأسعار"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    pricelists = PriceList.objects.filter(
        company=request.current_company
    ).prefetch_related('items')

    # الإحصائيات
    stats = {
        'total_count': pricelists.count(),
        'active_count': pricelists.filter(is_active=True).count(),
        'with_dates_count': pricelists.exclude(valid_from__isnull=True).count(),
    }

    context = {
        'title': _('تقرير قوائم الأسعار'),
        'pricelists': pricelists.order_by('-created_at'),
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('تقرير قوائم الأسعار'), 'url': ''},
        ]
    }

    return render(request, 'reports/core/pricelists_report.html', context)


@login_required
def export_pricelists_excel(request):
    """تصدير تقرير قوائم الأسعار إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    pricelists = PriceList.objects.filter(
        company=request.current_company
    )

    # إعداد البيانات
    headers = ['الكود', 'الاسم', 'تاريخ البداية', 'تاريخ النهاية', 'عدد الأصناف', 'الحالة']
    data = []

    for pricelist in pricelists:
        data.append([
            pricelist.code,
            pricelist.name,
            str(pricelist.valid_from) if pricelist.valid_from else '-',
            str(pricelist.valid_to) if pricelist.valid_to else '-',
            pricelist.items.count() if hasattr(pricelist, 'items') else 0,
            'نشط' if pricelist.is_active else 'غير نشط',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('تقرير قوائم الأسعار', headers, data,
                                column_widths=[15, 25, 15, 15, 15, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pricelists_report.xlsx"'

    return response


@login_required
def export_pricelists_pdf(request):
    """تصدير تقرير قوائم الأسعار إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:pricelists_report')


# ===========================
# Accounting Module Reports  
# ===========================

@login_required
def chart_of_accounts_report(request):
    """تقرير دليل الحسابات"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = ChartOfAccountsFilterForm(request.GET or None)

    accounts = Account.objects.filter(
        company=request.current_company
    ).select_related('parent')

    if form.is_valid():
        if form.cleaned_data.get('account_type'):
            accounts = accounts.filter(account_type=form.cleaned_data['account_type'])

        if form.cleaned_data.get('level'):
            accounts = accounts.filter(level=form.cleaned_data['level'])

        if form.cleaned_data.get('is_active') != '':
            is_active = form.cleaned_data['is_active'] == '1'
            accounts = accounts.filter(is_active=is_active)

    # الإحصائيات
    stats = {
        'total_count': accounts.count(),
        'active_count': accounts.filter(is_active=True).count(),
        'by_type': accounts.values('account_type').annotate(count=Count('id')).order_by('-count'),
    }

    context = {
        'title': _('تقرير دليل الحسابات'),
        'form': form,
        'accounts': accounts.order_by('code'),
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('دليل الحسابات'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/chart_of_accounts.html', context)


@login_required
def export_coa_excel(request):
    """تصدير دليل الحسابات إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    accounts = Account.objects.filter(
        company=request.current_company
    ).select_related('parent').order_by('code')

    # إعداد البيانات
    headers = ['الكود', 'اسم الحساب', 'الحساب الأب', 'النوع', 'المستوى', 'الحالة']
    data = []

    for account in accounts:
        data.append([
            account.code,
            account.name,
            account.parent.name if account.parent else '-',
            account.get_account_type_display(),
            account.level if hasattr(account, 'level') else '-',
            'نشط' if account.is_active else 'غير نشط',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('دليل الحسابات', headers, data,
                                column_widths=[15, 30, 25, 15, 10, 12])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="chart_of_accounts.xlsx"'

    return response


@login_required
def export_coa_pdf(request):
    """تصدير دليل الحسابات إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:chart_of_accounts_report')


@login_required
def trial_balance_report(request):
    """تقرير ميزان المراجعة"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = TrialBalanceFilterForm(request.GET or None)

    # الحصول على الأرصدة
    balances = AccountBalance.objects.filter(
        company=request.current_company
    ).select_related('account')

    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            balances = balances.filter(date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            balances = balances.filter(date__lte=form.cleaned_data['date_to'])

        if form.cleaned_data.get('account_type'):
            balances = balances.filter(account__account_type=form.cleaned_data['account_type'])

        if not form.cleaned_data.get('show_zero_balances'):
            balances = balances.exclude(balance=0)

    # الإحصائيات
    totals = balances.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit'),
        total_balance=Sum('balance')
    )

    context = {
        'title': _('ميزان المراجعة'),
        'form': form,
        'balances': balances.order_by('account__code'),
        'totals': totals,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('ميزان المراجعة'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/trial_balance.html', context)


@login_required
def export_trial_balance_excel(request):
    """تصدير ميزان المراجعة إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    balances = AccountBalance.objects.filter(
        company=request.current_company
    ).select_related('account').order_by('account__code')

    # إعداد البيانات
    headers = ['كود الحساب', 'اسم الحساب', 'مدين', 'دائن', 'الرصيد']
    data = []

    for balance in balances:
        data.append([
            balance.account.code,
            balance.account.name,
            float(balance.debit or 0),
            float(balance.credit or 0),
            float(balance.balance or 0),
        ])

    # إنشاء الملف
    wb = create_excel_workbook('ميزان المراجعة', headers, data,
                                column_widths=[15, 30, 15, 15, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="trial_balance.xlsx"'

    return response


@login_required
def export_trial_balance_pdf(request):
    """تصدير ميزان المراجعة إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:trial_balance_report')


@login_required
def general_ledger_report(request):
    """تقرير الأستاذ العام"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = GeneralLedgerFilterForm(request.GET or None, company=request.current_company)

    entries = JournalEntryLine.objects.filter(
        journal_entry__company=request.current_company,
        journal_entry__status='posted'
    ).select_related('account', 'journal_entry')

    if form.is_valid():
        if form.cleaned_data.get('account'):
            entries = entries.filter(account=form.cleaned_data['account'])

        if form.cleaned_data.get('date_from'):
            entries = entries.filter(journal_entry__entry_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            entries = entries.filter(journal_entry__entry_date__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    totals = entries.aggregate(
        total_debit=Sum('debit'),
        total_credit=Sum('credit')
    )

    context = {
        'title': _('تقرير الأستاذ العام'),
        'form': form,
        'entries': entries.order_by('journal_entry__entry_date')[:500],
        'totals': totals,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('الأستاذ العام'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/general_ledger.html', context)


@login_required
def export_general_ledger_excel(request):
    """تصدير الأستاذ العام إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    entries = JournalEntryLine.objects.filter(
        journal_entry__company=request.current_company,
        journal_entry__status='posted'
    ).select_related('account', 'journal_entry').order_by('journal_entry__entry_date')

    # إعداد البيانات
    headers = ['التاريخ', 'رقم القيد', 'الحساب', 'البيان', 'مدين', 'دائن']
    data = []

    for entry in entries:
        data.append([
            str(entry.journal_entry.entry_date),
            entry.journal_entry.entry_number,
            entry.account.name,
            entry.description or '-',
            float(entry.debit or 0),
            float(entry.credit or 0),
        ])

    # إنشاء الملف
    wb = create_excel_workbook('الأستاذ العام', headers, data,
                                column_widths=[12, 15, 25, 30, 15, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="general_ledger.xlsx"'

    return response


@login_required
def export_general_ledger_pdf(request):
    """تصدير الأستاذ العام إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:general_ledger_report')


@login_required
def journal_entries_report(request):
    """تقرير قيود اليومية"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = JournalEntriesFilterForm(request.GET or None)

    entries = JournalEntry.objects.filter(
        company=request.current_company
    ).prefetch_related('lines')

    if form.is_valid():
        if form.cleaned_data.get('entry_type'):
            entries = entries.filter(entry_type=form.cleaned_data['entry_type'])

        if form.cleaned_data.get('status'):
            entries = entries.filter(status=form.cleaned_data['status'])

        if form.cleaned_data.get('date_from'):
            entries = entries.filter(entry_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            entries = entries.filter(entry_date__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    stats = {
        'total_count': entries.count(),
        'posted_count': entries.filter(status='posted').count(),
        'draft_count': entries.filter(status='draft').count(),
        'total_amount': entries.aggregate(total=Sum('total_debit'))['total'] or 0,
    }

    context = {
        'title': _('تقرير قيود اليومية'),
        'form': form,
        'entries': entries.order_by('-entry_date')[:500],
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('قيود اليومية'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/journal_entries.html', context)


@login_required
def export_journal_entries_excel(request):
    """تصدير قيود اليومية إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    entries = JournalEntry.objects.filter(
        company=request.current_company
    ).order_by('-entry_date')

    # إعداد البيانات
    headers = ['رقم القيد', 'التاريخ', 'النوع', 'البيان', 'المبلغ', 'الحالة']
    data = []

    for entry in entries:
        data.append([
            entry.entry_number,
            str(entry.entry_date),
            entry.get_entry_type_display() if hasattr(entry, 'entry_type') else '-',
            entry.description or '-',
            float(entry.total_debit or 0),
            entry.get_status_display(),
        ])

    # إنشاء الملف
    wb = create_excel_workbook('قيود اليومية', headers, data,
                                column_widths=[15, 12, 12, 35, 15, 12])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="journal_entries.xlsx"'

    return response


@login_required
def export_journal_entries_pdf(request):
    """تصدير قيود اليومية إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:journal_entries_report')


@login_required
def receipts_payments_report(request):
    """تقرير المقبوضات والمدفوعات"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = ReceiptsPaymentsFilterForm(request.GET or None)

    # الحصول على سندات القبض والصرف
    receipts = ReceiptVoucher.objects.filter(
        company=request.current_company
    )
    payments = PaymentVoucher.objects.filter(
        company=request.current_company
    )

    if form.is_valid():
        if form.cleaned_data.get('voucher_type') == 'receipt':
            payments = payments.none()
        elif form.cleaned_data.get('voucher_type') == 'payment':
            receipts = receipts.none()

        if form.cleaned_data.get('date_from'):
            receipts = receipts.filter(voucher_date__gte=form.cleaned_data['date_from'])
            payments = payments.filter(voucher_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            receipts = receipts.filter(voucher_date__lte=form.cleaned_data['date_to'])
            payments = payments.filter(voucher_date__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    stats = {
        'receipts_count': receipts.count(),
        'payments_count': payments.count(),
        'total_receipts': receipts.aggregate(total=Sum('amount'))['total'] or 0,
        'total_payments': payments.aggregate(total=Sum('amount'))['total'] or 0,
    }
    stats['net_cash_flow'] = stats['total_receipts'] - stats['total_payments']

    context = {
        'title': _('تقرير المقبوضات والمدفوعات'),
        'form': form,
        'receipts': receipts.order_by('-voucher_date')[:100],
        'payments': payments.order_by('-voucher_date')[:100],
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('المقبوضات والمدفوعات'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/receipts_payments.html', context)


@login_required
def export_receipts_payments_excel(request):
    """تصدير المقبوضات والمدفوعات إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    receipts = ReceiptVoucher.objects.filter(
        company=request.current_company
    ).order_by('-voucher_date')

    payments = PaymentVoucher.objects.filter(
        company=request.current_company
    ).order_by('-voucher_date')

    # إنشاء workbook مع ورقتين
    wb = Workbook()

    # ورقة سندات القبض
    ws1 = wb.active
    ws1.title = 'سندات القبض'

    headers = ['رقم السند', 'التاريخ', 'المبلغ', 'من', 'البيان']
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, receipt in enumerate(receipts, 2):
        ws1.cell(row=row_num, column=1).value = receipt.voucher_number
        ws1.cell(row=row_num, column=2).value = str(receipt.voucher_date)
        ws1.cell(row=row_num, column=3).value = float(receipt.amount)
        ws1.cell(row=row_num, column=4).value = receipt.partner.name if hasattr(receipt, 'partner') and receipt.partner else '-'
        ws1.cell(row=row_num, column=5).value = receipt.description or '-'

    # ورقة سندات الصرف
    ws2 = wb.create_sheet('سندات الصرف')

    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, payment in enumerate(payments, 2):
        ws2.cell(row=row_num, column=1).value = payment.voucher_number
        ws2.cell(row=row_num, column=2).value = str(payment.voucher_date)
        ws2.cell(row=row_num, column=3).value = float(payment.amount)
        ws2.cell(row=row_num, column=4).value = payment.partner.name if hasattr(payment, 'partner') and payment.partner else '-'
        ws2.cell(row=row_num, column=5).value = payment.description or '-'

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="receipts_payments.xlsx"'

    return response


@login_required
def export_receipts_payments_pdf(request):
    """تصدير المقبوضات والمدفوعات إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:receipts_payments_report')


@login_required
def income_statement_report(request):
    """تقرير قائمة الدخل (الأرباح والخسائر)"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = IncomeStatementFilterForm(request.GET or None)

    date_from = form.cleaned_data.get('date_from') if form.is_valid() else date(date.today().year, 1, 1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else date.today()

    # حساب الإيرادات
    revenue_accounts = Account.objects.filter(
        company=request.current_company,
        account_type='revenue'
    )

    total_revenue = AccountBalance.objects.filter(
        account__in=revenue_accounts,
        date__range=[date_from, date_to]
    ).aggregate(total=Sum('credit'))['total'] or Decimal('0')

    # حساب المصروفات
    expense_accounts = Account.objects.filter(
        company=request.current_company,
        account_type='expense'
    )

    total_expenses = AccountBalance.objects.filter(
        account__in=expense_accounts,
        date__range=[date_from, date_to]
    ).aggregate(total=Sum('debit'))['total'] or Decimal('0')

    # صافي الربح / الخسارة
    net_income = total_revenue - total_expenses

    context = {
        'title': _('قائمة الدخل'),
        'form': form,
        'revenue_accounts': revenue_accounts,
        'expense_accounts': expense_accounts,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': net_income,
        'date_from': date_from,
        'date_to': date_to,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('قائمة الدخل'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/income_statement.html', context)


@login_required
def export_income_statement_excel(request):
    """تصدير قائمة الدخل إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # إعداد البيانات
    headers = ['البيان', 'المبلغ']
    data = [
        ['الإيرادات', 0],  # سيتم حسابها ديناميكياً
        ['المصروفات', 0],  # سيتم حسابها ديناميكياً
        ['صافي الربح/الخسارة', 0],
    ]

    # إنشاء الملف
    wb = create_excel_workbook('قائمة الدخل', headers, data,
                                column_widths=[30, 20])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="income_statement.xlsx"'

    return response


@login_required
def export_income_statement_pdf(request):
    """تصدير قائمة الدخل إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:income_statement_report')


@login_required
def balance_sheet_report(request):
    """تقرير المركز المالي (الميزانية العمومية)"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = BalanceSheetFilterForm(request.GET or None)

    as_of_date = form.cleaned_data.get('as_of_date') if form.is_valid() else date.today()

    # حساب الأصول
    asset_accounts = Account.objects.filter(
        company=request.current_company,
        account_type='asset'
    )

    total_assets = AccountBalance.objects.filter(
        account__in=asset_accounts,
        date__lte=as_of_date
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0')

    # حساب الخصوم
    liability_accounts = Account.objects.filter(
        company=request.current_company,
        account_type='liability'
    )

    total_liabilities = AccountBalance.objects.filter(
        account__in=liability_accounts,
        date__lte=as_of_date
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0')

    # حساب حقوق الملكية
    equity_accounts = Account.objects.filter(
        company=request.current_company,
        account_type='equity'
    )

    total_equity = AccountBalance.objects.filter(
        account__in=equity_accounts,
        date__lte=as_of_date
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0')

    context = {
        'title': _('المركز المالي'),
        'form': form,
        'asset_accounts': asset_accounts,
        'liability_accounts': liability_accounts,
        'equity_accounts': equity_accounts,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'as_of_date': as_of_date,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('المركز المالي'), 'url': ''},
        ]
    }

    return render(request, 'reports/accounting/balance_sheet.html', context)


@login_required
def export_balance_sheet_excel(request):
    """تصدير المركز المالي إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # إعداد البيانات
    headers = ['البيان', 'المبلغ']
    data = [
        ['الأصول', 0],
        ['الخصوم', 0],
        ['حقوق الملكية', 0],
    ]

    # إنشاء الملف
    wb = create_excel_workbook('المركز المالي', headers, data,
                                column_widths=[30, 20])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="balance_sheet.xlsx"'

    return response


@login_required
def export_balance_sheet_pdf(request):
    """تصدير المركز المالي إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:balance_sheet_report')


# ===========================
# Inventory Module Reports
# ===========================

@login_required
def stock_report(request):
    """تقرير الجرد (أرصدة المخزون)"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = StockReportFilterForm(request.GET or None, company=request.current_company)

    stock = ItemStock.objects.filter(
        warehouse__company=request.current_company
    ).select_related('item', 'warehouse', 'item__category')

    if form.is_valid():
        if form.cleaned_data.get('warehouse'):
            stock = stock.filter(warehouse=form.cleaned_data['warehouse'])

        if form.cleaned_data.get('category'):
            stock = stock.filter(item__category=form.cleaned_data['category'])

        if not form.cleaned_data.get('show_zero_stock'):
            stock = stock.exclude(quantity=0)

    # الإحصائيات
    stats = {
        'total_items': stock.count(),
        'total_quantity': stock.aggregate(total=Sum('quantity'))['total'] or 0,
        'total_value': stock.aggregate(
            total=Sum(F('quantity') * F('item__cost_price'))
        )['total'] or Decimal('0'),
    }

    context = {
        'title': _('تقرير الجرد'),
        'form': form,
        'stock': stock.order_by('item__code')[:500],
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('تقرير الجرد'), 'url': ''},
        ]
    }

    return render(request, 'reports/inventory/stock_report.html', context)


@login_required
def export_stock_excel(request):
    """تصدير تقرير الجرد إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    stock = ItemStock.objects.filter(
        warehouse__company=request.current_company
    ).select_related('item', 'warehouse').order_by('item__code')

    # إعداد البيانات
    headers = ['كود الصنف', 'اسم الصنف', 'المخزن', 'الكمية', 'الوحدة', 'التكلفة', 'القيمة']
    data = []

    for s in stock:
        value = float(s.quantity * (s.item.cost_price or Decimal('0')))
        data.append([
            s.item.code,
            s.item.name,
            s.warehouse.name,
            float(s.quantity),
            s.item.unit.name if s.item.unit else '-',
            float(s.item.cost_price or 0),
            value,
        ])

    # إنشاء الملف
    wb = create_excel_workbook('تقرير الجرد', headers, data,
                                column_widths=[15, 25, 20, 12, 12, 12, 15])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_report.xlsx"'

    return response


@login_required
def export_stock_pdf(request):
    """تصدير تقرير الجرد إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:stock_report')


@login_required
def stock_movement_report(request):
    """تقرير حركة المخزون"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = StockMovementFilterForm(request.GET or None, company=request.current_company)

    movements = StockMovement.objects.filter(
        warehouse__company=request.current_company
    ).select_related('item', 'warehouse')

    if form.is_valid():
        if form.cleaned_data.get('warehouse'):
            movements = movements.filter(warehouse=form.cleaned_data['warehouse'])

        if form.cleaned_data.get('category'):
            movements = movements.filter(item__category=form.cleaned_data['category'])

        if form.cleaned_data.get('movement_type'):
            movements = movements.filter(movement_type=form.cleaned_data['movement_type'])

        if form.cleaned_data.get('date_from'):
            movements = movements.filter(movement_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            movements = movements.filter(movement_date__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    stats = {
        'total_movements': movements.count(),
        'total_in': movements.filter(movement_type='in').aggregate(total=Sum('quantity'))['total'] or 0,
        'total_out': movements.filter(movement_type='out').aggregate(total=Sum('quantity'))['total'] or 0,
    }

    context = {
        'title': _('تقرير حركة المخزون'),
        'form': form,
        'movements': movements.order_by('-movement_date')[:500],
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('حركة المخزون'), 'url': ''},
        ]
    }

    return render(request, 'reports/inventory/stock_movement.html', context)


@login_required
def export_stock_movement_excel(request):
    """تصدير تقرير حركة المخزون إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    movements = StockMovement.objects.filter(
        warehouse__company=request.current_company
    ).select_related('item', 'warehouse').order_by('-movement_date')

    # إعداد البيانات
    headers = ['التاريخ', 'الصنف', 'المخزن', 'نوع الحركة', 'الكمية', 'المرجع']
    data = []

    for movement in movements:
        data.append([
            str(movement.movement_date),
            movement.item.name,
            movement.warehouse.name,
            movement.get_movement_type_display(),
            float(movement.quantity),
            movement.reference or '-',
        ])

    # إنشاء الملف
    wb = create_excel_workbook('حركة المخزون', headers, data,
                                column_widths=[12, 25, 20, 15, 12, 20])

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_movement.xlsx"'

    return response


@login_required
def export_stock_movement_pdf(request):
    """تصدير تقرير حركة المخزون إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:stock_movement_report')


@login_required
def stock_in_out_report(request):
    """تقرير الإدخالات والإخراجات"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    # الفلاتر
    form = StockInOutFilterForm(request.GET or None, company=request.current_company)

    stock_ins = StockIn.objects.filter(
        warehouse__company=request.current_company
    ).select_related('warehouse')

    stock_outs = StockOut.objects.filter(
        warehouse__company=request.current_company
    ).select_related('warehouse')

    if form.is_valid():
        if form.cleaned_data.get('warehouse'):
            stock_ins = stock_ins.filter(warehouse=form.cleaned_data['warehouse'])
            stock_outs = stock_outs.filter(warehouse=form.cleaned_data['warehouse'])

        if form.cleaned_data.get('transaction_type') == 'stock_in':
            stock_outs = stock_outs.none()
        elif form.cleaned_data.get('transaction_type') == 'stock_out':
            stock_ins = stock_ins.none()

        if form.cleaned_data.get('date_from'):
            stock_ins = stock_ins.filter(transaction_date__gte=form.cleaned_data['date_from'])
            stock_outs = stock_outs.filter(transaction_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):
            stock_ins = stock_ins.filter(transaction_date__lte=form.cleaned_data['date_to'])
            stock_outs = stock_outs.filter(transaction_date__lte=form.cleaned_data['date_to'])

    # الإحصائيات
    stats = {
        'stock_ins_count': stock_ins.count(),
        'stock_outs_count': stock_outs.count(),
        'total_in_value': stock_ins.aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
        'total_out_value': stock_outs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
    }

    context = {
        'title': _('تقرير الإدخالات والإخراجات'),
        'form': form,
        'stock_ins': stock_ins.order_by('-transaction_date')[:100],
        'stock_outs': stock_outs.order_by('-transaction_date')[:100],
        'stats': stats,
        'today': date.today(),
        'breadcrumbs': [
            {'title': _('التقارير'), 'url': reverse('reports:dashboard')},
            {'title': _('الإدخالات والإخراجات'), 'url': ''},
        ]
    }

    return render(request, 'reports/inventory/stock_in_out.html', context)


@login_required
def export_stock_in_out_excel(request):
    """تصدير تقرير الإدخالات والإخراجات إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('reports:dashboard')

    stock_ins = StockIn.objects.filter(
        warehouse__company=request.current_company
    ).select_related('warehouse').order_by('-transaction_date')

    stock_outs = StockOut.objects.filter(
        warehouse__company=request.current_company
    ).select_related('warehouse').order_by('-transaction_date')

    # إنشاء workbook مع ورقتين
    wb = Workbook()

    # ورقة الإدخالات
    ws1 = wb.active
    ws1.title = 'الإدخالات'

    headers = ['رقم العملية', 'التاريخ', 'المخزن', 'المبلغ', 'البيان']
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, stock_in in enumerate(stock_ins, 2):
        ws1.cell(row=row_num, column=1).value = stock_in.transaction_number
        ws1.cell(row=row_num, column=2).value = str(stock_in.transaction_date)
        ws1.cell(row=row_num, column=3).value = stock_in.warehouse.name
        ws1.cell(row=row_num, column=4).value = float(stock_in.total_amount or 0)
        ws1.cell(row=row_num, column=5).value = stock_in.notes or '-'

    # ورقة الإخراجات
    ws2 = wb.create_sheet('الإخراجات')

    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, stock_out in enumerate(stock_outs, 2):
        ws2.cell(row=row_num, column=1).value = stock_out.transaction_number
        ws2.cell(row=row_num, column=2).value = str(stock_out.transaction_date)
        ws2.cell(row=row_num, column=3).value = stock_out.warehouse.name
        ws2.cell(row=row_num, column=4).value = float(stock_out.total_amount or 0)
        ws2.cell(row=row_num, column=5).value = stock_out.notes or '-'

    # الاستجابة
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_in_out.xlsx"'

    return response


@login_required
def export_stock_in_out_pdf(request):
    """تصدير تقرير الإدخالات والإخراجات إلى PDF"""
    messages.info(request, 'سيتم إضافة تصدير PDF قريباً')
    return redirect('reports:stock_in_out_report')

