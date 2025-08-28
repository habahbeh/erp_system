# apps/base_data/views/import_views.py
"""
Views استيراد البيانات - Excel, CSV
Bootstrap 5 + RTL + استيراد متقدم مع التحقق
"""

import io
import csv
import json
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import FormView
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse
from django.utils.translation import gettext_lazy as _
from django.db import transaction
from django.core.exceptions import ValidationError
from django.views import View

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from ..models import Item, BusinessPartner, Warehouse, ItemCategory, UnitOfMeasure
from ..forms import ItemImportForm, PartnerImportForm, WarehouseImportForm
from apps.core.mixins import CompanyMixin
from apps.core.utils import generate_code


class BaseImportView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, FormView):
    """Base class للاستيراد"""
    model = None
    template_name = 'base_data/import/base_import.html'
    success_url = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'model_name': self.model.__name__,
            'sample_file_url': self.get_sample_file_url(),
        })
        return context

    def get_sample_file_url(self):
        """رابط ملف النموذج"""
        return None

    def form_valid(self, form):
        try:
            result = self.process_file(form.cleaned_data['file'], form.cleaned_data)

            if result['success']:
                messages.success(
                    self.request,
                    _('تم استيراد %(created)d سجل بنجاح. تم تحديث %(updated)d سجل. فشل %(failed)d سجل.') % {
                        'created': result['created'],
                        'updated': result['updated'],
                        'failed': result['failed']
                    }
                )
                if result['errors']:
                    self.request.session['import_errors'] = result['errors']
            else:
                messages.error(self.request, result['message'])

        except Exception as e:
            messages.error(self.request, _('خطأ في الاستيراد: %(error)s') % {'error': str(e)})

        return super().form_valid(form)

    def process_file(self, file, options):
        """معالجة الملف"""
        file_extension = file.name.lower().split('.')[-1]

        try:
            if file_extension in ['xlsx', 'xls']:
                return self.process_excel_file(file, options)
            elif file_extension == 'csv':
                return self.process_csv_file(file, options)
            else:
                return {
                    'success': False,
                    'message': _('نوع الملف غير مدعوم')
                }
        except Exception as e:
            return {
                'success': False,
                'message': _('خطأ في قراءة الملف: %(error)s') % {'error': str(e)}
            }

    def process_excel_file(self, file, options):
        """معالجة ملف Excel"""
        if not OPENPYXL_AVAILABLE:
            return {
                'success': False,
                'message': _('مكتبة Excel غير متوفرة')
            }

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # قراءة الرأس
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # قراءة البيانات
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # تجاهل الصفوف الفارغة
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_data[headers[i]] = value
                data.append(row_data)

        return self.process_data(data, options)

    def process_csv_file(self, file, options):
        """معالجة ملف CSV"""
        file.seek(0)
        content = file.read()

        # التحقق من الترميز
        try:
            content = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content = content.decode('windows-1256')
            except UnicodeDecodeError:
                content = content.decode('latin-1')

        # قراءة CSV
        reader = csv.DictReader(io.StringIO(content))
        data = list(reader)

        return self.process_data(data, options)

    def process_data(self, data, options):
        """معالجة البيانات"""
        created = 0
        updated = 0
        failed = 0
        errors = []

        update_existing = options.get('update_existing', False)
        validate_data = options.get('validate_data', True)

        with transaction.atomic():
            for i, row_data in enumerate(data, start=2):  # البداية من السطر 2
                try:
                    result = self.process_row(row_data, update_existing, validate_data)
                    if result == 'created':
                        created += 1
                    elif result == 'updated':
                        updated += 1

                except Exception as e:
                    failed += 1
                    errors.append({
                        'row': i,
                        'data': row_data,
                        'error': str(e)
                    })

                    if not options.get('skip_errors', False):
                        raise

        return {
            'success': True,
            'created': created,
            'updated': updated,
            'failed': failed,
            'errors': errors
        }

    def process_row(self, row_data, update_existing, validate_data):
        """معالجة صف واحد"""
        raise NotImplementedError


class ItemImportView(BaseImportView):
    """استيراد الأصناف"""
    model = Item
    form_class = ItemImportForm
    template_name = 'base_data/import_export/import_form.html'
    permission_required = 'base_data.add_item'
    success_url = reverse_lazy('base_data:item_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': _('استيراد الأصناف'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('الأصناف'), 'url': reverse('base_data:item_list')},
                {'title': _('استيراد'), 'active': True}
            ],
            'field_mapping': self.get_field_mapping(),
        })
        return context

    def get_field_mapping(self):
        """خريطة الحقول للمساعدة"""
        return {
            'code': _('كود الصنف'),
            'name': _('اسم الصنف'),
            'name_en': _('الاسم الإنجليزي'),
            'category': _('التصنيف'),
            'unit': _('وحدة القياس'),
            'barcode': _('الباركود'),
            'purchase_price': _('سعر الشراء'),
            'sale_price': _('سعر البيع'),
            'tax_rate': _('معدل الضريبة'),
            'manufacturer': _('الشركة المصنعة'),
            'weight': _('الوزن'),
            'minimum_quantity': _('أقل كمية'),
            'maximum_quantity': _('أكبر كمية'),
            'specifications': _('المواصفات'),
        }

    def process_row(self, row_data, update_existing, validate_data):
        # تنظيف البيانات
        cleaned_data = self.clean_row_data(row_data)

        # البحث عن الصنف الموجود
        existing_item = None
        if cleaned_data.get('code'):
            try:
                existing_item = Item.objects.get(
                    code=cleaned_data['code'],
                    company=self.request.user.company
                )
            except Item.DoesNotExist:
                pass

        # التحقق من التحديث
        if existing_item and not update_existing:
            raise ValidationError(_('الصنف بالكود %(code)s موجود مسبقاً') % {
                'code': cleaned_data['code']
            })

        # إنشاء أو تحديث الصنف
        if existing_item:
            # تحديث
            for key, value in cleaned_data.items():
                if value is not None:
                    setattr(existing_item, key, value)

            existing_item.updated_by = self.request.user

            if validate_data:
                existing_item.full_clean()

            existing_item.save()
            return 'updated'

        else:
            # إنشاء جديد
            item = Item(company=self.request.user.company, created_by=self.request.user)

            for key, value in cleaned_data.items():
                setattr(item, key, value)

            if not item.code:
                item.code = generate_code('ITEM', self.request.user.company)

            if validate_data:
                item.full_clean()

            item.save()
            return 'created'

    def clean_row_data(self, row_data):
        """تنظيف بيانات الصف"""
        cleaned = {}

        # تنظيف الحقول النصية
        text_fields = ['code', 'name', 'name_en', 'barcode', 'manufacturer', 'specifications']
        for field in text_fields:
            value = self.get_field_value(row_data, field)
            if value:
                cleaned[field] = str(value).strip()

        # تنظيف الحقول الرقمية
        numeric_fields = ['purchase_price', 'sale_price', 'tax_rate', 'weight', 'minimum_quantity', 'maximum_quantity']
        for field in numeric_fields:
            value = self.get_field_value(row_data, field)
            if value is not None and value != '':
                try:
                    cleaned[field] = float(value)
                except (ValueError, TypeError):
                    pass

        # معالجة التصنيف
        category_name = self.get_field_value(row_data, 'category')
        if category_name:
            try:
                category = ItemCategory.objects.get(
                    name=category_name.strip(),
                    company=self.request.user.company
                )
                cleaned['category'] = category
            except ItemCategory.DoesNotExist:
                pass

        # معالجة وحدة القياس
        unit_name = self.get_field_value(row_data, 'unit')
        if unit_name:
            try:
                unit = UnitOfMeasure.objects.get(
                    name=unit_name.strip(),
                    company=self.request.user.company
                )
                cleaned['unit'] = unit
            except UnitOfMeasure.DoesNotExist:
                pass

        return cleaned

    def get_field_value(self, row_data, field_name):
        """الحصول على قيمة الحقل مع دعم الأسماء المختلفة"""
        # قائمة الأسماء المحتملة للحقل
        field_aliases = {
            'code': ['code', 'كود', 'رمز', 'Code'],
            'name': ['name', 'اسم', 'الاسم', 'Name'],
            'name_en': ['name_en', 'english_name', 'الاسم الإنجليزي', 'English Name'],
            'category': ['category', 'تصنيف', 'التصنيف', 'Category'],
            'unit': ['unit', 'وحدة', 'وحدة القياس', 'Unit'],
            'barcode': ['barcode', 'باركود', 'الباركود', 'Barcode'],
            'purchase_price': ['purchase_price', 'سعر الشراء', 'Purchase Price'],
            'sale_price': ['sale_price', 'سعر البيع', 'Sale Price'],
            'tax_rate': ['tax_rate', 'معدل الضريبة', 'Tax Rate'],
            'manufacturer': ['manufacturer', 'شركة صنع', 'الشركة المصنعة', 'Manufacturer'],
            'weight': ['weight', 'وزن', 'الوزن', 'Weight'],
            'minimum_quantity': ['minimum_quantity', 'أقل كمية', 'Min Quantity'],
            'maximum_quantity': ['maximum_quantity', 'أكبر كمية', 'Max Quantity'],
            'specifications': ['specifications', 'مواصفات', 'المواصفات', 'Specifications'],
        }

        aliases = field_aliases.get(field_name, [field_name])

        for alias in aliases:
            if alias in row_data:
                return row_data[alias]

        return None


class PartnerImportView(BaseImportView):
    """استيراد الشركاء التجاريين"""
    model = BusinessPartner
    form_class = PartnerImportForm
    template_name = 'base_data/import_export/import_form.html'
    permission_required = 'base_data.add_businesspartner'
    success_url = reverse_lazy('base_data:partner_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'page_title': _('استيراد الشركاء التجاريين'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('الشركاء التجاريون'), 'url': reverse('base_data:partner_list')},
                {'title': _('استيراد'), 'active': True}
            ],
        })
        return context

    def process_row(self, row_data, update_existing, validate_data):
        cleaned_data = self.clean_partner_data(row_data)

        # البحث عن الشريك الموجود
        existing_partner = None
        if cleaned_data.get('code'):
            try:
                existing_partner = BusinessPartner.objects.get(
                    code=cleaned_data['code'],
                    company=self.request.user.company
                )
            except BusinessPartner.DoesNotExist:
                pass

        if existing_partner and not update_existing:
            raise ValidationError(_('الشريك بالكود %(code)s موجود مسبقاً') % {
                'code': cleaned_data['code']
            })

        if existing_partner:
            # تحديث
            for key, value in cleaned_data.items():
                if value is not None:
                    setattr(existing_partner, key, value)

            existing_partner.updated_by = self.request.user

            if validate_data:
                existing_partner.full_clean()

            existing_partner.save()
            return 'updated'

        else:
            # إنشاء جديد
            partner = BusinessPartner(
                company=self.request.user.company,
                created_by=self.request.user
            )

            for key, value in cleaned_data.items():
                setattr(partner, key, value)

            if not partner.code:
                prefix = 'CUST' if partner.partner_type in ['customer', 'both'] else 'SUPP'
                partner.code = generate_code(prefix, self.request.user.company)

            if validate_data:
                partner.full_clean()

            partner.save()
            return 'created'

    def clean_partner_data(self, row_data):
        """تنظيف بيانات الشريك"""
        cleaned = {}

        # الحقول النصية
        text_fields = [
            'code', 'name', 'name_en', 'contact_person', 'phone', 'mobile',
            'fax', 'email', 'website', 'address', 'city', 'region',
            'tax_number', 'commercial_register', 'payment_terms', 'notes'
        ]

        for field in text_fields:
            value = self.get_partner_field_value(row_data, field)
            if value:
                cleaned[field] = str(value).strip()

        # الحقول الرقمية
        numeric_fields = ['credit_limit', 'credit_period', 'discount_percentage', 'rating']
        for field in numeric_fields:
            value = self.get_partner_field_value(row_data, field)
            if value is not None and value != '':
                try:
                    cleaned[field] = float(value)
                except (ValueError, TypeError):
                    pass

        # نوع الشريك
        partner_type = self.get_partner_field_value(row_data, 'partner_type')
        if partner_type:
            type_map = {
                'عميل': 'customer',
                'مورد': 'supplier',
                'كليهما': 'both',
                'customer': 'customer',
                'supplier': 'supplier',
                'both': 'both'
            }
            cleaned['partner_type'] = type_map.get(partner_type.strip(), 'customer')
        else:
            cleaned['partner_type'] = 'customer'

        # نوع الحساب
        account_type = self.get_partner_field_value(row_data, 'account_type')
        if account_type:
            account_map = {
                'شخصي': 'individual',
                'شركة': 'company',
                'individual': 'individual',
                'company': 'company'
            }
            cleaned['account_type'] = account_map.get(account_type.strip(), 'company')

        return cleaned

    def get_partner_field_value(self, row_data, field_name):
        """الحصول على قيمة حقل الشريك"""
        field_aliases = {
            'code': ['code', 'كود', 'Code'],
            'name': ['name', 'اسم', 'الاسم', 'Name'],
            'name_en': ['name_en', 'english_name', 'الاسم الإنجليزي'],
            'partner_type': ['partner_type', 'نوع الشريك', 'Type'],
            'account_type': ['account_type', 'نوع الحساب', 'Account Type'],
            'contact_person': ['contact_person', 'جهة الاتصال', 'Contact Person'],
            'phone': ['phone', 'هاتف', 'الهاتف', 'Phone'],
            'mobile': ['mobile', 'موبايل', 'المحمول', 'Mobile'],
            'email': ['email', 'بريد إلكتروني', 'Email'],
            'city': ['city', 'مدينة', 'المدينة', 'City'],
            'region': ['region', 'منطقة', 'المنطقة', 'Region'],
            'tax_number': ['tax_number', 'رقم ضريبي', 'Tax Number'],
            'credit_limit': ['credit_limit', 'حد ائتمان', 'Credit Limit'],
        }

        aliases = field_aliases.get(field_name, [field_name])

        for alias in aliases:
            if alias in row_data:
                return row_data[alias]

        return None


class ImportErrorsView(LoginRequiredMixin, CompanyMixin, View):
    """عرض أخطاء الاستيراد"""
    template_name = 'base_data/import_export/import_errors.html'

    def get(self, request):
        errors = request.session.get('import_errors', [])

        if 'clear_errors' in request.GET:
            if 'import_errors' in request.session:
                del request.session['import_errors']
            return redirect('base_data:item_list')

        context = {
            'page_title': _('أخطاء الاستيراد'),
            'errors': errors,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('أخطاء الاستيراد'), 'active': True}
            ],
        }

        return render(request, self.template_name, context)


class DownloadSampleView(LoginRequiredMixin, CompanyMixin, View):
    """تحميل ملفات النماذج"""

    def get(self, request, model_type):
        if model_type == 'item':
            return self.download_item_sample()
        elif model_type == 'partner':
            return self.download_partner_sample()
        elif model_type == 'warehouse':
            return self.download_warehouse_sample()
        else:
            return JsonResponse({'error': _('نوع غير مدعوم')}, status=400)

    def download_item_sample(self):
        """تحميل نموذج الأصناف"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('نموذج الأصناف')

        headers = [
            _('كود الصنف'), _('اسم الصنف'), _('الاسم الإنجليزي'),
            _('التصنيف'), _('وحدة القياس'), _('الباركود'),
            _('سعر الشراء'), _('سعر البيع'), _('معدل الضريبة'),
            _('الشركة المصنعة'), _('الوزن'), _('أقل كمية'), _('أكبر كمية')
        ]

        # كتابة الرأس
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header

        # إضافة صف مثال
        sample_data = [
            'ITEM001', 'منتج تجريبي', 'Sample Product',
            'إلكترونيات', 'قطعة', '1234567890123',
            '100', '150', '15',
            'شركة تجريبية', '0.5', '10', '1000'
        ]

        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col).value = value

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sample_items.xlsx"'
        return response

    def download_partner_sample(self):
        """تحميل نموذج الشركاء"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('نموذج الشركاء')

        headers = [
            _('كود الشريك'), _('اسم الشريك'), _('نوع الشريك'),
            _('جهة الاتصال'), _('الهاتف'), _('الموبايل'), _('البريد الإلكتروني'),
            _('المدينة'), _('المنطقة'), _('الرقم الضريبي'), _('حد الائتمان')
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header

        # إضافة صف مثال
        sample_data = [
            'CUST001', 'شركة تجريبية', 'عميل',
            'أحمد محمد', '0112345678', '0501234567', 'test@example.com',
            'الرياض', 'الوسط', '123456789012345', '10000'
        ]

        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col).value = value

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sample_partners.xlsx"'
        return response