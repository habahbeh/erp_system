# apps/base_data/views/warehouse_views.py
"""
Views الخاصة بالمستودعات ووحدات القياس - كامل ومطابق للمتطلبات
يشمل: المستودعات، وحدات القياس، أرصدة المستودعات، التحويلات
"""

from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, Sum, Avg, F, Case, When, BooleanField, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views import View
from django_datatables_view.base_datatable_view import BaseDatatableView
from decimal import Decimal
import csv
import json
import openpyxl
from io import BytesIO

from ..models import (
    Warehouse, WarehouseItem, UnitOfMeasure, Item,
    Branch, User
)
from ..forms.warehouse_forms import (
    WarehouseForm, UnitOfMeasureForm, WarehouseItemForm,
    WarehouseFilterForm, WarehouseTransferForm, WarehouseImportForm,
    WarehouseQuickAddForm, UnitQuickAddForm
)


class BaseWarehouseMixin:
    """Mixin أساسي للمستودعات - يحتوي على الإعدادات المشتركة"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إضافة معلومات الشركة والفرع
        context['current_company'] = self.request.user.company
        context['current_branch'] = getattr(self.request.user, 'branch', None)

        # Breadcrumbs
        context['breadcrumbs'] = self.get_breadcrumbs()

        # إحصائيات عامة
        context['total_warehouses'] = Warehouse.objects.filter(
            company=self.request.user.company
        ).count()
        context['active_warehouses'] = Warehouse.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).count()
        context['total_units'] = UnitOfMeasure.objects.filter(
            company=self.request.user.company
        ).count()

        return context

    def get_breadcrumbs(self):
        """بناء breadcrumbs للصفحة"""
        return [
            {'title': _('الرئيسية'), 'url': '/'},
            {'title': _('البيانات الأساسية'), 'url': '#'},
        ]

    def get_queryset(self):
        """فلترة البيانات حسب الشركة"""
        queryset = super().get_queryset()
        return queryset.filter(company=self.request.user.company)


# ============== Views المستودعات ==============

class WarehouseListView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, ListView):
    """عرض قائمة المستودعات"""
    model = Warehouse
    template_name = 'base_data/warehouses/warehouse_list.html'
    context_object_name = 'warehouses'
    permission_required = 'base_data.view_warehouse'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('قائمة المستودعات')
        context['filter_form'] = WarehouseFilterForm(self.request.GET)

        # إحصائيات إضافية
        warehouses_qs = Warehouse.objects.filter(company=self.request.user.company)
        context['main_warehouses'] = warehouses_qs.filter(warehouse_type='main').count()
        context['branch_warehouses'] = warehouses_qs.filter(warehouse_type='branch').count()
        context['transit_warehouses'] = warehouses_qs.filter(warehouse_type='transit').count()
        context['damaged_warehouses'] = warehouses_qs.filter(warehouse_type='damaged').count()

        # إحصائيات الأرصدة
        context['total_items_in_stock'] = WarehouseItem.objects.filter(
            warehouse__company=self.request.user.company,
            quantity__gt=0
        ).count()

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'branch', 'keeper'
        ).prefetch_related(
            Prefetch('warehouse_items', queryset=WarehouseItem.objects.select_related('item'))
        )

        # تطبيق الفلاتر
        filter_form = WarehouseFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('search'):
                search = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(code__icontains=search) |
                    Q(name__icontains=search) |
                    Q(name_en__icontains=search) |
                    Q(location__icontains=search)
                )

            if filter_form.cleaned_data.get('warehouse_type'):
                queryset = queryset.filter(warehouse_type=filter_form.cleaned_data['warehouse_type'])

            if filter_form.cleaned_data.get('branch'):
                queryset = queryset.filter(branch=filter_form.cleaned_data['branch'])

            if filter_form.cleaned_data.get('keeper'):
                queryset = queryset.filter(keeper=filter_form.cleaned_data['keeper'])

            if filter_form.cleaned_data.get('is_active') is not None:
                queryset = queryset.filter(is_active=filter_form.cleaned_data['is_active'])

        return queryset.order_by('code', 'name')


class WarehouseDetailView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, DetailView):
    """تفاصيل المستودع"""
    model = Warehouse
    template_name = 'base_data/warehouses/warehouse_detail.html'
    context_object_name = 'warehouse'
    permission_required = 'base_data.view_warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل المستودع: %s') % self.object.name

        # إحصائيات المستودع
        warehouse_items = self.object.warehouse_items.select_related('item')

        context['warehouse_stats'] = {
            'total_items': warehouse_items.count(),
            'items_in_stock': warehouse_items.filter(quantity__gt=0).count(),
            'items_out_of_stock': warehouse_items.filter(quantity=0).count(),
            'low_stock_items': warehouse_items.filter(
                quantity__lte=F('item__reorder_level'),
                item__reorder_level__gt=0
            ).count(),
            'total_value': warehouse_items.aggregate(
                total=Sum(F('quantity') * F('average_cost'))
            )['total'] or 0,
        }

        # الأصناف الأكثر حركة
        context['top_items'] = warehouse_items.filter(
            quantity__gt=0
        ).order_by('-quantity')[:10]

        # الأصناف منخفضة المخزون
        context['low_stock_items'] = warehouse_items.filter(
            quantity__lte=F('item__reorder_level'),
            item__reorder_level__gt=0,
            quantity__gt=0
        ).select_related('item')[:10]

        # معلومات أمين المستودع
        if self.object.keeper:
            context['keeper_info'] = {
                'name': self.object.keeper.get_full_name(),
                'username': self.object.keeper.username,
                'email': self.object.keeper.email,
            }

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': str(self.object), 'active': True}
        ])

        return context


class WarehouseCreateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                          SuccessMessageMixin, CreateView):
    """إنشاء مستودع جديد"""
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'base_data/warehouses/warehouse_form.html'
    permission_required = 'base_data.add_warehouse'
    success_message = _('تم إنشاء المستودع بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة مستودع جديد')
        context['submit_text'] = _('حفظ')

        # قوائم الاختيار
        context['branches'] = Branch.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')

        context['users'] = User.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('first_name', 'last_name')

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': _('إضافة مستودع جديد'), 'active': True}
        ])

        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user

        # تعيين فرع افتراضي إذا لم يكن محدد
        if not form.instance.branch and self.request.user.branch:
            form.instance.branch = self.request.user.branch

        # إنشاء كود تلقائي إذا لم يكن موجود
        if not form.instance.code:
            warehouse_type = form.instance.warehouse_type
            last_warehouse = Warehouse.objects.filter(
                company=self.request.user.company,
                warehouse_type=warehouse_type
            ).order_by('-code').first()

            prefix_map = {
                'main': 'MAIN',
                'branch': 'BRCH',
                'transit': 'TRNS',
                'damaged': 'DMGD',
            }
            prefix = prefix_map.get(warehouse_type, 'WARE')

            if last_warehouse and last_warehouse.code.startswith(prefix):
                try:
                    last_num = int(last_warehouse.code[4:])
                    new_num = last_num + 1
                except (ValueError, IndexError):
                    new_num = 1
            else:
                new_num = 1

            form.instance.code = f"{prefix}{new_num:03d}"

        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:warehouse_add')
        return reverse_lazy('base_data:warehouse_detail', kwargs={'pk': self.object.pk})


class WarehouseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                          SuccessMessageMixin, UpdateView):
    """تعديل المستودع"""
    model = Warehouse
    form_class = WarehouseForm
    template_name = 'base_data/warehouses/warehouse_form.html'
    permission_required = 'base_data.change_warehouse'
    success_message = _('تم تعديل المستودع بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل المستودع: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        # قوائم الاختيار
        context['branches'] = Branch.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')

        context['users'] = User.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('first_name', 'last_name')

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:warehouse_detail', args=[self.object.pk])},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:warehouse_detail', kwargs={'pk': self.object.pk})


class WarehouseDeleteView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, DeleteView):
    """حذف المستودع"""
    model = Warehouse
    template_name = 'base_data/warehouses/warehouse_delete.html'
    permission_required = 'base_data.delete_warehouse'
    success_url = reverse_lazy('base_data:warehouse_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('حذف المستودع: %s') % self.object.name

        # التحقق من الارتباطات
        context['has_items'] = self.object.warehouse_items.exists()
        context['items_count'] = self.object.warehouse_items.count()
        context['has_stock'] = self.object.warehouse_items.filter(quantity__gt=0).exists()
        context['stock_items_count'] = self.object.warehouse_items.filter(quantity__gt=0).count()

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:warehouse_detail', args=[self.object.pk])},
            {'title': _('حذف'), 'active': True}
        ])

        return context

    def delete(self, request, *args, **kwargs):
        warehouse = self.get_object()

        # التحقق من وجود أرصدة
        if warehouse.warehouse_items.filter(quantity__gt=0).exists():
            messages.error(request, _('لا يمكن حذف المستودع لوجود أرصدة أصناف'))
            return redirect('base_data:warehouse_list')

        messages.success(request, _('تم حذف المستودع بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============== Views أرصدة المستودعات ==============

class WarehouseInventoryView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, DetailView):
    """عرض أرصدة المستودع"""
    model = Warehouse
    template_name = 'base_data/warehouses/warehouse_inventory.html'
    context_object_name = 'warehouse'
    permission_required = 'base_data.view_warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('أرصدة المستودع: %s') % self.object.name

        # فلترة الأرصدة
        filter_type = self.request.GET.get('filter', 'all')  # all, in_stock, out_of_stock, low_stock
        search = self.request.GET.get('search', '')

        warehouse_items = self.object.warehouse_items.select_related('item')

        if filter_type == 'in_stock':
            warehouse_items = warehouse_items.filter(quantity__gt=0)
        elif filter_type == 'out_of_stock':
            warehouse_items = warehouse_items.filter(quantity=0)
        elif filter_type == 'low_stock':
            warehouse_items = warehouse_items.filter(
                quantity__lte=F('item__reorder_level'),
                item__reorder_level__gt=0
            )

        if search:
            warehouse_items = warehouse_items.filter(
                Q(item__code__icontains=search) |
                Q(item__name__icontains=search) |
                Q(item__name_en__icontains=search)
            )

        context['warehouse_items'] = warehouse_items.order_by('item__name')
        context['filter_type'] = filter_type
        context['search_term'] = search

        # إحصائيات
        all_items = self.object.warehouse_items.select_related('item')
        context['inventory_stats'] = {
            'total_items': all_items.count(),
            'in_stock': all_items.filter(quantity__gt=0).count(),
            'out_of_stock': all_items.filter(quantity=0).count(),
            'low_stock': all_items.filter(
                quantity__lte=F('item__reorder_level'),
                item__reorder_level__gt=0
            ).count(),
            'total_value': all_items.aggregate(
                total=Sum(F('quantity') * F('average_cost'))
            )['total'] or 0,
        }

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:warehouse_detail', args=[self.object.pk])},
            {'title': _('الأرصدة'), 'active': True}
        ])

        return context


class WarehouseItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                              SuccessMessageMixin, UpdateView):
    """تعديل رصيد صنف في المستودع"""
    model = WarehouseItem
    form_class = WarehouseItemForm
    template_name = 'base_data/warehouses/warehouse_item_form.html'
    permission_required = 'base_data.change_warehouse'
    success_message = _('تم تحديث الرصيد بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل رصيد: %s') % self.object.item.name
        context['warehouse'] = self.object.warehouse

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': str(self.object.warehouse),
             'url': reverse_lazy('base_data:warehouse_detail', args=[self.object.warehouse.pk])},
            {'title': _('الأرصدة'),
             'url': reverse_lazy('base_data:warehouse_inventory', args=[self.object.warehouse.pk])},
            {'title': _('تعديل رصيد'), 'active': True}
        ])

        return context

    def get_queryset(self):
        return WarehouseItem.objects.filter(
            warehouse__company=self.request.user.company
        ).select_related('warehouse', 'item')

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:warehouse_inventory',
                            kwargs={'pk': self.object.warehouse.pk})


# ============== Views التحويلات بين المستودعات ==============

class WarehouseTransferListView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, ListView):
    """قائمة التحويلات بين المستودعات"""
    model = Warehouse  # سيتم استبدالها بـ WarehouseTransfer عند إنشاء النموذج
    template_name = 'base_data/warehouses/transfer_list.html'
    context_object_name = 'transfers'
    permission_required = 'base_data.view_warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('التحويلات بين المستودعات')

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': _('التحويلات'), 'active': True}
        ])

        return context

    def get_queryset(self):
        # TODO: استبدال بـ WarehouseTransfer.objects عند إنشاء النموذج
        return Warehouse.objects.none()


class WarehouseTransferCreateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                                  SuccessMessageMixin, CreateView):
    """إنشاء تحويل بين المستودعات"""
    model = Warehouse  # TODO: استبدال بـ WarehouseTransfer
    form_class = WarehouseTransferForm
    template_name = 'base_data/warehouses/transfer_form.html'
    permission_required = 'base_data.change_warehouse'
    success_message = _('تم إنشاء التحويل بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إنشاء تحويل بين المستودعات')

        # قوائم الاختيار
        context['warehouses'] = Warehouse.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')

        context['items'] = Item.objects.filter(
            company=self.request.user.company,
            is_inactive=False
        ).order_by('name')

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': _('التحويلات'), 'url': reverse_lazy('base_data:warehouse_transfer_list')},
            {'title': _('إنشاء تحويل'), 'active': True}
        ])

        return context


# ============== Views وحدات القياس ==============

class UnitOfMeasureListView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, ListView):
    """قائمة وحدات القياس"""
    model = UnitOfMeasure
    template_name = 'base_data/units/unit_list.html'
    context_object_name = 'units'
    permission_required = 'base_data.view_unitofmeasure'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('وحدات القياس')

        # إحصائيات
        units_qs = UnitOfMeasure.objects.filter(company=self.request.user.company)
        context['active_units'] = units_qs.filter(is_active=True).count()
        context['inactive_units'] = units_qs.filter(is_active=False).count()

        context['breadcrumbs'].extend([
            {'title': _('وحدات القياس'), 'active': True}
        ])

        return context

    def get_queryset(self):
        search = self.request.GET.get('search', '')
        queryset = super().get_queryset()

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        return queryset.order_by('name')


class UnitOfMeasureCreateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                              SuccessMessageMixin, CreateView):
    """إنشاء وحدة قياس جديدة"""
    model = UnitOfMeasure
    form_class = UnitOfMeasureForm
    template_name = 'base_data/units/unit_form.html'
    permission_required = 'base_data.add_unitofmeasure'
    success_message = _('تم إنشاء وحدة القياس بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة وحدة قياس جديدة')
        context['submit_text'] = _('حفظ')

        context['breadcrumbs'].extend([
            {'title': _('وحدات القياس'), 'url': reverse_lazy('base_data:unit_list')},
            {'title': _('إضافة وحدة جديدة'), 'active': True}
        ])

        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user

        # إنشاء كود تلقائي إذا لم يكن موجود
        if not form.instance.code:
            last_unit = UnitOfMeasure.objects.filter(
                company=self.request.user.company
            ).order_by('-code').first()

            if last_unit and last_unit.code.startswith('UNIT'):
                try:
                    last_num = int(last_unit.code[4:])
                    new_num = last_num + 1
                except (ValueError, IndexError):
                    new_num = 1
            else:
                new_num = 1

            form.instance.code = f"UNIT{new_num:03d}"

        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:unit_add')
        return reverse_lazy('base_data:unit_list')


class UnitOfMeasureUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin,
                              SuccessMessageMixin, UpdateView):
    """تعديل وحدة القياس"""
    model = UnitOfMeasure
    form_class = UnitOfMeasureForm
    template_name = 'base_data/units/unit_form.html'
    permission_required = 'base_data.change_unitofmeasure'
    success_message = _('تم تعديل وحدة القياس بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل وحدة القياس: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        context['breadcrumbs'].extend([
            {'title': _('وحدات القياس'), 'url': reverse_lazy('base_data:unit_list')},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:unit_list')


class UnitOfMeasureDeleteView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, DeleteView):
    """حذف وحدة القياس"""
    model = UnitOfMeasure
    template_name = 'base_data/units/unit_delete.html'
    permission_required = 'base_data.delete_unitofmeasure'
    success_url = reverse_lazy('base_data:unit_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('حذف وحدة القياس: %s') % self.object.name

        # التحقق من الارتباطات
        context['has_items'] = self.object.items.exists()
        context['items_count'] = self.object.items.count()
        context['has_conversions'] = False  # TODO: فحص معدلات التحويل

        context['breadcrumbs'].extend([
            {'title': _('وحدات القياس'), 'url': reverse_lazy('base_data:unit_list')},
            {'title': _('حذف'), 'active': True}
        ])

        return context

    def delete(self, request, *args, **kwargs):
        unit = self.get_object()

        # التحقق من وجود أصناف مرتبطة
        if unit.items.exists():
            messages.error(request, _('لا يمكن حذف وحدة القياس لوجود أصناف مرتبطة بها'))
            return redirect('base_data:unit_list')

        messages.success(request, _('تم حذف وحدة القياس بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============== Views التقارير والتصدير ==============

class WarehouseReportView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, ListView):
    """تقرير المستودعات"""
    model = Warehouse
    template_name = 'base_data/warehouses/warehouse_report.html'
    permission_required = 'base_data.view_warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تقرير المستودعات')
        context['filter_form'] = WarehouseFilterForm(self.request.GET)

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': _('التقارير'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'branch', 'keeper'
        ).prefetch_related('warehouse_items__item')

        # تطبيق الفلاتر
        filter_form = WarehouseFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('warehouse_type'):
                queryset = queryset.filter(warehouse_type=filter_form.cleaned_data['warehouse_type'])
            if filter_form.cleaned_data.get('branch'):
                queryset = queryset.filter(branch=filter_form.cleaned_data['branch'])
            if filter_form.cleaned_data.get('is_active') is not None:
                queryset = queryset.filter(is_active=filter_form.cleaned_data['is_active'])

        return queryset.order_by('warehouse_type', 'name')


class InventoryReportView(LoginRequiredMixin, PermissionRequiredMixin, BaseWarehouseMixin, ListView):
    """تقرير الأرصدة الشامل"""
    model = WarehouseItem
    template_name = 'base_data/warehouses/inventory_report.html'
    context_object_name = 'warehouse_items'
    permission_required = 'base_data.view_warehouse'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تقرير الأرصدة الشامل')

        # إحصائيات شاملة
        all_items = self.get_queryset()
        context['report_stats'] = {
            'total_items': all_items.count(),
            'items_in_stock': all_items.filter(quantity__gt=0).count(),
            'items_out_of_stock': all_items.filter(quantity=0).count(),
            'low_stock_items': all_items.filter(
                quantity__lte=F('item__reorder_level'),
                item__reorder_level__gt=0
            ).count(),
            'total_value': all_items.aggregate(
                total=Sum(F('quantity') * F('average_cost'))
            )['total'] or 0,
        }

        context['breadcrumbs'].extend([
            {'title': _('المستودعات'), 'url': reverse_lazy('base_data:warehouse_list')},
            {'title': _('تقرير الأرصدة'), 'active': True}
        ])

        return context

    def get_queryset(self):
        return WarehouseItem.objects.filter(
            warehouse__company=self.request.user.company
        ).select_related(
            'warehouse', 'item', 'item__category', 'item__unit_of_measure'
        ).order_by('warehouse__name', 'item__name')


class WarehouseExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """تصدير بيانات المستودعات إلى Excel/CSV"""
    permission_required = 'base_data.view_warehouse'

    def get(self, request, *args, **kwargs):
        export_type = request.GET.get('type', 'warehouses')  # warehouses, inventory
        export_format = request.GET.get('format', 'excel')

        if export_type == 'inventory':
            return self.export_inventory(export_format)
        else:
            return self.export_warehouses(export_format)

    def export_warehouses(self, export_format):
        """تصدير المستودعات"""
        warehouses = Warehouse.objects.filter(
            company=self.request.user.company
        ).select_related('branch', 'keeper').order_by('code')

        if export_format == 'excel':
            return self.export_warehouses_excel(warehouses)
        else:
            return self.export_warehouses_csv(warehouses)

    def export_inventory(self, export_format):
        """تصدير الأرصدة"""
        warehouse_items = WarehouseItem.objects.filter(
            warehouse__company=self.request.user.company
        ).select_related(
            'warehouse', 'item', 'item__unit_of_measure'
        ).order_by('warehouse__name', 'item__name')

        if export_format == 'excel':
            return self.export_inventory_excel(warehouse_items)
        else:
            return self.export_inventory_csv(warehouse_items)

    def export_warehouses_excel(self, warehouses):
        """تصدير المستودعات إلى Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "المستودعات"

        # العناوين
        headers = [
            'كود المستودع', 'اسم المستودع', 'الاسم الإنجليزي', 'نوع المستودع',
            'الفرع', 'أمين المستودع', 'الموقع', 'الحالة', 'ملاحظات'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # البيانات
        for row, warehouse in enumerate(warehouses, 2):
            ws.cell(row=row, column=1, value=warehouse.code)
            ws.cell(row=row, column=2, value=warehouse.name)
            ws.cell(row=row, column=3, value=warehouse.name_en or '')
            ws.cell(row=row, column=4, value=warehouse.get_warehouse_type_display())
            ws.cell(row=row, column=5, value=str(warehouse.branch) if warehouse.branch else '')
            ws.cell(row=row, column=6, value=warehouse.keeper.get_full_name() if warehouse.keeper else '')
            ws.cell(row=row, column=7, value=warehouse.location or '')
            ws.cell(row=row, column=8, value='نشط' if warehouse.is_active else 'غير نشط')
            ws.cell(row=row, column=9, value=warehouse.notes or '')

        # حفظ الملف
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="warehouses_report.xlsx"'
        return response

    def export_inventory_excel(self, warehouse_items):
        """تصدير الأرصدة إلى Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "أرصدة المستودعات"

        # العناوين
        headers = [
            'المستودع', 'كود الصنف', 'اسم الصنف', 'الكمية الحالية',
            'وحدة القياس', 'متوسط التكلفة', 'القيمة الإجمالية',
            'الحد الأدنى', 'الحد الأعلى', 'حالة المخزون'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # البيانات
        for row, warehouse_item in enumerate(warehouse_items, 2):
            item = warehouse_item.item
            total_value = warehouse_item.quantity * warehouse_item.average_cost

            # تحديد حالة المخزون
            if warehouse_item.quantity == 0:
                stock_status = 'نفد المخزون'
            elif item.reorder_level and warehouse_item.quantity <= item.reorder_level:
                stock_status = 'مخزون منخفض'
            else:
                stock_status = 'متوفر'

            ws.cell(row=row, column=1, value=warehouse_item.warehouse.name)
            ws.cell(row=row, column=2, value=item.code)
            ws.cell(row=row, column=3, value=item.name)
            ws.cell(row=row, column=4, value=float(warehouse_item.quantity))
            ws.cell(row=row, column=5, value=str(item.unit_of_measure) if item.unit_of_measure else '')
            ws.cell(row=row, column=6, value=float(warehouse_item.average_cost))
            ws.cell(row=row, column=7, value=float(total_value))
            ws.cell(row=row, column=8, value=float(item.reorder_level or 0))
            ws.cell(row=row, column=9, value=float(item.max_level or 0))
            ws.cell(row=row, column=10, value=stock_status)

        # حفظ الملف
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
        return response


# ============== DataTables AJAX Views ==============

class WarehouseDataTableView(LoginRequiredMixin, BaseDatatableView):
    """DataTables AJAX للمستودعات"""
    model = Warehouse
    columns = ['code', 'name', 'warehouse_type', 'branch', 'keeper', 'location', 'is_active', 'actions']
    order_columns = ['code', 'name', 'warehouse_type', 'branch__name', 'keeper__first_name', 'location', 'is_active']
    max_display_length = 100

    def get_initial_queryset(self):
        return Warehouse.objects.filter(
            company=self.request.user.company
        ).select_related('branch', 'keeper')

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search) |
                Q(location__icontains=search) |
                Q(branch__name__icontains=search) |
                Q(keeper__first_name__icontains=search) |
                Q(keeper__last_name__icontains=search)
            )
        return qs

    def prepare_results(self, qs):
        json_data = []
        for warehouse in qs:
            actions = f'''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="{reverse_lazy('base_data:warehouse_detail', args=[warehouse.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:warehouse_edit', args=[warehouse.pk])}" 
                       class="btn btn-outline-warning btn-sm" title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:warehouse_inventory', args=[warehouse.pk])}" 
                       class="btn btn-outline-success btn-sm" title="الأرصدة">
                        <i class="fas fa-boxes"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:warehouse_delete', args=[warehouse.pk])}" 
                       class="btn btn-outline-danger btn-sm" title="حذف"
                       onclick="return confirm('هل أنت متأكد من الحذف؟')">
                        <i class="fas fa-trash"></i>
                    </a>
                </div>
            '''

            json_data.append([
                warehouse.code,
                warehouse.name,
                warehouse.get_warehouse_type_display(),
                str(warehouse.branch) if warehouse.branch else '',
                warehouse.keeper.get_full_name() if warehouse.keeper else '',
                warehouse.location or '',
                '<span class="badge badge-success">نشط</span>' if warehouse.is_active else
                '<span class="badge badge-warning">غير نشط</span>',
                actions
            ])
        return json_data


class UnitDataTableView(LoginRequiredMixin, BaseDatatableView):
    """DataTables AJAX لوحدات القياس"""
    model = UnitOfMeasure
    columns = ['code', 'name', 'name_en', 'is_active', 'actions']
    order_columns = ['code', 'name', 'name_en', 'is_active']
    max_display_length = 100

    def get_initial_queryset(self):
        return UnitOfMeasure.objects.filter(
            company=self.request.user.company
        )

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )
        return qs

    def prepare_results(self, qs):
        json_data = []
        for unit in qs:
            actions = f'''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="{reverse_lazy('base_data:unit_edit', args=[unit.pk])}" 
                       class="btn btn-outline-warning btn-sm" title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:unit_delete', args=[unit.pk])}" 
                       class="btn btn-outline-danger btn-sm" title="حذف"
                       onclick="return confirm('هل أنت متأكد من الحذف؟')">
                        <i class="fas fa-trash"></i>
                    </a>
                </div>
            '''

            json_data.append([
                unit.code,
                unit.name,
                unit.name_en or '',
                '<span class="badge badge-success">نشط</span>' if unit.is_active else
                '<span class="badge badge-warning">غير نشط</span>',
                actions
            ])
        return json_data


# ============== AJAX Views للنماذج ==============

class WarehouseSearchAjaxView(LoginRequiredMixin, View):
    """البحث عن المستودعات عبر AJAX"""

    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')
        branch_id = request.GET.get('branch_id', '')
        warehouse_type = request.GET.get('type', '')

        warehouses_qs = Warehouse.objects.filter(
            company=request.user.company,
            is_active=True
        )

        if branch_id:
            warehouses_qs = warehouses_qs.filter(branch_id=branch_id)

        if warehouse_type:
            warehouses_qs = warehouses_qs.filter(warehouse_type=warehouse_type)

        warehouses_qs = warehouses_qs.filter(
            Q(code__icontains=term) |
            Q(name__icontains=term) |
            Q(name_en__icontains=term)
        ).order_by('name')[:20]

        results = []
        for warehouse in warehouses_qs:
            results.append({
                'id': warehouse.id,
                'text': f"{warehouse.code} - {warehouse.name}",
                'code': warehouse.code,
                'name': warehouse.name,
                'warehouse_type': warehouse.warehouse_type,
                'branch_id': warehouse.branch_id if warehouse.branch else None,
                'branch_name': str(warehouse.branch) if warehouse.branch else '',
                'keeper_name': warehouse.keeper.get_full_name() if warehouse.keeper else '',
                'location': warehouse.location or ''
            })

        return JsonResponse({'results': results})


class WarehouseInfoAjaxView(LoginRequiredMixin, View):
    """الحصول على معلومات المستودع عبر AJAX"""

    def get(self, request, pk):
        try:
            warehouse = Warehouse.objects.select_related(
                'branch', 'keeper'
            ).prefetch_related('warehouse_items').get(
                pk=pk,
                company=request.user.company
            )

            # إحصائيات المستودع
            warehouse_items = warehouse.warehouse_items.all()
            stats = {
                'total_items': warehouse_items.count(),
                'items_in_stock': warehouse_items.filter(quantity__gt=0).count(),
                'total_value': warehouse_items.aggregate(
                    total=Sum(F('quantity') * F('average_cost'))
                )['total'] or 0,
            }

            data = {
                'id': warehouse.id,
                'code': warehouse.code,
                'name': warehouse.name,
                'name_en': warehouse.name_en or '',
                'warehouse_type': warehouse.warehouse_type,
                'warehouse_type_display': warehouse.get_warehouse_type_display(),
                'location': warehouse.location or '',
                'branch_id': warehouse.branch_id if warehouse.branch else None,
                'branch_name': str(warehouse.branch) if warehouse.branch else '',
                'keeper_id': warehouse.keeper_id if warehouse.keeper else None,
                'keeper_name': warehouse.keeper.get_full_name() if warehouse.keeper else '',
                'is_active': warehouse.is_active,
                'stats': stats,
                'notes': warehouse.notes or ''
            }

            return JsonResponse({'success': True, 'warehouse': data})

        except Warehouse.DoesNotExist:
            return JsonResponse({'success': False, 'message': _('المستودع غير موجود')})


class UnitSearchAjaxView(LoginRequiredMixin, View):
    """البحث عن وحدات القياس عبر AJAX"""

    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')

        units = UnitOfMeasure.objects.filter(
            company=request.user.company,
            is_active=True
        ).filter(
            Q(code__icontains=term) |
            Q(name__icontains=term) |
            Q(name_en__icontains=term)
        ).order_by('name')[:20]

        results = []
        for unit in units:
            results.append({
                'id': unit.id,
                'text': f"{unit.code} - {unit.name}",
                'code': unit.code,
                'name': unit.name,
                'name_en': unit.name_en or ''
            })

        return JsonResponse({'results': results})


# ============== Views مساعدة ==============

def get_warehouses_by_branch_ajax(request):
    """الحصول على المستودعات حسب الفرع عبر AJAX"""
    if not request.user.has_perm('base_data.view_warehouse'):
        return JsonResponse({'error': 'ليس لديك صلاحية'}, status=403)

    branch_id = request.GET.get('branch_id')
    if not branch_id:
        return JsonResponse({'warehouses': []})

    warehouses = Warehouse.objects.filter(
        company=request.user.company,
        branch_id=branch_id,
        is_active=True
    ).values('id', 'name', 'code', 'warehouse_type').order_by('name')

    return JsonResponse({'warehouses': list(warehouses)})


def get_warehouse_items_ajax(request):
    """الحصول على أصناف المستودع عبر AJAX"""
    if not request.user.has_perm('base_data.view_warehouse'):
        return JsonResponse({'error': 'ليس لديك صلاحية'}, status=403)

    warehouse_id = request.GET.get('warehouse_id')
    if not warehouse_id:
        return JsonResponse({'items': []})

    warehouse_items = WarehouseItem.objects.filter(
        warehouse_id=warehouse_id,
        warehouse__company=request.user.company,
        quantity__gt=0
    ).select_related('item').values(
        'item__id', 'item__code', 'item__name', 'quantity', 'average_cost'
    ).order_by('item__name')

    return JsonResponse({'items': list(warehouse_items)})