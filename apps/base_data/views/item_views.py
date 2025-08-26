# apps/base_data/views/item_views.py
"""
Views الخاصة بالأصناف والمواد - كامل ومطابق للمتطلبات
يشمل: الأصناف، التصنيفات، معدلات التحويل، المكونات، المواد البديلة
"""

from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, Sum, Avg, F, Case, When, BooleanField, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views import View
from django_datatables_view.base_datatable_view import BaseDatatableView
from decimal import Decimal
import csv
import json
import openpyxl
from io import BytesIO

from ..models import (
    Item, ItemCategory, ItemConversion, ItemComponent, UnitOfMeasure,
    Warehouse, BusinessPartner
)

# استيراد النماذج من الملف المنفصل
from ..forms.item_forms import (
    ItemForm, ItemCategoryForm, ItemConversionForm, ItemComponentForm,
    ItemQuickAddForm, ItemFilterForm
)


class BaseItemMixin:
    """Mixin أساسي للأصناف - يحتوي على الإعدادات المشتركة"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إضافة معلومات الشركة والفرع
        context['current_company'] = self.request.user.company
        context['current_branch'] = getattr(self.request.user, 'branch', None)

        # Breadcrumbs
        context['breadcrumbs'] = self.get_breadcrumbs()

        # إحصائيات عامة
        context['total_items'] = Item.objects.filter(
            company=self.request.user.company
        ).count()
        context['active_items'] = Item.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).count()

        return context

    def get_breadcrumbs(self):
        """بناء breadcrumbs للصفحة"""
        return [
            {'title': _('الرئيسية'), 'url': '/'},
            {'title': _('البيانات الأساسية'), 'url': '#'},
        ]

    def get_queryset(self):
        """فلترة البيانات حسب الشركة"""
        queryset = super().get_queryset()
        return queryset.filter(company=self.request.user.company)


# ============== Views الأصناف ==============

class ItemListView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, ListView):
    """عرض قائمة الأصناف"""
    model = Item
    template_name = 'base_data/items/item_list.html'
    context_object_name = 'items'
    permission_required = 'base_data.view_item'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('قائمة الأصناف')
        context['filter_form'] = ItemFilterForm(
            self.request.GET,
            company=self.request.user.company
        )

        # إحصائيات إضافية
        context['categories_count'] = ItemCategory.objects.filter(
            company=self.request.user.company
        ).count()

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'category', 'unit'
        ).prefetch_related('substitute_items')

        # تطبيق الفلاتر
        filter_form = ItemFilterForm(
            self.request.GET,
            company=self.request.user.company
        )
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('search'):
                search = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(code__icontains=search) |
                    Q(name__icontains=search) |
                    Q(name_en__icontains=search)
                )

            if filter_form.cleaned_data.get('category'):
                queryset = queryset.filter(category=filter_form.cleaned_data['category'])

            if filter_form.cleaned_data.get('is_inactive'):
                queryset = queryset.filter(is_active=False)
            else:
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('code', 'name')


class ItemDetailView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DetailView):
    """تفاصيل الصنف"""
    model = Item
    template_name = 'base_data/items/item_detail.html'
    context_object_name = 'item'
    permission_required = 'base_data.view_item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل الصنف: %s') % self.object.name

        # معدلات التحويل
        context['conversions'] = self.object.conversions.filter(
            company=self.request.user.company
        ).select_related('from_unit', 'to_unit')

        # المكونات
        context['components'] = self.object.components.filter(
            company=self.request.user.company
        ).select_related('component_item', 'unit')

        # المواد البديلة
        context['substitutes'] = self.object.substitute_items.filter(
            company=self.request.user.company,
            is_active=True
        )

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'active': True}
        ])

        return context


class ItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin,
                     SuccessMessageMixin, CreateView):
    """إنشاء صنف جديد"""
    model = Item
    form_class = ItemForm
    template_name = 'base_data/items/item_form.html'
    permission_required = 'base_data.add_item'
    success_message = _('تم إنشاء الصنف بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة صنف جديد')
        context['submit_text'] = _('حفظ')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('إضافة صنف جديد'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:item_add')
        return reverse_lazy('base_data:item_detail', kwargs={'pk': self.object.pk})


class ItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin,
                     SuccessMessageMixin, UpdateView):
    """تعديل الصنف"""
    model = Item
    form_class = ItemForm
    template_name = 'base_data/items/item_form.html'
    permission_required = 'base_data.change_item'
    success_message = _('تم تعديل الصنف بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل الصنف: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:item_detail', args=[self.object.pk])},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:item_detail', kwargs={'pk': self.object.pk})


class ItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DeleteView):
    """حذف الصنف"""
    model = Item
    template_name = 'base_data/items/item_delete.html'
    permission_required = 'base_data.delete_item'
    success_url = reverse_lazy('base_data:item_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('حذف الصنف: %s') % self.object.name

        # التحقق من الارتباطات
        context['has_transactions'] = False  # يجب فحص الحركات المرتبطة
        context['has_components'] = self.object.components.exists()
        context['has_conversions'] = self.object.conversions.exists()

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:item_detail', args=[self.object.pk])},
            {'title': _('حذف'), 'active': True}
        ])

        return context

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف الصنف بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============== Views التصنيفات ==============

class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, ListView):
    """قائمة التصنيفات"""
    model = ItemCategory
    template_name = 'base_data/items/category_list.html'
    context_object_name = 'categories'
    permission_required = 'base_data.view_itemcategory'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تصنيفات الأصناف')

        # للفلاتر
        context['parent_categories'] = ItemCategory.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('التصنيفات'), 'active': True}
        ])

        return context

    def get_queryset(self):
        return super().get_queryset().annotate(
            items_count=Count('item_set')
        ).order_by('code', 'name')


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin,
                         SuccessMessageMixin, CreateView):
    """إنشاء تصنيف جديد"""
    model = ItemCategory
    form_class = ItemCategoryForm
    template_name = 'base_data/items/category_form.html'
    permission_required = 'base_data.add_itemcategory'
    success_message = _('تم إنشاء التصنيف بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة تصنيف جديد')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('التصنيفات'), 'url': reverse_lazy('base_data:category_list')},
            {'title': _('إضافة تصنيف'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:category_list')


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin,
                         SuccessMessageMixin, UpdateView):
    """تعديل التصنيف"""
    model = ItemCategory
    form_class = ItemCategoryForm
    template_name = 'base_data/items/category_form.html'
    permission_required = 'base_data.change_itemcategory'
    success_message = _('تم تعديل التصنيف بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل التصنيف: %s') % self.object.name

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('التصنيفات'), 'url': reverse_lazy('base_data:category_list')},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def get_success_url(self):
        return reverse_lazy('base_data:category_list')


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DeleteView):
    """حذف التصنيف"""
    model = ItemCategory
    template_name = 'base_data/items/category_delete.html'
    permission_required = 'base_data.delete_itemcategory'
    success_url = reverse_lazy('base_data:category_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('حذف التصنيف: %s') % self.object.name

        # التحقق من الارتباطات
        context['items_count'] = self.object.item_set.count()
        context['children_count'] = self.object.children.count()
        context['has_items'] = context['items_count'] > 0
        context['has_children'] = context['children_count'] > 0

        # إجمالي التصنيفات الفرعية (جميع المستويات)
        def count_descendants(category):
            count = 0
            for child in category.children.all():
                count += 1 + count_descendants(child)
            return count

        context['total_descendants'] = count_descendants(self.object)

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('التصنيفات'), 'url': reverse_lazy('base_data:category_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:category_edit', args=[self.object.pk])},
            {'title': _('حذف'), 'active': True}
        ])

        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من وجود أصناف مرتبطة
        if self.object.item_set.exists():
            messages.error(request, _('لا يمكن حذف التصنيف لوجود أصناف مرتبطة به'))
            return redirect('base_data:category_edit', pk=self.object.pk)

        # التحقق من وجود تصنيفات فرعية
        if self.object.children.exists():
            messages.error(request, _('لا يمكن حذف التصنيف لوجود تصنيفات فرعية'))
            return redirect('base_data:category_edit', pk=self.object.pk)

        category_name = self.object.name
        messages.success(request, _('تم حذف التصنيف "%s" بنجاح') % category_name)
        return super().delete(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        """معالجة طلبات AJAX للحذف"""
        if request.headers.get('Content-Type') == 'application/json':
            self.object = self.get_object()

            # التحقق من الارتباطات
            if self.object.item_set.exists() or self.object.children.exists():
                return JsonResponse({
                    'success': False,
                    'message': _('لا يمكن حذف التصنيف لوجود أصناف أو تصنيفات فرعية مرتبطة')
                })

            category_name = self.object.name
            self.object.delete()

            return JsonResponse({
                'success': True,
                'message': _('تم حذف التصنيف "%s" بنجاح') % category_name
            })

        return super().post(request, *args, **kwargs)


# ============== Views المكونات ==============

class ItemComponentsManageView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DetailView):
    """إدارة مكونات الصنف"""
    model = Item
    template_name = 'base_data/items/item_components.html'
    context_object_name = 'item'
    permission_required = 'base_data.change_item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إدارة مكونات: %s') % self.object.name

        # المكونات الحالية
        context['components'] = self.object.components.filter(
            company=self.request.user.company
        ).select_related('component_item', 'unit').order_by('component_item__name')

        # النموذج لإضافة مكون جديد
        context['component_form'] = ItemComponentForm()

        # الأصناف المتاحة للإضافة كمكونات
        context['available_items'] = Item.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).exclude(
            id=self.object.id
        ).exclude(
            id__in=self.object.components.values_list('component_item_id', flat=True)
        ).order_by('name')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:item_detail', args=[self.object.pk])},
            {'title': _('المكونات'), 'active': True}
        ])

        return context

    def post(self, request, *args, **kwargs):
        """إضافة أو تعديل أو حذف مكون"""
        self.object = self.get_object()
        action = request.POST.get('action')

        if action == 'add':
            form = ItemComponentForm(request.POST)
            if form.is_valid():
                component = form.save(commit=False)
                component.parent_item = self.object
                component.company = request.user.company
                component.created_by = request.user
                component.save()
                messages.success(request, _('تم إضافة المكون بنجاح'))
            else:
                messages.error(request, _('خطأ في البيانات المدخلة'))

        elif action == 'delete':
            component_id = request.POST.get('component_id')
            try:
                component = ItemComponent.objects.get(
                    id=component_id,
                    parent_item=self.object,
                    company=request.user.company
                )
                component.delete()
                messages.success(request, _('تم حذف المكون بنجاح'))
            except ItemComponent.DoesNotExist:
                messages.error(request, _('المكون غير موجود'))

        return redirect('base_data:item_components', pk=self.object.pk)


# ============== Views معدلات التحويل ==============

class ItemConversionsManageView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DetailView):
    """إدارة معدلات تحويل الصنف"""
    model = Item
    template_name = 'base_data/items/item_conversions.html'
    context_object_name = 'item'
    permission_required = 'base_data.change_item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إدارة معدلات التحويل: %s') % self.object.name

        # معدلات التحويل الحالية
        context['conversions'] = self.object.conversions.filter(
            company=self.request.user.company
        ).select_related('from_unit', 'to_unit').order_by('from_unit__name')

        # النموذج لإضافة معدل تحويل جديد
        context['conversion_form'] = ItemConversionForm()

        # الوحدات المتاحة
        context['units'] = UnitOfMeasure.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:item_detail', args=[self.object.pk])},
            {'title': _('معدلات التحويل'), 'active': True}
        ])

        return context

    def post(self, request, *args, **kwargs):
        """إضافة أو تعديل أو حذف معدل تحويل"""
        self.object = self.get_object()
        action = request.POST.get('action')

        if action == 'add':
            form = ItemConversionForm(request.POST)
            if form.is_valid():
                conversion = form.save(commit=False)
                conversion.item = self.object
                conversion.company = request.user.company
                conversion.created_by = request.user
                conversion.save()
                messages.success(request, _('تم إضافة معدل التحويل بنجاح'))
            else:
                messages.error(request, _('خطأ في البيانات المدخلة'))

        elif action == 'delete':
            conversion_id = request.POST.get('conversion_id')
            try:
                conversion = ItemConversion.objects.get(
                    id=conversion_id,
                    item=self.object,
                    company=request.user.company
                )
                conversion.delete()
                messages.success(request, _('تم حذف معدل التحويل بنجاح'))
            except ItemConversion.DoesNotExist:
                messages.error(request, _('معدل التحويل غير موجود'))

        return redirect('base_data:item_conversions', pk=self.object.pk)


# ============== Views المواد البديلة ==============

class ItemSubstitutesManageView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, DetailView):
    """إدارة المواد البديلة لصنف معين"""
    model = Item
    template_name = 'base_data/items/item_substitutes.html'
    context_object_name = 'item'
    permission_required = 'base_data.change_item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إدارة المواد البديلة: %s') % self.object.name

        # المواد البديلة الحالية
        context['current_substitutes'] = self.object.substitute_items.filter(
            is_active=True,
            company=self.request.user.company
        ).order_by('name')

        # الأصناف المتاحة للإضافة كبديلة
        context['available_items'] = Item.objects.filter(
            company=self.request.user.company,
            is_active=True,
            category=self.object.category  # نفس التصنيف
        ).exclude(
            id=self.object.id
        ).exclude(
            id__in=self.object.substitute_items.values_list('id', flat=True)
        ).order_by('name')

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:item_detail', args=[self.object.pk])},
            {'title': _('المواد البديلة'), 'active': True}
        ])

        return context

    def post(self, request, *args, **kwargs):
        """إضافة أو إزالة مواد بديلة"""
        self.object = self.get_object()
        action = request.POST.get('action')
        item_id = request.POST.get('item_id')

        if action == 'add' and item_id:
            try:
                substitute_item = Item.objects.get(id=item_id, company=request.user.company)
                self.object.substitute_items.add(substitute_item)
                messages.success(request, _('تم إضافة المادة البديلة بنجاح'))
            except Item.DoesNotExist:
                messages.error(request, _('الصنف غير موجود'))

        elif action == 'remove' and item_id:
            try:
                substitute_item = Item.objects.get(id=item_id, company=request.user.company)
                self.object.substitute_items.remove(substitute_item)
                messages.success(request, _('تم إزالة المادة البديلة بنجاح'))
            except Item.DoesNotExist:
                messages.error(request, _('الصنف غير موجود'))

        return redirect('base_data:item_substitutes', pk=self.object.pk)


# ============== Views التقارير والتصدير ==============

class ItemReportView(LoginRequiredMixin, PermissionRequiredMixin, BaseItemMixin, ListView):
    """تقرير الأصناف"""
    model = Item
    template_name = 'base_data/items/item_report.html'
    permission_required = 'base_data.view_item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تقرير الأصناف')
        context['filter_form'] = ItemFilterForm(
            self.request.GET,
            company=self.request.user.company
        )

        context['breadcrumbs'].extend([
            {'title': _('الأصناف'), 'url': reverse_lazy('base_data:item_list')},
            {'title': _('التقارير'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'category', 'unit'
        )

        # تطبيق الفلاتر
        filter_form = ItemFilterForm(
            self.request.GET,
            company=self.request.user.company
        )
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('category'):
                queryset = queryset.filter(category=filter_form.cleaned_data['category'])
            if filter_form.cleaned_data.get('is_inactive'):
                queryset = queryset.filter(is_active=False)
            else:
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('category__name', 'name')


class ItemExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """تصدير الأصناف إلى Excel/CSV"""
    permission_required = 'base_data.view_item'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('format', 'excel')

        # الحصول على البيانات
        items = Item.objects.filter(
            company=request.user.company
        ).select_related('category', 'unit').order_by('code')

        if export_format == 'excel':
            return self.export_excel(items)
        else:
            return self.export_csv(items)

    def export_excel(self, items):
        """تصدير إلى Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأصناف"

        # العناوين
        headers = [
            'كود الصنف', 'اسم الصنف', 'الاسم الإنجليزي', 'التصنيف',
            'وحدة القياس', 'سعر البيع', 'سعر الشراء',
            'الحالة', 'ملاحظات'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # البيانات
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.code)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.name_en or '')
            ws.cell(row=row, column=4, value=str(item.category) if item.category else '')
            ws.cell(row=row, column=5, value=str(item.unit) if item.unit else '')
            ws.cell(row=row, column=6, value=float(item.sale_price or 0))
            ws.cell(row=row, column=7, value=float(item.purchase_price or 0))
            ws.cell(row=row, column=8, value='غير نشط' if not item.is_active else 'نشط')
            ws.cell(row=row, column=9, value=item.notes or '')

        # حفظ الملف
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="items_report.xlsx"'
        return response

    def export_csv(self, items):
        """تصدير إلى CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="items_report.csv"'
        response.write('\ufeff')  # BOM for Arabic support

        writer = csv.writer(response)

        # العناوين
        writer.writerow([
            'كود الصنف', 'اسم الصنف', 'الاسم الإنجليزي', 'التصنيف',
            'وحدة القياس', 'سعر البيع', 'سعر الشراء',
            'الحالة', 'ملاحظات'
        ])

        # البيانات
        for item in items:
            writer.writerow([
                item.code,
                item.name,
                item.name_en or '',
                str(item.category) if item.category else '',
                str(item.unit) if item.unit else '',
                float(item.sale_price or 0),
                float(item.purchase_price or 0),
                'غير نشط' if not item.is_active else 'نشط',
                item.notes or ''
            ])

        return response


# ============== DataTables AJAX Views ==============

class ItemDataTableView(LoginRequiredMixin, BaseDatatableView):
    """DataTables AJAX للأصناف"""
    model = Item
    columns = ['code', 'name', 'category', 'unit', 'sale_price', 'is_active', 'actions']
    order_columns = ['code', 'name', 'category__name', 'unit__name', 'sale_price', 'is_active']
    max_display_length = 100

    def get_initial_queryset(self):
        return Item.objects.filter(
            company=self.request.user.company
        ).select_related('category', 'unit')

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )
        return qs

    def prepare_results(self, qs):
        json_data = []
        for item in qs:
            actions = f'''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="{reverse_lazy('base_data:item_detail', args=[item.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:item_edit', args=[item.pk])}" 
                       class="btn btn-outline-warning btn-sm" title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:item_delete', args=[item.pk])}" 
                       class="btn btn-outline-danger btn-sm" title="حذف"
                       onclick="return confirm('هل أنت متأكد من الحذف؟')">
                        <i class="fas fa-trash"></i>
                    </a>
                </div>
            '''

            json_data.append([
                item.code,
                item.name,
                str(item.category) if item.category else '',
                str(item.unit) if item.unit else '',
                f"{item.sale_price:.2f}" if item.sale_price else '',
                '<span class="badge badge-success">نشط</span>' if item.is_active else
                '<span class="badge badge-warning">غير نشط</span>',
                actions
            ])
        return json_data


class CategoryDataTableView(LoginRequiredMixin, BaseDatatableView):
    """DataTables AJAX للتصنيفات"""
    model = ItemCategory
    columns = ['code', 'name', 'name_en', 'parent', 'level', 'items_count', 'is_active', 'actions']
    order_columns = ['code', 'name', 'name_en', 'parent__name', 'level', 'items_count', 'is_active']
    max_display_length = 100

    def get_initial_queryset(self):
        return ItemCategory.objects.filter(
            company=self.request.user.company
        ).select_related('parent').annotate(
            items_count=Count('item_set')
        )

    def filter_queryset(self, qs):
        # البحث العام
        search = self.request.GET.get('search[value]')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        # فلاتر إضافية من النموذج
        parent_category = self.request.GET.get('parent_category')
        if parent_category:
            qs = qs.filter(parent_id=parent_category)

        level = self.request.GET.get('level')
        if level:
            qs = qs.filter(level=level)

        is_active = self.request.GET.get('is_active')
        if is_active == 'true':
            qs = qs.filter(is_active=True)
        elif is_active == 'false':
            qs = qs.filter(is_active=False)

        return qs

    def prepare_results(self, qs):
        json_data = []
        for category in qs:
            # أزرار الإجراءات
            actions = '<div class="btn-group btn-group-sm" role="group">'

            # زر العرض
            actions += f'''<a href="{reverse_lazy('base_data:category_edit', args=[category.pk])}" 
                              class="btn btn-outline-info btn-sm" title="عرض">
                              <i class="fas fa-eye"></i>
                           </a>'''

            # زر التعديل
            if self.request.user.has_perm('base_data.change_itemcategory'):
                actions += f'''<a href="{reverse_lazy('base_data:category_edit', args=[category.pk])}" 
                                  class="btn btn-outline-warning btn-sm" title="تعديل">
                                  <i class="fas fa-edit"></i>
                               </a>'''

            # زر الحذف
            if self.request.user.has_perm('base_data.delete_itemcategory'):
                actions += f'''<button type="button" 
                                      class="btn btn-outline-danger btn-sm" 
                                      title="حذف"
                                      onclick="showDeleteConfirm({category.pk}, '{category.name}')">
                                      <i class="fas fa-trash"></i>
                               </button>'''

            actions += '</div>'

            # بناء البيانات
            json_data.append([
                category.code,
                category.name,
                category.name_en or '',
                category.parent.name if category.parent else '',
                f'المستوى {category.level}',
                f'<span class="badge badge-info">{category.items_count}</span>',
                '<span class="badge badge-success">نشط</span>' if category.is_active else
                '<span class="badge badge-warning">غير نشط</span>',
                actions
            ])

        return json_data


# ============== AJAX Views للنماذج ==============

class ItemQuickAddView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """إضافة صنف سريع عبر AJAX"""
    permission_required = 'base_data.add_item'

    def post(self, request, *args, **kwargs):
        form = ItemQuickAddForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.company = request.user.company
            item.created_by = request.user
            item.save()

            return JsonResponse({
                'success': True,
                'item_id': item.id,
                'item_name': item.name,
                'message': _('تم إضافة الصنف بنجاح')
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })


class CategoryQuickAddView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """إضافة تصنيف سريع عبر AJAX"""
    permission_required = 'base_data.add_itemcategory'

    def post(self, request, *args, **kwargs):
        try:
            # البيانات الأساسية
            code = request.POST.get('code', '').strip()
            name = request.POST.get('name', '').strip()
            name_en = request.POST.get('name_en', '').strip()
            parent_id = request.POST.get('parent')

            # التحقق من البيانات المطلوبة
            if not code or not name:
                return JsonResponse({
                    'success': False,
                    'message': _('الكود والاسم مطلوبان')
                })

            # التحقق من تكرار الكود
            if ItemCategory.objects.filter(
                    company=request.user.company,
                    code=code
            ).exists():
                return JsonResponse({
                    'success': False,
                    'message': _('هذا الكود مستخدم من قبل')
                })

            # إنشاء التصنيف
            category = ItemCategory(
                company=request.user.company,
                created_by=request.user,
                code=code,
                name=name,
                name_en=name_en or None
            )

            # تعيين التصنيف الأب إذا كان موجوداً
            if parent_id:
                try:
                    parent = ItemCategory.objects.get(
                        id=parent_id,
                        company=request.user.company
                    )
                    category.parent = parent
                except ItemCategory.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': _('التصنيف الأب غير موجود')
                    })

            category.save()

            return JsonResponse({
                'success': True,
                'category_id': category.id,
                'category_name': category.name,
                'category_code': category.code,
                'message': _('تم إنشاء التصنيف بنجاح')
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': _('حدث خطأ أثناء الحفظ: %s') % str(e)
            })


class CategoryCheckCodeView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """التحقق من تكرار كود التصنيف عبر AJAX"""
    permission_required = 'base_data.add_itemcategory'

    def post(self, request, *args, **kwargs):
        code = request.POST.get('code', '').strip()
        category_id = request.POST.get('id')

        if not code:
            return JsonResponse({'exists': False})

        # بناء الاستعلام
        queryset = ItemCategory.objects.filter(
            company=request.user.company,
            code=code
        )

        # استبعاد التصنيف الحالي في حالة التعديل
        if category_id:
            queryset = queryset.exclude(id=category_id)

        exists = queryset.exists()

        return JsonResponse({'exists': exists})


class CategoryExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """تصدير التصنيفات إلى Excel/CSV"""
    permission_required = 'base_data.view_itemcategory'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('format', 'excel')

        # الحصول على البيانات
        categories = ItemCategory.objects.filter(
            company=request.user.company
        ).select_related('parent').annotate(
            items_count=Count('item_set'),
            children_count=Count('children')
        ).order_by('level', 'code')

        # تطبيق الفلاتر
        parent_category = request.GET.get('parent_category')
        if parent_category:
            categories = categories.filter(parent_id=parent_category)

        level = request.GET.get('level')
        if level:
            categories = categories.filter(level=level)

        is_active = request.GET.get('is_active')
        if is_active == 'true':
            categories = categories.filter(is_active=True)
        elif is_active == 'false':
            categories = categories.filter(is_active=False)

        if export_format == 'excel':
            return self.export_excel(categories)
        else:
            return self.export_csv(categories)

    def export_excel(self, categories):
        """تصدير إلى Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تصنيفات الأصناف"

        # العناوين
        headers = [
            'كود التصنيف', 'اسم التصنيف', 'الاسم الإنجليزي',
            'التصنيف الأب', 'المستوى', 'عدد الأصناف',
            'عدد التصنيفات الفرعية', 'الحالة', 'تاريخ الإنشاء'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # البيانات
        for row, category in enumerate(categories, 2):
            ws.cell(row=row, column=1, value=category.code)
            ws.cell(row=row, column=2, value=category.name)
            ws.cell(row=row, column=3, value=category.name_en or '')
            ws.cell(row=row, column=4, value=category.parent.name if category.parent else '')
            ws.cell(row=row, column=5, value=f'المستوى {category.level}')
            ws.cell(row=row, column=6, value=category.items_count)
            ws.cell(row=row, column=7, value=category.children_count)
            ws.cell(row=row, column=8, value='غير نشط' if not category.is_active else 'نشط')
            ws.cell(row=row, column=9,
                    value=category.created_at.strftime('%Y-%m-%d') if hasattr(category, 'created_at') else '')

        # حفظ الملف
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="categories_report.xlsx"'
        return response

    def export_csv(self, categories):
        """تصدير إلى CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="categories_report.csv"'
        response.write('\ufeff')  # BOM for Arabic support

        writer = csv.writer(response)

        # العناوين
        writer.writerow([
            'كود التصنيف', 'اسم التصنيف', 'الاسم الإنجليزي',
            'التصنيف الأب', 'المستوى', 'عدد الأصناف',
            'عدد التصنيفات الفرعية', 'الحالة', 'تاريخ الإنشاء'
        ])

        # البيانات
        for category in categories:
            writer.writerow([
                category.code,
                category.name,
                category.name_en or '',
                category.parent.name if category.parent else '',
                f'المستوى {category.level}',
                category.items_count,
                category.children_count,
                'غير نشط' if not category.is_active else 'نشط',
                category.created_at.strftime('%Y-%m-%d') if hasattr(category, 'created_at') else ''
            ])

        return response


class ItemSearchAjaxView(LoginRequiredMixin, View):
    """البحث عن الأصناف عبر AJAX"""

    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')
        items = Item.objects.filter(
            company=request.user.company,
            is_active=True
        ).filter(
            Q(code__icontains=term) |
            Q(name__icontains=term)
        ).order_by('name')[:20]

        results = []
        for item in items:
            results.append({
                'id': item.id,
                'text': f"{item.code} - {item.name}",
                'code': item.code,
                'name': item.name,
                'unit': str(item.unit) if item.unit else '',
                'sale_price': float(item.sale_price or 0),
                'purchase_price': float(item.purchase_price or 0)
            })

        return JsonResponse({'results': results})


class ItemInfoAjaxView(LoginRequiredMixin, View):
    """الحصول على معلومات الصنف عبر AJAX"""

    def get(self, request, pk):
        try:
            item = Item.objects.select_related(
                'category', 'unit'
            ).get(
                pk=pk,
                company=request.user.company
            )

            data = {
                'id': item.id,
                'code': item.code,
                'name': item.name,
                'name_en': item.name_en or '',
                'category': str(item.category) if item.category else '',
                'unit': str(item.unit) if item.unit else '',
                'sale_price': float(item.sale_price or 0),
                'purchase_price': float(item.purchase_price or 0),
                'tax_rate': float(item.tax_rate or 0),
                'minimum_quantity': float(item.minimum_quantity or 0),
                'maximum_quantity': float(item.maximum_quantity or 0) if item.maximum_quantity else None,
                'is_active': item.is_active,
                'barcode': item.barcode or '',
                'notes': item.notes or ''
            }

            return JsonResponse({'success': True, 'item': data})

        except Item.DoesNotExist:
            return JsonResponse({'success': False, 'message': _('الصنف غير موجود')})