# apps/base_data/views/partner_views.py
"""
Views الخاصة بالشركاء التجاريين - كامل ومطابق للمتطلبات
يشمل: العملاء، الموردين، معلومات الاتصال، الحسابات المحاسبية
"""

from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Count, Sum, Avg, F, Case, When, BooleanField, Prefetch
from django.http import JsonResponse, HttpResponse
from django.views import View
from django_datatables_view.base_datatable_view import BaseDatatableView
from decimal import Decimal
import csv
import json
import openpyxl
from io import BytesIO

from ..models import BusinessPartner, Customer, Supplier
from ..forms.partner_forms import (
    BusinessPartnerForm, CustomerForm, SupplierForm,
    PartnerQuickAddForm, PartnerFilterForm, ContactInfoForm
)


class BasePartnerMixin:
    """Mixin أساسي للشركاء - يحتوي على الإعدادات المشتركة"""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إضافة معلومات الشركة والفرع
        context['current_company'] = self.request.user.company
        context['current_branch'] = getattr(self.request.user, 'branch', None)

        # Breadcrumbs
        context['breadcrumbs'] = self.get_breadcrumbs()

        # إحصائيات عامة
        context['total_partners'] = BusinessPartner.objects.filter(
            company=self.request.user.company
        ).count()
        context['active_partners'] = BusinessPartner.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).count()
        context['customers_count'] = BusinessPartner.objects.filter(
            company=self.request.user.company,
            partner_type__in=['customer', 'both']
        ).count()
        context['suppliers_count'] = BusinessPartner.objects.filter(
            company=self.request.user.company,
            partner_type__in=['supplier', 'both']
        ).count()

        return context

    def get_breadcrumbs(self):
        """بناء breadcrumbs للصفحة"""
        return [
            {'title': _('الرئيسية'), 'url': '/'},
            {'title': _('البيانات الأساسية'), 'url': '#'},
        ]

    def get_queryset(self):
        """فلترة البيانات حسب الشركة"""
        queryset = super().get_queryset()
        return queryset.filter(company=self.request.user.company)


# ============== Views الشركاء التجاريين ==============

class BusinessPartnerListView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, ListView):
    """عرض قائمة الشركاء التجاريين"""
    model = BusinessPartner
    template_name = 'base_data/partners/partner_list.html'
    context_object_name = 'partners'
    permission_required = 'base_data.view_businesspartner'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('قائمة الشركاء التجاريين')
        context['filter_form'] = PartnerFilterForm(self.request.GET)

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related('salesperson')

        # تطبيق الفلاتر
        filter_form = PartnerFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('search'):
                search = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(code__icontains=search) |
                    Q(name__icontains=search) |
                    Q(name_en__icontains=search) |
                    Q(phone__icontains=search) |
                    Q(mobile__icontains=search) |
                    Q(email__icontains=search)
                )

            if filter_form.cleaned_data.get('partner_type'):
                queryset = queryset.filter(partner_type=filter_form.cleaned_data['partner_type'])

            if filter_form.cleaned_data.get('account_type'):
                queryset = queryset.filter(account_type=filter_form.cleaned_data['account_type'])

            if filter_form.cleaned_data.get('city'):
                queryset = queryset.filter(city__icontains=filter_form.cleaned_data['city'])

            if filter_form.cleaned_data.get('is_active'):
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('partner_type', 'code', 'name')


class BusinessPartnerDetailView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, DetailView):
    """تفاصيل الشريك التجاري"""
    model = BusinessPartner
    template_name = 'base_data/partners/partner_detail.html'
    context_object_name = 'partner'
    permission_required = 'base_data.view_businesspartner'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل الشريك: %s') % self.object.name

        # معلومات إضافية
        context['is_customer'] = self.object.is_customer()
        context['is_supplier'] = self.object.is_supplier()

        # إحصائيات الحساب (يمكن إضافتها لاحقاً)
        # context['total_sales'] = ...
        # context['total_purchases'] = ...
        # context['current_balance'] = ...

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': str(self.object), 'active': True}
        ])

        return context


class BusinessPartnerCreateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                                SuccessMessageMixin, CreateView):
    """إنشاء شريك تجاري جديد"""
    model = BusinessPartner
    form_class = BusinessPartnerForm
    template_name = 'base_data/partners/partner_form.html'
    permission_required = 'base_data.add_businesspartner'
    success_message = _('تم إنشاء الشريك التجاري بنجاح')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة شريك تجاري جديد')
        context['submit_text'] = _('حفظ')

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': _('إضافة شريك جديد'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:partner_add')
        return reverse_lazy('base_data:partner_detail', kwargs={'pk': self.object.pk})


class BusinessPartnerUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                                SuccessMessageMixin, UpdateView):
    """تعديل الشريك التجاري"""
    model = BusinessPartner
    form_class = BusinessPartnerForm
    template_name = 'base_data/partners/partner_form.html'
    permission_required = 'base_data.change_businesspartner'
    success_message = _('تم تعديل الشريك التجاري بنجاح')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل الشريك: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:partner_detail', args=[self.object.pk])},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:partner_detail', kwargs={'pk': self.object.pk})


class BusinessPartnerDeleteView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, DeleteView):
    """حذف الشريك التجاري"""
    model = BusinessPartner
    template_name = 'base_data/partners/partner_delete.html'
    permission_required = 'base_data.delete_businesspartner'
    success_url = reverse_lazy('base_data:partner_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('حذف الشريك: %s') % self.object.name

        # التحقق من الارتباطات
        context['has_transactions'] = False  # يجب فحص الحركات المرتبطة
        context['is_customer'] = self.object.is_customer()
        context['is_supplier'] = self.object.is_supplier()

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:partner_detail', args=[self.object.pk])},
            {'title': _('حذف'), 'active': True}
        ])

        return context

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف الشريك التجاري بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============== Views العملاء ==============

class CustomerListView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, ListView):
    """قائمة العملاء"""
    model = BusinessPartner
    template_name = 'base_data/partners/customer_list.html'
    context_object_name = 'customers'
    permission_required = 'base_data.view_businesspartner'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('قائمة العملاء')
        context['filter_form'] = PartnerFilterForm(self.request.GET)

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': _('العملاء'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            partner_type__in=['customer', 'both']
        ).select_related('salesperson')

        # تطبيق الفلاتر
        filter_form = PartnerFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('search'):
                search = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(code__icontains=search) |
                    Q(name__icontains=search) |
                    Q(name_en__icontains=search) |
                    Q(phone__icontains=search) |
                    Q(mobile__icontains=search) |
                    Q(email__icontains=search)
                )

            if filter_form.cleaned_data.get('account_type'):
                queryset = queryset.filter(account_type=filter_form.cleaned_data['account_type'])

            if filter_form.cleaned_data.get('city'):
                queryset = queryset.filter(city__icontains=filter_form.cleaned_data['city'])

            if filter_form.cleaned_data.get('is_active'):
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('code', 'name')


class CustomerDetailView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, DetailView):
    """تفاصيل العميل"""
    model = BusinessPartner
    template_name = 'base_data/partners/customer_detail.html'
    context_object_name = 'customer'
    permission_required = 'base_data.view_businesspartner'

    def get_queryset(self):
        return super().get_queryset().filter(partner_type__in=['customer', 'both'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل العميل: %s') % self.object.name

        # معلومات خاصة بالعميل
        context['credit_available'] = self.object.credit_limit - 0  # يجب حساب الرصيد الحالي

        context['breadcrumbs'].extend([
            {'title': _('العملاء'), 'url': reverse_lazy('base_data:customer_list')},
            {'title': str(self.object), 'active': True}
        ])

        return context


class CustomerCreateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                         SuccessMessageMixin, CreateView):
    """إنشاء عميل جديد"""
    model = BusinessPartner
    form_class = CustomerForm
    template_name = 'base_data/partners/customer_form.html'
    permission_required = 'base_data.add_businesspartner'
    success_message = _('تم إنشاء العميل بنجاح')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة عميل جديد')
        context['submit_text'] = _('حفظ')

        context['breadcrumbs'].extend([
            {'title': _('العملاء'), 'url': reverse_lazy('base_data:customer_list')},
            {'title': _('إضافة عميل جديد'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:customer_add')
        return reverse_lazy('base_data:customer_detail', kwargs={'pk': self.object.pk})


class CustomerUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                         SuccessMessageMixin, UpdateView):
    """تعديل العميل"""
    model = BusinessPartner
    form_class = CustomerForm
    template_name = 'base_data/partners/customer_form.html'
    permission_required = 'base_data.change_businesspartner'
    success_message = _('تم تعديل العميل بنجاح')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        return kwargs

    def get_queryset(self):
        return super().get_queryset().filter(partner_type__in=['customer', 'both'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل العميل: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        context['breadcrumbs'].extend([
            {'title': _('العملاء'), 'url': reverse_lazy('base_data:customer_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:customer_detail', args=[self.object.pk])},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:customer_detail', kwargs={'pk': self.object.pk})


# ============== Views الموردين ==============

class SupplierListView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, ListView):
    """قائمة الموردين"""
    model = BusinessPartner
    template_name = 'base_data/partners/supplier_list.html'
    context_object_name = 'suppliers'
    permission_required = 'base_data.view_businesspartner'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('قائمة الموردين')
        context['filter_form'] = PartnerFilterForm(self.request.GET)

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': _('الموردين'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            partner_type__in=['supplier', 'both']
        )

        # تطبيق الفلاتر
        filter_form = PartnerFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('search'):
                search = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(code__icontains=search) |
                    Q(name__icontains=search) |
                    Q(name_en__icontains=search) |
                    Q(phone__icontains=search) |
                    Q(mobile__icontains=search) |
                    Q(email__icontains=search)
                )

            if filter_form.cleaned_data.get('account_type'):
                queryset = queryset.filter(account_type=filter_form.cleaned_data['account_type'])

            if filter_form.cleaned_data.get('city'):
                queryset = queryset.filter(city__icontains=filter_form.cleaned_data['city'])

            if filter_form.cleaned_data.get('is_active'):
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('code', 'name')


class SupplierDetailView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, DetailView):
    """تفاصيل المورد"""
    model = BusinessPartner
    template_name = 'base_data/partners/supplier_detail.html'
    context_object_name = 'supplier'
    permission_required = 'base_data.view_businesspartner'

    def get_queryset(self):
        return super().get_queryset().filter(partner_type__in=['supplier', 'both'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل المورد: %s') % self.object.name

        # معلومات خاصة بالمورد
        context['rating_stars'] = range(self.object.rating)
        context['remaining_stars'] = range(5 - self.object.rating)

        context['breadcrumbs'].extend([
            {'title': _('الموردين'), 'url': reverse_lazy('base_data:supplier_list')},
            {'title': str(self.object), 'active': True}
        ])

        return context


class SupplierCreateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                         SuccessMessageMixin, CreateView):
    """إنشاء مورد جديد"""
    model = BusinessPartner
    form_class = SupplierForm
    template_name = 'base_data/partners/supplier_form.html'
    permission_required = 'base_data.add_businesspartner'
    success_message = _('تم إنشاء المورد بنجاح')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إضافة مورد جديد')
        context['submit_text'] = _('حفظ')

        context['breadcrumbs'].extend([
            {'title': _('الموردين'), 'url': reverse_lazy('base_data:supplier_list')},
            {'title': _('إضافة مورد جديد'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.company = self.request.user.company
        form.instance.created_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        if 'save_add_another' in self.request.POST:
            return reverse_lazy('base_data:supplier_add')
        return reverse_lazy('base_data:supplier_detail', kwargs={'pk': self.object.pk})


class SupplierUpdateView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin,
                         SuccessMessageMixin, UpdateView):
    """تعديل المورد"""
    model = BusinessPartner
    form_class = SupplierForm
    template_name = 'base_data/partners/supplier_form.html'
    permission_required = 'base_data.change_businesspartner'
    success_message = _('تم تعديل المورد بنجاح')

    def get_queryset(self):
        return super().get_queryset().filter(partner_type__in=['supplier', 'both'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل المورد: %s') % self.object.name
        context['submit_text'] = _('حفظ التعديلات')

        context['breadcrumbs'].extend([
            {'title': _('الموردين'), 'url': reverse_lazy('base_data:supplier_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:supplier_detail', args=[self.object.pk])},
            {'title': _('تعديل'), 'active': True}
        ])

        return context

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('base_data:supplier_detail', kwargs={'pk': self.object.pk})


# ============== Views معلومات الاتصال ==============

class ContactInfoManageView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, DetailView):
    """إدارة معلومات الاتصال للشريك"""
    model = BusinessPartner
    template_name = 'base_data/partners/contact_info.html'
    context_object_name = 'partner'
    permission_required = 'base_data.change_businesspartner'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('إدارة معلومات الاتصال: %s') % self.object.name
        context['contact_form'] = ContactInfoForm(instance=self.object)

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': str(self.object), 'url': reverse_lazy('base_data:partner_detail', args=[self.object.pk])},
            {'title': _('معلومات الاتصال'), 'active': True}
        ])

        return context

    def post(self, request, *args, **kwargs):
        """تحديث معلومات الاتصال"""
        self.object = self.get_object()
        form = ContactInfoForm(request.POST, instance=self.object)

        if form.is_valid():
            form.instance.updated_by = request.user
            form.save()

            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': _('تم تحديث معلومات الاتصال بنجاح')
                })
            else:
                messages.success(request, _('تم تحديث معلومات الاتصال بنجاح'))
        else:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            else:
                messages.error(request, _('خطأ في البيانات المدخلة'))

        return redirect('base_data:partner_contact_info', pk=self.object.pk)


# ============== Views التقارير والتصدير ==============

class PartnerReportView(LoginRequiredMixin, PermissionRequiredMixin, BasePartnerMixin, ListView):
    """تقرير الشركاء التجاريين"""
    model = BusinessPartner
    template_name = 'base_data/partners/partner_report.html'
    permission_required = 'base_data.view_businesspartner'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تقرير الشركاء التجاريين')
        context['filter_form'] = PartnerFilterForm(self.request.GET)

        context['breadcrumbs'].extend([
            {'title': _('الشركاء التجاريون'), 'url': reverse_lazy('base_data:partner_list')},
            {'title': _('التقارير'), 'active': True}
        ])

        return context

    def get_queryset(self):
        queryset = super().get_queryset().select_related('salesperson')

        # تطبيق الفلاتر
        filter_form = PartnerFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('partner_type'):
                queryset = queryset.filter(partner_type=filter_form.cleaned_data['partner_type'])
            if filter_form.cleaned_data.get('account_type'):
                queryset = queryset.filter(account_type=filter_form.cleaned_data['account_type'])
            if filter_form.cleaned_data.get('city'):
                queryset = queryset.filter(city__icontains=filter_form.cleaned_data['city'])
            if filter_form.cleaned_data.get('is_active'):
                queryset = queryset.filter(is_active=True)

        return queryset.order_by('partner_type', 'name')


class PartnerExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """تصدير الشركاء إلى Excel/CSV"""
    permission_required = 'base_data.view_businesspartner'

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('format', 'excel')

        # الحصول على البيانات
        partners = BusinessPartner.objects.filter(
            company=request.user.company
        ).select_related('salesperson').order_by('partner_type', 'code')

        if export_format == 'excel':
            return self.export_excel(partners)
        else:
            return self.export_csv(partners)

    def export_excel(self, partners):
        """تصدير إلى Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الشركاء التجاريون"

        # العناوين
        headers = [
            'نوع الشريك', 'الكود', 'الاسم', 'الاسم الإنجليزي', 'نوع الحساب',
            'الهاتف', 'الموبايل', 'البريد الإلكتروني', 'المدينة', 'حد الائتمان',
            'مندوب المبيعات', 'الحالة'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # البيانات
        for row, partner in enumerate(partners, 2):
            ws.cell(row=row, column=1, value=partner.get_partner_type_display())
            ws.cell(row=row, column=2, value=partner.code)
            ws.cell(row=row, column=3, value=partner.name)
            ws.cell(row=row, column=4, value=partner.name_en or '')
            ws.cell(row=row, column=5, value=partner.get_account_type_display())
            ws.cell(row=row, column=6, value=partner.phone or '')
            ws.cell(row=row, column=7, value=partner.mobile or '')
            ws.cell(row=row, column=8, value=partner.email or '')
            ws.cell(row=row, column=9, value=partner.city or '')
            ws.cell(row=row, column=10, value=float(partner.credit_limit or 0))
            ws.cell(row=row, column=11, value=str(partner.salesperson) if partner.salesperson else '')
            ws.cell(row=row, column=12, value='نشط' if partner.is_active else 'غير نشط')

        # حفظ الملف
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="partners_report.xlsx"'
        return response

    def export_csv(self, partners):
        """تصدير إلى CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="partners_report.csv"'
        response.write('\ufeff')  # BOM for Arabic support

        writer = csv.writer(response)

        # العناوين
        writer.writerow([
            'نوع الشريك', 'الكود', 'الاسم', 'الاسم الإنجليزي', 'نوع الحساب',
            'الهاتف', 'الموبايل', 'البريد الإلكتروني', 'المدينة', 'حد الائتمان',
            'مندوب المبيعات', 'الحالة'
        ])

        # البيانات
        for partner in partners:
            writer.writerow([
                partner.get_partner_type_display(),
                partner.code,
                partner.name,
                partner.name_en or '',
                partner.get_account_type_display(),
                partner.phone or '',
                partner.mobile or '',
                partner.email or '',
                partner.city or '',
                float(partner.credit_limit or 0),
                str(partner.salesperson) if partner.salesperson else '',
                'نشط' if partner.is_active else 'غير نشط'
            ])

        return response


# ============== DataTables AJAX Views ==============

class BusinessPartnerDataTableView(LoginRequiredMixin, BaseDatatableView):
    """DataTables AJAX للشركاء التجاريين"""
    model = BusinessPartner
    columns = ['code', 'name', 'partner_type', 'phone', 'email', 'city', 'credit_limit', 'is_active', 'actions']
    order_columns = ['code', 'name', 'partner_type', 'phone', 'email', 'city', 'credit_limit', 'is_active']
    max_display_length = 100

    def get_initial_queryset(self):
        return BusinessPartner.objects.filter(
            company=self.request.user.company
        ).select_related('salesperson')

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]')
        if search:
            qs = qs.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search) |
                Q(phone__icontains=search) |
                Q(mobile__icontains=search) |
                Q(email__icontains=search)
            )

        # فلاتر إضافية
        partner_type = self.request.GET.get('partner_type')
        if partner_type:
            qs = qs.filter(partner_type=partner_type)

        account_type = self.request.GET.get('account_type')
        if account_type:
            qs = qs.filter(account_type=account_type)

        city = self.request.GET.get('city')
        if city:
            qs = qs.filter(city__icontains=city)

        is_active = self.request.GET.get('is_active')
        if is_active == 'true':
            qs = qs.filter(is_active=True)
        elif is_active == 'false':
            qs = qs.filter(is_active=False)

        return qs

    def prepare_results(self, qs):
        json_data = []
        for partner in qs:
            actions = f'''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="{reverse_lazy('base_data:partner_detail', args=[partner.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:partner_edit', args=[partner.pk])}" 
                       class="btn btn-outline-warning btn-sm" title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                    <a href="{reverse_lazy('base_data:partner_delete', args=[partner.pk])}" 
                       class="btn btn-outline-danger btn-sm" title="حذف"
                       onclick="return confirm('هل أنت متأكد من الحذف؟')">
                        <i class="fas fa-trash"></i>
                    </a>
                </div>
            '''

            json_data.append([
                partner.code,
                partner.name,
                partner.get_partner_type_display(),
                partner.phone or partner.mobile or '',
                partner.email or '',
                partner.city or '',
                f"{partner.credit_limit:.2f}" if partner.credit_limit else '0.00',
                '<span class="badge badge-success">نشط</span>' if partner.is_active else
                '<span class="badge badge-warning">غير نشط</span>',
                actions
            ])
        return json_data


class CustomerDataTableView(BusinessPartnerDataTableView):
    """DataTables AJAX للعملاء"""

    def get_initial_queryset(self):
        return super().get_initial_queryset().filter(
            partner_type__in=['customer', 'both']
        )


class SupplierDataTableView(BusinessPartnerDataTableView):
    """DataTables AJAX للموردين"""

    def get_initial_queryset(self):
        return super().get_initial_queryset().filter(
            partner_type__in=['supplier', 'both']
        )


# ============== AJAX Views للنماذج ==============

class PartnerQuickAddView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """إضافة شريك سريع عبر AJAX"""
    permission_required = 'base_data.add_businesspartner'

    def post(self, request, *args, **kwargs):
        form = PartnerQuickAddForm(request.POST)
        if form.is_valid():
            partner = form.save(commit=False)
            partner.company = request.user.company
            partner.created_by = request.user
            partner.save()

            return JsonResponse({
                'success': True,
                'partner_id': partner.id,
                'partner_name': partner.name,
                'partner_code': partner.code,
                'partner_type': partner.get_partner_type_display(),
                'message': _('تم إضافة الشريك بنجاح')
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })


class PartnerSearchAjaxView(LoginRequiredMixin, View):
    """البحث عن الشركاء عبر AJAX"""

    def get(self, request, *args, **kwargs):
        term = request.GET.get('term', '')
        partner_type = request.GET.get('type', '')  # customer, supplier, both

        queryset = BusinessPartner.objects.filter(
            company=request.user.company,
            is_active=True
        ).filter(
            Q(code__icontains=term) |
            Q(name__icontains=term) |
            Q(name_en__icontains=term) |
            Q(phone__icontains=term) |
            Q(mobile__icontains=term)
        )

        # فلترة حسب نوع الشريك
        if partner_type == 'customer':
            queryset = queryset.filter(partner_type__in=['customer', 'both'])
        elif partner_type == 'supplier':
            queryset = queryset.filter(partner_type__in=['supplier', 'both'])

        queryset = queryset.order_by('name')[:20]

        results = []
        for partner in queryset:
            results.append({
                'id': partner.id,
                'text': f"{partner.code} - {partner.name}",
                'code': partner.code,
                'name': partner.name,
                'name_en': partner.name_en or '',
                'partner_type': partner.partner_type,
                'partner_type_display': partner.get_partner_type_display(),
                'phone': partner.phone or '',
                'mobile': partner.mobile or '',
                'email': partner.email or '',
                'credit_limit': float(partner.credit_limit or 0),
                'is_customer': partner.is_customer(),
                'is_supplier': partner.is_supplier()
            })

        return JsonResponse({'results': results})


class PartnerInfoAjaxView(LoginRequiredMixin, View):
    """الحصول على معلومات الشريك عبر AJAX"""

    def get(self, request, pk):
        try:
            partner = BusinessPartner.objects.select_related('salesperson').get(
                pk=pk,
                company=request.user.company
            )

            data = {
                'id': partner.id,
                'code': partner.code,
                'name': partner.name,
                'name_en': partner.name_en or '',
                'partner_type': partner.partner_type,
                'partner_type_display': partner.get_partner_type_display(),
                'account_type': partner.account_type,
                'account_type_display': partner.get_account_type_display(),
                'contact_person': partner.contact_person or '',
                'phone': partner.phone or '',
                'mobile': partner.mobile or '',
                'fax': partner.fax or '',
                'email': partner.email or '',
                'website': partner.website or '',
                'address': partner.address or '',
                'city': partner.city or '',
                'region': partner.region or '',
                'tax_number': partner.tax_number or '',
                'commercial_register': partner.commercial_register or '',
                'credit_limit': float(partner.credit_limit or 0),
                'credit_period': partner.credit_period,
                'discount_percentage': float(partner.discount_percentage or 0),
                'rating': partner.rating,
                'is_active': partner.is_active,
                'is_customer': partner.is_customer(),
                'is_supplier': partner.is_supplier(),
                'salesperson': str(partner.salesperson) if partner.salesperson else '',
                'notes': partner.notes or ''
            }

            return JsonResponse({'success': True, 'partner': data})

        except BusinessPartner.DoesNotExist:
            return JsonResponse({'success': False, 'message': _('الشريك غير موجود')})