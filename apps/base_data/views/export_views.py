# apps/base_data/views/export_views.py
"""
Views تصدير البيانات - Excel, PDF, CSV
Bootstrap 5 + RTL + تصدير متقدم
"""

import io
import csv
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views import View
from django.utils.translation import gettext_lazy as _
from django.template.loader import render_to_string
from django.conf import settings

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from ..models import Item, BusinessPartner, Warehouse, ItemCategory, UnitOfMeasure, WarehouseItem
from ..forms import ExportForm
from apps.core.mixins import CompanyMixin, AjaxResponseMixin


class BaseExportView(LoginRequiredMixin, CompanyMixin, View):
    """Base class للتصدير"""
    model = None
    fields = []
    filename_prefix = 'export'

    def get_queryset(self):
        """الحصول على البيانات للتصدير"""
        return self.model.objects.filter(company=self.request.user.company)

    def get_filename(self, format_type):
        """إنشاء اسم الملف"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{self.filename_prefix}_{timestamp}.{format_type}"

    def prepare_data(self, queryset):
        """تحضير البيانات للتصدير"""
        data = []
        for obj in queryset:
            row = {}
            for field in self.fields:
                if '__' in field:
                    # Related field
                    value = obj
                    for attr in field.split('__'):
                        value = getattr(value, attr, '') if value else ''
                else:
                    value = getattr(obj, field, '')

                # تنسيق القيم
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    value = _('نعم') if value else _('لا')
                elif value is None:
                    value = ''

                row[field] = str(value)
            data.append(row)
        return data


class ItemExportView(BaseExportView, PermissionRequiredMixin):
    """تصدير الأصناف"""
    model = Item
    permission_required = 'base_data.view_item'
    filename_prefix = 'items'
    fields = [
        'code', 'name', 'name_en', 'category__name', 'unit__name',
        'barcode', 'purchase_price', 'sale_price', 'tax_rate',
        'manufacturer', 'weight', 'minimum_quantity', 'maximum_quantity',
        'is_active', 'is_inactive', 'created_at'
    ]

    def get_queryset(self):
        queryset = super().get_queryset().select_related('category', 'unit')

        # تطبيق الفلاتر
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)

        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('code', 'name')

    def get(self, request):
        format_type = request.GET.get('format', 'xlsx')

        if format_type == 'xlsx':
            return self.export_excel()
        elif format_type == 'csv':
            return self.export_csv()
        elif format_type == 'pdf':
            return self.export_pdf()
        else:
            return JsonResponse({'error': _('صيغة غير مدعومة')}, status=400)

    def export_excel(self):
        """تصدير Excel"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        queryset = self.get_queryset()
        data = self.prepare_data(queryset)

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('الأصناف')

        # تنسيق الرأس
        headers = {
            'code': _('الكود'),
            'name': _('الاسم'),
            'name_en': _('الاسم الإنجليزي'),
            'category__name': _('التصنيف'),
            'unit__name': _('الوحدة'),
            'barcode': _('الباركود'),
            'purchase_price': _('سعر الشراء'),
            'sale_price': _('سعر البيع'),
            'tax_rate': _('معدل الضريبة'),
            'manufacturer': _('الشركة المصنعة'),
            'weight': _('الوزن'),
            'minimum_quantity': _('أقل كمية'),
            'maximum_quantity': _('أكبر كمية'),
            'is_active': _('نشط'),
            'is_inactive': _('غير فعال'),
            'created_at': _('تاريخ الإنشاء'),
        }

        # كتابة الرأس
        for col, field in enumerate(self.fields, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = headers.get(field, field)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # كتابة البيانات
        for row, item_data in enumerate(data, 2):
            for col, field in enumerate(self.fields, 1):
                ws.cell(row=row, column=col).value = item_data.get(field, '')

        # تنسيق العمودات
        for col in range(1, len(self.fields) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("xlsx")}"'
        return response

    def export_csv(self):
        """تصدير CSV"""
        queryset = self.get_queryset()
        data = self.prepare_data(queryset)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("csv")}"'

        # إضافة BOM للدعم الصحيح للعربية
        response.write('\ufeff')

        writer = csv.writer(response)

        # كتابة الرأس
        headers = [
            _('الكود'), _('الاسم'), _('الاسم الإنجليزي'), _('التصنيف'),
            _('الوحدة'), _('الباركود'), _('سعر الشراء'), _('سعر البيع'),
            _('معدل الضريبة'), _('الشركة المصنعة'), _('الوزن'),
            _('أقل كمية'), _('أكبر كمية'), _('نشط'), _('غير فعال'), _('تاريخ الإنشاء')
        ]
        writer.writerow(headers)

        # كتابة البيانات
        for item_data in data:
            row = [item_data.get(field, '') for field in self.fields]
            writer.writerow(row)

        return response

    def export_pdf(self):
        """تصدير PDF"""
        if not REPORTLAB_AVAILABLE:
            return JsonResponse({'error': _('مكتبة PDF غير متوفرة')}, status=500)

        queryset = self.get_queryset()[:100]  # تحديد عدد السجلات

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("pdf")}"'

        # إنشاء PDF
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []

        # العنوان
        title = Paragraph(f'<b>{_("قائمة الأصناف")}</b>', getSampleStyleSheet()['Title'])
        story.append(title)
        story.append(Spacer(1, 12))

        # البيانات
        table_data = []
        headers = [_('الكود'), _('الاسم'), _('التصنيف'), _('الوحدة'), _('سعر البيع')]
        table_data.append(headers)

        for item in queryset:
            row = [
                item.code or '',
                item.name[:30] + '...' if len(item.name) > 30 else item.name,
                item.category.name if item.category else '',
                item.unit.name if item.unit else '',
                str(item.sale_price) if item.sale_price else ''
            ]
            table_data.append(row)

        # إنشاء الجدول
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)
        doc.build(story)

        return response


class PartnerExportView(BaseExportView, PermissionRequiredMixin):
    """تصدير الشركاء التجاريين"""
    model = BusinessPartner
    permission_required = 'base_data.view_businesspartner'
    filename_prefix = 'partners'
    fields = [
        'code', 'name', 'name_en', 'partner_type', 'account_type',
        'contact_person', 'phone', 'mobile', 'email', 'city', 'region',
        'tax_number', 'commercial_register', 'credit_limit', 'is_active', 'created_at'
    ]

    def get_queryset(self):
        queryset = super().get_queryset()

        partner_type = self.request.GET.get('partner_type')
        if partner_type:
            if partner_type == 'both':
                queryset = queryset.filter(partner_type='both')
            else:
                queryset = queryset.filter(partner_type__in=[partner_type, 'both'])

        return queryset.order_by('name')

    def get(self, request):
        format_type = request.GET.get('format', 'xlsx')

        if format_type == 'xlsx':
            return self.export_excel()
        elif format_type == 'csv':
            return self.export_csv()
        else:
            return JsonResponse({'error': _('صيغة غير مدعومة')}, status=400)

    def export_excel(self):
        """تصدير Excel للشركاء"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        queryset = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('الشركاء التجاريون')

        headers = [
            _('الكود'), _('الاسم'), _('الاسم الإنجليزي'), _('النوع'), _('نوع الحساب'),
            _('جهة الاتصال'), _('الهاتف'), _('الموبايل'), _('البريد الإلكتروني'),
            _('المدينة'), _('المنطقة'), _('الرقم الضريبي'), _('السجل التجاري'),
            _('حد الائتمان'), _('نشط'), _('تاريخ الإنشاء')
        ]

        # كتابة الرأس
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # كتابة البيانات
        for row, partner in enumerate(queryset, 2):
            data = [
                partner.code or '',
                partner.name,
                partner.name_en or '',
                partner.get_partner_type_display(),
                partner.get_account_type_display(),
                partner.contact_person or '',
                partner.phone or '',
                partner.mobile or '',
                partner.email or '',
                partner.city or '',
                partner.region or '',
                partner.tax_number or '',
                partner.commercial_register or '',
                str(partner.credit_limit) if partner.credit_limit else '',
                _('نعم') if partner.is_active else _('لا'),
                partner.created_at.strftime('%Y-%m-%d') if partner.created_at else ''
            ]

            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col).value = value

        # تنسيق العمودات
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("xlsx")}"'
        return response

    def export_csv(self):
        """تصدير CSV للشركاء"""
        queryset = self.get_queryset()

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("csv")}"'
        response.write('\ufeff')

        writer = csv.writer(response)

        headers = [
            _('الكود'), _('الاسم'), _('الاسم الإنجليزي'), _('النوع'),
            _('جهة الاتصال'), _('الهاتف'), _('الموبايل'), _('البريد الإلكتروني'),
            _('المدينة'), _('المنطقة'), _('الرقم الضريبي'), _('حد الائتمان'), _('نشط')
        ]
        writer.writerow(headers)

        for partner in queryset:
            row = [
                partner.code or '',
                partner.name,
                partner.name_en or '',
                partner.get_partner_type_display(),
                partner.contact_person or '',
                partner.phone or '',
                partner.mobile or '',
                partner.email or '',
                partner.city or '',
                partner.region or '',
                partner.tax_number or '',
                str(partner.credit_limit) if partner.credit_limit else '',
                _('نعم') if partner.is_active else _('لا')
            ]
            writer.writerow(row)

        return response


class WarehouseExportView(BaseExportView, PermissionRequiredMixin):
    """تصدير المستودعات"""
    model = Warehouse
    permission_required = 'base_data.view_warehouse'
    filename_prefix = 'warehouses'
    fields = [
        'code', 'name', 'warehouse_type', 'location', 'branch__name',
        'keeper__get_full_name', 'is_active', 'created_at'
    ]

    def get(self, request):
        format_type = request.GET.get('format', 'xlsx')

        if format_type == 'xlsx':
            return self.export_excel()
        elif format_type == 'csv':
            return self.export_csv()
        else:
            return JsonResponse({'error': _('صيغة غير مدعومة')}, status=400)

    def export_excel(self):
        """تصدير Excel للمستودعات"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        queryset = self.get_queryset().select_related('branch', 'keeper')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('المستودعات')

        headers = [_('الكود'), _('الاسم'), _('النوع'), _('الموقع'), _('الفرع'), _('الأمين'), _('نشط'),
                   _('تاريخ الإنشاء')]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row, warehouse in enumerate(queryset, 2):
            data = [
                warehouse.code or '',
                warehouse.name,
                warehouse.get_warehouse_type_display(),
                warehouse.location or '',
                warehouse.branch.name if warehouse.branch else '',
                warehouse.keeper.get_full_name() if warehouse.keeper else '',
                _('نعم') if warehouse.is_active else _('لا'),
                warehouse.created_at.strftime('%Y-%m-%d') if warehouse.created_at else ''
            ]

            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col).value = value

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("xlsx")}"'
        return response


class StockReportExportView(BaseExportView, PermissionRequiredMixin):
    """تصدير تقرير المخزون"""
    permission_required = 'base_data.view_warehouse'
    filename_prefix = 'stock_report'

    def get(self, request):
        warehouse_id = request.GET.get('warehouse_id')
        show_zero = request.GET.get('show_zero', '0') == '1'
        format_type = request.GET.get('format', 'xlsx')

        # بناء الاستعلام
        queryset = WarehouseItem.objects.filter(
            warehouse__company=request.user.company
        ).select_related('item', 'warehouse', 'item__category', 'item__unit')

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        if not show_zero:
            queryset = queryset.filter(quantity__gt=0)

        queryset = queryset.order_by('warehouse__name', 'item__name')

        if format_type == 'xlsx':
            return self.export_stock_excel(queryset)
        elif format_type == 'csv':
            return self.export_stock_csv(queryset)
        else:
            return JsonResponse({'error': _('صيغة غير مدعومة')}, status=400)

    def export_stock_excel(self, queryset):
        """تصدير تقرير المخزون Excel"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': _('مكتبة Excel غير متوفرة')}, status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('تقرير المخزون')

        headers = [
            _('المستودع'), _('كود الصنف'), _('اسم الصنف'), _('التصنيف'),
            _('الوحدة'), _('الكمية'), _('متوسط التكلفة'), _('إجمالي القيمة')
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        total_value = 0
        for row, wi in enumerate(queryset, 2):
            value = wi.quantity * wi.average_cost
            total_value += value

            data = [
                wi.warehouse.name,
                wi.item.code or '',
                wi.item.name,
                wi.item.category.name if wi.item.category else '',
                wi.item.unit.name if wi.item.unit else '',
                float(wi.quantity),
                float(wi.average_cost),
                float(value)
            ]

            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col).value = val

        # إضافة الإجمالي
        total_row = queryset.count() + 2
        ws.cell(row=total_row, column=7).value = _('الإجمالي:')
        ws.cell(row=total_row, column=8).value = float(total_value)
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8).font = Font(bold=True)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_filename("xlsx")}"'
        return response


class CustomExportView(LoginRequiredMixin, CompanyMixin, AjaxResponseMixin, View):
    """تصدير مخصص حسب الطلب"""

    def post(self, request):
        model_name = request.POST.get('model')
        format_type = request.POST.get('format', 'xlsx')
        selected_fields = request.POST.getlist('fields[]')
        filters = request.POST.dict()

        # تحديد النموذج
        model_map = {
            'item': Item,
            'partner': BusinessPartner,
            'warehouse': Warehouse,
            'category': ItemCategory,
            'unit': UnitOfMeasure
        }

        model = model_map.get(model_name)
        if not model:
            return JsonResponse({'error': _('نموذج غير مدعوم')}, status=400)

        # التحقق من الصلاحيات
        permission_map = {
            'item': 'base_data.view_item',
            'partner': 'base_data.view_businesspartner',
            'warehouse': 'base_data.view_warehouse',
            'category': 'base_data.view_itemcategory',
            'unit': 'base_data.view_unitofmeasure'
        }

        required_permission = permission_map.get(model_name)
        if required_permission and not request.user.has_perm(required_permission):
            return JsonResponse({'error': _('ليس لديك صلاحية')}, status=403)

        # بناء الاستعلام
        queryset = model.objects.filter(company=request.user.company)

        # تطبيق الفلاتر (يمكن توسيعها)
        if filters.get('is_active') == '1':
            queryset = queryset.filter(is_active=True)
        elif filters.get('is_active') == '0':
            queryset = queryset.filter(is_active=False)

        # التصدير
        try:
            if format_type == 'json':
                return self.export_json(queryset, selected_fields)
            else:
                return JsonResponse({'error': _('صيغة غير مدعومة')}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    def export_json(self, queryset, fields):
        """تصدير JSON"""
        data = []
        for obj in queryset[:1000]:  # تحديد العدد
            row = {}
            for field in fields:
                if hasattr(obj, field):
                    value = getattr(obj, field)
                    if hasattr(value, 'strftime'):
                        value = value.isoformat()
                    elif isinstance(value, bool):
                        value = value
                    elif value is None:
                        value = None
                    else:
                        value = str(value)
                    row[field] = value
            data.append(row)

        response = JsonResponse({
            'success': True,
            'data': data,
            'count': len(data)
        })

        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response