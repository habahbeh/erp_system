# apps/purchases/views/order_views.py
"""
Views for Purchase Orders
Complete CRUD operations + Approval Workflow + Convert to Invoice
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from django.core.exceptions import PermissionDenied
from decimal import Decimal
from datetime import datetime, date

from ..models import PurchaseOrder, PurchaseOrderItem, PurchaseRequest
from ..forms import (
    PurchaseOrderForm,
    PurchaseOrderItemForm,
    PurchaseOrderItemFormSet,
    PurchaseOrderFilterForm,
)
from apps.core.models import BusinessPartner, Item


class PurchaseOrderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة أوامر الشراء"""
    model = PurchaseOrder
    template_name = 'purchases/orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 50
    permission_required = 'purchases.view_purchaseorder'

    def get_queryset(self):
        queryset = PurchaseOrder.objects.filter(
            company=self.request.current_company
        ).select_related(
            'supplier', 'warehouse', 'currency', 'purchase_request',
            'requested_by', 'approved_by', 'created_by'
        ).prefetch_related('lines').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(purchase_request__number__icontains=search)
            )

        # فلترة حسب المورد
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        # فلترة حسب الحالة
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('أوامر الشراء')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('أوامر الشراء'), 'url': ''},
        ]

        # الإحصائيات
        orders = self.get_queryset()
        context['stats'] = {
            'total_count': orders.count(),
            'draft_count': orders.filter(status='draft').count(),
            'pending_count': orders.filter(status='pending_approval').count(),
            'approved_count': orders.filter(status='approved').count(),
            'sent_count': orders.filter(status='sent').count(),
            'completed_count': orders.filter(status='completed').count(),
            'total_amount': orders.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0.000'),
        }

        # قائمة الموردين للفلترة
        context['suppliers'] = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['supplier', 'both'],
            is_active=True
        ).order_by('name')

        # قائمة المستودعات للفلترة
        from apps.core.models import Warehouse
        context['warehouses'] = Warehouse.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('name')

        # Filter form
        context['filter_form'] = PurchaseOrderFilterForm(
            self.request.GET,
            company=self.request.current_company
        )

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('purchases.add_purchaseorder')
        context['can_edit'] = self.request.user.has_perm('purchases.change_purchaseorder')
        context['can_delete'] = self.request.user.has_perm('purchases.delete_purchaseorder')
        context['can_approve'] = self.request.user.has_perm('purchases.change_purchaseorder')

        return context


class PurchaseOrderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل أمر الشراء"""
    model = PurchaseOrder
    template_name = 'purchases/orders/order_detail.html'
    context_object_name = 'order'
    permission_required = 'purchases.view_purchaseorder'

    def get_queryset(self):
        return PurchaseOrder.objects.filter(
            company=self.request.current_company
        ).select_related(
            'supplier', 'warehouse', 'currency', 'purchase_request',
            'requested_by', 'approved_by', 'created_by'
        ).prefetch_related('lines__item')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        order = self.object
        context['title'] = f'{_("أمر شراء")} {order.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('أوامر الشراء'), 'url': reverse('purchases:order_list')},
            {'title': order.number, 'url': ''},
        ]

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('purchases.change_purchaseorder') and
            order.status in ['draft', 'rejected']
        )
        context['can_delete'] = (
            self.request.user.has_perm('purchases.delete_purchaseorder') and
            order.status == 'draft'
        )
        context['can_approve'] = (
            self.request.user.has_perm('purchases.change_purchaseorder') and
            order.status == 'pending_approval'
        )
        context['can_reject'] = (
            self.request.user.has_perm('purchases.change_purchaseorder') and
            order.status == 'pending_approval'
        )
        context['can_send'] = (
            self.request.user.has_perm('purchases.change_purchaseorder') and
            order.status == 'approved'
        )
        context['can_convert'] = (
            self.request.user.has_perm('purchases.add_purchaseinvoice') and
            order.status in ['sent', 'approved', 'partial'] and
            not order.is_invoiced
        )

        return context


class PurchaseOrderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء أمر شراء جديد"""
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'purchases/orders/order_form.html'
    permission_required = 'purchases.add_purchaseorder'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة أمر شراء')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('أوامر الشراء'), 'url': reverse('purchases:order_list')},
            {'title': _('إضافة أمر'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseOrderItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseOrderItemFormSet(
                instance=self.object,
                company=self.request.current_company
            )

        # بيانات للجافاسكربت - AJAX Live Search
        context['use_live_search'] = True  # تفعيل البحث المباشر
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        # إذا كان من طلب شراء
        request_id = self.request.GET.get('from_request')
        if request_id:
            try:
                purchase_request = PurchaseRequest.objects.get(
                    pk=request_id,
                    company=self.request.current_company
                )
                context['purchase_request'] = purchase_request
            except PurchaseRequest.DoesNotExist:
                pass

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        # ربط الأمر بالشركة والفرع
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # حساب المجاميع
            self.object.calculate_total()
            self.object.save()

            messages.success(
                self.request,
                _('تم إضافة أمر الشراء بنجاح')
            )
            return redirect('purchases:order_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseOrderUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل أمر شراء"""
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'purchases/orders/order_form.html'
    permission_required = 'purchases.change_purchaseorder'

    def get_queryset(self):
        # يمكن تعديل الأمر في حالة مسودة أو مرفوض فقط
        return PurchaseOrder.objects.filter(
            company=self.request.current_company,
            status__in=['draft', 'rejected']
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        order = self.object
        context['title'] = f'{_("تعديل أمر")} {order.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('أوامر الشراء'), 'url': reverse('purchases:order_list')},
            {'title': order.number, 'url': reverse('purchases:order_detail', kwargs={'pk': order.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseOrderItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseOrderItemFormSet(
                instance=self.object,
                company=self.request.current_company
            )

        # بيانات للجافاسكربت - AJAX Live Search
        context['use_live_search'] = True  # تفعيل البحث المباشر
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # إعادة حساب المجاميع
            self.object.calculate_total()
            self.object.save()

            messages.success(
                self.request,
                _('تم تعديل أمر الشراء بنجاح')
            )
            return redirect('purchases:order_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseOrderDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف أمر شراء"""
    model = PurchaseOrder
    template_name = 'purchases/orders/order_confirm_delete.html'
    success_url = reverse_lazy('purchases:order_list')
    permission_required = 'purchases.delete_purchaseorder'

    def get_queryset(self):
        # يمكن حذف الأمر في حالة مسودة فقط
        return PurchaseOrder.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف أمر الشراء بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# Workflow Actions
# ============================================================================

@login_required
@permission_required('purchases.change_purchaseorder', raise_exception=True)
@transaction.atomic
def submit_for_approval(request, pk):
    """إرسال أمر الشراء للاعتماد"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status != 'draft':
        messages.error(request, _('يمكن إرسال الأمر للاعتماد في حالة مسودة فقط'))
        return redirect('purchases:order_detail', pk=pk)

    try:
        order.submit_for_approval()
        messages.success(
            request,
            _('تم إرسال أمر الشراء للاعتماد بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الإرسال")}: {str(e)}')

    return redirect('purchases:order_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaseorder', raise_exception=True)
@transaction.atomic
def approve_order(request, pk):
    """اعتماد أمر الشراء"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status != 'pending_approval':
        messages.error(request, _('يمكن اعتماد الأمر في حالة بانتظار الموافقة فقط'))
        return redirect('purchases:order_detail', pk=pk)

    try:
        order.approve(user=request.user)
        messages.success(
            request,
            _('تم اعتماد أمر الشراء بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الاعتماد")}: {str(e)}')

    return redirect('purchases:order_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaseorder', raise_exception=True)
@transaction.atomic
def reject_order(request, pk):
    """رفض أمر الشراء"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status != 'pending_approval':
        messages.error(request, _('يمكن رفض الأمر في حالة بانتظار الموافقة فقط'))
        return redirect('purchases:order_detail', pk=pk)

    # الحصول على سبب الرفض
    rejection_reason = request.POST.get('rejection_reason', '')

    try:
        order.reject(user=request.user, reason=rejection_reason)
        messages.success(
            request,
            _('تم رفض أمر الشراء')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الرفض")}: {str(e)}')

    return redirect('purchases:order_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaseorder', raise_exception=True)
@transaction.atomic
def send_to_supplier(request, pk):
    """إرسال أمر الشراء للمورد"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status not in ['approved', 'sent']:
        messages.error(request, _('يجب أن يكون الأمر معتمد لإرساله للمورد'))
        return redirect('purchases:order_detail', pk=pk)

    try:
        order.send_to_supplier()
        messages.success(
            request,
            _('تم تعليم الأمر كمرسل للمورد بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الإرسال")}: {str(e)}')

    return redirect('purchases:order_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaseorder', raise_exception=True)
@transaction.atomic
def cancel_order(request, pk):
    """إلغاء أمر الشراء"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status in ['completed', 'cancelled']:
        messages.error(request, _('لا يمكن إلغاء أمر مكتمل أو ملغي'))
        return redirect('purchases:order_detail', pk=pk)

    try:
        order.cancel()
        messages.success(
            request,
            _('تم إلغاء أمر الشراء بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الإلغاء")}: {str(e)}')

    return redirect('purchases:order_detail', pk=pk)


@login_required
@permission_required('purchases.add_purchaseinvoice', raise_exception=True)
@transaction.atomic
def convert_to_invoice(request, pk):
    """تحويل أمر الشراء إلى فاتورة"""
    order = get_object_or_404(
        PurchaseOrder,
        pk=pk,
        company=request.current_company
    )

    if order.status not in ['sent', 'approved', 'partial']:
        messages.error(
            request,
            _('يجب أن يكون الأمر معتمد أو مرسل لتحويله إلى فاتورة')
        )
        return redirect('purchases:order_detail', pk=pk)

    try:
        invoice = order.convert_to_invoice(user=request.user)
        messages.success(
            request,
            _('تم تحويل أمر الشراء إلى فاتورة بنجاح')
        )
        return redirect('purchases:invoice_detail', pk=invoice.pk)
    except Exception as e:
        messages.error(request, f'{_("خطأ في التحويل")}: {str(e)}')
        return redirect('purchases:order_detail', pk=pk)


# ============================================================================
# AJAX & Export
# ============================================================================

@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def order_datatable_ajax(request):
    """AJAX endpoint for DataTables"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    queryset = PurchaseOrder.objects.filter(
        company=request.current_company
    ).select_related('supplier', 'currency', 'warehouse')

    # البحث
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(supplier__name__icontains=search_value)
        )

    # الفلاتر
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    supplier = request.GET.get('supplier')
    if supplier:
        queryset = queryset.filter(supplier_id=supplier)

    warehouse = request.GET.get('warehouse')
    if warehouse:
        queryset = queryset.filter(warehouse_id=warehouse)

    date_from = request.GET.get('date_from')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)

    date_to = request.GET.get('date_to')
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # العدد الكلي
    total_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    queryset = queryset[start:start + length]

    # البيانات
    data = []
    for order in queryset:
        data.append({
            'number': order.number,
            'date': order.date.strftime('%Y-%m-%d'),
            'supplier': order.supplier.name,
            'warehouse': order.warehouse.name if order.warehouse else '-',
            'expected_delivery_date': order.expected_delivery_date.strftime('%Y-%m-%d') if order.expected_delivery_date else None,
            'total_amount': float(order.total_amount),
            'status': order.get_status_display(),
            'status_code': order.status,
            'pk': order.pk,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def export_orders_excel(request):
    """تصدير أوامر الشراء إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    # استرجاع البيانات
    queryset = PurchaseOrder.objects.filter(
        company=request.current_company
    ).select_related(
        'supplier', 'currency'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر من GET parameters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "أوامر الشراء"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم الأمر', 'التاريخ', 'المورد', 'تاريخ التسليم المتوقع',
        'الإجمالي', 'الضريبة', 'الإجمالي مع الضريبة',
        'العملة', 'الحالة'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, order in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=order.number).border = border
        ws.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=order.supplier.name).border = border
        ws.cell(row=row_num, column=4, value=order.expected_delivery_date.strftime('%Y-%m-%d') if order.expected_delivery_date else '').border = border
        ws.cell(row=row_num, column=5, value=float(order.total_amount)).border = border
        ws.cell(row=row_num, column=6, value='').border = border  # Tax amount not stored in PurchaseOrder
        ws.cell(row=row_num, column=7, value=float(order.total_amount)).border = border
        ws.cell(row=row_num, column=8, value=order.currency.code).border = border
        ws.cell(row=row_num, column=9, value=order.get_status_display()).border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 15

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"purchase_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ============================================
# AJAX Endpoints للتحسينات الجديدة
# ============================================

@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def get_supplier_item_price_ajax(request):
    """
    جلب آخر سعر شراء من المورد لمادة معينة
    يُستخدم لملء السعر تلقائياً في أوامر الشراء
    """
    supplier_id = request.GET.get('supplier_id')
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not supplier_id or not item_id:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        from apps.core.models import PartnerItemPrice

        # البحث عن آخر سعر شراء
        filter_params = {
            'company': request.current_company,
            'partner_id': supplier_id,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        price_record = PartnerItemPrice.objects.filter(
            **filter_params
        ).first()

        if price_record and price_record.last_purchase_price:
            return JsonResponse({
                'success': True,
                'has_price': True,
                'last_price': str(price_record.last_purchase_price),
                'last_date': price_record.last_purchase_date.strftime('%Y-%m-%d') if price_record.last_purchase_date else None,
                'last_quantity': str(price_record.last_purchase_quantity) if price_record.last_purchase_quantity else None,
                'total_purchased': str(price_record.total_purchased_quantity),
            })
        else:
            return JsonResponse({
                'success': True,
                'has_price': False,
                'message': 'لا توجد مشتريات سابقة من هذا المورد لهذه المادة'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def get_item_stock_multi_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة من جميع الفروع
    يعرض الكميات المتوفرة في كل مخزن مع تفاصيل الفرع
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث عن الكميات في جميع المخازن
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        stock_records = ItemStock.objects.filter(
            **filter_params
        ).select_related(
            'warehouse', 'warehouse__branch'
        ).order_by('warehouse__branch__name', 'warehouse__name')

        # تجميع البيانات
        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            branches_data.append({
                'branch_name': stock.warehouse.branch.name,
                'warehouse_name': stock.warehouse.name,
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost),
            })

        return JsonResponse({
            'success': True,
            'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def get_item_stock_current_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة في الفرع الحالي فقط
    يستخدم للعرض السريع في جدول أمر الشراء
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث في الفرع الحالي
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
            'warehouse__branch': request.current_branch,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        # حساب الإجمالي في الفرع الحالي
        stock_aggregate = ItemStock.objects.filter(
            **filter_params
        ).aggregate(
            total_qty=Sum('quantity'),
            total_reserved=Sum('reserved_quantity')
        )

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True,
            'quantity': str(total_qty),
            'reserved': str(total_reserved),
            'available': str(available),
            'has_stock': total_qty > 0,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseorder', raise_exception=True)
def item_search_ajax(request):
    """
    AJAX Live Search للمواد
    يُستخدم للبحث المباشر بدلاً من تحميل جميع المواد
    """
    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({
            'success': False,
            'message': 'يرجى إدخال حرفين على الأقل للبحث'
        })

    try:
        items = Item.objects.filter(
            company=request.current_company,
            is_active=True
        ).filter(
            Q(name__icontains=term) |
            Q(code__icontains=term) |
            Q(barcode__icontains=term)
        ).annotate(
            # كمية المخزون في الفرع الحالي
            current_branch_stock=Sum(
                'stock_records__quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)
            ),
            # الكمية المحجوزة في الفرع الحالي
            current_branch_reserved=Sum(
                'stock_records__reserved_quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)
            ),
            # إجمالي المخزون في كل الفروع
            total_stock=Sum('stock_records__quantity'),
        ).select_related(
            'category', 'base_uom'
        )[:limit]

        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'name': item.name,
                'code': item.code,
                'barcode': item.barcode or '',
                'category_name': item.category.name if item.category else '',
                'tax_rate': str(item.tax_rate),
                'base_uom_name': item.base_uom.name if item.base_uom else '',
                'base_uom_code': item.base_uom.code if item.base_uom else '',
                'current_branch_stock': str(item.current_branch_stock or 0),
                'current_branch_reserved': str(item.current_branch_reserved or 0),
                'total_stock': str(item.total_stock or 0),
            })

        return JsonResponse({
            'success': True,
            'items': items_data,
            'count': len(items_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.add_purchaseorder', raise_exception=True)
@transaction.atomic
def save_order_draft_ajax(request):
    """
    حفظ أمر الشراء كمسودة (Auto-save)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        from django.utils import timezone

        order_id = request.POST.get('order_id')

        # بيانات أمر الشراء الأساسية
        order_data = {
            'supplier_id': request.POST.get('supplier'),
            'warehouse_id': request.POST.get('warehouse'),
            'date': request.POST.get('date'),
            'expected_delivery_date': request.POST.get('expected_delivery_date'),
            'number': request.POST.get('number'),
            'reference': request.POST.get('reference'),
            'notes': request.POST.get('notes'),
        }

        # التحقق من البيانات الأساسية
        if not order_data['supplier_id'] or not order_data['warehouse_id']:
            return JsonResponse({
                'success': False,
                'message': 'يرجى اختيار المورد والمخزن'
            })

        # حفظ أو تحديث أمر الشراء
        if order_id:
            order = get_object_or_404(
                PurchaseOrder,
                pk=order_id,
                company=request.current_company,
                status='draft'
            )
            for key, value in order_data.items():
                if value:
                    setattr(order, key.replace('_id', ''), value)
            order.save()
        else:
            order = PurchaseOrder.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                created_by=request.user,
                status='draft',
                **order_data
            )

        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_number': order.number,
            'message': 'تم حفظ المسودة بنجاح',
            'saved_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
