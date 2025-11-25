# apps/purchases/views/invoice_views.py
"""
Views for Purchase Invoices
Complete CRUD operations + Posting functionality
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from decimal import Decimal
from datetime import datetime, date
import json

from ..models import PurchaseInvoice, PurchaseInvoiceItem
from ..forms import (
    PurchaseInvoiceForm,
    PurchaseInvoiceItemForm,
    PurchaseInvoiceItemFormSet,
)
from apps.core.models import BusinessPartner, Item


class PurchaseInvoiceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة فواتير المشتريات"""
    model = PurchaseInvoice
    template_name = 'purchases/invoices/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 50
    permission_required = 'purchases.view_purchaseinvoice'

    def get_queryset(self):
        queryset = PurchaseInvoice.objects.filter(
            company=self.request.current_company
        ).select_related(
            'supplier', 'warehouse', 'currency', 'payment_method', 'branch'
        ).prefetch_related('lines').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(receipt_number__icontains=search) |
                Q(supplier_invoice_number__icontains=search)
            )

        # فلترة حسب المورد
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        # فلترة حسب المستودع
        warehouse_id = self.request.GET.get('warehouse')
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        # فلترة حسب الحالة
        is_posted = self.request.GET.get('is_posted')
        if is_posted == '1':
            queryset = queryset.filter(is_posted=True)
        elif is_posted == '0':
            queryset = queryset.filter(is_posted=False)

        # فلترة حسب نوع الفاتورة
        invoice_type = self.request.GET.get('invoice_type')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['title'] = _('فواتير المشتريات')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('فواتير المشتريات'), 'url': ''},
        ]

        # الإحصائيات
        invoices = self.get_queryset()
        context['stats'] = {
            'total_count': invoices.count(),
            'posted_count': invoices.filter(is_posted=True).count(),
            'draft_count': invoices.filter(is_posted=False).count(),
            'purchase_count': invoices.filter(invoice_type='purchase').count(),
            'return_count': invoices.filter(invoice_type='return').count(),
            'total_amount': invoices.aggregate(
                total=Sum('total_with_tax')
            )['total'] or Decimal('0.000'),
        }

        # قائمة الموردين للفلترة
        context['suppliers'] = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['supplier', 'both'],
            is_active=True
        ).order_by('name')

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('purchases.add_purchaseinvoice')
        context['can_edit'] = self.request.user.has_perm('purchases.change_purchaseinvoice')
        context['can_delete'] = self.request.user.has_perm('purchases.delete_purchaseinvoice')
        context['can_post'] = self.request.user.has_perm('purchases.add_purchaseinvoice')

        return context


class PurchaseInvoiceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل فاتورة المشتريات"""
    model = PurchaseInvoice
    template_name = 'purchases/invoices/invoice_detail.html'
    context_object_name = 'invoice'
    permission_required = 'purchases.view_purchaseinvoice'

    def get_queryset(self):
        return PurchaseInvoice.objects.filter(
            company=self.request.current_company
        ).select_related(
            'supplier', 'warehouse', 'currency', 'payment_method',
            'branch', 'journal_entry', 'discount_account', 'supplier_account'
        ).prefetch_related('lines__item', 'lines__unit')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        invoice = self.object
        context['title'] = f'{_("فاتورة مشتريات")} {invoice.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('فواتير المشتريات'), 'url': reverse('purchases:invoice_list')},
            {'title': invoice.number, 'url': ''},
        ]

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('purchases.change_purchaseinvoice') and
            not invoice.is_posted
        )
        context['can_delete'] = (
            self.request.user.has_perm('purchases.delete_purchaseinvoice') and
            not invoice.is_posted
        )
        context['can_post'] = (
            self.request.user.has_perm('purchases.add_purchaseinvoice') and
            not invoice.is_posted
        )
        context['can_unpost'] = (
            self.request.user.has_perm('purchases.change_purchaseinvoice') and
            invoice.is_posted
        )

        return context


class PurchaseInvoiceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء فاتورة مشتريات جديدة"""
    model = PurchaseInvoice
    form_class = PurchaseInvoiceForm
    template_name = 'purchases/invoices/invoice_form.html'
    permission_required = 'purchases.add_purchaseinvoice'

    def post(self, request, *args, **kwargs):
        """Override post to debug the entire flow"""
        import sys
        sys.stdout.write("\n" + "="*100 + "\n")
        sys.stdout.write("🔥 POST METHOD CALLED - Starting invoice creation\n")
        sys.stdout.write(f"POST keys: {list(request.POST.keys())[:10]}...\n")
        sys.stdout.write(f"Number of POST items: {len(request.POST)}\n")

        # Check if there are any invoice items
        item_count = 0
        for key in request.POST.keys():
            if 'items-' in key and '-item' in key and '-DELETE' not in key:
                item_count += 1
                sys.stdout.write(f"Found item field: {key} = {request.POST.get(key)}\n")

        sys.stdout.write(f"Total item-related fields: {item_count}\n")
        sys.stdout.write("="*100 + "\n")
        sys.stdout.flush()

        # Call parent post method
        return super().post(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة فاتورة مشتريات')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('فواتير المشتريات'), 'url': reverse('purchases:invoice_list')},
            {'title': _('إضافة فاتورة'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseInvoiceItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )
        else:
            context['formset'] = PurchaseInvoiceItemFormSet(
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )

        # ✅ PERFORMANCE IMPROVEMENT: No longer loading all items
        # Now using AJAX Live Search instead
        context['items_data'] = []  # Empty - will use AJAX search
        context['use_live_search'] = True  # Flag for JavaScript

        return context

    @transaction.atomic
    def form_valid(self, form):
        import sys
        sys.stdout.write("\n" + "=" * 80 + "\n")
        sys.stdout.write("🔍 form_valid called!\n")
        sys.stdout.write(f"POST data keys: {list(self.request.POST.keys())}\n")
        sys.stdout.flush()

        context = self.get_context_data()
        formset = context['formset']

        # ربط الفاتورة بالشركة والفرع
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user

        sys.stdout.write(f"✅ Form is valid: {form.is_valid()}\n")
        sys.stdout.write(f"📋 Formset is valid: {formset.is_valid()}\n")
        sys.stdout.flush()

        if not formset.is_valid():
            sys.stdout.write(f"❌ Formset errors: {formset.errors}\n")
            sys.stdout.write(f"❌ Non-form errors: {formset.non_form_errors()}\n")
            sys.stdout.flush()

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # حساب المجاميع
            self.object.calculate_totals()
            self.object.save()

            messages.success(
                self.request,
                _('تم إضافة فاتورة المشتريات بنجاح')
            )
            return redirect('purchases:invoice_detail', pk=self.object.pk)
        else:
            # عرض أخطاء الـ formset بوضوح
            error_messages = []

            # أخطاء النموذج الرئيسي
            if form.errors:
                error_messages.append(f"أخطاء النموذج الرئيسي: {form.errors.as_text()}")

            # أخطاء الـ formset
            if formset.non_form_errors():
                error_messages.append(f"أخطاء عامة: {formset.non_form_errors()}")

            # أخطاء كل صف
            for i, form_item in enumerate(formset):
                if form_item.errors:
                    error_messages.append(f"الصف {i+1}: {form_item.errors.as_text()}")

            # عرض جميع الأخطاء
            for msg in error_messages:
                messages.error(self.request, msg)

            if not error_messages:
                messages.error(
                    self.request,
                    _('يرجى تصحيح الأخطاء في النموذج')
                )

            return self.render_to_response(self.get_context_data(form=form))

    def form_invalid(self, form):
        """Called when main form validation fails"""
        import sys
        sys.stdout.write("\n" + "=" * 80 + "\n")
        sys.stdout.write("❌ form_invalid called - Main form has validation errors!\n")
        sys.stdout.write(f"Form errors: {dict(form.errors)}\n")
        sys.stdout.flush()

        context = self.get_context_data(form=form)
        formset = context['formset']

        sys.stdout.write(f"Formset is valid: {formset.is_valid()}\n")
        if not formset.is_valid():
            sys.stdout.write(f"Formset errors: {formset.errors}\n")
            sys.stdout.write(f"Formset non-form errors: {formset.non_form_errors()}\n")
        sys.stdout.write("=" * 80 + "\n")
        sys.stdout.flush()

        return super().form_invalid(form)


class PurchaseInvoiceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل فاتورة مشتريات"""
    model = PurchaseInvoice
    form_class = PurchaseInvoiceForm
    template_name = 'purchases/invoices/invoice_form.html'
    permission_required = 'purchases.change_purchaseinvoice'

    def get_queryset(self):
        return PurchaseInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=False  # لا يمكن تعديل فاتورة مرحلة
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        invoice = self.object
        context['title'] = f'{_("تعديل فاتورة")} {invoice.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('فواتير المشتريات'), 'url': reverse('purchases:invoice_list')},
            {'title': invoice.number, 'url': reverse('purchases:invoice_detail', kwargs={'pk': invoice.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseInvoiceItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseInvoiceItemFormSet(
                instance=self.object,
                company=self.request.current_company
            )

        # ✅ PERFORMANCE IMPROVEMENT: No longer loading all items
        # Now using AJAX Live Search instead
        context['items_data'] = []  # Empty - will use AJAX search
        context['use_live_search'] = True  # Flag for JavaScript

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # إعادة حساب المجاميع
            self.object.calculate_totals()
            self.object.save()

            messages.success(
                self.request,
                _('تم تعديل فاتورة المشتريات بنجاح')
            )
            return redirect('purchases:invoice_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseInvoiceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف فاتورة مشتريات"""
    model = PurchaseInvoice
    template_name = 'purchases/invoices/invoice_confirm_delete.html'
    success_url = reverse_lazy('purchases:invoice_list')
    permission_required = 'purchases.delete_purchaseinvoice'

    def get_queryset(self):
        return PurchaseInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=False  # لا يمكن حذف فاتورة مرحلة
        )

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف فاتورة المشتريات بنجاح'))
        return super().delete(request, *args, **kwargs)


@login_required
@permission_required('purchases.add_purchaseinvoice', raise_exception=True)
@transaction.atomic
def post_invoice(request, pk):
    """ترحيل فاتورة المشتريات"""
    invoice = get_object_or_404(
        PurchaseInvoice,
        pk=pk,
        company=request.current_company
    )

    if invoice.is_posted:
        messages.error(request, _('الفاتورة مرحلة مسبقاً'))
        return redirect('purchases:invoice_detail', pk=pk)

    try:
        invoice.post(user=request.user)
        messages.success(
            request,
            _('تم ترحيل الفاتورة وإنشاء سند الإدخال والقيد المحاسبي بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الترحيل")}: {str(e)}')

    return redirect('purchases:invoice_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaseinvoice', raise_exception=True)
@transaction.atomic
def unpost_invoice(request, pk):
    """إلغاء ترحيل فاتورة المشتريات"""
    invoice = get_object_or_404(
        PurchaseInvoice,
        pk=pk,
        company=request.current_company
    )

    if not invoice.is_posted:
        messages.error(request, _('الفاتورة غير مرحلة'))
        return redirect('purchases:invoice_detail', pk=pk)

    try:
        invoice.unpost()
        messages.success(
            request,
            _('تم إلغاء ترحيل الفاتورة وحذف سند الإدخال والقيد المحاسبي بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في إلغاء الترحيل")}: {str(e)}')

    return redirect('purchases:invoice_detail', pk=pk)


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def invoice_datatable_ajax(request):
    """AJAX endpoint for DataTables"""
    from django.urls import reverse

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # الفلاتر من الواجهة
    invoice_type = request.GET.get('invoice_type', '')
    posted = request.GET.get('posted', '')
    supplier_id = request.GET.get('supplier', '')
    warehouse_id = request.GET.get('warehouse', '')
    search_filter = request.GET.get('search_filter', '')

    queryset = PurchaseInvoice.objects.filter(
        company=request.current_company
    ).select_related('supplier', 'warehouse', 'currency')

    # تطبيق الفلاتر
    if invoice_type:
        queryset = queryset.filter(invoice_type=invoice_type)

    if posted == '1':
        queryset = queryset.filter(is_posted=True)
    elif posted == '0':
        queryset = queryset.filter(is_posted=False)

    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    if warehouse_id:
        queryset = queryset.filter(warehouse_id=warehouse_id)

    # البحث العام
    if search_filter:
        queryset = queryset.filter(
            Q(number__icontains=search_filter) |
            Q(supplier__name__icontains=search_filter) |
            Q(receipt_number__icontains=search_filter)
        )

    # البحث من DataTable
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(supplier__name__icontains=search_value) |
            Q(receipt_number__icontains=search_value)
        )

    # العدد الكلي
    total_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    if length > 0:
        queryset = queryset[start:start + length]

    # البيانات - يجب أن تكون array وليس dictionary
    data = []
    for invoice in queryset:
        # تحديد نوع الفاتورة
        invoice_type_display = 'شراء' if invoice.invoice_type == 'purchase' else 'مرتجع'

        # تحديد الحالة
        if invoice.is_posted:
            status_badge = '<span class="badge bg-success">مرحلة</span>'
        else:
            status_badge = '<span class="badge bg-warning text-dark">مسودة</span>'

        # أزرار الإجراءات
        actions = f'''
        <div class="btn-group" role="group">
            <a href="{reverse('purchases:invoice_detail', args=[invoice.pk])}"
               class="btn btn-sm btn-info"
               data-bs-toggle="tooltip"
               title="عرض">
                <i class="fas fa-eye"></i>
            </a>
        '''

        if not invoice.is_posted:
            actions += f'''
            <a href="{reverse('purchases:invoice_update', args=[invoice.pk])}"
               class="btn btn-sm btn-primary"
               data-bs-toggle="tooltip"
               title="تعديل">
                <i class="fas fa-edit"></i>
            </a>
            <button onclick="postInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-success"
                    data-bs-toggle="tooltip"
                    title="ترحيل">
                <i class="fas fa-check"></i>
            </button>
            <button onclick="deleteInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-danger"
                    data-bs-toggle="tooltip"
                    title="حذف">
                <i class="fas fa-trash"></i>
            </button>
            '''
        else:
            actions += f'''
            <button onclick="unpostInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-warning"
                    data-bs-toggle="tooltip"
                    title="إلغاء الترحيل">
                <i class="fas fa-undo"></i>
            </button>
            '''

        actions += '</div>'

        # إضافة البيانات كـ array وليس dictionary
        data.append([
            invoice.number or '-',  # 0: رقم الفاتورة
            invoice.date.strftime('%Y-%m-%d'),  # 1: التاريخ
            invoice.supplier.name if invoice.supplier else '-',  # 2: المورد
            invoice_type_display,  # 3: النوع
            invoice.warehouse.name if invoice.warehouse else '-',  # 4: المستودع
            f'{float(invoice.total_with_tax):,.3f}',  # 5: الإجمالي
            status_badge,  # 6: الحالة
            actions  # 7: الإجراءات
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def export_invoices_excel(request):
    """تصدير فواتير المشتريات إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    # استرجاع البيانات
    queryset = PurchaseInvoice.objects.filter(
        company=request.current_company
    ).select_related(
        'supplier', 'warehouse', 'currency'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر من GET parameters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    is_posted = request.GET.get('is_posted')
    if is_posted == '1':
        queryset = queryset.filter(is_posted=True)
    elif is_posted == '0':
        queryset = queryset.filter(is_posted=False)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "فواتير المشتريات"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم الفاتورة', 'التاريخ', 'المورد', 'المستودع',
        'رقم الإيصال', 'الإجمالي', 'الضريبة', 'الإجمالي مع الضريبة',
        'العملة', 'الحالة'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, invoice in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=invoice.number).border = border
        ws.cell(row=row_num, column=2, value=invoice.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=invoice.supplier.name).border = border
        ws.cell(row=row_num, column=4, value=invoice.warehouse.name).border = border
        ws.cell(row=row_num, column=5, value=invoice.receipt_number).border = border
        ws.cell(row=row_num, column=6, value=float(invoice.total_amount)).border = border
        ws.cell(row=row_num, column=7, value=float(invoice.tax_amount)).border = border
        ws.cell(row=row_num, column=8, value=float(invoice.total_with_tax)).border = border
        ws.cell(row=row_num, column=9, value=invoice.currency.code).border = border
        ws.cell(row=row_num, column=10, value='مرحل' if invoice.is_posted else 'مسودة').border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"purchase_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ============================================
# AJAX Endpoints for Invoice Form Enhancements
# ============================================

@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def get_supplier_item_price_ajax(request):
    """
    جلب آخر سعر شراء من المورد لمادة معينة
    يستخدم في فاتورة المشتريات للتعبئة التلقائية للسعر
    """
    supplier_id = request.GET.get('supplier_id')
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not supplier_id or not item_id:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        from apps.core.models import PartnerItemPrice

        # البحث عن آخر سعر شراء
        filter_params = {
            'company': request.current_company,
            'partner_id': supplier_id,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        price_record = PartnerItemPrice.objects.filter(
            **filter_params
        ).first()

        if price_record and price_record.last_purchase_price:
            return JsonResponse({
                'success': True,
                'has_price': True,
                'last_price': str(price_record.last_purchase_price),
                'last_date': price_record.last_purchase_date.strftime('%Y-%m-%d') if price_record.last_purchase_date else None,
                'last_quantity': str(price_record.last_purchase_quantity) if price_record.last_purchase_quantity else None,
                'total_purchased': str(price_record.total_purchased_quantity),
            })
        else:
            return JsonResponse({
                'success': True,
                'has_price': False,
                'message': 'لا توجد مشتريات سابقة من هذا المورد لهذه المادة'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def get_item_stock_multi_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة من جميع الفروع
    يعرض الكميات المتوفرة في كل مخزن مع تفاصيل الفرع
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث عن الكميات في جميع المخازن
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        stock_records = ItemStock.objects.filter(
            **filter_params
        ).select_related(
            'warehouse', 'warehouse__branch'
        ).order_by('warehouse__branch__name', 'warehouse__name')

        # تجميع البيانات
        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            branches_data.append({
                'branch_name': stock.warehouse.branch.name,
                'warehouse_name': stock.warehouse.name,
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost),
            })

        return JsonResponse({
            'success': True,
            'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def get_item_stock_current_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة في الفرع الحالي فقط
    يستخدم للعرض السريع في جدول الفاتورة
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث في الفرع الحالي
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
            'warehouse__branch': request.current_branch,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        # حساب الإجمالي في الفرع الحالي
        stock_aggregate = ItemStock.objects.filter(
            **filter_params
        ).aggregate(
            total_qty=Sum('quantity'),
            total_reserved=Sum('reserved_quantity')
        )

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True,
            'quantity': str(total_qty),
            'reserved': str(total_reserved),
            'available': str(available),
            'has_stock': total_qty > 0,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def item_search_ajax(request):
    """
    AJAX Live Search للمواد
    يُستخدم للبحث المباشر بدلاً من تحميل جميع المواد
    """
    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({
            'success': False,
            'message': 'يرجى إدخال حرفين على الأقل للبحث'
        })

    try:
        from apps.core.models import ItemVariant, PartnerItemPrice

        # البحث في المتغيرات بدلاً من المواد الأساسية
        variants = ItemVariant.objects.filter(
            item__company=request.current_company,
            item__is_active=True,
            is_active=True
        ).filter(
            Q(item__name__icontains=term) |
            Q(item__code__icontains=term) |
            Q(item__barcode__icontains=term) |
            Q(code__icontains=term) |
            Q(barcode__icontains=term)
        ).annotate(
            # كمية المخزون في الفرع الحالي للمتغير
            current_branch_stock=Sum(
                'stock_records__quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)
            ),
            # الكمية المحجوزة في الفرع الحالي
            current_branch_reserved=Sum(
                'stock_records__reserved_quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)
            ),
            # إجمالي المخزون في كل الفروع
            total_stock=Sum('stock_records__quantity'),
        ).select_related(
            'item__category', 'item__base_uom', 'item'
        )[:limit]

        items_data = []
        for variant in variants:
            item = variant.item

            # Get last purchase price from PartnerItemPrice if available
            last_purchase_price = None
            try:
                # 1. البحث عن آخر سعر شراء فعلي
                price_record = PartnerItemPrice.objects.filter(
                    company=request.current_company,
                    item=item,
                    item_variant=variant,
                    last_purchase_price__isnull=False
                ).order_by('-last_purchase_date').first()

                if price_record:
                    last_purchase_price = str(price_record.last_purchase_price)
                elif variant.cost_price:
                    # 2. استخدم سعر التكلفة من المتغير إذا لم يكن هناك سعر شراء
                    last_purchase_price = str(variant.cost_price)
                elif variant.base_price:
                    # 3. أو السعر الأساسي
                    last_purchase_price = str(variant.base_price)
                else:
                    # 4. البحث في قوائم الأسعار (أول سعر متوفر)
                    from apps.core.models import PriceListItem
                    from datetime import date

                    price_list_item = PriceListItem.objects.filter(
                        item=item,
                        variant=variant,
                        price__isnull=False
                    ).filter(
                        Q(start_date__isnull=True) | Q(start_date__lte=date.today())
                    ).filter(
                        Q(end_date__isnull=True) | Q(end_date__gte=date.today())
                    ).order_by('price').first()

                    if price_list_item:
                        last_purchase_price = str(price_list_item.price)
            except Exception as e:
                # Fallback في حالة حدوث خطأ
                if variant.cost_price:
                    last_purchase_price = str(variant.cost_price)
                elif variant.base_price:
                    last_purchase_price = str(variant.base_price)
                else:
                    try:
                        from apps.core.models import PriceListItem
                        from datetime import date

                        price_list_item = PriceListItem.objects.filter(
                            item=item,
                            variant=variant,
                            price__isnull=False
                        ).order_by('price').first()

                        if price_list_item:
                            last_purchase_price = str(price_list_item.price)
                    except:
                        pass

            # بناء اسم مركب: المادة + المتغير
            display_name = item.name
            if variant.code and variant.code != item.code:
                display_name += f" - {variant.code}"

            items_data.append({
                'id': item.id,
                'variant_id': variant.id,
                'name': display_name,  # الاسم المركب
                'item_name': item.name,  # اسم المادة فقط
                'variant_code': variant.code,  # كود المتغير
                'code': item.code,
                'barcode': variant.barcode or item.barcode or '',
                'description': item.description or '',
                'category_name': item.category.name if item.category else '',
                'tax_rate': str(item.tax_rate),
                'base_uom_name': item.base_uom.name if item.base_uom else '',
                'base_uom_code': item.base_uom.code if item.base_uom else '',
                'base_uom_id': item.base_uom.id if item.base_uom else None,
                'last_purchase_price': last_purchase_price,
                'current_branch_stock': str(variant.current_branch_stock or 0),
                'current_branch_reserved': str(variant.current_branch_reserved or 0),
                'total_stock': str(variant.total_stock or 0),
            })

        return JsonResponse({
            'success': True,
            'items': items_data,
            'count': len(items_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.add_purchaseinvoice', raise_exception=True)
@transaction.atomic
def save_invoice_draft_ajax(request):
    """
    حفظ الفاتورة كمسودة (Auto-save)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        invoice_id = request.POST.get('invoice_id')

        # بيانات الفاتورة الأساسية
        invoice_data = {
            'supplier_id': request.POST.get('supplier'),
            'warehouse_id': request.POST.get('warehouse'),
            'date': request.POST.get('date'),
            'number': request.POST.get('number'),
            'supplier_invoice_number': request.POST.get('supplier_invoice_number'),
            'supplier_invoice_date': request.POST.get('supplier_invoice_date'),
            'reference': request.POST.get('reference'),
            'notes': request.POST.get('notes'),
            'discount_type': request.POST.get('discount_type', 'none'),
            'discount_value': request.POST.get('discount_value', 0),
        }

        # التحقق من البيانات الأساسية
        if not invoice_data['supplier_id'] or not invoice_data['warehouse_id']:
            return JsonResponse({
                'success': False,
                'message': 'يرجى اختيار المورد والمخزن'
            })

        # حفظ أو تحديث الفاتورة
        if invoice_id:
            invoice = get_object_or_404(
                PurchaseInvoice,
                pk=invoice_id,
                company=request.current_company,
                is_posted=False
            )
            for key, value in invoice_data.items():
                if value:
                    setattr(invoice, key.replace('_id', ''), value)
            invoice.save()
        else:
            invoice = PurchaseInvoice.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                created_by=request.user,
                **invoice_data
            )

        # حفظ الأسطر (بسيط - فقط للتجربة)
        # في production يجب معالجة formset كامل

        return JsonResponse({
            'success': True,
            'invoice_id': invoice.id,
            'invoice_number': invoice.number,
            'message': 'تم حفظ المسودة بنجاح',
            'saved_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaseinvoice', raise_exception=True)
def get_item_uom_conversions_ajax(request):
    """
    AJAX endpoint للحصول على قائمة تحويلات الوحدات لمادة معينة
    يُستخدم لملء dropdown وحدة الشراء
    """
    from apps.core.models import ItemUoMConversion

    item_id = request.GET.get('item_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        # جلب التحويلات من ItemUoMConversion
        conversions = ItemUoMConversion.objects.filter(
            item__company=request.current_company,
            item_id=item_id,
            is_active=True
        ).select_related('from_uom', 'to_uom').values(
            'id',
            'from_uom_id',
            'from_uom__name',
            'from_uom__code',
            'to_uom_id',
            'to_uom__name',
            'to_uom__code',
            'conversion_rate'
        )

        conversions_list = list(conversions)

        return JsonResponse({
            'success': True,
            'conversions': conversions_list,
            'count': len(conversions_list)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
