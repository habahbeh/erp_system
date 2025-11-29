# apps/purchases/views/request_views.py
"""
Views for Purchase Requests
Complete CRUD operations + Workflow (Submit, Approve, Reject)
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from decimal import Decimal
from datetime import datetime, date

from ..models import PurchaseRequest, PurchaseRequestItem
from ..forms import (
    PurchaseRequestForm,
    PurchaseRequestItemForm,
    PurchaseRequestItemFormSet,
    PurchaseRequestFilterForm,
)
from apps.core.models import Item, User


class PurchaseRequestListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة طلبات الشراء"""
    model = PurchaseRequest
    template_name = 'purchases/requests/request_list.html'
    context_object_name = 'requests'
    paginate_by = 50
    permission_required = 'purchases.view_purchaserequest'

    def get_queryset(self):
        queryset = PurchaseRequest.objects.filter(
            company=self.request.current_company
        ).select_related(
            'requested_by', 'created_by'
        ).prefetch_related('lines').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(department__name__icontains=search) |
                Q(purpose__icontains=search) |
                Q(requested_by__first_name__icontains=search) |
                Q(requested_by__last_name__icontains=search)
            )

        # فلترة حسب الموظف الطالب
        requested_by_id = self.request.GET.get('requested_by')
        if requested_by_id:
            queryset = queryset.filter(requested_by_id=requested_by_id)

        # فلترة حسب الحالة
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('طلبات الشراء')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات الشراء'), 'url': ''},
        ]

        # الإحصائيات
        requests = self.get_queryset()
        context['stats'] = {
            'total_count': requests.count(),
            'draft_count': requests.filter(status='draft').count(),
            'submitted_count': requests.filter(status='submitted').count(),
            'approved_count': requests.filter(status='approved').count(),
            'rejected_count': requests.filter(status='rejected').count(),
            'ordered_count': requests.filter(status='ordered').count(),
        }

        # قائمة المستخدمين النشطين للفلترة
        context['employees'] = User.objects.filter(
            is_active=True
        ).order_by('first_name', 'last_name')

        # Filter form
        context['filter_form'] = PurchaseRequestFilterForm(
            self.request.GET,
            company=self.request.current_company
        )

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('purchases.add_purchaserequest')
        context['can_edit'] = self.request.user.has_perm('purchases.change_purchaserequest')
        context['can_delete'] = self.request.user.has_perm('purchases.delete_purchaserequest')
        context['can_approve'] = self.request.user.has_perm('purchases.change_purchaserequest')

        return context


class PurchaseRequestDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل طلب الشراء"""
    model = PurchaseRequest
    template_name = 'purchases/requests/request_detail.html'
    context_object_name = 'request_obj'
    permission_required = 'purchases.view_purchaserequest'

    def get_queryset(self):
        return PurchaseRequest.objects.filter(
            company=self.request.current_company
        ).select_related(
            'requested_by', 'created_by'
        ).prefetch_related('lines__item')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        request_obj = self.object
        context['title'] = f'{_("طلب شراء")} {request_obj.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات الشراء'), 'url': reverse('purchases:request_list')},
            {'title': request_obj.number, 'url': ''},
        ]

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('purchases.change_purchaserequest') and
            request_obj.status in ['draft', 'rejected']
        )
        context['can_delete'] = (
            self.request.user.has_perm('purchases.delete_purchaserequest') and
            request_obj.status == 'draft'
        )
        context['can_submit'] = (
            self.request.user.has_perm('purchases.change_purchaserequest') and
            request_obj.status == 'draft'
        )
        context['can_approve'] = (
            self.request.user.has_perm('purchases.change_purchaserequest') and
            request_obj.status == 'submitted'
        )
        context['can_reject'] = (
            self.request.user.has_perm('purchases.change_purchaserequest') and
            request_obj.status == 'submitted'
        )
        context['can_create_order'] = (
            self.request.user.has_perm('purchases.add_purchaseorder') and
            request_obj.status == 'approved'
        )

        # أوامر الشراء المرتبطة
        context['related_orders'] = request_obj.orders.all()

        return context


class PurchaseRequestCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء طلب شراء جديد"""
    model = PurchaseRequest
    form_class = PurchaseRequestForm
    template_name = 'purchases/requests/request_form.html'
    permission_required = 'purchases.add_purchaserequest'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة طلب شراء')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات الشراء'), 'url': reverse('purchases:request_list')},
            {'title': _('إضافة طلب'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseRequestItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )
        else:
            context['formset'] = PurchaseRequestItemFormSet(
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )

        # AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        # ربط الطلب بالشركة
        form.instance.company = self.request.current_company
        form.instance.created_by = self.request.user

        # تعيين القسم تلقائياً من الموظف
        if form.instance.requested_by_employee:
            form.instance.department = form.instance.requested_by_employee.department

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            messages.success(
                self.request,
                _('تم إضافة طلب الشراء بنجاح')
            )
            return redirect('purchases:request_detail', pk=self.object.pk)
        else:
            # طباعة الأخطاء للتشخيص
            print("Form errors:", form.errors)
            print("Formset errors:", formset.errors)
            print("Formset non-form errors:", formset.non_form_errors())

            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseRequestUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل طلب شراء"""
    model = PurchaseRequest
    form_class = PurchaseRequestForm
    template_name = 'purchases/requests/request_form.html'
    permission_required = 'purchases.change_purchaserequest'

    def get_queryset(self):
        # يمكن تعديل الطلب في حالة مسودة أو مرفوض فقط
        return PurchaseRequest.objects.filter(
            company=self.request.current_company,
            status__in=['draft', 'rejected']
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        request_obj = self.object
        context['title'] = f'{_("تعديل طلب")} {request_obj.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات الشراء'), 'url': reverse('purchases:request_list')},
            {'title': request_obj.number, 'url': reverse('purchases:request_detail', kwargs={'pk': request_obj.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseRequestItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )
        else:
            context['formset'] = PurchaseRequestItemFormSet(
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )

        # AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        # تعيين القسم تلقائياً من الموظف
        if form.instance.requested_by_employee:
            form.instance.department = form.instance.requested_by_employee.department

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            messages.success(
                self.request,
                _('تم تعديل طلب الشراء بنجاح')
            )
            return redirect('purchases:request_detail', pk=self.object.pk)
        else:
            # طباعة الأخطاء للتشخيص
            print("Form errors:", form.errors)
            print("Formset errors:", formset.errors)
            print("Formset non-form errors:", formset.non_form_errors())

            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseRequestDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف طلب شراء"""
    model = PurchaseRequest
    template_name = 'purchases/requests/request_confirm_delete.html'
    success_url = reverse_lazy('purchases:request_list')
    permission_required = 'purchases.delete_purchaserequest'

    def get_queryset(self):
        # يمكن حذف الطلب في حالة مسودة فقط
        return PurchaseRequest.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف طلب الشراء بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# Workflow Actions
# ============================================================================

@login_required
@permission_required('purchases.change_purchaserequest', raise_exception=True)
@transaction.atomic
def submit_request(request, pk):
    """تقديم طلب الشراء"""
    request_obj = get_object_or_404(
        PurchaseRequest,
        pk=pk,
        company=request.current_company
    )

    if request_obj.status != 'draft':
        messages.error(request, _('يمكن تقديم الطلب في حالة مسودة فقط'))
        return redirect('purchases:request_detail', pk=pk)

    try:
        request_obj.submit()
        messages.success(
            request,
            _('تم تقديم طلب الشراء بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في التقديم")}: {str(e)}')

    return redirect('purchases:request_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaserequest', raise_exception=True)
@transaction.atomic
def approve_request(request, pk):
    """اعتماد طلب الشراء"""
    request_obj = get_object_or_404(
        PurchaseRequest,
        pk=pk,
        company=request.current_company
    )

    if request_obj.status != 'submitted':
        messages.error(request, _('يمكن اعتماد الطلب في حالة مقدم فقط'))
        return redirect('purchases:request_detail', pk=pk)

    try:
        # تمرير المستخدم الحالي لتسجيل من قام بالاعتماد
        request_obj.approve(user=request.user)
        messages.success(
            request,
            _('تم اعتماد طلب الشراء بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الاعتماد")}: {str(e)}')

    return redirect('purchases:request_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchaserequest', raise_exception=True)
@transaction.atomic
def reject_request(request, pk):
    """رفض طلب الشراء"""
    request_obj = get_object_or_404(
        PurchaseRequest,
        pk=pk,
        company=request.current_company
    )

    if request_obj.status != 'submitted':
        messages.error(request, _('يمكن رفض الطلب في حالة مقدم فقط'))
        return redirect('purchases:request_detail', pk=pk)

    try:
        # استلام سبب الرفض من الطلب (POST أو GET)
        rejection_reason = request.POST.get('reason', '') or request.GET.get('reason', '')

        # تمرير المستخدم وسبب الرفض
        request_obj.reject(user=request.user, reason=rejection_reason)
        messages.success(
            request,
            _('تم رفض طلب الشراء')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الرفض")}: {str(e)}')

    return redirect('purchases:request_detail', pk=pk)


@login_required
@permission_required('purchases.add_purchaseorder', raise_exception=True)
def create_order_from_request(request, pk):
    """إنشاء أمر شراء من طلب الشراء"""
    request_obj = get_object_or_404(
        PurchaseRequest,
        pk=pk,
        company=request.current_company
    )

    if request_obj.status != 'approved':
        messages.error(request, _('يجب اعتماد الطلب أولاً لإنشاء أمر شراء منه'))
        return redirect('purchases:request_detail', pk=pk)

    # إعادة توجيه إلى صفحة إنشاء أمر الشراء مع معرف الطلب
    return redirect(f"{reverse('purchases:order_create')}?from_request={pk}")


# ============================================================================
# AJAX & Export
# ============================================================================

@login_required
@permission_required('purchases.view_purchaserequest', raise_exception=True)
def request_datatable_ajax(request):
    """AJAX endpoint for DataTables"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # Additional filters
    status_filter = request.GET.get('status', '')
    requested_by_filter = request.GET.get('requested_by', '')
    date_from_filter = request.GET.get('date_from', '')
    date_to_filter = request.GET.get('date_to', '')
    search_filter = request.GET.get('search_filter', '')

    queryset = PurchaseRequest.objects.filter(
        company=request.current_company
    ).select_related('requested_by')

    # البحث العام (من شريط البحث في الجدول)
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(department__name__icontains=search_value) |
            Q(requested_by__first_name__icontains=search_value) |
            Q(requested_by__last_name__icontains=search_value) |
            Q(purpose__icontains=search_value)
        )

    # البحث من حقل البحث المخصص
    if search_filter:
        queryset = queryset.filter(
            Q(number__icontains=search_filter) |
            Q(department__name__icontains=search_filter) |
            Q(requested_by__first_name__icontains=search_filter) |
            Q(requested_by__last_name__icontains=search_filter) |
            Q(purpose__icontains=search_filter)
        )

    # فلترة الحالة
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    # فلترة الموظف الطالب
    if requested_by_filter:
        queryset = queryset.filter(requested_by_id=requested_by_filter)

    # فلترة التاريخ من
    if date_from_filter:
        queryset = queryset.filter(date__gte=date_from_filter)

    # فلترة التاريخ إلى
    if date_to_filter:
        queryset = queryset.filter(date__lte=date_to_filter)

    # العدد الكلي قبل الفلترة
    total_records = PurchaseRequest.objects.filter(
        company=request.current_company
    ).count()

    # العدد بعد الفلترة
    filtered_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    if length > 0:
        queryset = queryset[start:start + length]
    else:
        queryset = queryset[start:]

    # البيانات
    data = []
    for req in queryset:
        data.append({
            'number': req.number,
            'date': req.date.strftime('%Y-%m-%d'),
            'requested_by': req.requested_by.get_full_name() if req.requested_by else '-',
            'department': req.department.name if req.department else '-',
            'required_date': req.required_date.strftime('%Y-%m-%d') if req.required_date else '',
            'status': req.get_status_display(),
            'status_code': req.status,
            'pk': req.pk,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@login_required
@permission_required('purchases.view_purchaserequest', raise_exception=True)
def export_requests_excel(request):
    """تصدير طلبات الشراء إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    # استرجاع البيانات
    queryset = PurchaseRequest.objects.filter(
        company=request.current_company
    ).select_related(
        'requested_by'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر من GET parameters
    requested_by_id = request.GET.get('requested_by')
    if requested_by_id:
        queryset = queryset.filter(requested_by_id=requested_by_id)

    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "طلبات الشراء"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم الطلب', 'التاريخ', 'طلب بواسطة', 'القسم',
        'الغرض', 'التاريخ المطلوب', 'الحالة'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, req in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=req.number).border = border
        ws.cell(row=row_num, column=2, value=req.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=req.requested_by.get_full_name() if req.requested_by else '').border = border
        ws.cell(row=row_num, column=4, value=req.department.name if req.department else '').border = border
        ws.cell(row=row_num, column=5, value=req.purpose or '').border = border
        ws.cell(row=row_num, column=6, value=req.required_date.strftime('%Y-%m-%d') if req.required_date else '').border = border
        ws.cell(row=row_num, column=7, value=req.get_status_display()).border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"purchase_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ============================================================================
# AJAX Endpoints for Live Search & Stock Display
# ============================================================================

@login_required
@permission_required('purchases.view_purchaserequest', raise_exception=True)
def get_item_stock_multi_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة من جميع الفروع
    يعرض الكميات المتوفرة في كل مخزن مع تفاصيل الفرع
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث عن الكميات في جميع المخازن
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        stock_records = ItemStock.objects.filter(
            **filter_params
        ).select_related(
            'warehouse', 'warehouse__branch'
        ).order_by('warehouse__branch__name', 'warehouse__name')

        # تجميع البيانات
        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            # التحقق من وجود branch
            branch_name = 'غير محدد'
            if stock.warehouse and stock.warehouse.branch:
                branch_name = stock.warehouse.branch.name

            branches_data.append({
                'branch_name': branch_name,
                'warehouse_name': stock.warehouse.name if stock.warehouse else 'غير محدد',
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost or 0),
            })

        return JsonResponse({
            'success': True,
            'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaserequest', raise_exception=True)
def get_item_stock_current_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة في الفرع الحالي فقط
    يستخدم للعرض السريع في جدول طلب الشراء
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث في الفرع الحالي
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
            'warehouse__branch': request.current_branch,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        # حساب الإجمالي في الفرع الحالي
        stock_aggregate = ItemStock.objects.filter(
            **filter_params
        ).aggregate(
            total_qty=Sum('quantity'),
            total_reserved=Sum('reserved_quantity')
        )

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True,
            'quantity': str(total_qty),
            'reserved': str(total_reserved),
            'available': str(available),
            'has_stock': total_qty > 0,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_purchaserequest', raise_exception=True)
def item_search_ajax(request):
    """
    AJAX Live Search للمواد
    يُستخدم للبحث المباشر بدلاً من تحميل جميع المواد
    """
    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({
            'success': False,
            'message': 'يرجى إدخال حرفين على الأقل للبحث'
        })

    try:
        items = Item.objects.filter(
            company=request.current_company,
            is_active=True
        ).filter(
            Q(name__icontains=term) |
            Q(code__icontains=term) |
            Q(barcode__icontains=term)
        ).annotate(
            # كمية المخزون في الفرع الحالي (itemstock هو الـ related_name الصحيح)
            current_branch_stock=Sum(
                'itemstock__quantity',
                filter=Q(itemstock__warehouse__branch=request.current_branch)
            ),
            # الكمية المحجوزة في الفرع الحالي
            current_branch_reserved=Sum(
                'itemstock__reserved_quantity',
                filter=Q(itemstock__warehouse__branch=request.current_branch)
            ),
            # إجمالي المخزون في كل الفروع
            total_stock=Sum('itemstock__quantity'),
        ).select_related(
            'category', 'base_uom'
        )[:limit]

        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'name': item.name,
                'code': item.code,
                'barcode': item.barcode or '',
                'category_name': item.category.name if item.category else '',
                'tax_rate': str(item.tax_rate),
                'base_uom_name': item.base_uom.name if item.base_uom else '',
                'base_uom_code': item.base_uom.code if item.base_uom else '',
                'current_branch_stock': str(item.current_branch_stock or 0),
                'current_branch_reserved': str(item.current_branch_reserved or 0),
                'total_stock': str(item.total_stock or 0),
            })

        return JsonResponse({
            'success': True,
            'items': items_data,
            'count': len(items_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.add_purchaserequest', raise_exception=True)
@transaction.atomic
def save_request_draft_ajax(request):
    """
    حفظ طلب الشراء كمسودة (Auto-save)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        from django.utils import timezone

        request_id = request.POST.get('request_id')

        # بيانات طلب الشراء الأساسية
        request_data = {
            'requested_by_id': request.POST.get('requested_by') or request.user.id,
            'department_id': request.POST.get('department'),
            'date': request.POST.get('date'),
            'required_date': request.POST.get('required_date'),
            'number': request.POST.get('number'),
            'purpose': request.POST.get('purpose'),
            'notes': request.POST.get('notes'),
        }

        # التحقق من البيانات الأساسية
        if not request_data['department_id']:
            return JsonResponse({
                'success': False,
                'message': 'يرجى اختيار القسم'
            })

        # حفظ أو تحديث طلب الشراء
        if request_id:
            purchase_request = get_object_or_404(
                PurchaseRequest,
                pk=request_id,
                company=request.current_company,
                status='draft'
            )
            for key, value in request_data.items():
                if value:
                    setattr(purchase_request, key.replace('_id', ''), value)
            purchase_request.save()
        else:
            purchase_request = PurchaseRequest.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                created_by=request.user,
                status='draft',
                **request_data
            )

        return JsonResponse({
            'success': True,
            'request_id': purchase_request.id,
            'request_number': purchase_request.number,
            'message': 'تم حفظ المسودة بنجاح',
            'saved_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_employee_department_ajax(request):
    """جلب معلومات القسم للموظف المحدد"""
    employee_id = request.GET.get('employee_id')

    if not employee_id:
        return JsonResponse({
            'success': False,
            'error': 'يجب تحديد الموظف'
        }, status=400)

    try:
        from apps.hr.models import Employee

        employee = Employee.objects.select_related('department').get(
            id=employee_id,
            company=request.current_company,
            is_active=True
        )

        return JsonResponse({
            'success': True,
            'department': {
                'id': employee.department.id if employee.department else None,
                'name': employee.department.name if employee.department else 'غير محدد',
                'name_en': employee.department.name_en if employee.department and hasattr(employee.department, 'name_en') else ''
            },
            'employee': {
                'id': employee.id,
                'name': employee.get_full_name()
            }
        })

    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'الموظف غير موجود'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
