# apps/purchases/views/goods_receipt_views.py
"""
Views for Goods Receipt (استلام البضاعة)
"""

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.db import transaction
from django.db.models import Q, Sum
from decimal import Decimal

from ..models import GoodsReceipt, GoodsReceiptLine, PurchaseOrder
from ..forms import GoodsReceiptForm, GoodsReceiptLineFormSet
from apps.core.models import Item


class GoodsReceiptListView(LoginRequiredMixin, ListView):
    """قائمة محاضر استلام البضاعة"""
    model = GoodsReceipt
    template_name = 'purchases/goods_receipt/goods_receipt_list.html'
    context_object_name = 'goods_receipts'
    paginate_by = 50

    def get_queryset(self):
        queryset = GoodsReceipt.objects.filter(
            company=self.request.current_company
        ).select_related(
            'purchase_order', 'supplier', 'warehouse', 'received_by'
        ).order_by('-date', '-number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('محاضر استلام البضاعة')
        return context


class GoodsReceiptDetailView(LoginRequiredMixin, DetailView):
    """تفاصيل محضر استلام البضاعة"""
    model = GoodsReceipt
    template_name = 'purchases/goods_receipt/goods_receipt_detail.html'
    context_object_name = 'goods_receipt'

    def get_queryset(self):
        return GoodsReceipt.objects.filter(
            company=self.request.current_company
        ).select_related(
            'purchase_order', 'supplier', 'warehouse', 'received_by', 'invoice'
        ).prefetch_related('lines__item')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تفاصيل محضر الاستلام')
        return context


class GoodsReceiptCreateView(LoginRequiredMixin, CreateView):
    """إنشاء محضر استلام جديد"""
    model = GoodsReceipt
    form_class = GoodsReceiptForm
    template_name = 'purchases/goods_receipt/goods_receipt_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('محضر استلام جديد')

        if self.request.POST:
            context['lines'] = GoodsReceiptLineFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )
        else:
            context['lines'] = GoodsReceiptLineFormSet(
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )

        # AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        lines = context['lines']

        with transaction.atomic():
            # حفظ الرأس
            self.object = form.save(commit=False)
            self.object.company = self.request.current_company
            self.object.branch = self.request.current_branch
            self.object.created_by = self.request.user
            self.object.save()

            # حفظ السطور
            if lines.is_valid():
                lines.instance = self.object
                lines.save()
            else:
                return self.form_invalid(form)

        messages.success(
            self.request,
            _('تم إنشاء محضر الاستلام %(number)s بنجاح') % {'number': self.object.number}
        )
        return redirect('purchases:goods_receipt_detail', pk=self.object.pk)


class GoodsReceiptUpdateView(LoginRequiredMixin, UpdateView):
    """تعديل محضر استلام"""
    model = GoodsReceipt
    form_class = GoodsReceiptForm
    template_name = 'purchases/goods_receipt/goods_receipt_form.html'

    def get_queryset(self):
        return GoodsReceipt.objects.filter(
            company=self.request.current_company,
            status='draft'  # يمكن تعديل المسودات فقط
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('تعديل محضر الاستلام')

        if self.request.POST:
            context['lines'] = GoodsReceiptLineFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )
        else:
            context['lines'] = GoodsReceiptLineFormSet(
                instance=self.object,
                company=self.request.current_company,
                prefix='lines'
            )

        # AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []  # فارغ لأن البحث سيكون عبر AJAX

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        lines = context['lines']

        with transaction.atomic():
            self.object = form.save()

            if lines.is_valid():
                lines.instance = self.object
                lines.save()
            else:
                return self.form_invalid(form)

        messages.success(self.request, _('تم تعديل محضر الاستلام بنجاح'))
        return redirect('purchases:goods_receipt_detail', pk=self.object.pk)


class GoodsReceiptDeleteView(LoginRequiredMixin, DeleteView):
    """حذف محضر استلام"""
    model = GoodsReceipt
    template_name = 'purchases/goods_receipt/goods_receipt_confirm_delete.html'
    success_url = reverse_lazy('purchases:goods_receipt_list')

    def get_queryset(self):
        return GoodsReceipt.objects.filter(
            company=self.request.current_company,
            status='draft',  # يمكن حذف المسودات فقط
            is_posted=False
        )

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        messages.success(
            request,
            _('تم حذف محضر الاستلام %(number)s بنجاح') % {'number': self.object.number}
        )
        return super().delete(request, *args, **kwargs)


@login_required
@permission_required('purchases.change_goodsreceipt', raise_exception=True)
def confirm_goods_receipt(request, pk):
    """تأكيد محضر الاستلام"""
    goods_receipt = get_object_or_404(
        GoodsReceipt,
        pk=pk,
        company=request.current_company
    )

    try:
        goods_receipt.confirm(user=request.user)
        messages.success(
            request,
            _('تم تأكيد محضر الاستلام %(number)s') % {'number': goods_receipt.number}
        )
    except Exception as e:
        messages.error(request, str(e))

    return redirect('purchases:goods_receipt_detail', pk=pk)


@login_required
@permission_required('purchases.change_goodsreceipt', raise_exception=True)
def post_goods_receipt(request, pk):
    """ترحيل محضر الاستلام للمخزون"""
    goods_receipt = get_object_or_404(
        GoodsReceipt,
        pk=pk,
        company=request.current_company
    )

    try:
        stock_in = goods_receipt.post(user=request.user)
        messages.success(
            request,
            _('تم ترحيل محضر الاستلام %(number)s للمخزون') % {'number': goods_receipt.number}
        )
    except Exception as e:
        messages.error(request, str(e))

    return redirect('purchases:goods_receipt_detail', pk=pk)


@login_required
@permission_required('purchases.change_goodsreceipt', raise_exception=True)
def unpost_goods_receipt(request, pk):
    """إلغاء ترحيل محضر الاستلام"""
    goods_receipt = get_object_or_404(
        GoodsReceipt,
        pk=pk,
        company=request.current_company
    )

    try:
        goods_receipt.unpost()
        messages.success(
            request,
            _('تم إلغاء ترحيل محضر الاستلام %(number)s') % {'number': goods_receipt.number}
        )
    except Exception as e:
        messages.error(request, str(e))

    return redirect('purchases:goods_receipt_detail', pk=pk)


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def goods_receipt_datatable_ajax(request):
    """AJAX endpoint for DataTables"""
    # TODO: Implement DataTables server-side processing
    return JsonResponse({'data': [], 'recordsTotal': 0, 'recordsFiltered': 0})


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def export_goods_receipts_excel(request):
    """تصدير محاضر الاستلام لملف Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime

    # استرجاع البيانات
    queryset = GoodsReceipt.objects.filter(
        company=request.current_company
    ).select_related(
        'purchase_order', 'supplier', 'warehouse', 'received_by'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر من GET parameters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "محاضر الاستلام"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم المحضر', 'التاريخ', 'رقم أمر الشراء', 'المورد',
        'المستودع', 'رقم إيصال التسليم', 'تاريخ التسليم',
        'حالة الفحص', 'الحالة', 'استلمها'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, receipt in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=receipt.number).border = border
        ws.cell(row=row_num, column=2, value=receipt.date.strftime('%Y-%m-%d') if receipt.date else '').border = border
        ws.cell(row=row_num, column=3, value=receipt.purchase_order.number if receipt.purchase_order else '').border = border
        ws.cell(row=row_num, column=4, value=receipt.supplier.name if receipt.supplier else '').border = border
        ws.cell(row=row_num, column=5, value=receipt.warehouse.name if receipt.warehouse else '').border = border
        ws.cell(row=row_num, column=6, value=receipt.delivery_note_number or '').border = border
        ws.cell(row=row_num, column=7, value=receipt.delivery_date.strftime('%Y-%m-%d') if receipt.delivery_date else '').border = border
        ws.cell(row=row_num, column=8, value=receipt.get_quality_check_status_display()).border = border
        ws.cell(row=row_num, column=9, value=receipt.get_status_display()).border = border
        ws.cell(row=row_num, column=10, value=receipt.received_by.get_full_name() if receipt.received_by else '').border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"goods_receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ============================================================================
# AJAX Endpoints for Live Search & Stock Display
# ============================================================================

@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def get_purchase_order_item_price_ajax(request):
    """
    جلب معلومات المادة من أمر الشراء
    يُستخدم لملء البيانات تلقائياً في محضر الاستلام
    """
    purchase_order_id = request.GET.get('purchase_order_id')
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not purchase_order_id or not item_id:
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        from ..models import PurchaseOrderItem

        # البحث عن بند في أمر الشراء
        filter_params = {
            'order__company': request.current_company,
            'order_id': purchase_order_id,
            'item_id': item_id,
        }

        order_line = PurchaseOrderItem.objects.filter(
            **filter_params
        ).select_related('unit').first()

        if order_line:
            received_qty = order_line.received_quantity or Decimal('0')
            return JsonResponse({
                'success': True,
                'has_price': True,
                'price': str(order_line.unit_price),
                'quantity': str(order_line.quantity),
                'received_quantity': str(received_qty),
                'remaining_quantity': str(order_line.quantity - received_qty),
                'uom': order_line.unit.name if order_line.unit else None,
            })
        else:
            return JsonResponse({
                'success': True,
                'has_price': False,
                'message': 'لا توجد هذه المادة في أمر الشراء'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def get_item_stock_multi_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة من جميع الفروع
    يعرض الكميات المتوفرة في كل مخزن مع تفاصيل الفرع
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث عن الكميات في جميع المخازن
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        stock_records = ItemStock.objects.filter(
            **filter_params
        ).select_related(
            'warehouse', 'warehouse__branch'
        ).order_by('warehouse__branch__name', 'warehouse__name')

        # تجميع البيانات
        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            # التحقق من وجود branch
            branch_name = 'غير محدد'
            if stock.warehouse and stock.warehouse.branch:
                branch_name = stock.warehouse.branch.name

            branches_data.append({
                'branch_name': branch_name,
                'warehouse_name': stock.warehouse.name if stock.warehouse else 'غير محدد',
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost or 0),
            })

        return JsonResponse({
            'success': True,
            'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def get_item_stock_current_branch_ajax(request):
    """
    جلب رصيد المخزون لمادة معينة في الفرع الحالي فقط
    يستخدم للعرض السريع في جدول محضر الاستلام
    """
    item_id = request.GET.get('item_id')
    variant_id = request.GET.get('variant_id')

    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        from apps.inventory.models import ItemStock

        # البحث في الفرع الحالي
        filter_params = {
            'company': request.current_company,
            'item_id': item_id,
            'warehouse__branch': request.current_branch,
        }

        if variant_id:
            filter_params['item_variant_id'] = variant_id

        # حساب الإجمالي في الفرع الحالي
        stock_aggregate = ItemStock.objects.filter(
            **filter_params
        ).aggregate(
            total_qty=Sum('quantity'),
            total_reserved=Sum('reserved_quantity')
        )

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True,
            'quantity': str(total_qty),
            'reserved': str(total_reserved),
            'available': str(available),
            'has_stock': total_qty > 0,
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def item_search_ajax(request):
    """
    AJAX Live Search للمواد
    يُستخدم للبحث المباشر بدلاً من تحميل جميع المواد
    """
    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({
            'success': False,
            'message': 'يرجى إدخال حرفين على الأقل للبحث'
        })

    try:
        items = Item.objects.filter(
            company=request.current_company,
            is_active=True
        ).filter(
            Q(name__icontains=term) |
            Q(code__icontains=term) |
            Q(barcode__icontains=term)
        ).annotate(
            # كمية المخزون في الفرع الحالي (itemstock هو الـ related_name الصحيح)
            current_branch_stock=Sum(
                'itemstock__quantity',
                filter=Q(itemstock__warehouse__branch=request.current_branch)
            ),
            # الكمية المحجوزة في الفرع الحالي
            current_branch_reserved=Sum(
                'itemstock__reserved_quantity',
                filter=Q(itemstock__warehouse__branch=request.current_branch)
            ),
            # إجمالي المخزون في كل الفروع
            total_stock=Sum('itemstock__quantity'),
        ).select_related(
            'category', 'base_uom'
        )[:limit]

        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'name': item.name,
                'code': item.code,
                'barcode': item.barcode or '',
                'category_name': item.category.name if item.category else '',
                'tax_rate': str(item.tax_rate),
                'base_uom_name': item.base_uom.name if item.base_uom else '',
                'base_uom_code': item.base_uom.code if item.base_uom else '',
                'current_branch_stock': str(item.current_branch_stock or 0),
                'current_branch_reserved': str(item.current_branch_reserved or 0),
                'total_stock': str(item.total_stock or 0),
            })

        return JsonResponse({
            'success': True,
            'items': items_data,
            'count': len(items_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.view_goodsreceipt', raise_exception=True)
def get_purchase_order_lines_ajax(request):
    """
    جلب بنود أمر الشراء لتعبئتها في محضر الاستلام
    """
    purchase_order_id = request.GET.get('purchase_order_id')

    if not purchase_order_id:
        return JsonResponse({'error': 'Missing purchase_order_id'}, status=400)

    try:
        from ..models import PurchaseOrderItem

        # جلب بنود أمر الشراء
        order_lines = PurchaseOrderItem.objects.filter(
            order__company=request.current_company,
            order_id=purchase_order_id
        ).select_related('item', 'item__base_uom', 'unit')

        lines_data = []
        for line in order_lines:
            # حساب الكمية المتبقية
            received_qty = line.received_quantity or Decimal('0')
            remaining_qty = line.quantity - received_qty

            # تخطي البنود المستلمة بالكامل
            if remaining_qty <= 0:
                continue

            lines_data.append({
                'id': line.id,
                'item_id': line.item.id,
                'item_name': line.item.name,
                'item_code': line.item.code,
                'ordered_quantity': str(line.quantity),
                'received_quantity': str(received_qty),
                'remaining_quantity': str(remaining_qty),
                'price': str(line.unit_price),
                'uom_id': line.unit.id if line.unit else (line.item.base_uom.id if line.item.base_uom else None),
                'uom_name': line.unit.name if line.unit else (line.item.base_uom.name if line.item.base_uom else ''),
            })

        return JsonResponse({
            'success': True,
            'lines': lines_data,
            'count': len(lines_data)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required('purchases.add_goodsreceipt', raise_exception=True)
@transaction.atomic
def save_receipt_draft_ajax(request):
    """
    حفظ محضر الاستلام كمسودة (Auto-save)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        from django.utils import timezone

        receipt_id = request.POST.get('receipt_id')

        # بيانات محضر الاستلام الأساسية
        receipt_data = {
            'purchase_order_id': request.POST.get('purchase_order'),
            'warehouse_id': request.POST.get('warehouse'),
            'receipt_date': request.POST.get('receipt_date'),
            'number': request.POST.get('number'),
            'reference': request.POST.get('reference'),
            'notes': request.POST.get('notes'),
        }

        # التحقق من البيانات الأساسية
        if not receipt_data['purchase_order_id'] or not receipt_data['warehouse_id']:
            return JsonResponse({
                'success': False,
                'message': 'يرجى اختيار أمر الشراء والمخزن'
            })

        # حفظ أو تحديث محضر الاستلام
        if receipt_id:
            receipt = get_object_or_404(
                GoodsReceipt,
                pk=receipt_id,
                company=request.current_company,
                status='draft'
            )
            for key, value in receipt_data.items():
                if value:
                    setattr(receipt, key.replace('_id', ''), value)
            receipt.save()
        else:
            receipt = GoodsReceipt.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                created_by=request.user,
                status='draft',
                **receipt_data
            )

        return JsonResponse({
            'success': True,
            'receipt_id': receipt.id,
            'receipt_number': receipt.number,
            'message': 'تم حفظ المسودة بنجاح',
            'saved_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
