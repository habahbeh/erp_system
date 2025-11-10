# apps/purchases/views/report_views.py
"""
Views التقارير - Purchase Reports
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Q, Sum, Count, Avg, F, Max, Min, Value, DecimalField
from django.db.models.functions import TruncMonth, TruncYear, TruncDate, Coalesce
from django.utils.translation import gettext_lazy as _
from datetime import date, timedelta, datetime
from decimal import Decimal
import io

from apps.core.decorators import permission_required_with_message
from apps.core.models import BusinessPartner, Item
from ..models import (
    PurchaseInvoice, PurchaseOrder, PurchaseRequest,
    PurchaseContract, GoodsReceipt, PurchaseQuotation
)


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def purchases_summary_report(request):
    """تقرير ملخص المشتريات"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    invoice_type = request.GET.get('invoice_type', 'purchase')

    # الفواتير
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type=invoice_type
    ).select_related('supplier', 'currency', 'branch')

    if supplier_id:
        invoices = invoices.filter(supplier_id=supplier_id)

    if date_from:
        invoices = invoices.filter(date__gte=date_from)

    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # الإحصائيات
    stats = invoices.aggregate(
        total_count=Count('id'),
        total_amount=Coalesce(Sum('total_with_tax'), Decimal('0')),
        total_tax=Coalesce(Sum('tax_amount'), Decimal('0')),
        total_discount=Coalesce(Sum('discount_amount'), Decimal('0')),
        avg_invoice=Coalesce(Avg('total_with_tax'), Decimal('0'))
    )

    # حسب المورد
    by_supplier = invoices.values(
        'supplier__id',
        'supplier__name',
        'supplier__code'
    ).annotate(
        count=Count('id'),
        total_amount=Coalesce(Sum('total_with_tax'), Decimal('0')),
        avg_amount=Coalesce(Avg('total_with_tax'), Decimal('0'))
    ).order_by('-total_amount')[:20]

    # حسب الشهر
    by_month = invoices.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        count=Count('id'),
        total_amount=Coalesce(Sum('total_with_tax'), Decimal('0'))
    ).order_by('month')

    # الموردين
    suppliers = BusinessPartner.objects.filter(
        company=company,
        partner_type__in=['supplier', 'both'],
        is_active=True
    ).order_by('name')

    context = {
        'title': _('تقرير ملخص المشتريات'),
        'invoices': invoices.order_by('-date')[:100],
        'stats': stats,
        'by_supplier': by_supplier,
        'by_month': list(by_month),
        'suppliers': suppliers,
        'selected_supplier': supplier_id,
        'date_from': date_from,
        'date_to': date_to,
        'invoice_type': invoice_type,
        'today': datetime.now(),
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': '#'},
            {'title': _('ملخص المشتريات'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/purchases_summary.html', context)


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def supplier_performance_report(request):
    """تقرير أداء الموردين"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # الفواتير المرحلة
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type='purchase'
    )

    if date_from:
        invoices = invoices.filter(date__gte=date_from)
    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # أداء الموردين
    supplier_performance = invoices.values(
        'supplier__id',
        'supplier__name',
        'supplier__code',
        'supplier__phone',
        'supplier__email'
    ).annotate(
        # عدد الفواتير
        invoice_count=Count('id'),

        # القيم المالية
        total_purchases=Coalesce(Sum('total_with_tax'), Decimal('0')),
        avg_invoice_value=Coalesce(Avg('total_with_tax'), Decimal('0')),

        # التخفيضات
        total_discounts=Coalesce(Sum('discount_amount'), Decimal('0')),

        # الضرائب
        total_taxes=Coalesce(Sum('tax_amount'), Decimal('0')),

    ).order_by('-total_purchases')

    # الأوامر
    orders = PurchaseOrder.objects.filter(
        company=company,
        status__in=['approved', 'sent', 'partial', 'completed']
    )

    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)

    # محاضر الاستلام
    receipts = GoodsReceipt.objects.filter(
        company=company,
        is_posted=True
    )

    if date_from:
        receipts = receipts.filter(date__gte=date_from)
    if date_to:
        receipts = receipts.filter(date__lte=date_to)

    # إثراء البيانات بعدد الأوامر ومحاضر الاستلام
    for supplier in supplier_performance:
        supplier_id = supplier['supplier__id']

        # عدد الأوامر
        supplier['order_count'] = orders.filter(supplier_id=supplier_id).count()

        # عدد محاضر الاستلام
        supplier['receipt_count'] = receipts.filter(supplier_id=supplier_id).count()

        # متوسط وقت التسليم (إذا توفر)
        supplier['avg_delivery_time'] = '-'

    # الإحصائيات الإجمالية
    total_purchases = sum(s['total_purchases'] for s in supplier_performance)
    total_invoices = sum(s['invoice_count'] for s in supplier_performance)

    total_stats = {
        'total_suppliers': supplier_performance.count(),
        'total_purchases': total_purchases,
        'total_invoices': total_invoices,
        'avg_invoice_value': total_purchases / total_invoices if total_invoices > 0 else 0,
        'total_orders': orders.count(),
        'total_receipts': receipts.count(),
    }

    context = {
        'title': _('تقرير أداء الموردين'),
        'suppliers': supplier_performance,
        'total_stats': total_stats,
        'date_from': date_from,
        'date_to': date_to,
        'today': datetime.now(),
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': '#'},
            {'title': _('أداء الموردين'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/supplier_performance.html', context)


@login_required
@permission_required_with_message('purchases.view_purchaseorder')
def purchase_orders_report(request):
    """تقرير أوامر الشراء"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    orders = PurchaseOrder.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'branch', 'created_by')

    if status:
        orders = orders.filter(status=status)

    if supplier_id:
        orders = orders.filter(supplier_id=supplier_id)

    if date_from:
        orders = orders.filter(date__gte=date_from)

    if date_to:
        orders = orders.filter(date__lte=date_to)

    # الإحصائيات
    stats = orders.aggregate(
        total_count=Count('id'),
        total_sum=Coalesce(Sum('total_amount'), Decimal('0')),
        avg_order=Coalesce(Avg('total_amount'), Decimal('0'))
    )

    # حسب الحالة
    by_status = orders.values('status').annotate(
        count=Count('id'),
        total_sum=Coalesce(Sum('total_amount'), Decimal('0'))
    ).order_by('-count')

    # حسب المورد
    by_supplier = orders.values(
        'supplier__id',
        'supplier__name'
    ).annotate(
        count=Count('id'),
        total_sum=Coalesce(Sum('total_amount'), Decimal('0'))
    ).order_by('-total_sum')[:15]

    # الموردين
    suppliers = BusinessPartner.objects.filter(
        company=company,
        partner_type__in=['supplier', 'both'],
        is_active=True
    ).order_by('name')

    context = {
        'title': _('تقرير أوامر الشراء'),
        'orders': orders.order_by('-date')[:100],
        'stats': stats,
        'by_status': by_status,
        'by_supplier': by_supplier,
        'suppliers': suppliers,
        'selected_status': status,
        'selected_supplier': supplier_id,
        'date_from': date_from,
        'date_to': date_to,
        'today': datetime.now(),
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': '#'},
            {'title': _('أوامر الشراء'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/purchase_orders.html', context)


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def items_purchases_report(request):
    """تقرير مشتريات المواد"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    item_id = request.GET.get('item')
    category_id = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    from apps.purchases.models import PurchaseInvoiceItem

    # سطور الفواتير المرحلة
    invoice_lines = PurchaseInvoiceItem.objects.filter(
        invoice__company=company,
        invoice__is_posted=True,
        invoice__invoice_type='purchase'
    ).select_related('item', 'invoice', 'invoice__supplier', 'unit')

    if item_id:
        invoice_lines = invoice_lines.filter(item_id=item_id)

    if category_id:
        invoice_lines = invoice_lines.filter(item__category_id=category_id)

    if date_from:
        invoice_lines = invoice_lines.filter(invoice__date__gte=date_from)

    if date_to:
        invoice_lines = invoice_lines.filter(invoice__date__lte=date_to)

    # حسب المادة
    by_item = invoice_lines.values(
        'item__id',
        'item__code',
        'item__name',
        'unit__code'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        avg_price=Coalesce(Avg('unit_price'), Decimal('0')),
        invoice_count=Count('invoice', distinct=True),
        supplier_count=Count('invoice__supplier', distinct=True)
    ).order_by('-total_amount')[:50]

    # حسب الفئة
    by_category = invoice_lines.filter(
        item__category__isnull=False
    ).values(
        'item__category__id',
        'item__category__name'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        item_count=Count('item', distinct=True)
    ).order_by('-total_amount')

    # الإحصائيات
    stats = {
        'total_items': by_item.count(),
        'total_quantity': sum(item['quantity'] for item in by_item),
        'total_amount': sum(item['total_amount'] for item in by_item),
        'total_categories': by_category.count(),
    }

    # المواد للفلتر
    from apps.core.models import ItemCategory
    items = Item.objects.filter(
        company=company,
        is_active=True
    ).order_by('name')[:200]

    categories = ItemCategory.objects.filter(
        company=company,
        is_active=True
    ).order_by('name')

    context = {
        'title': _('تقرير مشتريات المواد'),
        'by_item': by_item,
        'by_category': by_category,
        'stats': stats,
        'items': items,
        'categories': categories,
        'selected_item': item_id,
        'selected_category': category_id,
        'date_from': date_from,
        'date_to': date_to,
        'today': datetime.now(),
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': '#'},
            {'title': _('مشتريات المواد'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/items_purchases.html', context)


@login_required
@permission_required_with_message('purchases.view_purchasecontract')
def contracts_report(request):
    """تقرير العقود"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company
    today = date.today()

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')

    contracts = PurchaseContract.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'created_by')

    if status:
        contracts = contracts.filter(status=status)

    if supplier_id:
        contracts = contracts.filter(supplier_id=supplier_id)

    # الإحصائيات
    stats = contracts.aggregate(
        total_count=Count('id'),
        total_value=Coalesce(Sum('contract_value'), Decimal('0')),
        avg_value=Coalesce(Avg('contract_value'), Decimal('0'))
    )

    # حسب الحالة
    by_status = contracts.values('status').annotate(
        count=Count('id'),
        total_value=Coalesce(Sum('contract_value'), Decimal('0'))
    ).order_by('-count')

    # العقود المنتهية قريباً (30 يوم)
    expiring_soon = contracts.filter(
        status='active',
        end_date__lte=today + timedelta(days=30),
        end_date__gte=today
    ).order_by('end_date')

    # العقود المنتهية
    expired = contracts.filter(
        status='active',
        end_date__lt=today
    ).count()

    # الموردين
    suppliers = BusinessPartner.objects.filter(
        company=company,
        partner_type__in=['supplier', 'both'],
        is_active=True
    ).order_by('name')

    context = {
        'title': _('تقرير العقود'),
        'contracts': contracts.order_by('-start_date')[:100],
        'stats': stats,
        'by_status': by_status,
        'expiring_soon': expiring_soon,
        'expired_count': expired,
        'suppliers': suppliers,
        'selected_status': status,
        'selected_supplier': supplier_id,
        'today': datetime.now(),
        'today_date': today,
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': '#'},
            {'title': _('العقود'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/contracts.html', context)


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def reports_list(request):
    """قائمة التقارير المتاحة"""

    reports = [
        {
            'title': 'ملخص المشتريات',
            'description': 'تقرير شامل بجميع فواتير المشتريات والإحصائيات',
            'icon': 'fa-file-invoice-dollar',
            'color': 'primary',
            'url': reverse('purchases:purchases_summary_report')
        },
        {
            'title': 'أداء الموردين',
            'description': 'تحليل أداء الموردين وإحصائيات التعامل',
            'icon': 'fa-chart-line',
            'color': 'success',
            'url': reverse('purchases:supplier_performance_report')
        },
        {
            'title': 'أوامر الشراء',
            'description': 'تقرير تفصيلي بأوامر الشراء وحالاتها',
            'icon': 'fa-clipboard-list',
            'color': 'info',
            'url': reverse('purchases:purchase_orders_report')
        },
        {
            'title': 'مشتريات المواد',
            'description': 'تقرير بمشتريات المواد والكميات',
            'icon': 'fa-boxes',
            'color': 'warning',
            'url': reverse('purchases:items_purchases_report')
        },
        {
            'title': 'العقود',
            'description': 'تقرير بعقود الشراء وحالاتها',
            'icon': 'fa-file-contract',
            'color': 'secondary',
            'url': reverse('purchases:contracts_report')
        },
    ]

    context = {
        'title': _('تقارير المشتريات'),
        'reports': reports,
        'breadcrumbs': [
            {'title': _('المشتريات'), 'url': reverse('purchases:dashboard')},
            {'title': _('التقارير'), 'url': ''},
        ]
    }

    return render(request, 'purchases/reports/reports_list.html', context)


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_purchases_summary_excel(request):
    """تصدير تقرير ملخص المشتريات إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    invoice_type = request.GET.get('invoice_type', 'purchase')

    # الفواتير
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type=invoice_type
    ).select_related('supplier', 'currency', 'branch')

    if supplier_id:
        invoices = invoices.filter(supplier_id=supplier_id)
    if date_from:
        invoices = invoices.filter(date__gte=date_from)
    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # الإحصائيات
    stats = invoices.aggregate(
        total_count=Count('id'),
        total_amount=Coalesce(Sum('total_with_tax'), Decimal('0')),
        total_tax=Coalesce(Sum('tax_amount'), Decimal('0')),
        total_discount=Coalesce(Sum('discount_amount'), Decimal('0')),
        avg_invoice=Coalesce(Avg('total_with_tax'), Decimal('0'))
    )

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "ملخص المشتريات"

    # التنسيق
    title_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # العنوان
    ws.merge_cells(f'A{current_row}:H{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "تقرير ملخص المشتريات"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # معلومات الشركة
    ws[f'A{current_row}'] = "الشركة:"
    ws[f'B{current_row}'] = company.name
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws[f'A{current_row}'] = "تاريخ التقرير:"
    ws[f'B{current_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2

    # الإحصائيات
    ws.merge_cells(f'A{current_row}:H{current_row}')
    stats_cell = ws[f'A{current_row}']
    stats_cell.value = "الإحصائيات الإجمالية"
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws[f'A{current_row}'] = "عدد الفواتير:"
    ws[f'B{current_row}'] = stats['total_count']
    ws[f'D{current_row}'] = "إجمالي المبلغ:"
    ws[f'E{current_row}'] = float(stats['total_amount'])
    current_row += 1

    ws[f'A{current_row}'] = "إجمالي الضريبة:"
    ws[f'B{current_row}'] = float(stats['total_tax'])
    ws[f'D{current_row}'] = "إجمالي الخصم:"
    ws[f'E{current_row}'] = float(stats['total_discount'])
    current_row += 1

    ws[f'A{current_row}'] = "متوسط الفاتورة:"
    ws[f'B{current_row}'] = float(stats['avg_invoice'])
    current_row += 2

    # جدول الفواتير
    headers = ['رقم الفاتورة', 'التاريخ', 'المورد', 'المبلغ قبل الضريبة', 'الضريبة', 'الخصم', 'المبلغ الإجمالي', 'الحالة']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات
    for invoice in invoices.order_by('-date'):
        ws.cell(row=current_row, column=1, value=invoice.number)
        ws.cell(row=current_row, column=2, value=invoice.date.strftime('%Y-%m-%d') if invoice.date else '')
        ws.cell(row=current_row, column=3, value=invoice.supplier.name if invoice.supplier else '')
        ws.cell(row=current_row, column=4, value=float(invoice.subtotal_after_discount))
        ws.cell(row=current_row, column=5, value=float(invoice.tax_amount))
        ws.cell(row=current_row, column=6, value=float(invoice.discount_amount))
        ws.cell(row=current_row, column=7, value=float(invoice.total_with_tax))
        ws.cell(row=current_row, column=8, value='مرحلة' if invoice.is_posted else 'مسودة')

        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    # ضبط عرض الأعمدة
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="purchases_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_purchases_summary_pdf(request):
    """تصدير تقرير ملخص المشتريات إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    invoice_type = request.GET.get('invoice_type', 'purchase')

    # الفواتير
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type=invoice_type
    ).select_related('supplier', 'currency', 'branch')

    if supplier_id:
        invoices = invoices.filter(supplier_id=supplier_id)
    if date_from:
        invoices = invoices.filter(date__gte=date_from)
    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # الإحصائيات
    stats = invoices.aggregate(
        total_count=Count('id'),
        total_amount=Coalesce(Sum('total_with_tax'), Decimal('0')),
        total_tax=Coalesce(Sum('tax_amount'), Decimal('0')),
        total_discount=Coalesce(Sum('discount_amount'), Decimal('0')),
        avg_invoice=Coalesce(Avg('total_with_tax'), Decimal('0'))
    )

    # تسجيل الخط العربي
    try:
        from django.conf import settings
        import os

        # Try different font paths
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]

        arabic_font = 'Helvetica'  # Default fallback

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

    except:
        arabic_font = 'Helvetica'

    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح في PDF"""
        if not text:
            return ''
        text_str = str(text)
        if arabic_font == 'ArabicFont':
            try:
                reshaped_text = arabic_reshaper.reshape(text_str)
                return get_display(reshaped_text)
            except Exception as e:
                # If reshaping fails, return original text
                return text_str
        return text_str

    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # نمط العنوان
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        textColor=colors.HexColor('#0D6EFD'),
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # نمط النص العادي
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_RIGHT
    )

    # نمط لمحتوى الجدول
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )

    # نمط لرؤوس الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=9,
        alignment=TA_CENTER,
        leading=11,
        textColor=colors.whitesmoke
    )

    # العنوان
    title = Paragraph(arabic_text('تقرير ملخص المشتريات'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # معلومات الشركة
    info_style = ParagraphStyle(
        'InfoStyle',
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_CENTER
    )

    info_data = [[
        Paragraph(arabic_text(f"الشركة: {company.name}"), info_style),
        Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), info_style),
    ]]

    info_table = Table(info_data, colWidths=[4*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # الإحصائيات
    stats_data = [
        [Paragraph(arabic_text('الإحصائيات الإجمالية'), table_header_style)],
        [Paragraph(arabic_text(f"عدد الفواتير: {stats['total_count']}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي المبلغ: {stats['total_amount']:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي الضريبة: {stats['total_tax']:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي الخصم: {stats['total_discount']:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"متوسط الفاتورة: {stats['avg_invoice']:,.2f}"), table_cell_style)],
    ]

    stats_table = Table(stats_data, colWidths=[8*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول الفواتير - Headers
    table_data = [[
        Paragraph(arabic_text('الحالة'), table_header_style),
        Paragraph(arabic_text('المبلغ الإجمالي'), table_header_style),
        Paragraph(arabic_text('الخصم'), table_header_style),
        Paragraph(arabic_text('الضريبة'), table_header_style),
        Paragraph(arabic_text('المبلغ بعد الخصم'), table_header_style),
        Paragraph(arabic_text('المورد'), table_header_style),
        Paragraph(arabic_text('التاريخ'), table_header_style),
        Paragraph(arabic_text('رقم الفاتورة'), table_header_style),
    ]]

    # جدول الفواتير - Data
    for invoice in invoices.order_by('-date')[:50]:  # أول 50 فاتورة
        table_data.append([
            Paragraph(arabic_text('مرحلة' if invoice.is_posted else 'مسودة'), table_cell_style),
            Paragraph(f"{float(invoice.total_with_tax):,.2f}", table_cell_style),
            Paragraph(f"{float(invoice.discount_amount):,.2f}", table_cell_style),
            Paragraph(f"{float(invoice.tax_amount):,.2f}", table_cell_style),
            Paragraph(f"{float(invoice.subtotal_after_discount):,.2f}", table_cell_style),
            Paragraph(arabic_text(invoice.supplier.name if invoice.supplier else ''), table_cell_style),
            Paragraph(invoice.date.strftime('%Y-%m-%d') if invoice.date else '', table_cell_style),
            Paragraph(arabic_text(invoice.number), table_cell_style),
        ])

    invoice_table = Table(table_data, colWidths=[0.8*inch, 1*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.5*inch, 1*inch, 1.2*inch])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(invoice_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="purchases_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_supplier_performance_excel(request):
    """تصدير تقرير أداء الموردين إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # الفواتير المرحلة
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type='purchase'
    )

    if date_from:
        invoices = invoices.filter(date__gte=date_from)
    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # أداء الموردين
    supplier_performance = invoices.values(
        'supplier__id',
        'supplier__name',
        'supplier__code',
        'supplier__phone',
        'supplier__email'
    ).annotate(
        invoice_count=Count('id'),
        total_purchases=Coalesce(Sum('total_with_tax'), Decimal('0')),
        avg_invoice_value=Coalesce(Avg('total_with_tax'), Decimal('0')),
        total_discounts=Coalesce(Sum('discount_amount'), Decimal('0')),
        total_taxes=Coalesce(Sum('tax_amount'), Decimal('0')),
    ).order_by('-total_purchases')

    # الأوامر
    orders = PurchaseOrder.objects.filter(
        company=company,
        status__in=['approved', 'sent', 'partial', 'completed']
    )
    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "أداء الموردين"

    # التنسيق
    title_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # العنوان
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "تقرير أداء الموردين"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # معلومات الشركة
    ws[f'A{current_row}'] = "الشركة:"
    ws[f'B{current_row}'] = company.name
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws[f'A{current_row}'] = "تاريخ التقرير:"
    ws[f'B{current_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2

    # الإحصائيات
    total_suppliers = supplier_performance.count()
    total_purchases = sum(s['total_purchases'] for s in supplier_performance)
    total_invoices = sum(s['invoice_count'] for s in supplier_performance)

    ws.merge_cells(f'A{current_row}:J{current_row}')
    stats_cell = ws[f'A{current_row}']
    stats_cell.value = "الإحصائيات الإجمالية"
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws[f'A{current_row}'] = "عدد الموردين:"
    ws[f'B{current_row}'] = total_suppliers
    ws[f'D{current_row}'] = "إجمالي المشتريات:"
    ws[f'E{current_row}'] = float(total_purchases)
    current_row += 1

    ws[f'A{current_row}'] = "عدد الفواتير:"
    ws[f'B{current_row}'] = total_invoices
    ws[f'D{current_row}'] = "متوسط قيمة الفاتورة:"
    ws[f'E{current_row}'] = float(total_purchases / total_invoices) if total_invoices > 0 else 0
    current_row += 2

    # جدول الموردين
    headers = ['كود المورد', 'اسم المورد', 'الهاتف', 'البريد الإلكتروني',
               'عدد الفواتير', 'إجمالي المشتريات', 'متوسط الفاتورة',
               'إجمالي الخصومات', 'إجمالي الضرائب', 'عدد الأوامر']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات
    for supplier in supplier_performance:
        supplier_id = supplier['supplier__id']
        order_count = orders.filter(supplier_id=supplier_id).count()

        ws.cell(row=current_row, column=1, value=supplier['supplier__code'] or '')
        ws.cell(row=current_row, column=2, value=supplier['supplier__name'] or '')
        ws.cell(row=current_row, column=3, value=supplier['supplier__phone'] or '')
        ws.cell(row=current_row, column=4, value=supplier['supplier__email'] or '')
        ws.cell(row=current_row, column=5, value=supplier['invoice_count'])
        ws.cell(row=current_row, column=6, value=float(supplier['total_purchases']))
        ws.cell(row=current_row, column=7, value=float(supplier['avg_invoice_value']))
        ws.cell(row=current_row, column=8, value=float(supplier['total_discounts']))
        ws.cell(row=current_row, column=9, value=float(supplier['total_taxes']))
        ws.cell(row=current_row, column=10, value=order_count)

        for col in range(1, 11):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    # ضبط عرض الأعمدة
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="supplier_performance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_supplier_performance_pdf(request):
    """تصدير تقرير أداء الموردين إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # الفواتير المرحلة
    invoices = PurchaseInvoice.objects.filter(
        company=company,
        is_posted=True,
        invoice_type='purchase'
    )

    if date_from:
        invoices = invoices.filter(date__gte=date_from)
    if date_to:
        invoices = invoices.filter(date__lte=date_to)

    # أداء الموردين
    supplier_performance = invoices.values(
        'supplier__id',
        'supplier__name',
        'supplier__code',
    ).annotate(
        invoice_count=Count('id'),
        total_purchases=Coalesce(Sum('total_with_tax'), Decimal('0')),
        avg_invoice_value=Coalesce(Avg('total_with_tax'), Decimal('0')),
        total_discounts=Coalesce(Sum('discount_amount'), Decimal('0')),
    ).order_by('-total_purchases')

    # تسجيل الخط العربي
    try:
        from django.conf import settings
        import os

        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            'C:\\Windows\\Fonts\\arial.ttf',
        ]

        arabic_font = 'Helvetica'

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

    except:
        arabic_font = 'Helvetica'

    def arabic_text(text):
        if not text:
            return ''
        if arabic_font == 'ArabicFont':
            try:
                reshaped_text = arabic_reshaper.reshape(str(text))
                return get_display(reshaped_text)
            except:
                return str(text)
        return str(text)

    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # نمط العنوان
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        textColor=colors.HexColor('#28A745'),
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # نمط لمحتوى الجدول
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )

    # نمط لرؤوس الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=9,
        alignment=TA_CENTER,
        leading=11,
        textColor=colors.whitesmoke
    )

    # العنوان
    title = Paragraph(arabic_text('تقرير أداء الموردين'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # معلومات الشركة
    info_style = ParagraphStyle(
        'InfoStyle',
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_CENTER
    )

    info_data = [[
        Paragraph(arabic_text(f"الشركة: {company.name}"), info_style),
        Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), info_style),
    ]]

    info_table = Table(info_data, colWidths=[4*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # الإحصائيات
    total_suppliers = supplier_performance.count()
    total_purchases = sum(s['total_purchases'] for s in supplier_performance)
    total_invoices = sum(s['invoice_count'] for s in supplier_performance)

    stats_data = [
        [Paragraph(arabic_text('الإحصائيات الإجمالية'), table_header_style)],
        [Paragraph(arabic_text(f"عدد الموردين: {total_suppliers}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي المشتريات: {total_purchases:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"عدد الفواتير: {total_invoices}"), table_cell_style)],
        [Paragraph(arabic_text(f"متوسط قيمة الفاتورة: {(total_purchases / total_invoices if total_invoices > 0 else 0):,.2f}"), table_cell_style)],
    ]

    stats_table = Table(stats_data, colWidths=[8*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول الموردين
    table_data = [[
        Paragraph(arabic_text('الخصومات'), table_header_style),
        Paragraph(arabic_text('متوسط الفاتورة'), table_header_style),
        Paragraph(arabic_text('إجمالي المشتريات'), table_header_style),
        Paragraph(arabic_text('عدد الفواتير'), table_header_style),
        Paragraph(arabic_text('الكود'), table_header_style),
        Paragraph(arabic_text('اسم المورد'), table_header_style),
    ]]

    for supplier in list(supplier_performance)[:30]:  # أول 30 مورد
        table_data.append([
            Paragraph(f"{float(supplier['total_discounts']):,.0f}", table_cell_style),
            Paragraph(f"{float(supplier['avg_invoice_value']):,.0f}", table_cell_style),
            Paragraph(f"{float(supplier['total_purchases']):,.0f}", table_cell_style),
            Paragraph(f"{supplier['invoice_count']}", table_cell_style),
            Paragraph(arabic_text(supplier['supplier__code'] or '-'), table_cell_style),
            Paragraph(arabic_text(supplier['supplier__name'] or ''), table_cell_style),
        ])

    supplier_table = Table(table_data, colWidths=[1*inch, 1.2*inch, 1.5*inch, 1*inch, 1*inch, 2.5*inch])
    supplier_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(supplier_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="supplier_performance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseorder')
def export_purchase_orders_excel(request):
    """تصدير تقرير أوامر الشراء إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    orders = PurchaseOrder.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'branch', 'created_by')

    if status:
        orders = orders.filter(status=status)
    if supplier_id:
        orders = orders.filter(supplier_id=supplier_id)
    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)

    # الإحصائيات
    stats = orders.aggregate(
        total_count=Count('id'),
        total_sum=Coalesce(Sum('total_amount'), Decimal('0')),
        avg_order=Coalesce(Avg('total_amount'), Decimal('0'))
    )

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "أوامر الشراء"

    # التنسيق
    title_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # العنوان
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "تقرير أوامر الشراء"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # معلومات الشركة
    ws[f'A{current_row}'] = "الشركة:"
    ws[f'B{current_row}'] = company.name
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws[f'A{current_row}'] = "تاريخ التقرير:"
    ws[f'B{current_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2

    # الإحصائيات
    ws.merge_cells(f'A{current_row}:J{current_row}')
    stats_cell = ws[f'A{current_row}']
    stats_cell.value = "الإحصائيات الإجمالية"
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws[f'A{current_row}'] = "عدد الأوامر:"
    ws[f'B{current_row}'] = stats['total_count']
    ws[f'D{current_row}'] = "إجمالي القيمة:"
    ws[f'E{current_row}'] = float(stats['total_sum'])
    current_row += 1

    ws[f'A{current_row}'] = "متوسط قيمة الأمر:"
    ws[f'B{current_row}'] = float(stats['avg_order'])
    current_row += 2

    # جدول الأوامر
    headers = ['التاريخ', 'رقم الأمر', 'المورد', 'الفرع', 'المبلغ الإجمالي', 'الحالة', 'المستخدم']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات
    status_translations = {
        'draft': 'مسودة',
        'pending_approval': 'بانتظار الموافقة',
        'approved': 'معتمد',
        'rejected': 'مرفوض',
        'sent': 'مرسل للمورد',
        'partial': 'استلام جزئي',
        'completed': 'مكتمل',
        'cancelled': 'ملغي'
    }

    for order in orders.order_by('-date'):
        ws.cell(row=current_row, column=1, value=order.date.strftime('%Y-%m-%d') if order.date else '')
        ws.cell(row=current_row, column=2, value=order.number)
        ws.cell(row=current_row, column=3, value=order.supplier.name if order.supplier else '')
        ws.cell(row=current_row, column=4, value=order.branch.name if order.branch else '')
        ws.cell(row=current_row, column=5, value=float(order.total_amount))
        ws.cell(row=current_row, column=6, value=status_translations.get(order.status, order.status))
        ws.cell(row=current_row, column=7, value=order.created_by.get_full_name() if order.created_by else '')

        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    # ضبط عرض الأعمدة
    column_widths = [15, 18, 25, 20, 20, 15, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseorder')
def export_purchase_orders_pdf(request):
    """تصدير تقرير أوامر الشراء إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    orders = PurchaseOrder.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'branch', 'created_by')

    if status:
        orders = orders.filter(status=status)
    if supplier_id:
        orders = orders.filter(supplier_id=supplier_id)
    if date_from:
        orders = orders.filter(date__gte=date_from)
    if date_to:
        orders = orders.filter(date__lte=date_to)

    # الإحصائيات
    stats = orders.aggregate(
        total_count=Count('id'),
        total_sum=Coalesce(Sum('total_amount'), Decimal('0')),
        avg_order=Coalesce(Avg('total_amount'), Decimal('0'))
    )

    # تسجيل الخط العربي
    try:
        from django.conf import settings
        import os

        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            'C:\\Windows\\Fonts\\arial.ttf',
        ]

        arabic_font = 'Helvetica'

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

    except:
        arabic_font = 'Helvetica'

    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح في PDF"""
        if not text:
            return ''
        text_str = str(text)
        if arabic_font == 'ArabicFont':
            try:
                reshaped_text = arabic_reshaper.reshape(text_str)
                return get_display(reshaped_text)
            except Exception as e:
                # If reshaping fails, return original text
                return text_str
        return text_str

    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # نمط العنوان
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        textColor=colors.HexColor('#17A2B8'),
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # نمط لمحتوى الجدول
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )

    # نمط لرؤوس الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=9,
        alignment=TA_CENTER,
        leading=11,
        textColor=colors.whitesmoke
    )

    # العنوان
    title = Paragraph(arabic_text('تقرير أوامر الشراء'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # معلومات الشركة
    info_style = ParagraphStyle(
        'InfoStyle',
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_CENTER
    )

    info_data = [[
        Paragraph(arabic_text(f"الشركة: {company.name}"), info_style),
        Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), info_style),
    ]]

    info_table = Table(info_data, colWidths=[4*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # الإحصائيات
    stats_data = [
        [Paragraph(arabic_text('الإحصائيات الإجمالية'), table_header_style)],
        [Paragraph(arabic_text(f"عدد الأوامر: {stats['total_count']}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي القيمة: {stats['total_sum']:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"متوسط قيمة الأمر: {stats['avg_order']:,.2f}"), table_cell_style)],
    ]

    stats_table = Table(stats_data, colWidths=[8*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول الأوامر - Headers
    status_translations = {
        'draft': 'مسودة',
        'pending_approval': 'بانتظار الموافقة',
        'approved': 'معتمد',
        'rejected': 'مرفوض',
        'sent': 'مرسل للمورد',
        'partial': 'استلام جزئي',
        'completed': 'مكتمل',
        'cancelled': 'ملغي'
    }

    table_data = [[
        Paragraph(arabic_text('المستخدم'), table_header_style),
        Paragraph(arabic_text('الحالة'), table_header_style),
        Paragraph(arabic_text('الإجمالي'), table_header_style),
        Paragraph(arabic_text('الفرع'), table_header_style),
        Paragraph(arabic_text('المورد'), table_header_style),
        Paragraph(arabic_text('التاريخ'), table_header_style),
        Paragraph(arabic_text('رقم الأمر'), table_header_style),
    ]]

    # جدول الأوامر - Data
    for order in orders.order_by('-date')[:50]:  # أول 50 أمر
        table_data.append([
            Paragraph(arabic_text(order.created_by.get_full_name() if order.created_by else ''), table_cell_style),
            Paragraph(arabic_text(status_translations.get(order.status, order.status)), table_cell_style),
            Paragraph(f"{float(order.total_amount):,.2f}", table_cell_style),
            Paragraph(arabic_text(order.branch.name if order.branch else ''), table_cell_style),
            Paragraph(arabic_text(order.supplier.name if order.supplier else ''), table_cell_style),
            Paragraph(order.date.strftime('%Y-%m-%d') if order.date else '', table_cell_style),
            Paragraph(arabic_text(order.number), table_cell_style),
        ])

    order_table = Table(table_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.5*inch, 1*inch, 1.2*inch])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(order_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_items_purchases_excel(request):
    """تصدير تقرير مشتريات المواد إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    item_id = request.GET.get('item')
    category_id = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    from apps.purchases.models import PurchaseInvoiceItem

    # سطور الفواتير المرحلة
    invoice_lines = PurchaseInvoiceItem.objects.filter(
        invoice__company=company,
        invoice__is_posted=True,
        invoice__invoice_type='purchase'
    ).select_related('item', 'invoice', 'invoice__supplier', 'unit')

    if item_id:
        invoice_lines = invoice_lines.filter(item_id=item_id)
    if category_id:
        invoice_lines = invoice_lines.filter(item__category_id=category_id)
    if date_from:
        invoice_lines = invoice_lines.filter(invoice__date__gte=date_from)
    if date_to:
        invoice_lines = invoice_lines.filter(invoice__date__lte=date_to)

    # حسب المادة
    by_item = invoice_lines.values(
        'item__id',
        'item__code',
        'item__name',
        'unit__code'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        avg_price=Coalesce(Avg('unit_price'), Decimal('0')),
        invoice_count=Count('invoice', distinct=True),
        supplier_count=Count('invoice__supplier', distinct=True)
    ).order_by('-total_amount')

    # حسب الفئة
    by_category = invoice_lines.filter(
        item__category__isnull=False
    ).values(
        'item__category__id',
        'item__category__name'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        item_count=Count('item', distinct=True)
    ).order_by('-total_amount')

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "مشتريات المواد"

    # التنسيق
    title_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # العنوان
    ws.merge_cells(f'A{current_row}:H{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "تقرير مشتريات المواد"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # معلومات الشركة
    ws[f'A{current_row}'] = "الشركة:"
    ws[f'B{current_row}'] = company.name
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws[f'A{current_row}'] = "تاريخ التقرير:"
    ws[f'B{current_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2

    # الإحصائيات
    total_items = by_item.count()
    total_quantity = sum(item['quantity'] for item in by_item)
    total_amount = sum(item['total_amount'] for item in by_item)
    total_categories = by_category.count()

    ws.merge_cells(f'A{current_row}:H{current_row}')
    stats_cell = ws[f'A{current_row}']
    stats_cell.value = "الإحصائيات الإجمالية"
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws[f'A{current_row}'] = "عدد المواد:"
    ws[f'B{current_row}'] = total_items
    ws[f'D{current_row}'] = "إجمالي الكمية:"
    ws[f'E{current_row}'] = float(total_quantity)
    current_row += 1

    ws[f'A{current_row}'] = "إجمالي المبلغ:"
    ws[f'B{current_row}'] = float(total_amount)
    ws[f'D{current_row}'] = "عدد الفئات:"
    ws[f'E{current_row}'] = total_categories
    current_row += 2

    # جدول المواد
    ws.merge_cells(f'A{current_row}:H{current_row}')
    section_cell = ws[f'A{current_row}']
    section_cell.value = "المشتريات حسب المادة"
    section_cell.font = Font(bold=True, size=11)
    section_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    headers = ['كود المادة', 'اسم المادة', 'الوحدة', 'الكمية',
               'إجمالي المبلغ', 'متوسط السعر', 'عدد الفواتير', 'عدد الموردين']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات - المواد
    for item in by_item:
        ws.cell(row=current_row, column=1, value=item['item__code'] or '')
        ws.cell(row=current_row, column=2, value=item['item__name'] or '')
        ws.cell(row=current_row, column=3, value=item['unit__code'] or '')
        ws.cell(row=current_row, column=4, value=float(item['quantity']))
        ws.cell(row=current_row, column=5, value=float(item['total_amount']))
        ws.cell(row=current_row, column=6, value=float(item['avg_price']))
        ws.cell(row=current_row, column=7, value=item['invoice_count'])
        ws.cell(row=current_row, column=8, value=item['supplier_count'])

        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    current_row += 1

    # جدول الفئات
    ws.merge_cells(f'A{current_row}:E{current_row}')
    section_cell = ws[f'A{current_row}']
    section_cell.value = "المشتريات حسب الفئة"
    section_cell.font = Font(bold=True, size=11)
    section_cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    section_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    category_headers = ['اسم الفئة', 'عدد المواد', 'الكمية', 'إجمالي المبلغ']

    for col_num, header in enumerate(category_headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات - الفئات
    for category in by_category:
        ws.cell(row=current_row, column=1, value=category['item__category__name'] or '')
        ws.cell(row=current_row, column=2, value=category['item_count'])
        ws.cell(row=current_row, column=3, value=float(category['quantity']))
        ws.cell(row=current_row, column=4, value=float(category['total_amount']))

        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="items_purchases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required_with_message('purchases.view_purchaseinvoice')
def export_items_purchases_pdf(request):
    """تصدير تقرير مشتريات المواد إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company

    # الفلاتر
    item_id = request.GET.get('item')
    category_id = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    from apps.purchases.models import PurchaseInvoiceItem

    # سطور الفواتير المرحلة
    invoice_lines = PurchaseInvoiceItem.objects.filter(
        invoice__company=company,
        invoice__is_posted=True,
        invoice__invoice_type='purchase'
    ).select_related('item', 'invoice', 'invoice__supplier', 'unit')

    if item_id:
        invoice_lines = invoice_lines.filter(item_id=item_id)
    if category_id:
        invoice_lines = invoice_lines.filter(item__category_id=category_id)
    if date_from:
        invoice_lines = invoice_lines.filter(invoice__date__gte=date_from)
    if date_to:
        invoice_lines = invoice_lines.filter(invoice__date__lte=date_to)

    # حسب المادة
    by_item = invoice_lines.values(
        'item__id',
        'item__code',
        'item__name',
        'unit__code'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        avg_price=Coalesce(Avg('unit_price'), Decimal('0')),
        invoice_count=Count('invoice', distinct=True),
    ).order_by('-total_amount')

    # حسب الفئة
    by_category = invoice_lines.filter(
        item__category__isnull=False
    ).values(
        'item__category__name'
    ).annotate(
        quantity=Coalesce(Sum('quantity'), Decimal('0')),
        total_amount=Coalesce(Sum(F('subtotal') + F('tax_amount')), Decimal('0')),
        item_count=Count('item', distinct=True)
    ).order_by('-total_amount')

    # تسجيل الخط العربي
    try:
        from django.conf import settings
        import os

        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            'C:\\Windows\\Fonts\\arial.ttf',
        ]

        arabic_font = 'Helvetica'

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

    except:
        arabic_font = 'Helvetica'

    def arabic_text(text):
        if not text:
            return ''
        if arabic_font == 'ArabicFont':
            try:
                reshaped_text = arabic_reshaper.reshape(str(text))
                return get_display(reshaped_text)
            except:
                return str(text)
        return str(text)

    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # نمط العنوان
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        textColor=colors.HexColor('#17A2B8'),
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # نمط لمحتوى الجدول
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )

    # نمط لرؤوس الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=9,
        alignment=TA_CENTER,
        leading=11,
        textColor=colors.whitesmoke
    )

    # العنوان
    title = Paragraph(arabic_text('تقرير مشتريات المواد'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # معلومات الشركة
    info_style = ParagraphStyle(
        'InfoStyle',
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_CENTER
    )

    info_data = [[
        Paragraph(arabic_text(f"الشركة: {company.name}"), info_style),
        Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), info_style),
    ]]

    info_table = Table(info_data, colWidths=[4*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # الإحصائيات
    total_items = by_item.count()
    total_quantity = sum(item['quantity'] for item in by_item)
    total_amount = sum(item['total_amount'] for item in by_item)
    total_categories = by_category.count()

    stats_data = [
        [Paragraph(arabic_text('الإحصائيات الإجمالية'), table_header_style)],
        [Paragraph(arabic_text(f"عدد المواد: {total_items}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي الكمية: {total_quantity:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي المبلغ: {total_amount:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"عدد الفئات: {total_categories}"), table_cell_style)],
    ]

    stats_table = Table(stats_data, colWidths=[8*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول المواد
    table_data = [[
        Paragraph(arabic_text('عدد الفواتير'), table_header_style),
        Paragraph(arabic_text('متوسط السعر'), table_header_style),
        Paragraph(arabic_text('إجمالي المبلغ'), table_header_style),
        Paragraph(arabic_text('الكمية'), table_header_style),
        Paragraph(arabic_text('الوحدة'), table_header_style),
        Paragraph(arabic_text('الكود'), table_header_style),
        Paragraph(arabic_text('اسم المادة'), table_header_style),
    ]]

    for item in list(by_item)[:30]:  # أول 30 مادة
        table_data.append([
            Paragraph(f"{item['invoice_count']}", table_cell_style),
            Paragraph(f"{float(item['avg_price']):,.0f}", table_cell_style),
            Paragraph(f"{float(item['total_amount']):,.0f}", table_cell_style),
            Paragraph(f"{float(item['quantity']):,.0f}", table_cell_style),
            Paragraph(arabic_text(item['unit__code'] or '-'), table_cell_style),
            Paragraph(arabic_text(item['item__code'] or '-'), table_cell_style),
            Paragraph(arabic_text(item['item__name'] or ''), table_cell_style),
        ])

    items_table = Table(table_data, colWidths=[0.8*inch, 1*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1*inch, 2.5*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(items_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="items_purchases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


@login_required
@permission_required_with_message('purchases.view_purchasecontract')
def export_contracts_excel(request):
    """تصدير تقرير العقود إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company
    today = date.today()

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')

    contracts = PurchaseContract.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'created_by')

    if status:
        contracts = contracts.filter(status=status)
    if supplier_id:
        contracts = contracts.filter(supplier_id=supplier_id)

    # الإحصائيات
    stats = contracts.aggregate(
        total_count=Count('id'),
        total_value=Coalesce(Sum('contract_value'), Decimal('0')),
        avg_value=Coalesce(Avg('contract_value'), Decimal('0'))
    )

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير العقود"

    # التنسيق
    title_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # العنوان
    ws.merge_cells(f'A{current_row}:I{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "تقرير عقود الشراء"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # معلومات الشركة
    ws[f'A{current_row}'] = "الشركة:"
    ws[f'B{current_row}'] = company.name
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    ws[f'A{current_row}'] = "تاريخ التقرير:"
    ws[f'B{current_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 2

    # الإحصائيات
    ws.merge_cells(f'A{current_row}:I{current_row}')
    stats_cell = ws[f'A{current_row}']
    stats_cell.value = "الإحصائيات الإجمالية"
    stats_cell.font = Font(bold=True, size=12)
    stats_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    stats_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws[f'A{current_row}'] = "عدد العقود:"
    ws[f'B{current_row}'] = stats['total_count']
    ws[f'D{current_row}'] = "إجمالي القيمة:"
    ws[f'E{current_row}'] = float(stats['total_value'])
    current_row += 1

    ws[f'A{current_row}'] = "متوسط قيمة العقد:"
    ws[f'B{current_row}'] = float(stats['avg_value'])
    current_row += 2

    # جدول العقود
    headers = ['رقم العقد', 'المورد', 'تاريخ العقد', 'تاريخ البدء', 'تاريخ الانتهاء',
               'قيمة العقد', 'الحالة', 'معتمد', 'المستخدم']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    current_row += 1

    # البيانات
    status_translations = {
        'draft': 'مسودة',
        'active': 'نشط',
        'suspended': 'معلق',
        'completed': 'مكتمل',
        'terminated': 'منهي',
        'expired': 'منتهي الصلاحية'
    }

    for contract in contracts.order_by('-start_date'):
        ws.cell(row=current_row, column=1, value=contract.number)
        ws.cell(row=current_row, column=2, value=contract.supplier.name if contract.supplier else '')
        ws.cell(row=current_row, column=3, value=contract.contract_date.strftime('%Y-%m-%d') if contract.contract_date else '')
        ws.cell(row=current_row, column=4, value=contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '')
        ws.cell(row=current_row, column=5, value=contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '')
        ws.cell(row=current_row, column=6, value=float(contract.contract_value))
        ws.cell(row=current_row, column=7, value=status_translations.get(contract.status, contract.status))
        ws.cell(row=current_row, column=8, value='نعم' if contract.approved else 'لا')
        ws.cell(row=current_row, column=9, value=contract.created_by.get_full_name() if contract.created_by else '')

        for col in range(1, 10):
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        current_row += 1

    # ضبط عرض الأعمدة
    column_widths = [18, 25, 15, 15, 15, 18, 15, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required_with_message('purchases.view_purchasecontract')
def export_contracts_pdf(request):
    """تصدير تقرير العقود إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import arabic_reshaper
    from bidi.algorithm import get_display

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('purchases:dashboard')

    company = request.current_company
    today = date.today()

    # الفلاتر
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')

    contracts = PurchaseContract.objects.filter(
        company=company
    ).select_related('supplier', 'currency', 'created_by')

    if status:
        contracts = contracts.filter(status=status)
    if supplier_id:
        contracts = contracts.filter(supplier_id=supplier_id)

    # الإحصائيات
    stats = contracts.aggregate(
        total_count=Count('id'),
        total_value=Coalesce(Sum('contract_value'), Decimal('0')),
        avg_value=Coalesce(Avg('contract_value'), Decimal('0'))
    )

    # تسجيل الخط العربي
    try:
        from django.conf import settings
        import os

        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            'C:\\Windows\\Fonts\\arial.ttf',
        ]

        arabic_font = 'Helvetica'

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

    except:
        arabic_font = 'Helvetica'

    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح في PDF"""
        if not text:
            return ''
        if arabic_font == 'ArabicFont':
            try:
                reshaped_text = arabic_reshaper.reshape(str(text))
                return get_display(reshaped_text)
            except:
                return str(text)
        return str(text)

    # إنشاء PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    # نمط العنوان
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=arabic_font,
        fontSize=18,
        textColor=colors.HexColor('#6C757D'),
        alignment=TA_CENTER,
        spaceAfter=20
    )

    # نمط لمحتوى الجدول
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=7,
        alignment=TA_CENTER,
        leading=9
    )

    # نمط لرؤوس الجدول
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName=arabic_font,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
        textColor=colors.whitesmoke
    )

    # العنوان
    title = Paragraph(arabic_text('تقرير عقود الشراء'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))

    # معلومات الشركة
    info_style = ParagraphStyle(
        'InfoStyle',
        fontName=arabic_font,
        fontSize=10,
        alignment=TA_CENTER
    )

    info_data = [[
        Paragraph(arabic_text(f"الشركة: {company.name}"), info_style),
        Paragraph(arabic_text(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), info_style),
    ]]

    info_table = Table(info_data, colWidths=[4*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))

    # الإحصائيات
    stats_data = [
        [Paragraph(arabic_text('الإحصائيات الإجمالية'), table_header_style)],
        [Paragraph(arabic_text(f"عدد العقود: {stats['total_count']}"), table_cell_style)],
        [Paragraph(arabic_text(f"إجمالي القيمة: {stats['total_value']:,.2f}"), table_cell_style)],
        [Paragraph(arabic_text(f"متوسط قيمة العقد: {stats['avg_value']:,.2f}"), table_cell_style)],
    ]

    stats_table = Table(stats_data, colWidths=[8*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول العقود - Headers
    status_translations = {
        'draft': 'مسودة',
        'active': 'نشط',
        'suspended': 'معلق',
        'completed': 'مكتمل',
        'terminated': 'منهي',
        'expired': 'منتهي الصلاحية'
    }

    table_data = [[
        Paragraph(arabic_text('المستخدم'), table_header_style),
        Paragraph(arabic_text('معتمد'), table_header_style),
        Paragraph(arabic_text('الحالة'), table_header_style),
        Paragraph(arabic_text('القيمة'), table_header_style),
        Paragraph(arabic_text('الانتهاء'), table_header_style),
        Paragraph(arabic_text('البدء'), table_header_style),
        Paragraph(arabic_text('المورد'), table_header_style),
        Paragraph(arabic_text('رقم العقد'), table_header_style),
    ]]

    # جدول العقود - Data
    for contract in contracts.order_by('-start_date')[:50]:  # أول 50 عقد
        table_data.append([
            Paragraph(arabic_text(contract.created_by.get_full_name() if contract.created_by else ''), table_cell_style),
            Paragraph(arabic_text('نعم' if contract.approved else 'لا'), table_cell_style),
            Paragraph(arabic_text(status_translations.get(contract.status, contract.status)), table_cell_style),
            Paragraph(f"{float(contract.contract_value):,.0f}", table_cell_style),
            Paragraph(contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '', table_cell_style),
            Paragraph(contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '', table_cell_style),
            Paragraph(arabic_text(contract.supplier.name if contract.supplier else ''), table_cell_style),
            Paragraph(arabic_text(contract.number), table_cell_style),
        ])

    contracts_table = Table(table_data, colWidths=[1*inch, 0.7*inch, 0.9*inch, 1*inch, 0.9*inch, 0.9*inch, 1.5*inch, 1.2*inch])
    contracts_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(contracts_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="contracts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response
