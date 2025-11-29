# apps/purchases/views/quotation_views.py
"""
Views for Purchase Quotations
Complete CRUD operations + RFQ Workflow + Quotation Evaluation + Award Logic
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Count, F, Avg
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from django.core.exceptions import PermissionDenied, ValidationError
from decimal import Decimal
from datetime import datetime, date
from io import BytesIO

from ..models import (
    PurchaseQuotationRequest, PurchaseQuotationRequestItem,
    PurchaseQuotation, PurchaseQuotationItem,
    PurchaseRequest, PurchaseOrder
)
from ..forms import (
    PurchaseQuotationRequestForm,
    PurchaseQuotationRequestItemForm,
    PurchaseQuotationRequestItemFormSet,
    PurchaseQuotationForm,
    PurchaseQuotationItemForm,
    PurchaseQuotationItemFormSet,
)
from apps.core.models import BusinessPartner, Item, Currency


# ============================================================================
# PURCHASE QUOTATION REQUEST (RFQ) - قائمة طلبات عروض الأسعار
# ============================================================================

class PurchaseQuotationRequestListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة طلبات عروض الأسعار (RFQ)"""
    model = PurchaseQuotationRequest
    template_name = 'purchases/quotations/rfq_list.html'
    context_object_name = 'rfqs'
    paginate_by = 50
    permission_required = 'purchases.view_purchasequotationrequest'

    def get_queryset(self):
        queryset = PurchaseQuotationRequest.objects.filter(
            company=self.request.current_company
        ).select_related(
            'purchase_request', 'awarded_quotation__supplier',
            'created_by'
        ).prefetch_related('items', 'quotations').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(subject__icontains=search) |
                Q(description__icontains=search)
            )

        # فلترة حسب الحالة
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        # فلترة حسب موعد الاستحقاق
        deadline_expired = self.request.GET.get('deadline_expired')
        if deadline_expired == 'true':
            queryset = queryset.filter(submission_deadline__lt=date.today())

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('طلبات عروض الأسعار')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات عروض الأسعار'), 'url': ''},
        ]

        # الإحصائيات
        rfqs = self.get_queryset()
        context['stats'] = {
            'total_count': rfqs.count(),
            'draft_count': rfqs.filter(status='draft').count(),
            'sent_count': rfqs.filter(status='sent').count(),
            'receiving_count': rfqs.filter(status='receiving').count(),
            'evaluating_count': rfqs.filter(status='evaluating').count(),
            'awarded_count': rfqs.filter(status='awarded').count(),
            'expired_count': rfqs.filter(
                submission_deadline__lt=date.today(),
                status__in=['draft', 'sent', 'receiving']
            ).count(),
        }

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('purchases.add_purchasequotationrequest')
        context['can_edit'] = self.request.user.has_perm('purchases.change_purchasequotationrequest')
        context['can_delete'] = self.request.user.has_perm('purchases.delete_purchasequotationrequest')
        context['can_send'] = self.request.user.has_perm('purchases.change_purchasequotationrequest')

        return context


class PurchaseQuotationRequestDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل طلب عرض أسعار مع العروض المستلمة"""
    model = PurchaseQuotationRequest
    template_name = 'purchases/quotations/rfq_detail.html'
    context_object_name = 'rfq'
    permission_required = 'purchases.view_purchasequotationrequest'

    def get_queryset(self):
        return PurchaseQuotationRequest.objects.filter(
            company=self.request.current_company
        ).select_related(
            'purchase_request', 'awarded_quotation__supplier', 'created_by', 'currency'
        ).prefetch_related(
            'items__item__base_uom',
            'quotations__supplier'
        )

    def get_object(self, queryset=None):
        """الحصول على الكائن والتأكد من وجود عملة"""
        obj = super().get_object(queryset)

        # إذا لم يكن للكائن عملة، قم بتعيين العملة الافتراضية
        if not obj.currency:
            default_currency = Currency.objects.filter(
                is_base=True,
                is_active=True
            ).first()
            if default_currency:
                obj.currency = default_currency
                obj.save()

        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        rfq = self.object
        context['title'] = f'{_("طلب عرض أسعار")} {rfq.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات عروض الأسعار'), 'url': reverse('purchases:rfq_list')},
            {'title': rfq.number, 'url': ''},
        ]

        # حساب الإجمالي التقديري للأصناف
        items_total = Decimal('0.000')
        for item in rfq.items.all():
            if item.line_total:
                items_total += item.line_total
            elif item.estimated_price and item.quantity:
                items_total += item.estimated_price * item.quantity
        context['estimated_total'] = items_total

        # العروض المستلمة
        context['quotations'] = rfq.quotations.select_related('supplier').order_by('-score', '-total_amount')

        # إحصائيات العروض
        context['quotation_stats'] = {
            'total_count': rfq.quotations.count(),
            'sent_count': rfq.quotations.filter(status='sent').count(),
            'received_count': rfq.quotations.filter(status='received').count(),
            'under_evaluation_count': rfq.quotations.filter(status='under_evaluation').count(),
            'accepted_count': rfq.quotations.filter(status='accepted').count(),
            'rejected_count': rfq.quotations.filter(status='rejected').count(),
            'awarded_count': rfq.quotations.filter(is_awarded=True).count(),
        }

        # متوسط التقييم
        avg_score = rfq.quotations.filter(score__isnull=False).aggregate(avg=Avg('score'))['avg']
        context['average_score'] = avg_score or 0

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('purchases.change_purchasequotationrequest') and
            rfq.status in ['draft', 'receiving']
        )
        context['can_delete'] = (
            self.request.user.has_perm('purchases.delete_purchasequotationrequest') and
            rfq.status == 'draft'
        )
        context['can_send'] = (
            self.request.user.has_perm('purchases.change_purchasequotationrequest') and
            rfq.status == 'draft'
        )
        context['can_evaluate'] = (
            self.request.user.has_perm('purchases.change_purchasequotationrequest') and
            rfq.status in ['receiving', 'evaluating']
        )
        context['can_award'] = (
            self.request.user.has_perm('purchases.change_purchasequotationrequest') and
            rfq.status in ['receiving', 'evaluating'] and
            not rfq.is_awarded
        )
        context['can_cancel'] = (
            self.request.user.has_perm('purchases.change_purchasequotationrequest') and
            rfq.status != 'awarded'
        )

        return context


class PurchaseQuotationRequestCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء طلب عرض أسعار جديد"""
    model = PurchaseQuotationRequest
    form_class = PurchaseQuotationRequestForm
    template_name = 'purchases/quotations/rfq_form.html'
    permission_required = 'purchases.add_purchasequotationrequest'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة طلب عرض أسعار')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات عروض الأسعار'), 'url': reverse('purchases:rfq_list')},
            {'title': _('إضافة طلب'), 'url': ''},
        ]

        # تفعيل البحث المباشر AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []

        if self.request.POST:
            context['formset'] = PurchaseQuotationRequestItemFormSet(
                self.request.POST,
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseQuotationRequestItemFormSet(
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )

        # بيانات للجافاسكربت
        context['items_data'] = list(
            Item.objects.filter(
                company=self.request.current_company,
                is_active=True
            ).values(
                'id', 'name', 'code', 'barcode',
                'base_uom__name',
                'base_uom__code'
            )
        )

        # إذا كان من طلب شراء
        request_id = self.request.GET.get('from_request')
        if request_id:
            try:
                purchase_request = PurchaseRequest.objects.get(
                    pk=request_id,
                    company=self.request.current_company
                )
                context['purchase_request'] = purchase_request
            except PurchaseRequest.DoesNotExist:
                pass

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        # ربط الطلب بالشركة
        form.instance.company = self.request.current_company
        form.instance.created_by = self.request.user

        # التأكد من وجود عملة
        if not form.instance.currency_id:
            default_currency = Currency.objects.filter(
                is_base=True,
                is_active=True
            ).first()
            if default_currency:
                form.instance.currency = default_currency

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # حفظ الموردين المختارين
            selected_suppliers = form.cleaned_data.get('suppliers')
            if selected_suppliers:
                # حذف الموردين السابقين (إن وجدوا)
                self.object.quotations.all().delete()

                # إنشاء PurchaseQuotation لكل مورد مختار
                for supplier in selected_suppliers:
                    PurchaseQuotation.objects.create(
                        company=self.request.current_company,
                        quotation_request=self.object,
                        supplier=supplier,
                        date=self.object.date,
                        valid_until=self.object.submission_deadline,
                        currency=self.object.currency,  # تمرير العملة من طلب العرض
                        status='draft',
                        created_by=self.request.user
                    )

            messages.success(
                self.request,
                _('تم إضافة طلب عرض الأسعار بنجاح')
            )
            return redirect('purchases:rfq_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseQuotationRequestUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل طلب عرض أسعار (في حالة مسودة فقط)"""
    model = PurchaseQuotationRequest
    form_class = PurchaseQuotationRequestForm
    template_name = 'purchases/quotations/rfq_form.html'
    permission_required = 'purchases.change_purchasequotationrequest'

    def get_queryset(self):
        # يمكن تعديل الطلب في حالة مسودة أو استقبال عروض فقط
        return PurchaseQuotationRequest.objects.filter(
            company=self.request.current_company,
            status__in=['draft', 'receiving']
        )

    def get_object(self, queryset=None):
        """الحصول على الكائن والتأكد من وجود عملة"""
        obj = super().get_object(queryset)

        # إذا لم يكن للكائن عملة، قم بتعيين العملة الافتراضية
        if not obj.currency:
            default_currency = Currency.objects.filter(
                is_base=True,
                is_active=True
            ).first()
            if default_currency:
                obj.currency = default_currency
                obj.save()

        return obj

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        rfq = self.object
        context['title'] = f'{_("تعديل طلب عرض أسعار")} {rfq.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('طلبات عروض الأسعار'), 'url': reverse('purchases:rfq_list')},
            {'title': rfq.number, 'url': reverse('purchases:rfq_detail', kwargs={'pk': rfq.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseQuotationRequestItemFormSet(
                self.request.POST,
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseQuotationRequestItemFormSet(
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )

        # تفعيل البحث المباشر AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []

        # جلب الموردين المحددين مسبقاً
        if not self.request.POST:
            selected_suppliers = rfq.quotations.values_list('supplier_id', flat=True)
            context['form'].fields['suppliers'].initial = list(selected_suppliers)

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # تحديث الموردين المختارين
            selected_suppliers = form.cleaned_data.get('suppliers')
            if selected_suppliers is not None:
                # حذف الموردين السابقين فقط في حالة draft
                if self.object.status == 'draft':
                    self.object.quotations.filter(status='draft').delete()

                    # إنشاء PurchaseQuotation لكل مورد مختار
                    for supplier in selected_suppliers:
                        PurchaseQuotation.objects.create(
                            company=self.request.current_company,
                            quotation_request=self.object,
                            supplier=supplier,
                            date=self.object.date,
                            valid_until=self.object.submission_deadline,
                            currency=self.object.currency,  # تمرير العملة من طلب العرض
                            status='draft',
                            created_by=self.request.user
                        )

            messages.success(
                self.request,
                _('تم تعديل طلب عرض الأسعار بنجاح')
            )
            return redirect('purchases:rfq_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseQuotationRequestDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف طلب عرض أسعار (في حالة مسودة فقط)"""
    model = PurchaseQuotationRequest
    template_name = 'purchases/quotations/rfq_confirm_delete.html'
    success_url = reverse_lazy('purchases:rfq_list')
    permission_required = 'purchases.delete_purchasequotationrequest'

    def get_queryset(self):
        # يمكن حذف الطلب في حالة مسودة فقط
        return PurchaseQuotationRequest.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف طلب عرض الأسعار بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# PURCHASE QUOTATION - عروض الأسعار
# ============================================================================

class PurchaseQuotationListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة عروض الأسعار"""
    model = PurchaseQuotation
    template_name = 'purchases/quotations/quotation_list.html'
    context_object_name = 'quotations'
    paginate_by = 50
    permission_required = 'purchases.view_purchasequotation'

    def get_queryset(self):
        queryset = PurchaseQuotation.objects.filter(
            company=self.request.current_company
        ).select_related(
            'quotation_request', 'supplier', 'currency', 'created_by'
        ).prefetch_related('lines').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(supplier__name__icontains=search) |
                Q(quotation_request__number__icontains=search)
            )

        # فلترة حسب المورد
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        # فلترة حسب الحالة
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # فلترة حسب صلاحية العرض
        valid_only = self.request.GET.get('valid_only')
        if valid_only == 'true':
            queryset = queryset.filter(valid_until__gte=date.today())

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('عروض الأسعار')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': ''},
        ]

        # الإحصائيات
        quotations = self.get_queryset()
        context['stats'] = {
            'total_count': quotations.count(),
            'draft_count': quotations.filter(status='draft').count(),
            'sent_count': quotations.filter(status='sent').count(),
            'received_count': quotations.filter(status='received').count(),
            'under_evaluation_count': quotations.filter(status='under_evaluation').count(),
            'accepted_count': quotations.filter(status='accepted').count(),
            'rejected_count': quotations.filter(status='rejected').count(),
            'awarded_count': quotations.filter(is_awarded=True).count(),
            'expired_count': quotations.filter(
                valid_until__lt=date.today(),
                status__in=['sent', 'received', 'under_evaluation']
            ).count(),
            'total_amount': quotations.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0.000'),
        }

        # قائمة الموردين للفلترة
        context['suppliers'] = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['supplier', 'both'],
            is_active=True
        ).order_by('name')

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('purchases.add_purchasequotation')
        context['can_edit'] = self.request.user.has_perm('purchases.change_purchasequotation')
        context['can_delete'] = self.request.user.has_perm('purchases.delete_purchasequotation')
        context['can_evaluate'] = self.request.user.has_perm('purchases.change_purchasequotation')

        return context


class PurchaseQuotationDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل عرض السعر"""
    model = PurchaseQuotation
    template_name = 'purchases/quotations/quotation_detail.html'
    context_object_name = 'quotation'
    permission_required = 'purchases.view_purchasequotation'

    def get_queryset(self):
        return PurchaseQuotation.objects.filter(
            company=self.request.current_company
        ).select_related(
            'quotation_request', 'supplier', 'currency', 'created_by'
        ).prefetch_related('lines__item', 'lines__rfq_item')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        quotation = self.object
        context['title'] = f'{_("عرض سعر")} {quotation.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': reverse('purchases:quotation_list')},
            {'title': quotation.number, 'url': ''},
        ]

        # عروض أخرى لنفس الطلب
        context['other_quotations'] = quotation.quotation_request.quotations.exclude(
            pk=quotation.pk
        ).order_by('-score', '-total_amount')

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('purchases.change_purchasequotation') and
            quotation.status in ['draft', 'received']
        )
        context['can_delete'] = (
            self.request.user.has_perm('purchases.delete_purchasequotation') and
            quotation.status == 'draft'
        )
        context['can_evaluate'] = (
            self.request.user.has_perm('purchases.change_purchasequotation') and
            quotation.status in ['received', 'under_evaluation']
        )
        context['can_award'] = (
            self.request.user.has_perm('purchases.change_purchasequotation') and
            quotation.status in ['received', 'under_evaluation'] and
            not quotation.is_awarded
        )
        context['can_reject'] = (
            self.request.user.has_perm('purchases.change_purchasequotation') and
            quotation.status in ['received', 'under_evaluation'] and
            not quotation.is_awarded
        )
        context['can_convert'] = (
            self.request.user.has_perm('purchases.add_purchaseorder') and
            quotation.is_awarded and
            quotation.status == 'awarded'
        )

        return context


class PurchaseQuotationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء عرض سعر جديد (إدخال عرض من مورد)"""
    model = PurchaseQuotation
    form_class = PurchaseQuotationForm
    template_name = 'purchases/quotations/quotation_form.html'
    permission_required = 'purchases.add_purchasequotation'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة عرض سعر')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': reverse('purchases:quotation_list')},
            {'title': _('إضافة عرض'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseQuotationItemFormSet(
                self.request.POST,
                self.request.FILES,
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseQuotationItemFormSet(
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )

        # تفعيل البحث المباشر AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []

        # إذا كان من طلب عرض أسعار
        rfq_id = self.request.GET.get('from_rfq')
        if rfq_id:
            try:
                rfq = PurchaseQuotationRequest.objects.get(
                    pk=rfq_id,
                    company=self.request.current_company
                )
                context['rfq'] = rfq
            except PurchaseQuotationRequest.DoesNotExist:
                pass

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        # ربط العرض بالشركة
        form.instance.company = self.request.current_company
        form.instance.created_by = self.request.user

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # حساب المجاميع
            self.object.calculate_totals()

            messages.success(
                self.request,
                _('تم إضافة عرض السعر بنجاح')
            )
            return redirect('purchases:quotation_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseQuotationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل عرض سعر"""
    model = PurchaseQuotation
    form_class = PurchaseQuotationForm
    template_name = 'purchases/quotations/quotation_form.html'
    permission_required = 'purchases.change_purchasequotation'

    def get_queryset(self):
        # يمكن تعديل العرض في حالة مسودة أو مستلم فقط
        return PurchaseQuotation.objects.filter(
            company=self.request.current_company,
            status__in=['draft', 'received']
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        quotation = self.object
        context['title'] = f'{_("تعديل عرض سعر")} {quotation.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': reverse('purchases:quotation_list')},
            {'title': quotation.number, 'url': reverse('purchases:quotation_detail', kwargs={'pk': quotation.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = PurchaseQuotationItemFormSet(
                self.request.POST,
                self.request.FILES,
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )
        else:
            context['formset'] = PurchaseQuotationItemFormSet(
                instance=self.object,
                prefix='lines',
                company=self.request.current_company
            )

        # تفعيل البحث المباشر AJAX Live Search
        context['use_live_search'] = True
        context['items_data'] = []

        return context

    @transaction.atomic
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()

            # حساب المجاميع
            self.object.calculate_totals()

            messages.success(
                self.request,
                _('تم تعديل عرض السعر بنجاح')
            )
            return redirect('purchases:quotation_detail', pk=self.object.pk)
        else:
            messages.error(
                self.request,
                _('يرجى تصحيح الأخطاء في النموذج')
            )
            return self.render_to_response(self.get_context_data(form=form))


class PurchaseQuotationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف عرض سعر (في حالة مسودة فقط)"""
    model = PurchaseQuotation
    template_name = 'purchases/quotations/quotation_confirm_delete.html'
    success_url = reverse_lazy('purchases:quotation_list')
    permission_required = 'purchases.delete_purchasequotation'

    def get_queryset(self):
        # يمكن حذف العرض في حالة مسودة فقط
        return PurchaseQuotation.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('تم حذف عرض السعر بنجاح'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# RFQ Workflow Actions - إجراءات تدفق طلبات عروض الأسعار
# ============================================================================

@login_required
@permission_required('purchases.change_purchasequotationrequest', raise_exception=True)
@transaction.atomic
def send_rfq_to_suppliers(request, pk):
    """إرسال طلب العرض للموردين المحددين"""
    rfq = get_object_or_404(
        PurchaseQuotationRequest,
        pk=pk,
        company=request.current_company
    )

    if rfq.status != 'draft':
        messages.error(request, _('يمكن إرسال الطلبات في حالة مسودة فقط'))
        return redirect('purchases:rfq_detail', pk=pk)

    if request.method == 'POST':
        supplier_ids = request.POST.getlist('suppliers')

        if not supplier_ids:
            messages.error(request, _('يجب اختيار مورد واحد على الأقل'))
            return redirect('purchases:rfq_detail', pk=pk)

        try:
            rfq.send_to_suppliers(supplier_ids=supplier_ids, user=request.user)
            messages.success(
                request,
                _('تم إرسال طلب العرض للموردين بنجاح')
            )
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'{_("خطأ في الإرسال")}: {str(e)}')

    return redirect('purchases:rfq_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchasequotationrequest', raise_exception=True)
@transaction.atomic
def cancel_rfq(request, pk):
    """إلغاء طلب عرض أسعار"""
    rfq = get_object_or_404(
        PurchaseQuotationRequest,
        pk=pk,
        company=request.current_company
    )

    if rfq.status == 'awarded':
        messages.error(request, _('لا يمكن إلغاء طلب عرض تمت ترسيته'))
        return redirect('purchases:rfq_detail', pk=pk)

    try:
        rfq.status = 'cancelled'
        rfq.save()
        messages.success(
            request,
            _('تم إلغاء طلب العرض بنجاح')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في الإلغاء")}: {str(e)}')

    return redirect('purchases:rfq_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchasequotationrequest', raise_exception=True)
@transaction.atomic
def mark_rfq_as_evaluating(request, pk):
    """تحديث حالة طلب العرض إلى قيد التقييم"""
    rfq = get_object_or_404(
        PurchaseQuotationRequest,
        pk=pk,
        company=request.current_company
    )

    if rfq.status not in ['receiving', 'sent']:
        messages.error(request, _('يمكن تقييم العروض في حالة الاستقبال فقط'))
        return redirect('purchases:rfq_detail', pk=pk)

    try:
        rfq.status = 'evaluating'
        rfq.save()

        # تحديث حالة جميع العروض المستلمة
        rfq.quotations.filter(status='sent').update(status='under_evaluation')

        messages.success(
            request,
            _('تم تحديث حالة طلب العرض إلى قيد التقييم')
        )
    except Exception as e:
        messages.error(request, f'{_("خطأ في التحديث")}: {str(e)}')

    return redirect('purchases:rfq_detail', pk=pk)


# ============================================================================
# Quotation Workflow Actions - إجراءات تدفق عروض الأسعار
# ============================================================================

@login_required
@permission_required('purchases.change_purchasequotation', raise_exception=True)
@transaction.atomic
def evaluate_quotation(request, pk):
    """تقييم عرض سعر"""
    quotation = get_object_or_404(
        PurchaseQuotation,
        pk=pk,
        company=request.current_company
    )

    if quotation.status not in ['sent', 'received']:
        messages.error(request, _('يمكن تقييم العروض المستلمة فقط'))
        return redirect('purchases:quotation_detail', pk=pk)

    if request.method == 'POST':
        score = request.POST.get('score', '')
        evaluation_notes = request.POST.get('evaluation_notes', '')

        try:
            if score:
                quotation.score = Decimal(score)
            quotation.evaluation_notes = evaluation_notes
            quotation.status = 'under_evaluation'
            quotation.save()

            messages.success(
                request,
                _('تم حفظ تقييم العرض بنجاح')
            )
        except Exception as e:
            messages.error(request, f'{_("خطأ في التقييم")}: {str(e)}')

        return redirect('purchases:quotation_detail', pk=pk)

    context = {
        'quotation': quotation,
        'title': _('تقييم عرض السعر'),
        'breadcrumbs': [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': reverse('purchases:quotation_list')},
            {'title': quotation.number, 'url': reverse('purchases:quotation_detail', kwargs={'pk': quotation.pk})},
            {'title': _('تقييم'), 'url': ''},
        ],
    }

    return render(request, 'purchases/quotations/quotation_evaluate.html', context)


@login_required
@permission_required('purchases.change_purchasequotation', raise_exception=True)
@transaction.atomic
def award_quotation(request, pk):
    """ترسية عرض سعر (تحديده كفائز)"""
    quotation = get_object_or_404(
        PurchaseQuotation,
        pk=pk,
        company=request.current_company
    )

    if quotation.is_awarded:
        messages.error(request, _('هذا العرض فائز بالفعل'))
        return redirect('purchases:quotation_detail', pk=pk)

    if quotation.status not in ['received', 'under_evaluation']:
        messages.error(request, _('يمكن ترسية العروض المستلمة أو تحت التقييم فقط'))
        return redirect('purchases:quotation_detail', pk=pk)

    try:
        quotation.mark_as_awarded(user=request.user)
        messages.success(
            request,
            _('تم ترسية العرض بنجاح')
        )
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'{_("خطأ في الترسية")}: {str(e)}')

    return redirect('purchases:quotation_detail', pk=pk)


@login_required
@permission_required('purchases.change_purchasequotation', raise_exception=True)
@transaction.atomic
def reject_quotation(request, pk):
    """رفض عرض سعر"""
    quotation = get_object_or_404(
        PurchaseQuotation,
        pk=pk,
        company=request.current_company
    )

    if quotation.is_awarded:
        messages.error(request, _('لا يمكن رفض عرض فائز'))
        return redirect('purchases:quotation_detail', pk=pk)

    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')

        try:
            quotation.status = 'rejected'
            quotation.rejection_reason = rejection_reason
            quotation.save()

            messages.success(
                request,
                _('تم رفض العرض بنجاح')
            )
            return redirect('purchases:quotation_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'{_("خطأ في الرفض")}: {str(e)}')

    context = {
        'quotation': quotation,
        'title': _('رفض عرض السعر'),
        'breadcrumbs': [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المشتريات'), 'url': '#'},
            {'title': _('عروض الأسعار'), 'url': reverse('purchases:quotation_list')},
            {'title': quotation.number, 'url': reverse('purchases:quotation_detail', kwargs={'pk': quotation.pk})},
            {'title': _('رفض'), 'url': ''},
        ],
    }

    return render(request, 'purchases/quotations/quotation_reject.html', context)


@login_required
@permission_required('purchases.add_purchaseorder', raise_exception=True)
@transaction.atomic
def convert_quotation_to_order(request, pk):
    """تحويل العرض الفائز إلى أمر شراء"""
    quotation = get_object_or_404(
        PurchaseQuotation,
        pk=pk,
        company=request.current_company
    )

    if not quotation.is_awarded:
        messages.error(request, _('يمكن تحويل العروض الفائزة فقط'))
        return redirect('purchases:quotation_detail', pk=pk)

    try:
        order = quotation.convert_to_purchase_order(user=request.user)
        messages.success(
            request,
            _('تم تحويل العرض إلى أمر شراء بنجاح')
        )
        return redirect('purchases:order_detail', pk=order.pk)
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'{_("خطأ في التحويل")}: {str(e)}')

    return redirect('purchases:quotation_detail', pk=pk)


# ============================================================================
# AJAX & Export - البيانات والتصدير
# ============================================================================

@login_required
@permission_required('purchases.view_purchasequotationrequest', raise_exception=True)
def rfq_datatable_ajax(request):
    """AJAX endpoint لجداول طلبات عروض الأسعار"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # الفلاتر الإضافية
    status_filter = request.GET.get('status', '')
    supplier_filter = request.GET.get('supplier', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_filter = request.GET.get('search_filter', '')

    queryset = PurchaseQuotationRequest.objects.filter(
        company=request.current_company
    ).select_related('purchase_request')

    # تطبيق الفلاتر
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if supplier_filter:
        queryset = queryset.filter(quotations__supplier__id=supplier_filter).distinct()

    if date_from:
        queryset = queryset.filter(date__gte=date_from)

    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # البحث
    if search_value or search_filter:
        search_term = search_filter if search_filter else search_value
        queryset = queryset.filter(
            Q(number__icontains=search_term) |
            Q(subject__icontains=search_term)
        )

    # العدد الكلي
    total_records = PurchaseQuotationRequest.objects.filter(company=request.current_company).count()
    filtered_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    queryset = queryset[start:start + length]

    # البيانات
    data = []
    for rfq in queryset:
        # عدد الموردين = عدد عروض الأسعار المستلمة
        suppliers_count = rfq.quotations.values('supplier').distinct().count()
        data.append({
            'number': rfq.number,
            'date': rfq.date.strftime('%Y-%m-%d'),
            'subject': rfq.subject,
            'suppliers_count': suppliers_count,
            'submission_deadline': rfq.submission_deadline.strftime('%Y-%m-%d') if rfq.submission_deadline else '',
            'status': rfq.get_status_display(),
            'status_code': rfq.status,
            'pk': rfq.pk,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@login_required
@permission_required('purchases.view_purchasequotation', raise_exception=True)
def quotation_datatable_ajax(request):
    """AJAX endpoint لجداول عروض الأسعار"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    queryset = PurchaseQuotation.objects.filter(
        company=request.current_company
    ).select_related('supplier', 'quotation_request')

    # البحث
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(supplier__name__icontains=search_value)
        )

    # العدد الكلي
    total_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    queryset = queryset[start:start + length]

    # البيانات
    data = []
    for quotation in queryset:
        data.append({
            'number': quotation.number,
            'date': quotation.date.strftime('%Y-%m-%d'),
            'supplier': quotation.supplier.name,
            'total': float(quotation.total_amount),
            'score': float(quotation.score) if quotation.score else '',
            'status': quotation.get_status_display(),
            'status_code': quotation.status,
            'is_awarded': quotation.is_awarded,
            'pk': quotation.pk,
        })

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


@login_required
@permission_required('purchases.view_purchasequotationrequest', raise_exception=True)
def export_rfqs_excel(request):
    """تصدير طلبات عروض الأسعار إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # استرجاع البيانات
    queryset = PurchaseQuotationRequest.objects.filter(
        company=request.current_company
    ).select_related(
        'purchase_request'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "طلبات عروض الأسعار"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم الطلب', 'التاريخ', 'الموضوع', 'موعد الاستحقاق',
        'عدد العروض', 'الحالة'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, rfq in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=rfq.number).border = border
        ws.cell(row=row_num, column=2, value=rfq.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=rfq.subject).border = border
        ws.cell(row=row_num, column=4, value=rfq.submission_deadline.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=5, value=rfq.quotations.count()).border = border
        ws.cell(row=row_num, column=6, value=rfq.get_status_display()).border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"rfqs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
@permission_required('purchases.view_purchasequotation', raise_exception=True)
def export_quotations_excel(request):
    """تصدير عروض الأسعار إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # استرجاع البيانات
    queryset = PurchaseQuotation.objects.filter(
        company=request.current_company
    ).select_related(
        'supplier', 'quotation_request'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "عروض الأسعار"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم العرض', 'التاريخ', 'المورد', 'الإجمالي',
        'التقييم', 'فائز', 'الحالة'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, quotation in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=quotation.number).border = border
        ws.cell(row=row_num, column=2, value=quotation.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=quotation.supplier.name).border = border
        ws.cell(row=row_num, column=4, value=float(quotation.total_amount)).border = border
        ws.cell(row=row_num, column=5, value=float(quotation.score) if quotation.score else '').border = border
        ws.cell(row=row_num, column=6, value='نعم' if quotation.is_awarded else 'لا').border = border
        ws.cell(row=row_num, column=7, value=quotation.get_status_display()).border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"quotations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


# ============================================================================
# AJAX - جلب أصناف طلب الشراء
# ============================================================================

@login_required
def get_purchase_request_items_ajax(request, request_id):
    """جلب أصناف طلب الشراء عبر AJAX"""
    try:
        # التحقق من وجود طلب الشراء وأنه تابع للشركة الحالية
        purchase_request = get_object_or_404(
            PurchaseRequest,
            id=request_id,
            company=request.current_company
        )

        # جلب أصناف طلب الشراء
        items = purchase_request.lines.select_related('item', 'item__base_uom').all()

        # تحويل البيانات إلى قائمة
        items_data = []
        for item in items:
            item_data = {
                'item_id': item.item.id if item.item else None,
                'item_name': item.item.name if item.item else '',
                'item_code': item.item.code if item.item else '',
                'item_description': item.item_description,
                'quantity': str(item.quantity),
                'unit': item.unit if item.unit else (item.item.base_uom.name if item.item and item.item.base_uom else ''),
                'estimated_price': str(item.estimated_price) if item.estimated_price else '0.000',
                'notes': item.notes if item.notes else '',
            }
            items_data.append(item_data)

        # معلومات إضافية عن طلب الشراء
        request_data = {
            'number': purchase_request.number,
            'date': purchase_request.date.strftime('%Y-%m-%d'),
            'required_date': purchase_request.required_date.strftime('%Y-%m-%d') if purchase_request.required_date else None,
            'purpose': purchase_request.purpose,
            'notes': purchase_request.notes,
            'items': items_data
        }

        return JsonResponse({
            'success': True,
            'data': request_data
        })

    except PurchaseRequest.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'طلب الشراء غير موجود'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# AJAX Endpoints for RFQs & Quotations - Live Search & Stock Display
# ============================================================================

@login_required
@permission_required('purchases.view_purchasequotationrequest', raise_exception=True)
def rfq_get_item_stock_multi_branch_ajax(request):
    """جلب رصيد المخزون من جميع الفروع - RFQ"""
    from apps.inventory.models import ItemStock
    from decimal import Decimal

    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        stock_records = ItemStock.objects.filter(
            company=request.current_company,
            item_id=item_id
        ).select_related('warehouse', 'warehouse__branch').order_by('warehouse__branch__name')

        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            # التحقق من وجود branch
            branch_name = 'غير محدد'
            if stock.warehouse and stock.warehouse.branch:
                branch_name = stock.warehouse.branch.name

            branches_data.append({
                'branch_name': branch_name,
                'warehouse_name': stock.warehouse.name if stock.warehouse else 'غير محدد',
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost or 0),
            })

        return JsonResponse({
            'success': True,
            'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('purchases.view_purchasequotationrequest', raise_exception=True)
def rfq_get_item_stock_current_branch_ajax(request):
    """جلب رصيد الفرع الحالي - RFQ"""
    from apps.inventory.models import ItemStock
    from django.db.models import Sum
    from decimal import Decimal

    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        stock_aggregate = ItemStock.objects.filter(
            company=request.current_company,
            item_id=item_id,
            warehouse__branch=request.current_branch
        ).aggregate(total_qty=Sum('quantity'), total_reserved=Sum('reserved_quantity'))

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True,
            'quantity': str(total_qty),
            'reserved': str(total_reserved),
            'available': str(available),
            'has_stock': total_qty > 0,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('purchases.view_purchasequotationrequest', raise_exception=True)
def rfq_item_search_ajax(request):
    """AJAX Live Search - RFQ"""
    from apps.core.models import Item
    from django.db.models import Q, Sum

    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({'success': False, 'message': 'يرجى إدخال حرفين على الأقل'})

    try:
        items = Item.objects.filter(
            company=request.current_company, is_active=True
        ).filter(
            Q(name__icontains=term) | Q(code__icontains=term) | Q(barcode__icontains=term)
        ).annotate(
            current_branch_stock=Sum('stock_records__quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)),
            current_branch_reserved=Sum('stock_records__reserved_quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)),
            total_stock=Sum('stock_records__quantity'),
        ).select_related('category', 'base_uom')[:limit]

        items_data = [{
            'id': item.id, 'name': item.name, 'code': item.code,
            'barcode': item.barcode or '', 'category_name': item.category.name if item.category else '',
            'tax_rate': str(item.tax_rate), 'base_uom_name': item.base_uom.name if item.base_uom else '',
            'base_uom_code': item.base_uom.code if item.base_uom else '',
            'current_branch_stock': str(item.current_branch_stock or 0),
            'current_branch_reserved': str(item.current_branch_reserved or 0),
            'total_stock': str(item.total_stock or 0),
        } for item in items]

        return JsonResponse({'success': True, 'items': items_data, 'count': len(items_data)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('purchases.view_purchasequotation', raise_exception=True)
def quotation_get_item_stock_multi_branch_ajax(request):
    """جلب رصيد المخزون من جميع الفروع - Quotation"""
    from apps.inventory.models import ItemStock
    from decimal import Decimal

    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        stock_records = ItemStock.objects.filter(
            company=request.current_company, item_id=item_id
        ).select_related('warehouse', 'warehouse__branch').order_by('warehouse__branch__name')

        branches_data = []
        total_quantity = Decimal('0')
        total_available = Decimal('0')

        for stock in stock_records:
            available = stock.quantity - stock.reserved_quantity
            total_quantity += stock.quantity
            total_available += available

            # التحقق من وجود branch
            branch_name = 'غير محدد'
            if stock.warehouse and stock.warehouse.branch:
                branch_name = stock.warehouse.branch.name

            branches_data.append({
                'branch_name': branch_name,
                'warehouse_name': stock.warehouse.name if stock.warehouse else 'غير محدد',
                'quantity': str(stock.quantity),
                'reserved': str(stock.reserved_quantity),
                'available': str(available),
                'average_cost': str(stock.average_cost or 0),
            })

        return JsonResponse({
            'success': True, 'has_stock': len(branches_data) > 0,
            'branches': branches_data,
            'total_quantity': str(total_quantity),
            'total_available': str(total_available),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('purchases.view_purchasequotation', raise_exception=True)
def quotation_get_item_stock_current_branch_ajax(request):
    """جلب رصيد الفرع الحالي - Quotation"""
    from apps.inventory.models import ItemStock
    from django.db.models import Sum
    from decimal import Decimal

    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'Missing item_id'}, status=400)

    try:
        stock_aggregate = ItemStock.objects.filter(
            company=request.current_company, item_id=item_id,
            warehouse__branch=request.current_branch
        ).aggregate(total_qty=Sum('quantity'), total_reserved=Sum('reserved_quantity'))

        total_qty = stock_aggregate['total_qty'] or Decimal('0')
        total_reserved = stock_aggregate['total_reserved'] or Decimal('0')
        available = total_qty - total_reserved

        return JsonResponse({
            'success': True, 'quantity': str(total_qty),
            'reserved': str(total_reserved), 'available': str(available),
            'has_stock': total_qty > 0,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required('purchases.view_purchasequotation', raise_exception=True)
def quotation_item_search_ajax(request):
    """AJAX Live Search - Quotation"""
    from apps.core.models import Item
    from django.db.models import Q, Sum

    term = request.GET.get('term', '').strip()
    limit = int(request.GET.get('limit', 20))

    if len(term) < 2:
        return JsonResponse({'success': False, 'message': 'يرجى إدخال حرفين على الأقل'})

    try:
        items = Item.objects.filter(
            company=request.current_company, is_active=True
        ).filter(
            Q(name__icontains=term) | Q(code__icontains=term) | Q(barcode__icontains=term)
        ).annotate(
            current_branch_stock=Sum('stock_records__quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)),
            current_branch_reserved=Sum('stock_records__reserved_quantity',
                filter=Q(stock_records__warehouse__branch=request.current_branch)),
            total_stock=Sum('stock_records__quantity'),
        ).select_related('category', 'base_uom')[:limit]

        items_data = [{
            'id': item.id, 'name': item.name, 'code': item.code,
            'barcode': item.barcode or '', 'category_name': item.category.name if item.category else '',
            'tax_rate': str(item.tax_rate), 'base_uom_name': item.base_uom.name if item.base_uom else '',
            'base_uom_code': item.base_uom.code if item.base_uom else '',
            'current_branch_stock': str(item.current_branch_stock or 0),
            'current_branch_reserved': str(item.current_branch_reserved or 0),
            'total_stock': str(item.total_stock or 0),
        } for item in items]

        return JsonResponse({'success': True, 'items': items_data, 'count': len(items_data)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
