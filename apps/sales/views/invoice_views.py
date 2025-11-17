# apps/sales/views/invoice_views.py
"""
Views لفواتير المبيعات
Complete CRUD operations + Post/Unpost + Print
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db.models import Q, Sum, Count, F, Avg
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from django.core.exceptions import PermissionDenied, ValidationError
from decimal import Decimal
from datetime import datetime, date

from ..models import SalesInvoice, InvoiceItem
from ..forms import (
    SalesInvoiceForm,
    InvoiceItemForm,
    InvoiceItemFormSet,
)
from apps.core.models import BusinessPartner, Item


class SalesInvoiceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة فواتير المبيعات"""
    model = SalesInvoice
    template_name = 'sales/invoices/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 50
    permission_required = 'sales.view_salesinvoice'

    def get_queryset(self):
        queryset = SalesInvoice.objects.filter(
            company=self.request.current_company
        ).select_related(
            'customer', 'salesperson', 'warehouse', 'currency',
            'payment_method', 'created_by', 'branch'
        ).prefetch_related('lines').order_by('-date', '-number')

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__code__icontains=search) |
                Q(recipient_name__icontains=search) |
                Q(recipient_phone__icontains=search)
            )

        # فلترة حسب العميل
        customer_id = self.request.GET.get('customer')
        if customer_id:
            queryset = queryset.filter(customer_id=customer_id)

        # فلترة حسب المندوب
        salesperson_id = self.request.GET.get('salesperson')
        if salesperson_id:
            queryset = queryset.filter(salesperson_id=salesperson_id)

        # فلترة حسب نوع الفاتورة
        invoice_type = self.request.GET.get('invoice_type')
        if invoice_type:
            queryset = queryset.filter(invoice_type=invoice_type)

        # فلترة حسب حالة الدفع
        payment_status = self.request.GET.get('payment_status')
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)

        # فلترة حسب حالة الترحيل
        is_posted = self.request.GET.get('is_posted')
        if is_posted == 'true':
            queryset = queryset.filter(is_posted=True)
        elif is_posted == 'false':
            queryset = queryset.filter(is_posted=False)

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        # فلترة حسب المستودع
        warehouse_id = self.request.GET.get('warehouse')
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('فواتير المبيعات')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المبيعات'), 'url': '#'},
            {'title': _('فواتير المبيعات'), 'url': ''},
        ]

        # الإحصائيات
        invoices = self.get_queryset()

        # احصائيات عامة
        context['stats'] = {
            'total_count': invoices.count(),
            'posted_count': invoices.filter(is_posted=True).count(),
            'draft_count': invoices.filter(is_posted=False).count(),
        }

        # احصائيات الدفع
        context['payment_stats'] = {
            'unpaid_count': invoices.filter(payment_status='unpaid').count(),
            'partial_count': invoices.filter(payment_status='partial').count(),
            'paid_count': invoices.filter(payment_status='paid').count(),
        }

        # احصائيات مالية
        posted_invoices = invoices.filter(is_posted=True)
        context['financial_stats'] = {
            'total_amount': posted_invoices.aggregate(
                total=Sum('total_with_tax')
            )['total'] or Decimal('0.000'),
            'paid_amount': posted_invoices.aggregate(
                total=Sum('paid_amount')
            )['total'] or Decimal('0.000'),
            'remaining_amount': posted_invoices.aggregate(
                total=Sum('remaining_amount')
            )['total'] or Decimal('0.000'),
            'average_invoice': posted_invoices.aggregate(
                avg=Avg('total_with_tax')
            )['avg'] or Decimal('0.000'),
        }

        # احصائيات حسب نوع الفاتورة
        context['type_stats'] = {
            'cash_sale_count': invoices.filter(invoice_type='cash_sale').count(),
            'credit_sale_count': invoices.filter(invoice_type='credit_sale').count(),
            'installment_count': invoices.filter(invoice_type='installment').count(),
            'return_count': invoices.filter(invoice_type='return').count(),
        }

        # قائمة العملاء للفلترة
        context['customers'] = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['customer', 'both'],
            is_active=True
        ).order_by('name')

        # قائمة المندوبين للفلترة
        from apps.core.models import User
        context['salespersons'] = User.objects.filter(
            is_active=True
        ).order_by('first_name', 'last_name')

        # قائمة المستودعات للفلترة
        from apps.core.models import Warehouse
        context['warehouses'] = Warehouse.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('name')

        # خيارات الفلترة
        context['invoice_types'] = SalesInvoice.INVOICE_TYPES
        context['payment_statuses'] = SalesInvoice.PAYMENT_STATUS_CHOICES

        # الصلاحيات
        context['can_add'] = self.request.user.has_perm('sales.add_salesinvoice')
        context['can_edit'] = self.request.user.has_perm('sales.change_salesinvoice')
        context['can_delete'] = self.request.user.has_perm('sales.delete_salesinvoice')
        context['can_post'] = self.request.user.has_perm('sales.change_salesinvoice')

        # Query parameters للفلاتر
        context['current_filters'] = {
            'search': self.request.GET.get('search', ''),
            'customer': self.request.GET.get('customer', ''),
            'salesperson': self.request.GET.get('salesperson', ''),
            'invoice_type': self.request.GET.get('invoice_type', ''),
            'payment_status': self.request.GET.get('payment_status', ''),
            'is_posted': self.request.GET.get('is_posted', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'warehouse': self.request.GET.get('warehouse', ''),
        }

        return context


class SalesInvoiceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء فاتورة مبيعات جديدة"""
    model = SalesInvoice
    form_class = SalesInvoiceForm
    template_name = 'sales/invoices/invoice_form.html'
    permission_required = 'sales.add_salesinvoice'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = _('إضافة فاتورة مبيعات')
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المبيعات'), 'url': '#'},
            {'title': _('فواتير المبيعات'), 'url': reverse('sales:invoice_list')},
            {'title': _('إضافة فاتورة'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = InvoiceItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company
            )
        else:
            context['formset'] = InvoiceItemFormSet(
                instance=self.object,
                company=self.request.current_company
            )

        # بيانات للجافاسكربت
        # ملاحظة: السعر سيتم جلبه من قائمة الأسعار عبر AJAX
        context['items_data'] = list(
            Item.objects.filter(
                company=self.request.current_company,
                is_active=True
            ).values(
                'id', 'name', 'code', 'barcode',
                'tax_rate',
                'unit_of_measure__name',
                'unit_of_measure__code'
            )
        )

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST request with form and formset"""
        self.object = None
        form = self.get_form()

        # Create formset from POST data
        formset = InvoiceItemFormSet(
            self.request.POST,
            instance=self.object,
            company=self.request.current_company
        )

        # Validate both form and formset
        if form.is_valid() and formset.is_valid():
            return self.form_valid(form, formset)
        else:
            return self.form_invalid(form, formset)

    @transaction.atomic
    def form_valid(self, form, formset):
        """Save both form and formset"""
        # ربط الفاتورة بالشركة والفرع
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user

        # حفظ الفاتورة
        self.object = form.save()

        # ربط الفورمسيت بالفاتورة وحفظه
        formset.instance = self.object
        formset.save()

        # حساب المجاميع
        self.object.calculate_totals()

        # تحديث حالة الدفع
        self.object.update_payment_status()

        # حساب العمولة
        self.object.calculate_commission()

        self.object.save()

        # التحقق من حد الائتمان للعميل (للمبيعات الآجلة)
        if self.object.invoice_type in ['credit_sale', 'installment']:
            credit_check = self.object.customer.check_credit_limit(self.object.total_with_tax)
            if not credit_check['allowed']:
                messages.warning(
                    self.request,
                    _('تحذير: ') + credit_check['message']
                )

        messages.success(
            self.request,
            _('تم إضافة فاتورة المبيعات بنجاح')
        )
        return redirect('sales:invoice_detail', pk=self.object.pk)

    def form_invalid(self, form, formset):
        """Handle invalid form or formset"""
        # Return with form and formset context to show inline errors
        context = self.get_context_data()
        context['form'] = form
        context['formset'] = formset
        return self.render_to_response(context)


class SalesInvoiceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل فاتورة مبيعات"""
    model = SalesInvoice
    form_class = SalesInvoiceForm
    template_name = 'sales/invoices/invoice_form.html'
    permission_required = 'sales.change_salesinvoice'

    def get_queryset(self):
        # يمكن تعديل الفاتورة إذا لم تكن مرحّلة فقط
        return SalesInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=False
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.current_company
        kwargs['branch'] = self.request.current_branch
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        invoice = self.object
        context['title'] = f'{_("تعديل فاتورة")} {invoice.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المبيعات'), 'url': '#'},
            {'title': _('فواتير المبيعات'), 'url': reverse('sales:invoice_list')},
            {'title': invoice.number, 'url': reverse('sales:invoice_detail', kwargs={'pk': invoice.pk})},
            {'title': _('تعديل'), 'url': ''},
        ]

        if self.request.POST:
            context['formset'] = InvoiceItemFormSet(
                self.request.POST,
                instance=self.object,
                company=self.request.current_company
            )
        else:
            context['formset'] = InvoiceItemFormSet(
                instance=self.object,
                company=self.request.current_company
            )

        # بيانات للجافاسكربت
        # ملاحظة: السعر سيتم جلبه من قائمة الأسعار عبر AJAX
        context['items_data'] = list(
            Item.objects.filter(
                company=self.request.current_company,
                is_active=True
            ).values(
                'id', 'name', 'code', 'barcode',
                'tax_rate',
                'unit_of_measure__name',
                'unit_of_measure__code'
            )
        )

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST request with form and formset"""
        self.object = self.get_object()
        form = self.get_form()

        # Create formset from POST data
        formset = InvoiceItemFormSet(
            self.request.POST,
            instance=self.object,
            company=self.request.current_company
        )

        # Validate both form and formset
        if form.is_valid() and formset.is_valid():
            return self.form_valid(form, formset)
        else:
            return self.form_invalid(form, formset)

    @transaction.atomic
    def form_valid(self, form, formset):
        """Save both form and formset"""
        # حفظ الفاتورة
        self.object = form.save()

        # ربط الفورمسيت بالفاتورة وحفظه
        formset.instance = self.object
        formset.save()

        # إعادة حساب المجاميع
        self.object.calculate_totals()

        # تحديث حالة الدفع
        self.object.update_payment_status()

        # حساب العمولة
        self.object.calculate_commission()

        self.object.save()

        # التحقق من حد الائتمان للعميل (للمبيعات الآجلة)
        if self.object.invoice_type in ['credit_sale', 'installment']:
            credit_check = self.object.customer.check_credit_limit(self.object.total_with_tax)
            if not credit_check['allowed']:
                messages.warning(
                    self.request,
                    _('تحذير: ') + credit_check['message']
                )

        messages.success(
            self.request,
            _('تم تعديل فاتورة المبيعات بنجاح')
        )
        return redirect('sales:invoice_detail', pk=self.object.pk)

    def form_invalid(self, form, formset):
        """Handle invalid form or formset"""
        # Return with form and formset context to show inline errors
        context = self.get_context_data()
        context['form'] = form
        context['formset'] = formset
        return self.render_to_response(context)


class SalesInvoiceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """تفاصيل فاتورة المبيعات"""
    model = SalesInvoice
    template_name = 'sales/invoices/invoice_detail.html'
    context_object_name = 'invoice'
    permission_required = 'sales.view_salesinvoice'

    def get_queryset(self):
        return SalesInvoice.objects.filter(
            company=self.request.current_company
        ).select_related(
            'customer', 'salesperson', 'warehouse', 'currency',
            'payment_method', 'created_by', 'branch'
        ).prefetch_related('lines__item__unit_of_measure', 'installments', 'commission')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        invoice = self.object
        context['title'] = f'{_("فاتورة مبيعات")} {invoice.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المبيعات'), 'url': '#'},
            {'title': _('فواتير المبيعات'), 'url': reverse('sales:invoice_list')},
            {'title': invoice.number, 'url': ''},
        ]

        # الصلاحيات
        context['can_edit'] = (
            self.request.user.has_perm('sales.change_salesinvoice') and
            not invoice.is_posted
        )
        context['can_delete'] = (
            self.request.user.has_perm('sales.delete_salesinvoice') and
            not invoice.is_posted
        )
        context['can_post'] = (
            self.request.user.has_perm('sales.change_salesinvoice') and
            not invoice.is_posted
        )
        context['can_unpost'] = (
            self.request.user.has_perm('sales.change_salesinvoice') and
            invoice.is_posted
        )
        context['can_print'] = self.request.user.has_perm('sales.view_salesinvoice')

        # معلومات إضافية
        context['has_installments'] = invoice.installments.exists()
        context['has_commission'] = hasattr(invoice, 'commission')

        # إحصائيات السطور
        lines = invoice.lines.all()
        context['lines_stats'] = {
            'total_items': lines.count(),
            'total_quantity': sum(line.quantity for line in lines),
        }

        return context


class SalesInvoiceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف فاتورة مبيعات"""
    model = SalesInvoice
    template_name = 'sales/invoices/invoice_confirm_delete.html'
    success_url = reverse_lazy('sales:invoice_list')
    permission_required = 'sales.delete_salesinvoice'

    def get_queryset(self):
        # يمكن حذف الفاتورة إذا لم تكن مرحّلة فقط
        return SalesInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=False
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        invoice = self.object
        context['title'] = f'{_("حذف فاتورة")} {invoice.number}'
        context['breadcrumbs'] = [
            {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
            {'title': _('المبيعات'), 'url': '#'},
            {'title': _('فواتير المبيعات'), 'url': reverse('sales:invoice_list')},
            {'title': invoice.number, 'url': reverse('sales:invoice_detail', kwargs={'pk': invoice.pk})},
            {'title': _('حذف'), 'url': ''},
        ]

        # التحقق من الحذف الآمن
        context['can_delete_safely'] = True
        context['warnings'] = []

        # التحقق من وجود أقساط
        if invoice.installments.exists():
            context['warnings'].append(_('تحتوي هذه الفاتورة على أقساط'))

        # التحقق من وجود عمولة
        if hasattr(invoice, 'commission'):
            context['warnings'].append(_('تحتوي هذه الفاتورة على عمولة مندوب'))

        # التحقق من وجود دفعات
        if invoice.paid_amount > 0:
            context['warnings'].append(_('تم دفع مبلغ من هذه الفاتورة'))

        return context

    def delete(self, request, *args, **kwargs):
        invoice = self.get_object()

        # التحقق من أن الفاتورة غير مرحّلة
        if invoice.is_posted:
            messages.error(
                request,
                _('لا يمكن حذف فاتورة مرحّلة. يرجى إلغاء الترحيل أولاً')
            )
            return redirect('sales:invoice_detail', pk=invoice.pk)

        messages.success(request, _('تم حذف فاتورة المبيعات بنجاح'))
        return super().delete(request, *args, **kwargs)


class SalesInvoicePostView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """ترحيل فاتورة مبيعات"""
    permission_required = 'sales.change_salesinvoice'

    def get_queryset(self):
        """الحصول على الفواتير غير المرحّلة فقط"""
        return SalesInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=False
        )

    def get(self, request, pk):
        """إعادة التوجيه إلى صفحة التفاصيل"""
        return redirect('sales:invoice_detail', pk=pk)

    def post(self, request, pk):
        """ترحيل الفاتورة"""
        try:
            invoice = self.get_queryset().get(pk=pk)
        except SalesInvoice.DoesNotExist:
            messages.error(
                request,
                _('الفاتورة غير موجودة أو مرحّلة مسبقاً')
            )
            return redirect('sales:invoice_list')

        try:
            # استدعاء method الترحيل من الـ model
            stock_out, journal_entry = invoice.post(user=request.user)

            # رسالة النجاح مع تفاصيل
            messages.success(
                request,
                _(f'تم ترحيل فاتورة المبيعات {invoice.number} بنجاح')
            )
            messages.info(
                request,
                _(f'تم إنشاء سند إخراج: {stock_out.number}')
            )
            messages.info(
                request,
                _(f'تم إنشاء قيد محاسبي: {journal_entry.number}')
            )

        except ValidationError as e:
            # عرض رسائل الخطأ
            if hasattr(e, 'message_dict'):
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            elif hasattr(e, 'messages'):
                for error in e.messages:
                    messages.error(request, error)
            else:
                messages.error(request, str(e))

        except Exception as e:
            # خطأ غير متوقع
            messages.error(
                request,
                _(f'حدث خطأ أثناء ترحيل الفاتورة: {str(e)}')
            )

        return redirect('sales:invoice_detail', pk=invoice.pk)


class SalesInvoiceUnpostView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """إلغاء ترحيل فاتورة مبيعات"""
    permission_required = 'sales.change_salesinvoice'

    def get_queryset(self):
        """الحصول على الفواتير المرحّلة فقط"""
        return SalesInvoice.objects.filter(
            company=self.request.current_company,
            is_posted=True
        )

    def get(self, request, pk):
        """إعادة التوجيه إلى صفحة التفاصيل"""
        return redirect('sales:invoice_detail', pk=pk)

    def post(self, request, pk):
        """إلغاء ترحيل الفاتورة"""
        try:
            invoice = self.get_queryset().get(pk=pk)
        except SalesInvoice.DoesNotExist:
            messages.error(
                request,
                _('الفاتورة غير موجودة أو غير مرحّلة')
            )
            return redirect('sales:invoice_list')

        # التحقق من عدم وجود مدفوعات
        if hasattr(invoice, 'paid_amount') and invoice.paid_amount > 0:
            messages.error(
                request,
                _('لا يمكن إلغاء ترحيل فاتورة تم دفع مبلغ منها. يرجى إلغاء المدفوعات أولاً')
            )
            return redirect('sales:invoice_detail', pk=invoice.pk)

        try:
            # استدعاء method إلغاء الترحيل من الـ model
            invoice.unpost()

            # رسالة النجاح
            messages.success(
                request,
                _(f'تم إلغاء ترحيل فاتورة المبيعات {invoice.number} بنجاح')
            )
            messages.info(
                request,
                _('تم حذف سند الإخراج والقيد المحاسبي المرتبطين')
            )

        except ValidationError as e:
            # عرض رسائل الخطأ
            if hasattr(e, 'message_dict'):
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            elif hasattr(e, 'messages'):
                for error in e.messages:
                    messages.error(request, error)
            else:
                messages.error(request, str(e))

        except Exception as e:
            # خطأ غير متوقع
            messages.error(
                request,
                _(f'حدث خطأ أثناء إلغاء ترحيل الفاتورة: {str(e)}')
            )

        return redirect('sales:invoice_detail', pk=invoice.pk)


@login_required
@permission_required('sales.view_salesinvoice', raise_exception=True)
def invoice_datatable_ajax(request):
    """AJAX endpoint for DataTables"""
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # الفلاتر من الواجهة
    invoice_type = request.GET.get('invoice_type', '')
    is_posted = request.GET.get('is_posted', '')
    payment_status = request.GET.get('payment_status', '')
    customer_id = request.GET.get('customer', '')
    salesperson_id = request.GET.get('salesperson', '')
    warehouse_id = request.GET.get('warehouse', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_filter = request.GET.get('search_filter', '')

    queryset = SalesInvoice.objects.filter(
        company=request.current_company
    ).select_related('customer', 'warehouse', 'salesperson', 'currency')

    # تطبيق الفلاتر
    if invoice_type:
        queryset = queryset.filter(invoice_type=invoice_type)

    if is_posted == '1':
        queryset = queryset.filter(is_posted=True)
    elif is_posted == '0':
        queryset = queryset.filter(is_posted=False)

    if payment_status:
        queryset = queryset.filter(payment_status=payment_status)

    if customer_id:
        queryset = queryset.filter(customer_id=customer_id)

    if salesperson_id:
        queryset = queryset.filter(salesperson_id=salesperson_id)

    if warehouse_id:
        queryset = queryset.filter(warehouse_id=warehouse_id)

    if date_from:
        queryset = queryset.filter(date__gte=date_from)

    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # البحث العام
    if search_filter:
        queryset = queryset.filter(
            Q(number__icontains=search_filter) |
            Q(customer__name__icontains=search_filter) |
            Q(receipt_number__icontains=search_filter)
        )

    # البحث من DataTable
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(customer__name__icontains=search_value) |
            Q(receipt_number__icontains=search_value)
        )

    # العدد الكلي
    total_records = queryset.count()

    # الترتيب
    queryset = queryset.order_by('-date', '-number')

    # Pagination
    if length > 0:
        queryset = queryset[start:start + length]

    # البيانات - يجب أن تكون array وليس dictionary
    data = []
    for invoice in queryset:
        # تحديد نوع الفاتورة
        invoice_type_display = 'مبيعات' if invoice.invoice_type == 'sales' else 'مرتجع'

        # تحديد حالة الترحيل
        if invoice.is_posted:
            status_badge = '<span class="badge bg-success">مرحلة</span>'
        else:
            status_badge = '<span class="badge bg-warning text-dark">مسودة</span>'

        # تحديد حالة الدفع
        if invoice.payment_status == 'paid':
            payment_badge = '<span class="badge bg-success">مدفوعة</span>'
        elif invoice.payment_status == 'partial':
            payment_badge = '<span class="badge bg-warning text-dark">دفع جزئي</span>'
        else:
            payment_badge = '<span class="badge bg-danger">غير مدفوعة</span>'

        # أزرار الإجراءات
        actions = f'''
        <div class="btn-group" role="group">
            <a href="{reverse('sales:invoice_detail', args=[invoice.pk])}"
               class="btn btn-sm btn-info"
               data-bs-toggle="tooltip"
               title="عرض">
                <i class="fas fa-eye"></i>
            </a>
        '''

        if not invoice.is_posted:
            actions += f'''
            <a href="{reverse('sales:invoice_update', args=[invoice.pk])}"
               class="btn btn-sm btn-primary"
               data-bs-toggle="tooltip"
               title="تعديل">
                <i class="fas fa-edit"></i>
            </a>
            <button onclick="postInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-success"
                    data-bs-toggle="tooltip"
                    title="ترحيل">
                <i class="fas fa-check"></i>
            </button>
            <button onclick="deleteInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-danger"
                    data-bs-toggle="tooltip"
                    title="حذف">
                <i class="fas fa-trash"></i>
            </button>
            '''
        else:
            actions += f'''
            <button onclick="unpostInvoice({invoice.pk}, '{invoice.number}')"
                    class="btn btn-sm btn-warning"
                    data-bs-toggle="tooltip"
                    title="إلغاء الترحيل">
                <i class="fas fa-undo"></i>
            </button>
            '''

        actions += '</div>'

        # إضافة البيانات كـ array وليس dictionary
        data.append([
            invoice.number or '-',  # 0: رقم الفاتورة
            invoice.date.strftime('%Y-%m-%d'),  # 1: التاريخ
            invoice.customer.name if invoice.customer else '-',  # 2: الزبون
            invoice_type_display,  # 3: النوع
            invoice.warehouse.name if invoice.warehouse else '-',  # 4: المستودع
            f'{float(invoice.total_with_tax):,.3f}',  # 5: الإجمالي
            payment_badge,  # 6: حالة الدفع
            status_badge,  # 7: حالة الترحيل
            actions  # 8: الإجراءات
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    })


@login_required
@permission_required('sales.view_salesinvoice', raise_exception=True)
def export_invoices_excel(request):
    """تصدير فواتير المبيعات إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    # استرجاع البيانات
    queryset = SalesInvoice.objects.filter(
        company=request.current_company
    ).select_related(
        'customer', 'warehouse', 'salesperson', 'currency'
    ).order_by('-date', '-number')

    # تطبيق الفلاتر من GET parameters
    customer_id = request.GET.get('customer')
    if customer_id:
        queryset = queryset.filter(customer_id=customer_id)

    salesperson_id = request.GET.get('salesperson')
    if salesperson_id:
        queryset = queryset.filter(salesperson_id=salesperson_id)

    warehouse_id = request.GET.get('warehouse')
    if warehouse_id:
        queryset = queryset.filter(warehouse_id=warehouse_id)

    invoice_type = request.GET.get('invoice_type')
    if invoice_type:
        queryset = queryset.filter(invoice_type=invoice_type)

    is_posted = request.GET.get('is_posted')
    if is_posted == '1':
        queryset = queryset.filter(is_posted=True)
    elif is_posted == '0':
        queryset = queryset.filter(is_posted=False)

    payment_status = request.GET.get('payment_status')
    if payment_status:
        queryset = queryset.filter(payment_status=payment_status)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "فواتير المبيعات"

    # الأنماط
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # العناوين
    headers = [
        'رقم الفاتورة', 'التاريخ', 'الزبون', 'المندوب', 'المستودع',
        'رقم الإيصال', 'الإجمالي', 'الضريبة', 'الإجمالي مع الضريبة',
        'العملة', 'حالة الترحيل', 'حالة الدفع'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # البيانات
    for row_num, invoice in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=invoice.number).border = border
        ws.cell(row=row_num, column=2, value=invoice.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row_num, column=3, value=invoice.customer.name).border = border
        ws.cell(row=row_num, column=4, value=invoice.salesperson.get_full_name() if invoice.salesperson else '-').border = border
        ws.cell(row=row_num, column=5, value=invoice.warehouse.name).border = border
        ws.cell(row=row_num, column=6, value=invoice.receipt_number).border = border
        ws.cell(row=row_num, column=7, value=float(invoice.total_amount)).border = border
        ws.cell(row=row_num, column=8, value=float(invoice.tax_amount)).border = border
        ws.cell(row=row_num, column=9, value=float(invoice.total_with_tax)).border = border
        ws.cell(row=row_num, column=10, value=invoice.currency.code).border = border
        ws.cell(row=row_num, column=11, value='مرحل' if invoice.is_posted else 'مسودة').border = border

        # حالة الدفع
        payment_status_text = {
            'unpaid': 'غير مدفوعة',
            'partial': 'دفع جزئي',
            'paid': 'مدفوعة'
        }.get(invoice.payment_status, '-')
        ws.cell(row=row_num, column=12, value=payment_status_text).border = border

    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 15

    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # إرجاع الملف
    filename = f"sales_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
