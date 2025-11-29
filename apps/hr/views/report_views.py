# apps/hr/views/report_views.py
"""
تقارير الموارد البشرية - المرحلة الرابعة
HR Reports Views - Phase 4
مع دعم تصدير Excel
"""

from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Count, Avg, Q, F, Min, Max
from django.db.models.functions import TruncMonth, ExtractMonth, ExtractYear
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
import json
from io import BytesIO

# استيراد مكتبات التصدير
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ..models import (
    Employee, Department, JobTitle, JobGrade, EmployeeContract,
    Attendance, LeaveRequest, LeaveBalance, LeaveType,
    Advance, AdvanceInstallment, Overtime, EarlyLeave,
    Payroll, PayrollDetail
)


class HRDashboardReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """لوحة تحكم التقارير الرئيسية"""
    template_name = 'hr/reports/dashboard.html'
    permission_required = 'hr.view_employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()
        current_month_start = today.replace(day=1)

        # إحصائيات الموظفين
        employee_stats = Employee.objects.filter(company=company).aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='active', is_active=True)),
            inactive=Count('id', filter=Q(status='inactive') | Q(is_active=False)),
            on_leave=Count('id', filter=Q(status='on_leave')),
            terminated=Count('id', filter=Q(status='terminated')),
            male=Count('id', filter=Q(gender='male', is_active=True)),
            female=Count('id', filter=Q(gender='female', is_active=True)),
        )

        # إحصائيات الأقسام
        department_stats = Department.objects.filter(company=company, is_active=True).annotate(
            emp_count=Count('employees', filter=Q(employees__is_active=True, employees__status='active'))
        ).order_by('-emp_count')[:10]

        # إحصائيات الحضور للشهر الحالي
        attendance_stats = Attendance.objects.filter(
            employee__company=company,
            date__gte=current_month_start,
            date__lte=today
        ).aggregate(
            total_records=Count('id'),
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            on_leave=Count('id', filter=Q(status='leave')),
            total_late_minutes=Sum('late_minutes'),
            total_overtime_hours=Sum('overtime_hours'),
        )

        # طلبات الإجازات الحالية
        leave_stats = LeaveRequest.objects.filter(
            employee__company=company,
            created_at__gte=current_month_start
        ).aggregate(
            total=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            approved=Count('id', filter=Q(status='approved')),
            rejected=Count('id', filter=Q(status='rejected')),
            total_days=Sum('days', filter=Q(status='approved')),
        )

        # إحصائيات الرواتب للشهر الحالي
        payroll_stats = PayrollDetail.objects.filter(
            payroll__company=company,
            payroll__period_year=today.year,
            payroll__period_month=today.month,
            payroll__status__in=['approved', 'paid']
        ).aggregate(
            total_basic=Sum('basic_salary'),
            total_allowances=Sum('total_allowances'),
            total_deductions=Sum('total_deductions'),
            total_net=Sum('net_salary'),
            employee_count=Count('id'),
        )

        # السلف المستحقة
        advance_stats = Advance.objects.filter(
            employee__company=company,
            status='disbursed'
        ).aggregate(
            total_amount=Sum('amount'),
            remaining=Sum('remaining_amount'),
            count=Count('id'),
        )

        context.update({
            'title': _('لوحة تقارير الموارد البشرية'),
            'employee_stats': employee_stats,
            'department_stats': department_stats,
            'attendance_stats': attendance_stats,
            'leave_stats': leave_stats,
            'payroll_stats': payroll_stats,
            'advance_stats': advance_stats,
            'current_month': today.strftime('%B %Y'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': None},
            ],
        })
        return context


class EmployeeReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير الموظفين"""
    template_name = 'hr/reports/employee_report.html'
    permission_required = 'hr.view_employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company

        # الفلاتر
        department_id = self.request.GET.get('department')
        status = self.request.GET.get('status', 'active')
        job_title_id = self.request.GET.get('job_title')

        # بناء الاستعلام
        employees = Employee.objects.filter(company=company).select_related(
            'department', 'job_title', 'job_grade'
        )

        if department_id:
            employees = employees.filter(department_id=department_id)
        if status and status != 'all':
            employees = employees.filter(status=status)
        if job_title_id:
            employees = employees.filter(job_title_id=job_title_id)

        # إحصائيات
        stats = employees.aggregate(
            total=Count('id'),
            male=Count('id', filter=Q(gender='male')),
            female=Count('id', filter=Q(gender='female')),
        )

        # توزيع حسب القسم
        by_department = employees.values(
            'department__name'
        ).annotate(count=Count('id')).order_by('-count')

        # توزيع حسب المسمى
        by_job_title = employees.values(
            'job_title__name'
        ).annotate(count=Count('id')).order_by('-count')

        # توزيع حسب الدرجة
        by_job_grade = employees.values(
            'job_grade__name'
        ).annotate(count=Count('id')).order_by('-count')

        context.update({
            'title': _('تقرير الموظفين'),
            'employees': employees[:100],  # Limit for performance
            'stats': stats,
            'by_department': list(by_department[:10]),
            'by_job_title': list(by_job_title[:10]),
            'by_job_grade': list(by_job_grade[:10]),
            'departments': Department.objects.filter(company=company, is_active=True),
            'job_titles': JobTitle.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'selected_status': status,
            'selected_job_title': job_title_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير الموظفين'), 'url': None},
            ],
        })
        return context


class AttendanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير الحضور والانصراف"""
    template_name = 'hr/reports/attendance_report.html'
    permission_required = 'hr.view_attendance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        department_id = self.request.GET.get('department')
        employee_id = self.request.GET.get('employee')

        # افتراضي: الشهر الحالي
        if not from_date:
            from_date = today.replace(day=1).isoformat()
        if not to_date:
            to_date = today.isoformat()

        # بناء الاستعلام
        attendance = Attendance.objects.filter(
            employee__company=company,
            date__gte=from_date,
            date__lte=to_date
        ).select_related('employee', 'employee__department')

        if department_id:
            attendance = attendance.filter(employee__department_id=department_id)
        if employee_id:
            attendance = attendance.filter(employee_id=employee_id)

        # إحصائيات عامة
        stats = attendance.aggregate(
            total_records=Count('id'),
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            on_leave=Count('id', filter=Q(status='leave')),
            total_late_minutes=Sum('late_minutes'),
            total_overtime_hours=Sum('overtime_hours'),
            avg_working_hours=Avg('working_hours'),
        )

        # تفصيل حسب الموظف
        by_employee = attendance.values(
            'employee__id',
            'employee__first_name',
            'employee__last_name',
            'employee__employee_number'
        ).annotate(
            total_days=Count('id'),
            present_days=Count('id', filter=Q(status='present')),
            absent_days=Count('id', filter=Q(status='absent')),
            late_days=Count('id', filter=Q(status='late')),
            total_late_minutes=Sum('late_minutes'),
            total_overtime_hours=Sum('overtime_hours'),
        ).order_by('employee__first_name')

        # تفصيل حسب القسم
        by_department = attendance.values(
            'employee__department__name'
        ).annotate(
            total_records=Count('id'),
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            attendance_rate=Count('id', filter=Q(status__in=['present', 'late'])) * 100 / Count('id')
        ).order_by('-attendance_rate')

        # التوزيع اليومي
        daily_trend = attendance.values('date').annotate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
        ).order_by('date')

        context.update({
            'title': _('تقرير الحضور والانصراف'),
            'stats': stats,
            'by_employee': list(by_employee[:50]),
            'by_department': list(by_department),
            'daily_trend': json.dumps(list(daily_trend), default=str),
            'from_date': from_date,
            'to_date': to_date,
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'selected_employee': employee_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير الحضور'), 'url': None},
            ],
        })
        return context


class LeaveReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير الإجازات"""
    template_name = 'hr/reports/leave_report.html'
    permission_required = 'hr.view_leaverequest'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        year = int(self.request.GET.get('year', today.year))
        department_id = self.request.GET.get('department')
        leave_type_id = self.request.GET.get('leave_type')

        # بناء الاستعلام
        leaves = LeaveRequest.objects.filter(
            employee__company=company,
            start_date__year=year
        ).select_related('employee', 'leave_type')

        if department_id:
            leaves = leaves.filter(employee__department_id=department_id)
        if leave_type_id:
            leaves = leaves.filter(leave_type_id=leave_type_id)

        # إحصائيات عامة
        stats = leaves.aggregate(
            total_requests=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            approved=Count('id', filter=Q(status='approved')),
            rejected=Count('id', filter=Q(status='rejected')),
            cancelled=Count('id', filter=Q(status='cancelled')),
            total_days=Sum('days', filter=Q(status='approved')),
        )

        # تفصيل حسب نوع الإجازة
        by_leave_type = leaves.filter(status='approved').values(
            'leave_type__name'
        ).annotate(
            count=Count('id'),
            total_days=Sum('days')
        ).order_by('-total_days')

        # تفصيل حسب القسم
        by_department = leaves.filter(status='approved').values(
            'employee__department__name'
        ).annotate(
            count=Count('id'),
            total_days=Sum('days')
        ).order_by('-total_days')

        # التوزيع الشهري
        monthly_trend = leaves.filter(status='approved').annotate(
            month=ExtractMonth('start_date')
        ).values('month').annotate(
            count=Count('id'),
            total_days=Sum('days')
        ).order_by('month')

        # أرصدة الإجازات
        # استخدام أسماء الحقول من النموذج (مع db_column لمطابقة قاعدة البيانات)
        balances = LeaveBalance.objects.filter(
            employee__company=company,
            year=year,
            employee__is_active=True,
            employee__status='active'
        ).select_related('employee', 'leave_type').values(
            'leave_type__name'
        ).annotate(
            total_balance=Sum(F('opening_balance') + F('carried_forward') + F('adjustment')),
            used_balance=Sum('used'),
            remaining=Sum(F('opening_balance') + F('carried_forward') + F('adjustment') - F('used'))
        ).order_by('leave_type__name')

        context.update({
            'title': _('تقرير الإجازات'),
            'stats': stats,
            'by_leave_type': list(by_leave_type),
            'by_department': list(by_department),
            'monthly_trend': json.dumps(list(monthly_trend)),
            'balances': list(balances),
            'year': year,
            'years': range(today.year - 5, today.year + 2),
            'departments': Department.objects.filter(company=company, is_active=True),
            'leave_types': LeaveType.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'selected_leave_type': leave_type_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير الإجازات'), 'url': None},
            ],
        })
        return context


class PayrollSummaryReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير ملخص الرواتب"""
    template_name = 'hr/reports/payroll_summary_report.html'
    permission_required = 'hr.view_payroll'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        year = int(self.request.GET.get('year', today.year))
        month = self.request.GET.get('month')
        department_id = self.request.GET.get('department')

        # بناء الاستعلام الأساسي
        payrolls = Payroll.objects.filter(
            company=company,
            period_year=year,
            status__in=['approved', 'paid']
        )

        if month:
            payrolls = payrolls.filter(period_month=int(month))

        details_filter = Q(payroll__company=company, payroll__period_year=year, payroll__status__in=['approved', 'paid'])
        if month:
            details_filter &= Q(payroll__period_month=int(month))
        if department_id:
            details_filter &= Q(employee__department_id=department_id)

        details = PayrollDetail.objects.filter(details_filter)

        # الإحصائيات العامة - استخدام الأعمدة الموجودة فعلياً في قاعدة البيانات
        stats = details.aggregate(
            employee_count=Count('id'),
            total_basic=Sum('basic_salary'),
            total_allowances=Sum('total_allowances'),
            total_deductions=Sum('total_deductions'),
            total_net=Sum('net_salary'),
            total_working_days=Sum('working_days'),
            total_actual_days=Sum('actual_days'),
        )

        # التفصيل الشهري
        monthly_data = PayrollDetail.objects.filter(
            payroll__company=company,
            payroll__period_year=year,
            payroll__status__in=['approved', 'paid']
        ).values(
            'payroll__period_month'
        ).annotate(
            employee_count=Count('id'),
            total_basic=Sum('basic_salary'),
            total_allowances=Sum('total_allowances'),
            total_deductions=Sum('total_deductions'),
            total_net=Sum('net_salary'),
        ).order_by('payroll__period_month')

        # التفصيل حسب القسم
        by_department = details.values(
            'employee__department__name'
        ).annotate(
            employee_count=Count('id'),
            total_basic=Sum('basic_salary'),
            total_allowances=Sum('total_allowances'),
            total_deductions=Sum('total_deductions'),
            total_net=Sum('net_salary'),
            avg_salary=Avg('net_salary'),
        ).order_by('-total_net')

        # تفصيل البدلات والخصومات - استخدام الإجماليات المتاحة فقط
        # ملاحظة: الجدول الحالي يحتوي على إجماليات فقط وليس تفاصيل
        allowances_breakdown = {
            'total': stats.get('total_allowances', 0) or 0
        }

        deductions_breakdown = {
            'total': stats.get('total_deductions', 0) or 0
        }

        context.update({
            'title': _('تقرير ملخص الرواتب'),
            'stats': stats,
            'monthly_data': json.dumps(list(monthly_data), default=str),
            'by_department': list(by_department),
            'allowances_breakdown': allowances_breakdown,
            'deductions_breakdown': deductions_breakdown,
            'year': year,
            'month': month,
            'years': range(today.year - 5, today.year + 2),
            'months': [
                (1, 'يناير'), (2, 'فبراير'), (3, 'مارس'), (4, 'أبريل'),
                (5, 'مايو'), (6, 'يونيو'), (7, 'يوليو'), (8, 'أغسطس'),
                (9, 'سبتمبر'), (10, 'أكتوبر'), (11, 'نوفمبر'), (12, 'ديسمبر')
            ],
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير الرواتب'), 'url': None},
            ],
        })
        return context


class AdvanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير السلف"""
    template_name = 'hr/reports/advance_report.html'
    permission_required = 'hr.view_advance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        year = int(self.request.GET.get('year', today.year))
        status = self.request.GET.get('status')
        department_id = self.request.GET.get('department')

        # بناء الاستعلام
        advances = Advance.objects.filter(
            employee__company=company,
            request_date__year=year
        ).select_related('employee', 'employee__department')

        if status:
            advances = advances.filter(status=status)
        if department_id:
            advances = advances.filter(employee__department_id=department_id)

        # الإحصائيات
        stats = advances.aggregate(
            total_count=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            approved=Count('id', filter=Q(status='approved')),
            disbursed=Count('id', filter=Q(status='disbursed')),
            fully_paid=Count('id', filter=Q(status='fully_paid')),
            rejected=Count('id', filter=Q(status='rejected')),
            total_amount=Sum('amount', filter=Q(status__in=['approved', 'disbursed', 'fully_paid'])),
            total_remaining=Sum('remaining_amount', filter=Q(status='disbursed')),
        )

        # التفصيل حسب القسم
        by_department = advances.filter(
            status__in=['approved', 'disbursed', 'fully_paid']
        ).values(
            'employee__department__name'
        ).annotate(
            count=Count('id'),
            total_amount=Sum('amount'),
            remaining=Sum('remaining_amount'),
        ).order_by('-total_amount')

        # التوزيع الشهري
        monthly_data = advances.filter(
            status__in=['approved', 'disbursed', 'fully_paid']
        ).annotate(
            month=ExtractMonth('request_date')
        ).values('month').annotate(
            count=Count('id'),
            total_amount=Sum('amount'),
        ).order_by('month')

        # السلف المعلقة
        pending_advances = advances.filter(status='disbursed').order_by('-remaining_amount')[:20]

        context.update({
            'title': _('تقرير السلف'),
            'stats': stats,
            'by_department': list(by_department),
            'monthly_data': json.dumps(list(monthly_data), default=str),
            'pending_advances': pending_advances,
            'year': year,
            'years': range(today.year - 5, today.year + 2),
            'status_choices': [
                ('pending', 'معلق'),
                ('approved', 'معتمد'),
                ('disbursed', 'مصروف'),
                ('fully_paid', 'مسدد بالكامل'),
                ('rejected', 'مرفوض'),
            ],
            'selected_status': status,
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير السلف'), 'url': None},
            ],
        })
        return context


class OvertimeReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير العمل الإضافي"""
    template_name = 'hr/reports/overtime_report.html'
    permission_required = 'hr.view_overtime'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        year = int(self.request.GET.get('year', today.year))
        month = self.request.GET.get('month')
        department_id = self.request.GET.get('department')

        # بناء الاستعلام
        overtime_qs = Overtime.objects.filter(
            employee__company=company,
            date__year=year
        ).select_related('employee', 'employee__department')

        if month:
            overtime_qs = overtime_qs.filter(date__month=int(month))
        if department_id:
            overtime_qs = overtime_qs.filter(employee__department_id=department_id)

        # الإحصائيات
        stats = overtime_qs.aggregate(
            total_records=Count('id'),
            pending=Count('id', filter=Q(status='pending')),
            approved=Count('id', filter=Q(status='approved')),
            rejected=Count('id', filter=Q(status='rejected')),
            total_hours=Sum('hours', filter=Q(status='approved')),
            total_amount=Sum('amount', filter=Q(status='approved')),
        )

        # التفصيل حسب الموظف
        by_employee = overtime_qs.filter(status='approved').values(
            'employee__id',
            'employee__first_name',
            'employee__last_name',
            'employee__employee_number'
        ).annotate(
            count=Count('id'),
            total_hours=Sum('hours'),
            total_amount=Sum('amount'),
        ).order_by('-total_hours')[:20]

        # التفصيل حسب القسم
        by_department = overtime_qs.filter(status='approved').values(
            'employee__department__name'
        ).annotate(
            count=Count('id'),
            total_hours=Sum('hours'),
            total_amount=Sum('amount'),
        ).order_by('-total_hours')

        # التوزيع الشهري
        monthly_data = overtime_qs.filter(status='approved').annotate(
            month=ExtractMonth('date')
        ).values('month').annotate(
            count=Count('id'),
            total_hours=Sum('hours'),
            total_amount=Sum('amount'),
        ).order_by('month')

        context.update({
            'title': _('تقرير العمل الإضافي'),
            'stats': stats,
            'by_employee': list(by_employee),
            'by_department': list(by_department),
            'monthly_data': json.dumps(list(monthly_data), default=str),
            'year': year,
            'month': month,
            'years': range(today.year - 5, today.year + 2),
            'months': [
                (1, 'يناير'), (2, 'فبراير'), (3, 'مارس'), (4, 'أبريل'),
                (5, 'مايو'), (6, 'يونيو'), (7, 'يوليو'), (8, 'أغسطس'),
                (9, 'سبتمبر'), (10, 'أكتوبر'), (11, 'نوفمبر'), (12, 'ديسمبر')
            ],
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير العمل الإضافي'), 'url': None},
            ],
        })
        return context


class DepartmentAnalysisReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير تحليل الأقسام"""
    template_name = 'hr/reports/department_analysis_report.html'
    permission_required = 'hr.view_department'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # جميع الأقسام مع إحصائياتها
        departments = Department.objects.filter(
            company=company,
            is_active=True
        ).annotate(
            employee_count=Count('employees', filter=Q(employees__is_active=True, employees__status='active')),
            male_count=Count('employees', filter=Q(employees__is_active=True, employees__gender='male')),
            female_count=Count('employees', filter=Q(employees__is_active=True, employees__gender='female')),
        ).order_by('-employee_count')

        # إجمالي الرواتب لكل قسم (آخر شهر)
        current_month = today.month
        current_year = today.year

        payroll_by_dept = PayrollDetail.objects.filter(
            payroll__company=company,
            payroll__period_year=current_year,
            payroll__period_month=current_month,
            payroll__status__in=['approved', 'paid']
        ).values(
            'employee__department__id',
            'employee__department__name'
        ).annotate(
            total_basic=Sum('basic_salary'),
            total_net=Sum('net_salary'),
            avg_salary=Avg('net_salary'),
        )

        # تحويل إلى قاموس للوصول السريع
        payroll_dict = {item['employee__department__id']: item for item in payroll_by_dept}

        # إضافة بيانات الرواتب للأقسام
        departments_data = []
        for dept in departments:
            dept_payroll = payroll_dict.get(dept.id, {})
            departments_data.append({
                'department': dept,
                'employee_count': dept.employee_count,
                'male_count': dept.male_count,
                'female_count': dept.female_count,
                'total_basic': dept_payroll.get('total_basic', 0),
                'total_net': dept_payroll.get('total_net', 0),
                'avg_salary': dept_payroll.get('avg_salary', 0),
            })

        # إحصائيات الحضور حسب القسم
        attendance_by_dept = Attendance.objects.filter(
            employee__company=company,
            date__year=current_year,
            date__month=current_month
        ).values(
            'employee__department__name'
        ).annotate(
            total=Count('id'),
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
        )

        context.update({
            'title': _('تقرير تحليل الأقسام'),
            'departments_data': departments_data,
            'attendance_by_dept': list(attendance_by_dept),
            'current_month': today.strftime('%B %Y'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تحليل الأقسام'), 'url': None},
            ],
        })
        return context


class ContractExpiryReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير انتهاء العقود"""
    template_name = 'hr/reports/contract_expiry_report.html'
    permission_required = 'hr.view_employeecontract'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # الفلاتر
        days_ahead = int(self.request.GET.get('days', 90))

        # العقود المنتهية صلاحيتها
        expiring_date = today + timedelta(days=days_ahead)

        contracts = EmployeeContract.objects.filter(
            company=company,
            status='active',
            end_date__lte=expiring_date
        ).select_related('employee', 'employee__department').order_by('end_date')

        # تقسيم العقود حسب الفترة
        expired = contracts.filter(end_date__lt=today)
        expiring_30 = contracts.filter(end_date__gte=today, end_date__lte=today + timedelta(days=30))
        expiring_60 = contracts.filter(end_date__gt=today + timedelta(days=30), end_date__lte=today + timedelta(days=60))
        expiring_90 = contracts.filter(end_date__gt=today + timedelta(days=60), end_date__lte=today + timedelta(days=90))

        # إحصائيات
        stats = {
            'total': contracts.count(),
            'expired': expired.count(),
            'expiring_30': expiring_30.count(),
            'expiring_60': expiring_60.count(),
            'expiring_90': expiring_90.count(),
        }

        context.update({
            'title': _('تقرير انتهاء العقود'),
            'contracts': contracts,
            'expired': expired,
            'expiring_30': expiring_30,
            'expiring_60': expiring_60,
            'expiring_90': expiring_90,
            'stats': stats,
            'days_ahead': days_ahead,
            'today': today,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('انتهاء العقود'), 'url': None},
            ],
        })
        return context


# ==========================================
# Excel Export Views
# ==========================================

class ExcelExportMixin:
    """Mixin للمساعدة في تصدير Excel"""

    def create_workbook(self, title):
        """إنشاء ملف Excel جديد"""
        if not EXCEL_AVAILABLE:
            return None
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel يسمح ب 31 حرف كحد أقصى
        ws.sheet_view.rightToLeft = True  # دعم RTL للعربية
        return wb, ws

    def style_header(self, ws, row=1, start_col=1, end_col=10):
        """تنسيق صف العناوين"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_fit_columns(self, ws):
        """تعديل عرض الأعمدة تلقائياً"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_excel_response(self, wb, filename):
        """إرجاع response للتحميل"""
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExportEmployeesExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير الموظفين إلى Excel"""
    permission_required = 'hr.view_employee'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company

        # الفلاتر
        department_id = request.GET.get('department')
        status = request.GET.get('status', 'active')

        # بناء الاستعلام
        employees = Employee.objects.filter(company=company).select_related(
            'department', 'job_title', 'job_grade'
        )

        if department_id:
            employees = employees.filter(department_id=department_id)
        if status and status != 'all':
            employees = employees.filter(status=status)

        # إنشاء Excel
        wb, ws = self.create_workbook('الموظفين')

        # العناوين
        headers = ['رقم الموظف', 'الاسم الأول', 'اسم العائلة', 'القسم', 'المسمى الوظيفي',
                   'الدرجة', 'الجنس', 'الحالة', 'تاريخ التعيين', 'البريد الإلكتروني', 'الهاتف']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        # البيانات
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp.employee_number)
            ws.cell(row=row, column=2, value=emp.first_name)
            ws.cell(row=row, column=3, value=emp.last_name)
            ws.cell(row=row, column=4, value=emp.department.name if emp.department else '-')
            ws.cell(row=row, column=5, value=emp.job_title.name if emp.job_title else '-')
            ws.cell(row=row, column=6, value=emp.job_grade.name if emp.job_grade else '-')
            ws.cell(row=row, column=7, value='ذكر' if emp.gender == 'male' else 'أنثى')
            ws.cell(row=row, column=8, value=dict(Employee.STATUS_CHOICES).get(emp.status, emp.status))
            ws.cell(row=row, column=9, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '-')
            ws.cell(row=row, column=10, value=emp.work_email or '-')
            ws.cell(row=row, column=11, value=emp.mobile or '-')

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'employees_report_{date.today()}.xlsx')


class ExportAttendanceExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير الحضور إلى Excel"""
    permission_required = 'hr.view_attendance'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        from_date = request.GET.get('from_date', today.replace(day=1).isoformat())
        to_date = request.GET.get('to_date', today.isoformat())
        department_id = request.GET.get('department')

        # بناء الاستعلام
        attendance = Attendance.objects.filter(
            employee__company=company,
            date__gte=from_date,
            date__lte=to_date
        ).select_related('employee', 'employee__department').order_by('date', 'employee__first_name')

        if department_id:
            attendance = attendance.filter(employee__department_id=department_id)

        # إنشاء Excel
        wb, ws = self.create_workbook('الحضور')

        # العناوين
        headers = ['التاريخ', 'رقم الموظف', 'اسم الموظف', 'القسم', 'وقت الحضور',
                   'وقت الانصراف', 'ساعات العمل', 'الحالة', 'دقائق التأخير', 'ساعات إضافية']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        status_map = {'present': 'حاضر', 'absent': 'غائب', 'late': 'متأخر', 'leave': 'إجازة', 'holiday': 'عطلة'}

        # البيانات
        for row, att in enumerate(attendance, 2):
            ws.cell(row=row, column=1, value=att.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=att.employee.employee_number)
            ws.cell(row=row, column=3, value=f'{att.employee.first_name} {att.employee.last_name}')
            ws.cell(row=row, column=4, value=att.employee.department.name if att.employee.department else '-')
            ws.cell(row=row, column=5, value=att.check_in.strftime('%H:%M') if att.check_in else '-')
            ws.cell(row=row, column=6, value=att.check_out.strftime('%H:%M') if att.check_out else '-')
            ws.cell(row=row, column=7, value=float(att.working_hours) if att.working_hours else 0)
            ws.cell(row=row, column=8, value=status_map.get(att.status, att.status))
            ws.cell(row=row, column=9, value=att.late_minutes or 0)
            ws.cell(row=row, column=10, value=float(att.overtime_hours) if att.overtime_hours else 0)

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'attendance_report_{from_date}_{to_date}.xlsx')


class ExportLeavesExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير الإجازات إلى Excel"""
    permission_required = 'hr.view_leaverequest'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        year = int(request.GET.get('year', today.year))

        # طلبات الإجازات
        leaves = LeaveRequest.objects.filter(
            employee__company=company,
            start_date__year=year
        ).select_related('employee', 'leave_type', 'employee__department').order_by('-start_date')

        # إنشاء Excel
        wb, ws = self.create_workbook('الإجازات')

        # العناوين
        headers = ['رقم الموظف', 'اسم الموظف', 'القسم', 'نوع الإجازة',
                   'تاريخ البداية', 'تاريخ النهاية', 'عدد الأيام', 'الحالة', 'السبب']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        status_map = {'pending': 'معلق', 'approved': 'موافق', 'rejected': 'مرفوض', 'cancelled': 'ملغي'}

        # البيانات
        for row, leave in enumerate(leaves, 2):
            ws.cell(row=row, column=1, value=leave.employee.employee_number)
            ws.cell(row=row, column=2, value=f'{leave.employee.first_name} {leave.employee.last_name}')
            ws.cell(row=row, column=3, value=leave.employee.department.name if leave.employee.department else '-')
            ws.cell(row=row, column=4, value=leave.leave_type.name if leave.leave_type else '-')
            ws.cell(row=row, column=5, value=leave.start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=leave.end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=leave.days)
            ws.cell(row=row, column=8, value=status_map.get(leave.status, leave.status))
            ws.cell(row=row, column=9, value=leave.reason or '-')

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'leaves_report_{year}.xlsx')


class ExportPayrollExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير الرواتب إلى Excel"""
    permission_required = 'hr.view_payroll'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        year = int(request.GET.get('year', today.year))
        month = request.GET.get('month')

        # بناء الاستعلام
        details_filter = Q(payroll__company=company, payroll__period_year=year, payroll__status__in=['approved', 'paid'])
        if month:
            details_filter &= Q(payroll__period_month=int(month))

        details = PayrollDetail.objects.filter(details_filter).select_related(
            'employee', 'employee__department', 'payroll'
        ).order_by('payroll__period_month', 'employee__first_name')

        # إنشاء Excel
        wb, ws = self.create_workbook('الرواتب')

        # العناوين
        headers = ['الشهر', 'رقم الموظف', 'اسم الموظف', 'القسم', 'الراتب الأساسي',
                   'البدلات', 'الخصومات', 'صافي الراتب', 'أيام العمل', 'الأيام الفعلية']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        month_names = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
                       'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']

        # البيانات
        for row, detail in enumerate(details, 2):
            ws.cell(row=row, column=1, value=month_names[detail.payroll.period_month - 1])
            ws.cell(row=row, column=2, value=detail.employee.employee_number)
            ws.cell(row=row, column=3, value=f'{detail.employee.first_name} {detail.employee.last_name}')
            ws.cell(row=row, column=4, value=detail.employee.department.name if detail.employee.department else '-')
            ws.cell(row=row, column=5, value=float(detail.basic_salary) if detail.basic_salary else 0)
            ws.cell(row=row, column=6, value=float(detail.total_allowances) if detail.total_allowances else 0)
            ws.cell(row=row, column=7, value=float(detail.total_deductions) if detail.total_deductions else 0)
            ws.cell(row=row, column=8, value=float(detail.net_salary) if detail.net_salary else 0)
            ws.cell(row=row, column=9, value=detail.working_days or 0)
            ws.cell(row=row, column=10, value=detail.actual_days or 0)

        # تنسيق الأرقام
        for row in range(2, ws.max_row + 1):
            for col in [5, 6, 7, 8]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'

        self.auto_fit_columns(ws)

        filename = f'payroll_report_{year}'
        if month:
            filename += f'_{month}'

        return self.get_excel_response(wb, f'{filename}.xlsx')


class ExportAdvancesExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير السلف إلى Excel"""
    permission_required = 'hr.view_advance'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        year = int(request.GET.get('year', today.year))
        status = request.GET.get('status')

        # بناء الاستعلام
        advances = Advance.objects.filter(
            employee__company=company,
            request_date__year=year
        ).select_related('employee', 'employee__department').order_by('-request_date')

        if status:
            advances = advances.filter(status=status)

        # إنشاء Excel
        wb, ws = self.create_workbook('السلف')

        # العناوين
        headers = ['رقم الموظف', 'اسم الموظف', 'القسم', 'تاريخ الطلب',
                   'المبلغ', 'المبلغ المتبقي', 'عدد الأقساط', 'الحالة', 'السبب']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        status_map = {'pending': 'معلق', 'approved': 'موافق', 'rejected': 'مرفوض',
                      'disbursed': 'مصروف', 'fully_paid': 'مسدد بالكامل'}

        # البيانات
        for row, adv in enumerate(advances, 2):
            ws.cell(row=row, column=1, value=adv.employee.employee_number)
            ws.cell(row=row, column=2, value=f'{adv.employee.first_name} {adv.employee.last_name}')
            ws.cell(row=row, column=3, value=adv.employee.department.name if adv.employee.department else '-')
            ws.cell(row=row, column=4, value=adv.request_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=float(adv.amount))
            ws.cell(row=row, column=6, value=float(adv.remaining_amount) if adv.remaining_amount else 0)
            ws.cell(row=row, column=7, value=adv.installments_count or 1)
            ws.cell(row=row, column=8, value=status_map.get(adv.status, adv.status))
            ws.cell(row=row, column=9, value=adv.reason or '-')

        # تنسيق الأرقام
        for row in range(2, ws.max_row + 1):
            for col in [5, 6]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'advances_report_{year}.xlsx')


class ExportOvertimeExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير العمل الإضافي إلى Excel"""
    permission_required = 'hr.view_overtime'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        year = int(request.GET.get('year', today.year))
        month = request.GET.get('month')

        # بناء الاستعلام
        overtime_qs = Overtime.objects.filter(
            employee__company=company,
            date__year=year
        ).select_related('employee', 'employee__department').order_by('-date')

        if month:
            overtime_qs = overtime_qs.filter(date__month=int(month))

        # إنشاء Excel
        wb, ws = self.create_workbook('العمل الإضافي')

        # العناوين
        headers = ['التاريخ', 'رقم الموظف', 'اسم الموظف', 'القسم',
                   'عدد الساعات', 'المبلغ', 'الحالة', 'السبب']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        status_map = {'pending': 'معلق', 'approved': 'موافق', 'rejected': 'مرفوض'}

        # البيانات
        for row, ot in enumerate(overtime_qs, 2):
            ws.cell(row=row, column=1, value=ot.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=ot.employee.employee_number)
            ws.cell(row=row, column=3, value=f'{ot.employee.first_name} {ot.employee.last_name}')
            ws.cell(row=row, column=4, value=ot.employee.department.name if ot.employee.department else '-')
            ws.cell(row=row, column=5, value=float(ot.hours))
            ws.cell(row=row, column=6, value=float(ot.amount) if ot.amount else 0)
            ws.cell(row=row, column=7, value=status_map.get(ot.status, ot.status))
            ws.cell(row=row, column=8, value=ot.reason or '-')

        # تنسيق الأرقام
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=6).number_format = '#,##0.00'

        self.auto_fit_columns(ws)

        filename = f'overtime_report_{year}'
        if month:
            filename += f'_{month}'

        return self.get_excel_response(wb, f'{filename}.xlsx')


class ExportContractsExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير العقود إلى Excel"""
    permission_required = 'hr.view_employeecontract'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        company = request.current_company
        today = date.today()

        # الفلاتر
        days_ahead = int(request.GET.get('days', 90))
        expiring_date = today + timedelta(days=days_ahead)

        # العقود
        contracts = EmployeeContract.objects.filter(
            company=company,
            status='active',
            end_date__lte=expiring_date
        ).select_related('employee', 'employee__department').order_by('end_date')

        # إنشاء Excel
        wb, ws = self.create_workbook('العقود')

        # العناوين
        headers = ['رقم الموظف', 'اسم الموظف', 'القسم', 'نوع العقد',
                   'تاريخ البداية', 'تاريخ النهاية', 'الأيام المتبقية', 'الحالة']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        # البيانات
        for row, contract in enumerate(contracts, 2):
            days_remaining = (contract.end_date - today).days if contract.end_date else 0
            status_text = 'منتهي' if days_remaining < 0 else f'{days_remaining} يوم'

            ws.cell(row=row, column=1, value=contract.employee.employee_number)
            ws.cell(row=row, column=2, value=f'{contract.employee.first_name} {contract.employee.last_name}')
            ws.cell(row=row, column=3, value=contract.employee.department.name if contract.employee.department else '-')
            ws.cell(row=row, column=4, value=contract.contract_type or '-')
            ws.cell(row=row, column=5, value=contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '-')
            ws.cell(row=row, column=6, value=contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '-')
            ws.cell(row=row, column=7, value=days_remaining)
            ws.cell(row=row, column=8, value=status_text)

            # تلوين الصفوف حسب الحالة
            if days_remaining < 0:
                fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
            elif days_remaining <= 30:
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'contracts_expiry_report_{today}.xlsx')


# ==========================================
# Biometric Reports - Phase 8
# ==========================================

class BiometricReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير سجلات البصمة"""
    template_name = 'hr/reports/biometric_report.html'
    permission_required = 'hr.view_attendance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        # استيراد نماذج البصمة
        from ..models import BiometricDevice, BiometricLog, BiometricSyncLog, EmployeeBiometricMapping

        # الفلاتر
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        device_id = self.request.GET.get('device')
        department_id = self.request.GET.get('department')
        employee_id = self.request.GET.get('employee')

        # افتراضي: الشهر الحالي
        if not from_date:
            from_date = today.replace(day=1).isoformat()
        if not to_date:
            to_date = today.isoformat()

        # إحصائيات الأجهزة
        device_stats = BiometricDevice.objects.filter(company=company).aggregate(
            total_devices=Count('id'),
            active_devices=Count('id', filter=Q(status='active', is_active=True)),
            offline_devices=Count('id', filter=Q(status='offline')),
            maintenance_devices=Count('id', filter=Q(status='maintenance')),
        )

        # إحصائيات الربط
        mapping_stats = EmployeeBiometricMapping.objects.filter(company=company).aggregate(
            total_mappings=Count('id'),
            enrolled=Count('id', filter=Q(is_enrolled=True)),
            active=Count('id', filter=Q(is_active=True)),
        )

        # بناء استعلام السجلات
        logs = BiometricLog.objects.filter(
            company=company,
            punch_time__date__gte=from_date,
            punch_time__date__lte=to_date
        ).select_related('device', 'employee', 'employee__department')

        if device_id:
            logs = logs.filter(device_id=device_id)
        if department_id:
            logs = logs.filter(employee__department_id=department_id)
        if employee_id:
            logs = logs.filter(employee_id=employee_id)

        # إحصائيات السجلات
        log_stats = logs.aggregate(
            total_logs=Count('id'),
            processed=Count('id', filter=Q(is_processed=True)),
            unprocessed=Count('id', filter=Q(is_processed=False)),
            in_punches=Count('id', filter=Q(punch_type='in')),
            out_punches=Count('id', filter=Q(punch_type='out')),
            by_fingerprint=Count('id', filter=Q(verification_type='fingerprint')),
            by_face=Count('id', filter=Q(verification_type='face')),
            by_card=Count('id', filter=Q(verification_type='card')),
        )

        # تفصيل حسب الجهاز
        by_device = logs.values(
            'device__name'
        ).annotate(
            total=Count('id'),
            in_count=Count('id', filter=Q(punch_type='in')),
            out_count=Count('id', filter=Q(punch_type='out')),
        ).order_by('-total')

        # تفصيل حسب الموظف
        by_employee = logs.filter(employee__isnull=False).values(
            'employee__id',
            'employee__first_name',
            'employee__last_name',
            'employee__employee_number',
            'employee__department__name'
        ).annotate(
            total_punches=Count('id'),
            in_punches=Count('id', filter=Q(punch_type='in')),
            out_punches=Count('id', filter=Q(punch_type='out')),
            first_in=Min('punch_time', filter=Q(punch_type='in')),
            last_out=Max('punch_time', filter=Q(punch_type='out')),
        ).order_by('-total_punches')[:50]

        # التوزيع اليومي
        daily_trend = logs.values('punch_time__date').annotate(
            total=Count('id'),
            in_count=Count('id', filter=Q(punch_type='in')),
            out_count=Count('id', filter=Q(punch_type='out')),
        ).order_by('punch_time__date')

        # سجلات المزامنة الأخيرة
        recent_syncs = BiometricSyncLog.objects.filter(
            company=company
        ).select_related('device').order_by('-created_at')[:10]

        # الأجهزة للفلتر
        devices = BiometricDevice.objects.filter(company=company, is_active=True)

        context.update({
            'title': _('تقرير سجلات البصمة'),
            'device_stats': device_stats,
            'mapping_stats': mapping_stats,
            'log_stats': log_stats,
            'by_device': list(by_device),
            'by_employee': list(by_employee),
            'daily_trend': json.dumps(list(daily_trend), default=str),
            'recent_syncs': recent_syncs,
            'from_date': from_date,
            'to_date': to_date,
            'devices': devices,
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_device': device_id,
            'selected_department': department_id,
            'selected_employee': employee_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير البصمة'), 'url': None},
            ],
        })
        return context


class BiometricAttendanceReportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """تقرير الحضور من البصمة - تفصيلي"""
    template_name = 'hr/reports/biometric_attendance_report.html'
    permission_required = 'hr.view_attendance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.request.current_company
        today = date.today()

        from ..models import BiometricLog, Employee

        # الفلاتر
        report_date = self.request.GET.get('date', today.isoformat())
        department_id = self.request.GET.get('department')

        # الموظفين مع معلومات البصمة لهذا اليوم
        employees = Employee.objects.filter(
            company=company,
            is_active=True,
            status='active'
        ).select_related('department', 'job_title')

        if department_id:
            employees = employees.filter(department_id=department_id)

        # جمع بيانات البصمة لكل موظف
        attendance_data = []
        for emp in employees:
            # سجلات البصمة لهذا الموظف في هذا اليوم
            logs = BiometricLog.objects.filter(
                employee=emp,
                punch_time__date=report_date
            ).order_by('punch_time')

            first_in = None
            last_out = None
            all_punches = []

            for log in logs:
                all_punches.append({
                    'time': log.punch_time,
                    'type': log.punch_type,
                    'verification': log.verification_type,
                    'device': log.device.name if log.device else '-'
                })
                if log.punch_type == 'in':
                    if first_in is None or log.punch_time < first_in:
                        first_in = log.punch_time
                elif log.punch_type == 'out':
                    if last_out is None or log.punch_time > last_out:
                        last_out = log.punch_time

            # حساب ساعات العمل
            working_hours = None
            if first_in and last_out:
                delta = last_out - first_in
                working_hours = delta.total_seconds() / 3600

            attendance_data.append({
                'employee': emp,
                'first_in': first_in,
                'last_out': last_out,
                'working_hours': working_hours,
                'punch_count': len(all_punches),
                'punches': all_punches,
                'status': 'present' if first_in else 'absent'
            })

        # إحصائيات
        present_count = sum(1 for d in attendance_data if d['status'] == 'present')
        absent_count = sum(1 for d in attendance_data if d['status'] == 'absent')
        total_hours = sum(d['working_hours'] or 0 for d in attendance_data)

        context.update({
            'title': _('تقرير الحضور اليومي من البصمة'),
            'attendance_data': attendance_data,
            'report_date': report_date,
            'stats': {
                'total': len(attendance_data),
                'present': present_count,
                'absent': absent_count,
                'total_hours': round(total_hours, 2),
                'avg_hours': round(total_hours / present_count, 2) if present_count > 0 else 0
            },
            'departments': Department.objects.filter(company=company, is_active=True),
            'selected_department': department_id,
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': '/'},
                {'title': _('الموارد البشرية'), 'url': '/hr/'},
                {'title': _('التقارير'), 'url': '/hr/reports/'},
                {'title': _('تقرير الحضور اليومي'), 'url': None},
            ],
        })
        return context


class ExportBiometricLogsExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير سجلات البصمة إلى Excel"""
    permission_required = 'hr.view_attendance'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        from ..models import BiometricLog

        company = request.current_company
        today = date.today()

        # الفلاتر
        from_date = request.GET.get('from_date', today.replace(day=1).isoformat())
        to_date = request.GET.get('to_date', today.isoformat())
        device_id = request.GET.get('device')

        # بناء الاستعلام
        logs = BiometricLog.objects.filter(
            company=company,
            punch_time__date__gte=from_date,
            punch_time__date__lte=to_date
        ).select_related('device', 'employee', 'employee__department').order_by('punch_time')

        if device_id:
            logs = logs.filter(device_id=device_id)

        # إنشاء Excel
        wb, ws = self.create_workbook('سجلات البصمة')

        # العناوين
        headers = ['التاريخ', 'الوقت', 'رقم الموظف', 'اسم الموظف', 'القسم',
                   'الجهاز', 'نوع البصمة', 'طريقة التحقق', 'الحالة']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        punch_types = {'in': 'حضور', 'out': 'انصراف', 'break_out': 'خروج استراحة',
                       'break_in': 'عودة استراحة', 'overtime_in': 'بداية إضافي', 'overtime_out': 'نهاية إضافي'}
        verification_types = {'fingerprint': 'بصمة', 'face': 'وجه', 'card': 'بطاقة', 'password': 'كلمة مرور'}

        # البيانات
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log.punch_time.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=log.punch_time.strftime('%H:%M:%S'))
            ws.cell(row=row, column=3, value=log.employee.employee_number if log.employee else log.device_user_id)
            ws.cell(row=row, column=4, value=f'{log.employee.first_name} {log.employee.last_name}' if log.employee else '-')
            ws.cell(row=row, column=5, value=log.employee.department.name if log.employee and log.employee.department else '-')
            ws.cell(row=row, column=6, value=log.device.name if log.device else '-')
            ws.cell(row=row, column=7, value=punch_types.get(log.punch_type, log.punch_type))
            ws.cell(row=row, column=8, value=verification_types.get(log.verification_type, log.verification_type))
            ws.cell(row=row, column=9, value='معالج' if log.is_processed else 'غير معالج')

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'biometric_logs_{from_date}_{to_date}.xlsx')


class ExportBiometricAttendanceExcelView(LoginRequiredMixin, PermissionRequiredMixin, ExcelExportMixin, View):
    """تصدير تقرير الحضور من البصمة إلى Excel"""
    permission_required = 'hr.view_attendance'

    def get(self, request):
        if not EXCEL_AVAILABLE:
            return HttpResponse('مكتبة openpyxl غير متوفرة', status=500)

        from ..models import BiometricLog

        company = request.current_company
        today = date.today()

        # الفلاتر
        from_date = request.GET.get('from_date', today.replace(day=1).isoformat())
        to_date = request.GET.get('to_date', today.isoformat())
        department_id = request.GET.get('department')

        # الموظفين
        employees = Employee.objects.filter(
            company=company,
            is_active=True,
            status='active'
        ).select_related('department')

        if department_id:
            employees = employees.filter(department_id=department_id)

        # إنشاء Excel
        wb, ws = self.create_workbook('حضور البصمة')

        # العناوين
        headers = ['رقم الموظف', 'اسم الموظف', 'القسم', 'التاريخ',
                   'أول حضور', 'آخر انصراف', 'ساعات العمل', 'عدد البصمات']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self.style_header(ws, row=1, start_col=1, end_col=len(headers))

        # جمع البيانات
        from datetime import datetime
        start = datetime.strptime(from_date, '%Y-%m-%d').date()
        end = datetime.strptime(to_date, '%Y-%m-%d').date()

        row_num = 2
        current = start
        while current <= end:
            for emp in employees:
                logs = BiometricLog.objects.filter(
                    employee=emp,
                    punch_time__date=current
                ).order_by('punch_time')

                if logs.exists():
                    first_in = None
                    last_out = None

                    for log in logs:
                        if log.punch_type == 'in':
                            if first_in is None or log.punch_time < first_in:
                                first_in = log.punch_time
                        elif log.punch_type == 'out':
                            if last_out is None or log.punch_time > last_out:
                                last_out = log.punch_time

                    working_hours = 0
                    if first_in and last_out:
                        delta = last_out - first_in
                        working_hours = round(delta.total_seconds() / 3600, 2)

                    ws.cell(row=row_num, column=1, value=emp.employee_number)
                    ws.cell(row=row_num, column=2, value=f'{emp.first_name} {emp.last_name}')
                    ws.cell(row=row_num, column=3, value=emp.department.name if emp.department else '-')
                    ws.cell(row=row_num, column=4, value=current.strftime('%Y-%m-%d'))
                    ws.cell(row=row_num, column=5, value=first_in.strftime('%H:%M:%S') if first_in else '-')
                    ws.cell(row=row_num, column=6, value=last_out.strftime('%H:%M:%S') if last_out else '-')
                    ws.cell(row=row_num, column=7, value=working_hours)
                    ws.cell(row=row_num, column=8, value=logs.count())

                    row_num += 1

            current += timedelta(days=1)

        self.auto_fit_columns(ws)

        return self.get_excel_response(wb, f'biometric_attendance_{from_date}_{to_date}.xlsx')


# Export views __all__
__all__ = [
    'HRDashboardReportView',
    'EmployeeReportView',
    'AttendanceReportView',
    'LeaveReportView',
    'PayrollSummaryReportView',
    'AdvanceReportView',
    'OvertimeReportView',
    'DepartmentAnalysisReportView',
    'ContractExpiryReportView',
    # Biometric Reports
    'BiometricReportView',
    'BiometricAttendanceReportView',
    # Excel Export Views
    'ExportEmployeesExcelView',
    'ExportAttendanceExcelView',
    'ExportLeavesExcelView',
    'ExportPayrollExcelView',
    'ExportAdvancesExcelView',
    'ExportOvertimeExcelView',
    'ExportContractsExcelView',
    'ExportBiometricLogsExcelView',
    'ExportBiometricAttendanceExcelView',
]
