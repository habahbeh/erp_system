# apps/accounting/views/voucher_views.py
"""
Voucher Views - إدارة السندات (القبض والصرف)
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.db.models import Q, Sum, Count
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import ValidationError

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import PaymentVoucher, ReceiptVoucher
from ..forms.voucher_forms import PaymentVoucherForm, ReceiptVoucherForm


# ========== Payment Voucher Views ==========

class PaymentVoucherListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة سندات الصرف"""

    model = PaymentVoucher
    template_name = 'accounting/vouchers/payment_voucher_list.html'
    context_object_name = 'vouchers'
    permission_required = 'accounting.view_paymentvoucher'
    paginate_by = 25

    def get_queryset(self):
        queryset = PaymentVoucher.objects.filter(
            company=self.request.current_company
        ).select_related('cash_account', 'expense_account', 'currency', 'created_by').order_by('-date', '-number')

        # الفلترة
        status = self.request.GET.get('status')
        payment_method = self.request.GET.get('payment_method')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        if date_from:
            queryset = queryset.filter(date__gte=date_from)

        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(beneficiary_name__icontains=search) |
                Q(description__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات سريعة
        company = self.request.current_company
        stats = PaymentVoucher.objects.filter(company=company).aggregate(
            total=Count('id'),
            draft=Count('id', filter=Q(status='draft')),
            posted=Count('id', filter=Q(status='posted')),
            total_amount=Sum('amount', filter=Q(status='posted'))
        )

        context.update({
            'title': _('سندات الصرف'),
            'can_add': self.request.user.has_perm('accounting.add_paymentvoucher'),
            'can_edit': self.request.user.has_perm('accounting.change_paymentvoucher'),
            'can_delete': self.request.user.has_perm('accounting.delete_paymentvoucher'),
            'status_choices': PaymentVoucher.STATUS_CHOICES,
            'payment_method_choices': PaymentVoucher.PAYMENT_METHODS,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات الصرف'), 'url': ''},
            ]
        })
        return context


class PaymentVoucherCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء سند صرف جديد"""

    model = PaymentVoucher
    form_class = PaymentVoucherForm
    template_name = 'accounting/vouchers/payment_voucher_form.html'
    permission_required = 'accounting.add_paymentvoucher'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'تم إنشاء سند الصرف {self.object.number} بنجاح')
        return response

    def get_success_url(self):
        return reverse('accounting:payment_voucher_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إنشاء سند صرف جديد'),
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات الصرف'), 'url': reverse('accounting:payment_voucher_list')},
                {'title': _('إنشاء جديد'), 'url': ''},
            ]
        })
        return context


class PaymentVoucherDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل سند الصرف"""

    model = PaymentVoucher
    template_name = 'accounting/vouchers/payment_voucher_detail.html'
    context_object_name = 'voucher'
    permission_required = 'accounting.view_paymentvoucher'

    def get_queryset(self):
        return PaymentVoucher.objects.filter(
            company=self.request.current_company
        ).select_related('cash_account', 'expense_account', 'currency', 'created_by', 'posted_by', 'journal_entry')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'سند الصرف {self.object.number}',
            'can_edit': self.request.user.has_perm('accounting.change_paymentvoucher') and self.object.can_edit(),
            'can_delete': self.request.user.has_perm('accounting.delete_paymentvoucher') and self.object.can_delete(),
            'can_post': self.request.user.has_perm('accounting.change_paymentvoucher') and self.object.can_post(),
            'can_unpost': self.request.user.has_perm('accounting.change_paymentvoucher') and self.object.can_unpost(),
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات الصرف'), 'url': reverse('accounting:payment_voucher_list')},
                {'title': self.object.number, 'url': ''},
            ]
        })
        return context


class PaymentVoucherUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل سند الصرف"""

    model = PaymentVoucher
    form_class = PaymentVoucherForm
    template_name = 'accounting/vouchers/payment_voucher_form.html'
    permission_required = 'accounting.change_paymentvoucher'

    def get_queryset(self):
        return PaymentVoucher.objects.filter(company=self.request.current_company)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        # التحقق من إمكانية التعديل
        if not self.object.can_edit():
            messages.error(self.request, _('لا يمكن تعديل سند مرحل'))
            return redirect('accounting:payment_voucher_detail', pk=self.object.pk)

        response = super().form_valid(form)
        messages.success(self.request, f'تم تحديث سند الصرف {self.object.number} بنجاح')
        return response

    def get_success_url(self):
        return reverse('accounting:payment_voucher_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل سند الصرف {self.object.number}',
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات الصرف'), 'url': reverse('accounting:payment_voucher_list')},
                {'title': f'تعديل {self.object.number}', 'url': ''},
            ]
        })
        return context


class PaymentVoucherDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف سند الصرف"""

    model = PaymentVoucher
    template_name = 'accounting/vouchers/payment_voucher_confirm_delete.html'
    permission_required = 'accounting.delete_paymentvoucher'
    success_url = reverse_lazy('accounting:payment_voucher_list')

    def get_queryset(self):
        return PaymentVoucher.objects.filter(company=self.request.current_company)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من إمكانية الحذف
        if not self.object.can_delete():
            messages.error(request, _('لا يمكن حذف سند مرحل'))
            return redirect('accounting:payment_voucher_detail', pk=self.object.pk)

        voucher_number = self.object.number
        messages.success(request, f'تم حذف سند الصرف {voucher_number} بنجاح')
        return super().delete(request, *args, **kwargs)


# ========== Receipt Voucher Views ==========

class ReceiptVoucherListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة سندات القبض"""

    model = ReceiptVoucher
    template_name = 'accounting/vouchers/receipt_voucher_list.html'
    context_object_name = 'vouchers'
    permission_required = 'accounting.view_receiptvoucher'
    paginate_by = 25

    def get_queryset(self):
        queryset = ReceiptVoucher.objects.filter(
            company=self.request.current_company
        ).select_related('cash_account', 'income_account', 'currency', 'created_by').order_by('-date', '-number')

        # الفلترة
        status = self.request.GET.get('status')
        receipt_method = self.request.GET.get('receipt_method')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if receipt_method:
            queryset = queryset.filter(receipt_method=receipt_method)

        if date_from:
            queryset = queryset.filter(date__gte=date_from)

        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(number__icontains=search) |
                Q(received_from__icontains=search) |
                Q(description__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات سريعة
        company = self.request.current_company
        stats = ReceiptVoucher.objects.filter(company=company).aggregate(
            total=Count('id'),
            draft=Count('id', filter=Q(status='draft')),
            posted=Count('id', filter=Q(status='posted')),
            total_amount=Sum('amount', filter=Q(status='posted'))
        )

        context.update({
            'title': _('سندات القبض'),
            'can_add': self.request.user.has_perm('accounting.add_receiptvoucher'),
            'can_edit': self.request.user.has_perm('accounting.change_receiptvoucher'),
            'can_delete': self.request.user.has_perm('accounting.delete_receiptvoucher'),
            'status_choices': ReceiptVoucher.STATUS_CHOICES,
            'receipt_method_choices': ReceiptVoucher.RECEIPT_METHODS,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات القبض'), 'url': ''},
            ]
        })
        return context


class ReceiptVoucherCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء سند قبض جديد"""

    model = ReceiptVoucher
    form_class = ReceiptVoucherForm
    template_name = 'accounting/vouchers/receipt_voucher_form.html'
    permission_required = 'accounting.add_receiptvoucher'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, f'تم إنشاء سند القبض {self.object.number} بنجاح')
        return response

    def get_success_url(self):
        return reverse('accounting:receipt_voucher_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إنشاء سند قبض جديد'),
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات القبض'), 'url': reverse('accounting:receipt_voucher_list')},
                {'title': _('إنشاء جديد'), 'url': ''},
            ]
        })
        return context


class ReceiptVoucherDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل سند القبض"""

    model = ReceiptVoucher
    template_name = 'accounting/vouchers/receipt_voucher_detail.html'
    context_object_name = 'voucher'
    permission_required = 'accounting.view_receiptvoucher'

    def get_queryset(self):
        return ReceiptVoucher.objects.filter(
            company=self.request.current_company
        ).select_related('cash_account', 'income_account', 'currency', 'created_by', 'posted_by', 'journal_entry')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'سند القبض {self.object.number}',
            'can_edit': self.request.user.has_perm('accounting.change_receiptvoucher') and self.object.can_edit(),
            'can_delete': self.request.user.has_perm('accounting.delete_receiptvoucher') and self.object.can_delete(),
            'can_post': self.request.user.has_perm('accounting.change_receiptvoucher') and self.object.can_post(),
            'can_unpost': self.request.user.has_perm('accounting.change_receiptvoucher') and self.object.can_unpost(),
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات القبض'), 'url': reverse('accounting:receipt_voucher_list')},
                {'title': self.object.number, 'url': ''},
            ]
        })
        return context


class ReceiptVoucherUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل سند القبض"""

    model = ReceiptVoucher
    form_class = ReceiptVoucherForm
    template_name = 'accounting/vouchers/receipt_voucher_form.html'
    permission_required = 'accounting.change_receiptvoucher'

    def get_queryset(self):
        return ReceiptVoucher.objects.filter(company=self.request.current_company)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        # التحقق من إمكانية التعديل
        if not self.object.can_edit():
            messages.error(self.request, _('لا يمكن تعديل سند مرحل'))
            return redirect('accounting:receipt_voucher_detail', pk=self.object.pk)

        response = super().form_valid(form)
        messages.success(self.request, f'تم تحديث سند القبض {self.object.number} بنجاح')
        return response

    def get_success_url(self):
        return reverse('accounting:receipt_voucher_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل سند القبض {self.object.number}',
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('سندات القبض'), 'url': reverse('accounting:receipt_voucher_list')},
                {'title': f'تعديل {self.object.number}', 'url': ''},
            ]
        })
        return context


class ReceiptVoucherDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف سند القبض"""

    model = ReceiptVoucher
    template_name = 'accounting/vouchers/receipt_voucher_confirm_delete.html'
    permission_required = 'accounting.delete_receiptvoucher'
    success_url = reverse_lazy('accounting:receipt_voucher_list')

    def get_queryset(self):
        return ReceiptVoucher.objects.filter(company=self.request.current_company)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من إمكانية الحذف
        if not self.object.can_delete():
            messages.error(request, _('لا يمكن حذف سند مرحل'))
            return redirect('accounting:receipt_voucher_detail', pk=self.object.pk)

        voucher_number = self.object.number
        messages.success(request, f'تم حذف سند القبض {voucher_number} بنجاح')
        return super().delete(request, *args, **kwargs)


# ========== Ajax Functions للترحيل ==========

@login_required
@permission_required_with_message('accounting.change_paymentvoucher')
@require_http_methods(["POST"])
def post_payment_voucher(request, pk):
    """ترحيل سند الصرف"""
    try:
        voucher = get_object_or_404(
            PaymentVoucher,
            pk=pk,
            company=request.current_company
        )

        if not voucher.can_post():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن ترحيل هذا السند'
            })

        voucher.status = 'confirmed'  # تأكيد السند أولاً
        voucher.save()

        voucher.post(user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم ترحيل سند الصرف {voucher.number} بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في ترحيل السند: {str(e)}'
        })


@login_required
@permission_required_with_message('accounting.change_paymentvoucher')
@require_http_methods(["POST"])
def unpost_payment_voucher(request, pk):
    """إلغاء ترحيل سند الصرف"""
    try:
        voucher = get_object_or_404(
            PaymentVoucher,
            pk=pk,
            company=request.current_company
        )

        if not voucher.can_unpost():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن إلغاء ترحيل هذا السند'
            })

        voucher.unpost()

        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء ترحيل سند الصرف {voucher.number} بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إلغاء ترحيل السند: {str(e)}'
        })


@login_required
@permission_required_with_message('accounting.change_receiptvoucher')
@require_http_methods(["POST"])
def post_receipt_voucher(request, pk):
    """ترحيل سند القبض"""
    try:
        voucher = get_object_or_404(
            ReceiptVoucher,
            pk=pk,
            company=request.current_company
        )

        if not voucher.can_post():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن ترحيل هذا السند'
            })

        voucher.status = 'confirmed'  # تأكيد السند أولاً
        voucher.save()

        voucher.post(user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم ترحيل سند القبض {voucher.number} بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في ترحيل السند: {str(e)}'
        })


@login_required
@permission_required_with_message('accounting.change_receiptvoucher')
@require_http_methods(["POST"])
def unpost_receipt_voucher(request, pk):
    """إلغاء ترحيل سند القبض"""
    try:
        voucher = get_object_or_404(
            ReceiptVoucher,
            pk=pk,
            company=request.current_company
        )

        if not voucher.can_unpost():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن إلغاء ترحيل هذا السند'
            })

        voucher.unpost()

        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء ترحيل سند القبض {voucher.number} بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إلغاء ترحيل السند: {str(e)}'
        })


# ========== Ajax DataTables للسندات ==========

@login_required
@permission_required('accounting.view_paymentvoucher')
def payment_voucher_datatable_ajax(request):
    """Ajax endpoint لجدول سندات الصرف"""
    from django.http import JsonResponse
    from django.core.paginator import Paginator
    from django.db.models import Q

    # الحصول على المعاملات
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '').strip()
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')

    # أعمدة الجدول
    columns = ['number', 'date', 'beneficiary_name', 'amount', 'payment_method', 'status']

    # الاستعلام الأساسي
    queryset = PaymentVoucher.objects.filter(
        company=request.current_company
    ).select_related('cash_account', 'currency', 'created_by')

    # البحث
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(beneficiary_name__icontains=search_value) |
            Q(description__icontains=search_value)
        )

    # الترتيب
    if order_column < len(columns):
        order_field = columns[order_column]
        if order_dir == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field)
    else:
        queryset = queryset.order_by('-date', '-number')

    # العدد الكلي
    total_records = PaymentVoucher.objects.filter(company=request.current_company).count()
    filtered_records = queryset.count()

    # التقسيم على صفحات
    vouchers = queryset[start:start + length]

    # بناء البيانات
    data = []
    for voucher in vouchers:
        # تحديد لون الحالة
        if voucher.status == 'draft':
            status_badge = '<span class="badge bg-secondary">مسودة</span>'
        elif voucher.status == 'confirmed':
            status_badge = '<span class="badge bg-warning">مؤكد</span>'
        elif voucher.status == 'posted':
            status_badge = '<span class="badge bg-success">مرحل</span>'
        else:
            status_badge = '<span class="badge bg-danger">ملغي</span>'

        # بناء أزرار الإجراءات
        actions = f'''
        <div class="btn-group" role="group">
            <a href="/accounting/payment-vouchers/{voucher.pk}/" 
               class="btn btn-outline-info btn-sm" title="عرض">
                <i class="fas fa-eye"></i>
            </a>
        '''

        if voucher.can_edit() and request.user.has_perm('accounting.change_paymentvoucher'):
            actions += f'''
            <a href="/accounting/payment-vouchers/{voucher.pk}/update/" 
               class="btn btn-outline-primary btn-sm" title="تعديل">
                <i class="fas fa-edit"></i>
            </a>
            '''

        if voucher.can_post() and request.user.has_perm('accounting.change_paymentvoucher'):
            actions += f'''
            <button type="button" class="btn btn-outline-success btn-sm" 
                    onclick="postVoucher({voucher.pk})" title="ترحيل">
                <i class="fas fa-check"></i>
            </button>
            '''
        elif voucher.can_unpost() and request.user.has_perm('accounting.change_paymentvoucher'):
            actions += f'''
            <button type="button" class="btn btn-outline-warning btn-sm" 
                    onclick="unpostVoucher({voucher.pk})" title="إلغاء ترحيل">
                <i class="fas fa-undo"></i>
            </button>
            '''

        if voucher.can_delete() and request.user.has_perm('accounting.delete_paymentvoucher'):
            actions += f'''
            <a href="/accounting/payment-vouchers/{voucher.pk}/delete/" 
               class="btn btn-outline-danger btn-sm" title="حذف">
                <i class="fas fa-trash"></i>
            </a>
            '''

        actions += '</div>'

        data.append([
            f'<a href="/accounting/payment-vouchers/{voucher.pk}/" class="text-decoration-none">{voucher.number}</a>',
            voucher.date.strftime('%Y/%m/%d'),
            voucher.beneficiary_name,
            f'{voucher.amount:,.2f}',
            voucher.get_payment_method_display(),
            status_badge,
            actions
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@login_required
@permission_required('accounting.view_receiptvoucher')
def receipt_voucher_datatable_ajax(request):
    """Ajax endpoint لجدول سندات القبض"""
    from django.http import JsonResponse
    from django.core.paginator import Paginator
    from django.db.models import Q

    # الحصول على المعاملات
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '').strip()
    order_column = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'desc')

    # أعمدة الجدول
    columns = ['number', 'date', 'received_from', 'amount', 'receipt_method', 'status']

    # الاستعلام الأساسي
    queryset = ReceiptVoucher.objects.filter(
        company=request.current_company
    ).select_related('cash_account', 'currency', 'created_by')

    # البحث
    if search_value:
        queryset = queryset.filter(
            Q(number__icontains=search_value) |
            Q(received_from__icontains=search_value) |
            Q(description__icontains=search_value)
        )

    # الترتيب
    if order_column < len(columns):
        order_field = columns[order_column]
        if order_dir == 'desc':
            order_field = f'-{order_field}'
        queryset = queryset.order_by(order_field)
    else:
        queryset = queryset.order_by('-date', '-number')

    # العدد الكلي
    total_records = ReceiptVoucher.objects.filter(company=request.current_company).count()
    filtered_records = queryset.count()

    # التقسيم على صفحات
    vouchers = queryset[start:start + length]

    # بناء البيانات
    data = []
    for voucher in vouchers:
        # تحديد لون الحالة
        if voucher.status == 'draft':
            status_badge = '<span class="badge bg-secondary">مسودة</span>'
        elif voucher.status == 'confirmed':
            status_badge = '<span class="badge bg-warning">مؤكد</span>'
        elif voucher.status == 'posted':
            status_badge = '<span class="badge bg-success">مرحل</span>'
        else:
            status_badge = '<span class="badge bg-danger">ملغي</span>'

        # بناء أزرار الإجراءات
        actions = f'''
        <div class="btn-group" role="group">
            <a href="/accounting/receipt-vouchers/{voucher.pk}/" 
               class="btn btn-outline-info btn-sm" title="عرض">
                <i class="fas fa-eye"></i>
            </a>
        '''

        if voucher.can_edit() and request.user.has_perm('accounting.change_receiptvoucher'):
            actions += f'''
            <a href="/accounting/receipt-vouchers/{voucher.pk}/update/" 
               class="btn btn-outline-primary btn-sm" title="تعديل">
                <i class="fas fa-edit"></i>
            </a>
            '''

        if voucher.can_post() and request.user.has_perm('accounting.change_receiptvoucher'):
            actions += f'''
            <button type="button" class="btn btn-outline-success btn-sm" 
                    onclick="postVoucher({voucher.pk})" title="ترحيل">
                <i class="fas fa-check"></i>
            </button>
            '''
        elif voucher.can_unpost() and request.user.has_perm('accounting.change_receiptvoucher'):
            actions += f'''
            <button type="button" class="btn btn-outline-warning btn-sm" 
                    onclick="unpostVoucher({voucher.pk})" title="إلغاء ترحيل">
                <i class="fas fa-undo"></i>
            </button>
            '''

        if voucher.can_delete() and request.user.has_perm('accounting.delete_receiptvoucher'):
            actions += f'''
            <a href="/accounting/receipt-vouchers/{voucher.pk}/delete/" 
               class="btn btn-outline-danger btn-sm" title="حذف">
                <i class="fas fa-trash"></i>
            </a>
            '''

        actions += '</div>'

        data.append([
            f'<a href="/accounting/receipt-vouchers/{voucher.pk}/" class="text-decoration-none">{voucher.number}</a>',
            voucher.date.strftime('%Y/%m/%d'),
            voucher.received_from,
            f'{voucher.amount:,.2f}',
            voucher.get_receipt_method_display(),
            status_badge,
            actions
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


# ========== Export Functions ==========

@login_required
@permission_required('accounting.view_paymentvoucher')
def export_payment_vouchers(request):
    """تصدير سندات الصرف إلى Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    # إنشاء workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سندات الصرف"

    # العناوين
    headers = [
        'رقم السند', 'التاريخ', 'المستفيد', 'المبلغ', 'العملة',
        'طريقة الدفع', 'الحالة', 'البيان', 'حساب الصندوق', 'تاريخ الإنشاء'
    ]

    # إضافة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # الحصول على البيانات
    vouchers = PaymentVoucher.objects.filter(
        company=request.current_company
    ).select_related('cash_account', 'currency').order_by('-date', '-number')

    # إضافة البيانات
    for row, voucher in enumerate(vouchers, 2):
        ws.cell(row=row, column=1, value=voucher.number)
        ws.cell(row=row, column=2, value=voucher.date.strftime('%Y/%m/%d'))
        ws.cell(row=row, column=3, value=voucher.beneficiary_name)
        ws.cell(row=row, column=4, value=float(voucher.amount))
        ws.cell(row=row, column=5, value=voucher.currency.name)
        ws.cell(row=row, column=6, value=voucher.get_payment_method_display())
        ws.cell(row=row, column=7, value=voucher.get_status_display())
        ws.cell(row=row, column=8, value=voucher.description)
        ws.cell(row=row, column=9, value=f"{voucher.cash_account.code} - {voucher.cash_account.name}")
        ws.cell(row=row, column=10, value=voucher.created_at.strftime('%Y/%m/%d %H:%M'))

    # تعديل عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # إنشاء الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"payment_vouchers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@permission_required('accounting.view_receiptvoucher')
def export_receipt_vouchers(request):
    """تصدير سندات القبض إلى Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    # إنشاء workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سندات القبض"

    # العناوين
    headers = [
        'رقم السند', 'التاريخ', 'مستلم من', 'المبلغ', 'العملة',
        'طريقة القبض', 'الحالة', 'البيان', 'حساب الصندوق', 'تاريخ الإنشاء'
    ]

    # إضافة العناوين
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # الحصول على البيانات
    vouchers = ReceiptVoucher.objects.filter(
        company=request.current_company
    ).select_related('cash_account', 'currency').order_by('-date', '-number')

    # إضافة البيانات
    for row, voucher in enumerate(vouchers, 2):
        ws.cell(row=row, column=1, value=voucher.number)
        ws.cell(row=row, column=2, value=voucher.date.strftime('%Y/%m/%d'))
        ws.cell(row=row, column=3, value=voucher.received_from)
        ws.cell(row=row, column=4, value=float(voucher.amount))
        ws.cell(row=row, column=5, value=voucher.currency.name)
        ws.cell(row=row, column=6, value=voucher.get_receipt_method_display())
        ws.cell(row=row, column=7, value=voucher.get_status_display())
        ws.cell(row=row, column=8, value=voucher.description)
        ws.cell(row=row, column=9, value=f"{voucher.cash_account.code} - {voucher.cash_account.name}")
        ws.cell(row=row, column=10, value=voucher.created_at.strftime('%Y/%m/%d %H:%M'))

    # تعديل عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # إنشاء الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"receipt_vouchers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response