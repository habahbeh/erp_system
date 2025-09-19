# apps/accounting/views/account_type_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.db.models import Q, Count
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.template.loader import render_to_string
import json
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import AccountType
from ..forms.account_type_forms import AccountTypeForm, AccountTypeImportForm


class AccountTypeListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """عرض قائمة أنواع الحسابات"""

    template_name = 'accounting/account_types/account_type_list.html'
    permission_required = 'accounting.view_accounttype'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('أنواع الحسابات'),
            'can_add': self.request.user.has_perm('accounting.add_accounttype'),
            'can_edit': self.request.user.has_perm('accounting.change_accounttype'),
            'can_delete': self.request.user.has_perm('accounting.delete_accounttype'),
            'add_url': reverse('accounting:account_type_create'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('المحاسبة'), 'url': '#'},
                {'title': _('أنواع الحسابات'), 'url': ''}
            ],
        })
        return context


class AccountTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء نوع حساب جديد"""

    model = AccountType
    form_class = AccountTypeForm
    template_name = 'accounting/account_types/account_type_form.html'
    permission_required = 'accounting.add_accounttype'
    success_url = reverse_lazy('accounting:account_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة نوع حساب جديد'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('أنواع الحسابات'), 'url': reverse('accounting:account_type_list')},
                {'title': _('إضافة جديد'), 'url': ''}
            ],
        })
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('تم إضافة نوع الحساب بنجاح'))
        return response


class AccountTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل نوع حساب"""

    model = AccountType
    form_class = AccountTypeForm
    template_name = 'accounting/account_types/account_type_form.html'
    permission_required = 'accounting.change_accounttype'
    success_url = reverse_lazy('accounting:account_type_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('تعديل نوع الحساب'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('أنواع الحسابات'), 'url': reverse('accounting:account_type_list')},
                {'title': _('تعديل'), 'url': ''}
            ],
        })
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('تم تحديث نوع الحساب بنجاح'))
        return response


class AccountTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف نوع حساب"""

    model = AccountType
    permission_required = 'accounting.delete_accounttype'
    success_url = reverse_lazy('accounting:account_type_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من وجود حسابات مرتبطة
        if self.object.account_set.exists():
            messages.error(request, _('لا يمكن حذف نوع الحساب لوجود حسابات مرتبطة به'))
            return redirect('accounting:account_type_list')

        messages.success(request, _('تم حذف نوع الحساب بنجاح'))
        return super().delete(request, *args, **kwargs)


# Ajax Views
@login_required
@permission_required_with_message('accounting.view_accounttype')
@require_http_methods(["GET"])
def account_type_datatable_ajax(request):
    """Ajax endpoint لجدول أنواع الحسابات"""

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # الفلاتر المخصصة
    type_category = request.GET.get('type_category', '')
    normal_balance = request.GET.get('normal_balance', '')
    search_filter = request.GET.get('search_filter', '')

    # البحث والفلترة
    queryset = AccountType.objects.all()

    # تطبيق الفلاتر
    if type_category:
        queryset = queryset.filter(type_category=type_category)

    if normal_balance:
        queryset = queryset.filter(normal_balance=normal_balance)

    # البحث العام
    if search_value or search_filter:
        search_term = search_value or search_filter
        queryset = queryset.filter(
            Q(code__icontains=search_term) |
            Q(name__icontains=search_term)
        )

    # الترتيب
    order_column = request.GET.get('order[0][column]')
    order_dir = request.GET.get('order[0][dir]')

    if order_column:
        columns = ['code', 'name', 'type_category', 'normal_balance']
        if int(order_column) < len(columns):
            order_field = columns[int(order_column)]
            if order_dir == 'desc':
                order_field = '-' + order_field
            queryset = queryset.order_by(order_field)

    # العد الإجمالي
    total_records = AccountType.objects.count()
    filtered_records = queryset.count()

    # الصفحات
    queryset = queryset[start:start + length]

    # إعداد البيانات
    data = []
    can_edit = request.user.has_perm('accounting.change_accounttype')
    can_delete = request.user.has_perm('accounting.delete_accounttype')

    for account_type in queryset:
        # أزرار الإجراءات
        actions = []
        if can_edit:
            actions.append(f'''
                <a href="{reverse('accounting:account_type_update', args=[account_type.pk])}" 
                   class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                    <i class="fas fa-edit"></i>
                </a>
            ''')

        if can_delete:
            actions.append(f'''
                <button type="button" class="btn btn-outline-danger btn-sm" 
                        onclick="deleteAccountType({account_type.pk})" title="حذف" data-bs-toggle="tooltip">
                    <i class="fas fa-trash"></i>
                </button>
            ''')

        actions_html = ' '.join(actions) if actions else '-'

        data.append([
            account_type.code,
            account_type.name,
            account_type.get_type_category_display(),
            account_type.get_normal_balance_display(),
            actions_html
        ])

    return JsonResponse({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@login_required
@permission_required_with_message('accounting.view_accounttype')
def export_account_types(request):
    """تصدير أنواع الحسابات إلى Excel"""

    # إنشاء workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "أنواع الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # إضافة الرأس
    headers = ['الرمز', 'الاسم', 'تصنيف الحساب', 'الرصيد الطبيعي']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # إضافة البيانات
    queryset = AccountType.objects.all().order_by('code')
    for row, account_type in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=account_type.code)
        ws.cell(row=row, column=2, value=account_type.name)
        ws.cell(row=row, column=3, value=account_type.get_type_category_display())
        ws.cell(row=row, column=4, value=account_type.get_normal_balance_display())

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="account_types.xlsx"'

    wb.save(response)
    return response


@login_required
@permission_required_with_message('accounting.add_accounttype')
def import_account_types(request):
    """استيراد أنواع الحسابات من Excel/CSV"""

    if request.method == 'POST':
        form = AccountTypeImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = form.cleaned_data['file']

                # قراءة الملف
                if file.name.lower().endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # التحقق من الأعمدة المطلوبة
                required_columns = ['code', 'name', 'type_category', 'normal_balance']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, _('الملف لا يحتوي على الأعمدة المطلوبة'))
                    return redirect('accounting:account_type_list')

                # استيراد البيانات
                created_count = 0
                updated_count = 0
                errors = []

                for index, row in df.iterrows():
                    try:
                        account_type, created = AccountType.objects.get_or_create(
                            code=row['code'].upper(),
                            defaults={
                                'name': row['name'],
                                'type_category': row['type_category'],
                                'normal_balance': row['normal_balance']
                            }
                        )

                        if created:
                            created_count += 1
                        else:
                            # تحديث البيانات الموجودة
                            account_type.name = row['name']
                            account_type.type_category = row['type_category']
                            account_type.normal_balance = row['normal_balance']
                            account_type.save()
                            updated_count += 1

                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')

                # رسائل النتائج
                if created_count > 0:
                    messages.success(request, f'تم إنشاء {created_count} نوع حساب جديد')
                if updated_count > 0:
                    messages.success(request, f'تم تحديث {updated_count} نوع حساب')
                if errors:
                    for error in errors[:5]:  # إظهار أول 5 أخطاء فقط
                        messages.error(request, error)

            except Exception as e:
                messages.error(request, f'خطأ في معالجة الملف: {str(e)}')

    return redirect('accounting:account_type_list')