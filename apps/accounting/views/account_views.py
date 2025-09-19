# apps/accounting/views/account_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.db.models import Q, Count, Case, When, IntegerField
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.template.loader import render_to_string
import json
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import Account, AccountType
from ..forms.account_forms import AccountForm, AccountImportForm, AccountFilterForm
from apps.core.decorators import company_required


class AccountListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """عرض قائمة الحسابات"""

    template_name = 'accounting/accounts/account_list.html'
    permission_required = 'accounting.view_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('دليل الحسابات'),
            'can_add': self.request.user.has_perm('accounting.add_account'),
            'can_edit': self.request.user.has_perm('accounting.change_account'),
            'can_delete': self.request.user.has_perm('accounting.delete_account'),
            'add_url': reverse('accounting:account_create'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('المحاسبة'), 'url': '#'},
                {'title': _('دليل الحسابات'), 'url': ''}
            ],
        })
        return context


class AccountCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء حساب جديد"""

    model = Account
    form_class = AccountForm
    template_name = 'accounting/accounts/account_form.html'
    permission_required = 'accounting.add_account'
    success_url = reverse_lazy('accounting:account_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة حساب جديد'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('دليل الحسابات'), 'url': reverse('accounting:account_list')},
                {'title': _('إضافة جديد'), 'url': ''}
            ],
        })
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.company = self.request.current_company
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        messages.success(self.request, _('تم إضافة الحساب بنجاح'))
        return response


class AccountUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل حساب"""

    model = Account
    form_class = AccountForm
    template_name = 'accounting/accounts/account_form.html'
    permission_required = 'accounting.change_account'
    success_url = reverse_lazy('accounting:account_list')

    def get_queryset(self):
        return Account.objects.filter(company=self.request.current_company)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('تعديل الحساب'),
            'breadcrumbs': [
                {'title': _('الرئيسية'), 'url': reverse('core:dashboard')},
                {'title': _('دليل الحسابات'), 'url': reverse('accounting:account_list')},
                {'title': _('تعديل'), 'url': ''}
            ],
        })
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        account = form.instance

        # التحقق الإضافي قبل الحفظ
        if account.children.exists() and account.accept_entries:
            messages.warning(
                self.request,
                _('تم تعديل "يقبل قيود مباشرة" تلقائياً لأن الحساب له حسابات فرعية')
            )

        return super().form_valid(form)


class AccountDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف حساب"""

    model = Account
    permission_required = 'accounting.delete_account'
    success_url = reverse_lazy('accounting:account_list')

    def get_queryset(self):
        return Account.objects.filter(company=self.request.current_company)

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من وجود حسابات فرعية
        if self.object.children.exists():
            messages.error(request, _('لا يمكن حذف الحساب لوجود حسابات فرعية'))
            return redirect('accounting:account_list')

        # التحقق من وجود قيود محاسبية (سيتم إضافته لاحقاً)
        # if self.object.journal_entries.exists():
        #     messages.error(request, _('لا يمكن حذف الحساب لوجود قيود محاسبية'))
        #     return redirect('accounting:account_list')

        messages.success(request, _('تم حذف الحساب بنجاح'))
        return super().delete(request, *args, **kwargs)


# Ajax Views
@company_required
@login_required
@permission_required_with_message('accounting.view_account')
@require_http_methods(["GET"])
def account_datatable_ajax(request):
    """Ajax endpoint لجدول الحسابات"""

    # التحقق من وجود current_company
    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # الفلاتر المخصصة
    account_type_id = request.GET.get('account_type', '')
    is_active = request.GET.get('is_active', '')
    has_children = request.GET.get('has_children', '')
    search_filter = request.GET.get('search_filter', '')

    # إذا كان طلب للحصول على أنواع الحسابات للفلتر
    if request.GET.get('get_account_types') == '1':
        from .models import AccountType
        account_types = AccountType.objects.all().values('id', 'name')
        return JsonResponse({'account_types': list(account_types)})

    try:
        # البحث والفلترة
        queryset = Account.objects.filter(
            company=request.current_company
        ).select_related('account_type', 'parent', 'currency').annotate(
            children_count=Count('children')
        )

        # تطبيق الفلاتر
        if account_type_id:
            queryset = queryset.filter(account_type_id=account_type_id)

        if is_active == '1':
            queryset = queryset.filter(is_suspended=False)
        elif is_active == '0':
            queryset = queryset.filter(is_suspended=True)

        if has_children == '1':
            queryset = queryset.filter(children_count__gt=0)
        elif has_children == '0':
            queryset = queryset.filter(children_count=0)

        # البحث العام
        if search_value or search_filter:
            search_term = search_value or search_filter
            queryset = queryset.filter(
                Q(code__icontains=search_term) |
                Q(name__icontains=search_term) |
                Q(name_en__icontains=search_term)
            )

        # الترتيب
        order_column = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]')

        if order_column:
            columns = ['code', 'name', 'account_type__name', 'parent__name', 'is_suspended']
            if int(order_column) < len(columns):
                order_field = columns[int(order_column)]
                if order_dir == 'desc':
                    order_field = '-' + order_field
                queryset = queryset.order_by(order_field)
        else:
            queryset = queryset.order_by('code')

        # العد الإجمالي
        total_records = Account.objects.filter(company=request.current_company).count()
        filtered_records = queryset.count()

        # الصفحات
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_edit = request.user.has_perm('accounting.change_account')
        can_delete = request.user.has_perm('accounting.delete_account')

        for account in queryset:
            # حالة الحساب
            if account.is_suspended:
                status_badge = '<span class="badge bg-danger">موقوف</span>'
            else:
                status_badge = '<span class="badge bg-success">نشط</span>'

            # الحساب الأب
            parent_name = account.parent.name if account.parent else '-'

            # عدد الحسابات الفرعية
            children_info = ''
            if account.children_count > 0:
                children_info = f' <span class="badge bg-info">{account.children_count} فرعي</span>'

            # أزرار الإجراءات
            actions = []
            if can_edit:
                actions.append(f'''
                    <a href="{reverse('accounting:account_update', args=[account.pk])}" 
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            if can_delete and account.children_count == 0:
                actions.append(f'''
                    <button type="button" class="btn btn-outline-danger btn-sm" 
                            onclick="deleteAccount({account.pk})" title="حذف" data-bs-toggle="tooltip">
                        <i class="fas fa-trash"></i>
                    </button>
                ''')

            actions_html = ' '.join(actions) if actions else '-'

            data.append([
                account.code,
                f"{account.name}{children_info}",
                account.account_type.name,
                parent_name,
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        # في حالة حدوث خطأ
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        })


@login_required
@permission_required_with_message('accounting.view_account')
def export_accounts(request):
    """تصدير الحسابات إلى Excel"""

    # إنشاء workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "دليل الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # إضافة الرأس
    headers = [
        'رمز الحساب', 'اسم الحساب', 'الاسم الإنجليزي', 'نوع الحساب',
        'الحساب الأب', 'العملة', 'طبيعة الحساب', 'الرصيد الافتتاحي', 'الحالة'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # إضافة البيانات
    queryset = Account.objects.filter(
        company=request.current_company
    ).select_related('account_type', 'parent', 'currency').order_by('code')

    for row, account in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=account.code)
        ws.cell(row=row, column=2, value=account.name)
        ws.cell(row=row, column=3, value=account.name_en or '')
        ws.cell(row=row, column=4, value=account.account_type.name)
        ws.cell(row=row, column=5, value=account.parent.name if account.parent else '')
        ws.cell(row=row, column=6, value=account.currency.name)
        ws.cell(row=row, column=7, value=account.get_nature_display())
        ws.cell(row=row, column=8, value=float(account.opening_balance))
        ws.cell(row=row, column=9, value='موقوف' if account.is_suspended else 'نشط')

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="chart_of_accounts.xlsx"'

    wb.save(response)
    return response


@login_required
@permission_required_with_message('accounting.add_account')
def import_accounts(request):
    """استيراد الحسابات من Excel/CSV"""

    if request.method == 'POST':
        form = AccountImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = form.cleaned_data['file']
                update_existing = form.cleaned_data.get('update_existing', False)

                # قراءة الملف
                if file.name.lower().endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # التحقق من الأعمدة المطلوبة
                required_columns = ['code', 'name', 'account_type_code', 'currency_code']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, _('الملف لا يحتوي على الأعمدة المطلوبة'))
                    return redirect('accounting:account_list')

                # استيراد البيانات
                created_count = 0
                updated_count = 0
                errors = []

                for index, row in df.iterrows():
                    try:
                        # البحث عن نوع الحساب
                        try:
                            account_type = AccountType.objects.get(code=row['account_type_code'])
                        except AccountType.DoesNotExist:
                            errors.append(f'الصف {index + 2}: نوع الحساب غير موجود')
                            continue

                        # البحث عن العملة
                        try:
                            currency = Currency.objects.get(
                                code=row['currency_code'],
                                company=request.current_company
                            )
                        except Currency.DoesNotExist:
                            errors.append(f'الصف {index + 2}: العملة غير موجودة')
                            continue

                        # البحث عن الحساب الأب (إذا وُجد)
                        parent = None
                        if 'parent_code' in row and pd.notna(row['parent_code']):
                            try:
                                parent = Account.objects.get(
                                    code=row['parent_code'],
                                    company=request.current_company
                                )
                            except Account.DoesNotExist:
                                errors.append(f'الصف {index + 2}: الحساب الأب غير موجود')
                                continue

                        # إنشاء أو تحديث الحساب
                        account_data = {
                            'name': row['name'],
                            'name_en': row.get('name_en', ''),
                            'account_type': account_type,
                            'parent': parent,
                            'currency': currency,
                            'company': request.current_company,
                            'created_by': request.user
                        }

                        account, created = Account.objects.get_or_create(
                            code=row['code'],
                            company=request.current_company,
                            defaults=account_data
                        )

                        if created:
                            created_count += 1
                        elif update_existing:
                            for key, value in account_data.items():
                                if key not in ['company', 'created_by']:
                                    setattr(account, key, value)
                            account.save()
                            updated_count += 1

                    except Exception as e:
                        errors.append(f'الصف {index + 2}: {str(e)}')

                # رسائل النتائج
                if created_count > 0:
                    messages.success(request, f'تم إنشاء {created_count} حساب جديد')
                if updated_count > 0:
                    messages.success(request, f'تم تحديث {updated_count} حساب')
                if errors:
                    for error in errors[:5]:  # إظهار أول 5 أخطاء فقط
                        messages.error(request, error)

            except Exception as e:
                messages.error(request, f'خطأ في معالجة الملف: {str(e)}')

    return redirect('accounting:account_list')