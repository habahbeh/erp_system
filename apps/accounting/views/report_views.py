# apps/accounting/views/report_views.py
"""
Report Views - التقارير المحاسبية والتصدير والاستيراد
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.generic import TemplateView
from django.db.models import Q, Sum, Case, When, DecimalField, F
from django.utils.translation import gettext_lazy as _
from django.urls import reverse
from django.db import transaction
from datetime import date, datetime
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.core.decorators import permission_required_with_message, company_required
from apps.core.mixins import CompanyMixin
from apps.core.models import Currency
from ..models import Account, AccountType, JournalEntry, JournalEntryLine, AccountBalance
from ..forms.account_forms import AccountImportForm
from dateutil.relativedelta import relativedelta


# ========== التقارير المحاسبية ==========

class GeneralLedgerView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """دفتر الأستاذ - عرض حركات حساب معين"""

    template_name = 'accounting/reports/general_ledger.html'
    permission_required = 'accounting.view_journalentry'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على معاملات البحث
        account_id = self.request.GET.get('account_id')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        # تحديد فترة افتراضية (بداية العام الحالي حتى اليوم)
        if not date_from:
            date_from = date.today().replace(month=1, day=1).strftime('%Y-%m-%d')
        if not date_to:
            date_to = date.today().strftime('%Y-%m-%d')

        # البيانات الأساسية
        account = None
        ledger_entries = []
        summary = {
            'opening_balance': Decimal('0.00'),
            'total_debit': Decimal('0.00'),
            'total_credit': Decimal('0.00'),
            'closing_balance': Decimal('0.00')
        }

        # إذا تم اختيار حساب
        if account_id:
            try:
                account = Account.objects.get(
                    id=account_id,
                    company=self.request.current_company
                )

                # الحصول على حركات الحساب
                ledger_entries = self.get_ledger_entries(account, date_from, date_to)

                # حساب الملخص
                summary = self.calculate_summary(account, ledger_entries, date_from)

            except Account.DoesNotExist:
                account = None

        # قائمة الحسابات للاختيار
        accounts = Account.objects.filter(
            company=self.request.current_company,
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        context.update({
            'title': _('دفتر الأستاذ'),
            'account': account,
            'accounts': accounts,
            'ledger_entries': ledger_entries,
            'summary': summary,
            'date_from': date_from,
            'date_to': date_to,
            'account_id': account_id,
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('دفتر الأستاذ'), 'url': ''},
            ]
        })

        return context

    def get_ledger_entries(self, account, date_from, date_to):
        """الحصول على حركات الحساب"""

        # الحصول على خطوط القيود
        lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=self.request.current_company,
            journal_entry__status='posted',
            journal_entry__entry_date__range=[date_from, date_to]
        ).select_related(
            'journal_entry',
            'journal_entry__created_by'
        ).order_by('journal_entry__entry_date', 'journal_entry__number')

        # تجميع البيانات
        entries = []
        running_balance = account.opening_balance or Decimal('0.00')

        for line in lines:
            # التأكد من عدم وجود قيم None
            debit_amount = line.debit_amount or Decimal('0.00')
            credit_amount = line.credit_amount or Decimal('0.00')

            # حساب الرصيد التراكمي
            if account.account_type.normal_balance == 'debit':
                # حسابات مدينة الطبيعة
                running_balance += debit_amount - credit_amount
            else:
                # حسابات دائنة الطبيعة
                running_balance += credit_amount - debit_amount

            entries.append({
                'entry_number': line.journal_entry.number,
                'entry_date': line.journal_entry.entry_date,
                'description': line.description,
                'reference': line.reference or line.journal_entry.reference,
                'debit_amount': debit_amount,
                'credit_amount': credit_amount,
                'balance': abs(running_balance),
                'balance_type': 'مدين' if running_balance >= 0 else 'دائن',
                'journal_entry_id': line.journal_entry.id,
                'created_by': line.journal_entry.created_by.get_full_name() or line.journal_entry.created_by.username
            })

        return entries

    def calculate_summary(self, account, entries, date_from):
        """حساب ملخص الحساب"""

        # الرصيد الافتتاحي
        opening_balance = account.opening_balance or Decimal('0.00')

        # مجاميع الحركة - التأكد من عدم وجود قيم None
        total_debit = sum(entry['debit_amount'] or Decimal('0.00') for entry in entries)
        total_credit = sum(entry['credit_amount'] or Decimal('0.00') for entry in entries)

        # الرصيد الختامي
        if account.account_type.normal_balance == 'debit':
            closing_balance = opening_balance + total_debit - total_credit
        else:
            closing_balance = opening_balance + total_credit - total_debit

        return {
            'opening_balance': opening_balance,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'closing_balance': closing_balance,
            'balance_type': 'مدين' if closing_balance >= 0 else 'دائن'
        }


class TrialBalanceView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """ميزان المراجعة"""

    template_name = 'accounting/reports/trial_balance.html'
    permission_required = 'accounting.view_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # معاملات التقرير
        as_of_date = self.request.GET.get('as_of_date', date.today().strftime('%Y-%m-%d'))
        show_zero_balances = self.request.GET.get('show_zero_balances') == 'on'
        account_level = int(self.request.GET.get('account_level', 4))

        # الحصول على بيانات ميزان المراجعة
        trial_balance_data = self.get_trial_balance_data(as_of_date, show_zero_balances, account_level)

        context.update({
            'title': _('ميزان المراجعة'),
            'trial_balance_data': trial_balance_data,
            'as_of_date': as_of_date,
            'show_zero_balances': show_zero_balances,
            'account_level': account_level,
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('ميزان المراجعة'), 'url': ''},
            ]
        })

        return context

    def get_trial_balance_data(self, as_of_date, show_zero_balances, account_level):
        """الحصول على بيانات ميزان المراجعة"""

        # الحصول على الحسابات حسب المستوى المطلوب
        accounts = Account.objects.filter(
            company=self.request.current_company,
            level__lte=account_level,
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        trial_balance = []
        total_debit = Decimal('0.00')
        total_credit = Decimal('0.00')

        for account in accounts:
            # حساب رصيد الحساب حتى التاريخ المحدد
            balance_data = self.calculate_account_balance(account, as_of_date)

            # تخطي الحسابات ذات الرصيد صفر إذا لم يطلب عرضها
            if not show_zero_balances and balance_data['debit_balance'] == 0 and balance_data['credit_balance'] == 0:
                continue

            trial_balance.append({
                'account': account,
                'debit_balance': balance_data['debit_balance'],
                'credit_balance': balance_data['credit_balance']
            })

            total_debit += balance_data['debit_balance']
            total_credit += balance_data['credit_balance']

        # حساب الفرق
        difference = total_debit - total_credit

        return {
            'accounts': trial_balance,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'difference': difference,  # إضافة الفرق
            'is_balanced': total_debit == total_credit
        }

    def calculate_account_balance(self, account, as_of_date):
        """حساب رصيد حساب حتى تاريخ معين"""

        # الرصيد الافتتاحي
        opening_balance = account.opening_balance or Decimal('0.00')

        # حساب حركات الحساب حتى التاريخ المحدد
        movements = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=self.request.current_company,
            journal_entry__status='posted',
            journal_entry__entry_date__lte=as_of_date
        ).aggregate(
            total_debit=Sum('debit_amount'),
            total_credit=Sum('credit_amount')
        )

        # التأكد من عدم وجود قيم None
        total_debit = movements['total_debit'] or Decimal('0.00')
        total_credit = movements['total_credit'] or Decimal('0.00')

        # حساب الرصيد النهائي حسب طبيعة الحساب
        if account.account_type.normal_balance == 'debit':
            # حساب مدين الطبيعة
            net_balance = opening_balance + total_debit - total_credit
            if net_balance >= 0:
                return {'debit_balance': net_balance, 'credit_balance': Decimal('0.00')}
            else:
                return {'debit_balance': Decimal('0.00'), 'credit_balance': abs(net_balance)}
        else:
            # حساب دائن الطبيعة
            net_balance = opening_balance + total_credit - total_debit
            if net_balance >= 0:
                return {'debit_balance': Decimal('0.00'), 'credit_balance': net_balance}
            else:
                return {'debit_balance': abs(net_balance), 'credit_balance': Decimal('0.00')}


@login_required
@permission_required('accounting.view_account')
def account_search_for_reports(request):
    """البحث عن الحسابات للتقارير"""

    query = request.GET.get('term', '')
    if len(query) < 2:
        return JsonResponse([])

    accounts = Account.objects.filter(
        company=request.current_company,
        accept_entries=True,
        is_suspended=False
    ).filter(
        Q(code__icontains=query) | Q(name__icontains=query)
    )[:20]

    results = []
    for account in accounts:
        results.append({
            'id': account.id,
            'text': f"{account.code} - {account.name}",
            'code': account.code,
            'name': account.name
        })

    return JsonResponse(results, safe=False)


# ========== التصدير والاستيراد (الكود القديم محفوظ) ==========

@login_required
@permission_required_with_message('accounting.view_accounttype')
def export_account_types(request):
    """تصدير أنواع الحسابات إلى Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "أنواع الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # إضافة الرأس
    headers = ['الرمز', 'الاسم', 'التصنيف', 'الرصيد الطبيعي']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # إضافة البيانات
    for row, account_type in enumerate(AccountType.objects.order_by('code'), 2):
        ws.cell(row=row, column=1, value=account_type.code).border = thin_border
        ws.cell(row=row, column=2, value=account_type.name).border = thin_border
        ws.cell(row=row, column=3, value=account_type.get_type_category_display()).border = thin_border
        ws.cell(row=row, column=4, value=account_type.get_normal_balance_display()).border = thin_border

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="account_types.xlsx"'

    wb.save(response)
    return response


@company_required
@login_required
@permission_required_with_message('accounting.view_account')
def export_accounts(request):
    """تصدير الحسابات إلى Excel"""

    company = request.current_company

    wb = Workbook()
    ws = wb.active
    ws.title = "دليل الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # إضافة الرأس
    headers = [
        'رمز الحساب', 'اسم الحساب', 'الاسم الإنجليزي', 'نوع الحساب',
        'الحساب الأب', 'العملة', 'طبيعة الحساب', 'الرصيد الافتتاحي',
        'المستوى', 'حساب نقدي', 'حساب بنكي', 'يقبل قيود', 'الحالة'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # إضافة البيانات
    queryset = Account.objects.filter(
        company=company
    ).select_related('account_type', 'parent', 'currency').order_by('code')

    for row, account in enumerate(queryset, 2):
        data = [
            account.code,
            account.name,
            account.name_en or '',
            account.account_type.name,
            account.parent.name if account.parent else '',
            account.currency.name,
            account.get_nature_display(),
            float(account.opening_balance),
            account.level,
            'نعم' if account.is_cash_account else 'لا',
            'نعم' if account.is_bank_account else 'لا',
            'نعم' if account.accept_entries else 'لا',
            'موقوف' if account.is_suspended else 'نشط'
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value).border = thin_border

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="accounts_{company.name}.xlsx"'

    wb.save(response)
    return response


@login_required
@permission_required_with_message('accounting.add_accounttype')
def import_account_types(request):
    """استيراد أنواع الحسابات"""

    if request.method == 'POST':
        # منطق الاستيراد هنا
        pass

    return redirect('accounting:account_type_list')


@company_required
@login_required
@permission_required_with_message('accounting.add_account')
def import_accounts(request):
    """استيراد الحسابات"""

    if request.method == 'POST':
        # منطق الاستيراد هنا
        pass

    return redirect('accounting:account_list')


class AccountStatementView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """كشف حساب العميل/المورد"""

    template_name = 'accounting/reports/account_statement.html'
    permission_required = 'accounting.view_journalentry'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الحصول على معاملات البحث
        account_id = self.request.GET.get('account_id')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        statement_type = self.request.GET.get('statement_type', 'all')  # all, customer, supplier

        # تحديد فترة افتراضية
        if not date_from:
            date_from = date.today().replace(month=1, day=1).strftime('%Y-%m-%d')
        if not date_to:
            date_to = date.today().strftime('%Y-%m-%d')

        # البيانات الأساسية
        account = None
        statement_entries = []
        summary = {
            'opening_balance': Decimal('0.00'),
            'total_debit': Decimal('0.00'),
            'total_credit': Decimal('0.00'),
            'closing_balance': Decimal('0.00'),
            'aging_analysis': {}
        }

        # إذا تم اختيار حساب
        if account_id:
            try:
                account = Account.objects.get(
                    id=account_id,
                    company=self.request.current_company
                )

                # الحصول على حركات الحساب
                statement_entries = self.get_statement_entries(account, date_from, date_to)

                # حساب الملخص
                summary = self.calculate_statement_summary(account, statement_entries, date_from, date_to)

            except Account.DoesNotExist:
                account = None

        # قائمة الحسابات للاختيار حسب النوع
        accounts_filter = Q(
            company=self.request.current_company,
            accept_entries=True,
            is_suspended=False
        )

        if statement_type == 'customer':
            # حسابات العملاء - عادة من نوع المدينون أو العملاء
            accounts_filter &= Q(
                Q(account_type__name__icontains='عميل') |
                Q(account_type__name__icontains='مدين') |
                Q(account_type__type_category='assets')
            )
        elif statement_type == 'supplier':
            # حسابات الموردين - عادة من نوع الدائنون أو الموردين
            accounts_filter &= Q(
                Q(account_type__name__icontains='مورد') |
                Q(account_type__name__icontains='دائن') |
                Q(account_type__type_category='liabilities')
            )

        accounts = Account.objects.filter(accounts_filter).order_by('code')

        context.update({
            'title': _('كشف حساب العميل/المورد'),
            'account': account,
            'accounts': accounts,
            'statement_entries': statement_entries,
            'summary': summary,
            'date_from': date_from,
            'date_to': date_to,
            'account_id': account_id,
            'statement_type': statement_type,
            'statement_types': [
                ('all', 'جميع الحسابات'),
                ('customer', 'العملاء فقط'),
                ('supplier', 'الموردين فقط')
            ],
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('كشف الحساب'), 'url': ''},
            ]
        })

        return context

    def get_statement_entries(self, account, date_from, date_to):
        """الحصول على حركات كشف الحساب"""

        # الحصول على خطوط القيود مع تفاصيل أكثر
        lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=self.request.current_company,
            journal_entry__status='posted',
            journal_entry__entry_date__range=[date_from, date_to]
        ).select_related(
            'journal_entry',
            'journal_entry__created_by'
        ).order_by('journal_entry__entry_date', 'journal_entry__number')

        # تجميع البيانات مع معلومات إضافية
        entries = []
        running_balance = account.opening_balance or Decimal('0.00')

        for line in lines:
            # التأكد من عدم وجود قيم None
            debit_amount = line.debit_amount or Decimal('0.00')
            credit_amount = line.credit_amount or Decimal('0.00')

            # حساب الرصيد التراكمي
            if account.account_type.normal_balance == 'debit':
                running_balance += debit_amount - credit_amount
            else:
                running_balance += credit_amount - debit_amount

            # تحديد نوع العملية
            transaction_type = self.determine_transaction_type(line, account)

            entries.append({
                'entry_number': line.journal_entry.number,
                'entry_date': line.journal_entry.entry_date,
                'description': line.description,
                'reference': line.reference or line.journal_entry.reference,
                'debit_amount': debit_amount,
                'credit_amount': credit_amount,
                'balance': abs(running_balance),
                'balance_type': 'مدين' if running_balance >= 0 else 'دائن',
                'journal_entry_id': line.journal_entry.id,
                'created_by': line.journal_entry.created_by.get_full_name() or line.journal_entry.created_by.username,
                'transaction_type': transaction_type,
                'days_since': (date.today() - line.journal_entry.entry_date).days
            })

        return entries

    def determine_transaction_type(self, line, account):
        """تحديد نوع العملية"""

        debit_amount = line.debit_amount or Decimal('0.00')
        credit_amount = line.credit_amount or Decimal('0.00')

        # للعملاء (حسابات مدينة الطبيعة)
        if account.account_type.normal_balance == 'debit':
            if debit_amount > 0:
                return 'مبيعات/فاتورة'
            else:
                return 'تحصيل/دفعة'
        # للموردين (حسابات دائنة الطبيعة)
        else:
            if credit_amount > 0:
                return 'مشتريات/فاتورة'
            else:
                return 'سداد/دفعة'

    def calculate_statement_summary(self, account, entries, date_from, date_to):
        """حساب ملخص كشف الحساب"""

        # الرصيد الافتتاحي
        opening_balance = account.opening_balance or Decimal('0.00')

        # مجاميع الحركة
        total_debit = sum(entry['debit_amount'] or Decimal('0.00') for entry in entries)
        total_credit = sum(entry['credit_amount'] or Decimal('0.00') for entry in entries)

        # الرصيد الختامي
        if account.account_type.normal_balance == 'debit':
            closing_balance = opening_balance + total_debit - total_credit
        else:
            closing_balance = opening_balance + total_credit - total_debit

        # تحليل الأعمار (Aging Analysis)
        aging_analysis = self.calculate_aging_analysis(entries, account)

        return {
            'opening_balance': opening_balance,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'closing_balance': closing_balance,
            'balance_type': 'مدين' if closing_balance >= 0 else 'دائن',
            'aging_analysis': aging_analysis,
            'transaction_count': len(entries),
            'period_days': (datetime.strptime(date_to, '%Y-%m-%d').date() -
                            datetime.strptime(date_from, '%Y-%m-%d').date()).days + 1
        }

    def calculate_aging_analysis(self, entries, account):
        """حساب تحليل أعمار الذمم"""

        aging = {
            'current': Decimal('0.00'),  # 0-30 يوم
            'period_1': Decimal('0.00'),  # 31-60 يوم
            'period_2': Decimal('0.00'),  # 61-90 يوم
            'period_3': Decimal('0.00'),  # 91-120 يوم
            'over_120': Decimal('0.00')  # أكثر من 120 يوم
        }

        for entry in entries:
            days = entry['days_since']
            amount = entry['debit_amount'] - entry['credit_amount']

            # للعملاء نحسب المبلغ المدين، للموردين نحسب المبلغ الدائن
            if account.account_type.normal_balance == 'credit':
                amount = entry['credit_amount'] - entry['debit_amount']

            if amount > 0:  # فقط الأرصدة المدينة للعملاء أو الدائنة للموردين
                if days <= 30:
                    aging['current'] += amount
                elif days <= 60:
                    aging['period_1'] += amount
                elif days <= 90:
                    aging['period_2'] += amount
                elif days <= 120:
                    aging['period_3'] += amount
                else:
                    aging['over_120'] += amount

        return aging


# إضافة View للمقارنة بين الحسابات
class AccountComparisonView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """مقارنة بين حسابات متعددة"""

    template_name = 'accounting/reports/account_comparison.html'
    permission_required = 'accounting.view_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # سنطورها لاحقاً
        context.update({
            'title': _('مقارنة الحسابات'),
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('مقارنة الحسابات'), 'url': ''},
            ]
        })

        return context


class IncomeStatementView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """قائمة الدخل - بيان الأرباح والخسائر"""

    template_name = 'accounting/reports/income_statement.html'
    permission_required = 'accounting.view_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # معاملات التقرير
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        comparison_type = self.request.GET.get('comparison_type', 'none')  # none, previous_period, previous_year

        # تحديد فترة افتراضية (بداية العام الحالي حتى اليوم)
        if not date_from:
            date_from = date.today().replace(month=1, day=1).strftime('%Y-%m-%d')
        if not date_to:
            date_to = date.today().strftime('%Y-%m-%d')

        # الحصول على بيانات قائمة الدخل
        income_data = self.get_income_statement_data(date_from, date_to)

        # بيانات المقارنة إذا طُلبت
        comparison_data = None
        if comparison_type != 'none':
            comparison_date_from, comparison_date_to = self.get_comparison_dates(
                date_from, date_to, comparison_type
            )
            comparison_data = self.get_income_statement_data(comparison_date_from, comparison_date_to)

        context.update({
            'title': _('قائمة الدخل'),
            'income_data': income_data,
            'comparison_data': comparison_data,
            'date_from': date_from,
            'date_to': date_to,
            'comparison_type': comparison_type,
            'comparison_types': [
                ('none', 'بدون مقارنة'),
                ('previous_period', 'الفترة السابقة'),
                ('previous_year', 'نفس الفترة من العام السابق')
            ],
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('قائمة الدخل'), 'url': ''},
            ]
        })

        return context

    def get_income_statement_data(self, date_from, date_to):
        """الحصول على بيانات قائمة الدخل"""

        # الإيرادات (Revenue Accounts)
        revenue_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='revenue',
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        revenues = []
        total_revenue = Decimal('0.00')

        for account in revenue_accounts:
            balance = self.calculate_period_balance(account, date_from, date_to)
            if balance != 0:  # عرض الحسابات التي لها حركة فقط
                revenues.append({
                    'account': account,
                    'balance': abs(balance)  # الإيرادات دائماً موجبة في التقرير
                })
                total_revenue += abs(balance)

        # تكلفة المبيعات (Cost of Goods Sold)
        cogs_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='expenses',
            account_type__name__icontains='تكلفة'
        ).order_by('code')

        cost_of_sales = []
        total_cogs = Decimal('0.00')

        for account in cogs_accounts:
            balance = self.calculate_period_balance(account, date_from, date_to)
            if balance != 0:
                cost_of_sales.append({
                    'account': account,
                    'balance': abs(balance)
                })
                total_cogs += abs(balance)

        # إجمالي الربح
        gross_profit = total_revenue - total_cogs

        # المصروفات التشغيلية (Operating Expenses)
        operating_expense_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='expenses',
            accept_entries=True,
            is_suspended=False
        ).exclude(
            account_type__name__icontains='تكلفة'
        ).order_by('code')

        operating_expenses = []
        total_operating_expenses = Decimal('0.00')

        for account in operating_expense_accounts:
            balance = self.calculate_period_balance(account, date_from, date_to)
            if balance != 0:
                operating_expenses.append({
                    'account': account,
                    'balance': abs(balance)
                })
                total_operating_expenses += abs(balance)

        # صافي الدخل
        net_income = gross_profit - total_operating_expenses

        return {
            'revenues': revenues,
            'total_revenue': total_revenue,
            'cost_of_sales': cost_of_sales,
            'total_cogs': total_cogs,
            'gross_profit': gross_profit,
            'gross_profit_margin': (gross_profit / total_revenue * 100) if total_revenue > 0 else 0,
            'operating_expenses': operating_expenses,
            'total_operating_expenses': total_operating_expenses,
            'net_income': net_income,
            'net_profit_margin': (net_income / total_revenue * 100) if total_revenue > 0 else 0
        }

    def calculate_period_balance(self, account, date_from, date_to):
        """حساب رصيد الحساب خلال فترة معينة"""

        movements = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=self.request.current_company,
            journal_entry__status='posted',
            journal_entry__entry_date__range=[date_from, date_to]
        ).aggregate(
            total_debit=Sum('debit_amount'),
            total_credit=Sum('credit_amount')
        )

        # التأكد من عدم وجود قيم None
        total_debit = movements['total_debit'] or Decimal('0.00')
        total_credit = movements['total_credit'] or Decimal('0.00')

        # للإيرادات والمصروفات نحسب صافي الحركة فقط (بدون رصيد افتتاحي)
        if account.account_type.normal_balance == 'credit':
            # حسابات الإيرادات (دائنة الطبيعة)
            return total_credit - total_debit
        else:
            # حسابات المصروفات (مدينة الطبيعة)
            return total_debit - total_credit

    def get_comparison_dates(self, date_from, date_to, comparison_type):
        """حساب تواريخ المقارنة"""
        from datetime import datetime, timedelta
        from dateutil.relativedelta import relativedelta

        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()

        if comparison_type == 'previous_period':
            # الفترة السابقة بنفس الطول
            period_length = (end_date - start_date).days + 1
            comparison_end = start_date - timedelta(days=1)
            comparison_start = comparison_end - timedelta(days=period_length - 1)
        elif comparison_type == 'previous_year':
            # نفس الفترة من العام السابق
            comparison_start = start_date - relativedelta(years=1)
            comparison_end = end_date - relativedelta(years=1)

        return comparison_start.strftime('%Y-%m-%d'), comparison_end.strftime('%Y-%m-%d')


class BalanceSheetView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """الميزانية العمومية - قائمة المركز المالي"""

    template_name = 'accounting/reports/balance_sheet.html'
    permission_required = 'accounting.view_account'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # معاملات التقرير
        as_of_date = self.request.GET.get('as_of_date', date.today().strftime('%Y-%m-%d'))
        comparison_type = self.request.GET.get('comparison_type', 'none')

        # الحصول على بيانات الميزانية
        balance_sheet_data = self.get_balance_sheet_data(as_of_date)

        # بيانات المقارنة إذا طُلبت
        comparison_data = None
        if comparison_type != 'none':
            if comparison_type == 'previous_month':
                comparison_date = (datetime.strptime(as_of_date, '%Y-%m-%d').date()
                                   - relativedelta(months=1)).strftime('%Y-%m-%d')
            elif comparison_type == 'previous_year':
                comparison_date = (datetime.strptime(as_of_date, '%Y-%m-%d').date()
                                   - relativedelta(years=1)).strftime('%Y-%m-%d')

            comparison_data = self.get_balance_sheet_data(comparison_date)

        context.update({
            'title': _('الميزانية العمومية'),
            'balance_sheet_data': balance_sheet_data,
            'comparison_data': comparison_data,
            'as_of_date': as_of_date,
            'comparison_type': comparison_type,
            'comparison_types': [
                ('none', 'بدون مقارنة'),
                ('previous_month', 'الشهر السابق'),
                ('previous_year', 'نفس التاريخ من العام السابق')
            ],
            'breadcrumbs': [
                {'title': _('المحاسبة'), 'url': reverse('accounting:dashboard')},
                {'title': _('التقارير'), 'url': '#'},
                {'title': _('الميزانية العمومية'), 'url': ''},
            ]
        })

        return context

    def get_balance_sheet_data(self, as_of_date):
        """الحصول على بيانات الميزانية العمومية"""

        # الأصول (Assets)
        asset_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='assets',
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        # تصنيف الأصول
        current_assets = []
        fixed_assets = []
        total_current_assets = Decimal('0.00')
        total_fixed_assets = Decimal('0.00')

        for account in asset_accounts:
            balance = self.calculate_account_balance_as_of(account, as_of_date)
            if abs(balance) > 0.01:  # عرض الحسابات التي لها رصيد فقط
                asset_item = {
                    'account': account,
                    'balance': abs(balance)
                }

                # تصنيف الأصول حسب الاسم أو الكود
                if any(keyword in account.name.lower() for keyword in ['صندوق', 'بنك', 'نقد', 'عميل', 'مدين']):
                    current_assets.append(asset_item)
                    total_current_assets += abs(balance)
                else:
                    fixed_assets.append(asset_item)
                    total_fixed_assets += abs(balance)

        total_assets = total_current_assets + total_fixed_assets

        # الخصوم (Liabilities)
        liability_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='liabilities',
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        current_liabilities = []
        long_term_liabilities = []
        total_current_liabilities = Decimal('0.00')
        total_long_term_liabilities = Decimal('0.00')

        for account in liability_accounts:
            balance = self.calculate_account_balance_as_of(account, as_of_date)
            if abs(balance) > 0.01:
                liability_item = {
                    'account': account,
                    'balance': abs(balance)
                }

                # تصنيف الخصوم
                if any(keyword in account.name.lower() for keyword in ['مورد', 'دائن', 'مستحق']):
                    current_liabilities.append(liability_item)
                    total_current_liabilities += abs(balance)
                else:
                    long_term_liabilities.append(liability_item)
                    total_long_term_liabilities += abs(balance)

        total_liabilities = total_current_liabilities + total_long_term_liabilities

        # حقوق الملكية (Equity)
        equity_accounts = Account.objects.filter(
            company=self.request.current_company,
            account_type__type_category='equity',
            accept_entries=True,
            is_suspended=False
        ).order_by('code')

        equity = []
        total_equity = Decimal('0.00')

        for account in equity_accounts:
            balance = self.calculate_account_balance_as_of(account, as_of_date)
            if abs(balance) > 0.01:
                equity.append({
                    'account': account,
                    'balance': abs(balance)
                })
                total_equity += abs(balance)

        # إجمالي الخصوم وحقوق الملكية
        total_liabilities_equity = total_liabilities + total_equity

        # التحقق من توازن الميزانية
        is_balanced = abs(total_assets - total_liabilities_equity) < 0.01

        # حساب النسب المئوية
        current_assets_percentage = 0
        fixed_assets_percentage = 0

        if total_assets > 0:
            current_assets_percentage = (total_current_assets / total_assets) * 100
            fixed_assets_percentage = (total_fixed_assets / total_assets) * 100

        return {
            'current_assets': current_assets,
            'fixed_assets': fixed_assets,
            'total_current_assets': total_current_assets,
            'total_fixed_assets': total_fixed_assets,
            'total_assets': total_assets,
            'current_liabilities': current_liabilities,
            'long_term_liabilities': long_term_liabilities,
            'total_current_liabilities': total_current_liabilities,
            'total_long_term_liabilities': total_long_term_liabilities,
            'total_liabilities': total_liabilities,
            'equity': equity,
            'total_equity': total_equity,
            'total_liabilities_equity': total_liabilities_equity,
            'is_balanced': is_balanced,
            'current_ratio': (total_current_assets / total_current_liabilities) if total_current_liabilities > 0 else 0,
            'debt_to_equity_ratio': (total_liabilities / total_equity) if total_equity > 0 else 0,
            # إضافة النسب المئوية المحسوبة
            'current_assets_percentage': current_assets_percentage,
            'fixed_assets_percentage': fixed_assets_percentage
        }

    def calculate_account_balance_as_of(self, account, as_of_date):
        """حساب رصيد الحساب حتى تاريخ معين (تراكمي)"""

        # الرصيد الافتتاحي
        opening_balance = account.opening_balance or Decimal('0.00')

        # حساب حركات الحساب حتى التاريخ المحدد
        movements = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__company=self.request.current_company,
            journal_entry__status='posted',
            journal_entry__entry_date__lte=as_of_date
        ).aggregate(
            total_debit=Sum('debit_amount'),
            total_credit=Sum('credit_amount')
        )

        # التأكد من عدم وجود قيم None
        total_debit = movements['total_debit'] or Decimal('0.00')
        total_credit = movements['total_credit'] or Decimal('0.00')

        # حساب الرصيد النهائي حسب طبيعة الحساب
        if account.account_type.normal_balance == 'debit':
            return opening_balance + total_debit - total_credit
        else:
            return opening_balance + total_credit - total_debit

# ========== تصدير التقارير الجديدة ==========

@login_required
@permission_required('accounting.view_journalentry')
def export_general_ledger(request):
    """تصدير دفتر الأستاذ إلى Excel"""

    account_id = request.GET.get('account_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not account_id:
        messages.error(request, 'يجب تحديد الحساب أولاً')
        return redirect('accounting:general_ledger')

    try:
        account = Account.objects.get(id=account_id, company=request.current_company)
    except Account.DoesNotExist:
        messages.error(request, 'الحساب غير موجود')
        return redirect('accounting:general_ledger')

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "دفتر الأستاذ"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # معلومات الحساب
    ws.cell(row=1, column=1, value=f"دفتر الأستاذ - {account.name}")
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment

    ws.cell(row=2, column=1, value=f"رمز الحساب: {account.code}")
    ws.cell(row=2, column=4, value=f"الفترة: {date_from} إلى {date_to}")

    # عناوين الأعمدة
    headers = ['التاريخ', 'رقم القيد', 'البيان', 'المرجع', 'مدين', 'دائن', 'الرصيد', 'نوع الرصيد']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # البيانات
    view = GeneralLedgerView()
    view.request = request
    ledger_entries = view.get_ledger_entries(account, date_from, date_to)

    for row, entry in enumerate(ledger_entries, 5):
        ws.cell(row=row, column=1, value=entry['entry_date'].strftime('%Y/%m/%d'))
        ws.cell(row=row, column=2, value=entry['entry_number'])
        ws.cell(row=row, column=3, value=entry['description'])
        ws.cell(row=row, column=4, value=entry['reference'])
        ws.cell(row=row, column=5, value=float(entry['debit_amount']) if entry['debit_amount'] else None)
        ws.cell(row=row, column=6, value=float(entry['credit_amount']) if entry['credit_amount'] else None)
        ws.cell(row=row, column=7, value=float(entry['balance']))
        ws.cell(row=row, column=8, value=entry['balance_type'])

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"general_ledger_{account.code}_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@permission_required('accounting.view_account')
def export_trial_balance(request):
    """تصدير ميزان المراجعة إلى Excel"""

    as_of_date = request.GET.get('as_of_date')
    show_zero_balances = request.GET.get('show_zero_balances') == 'on'
    account_level = int(request.GET.get('account_level', 4))

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "ميزان المراجعة"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # العنوان
    ws.cell(row=1, column=1, value=f"ميزان المراجعة - {request.current_company.name}")
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment

    ws.cell(row=2, column=1, value=f"حتى تاريخ: {as_of_date}")
    ws.cell(row=2, column=3, value=f"مستوى الحسابات: {account_level}")

    # عناوين الأعمدة
    headers = ['رمز الحساب', 'اسم الحساب', 'رصيد مدين', 'رصيد دائن']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # البيانات
    view = TrialBalanceView()
    view.request = request
    trial_balance_data = view.get_trial_balance_data(as_of_date, show_zero_balances, account_level)

    row = 5
    for item in trial_balance_data['accounts']:
        ws.cell(row=row, column=1, value=item['account'].code)
        ws.cell(row=row, column=2, value=item['account'].name)
        ws.cell(row=row, column=3, value=float(item['debit_balance']) if item['debit_balance'] else None)
        ws.cell(row=row, column=4, value=float(item['credit_balance']) if item['credit_balance'] else None)
        row += 1

    # الإجمالي
    ws.cell(row=row, column=1, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=row, column=2, value="").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(trial_balance_data['total_debit'])).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(trial_balance_data['total_credit'])).font = Font(bold=True)

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"trial_balance_{as_of_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@permission_required('accounting.view_journalentry')
def export_account_statement(request):
    """تصدير كشف الحساب إلى Excel"""

    account_id = request.GET.get('account_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    statement_type = request.GET.get('statement_type', 'all')

    if not account_id:
        messages.error(request, 'يجب تحديد الحساب أولاً')
        return redirect('accounting:account_statement')

    try:
        account = Account.objects.get(id=account_id, company=request.current_company)
    except Account.DoesNotExist:
        messages.error(request, 'الحساب غير موجود')
        return redirect('accounting:account_statement')

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحساب"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6f42c1", end_color="6f42c1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # معلومات الحساب
    ws.cell(row=1, column=1, value=f"كشف حساب - {account.name}")
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment

    ws.cell(row=2, column=1, value=f"رمز الحساب: {account.code}")
    ws.cell(row=2, column=4, value=f"نوع الحساب: {account.account_type.name}")
    ws.cell(row=2, column=7, value=f"الفترة: {date_from} إلى {date_to}")

    # عناوين الأعمدة
    headers = ['التاريخ', 'رقم القيد', 'البيان', 'نوع العملية', 'المرجع', 'مدين', 'دائن', 'الرصيد', 'العمر بالأيام']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # البيانات
    view = AccountStatementView()
    view.request = request
    statement_entries = view.get_statement_entries(account, date_from, date_to)
    summary = view.calculate_statement_summary(account, statement_entries, date_from, date_to)

    for row, entry in enumerate(statement_entries, 5):
        ws.cell(row=row, column=1, value=entry['entry_date'].strftime('%Y/%m/%d'))
        ws.cell(row=row, column=2, value=entry['entry_number'])
        ws.cell(row=row, column=3, value=entry['description'])
        ws.cell(row=row, column=4, value=entry['transaction_type'])
        ws.cell(row=row, column=5, value=entry['reference'])
        ws.cell(row=row, column=6, value=float(entry['debit_amount']) if entry['debit_amount'] else None)
        ws.cell(row=row, column=7, value=float(entry['credit_amount']) if entry['credit_amount'] else None)
        ws.cell(row=row, column=8, value=f"{float(entry['balance'])} {entry['balance_type']}")
        ws.cell(row=row, column=9, value=entry['days_since'])

    # إضافة ملخص في النهاية
    last_row = len(statement_entries) + 6
    ws.cell(row=last_row, column=1, value="ملخص الحساب").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{last_row}:I{last_row}')

    ws.cell(row=last_row + 1, column=1, value="الرصيد الافتتاحي:")
    ws.cell(row=last_row + 1, column=2, value=float(summary['opening_balance']))

    ws.cell(row=last_row + 2, column=1, value="إجمالي المدين:")
    ws.cell(row=last_row + 2, column=2, value=float(summary['total_debit']))

    ws.cell(row=last_row + 3, column=1, value="إجمالي الدائن:")
    ws.cell(row=last_row + 3, column=2, value=float(summary['total_credit']))

    ws.cell(row=last_row + 4, column=1, value="الرصيد الختامي:")
    ws.cell(row=last_row + 4, column=2, value=f"{float(summary['closing_balance'])} {summary['balance_type']}")

    # تحليل الأعمار
    ws.cell(row=last_row + 6, column=1, value="تحليل أعمار الذمم:").font = Font(bold=True)
    ws.cell(row=last_row + 7, column=1, value="الحالي (0-30):")
    ws.cell(row=last_row + 7, column=2, value=float(summary['aging_analysis']['current']))
    ws.cell(row=last_row + 8, column=1, value="31-60 يوم:")
    ws.cell(row=last_row + 8, column=2, value=float(summary['aging_analysis']['period_1']))
    ws.cell(row=last_row + 9, column=1, value="61-90 يوم:")
    ws.cell(row=last_row + 9, column=2, value=float(summary['aging_analysis']['period_2']))
    ws.cell(row=last_row + 10, column=1, value="91-120 يوم:")
    ws.cell(row=last_row + 10, column=2, value=float(summary['aging_analysis']['period_3']))
    ws.cell(row=last_row + 11, column=1, value="أكثر من 120:")
    ws.cell(row=last_row + 11, column=2, value=float(summary['aging_analysis']['over_120']))

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"account_statement_{account.code}_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@permission_required('accounting.view_account')
def export_income_statement(request):
    """تصدير قائمة الدخل إلى Excel"""

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    comparison_type = request.GET.get('comparison_type', 'none')

    if not date_from or not date_to:
        messages.error(request, 'يجب تحديد الفترة الزمنية أولاً')
        return redirect('accounting:income_statement')

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة الدخل"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # العنوان
    ws.cell(row=1, column=1, value=f"قائمة الدخل - {request.current_company.name}")
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment

    ws.cell(row=2, column=1, value=f"للفترة من: {date_from} إلى {date_to}")
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = header_alignment

    # عناوين الأعمدة
    headers = ['البيان', 'المبلغ']
    if comparison_type != 'none':
        headers.extend(['فترة المقارنة', 'التغير %'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # البيانات
    view = IncomeStatementView()
    view.request = request
    income_data = view.get_income_statement_data(date_from, date_to)

    row = 5

    # الإيرادات
    ws.cell(row=row, column=1, value="الإيرادات").font = Font(bold=True)
    row += 1

    for item in income_data['revenues']:
        ws.cell(row=row, column=1, value=f"    {item['account'].name}")
        ws.cell(row=row, column=2, value=float(item['balance']))
        row += 1

    ws.cell(row=row, column=1, value="إجمالي الإيرادات").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(income_data['total_revenue'])).font = Font(bold=True)
    row += 2

    # تكلفة المبيعات
    if income_data['cost_of_sales']:
        ws.cell(row=row, column=1, value="تكلفة المبيعات").font = Font(bold=True)
        row += 1

        for item in income_data['cost_of_sales']:
            ws.cell(row=row, column=1, value=f"    {item['account'].name}")
            ws.cell(row=row, column=2, value=float(item['balance']) * -1)  # سالب
            row += 1

        ws.cell(row=row, column=1, value="إجمالي تكلفة المبيعات").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(income_data['total_cogs']) * -1).font = Font(bold=True)
        row += 1

        ws.cell(row=row, column=1, value="إجمالي الربح").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(income_data['gross_profit'])).font = Font(bold=True)
        row += 2

    # المصروفات التشغيلية
    ws.cell(row=row, column=1, value="المصروفات التشغيلية").font = Font(bold=True)
    row += 1

    for item in income_data['operating_expenses']:
        ws.cell(row=row, column=1, value=f"    {item['account'].name}")
        ws.cell(row=row, column=2, value=float(item['balance']) * -1)  # سالب
        row += 1

    ws.cell(row=row, column=1, value="إجمالي المصروفات التشغيلية").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(income_data['total_operating_expenses']) * -1).font = Font(bold=True)
    row += 2

    # صافي الدخل
    ws.cell(row=row, column=1, value="صافي الدخل").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=float(income_data['net_income'])).font = Font(bold=True, size=14)

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"income_statement_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@permission_required('accounting.view_account')
def export_balance_sheet(request):
    """تصدير الميزانية العمومية إلى Excel"""

    as_of_date = request.GET.get('as_of_date')
    comparison_type = request.GET.get('comparison_type', 'none')

    if not as_of_date:
        messages.error(request, 'يجب تحديد التاريخ أولاً')
        return redirect('accounting:balance_sheet')

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "الميزانية العمومية"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6c757d", end_color="6c757d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # العنوان
    ws.cell(row=1, column=1, value=f"الميزانية العمومية - {request.current_company.name}")
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment

    ws.cell(row=2, column=1, value=f"كما في: {as_of_date}")
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = header_alignment

    # عناوين الأعمدة
    headers = ['البيان', 'المبلغ']
    if comparison_type != 'none':
        headers.extend(['فترة المقارنة', 'التغير %'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # البيانات
    view = BalanceSheetView()
    view.request = request
    balance_data = view.get_balance_sheet_data(as_of_date)

    row = 5

    # الأصول
    ws.cell(row=row, column=1, value="الأصول").font = Font(bold=True)
    row += 1

    # الأصول الجارية
    if balance_data['current_assets']:
        ws.cell(row=row, column=1, value="الأصول الجارية").font = Font(bold=True)
        row += 1

        for item in balance_data['current_assets']:
            ws.cell(row=row, column=1, value=f"    {item['account'].name}")
            ws.cell(row=row, column=2, value=float(item['balance']))
            row += 1

        ws.cell(row=row, column=1, value="إجمالي الأصول الجارية").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(balance_data['total_current_assets'])).font = Font(bold=True)
        row += 1

    # الأصول الثابتة
    if balance_data['fixed_assets']:
        ws.cell(row=row, column=1, value="الأصول الثابتة").font = Font(bold=True)
        row += 1

        for item in balance_data['fixed_assets']:
            ws.cell(row=row, column=1, value=f"    {item['account'].name}")
            ws.cell(row=row, column=2, value=float(item['balance']))
            row += 1

        ws.cell(row=row, column=1, value="إجمالي الأصول الثابتة").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(balance_data['total_fixed_assets'])).font = Font(bold=True)
        row += 1

    # إجمالي الأصول
    ws.cell(row=row, column=1, value="إجمالي الأصول").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=float(balance_data['total_assets'])).font = Font(bold=True, size=14)
    row += 3

    # الخصوم
    ws.cell(row=row, column=1, value="الخصوم").font = Font(bold=True)
    row += 1

    # الخصوم الجارية
    if balance_data['current_liabilities']:
        ws.cell(row=row, column=1, value="الخصوم الجارية").font = Font(bold=True)
        row += 1

        for item in balance_data['current_liabilities']:
            ws.cell(row=row, column=1, value=f"    {item['account'].name}")
            ws.cell(row=row, column=2, value=float(item['balance']))
            row += 1

        ws.cell(row=row, column=1, value="إجمالي الخصوم الجارية").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(balance_data['total_current_liabilities'])).font = Font(bold=True)
        row += 1

    # الخصوم طويلة الأجل
    if balance_data['long_term_liabilities']:
        ws.cell(row=row, column=1, value="الخصوم طويلة الأجل").font = Font(bold=True)
        row += 1

        for item in balance_data['long_term_liabilities']:
            ws.cell(row=row, column=1, value=f"    {item['account'].name}")
            ws.cell(row=row, column=2, value=float(item['balance']))
            row += 1

        ws.cell(row=row, column=1, value="إجمالي الخصوم طويلة الأجل").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(balance_data['total_long_term_liabilities'])).font = Font(bold=True)
        row += 1

    # إجمالي الخصوم
    ws.cell(row=row, column=1, value="إجمالي الخصوم").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(balance_data['total_liabilities'])).font = Font(bold=True)
    row += 2

    # حقوق الملكية
    ws.cell(row=row, column=1, value="حقوق الملكية").font = Font(bold=True)
    row += 1

    for item in balance_data['equity']:
        ws.cell(row=row, column=1, value=f"    {item['account'].name}")
        ws.cell(row=row, column=2, value=float(item['balance']))
        row += 1

    ws.cell(row=row, column=1, value="إجمالي حقوق الملكية").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(balance_data['total_equity'])).font = Font(bold=True)
    row += 2

    # إجمالي الخصوم وحقوق الملكية
    ws.cell(row=row, column=1, value="إجمالي الخصوم وحقوق الملكية").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value=float(balance_data['total_liabilities_equity'])).font = Font(bold=True, size=14)

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"balance_sheet_{as_of_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response