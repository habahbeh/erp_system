# apps/accounting/views/account_reports.py
"""
Account Reports - تقارير الحسابات والاستيراد والتصدير
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils.translation import gettext_lazy as _
from django.db import transaction
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.core.decorators import permission_required_with_message, company_required
from apps.core.models import Currency
from ..models import Account, AccountType
from ..forms.account_forms import AccountImportForm


@login_required
@permission_required_with_message('accounting.view_accounttype')
def export_account_types(request):
    """تصدير أنواع الحسابات إلى Excel"""

    # إنشاء workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "أنواع الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # إضافة الرأس
    headers = ['الرمز', 'الاسم', 'التصنيف', 'الرصيد الطبيعي', 'عدد الحسابات']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # إضافة البيانات
    from django.db.models import Count
    queryset = AccountType.objects.annotate(
        accounts_count=Count('accounts')
    ).order_by('code')

    for row, account_type in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=account_type.code).border = thin_border
        ws.cell(row=row, column=2, value=account_type.name).border = thin_border
        ws.cell(row=row, column=3, value=account_type.get_type_category_display()).border = thin_border
        ws.cell(row=row, column=4, value=account_type.get_normal_balance_display()).border = thin_border
        ws.cell(row=row, column=5, value=account_type.accounts_count).border = thin_border

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="account_types.xlsx"'

    wb.save(response)
    messages.success(request, 'تم تصدير أنواع الحسابات بنجاح')
    return response


@company_required
@login_required
@permission_required_with_message('accounting.view_account')
def export_accounts(request):
    """تصدير الحسابات إلى Excel"""

    company = request.current_company

    # إنشاء workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "دليل الحسابات"

    # تنسيق الرأس
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # إضافة الرأس
    headers = [
        'رمز الحساب', 'اسم الحساب', 'الاسم الإنجليزي', 'نوع الحساب',
        'رمز نوع الحساب', 'الحساب الأب', 'رمز الحساب الأب', 'العملة',
        'رمز العملة', 'طبيعة الحساب', 'الرصيد الافتتاحي', 'تاريخ الرصيد',
        'المستوى', 'حساب نقدي', 'حساب بنكي', 'يقبل قيود', 'الحالة', 'ملاحظات'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # إضافة البيانات
    queryset = Account.objects.filter(
        company=company
    ).select_related('account_type', 'parent', 'currency').order_by('code')

    for row, account in enumerate(queryset, 2):
        data = [
            account.code,
            account.name,
            account.name_en or '',
            account.account_type.name,
            account.account_type.code,
            account.parent.name if account.parent else '',
            account.parent.code if account.parent else '',
            account.currency.name,
            account.currency.code,
            account.get_nature_display(),
            float(account.opening_balance),
            account.opening_balance_date.strftime('%Y-%m-%d') if account.opening_balance_date else '',
            account.level,
            'نعم' if account.is_cash_account else 'لا',
            'نعم' if account.is_bank_account else 'لا',
            'نعم' if account.accept_entries else 'لا',
            'موقوف' if account.is_suspended else 'نشط',
            account.notes or ''
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value).border = thin_border

    # تعديل عرض الأعمدة
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="chart_of_accounts_{company.name}.xlsx"'

    wb.save(response)
    messages.success(request, 'تم تصدير دليل الحسابات بنجاح')
    return response


@login_required
@permission_required_with_message('accounting.add_accounttype')
def import_account_types(request):
    """استيراد أنواع الحسابات من Excel/CSV"""

    if request.method == 'POST':
        form = AccountImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = form.cleaned_data['file']
                update_existing = form.cleaned_data.get('update_existing', False)

                # قراءة الملف
                if file.name.lower().endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # التحقق من الأعمدة المطلوبة
                required_columns = ['code', 'name', 'type_category', 'normal_balance']
                missing_columns = [col for col in required_columns if col not in df.columns]

                if missing_columns:
                    messages.error(request, f'الملف لا يحتوي على الأعمدة المطلوبة: {", ".join(missing_columns)}')
                    return redirect('accounting:account_type_list')

                # تنظيف البيانات
                df = df.dropna(subset=required_columns)
                df['code'] = df['code'].astype(str).str.strip().str.upper()
                df['name'] = df['name'].astype(str).str.strip()

                # استيراد البيانات مع transaction
                with transaction.atomic():
                    created_count = 0
                    updated_count = 0
                    errors = []

                    for index, row in df.iterrows():
                        try:
                            # التحقق من صحة البيانات
                            if row['type_category'] not in dict(AccountType.ACCOUNT_TYPES):
                                errors.append(f'الصف {index + 2}: تصنيف حساب غير صحيح - {row["type_category"]}')
                                continue

                            if row['normal_balance'] not in ['debit', 'credit']:
                                errors.append(f'الصف {index + 2}: رصيد طبيعي غير صحيح - {row["normal_balance"]}')
                                continue

                            # إنشاء أو تحديث نوع الحساب
                            account_type_data = {
                                'name': row['name'],
                                'type_category': row['type_category'],
                                'normal_balance': row['normal_balance']
                            }

                            account_type, created = AccountType.objects.get_or_create(
                                code=row['code'],
                                defaults=account_type_data
                            )

                            if created:
                                created_count += 1
                            elif update_existing:
                                for key, value in account_type_data.items():
                                    setattr(account_type, key, value)
                                account_type.save()
                                updated_count += 1

                        except Exception as e:
                            errors.append(f'الصف {index + 2}: {str(e)}')

                # رسائل النتائج
                if created_count > 0:
                    messages.success(request, f'تم إنشاء {created_count} نوع حساب جديد')
                if updated_count > 0:
                    messages.success(request, f'تم تحديث {updated_count} نوع حساب')
                if errors:
                    for error in errors[:10]:  # إظهار أول 10 أخطاء
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f'وجد {len(errors) - 10} خطأ إضافي')

            except Exception as e:
                messages.error(request, f'خطأ في معالجة الملف: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')

    return redirect('accounting:account_type_list')


@company_required
@login_required
@permission_required_with_message('accounting.add_account')
def import_accounts(request):
    """استيراد الحسابات من Excel/CSV"""

    if request.method == 'POST':
        form = AccountImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = form.cleaned_data['file']
                update_existing = form.cleaned_data.get('update_existing', False)
                validate_hierarchy = form.cleaned_data.get('validate_hierarchy', True)

                # قراءة الملف
                if file.name.lower().endswith('.csv'):
                    df = pd.read_csv(file, encoding='utf-8-sig')
                else:
                    df = pd.read_excel(file)

                # التحقق من الأعمدة المطلوبة
                required_columns = ['code', 'name', 'account_type_code', 'currency_code']
                missing_columns = [col for col in required_columns if col not in df.columns]

                if missing_columns:
                    messages.error(request, f'الملف لا يحتوي على الأعمدة المطلوبة: {", ".join(missing_columns)}')
                    return redirect('accounting:account_list')

                # تنظيف البيانات
                df = df.dropna(subset=required_columns)
                df['code'] = df['code'].astype(str).str.strip().str.upper()
                df['name'] = df['name'].astype(str).str.strip()
                df['account_type_code'] = df['account_type_code'].astype(str).str.strip().str.upper()
                df['currency_code'] = df['currency_code'].astype(str).str.strip().str.upper()

                # ملء القيم الناقصة
                df['name_en'] = df.get('name_en', '').fillna('')
                df['parent_code'] = df.get('parent_code', '').fillna('')
                df['nature'] = df.get('nature', 'both').fillna('both')
                df['opening_balance'] = pd.to_numeric(df.get('opening_balance', 0), errors='coerce').fillna(0)
                df['is_suspended'] = df.get('is_suspended', False).fillna(False)
                df['notes'] = df.get('notes', '').fillna('')

                # استيراد البيانات مع transaction
                with transaction.atomic():
                    created_count = 0
                    updated_count = 0
                    errors = []

                    # إنشاء خريطة للحسابات المستوردة
                    imported_accounts = {}

                    # المرحلة الأولى: إنشاء الحسابات بدون آباء
                    for index, row in df.iterrows():
                        if pd.isna(row.get('parent_code')) or not str(row.get('parent_code')).strip():
                            result = create_or_update_account(
                                request, row, index, update_existing, imported_accounts, errors
                            )
                            if result:
                                if result['created']:
                                    created_count += 1
                                elif result['updated']:
                                    updated_count += 1
                                imported_accounts[row['code']] = result['account']

                    # المرحلة الثانية: إنشاء الحسابات مع آباء
                    remaining_rows = df[df['parent_code'].notna() & (df['parent_code'].str.strip() != '')]
                    max_iterations = 5  # لتجنب الدورات اللانهائية

                    for iteration in range(max_iterations):
                        processed_any = False
                        unprocessed_rows = []

                        for index, row in remaining_rows.iterrows():
                            parent_code = str(row['parent_code']).strip()

                            # البحث عن الأب في الحسابات المستوردة أو الموجودة
                            parent = imported_accounts.get(parent_code)
                            if not parent:
                                try:
                                    parent = Account.objects.get(
                                        code=parent_code,
                                        company=request.current_company
                                    )
                                except Account.DoesNotExist:
                                    unprocessed_rows.append((index, row))
                                    continue

                            # إنشاء الحساب
                            result = create_or_update_account(
                                request, row, index, update_existing, imported_accounts, errors, parent
                            )
                            if result:
                                if result['created']:
                                    created_count += 1
                                elif result['updated']:
                                    updated_count += 1
                                imported_accounts[row['code']] = result['account']
                                processed_any = True

                        if not unprocessed_rows:
                            break

                        if not processed_any:
                            # إضافة أخطاء للحسابات التي لم تُعالج
                            for index, row in unprocessed_rows:
                                errors.append(f'الصف {index + 2}: الحساب الأب غير موجود - {row["parent_code"]}')
                            break

                        # إعداد البيانات للتكرار التالي
                        remaining_rows = pd.DataFrame([row for index, row in unprocessed_rows])

                # رسائل النتائج
                if created_count > 0:
                    messages.success(request, f'تم إنشاء {created_count} حساب جديد')
                if updated_count > 0:
                    messages.success(request, f'تم تحديث {updated_count} حساب')
                if errors:
                    for error in errors[:15]:  # إظهار أول 15 خطأ
                        messages.error(request, error)
                    if len(errors) > 15:
                        messages.warning(request, f'وجد {len(errors) - 15} خطأ إضافي')

            except Exception as e:
                messages.error(request, f'خطأ في معالجة الملف: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # عرض نموذج الاستيراد
        form = AccountImportForm()
        return render(request, 'accounting/account/import_accounts.html', {
            'form': form,
            'title': 'استيراد الحسابات',
            'breadcrumbs': [
                {'title': 'الرئيسية', 'url': '/'},
                {'title': 'المحاسبة', 'url': '/accounting/'},
                {'title': 'دليل الحسابات', 'url': '/accounting/accounts/'},
                {'title': 'استيراد الحسابات', 'url': ''},
            ]
        })

    return redirect('accounting:account_list')


def create_or_update_account(request, row, index, update_existing, imported_accounts, errors, parent=None):
    """دالة مساعدة لإنشاء أو تحديث حساب"""

    try:
        # البحث عن نوع الحساب
        try:
            account_type = AccountType.objects.get(code=row['account_type_code'])
        except AccountType.DoesNotExist:
            errors.append(f'الصف {index + 2}: نوع الحساب غير موجود - {row["account_type_code"]}')
            return None

        # البحث عن العملة
        try:
            currency = Currency.objects.get(
                code=row['currency_code'],
                company=request.current_company
            )
        except Currency.DoesNotExist:
            try:
                # البحث في العملات الأساسية
                currency = Currency.objects.get(
                    code=row['currency_code'],
                    is_base_currency=True
                )
            except Currency.DoesNotExist:
                errors.append(f'الصف {index + 2}: العملة غير موجودة - {row["currency_code"]}')
                return None

        # إعداد بيانات الحساب
        account_data = {
            'name': row['name'],
            'name_en': row.get('name_en', ''),
            'account_type': account_type,
            'parent': parent,
            'currency': currency,
            'nature': row.get('nature', 'both'),
            'opening_balance': float(row.get('opening_balance', 0)),
            'is_suspended': bool(row.get('is_suspended', False)),
            'notes': row.get('notes', ''),
            'company': request.current_company,
            'branch': request.current_branch,
            'created_by': request.user
        }

        # إنشاء أو تحديث الحساب
        account, created = Account.objects.get_or_create(
            code=row['code'],
            company=request.current_company,
            defaults=account_data
        )

        if created:
            return {'account': account, 'created': True, 'updated': False}
        elif update_existing:
            for key, value in account_data.items():
                if key not in ['company', 'branch', 'created_by']:
                    setattr(account, key, value)
            account.save()
            return {'account': account, 'created': False, 'updated': True}
        else:
            return {'account': account, 'created': False, 'updated': False}

    except Exception as e:
        errors.append(f'الصف {index + 2}: {str(e)}')
        return None