# apps/assets/utils/export_utils.py
"""
دوال التصدير إلى Excel و PDF
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils.translation import gettext as _

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExcelExporter:
    """فئة لتصدير البيانات إلى Excel"""

    def __init__(self, company_name=''):
        """تهيئة الـ Exporter"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Run: pip install openpyxl")

        self.wb = Workbook()
        self.ws = self.wb.active
        self.company_name = company_name
        self.current_row = 1

        # الأنماط
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_title(self, title):
        """إضافة عنوان التقرير"""
        self.ws.cell(row=self.current_row, column=1, value=title)
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 1

    def add_info_row(self, label, value):
        """إضافة صف معلومات"""
        self.ws.cell(row=self.current_row, column=1, value=label)
        self.ws.cell(row=self.current_row, column=2, value=value)
        self.current_row += 1

    def add_empty_row(self):
        """إضافة صف فارغ"""
        self.current_row += 1

    def add_headers(self, headers):
        """إضافة رؤوس الأعمدة"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        self.current_row += 1

    def add_data_rows(self, data_rows):
        """إضافة صفوف البيانات"""
        for row_data in data_rows:
            for col_num, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
                cell.font = self.normal_font
                cell.border = self.border

                # تنسيق الأرقام
                if isinstance(value, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (float, Decimal)):
                        cell.number_format = '#,##0.000'
                else:
                    cell.alignment = Alignment(horizontal='right')

            self.current_row += 1

    def add_summary_row(self, summary_data, bold=True):
        """إضافة صف الإجمالي"""
        for col_num, value in enumerate(summary_data, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
            if bold:
                cell.font = Font(name='Arial', size=10, bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='right')

            if isinstance(value, (int, float, Decimal)) and col_num > 1:
                cell.number_format = '#,##0.000'

        self.current_row += 1

    def auto_adjust_columns(self):
        """ضبط عرض الأعمدة تلقائياً"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def get_response(self, filename):
        """الحصول على HttpResponse للتحميل"""
        self.auto_adjust_columns()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class PDFExporter:
    """فئة لتصدير البيانات إلى PDF"""

    def __init__(self, company_name='', orientation='portrait'):
        """تهيئة الـ PDF Exporter"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is not installed. Run: pip install reportlab")

        self.company_name = company_name
        self.orientation = orientation
        self.buffer = BytesIO()
        self.elements = []

        # تحديد حجم الصفحة
        if orientation == 'landscape':
            self.pagesize = landscape(A4)
        else:
            self.pagesize = A4

        # تحميل الخط العربي (Cairo) - اختياري
        try:
            pdfmetrics.registerFont(TTFont('Cairo', 'static/fonts/Cairo-Regular.ttf'))
            pdfmetrics.registerFont(TTFont('Cairo-Bold', 'static/fonts/Cairo-Bold.ttf'))
            self.arabic_font = 'Cairo'
            self.arabic_font_bold = 'Cairo-Bold'
        except:
            # fallback إلى Helvetica
            self.arabic_font = 'Helvetica'
            self.arabic_font_bold = 'Helvetica-Bold'

        # الأنماط
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=self.arabic_font_bold,
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Normal'],
            fontName=self.arabic_font,
            fontSize=10,
            alignment=TA_RIGHT,
            spaceAfter=6
        )

    def add_title(self, title):
        """إضافة عنوان التقرير"""
        self.elements.append(Paragraph(title, self.title_style))
        self.elements.append(Spacer(1, 0.3 * cm))

    def add_header_info(self, info_dict):
        """إضافة معلومات رأس التقرير"""
        for label, value in info_dict.items():
            text = f"<b>{label}:</b> {value}"
            self.elements.append(Paragraph(text, self.header_style))

        self.elements.append(Spacer(1, 0.5 * cm))

    def add_table(self, data, col_widths=None):
        """إضافة جدول"""
        if not data:
            return

        # إنشاء الجدول
        table = Table(data, colWidths=col_widths)

        # تنسيق الجدول
        table.setStyle(TableStyle([
            # رأس الجدول
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.arabic_font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

            # محتوى الجدول
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # الحدود
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # صف الإجمالي
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
            ('FONTNAME', (0, -1), (-1, -1), self.arabic_font_bold),
        ]))

        self.elements.append(table)
        self.elements.append(Spacer(1, 0.5 * cm))

    def add_spacer(self, height=0.5):
        """إضافة مسافة فارغة"""
        self.elements.append(Spacer(1, height * cm))

    def add_page_break(self):
        """إضافة فاصل صفحة"""
        self.elements.append(PageBreak())

    def get_response(self, filename):
        """الحصول على HttpResponse للتحميل"""
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.pagesize,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm
        )

        doc.build(self.elements)

        self.buffer.seek(0)

        response = HttpResponse(
            self.buffer.read(),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


def format_currency(value, currency_symbol='JOD'):
    """تنسيق العملة"""
    if value is None:
        return '0.000'
    return f"{float(value):,.3f} {currency_symbol}"


def format_date(date_obj):
    """تنسيق التاريخ"""
    if not date_obj:
        return '-'
    return date_obj.strftime('%Y-%m-%d')