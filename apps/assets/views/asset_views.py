# apps/assets/views/asset_views.py
"""
Views الأصول الثابتة - نسخة كاملة ومحسّنة
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField, Avg
from django.db.models.functions import Coalesce
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.utils import timezone
import json
from datetime import date, timedelta, datetime
from decimal import Decimal

from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import qrcode
import barcode
from barcode.writer import ImageWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import (
    Asset, AssetCategory, DepreciationMethod, AssetCondition,
    AssetAttachment, AssetValuation
)
from apps.accounting.models import Account, CostCenter, JournalEntry


# ==================== Asset CRUD ====================

class AssetListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة الأصول الثابتة"""

    model = Asset
    template_name = 'assets/assets/asset_list.html'
    context_object_name = 'assets'
    permission_required = 'assets.view_asset'
    paginate_by = 25

    def get_queryset(self):
        queryset = Asset.objects.filter(
            company=self.request.current_company
        ).select_related(
            'category', 'condition', 'depreciation_method',
            'cost_center', 'responsible_employee', 'supplier', 'currency'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        category = self.request.GET.get('category')
        branch = self.request.GET.get('branch')
        cost_center = self.request.GET.get('cost_center')
        condition = self.request.GET.get('condition')
        responsible = self.request.GET.get('responsible')
        search = self.request.GET.get('search')

        # فلترة حسب التاريخ
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if status:
            queryset = queryset.filter(status=status)

        if category:
            queryset = queryset.filter(category_id=category)

        if branch:
            queryset = queryset.filter(branch_id=branch)

        if cost_center:
            queryset = queryset.filter(cost_center_id=cost_center)

        if condition:
            queryset = queryset.filter(condition_id=condition)

        if responsible:
            queryset = queryset.filter(responsible_employee_id=responsible)

        if date_from:
            queryset = queryset.filter(purchase_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(purchase_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(asset_number__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(barcode__icontains=search) |
                Q(description__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-purchase_date')
        queryset = queryset.order_by(sort_by, '-asset_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات سريعة
        assets = Asset.objects.filter(company=self.request.current_company)
        stats = {
            'total': assets.count(),
            'active': assets.filter(status='active').count(),
            'under_maintenance': assets.filter(status='under_maintenance').count(),
            'total_value': assets.filter(status='active').aggregate(
                total=Coalesce(Sum('book_value'), Decimal('0'))
            )['total'],
        }

        # الفئات للفلترة
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        conditions = AssetCondition.objects.filter(is_active=True).order_by('name')

        context.update({
            'title': _('الأصول الثابتة'),
            'can_add': self.request.user.has_perm('assets.add_asset'),
            'can_edit': self.request.user.has_perm('assets.change_asset'),
            'can_delete': self.request.user.has_perm('assets.delete_asset'),
            'can_export': self.request.user.has_perm('assets.view_asset'),
            'status_choices': Asset.STATUS_CHOICES,
            'categories': categories,
            'conditions': conditions,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': ''},
            ]
        })
        return context


class AssetCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء أصل جديد"""

    model = Asset
    template_name = 'assets/assets/asset_form.html'
    permission_required = 'assets.add_asset'
    fields = [
        'name', 'name_en', 'category', 'condition', 'status',
        'purchase_date', 'purchase_invoice_number', 'supplier',
        'currency', 'original_cost', 'salvage_value',
        'depreciation_method', 'useful_life_months', 'depreciation_start_date',
        'depreciation_status', 'total_expected_units', 'unit_name',
        'cost_center', 'responsible_employee', 'physical_location',
        'serial_number', 'model', 'manufacturer',
        'warranty_start_date', 'warranty_end_date', 'warranty_provider',
        'barcode', 'is_leased', 'description', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # تخصيص الحقول
        company = self.request.current_company

        form.fields['category'].queryset = AssetCategory.objects.filter(
            company=company, is_active=True
        )
        form.fields['depreciation_method'].queryset = DepreciationMethod.objects.filter(
            is_active=True
        )
        form.fields['condition'].queryset = AssetCondition.objects.filter(
            is_active=True
        )

        if 'cost_center' in form.fields:
            form.fields['cost_center'].queryset = CostCenter.objects.filter(
                company=company, is_active=True
            )

        # القيم الافتراضية
        form.fields['purchase_date'].initial = date.today()
        form.fields['depreciation_start_date'].initial = date.today()
        form.fields['status'].initial = 'active'
        form.fields['depreciation_status'].initial = 'active'

        # إضافة classes للـ styling
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        form.instance.company = self.request.current_company
        form.instance.branch = self.request.current_branch
        form.instance.created_by = self.request.user

        # إذا لم تحدد العملة، استخدم عملة الشركة
        if not form.instance.currency:
            form.instance.currency = self.request.current_company.base_currency

        try:
            # حفظ الأصل
            self.object = form.save()

            # ✅ إنشاء القيد المحاسبي إذا طُلب
            create_journal = self.request.POST.get('create_journal_entry')
            journal_entry = None

            if create_journal:
                try:
                    journal_entry = self.object.create_purchase_journal_entry(user=self.request.user)
                    messages.success(
                        self.request,
                        f'✅ تم إنشاء الأصل {self.object.asset_number} والقيد المحاسبي {journal_entry.number} بنجاح'
                    )
                except ValidationError as e:
                    messages.warning(
                        self.request,
                        f'⚠️ تم إنشاء الأصل {self.object.asset_number} لكن فشل إنشاء القيد: {str(e)}'
                    )
            else:
                messages.success(
                    self.request,
                    f'✅ تم إنشاء الأصل {self.object.asset_number} بنجاح'
                )

            # Log the action
            self.log_action('create', self.object)

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ خطأ في التحقق: {e}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ حدث خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:asset_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة أصل جديد'),
            'submit_text': _('إنشاء الأصل'),
            'show_create_journal_option': True,  # ✅ إظهار خيار إنشاء القيد المحاسبي
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': _('إضافة أصل'), 'url': ''},
            ]
        })
        return context


class AssetUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل أصل"""

    model = Asset
    template_name = 'assets/assets/asset_form.html'
    permission_required = 'assets.change_asset'
    fields = [
        'name', 'name_en', 'category', 'condition', 'status',
        'purchase_date', 'purchase_invoice_number', 'supplier',
        'currency', 'original_cost', 'salvage_value',
        'depreciation_method', 'useful_life_months', 'depreciation_start_date',
        'depreciation_status', 'total_expected_units', 'unit_name',
        'cost_center', 'responsible_employee', 'physical_location',
        'serial_number', 'model', 'manufacturer',
        'warranty_start_date', 'warranty_end_date', 'warranty_provider',
        'barcode', 'is_leased', 'insurance_status', 'description', 'notes'
    ]

    def get_queryset(self):
        return Asset.objects.filter(company=self.request.current_company)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['category'].queryset = AssetCategory.objects.filter(
            company=company, is_active=True
        )
        form.fields['depreciation_method'].queryset = DepreciationMethod.objects.filter(
            is_active=True
        )
        form.fields['condition'].queryset = AssetCondition.objects.filter(
            is_active=True
        )

        if 'cost_center' in form.fields:
            form.fields['cost_center'].queryset = CostCenter.objects.filter(
                company=company, is_active=True
            )

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        # ✅ التحقق من إمكانية التعديل
        if not self.object.can_edit():
            messages.error(
                self.request,
                _('❌ لا يمكن تعديل هذا الأصل (مباع، مستبعد، أو له معاملات مرحّلة)')
            )
            return redirect('assets:asset_detail', pk=self.object.pk)

        try:
            old_values = {
                'status': self.object.status,
                'location': self.object.physical_location,
                'responsible': self.object.responsible_employee,
            }

            self.object = form.save()

            # Log changes
            self.log_action('update', self.object, old_values)

            messages.success(
                self.request,
                f'✅ تم تحديث الأصل {self.object.asset_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ خطأ في التحقق: {e}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ حدث خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:asset_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل الأصل {self.object.asset_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.object.asset_number, 'url': reverse('assets:asset_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class AssetDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل الأصل"""

    model = Asset
    template_name = 'assets/assets/asset_detail.html'
    context_object_name = 'asset'
    permission_required = 'assets.view_asset'

    def get_queryset(self):
        return Asset.objects.filter(
            company=self.request.current_company
        ).select_related(
            'category', 'condition', 'depreciation_method',
            'cost_center', 'responsible_employee', 'supplier',
            'currency', 'created_by'
        ).prefetch_related(
            'attachments', 'valuations', 'depreciation_records',
            'maintenances', 'transactions', 'transfers',
            'insurances', 'leases'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # سجل الإهلاك
        depreciation_records = self.object.depreciation_records.select_related(
            'fiscal_year', 'period'
        ).order_by('-depreciation_date')[:10]

        # الصيانات
        maintenances = self.object.maintenances.select_related(
            'maintenance_type', 'performed_by'
        ).order_by('-scheduled_date')[:5]

        # المرفقات
        attachments = self.object.attachments.order_by('-uploaded_at')

        # المعاملات
        transactions = self.object.transactions.select_related(
            'business_partner'
        ).order_by('-transaction_date')[:5]

        # التقييمات
        valuations = self.object.valuations.select_related(
            'created_by', 'approved_by'
        ).order_by('-valuation_date')[:5]

        # التحويلات
        transfers = self.object.transfers.select_related(
            'from_branch', 'to_branch'
        ).order_by('-transfer_date')[:5]

        # التأمينات
        insurances = self.object.insurances.select_related(
            'insurance_company'
        ).order_by('-start_date')[:5]

        # الإيجارات
        leases = self.object.leases.select_related(
            'lessor'
        ).order_by('-start_date')[:5]

        # إحصائيات الإهلاك
        depreciation_stats = {
            'total_depreciation': self.object.accumulated_depreciation,
            'remaining_value': self.object.book_value,
            'depreciation_percentage': self.object.get_depreciation_percentage(),
            'remaining_months': self.object.get_remaining_months(),
            'is_fully_depreciated': self.object.is_fully_depreciated(),
        }

        # حالات التحذير
        warnings = []
        if self.object.needs_physical_count():
            warnings.append({
                'type': 'warning',
                'icon': 'fa-clipboard-check',
                'message': 'الأصل يحتاج لجرد فعلي'
            })
        if self.object.is_warranty_valid():
            warnings.append({
                'type': 'info',
                'icon': 'fa-shield-alt',
                'message': f'الضمان ساري حتى {self.object.warranty_end_date}'
            })
        elif self.object.warranty_end_date and self.object.warranty_end_date < date.today():
            warnings.append({
                'type': 'danger',
                'icon': 'fa-shield-alt',
                'message': 'الضمان منتهي'
            })

        context.update({
            'title': f'الأصل {self.object.asset_number}',
            # ✅ استخدام methods من الـ model للصلاحيات
            'can_edit': self.request.user.has_perm('assets.change_asset') and self.object.can_edit(),
            'can_delete': self.request.user.has_perm('assets.delete_asset') and self.object.can_delete(),
            'can_depreciate': self.request.user.has_perm('assets.can_calculate_depreciation') and self.object.can_depreciate(),
            'can_calculate_depreciation': self.request.user.has_perm('assets.can_calculate_depreciation'),
            'can_sell': self.request.user.has_perm('assets.can_sell_asset'),
            'can_dispose': self.request.user.has_perm('assets.can_dispose_asset'),
            'can_transfer': self.request.user.has_perm('assets.can_transfer_asset'),
            'can_revalue': self.request.user.has_perm('assets.can_revalue_asset'),
            # ✅ معلومات محاسبية إضافية
            'current_book_value': self.object.current_book_value,
            'total_accumulated_depreciation': self.object.get_total_accumulated_depreciation(),
            # البيانات الموجودة
            'depreciation_records': depreciation_records,
            'maintenances': maintenances,
            'attachments': attachments,
            'transactions': transactions,
            'valuations': valuations,
            'transfers': transfers,
            'insurances': insurances,
            'leases': leases,
            'depreciation_stats': depreciation_stats,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.object.asset_number, 'url': ''},
            ]
        })
        return context


class AssetDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف أصل"""

    model = Asset
    template_name = 'assets/assets/asset_confirm_delete.html'
    permission_required = 'assets.delete_asset'
    success_url = reverse_lazy('assets:asset_list')

    def get_queryset(self):
        return Asset.objects.filter(company=self.request.current_company)

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # ✅ التحقق من إمكانية الحذف باستخدام الـ method
        if not self.object.can_delete():
            messages.error(
                request,
                _('❌ لا يمكن حذف هذا الأصل (له إهلاك، معاملات، صيانة، تأمين، أو جرد)')
            )
            return redirect('assets:asset_detail', pk=self.object.pk)

        asset_number = self.object.asset_number
        messages.success(request, f'✅ تم حذف الأصل {asset_number} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Categories ====================

class AssetCategoryListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة فئات الأصول"""

    model = AssetCategory
    template_name = 'assets/categories/category_list.html'
    context_object_name = 'categories'
    permission_required = 'assets.view_assetcategory'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetCategory.objects.filter(
            company=self.request.current_company
        ).select_related(
            'parent', 'default_depreciation_method',
            'asset_account', 'accumulated_depreciation_account'
        ).annotate(
            asset_count=Count('assets', filter=Q(assets__status='active'))
        )

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        return queryset.order_by('code')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company
        )
        stats = {
            'total': categories.count(),
            'with_assets': categories.annotate(
                count=Count('assets', filter=Q(assets__status='active'))
            ).filter(count__gt=0).count(),
        }

        context.update({
            'title': _('فئات الأصول'),
            'can_add': self.request.user.has_perm('assets.add_assetcategory'),
            'can_edit': self.request.user.has_perm('assets.change_assetcategory'),
            'can_delete': self.request.user.has_perm('assets.delete_assetcategory'),
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('فئات الأصول'), 'url': ''},
            ]
        })
        return context


class AssetCategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء فئة أصول جديدة"""

    model = AssetCategory
    template_name = 'assets/categories/category_form.html'
    permission_required = 'assets.add_assetcategory'
    fields = [
        'code', 'name', 'name_en', 'parent',
        'asset_account', 'accumulated_depreciation_account',
        'depreciation_expense_account', 'loss_on_disposal_account',
        'gain_on_sale_account', 'maintenance_expense_account',
        'default_depreciation_method', 'default_useful_life_months',
        'default_salvage_value_rate', 'default_physical_count_frequency',
        'description'
    ]
    success_url = reverse_lazy('assets:category_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        # تخصيص الحقول
        form.fields['parent'].queryset = AssetCategory.objects.filter(
            company=company, is_active=True
        )
        form.fields['parent'].required = False

        # الحسابات المحاسبية
        accounts = Account.objects.filter(
            company=company,
            accept_entries=True,
            is_active=True
        )

        for field_name in [
            'asset_account', 'accumulated_depreciation_account',
            'depreciation_expense_account', 'loss_on_disposal_account',
            'gain_on_sale_account', 'maintenance_expense_account'
        ]:
            if field_name in form.fields:
                form.fields[field_name].queryset = accounts
                form.fields[field_name].required = False

        form.fields['default_depreciation_method'].queryset = DepreciationMethod.objects.filter(
            is_active=True
        )
        form.fields['default_depreciation_method'].required = False

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.created_by = self.request.user
            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء الفئة {self.object.code} بنجاح'
            )
            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'❌ حدث خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة فئة أصول'),
            'submit_text': _('إنشاء الفئة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('فئات الأصول'), 'url': reverse('assets:category_list')},
                {'title': _('إضافة فئة'), 'url': ''},
            ]
        })
        return context


class AssetCategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل فئة أصول"""

    model = AssetCategory
    template_name = 'assets/categories/category_form.html'
    permission_required = 'assets.change_assetcategory'
    fields = [
        'code', 'name', 'name_en', 'parent',
        'asset_account', 'accumulated_depreciation_account',
        'depreciation_expense_account', 'loss_on_disposal_account',
        'gain_on_sale_account', 'maintenance_expense_account',
        'default_depreciation_method', 'default_useful_life_months',
        'default_salvage_value_rate', 'default_physical_count_frequency',
        'description'
    ]
    success_url = reverse_lazy('assets:category_list')

    def get_queryset(self):
        return AssetCategory.objects.filter(company=self.request.current_company)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        # منع اختيار نفسه كأب
        form.fields['parent'].queryset = AssetCategory.objects.filter(
            company=company, is_active=True
        ).exclude(pk=self.object.pk)
        form.fields['parent'].required = False

        # الحسابات المحاسبية
        accounts = Account.objects.filter(
            company=company,
            accept_entries=True,
            is_active=True
        )

        for field_name in [
            'asset_account', 'accumulated_depreciation_account',
            'depreciation_expense_account', 'loss_on_disposal_account',
            'gain_on_sale_account', 'maintenance_expense_account'
        ]:
            if field_name in form.fields:
                form.fields[field_name].queryset = accounts
                form.fields[field_name].required = False

        form.fields['default_depreciation_method'].queryset = DepreciationMethod.objects.filter(
            is_active=True
        )
        form.fields['default_depreciation_method'].required = False

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث الفئة {self.object.code} بنجاح'
            )
            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'❌ حدث خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل الفئة {self.object.code}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('فئات الأصول'), 'url': reverse('assets:category_list')},
                {'title': self.object.code, 'url': reverse('assets:category_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class AssetCategoryDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل فئة الأصول"""

    model = AssetCategory
    template_name = 'assets/categories/category_detail.html'
    context_object_name = 'category'
    permission_required = 'assets.view_assetcategory'

    def get_queryset(self):
        return AssetCategory.objects.filter(
            company=self.request.current_company
        ).select_related(
            'parent', 'default_depreciation_method',
            'asset_account', 'accumulated_depreciation_account',
            'depreciation_expense_account', 'loss_on_disposal_account',
            'gain_on_sale_account', 'maintenance_expense_account'
        ).prefetch_related('assets', 'children')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات الفئة
        assets = self.object.assets.filter(status='active')
        assets_stats = assets.aggregate(
            total_count=Count('id'),
            total_original_cost=Coalesce(Sum('original_cost'), Decimal('0')),
            total_accumulated_depreciation=Coalesce(Sum('accumulated_depreciation'), Decimal('0')),
            total_book_value=Coalesce(Sum('book_value'), Decimal('0')),
            avg_age_months=Avg(
                (timezone.now().year - F('purchase_date__year')) * 12 +
                (timezone.now().month - F('purchase_date__month'))
            )
        )

        # الفئات الفرعية
        children = self.object.children.annotate(
            asset_count=Count('assets', filter=Q(assets__status='active'))
        )

        # آخر الأصول المضافة
        recent_assets = assets.order_by('-created_at')[:10]

        context.update({
            'title': f'الفئة {self.object.code}',
            'can_edit': self.request.user.has_perm('assets.change_assetcategory'),
            'can_delete': self.request.user.has_perm('assets.delete_assetcategory'),
            'assets_stats': assets_stats,
            'children': children,
            'recent_assets': recent_assets,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('فئات الأصول'), 'url': reverse('assets:category_list')},
                {'title': self.object.code, 'url': ''},
            ]
        })
        return context


class AssetCategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف فئة أصول"""

    model = AssetCategory
    template_name = 'assets/categories/category_confirm_delete.html'
    permission_required = 'assets.delete_assetcategory'
    success_url = reverse_lazy('assets:category_list')

    def get_queryset(self):
        return AssetCategory.objects.filter(company=self.request.current_company)

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق من إمكانية الحذف
        errors = []

        if self.object.assets.exists():
            errors.append(f'لا يمكن حذف فئة لديها {self.object.assets.count()} أصول')

        if self.object.children.exists():
            errors.append(f'لا يمكن حذف فئة لديها {self.object.children.count()} فئات فرعية')

        if errors:
            for error in errors:
                messages.error(request, f'❌ {error}')
            return redirect('assets:category_detail', pk=self.object.pk)

        category_code = self.object.code
        messages.success(request, f'✅ تم حذف الفئة {category_code} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Depreciation Methods ====================

class DepreciationMethodListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة طرق الإهلاك"""

    model = DepreciationMethod
    template_name = 'assets/categories/depreciation_method_list.html'
    context_object_name = 'methods'
    permission_required = 'assets.view_depreciationmethod'
    paginate_by = 50

    def get_queryset(self):
        queryset = DepreciationMethod.objects.all().annotate(
            usage_count=Count('asset')
        )

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        # الفلترة
        method_type = self.request.GET.get('method_type')
        if method_type:
            queryset = queryset.filter(method_type=method_type)

        is_active = self.request.GET.get('is_active')
        if is_active:
            queryset = queryset.filter(is_active=is_active == '1')

        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('طرق الإهلاك'),
            'can_add': self.request.user.has_perm('assets.add_depreciationmethod'),
            'can_edit': self.request.user.has_perm('assets.change_depreciationmethod'),
            'can_delete': self.request.user.has_perm('assets.delete_depreciationmethod'),
            'method_types': DepreciationMethod.METHOD_TYPES,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('طرق الإهلاك'), 'url': ''},
            ]
        })
        return context


class DepreciationMethodCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء طريقة إهلاك"""

    model = DepreciationMethod
    template_name = 'assets/categories/depreciation_method_form.html'
    permission_required = 'assets.add_depreciationmethod'
    fields = ['code', 'name', 'name_en', 'method_type', 'rate_percentage', 'description', 'is_active']
    success_url = reverse_lazy('assets:depreciation_method_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        # Help texts
        form.fields['rate_percentage'].help_text = 'للقسط المتناقص فقط - مثال: 200 للقسط المتناقص المضاعف'

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم إنشاء طريقة الإهلاك {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة طريقة إهلاك'),
            'submit_text': _('إنشاء الطريقة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('طرق الإهلاك'), 'url': reverse('assets:depreciation_method_list')},
                {'title': _('إضافة طريقة'), 'url': ''},
            ]
        })
        return context


class DepreciationMethodUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل طريقة إهلاك"""

    model = DepreciationMethod
    template_name = 'assets/categories/depreciation_method_form.html'
    permission_required = 'assets.change_depreciationmethod'
    fields = ['code', 'name', 'name_en', 'method_type', 'rate_percentage', 'description', 'is_active']
    success_url = reverse_lazy('assets:depreciation_method_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم تحديث طريقة الإهلاك {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل طريقة الإهلاك {self.object.name}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('طرق الإهلاك'), 'url': reverse('assets:depreciation_method_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class DepreciationMethodDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """عرض تفاصيل طريقة الإهلاك"""

    model = DepreciationMethod
    template_name = 'assets/categories/depreciation_method_detail.html'
    context_object_name = 'method'
    permission_required = 'assets.view_depreciationmethod'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الأصول المستخدمة لهذه الطريقة
        assets = Asset.objects.filter(
            depreciation_method=self.object,
            status='active'
        ).select_related('category', 'condition')

        # إحصائيات
        stats = assets.aggregate(
            total_count=Count('id'),
            total_value=Coalesce(Sum('book_value'), Decimal('0')),
            avg_useful_life=Avg('useful_life_months')
        )

        context.update({
            'title': f'طريقة الإهلاك: {self.object.name}',
            'can_edit': self.request.user.has_perm('assets.change_depreciationmethod'),
            'can_delete': self.request.user.has_perm('assets.delete_depreciationmethod'),
            'assets': assets[:20],  # أول 20 أصل
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('طرق الإهلاك'), 'url': reverse('assets:depreciation_method_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class DepreciationMethodDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف طريقة إهلاك"""

    model = DepreciationMethod
    template_name = 'assets/categories/depreciation_method_confirm_delete.html'
    permission_required = 'assets.delete_depreciationmethod'
    success_url = reverse_lazy('assets:depreciation_method_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق
        asset_count = Asset.objects.filter(depreciation_method=self.object).count()
        if asset_count > 0:
            messages.error(
                request,
                f'❌ لا يمكن حذف طريقة إهلاك مستخدمة في {asset_count} أصول'
            )
            return redirect('assets:depreciation_method_list')

        method_name = self.object.name
        messages.success(request, f'✅ تم حذف طريقة الإهلاك {method_name} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Conditions ====================

class AssetConditionListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """قائمة حالات الأصول"""

    model = AssetCondition
    template_name = 'assets/categories/condition_list.html'
    context_object_name = 'conditions'
    permission_required = 'assets.view_assetcondition'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetCondition.objects.all().annotate(
            usage_count=Count('assets', filter=Q(assets__status='active'))
        )

        # البحث
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        is_active = self.request.GET.get('is_active')
        if is_active:
            queryset = queryset.filter(is_active=is_active == '1')

        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('حالات الأصول'),
            'can_add': self.request.user.has_perm('assets.add_assetcondition'),
            'can_edit': self.request.user.has_perm('assets.change_assetcondition'),
            'can_delete': self.request.user.has_perm('assets.delete_assetcondition'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('حالات الأصول'), 'url': ''},
            ]
        })
        return context


class AssetConditionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """إنشاء حالة أصل"""

    model = AssetCondition
    template_name = 'assets/categories/condition_form.html'
    permission_required = 'assets.add_assetcondition'
    fields = ['name', 'name_en', 'color_code', 'description', 'is_active']
    success_url = reverse_lazy('assets:condition_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput']:
                field.widget.attrs.update({'class': 'form-control'})

        # Color picker
        form.fields['color_code'].widget.attrs.update({
            'type': 'color',
            'class': 'form-control form-control-color'
        })

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم إنشاء الحالة {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة حالة أصل'),
            'submit_text': _('إنشاء الحالة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('حالات الأصول'), 'url': reverse('assets:condition_list')},
                {'title': _('إضافة حالة'), 'url': ''},
            ]
        })
        return context


class AssetConditionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """تعديل حالة أصل"""

    model = AssetCondition
    template_name = 'assets/categories/condition_form.html'
    permission_required = 'assets.change_assetcondition'
    fields = ['name', 'name_en', 'color_code', 'description', 'is_active']
    success_url = reverse_lazy('assets:condition_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput']:
                field.widget.attrs.update({'class': 'form-control'})

        # Color picker
        form.fields['color_code'].widget.attrs.update({
            'type': 'color',
            'class': 'form-control form-control-color'
        })

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم تحديث الحالة {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل حالة الأصل {self.object.name}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('حالات الأصول'), 'url': reverse('assets:condition_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class AssetConditionDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """عرض تفاصيل حالة الأصل"""

    model = AssetCondition
    template_name = 'assets/categories/condition_detail.html'
    context_object_name = 'condition'
    permission_required = 'assets.view_assetcondition'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الأصول في هذه الحالة
        assets = Asset.objects.filter(
            condition=self.object,
            status='active'
        ).select_related('category', 'depreciation_method')

        # إحصائيات
        stats = assets.aggregate(
            total_count=Count('id'),
            total_value=Coalesce(Sum('book_value'), Decimal('0')),
            avg_age_months=Avg(
                (timezone.now().year - F('purchase_date__year')) * 12 +
                (timezone.now().month - F('purchase_date__month'))
            )
        )

        context.update({
            'title': f'حالة الأصل: {self.object.name}',
            'can_edit': self.request.user.has_perm('assets.change_assetcondition'),
            'can_delete': self.request.user.has_perm('assets.delete_assetcondition'),
            'assets': assets[:20],  # أول 20 أصل
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('حالات الأصول'), 'url': reverse('assets:condition_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class AssetConditionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """حذف حالة أصل"""

    model = AssetCondition
    template_name = 'assets/categories/condition_confirm_delete.html'
    permission_required = 'assets.delete_assetcondition'
    success_url = reverse_lazy('assets:condition_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق
        asset_count = Asset.objects.filter(condition=self.object).count()
        if asset_count > 0:
            messages.error(
                request,
                f'❌ لا يمكن حذف حالة مستخدمة في {asset_count} أصول'
            )
            return redirect('assets:condition_list')

        condition_name = self.object.name
        messages.success(request, f'✅ تم حذف الحالة {condition_name} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Attachments ====================

class AssetAttachmentListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة مرفقات الأصول"""

    model = AssetAttachment
    template_name = 'assets/assets/attachment_list.html'
    context_object_name = 'attachments'
    permission_required = 'assets.view_asset'
    paginate_by = 50

    def get_queryset(self):
        asset_id = self.kwargs.get('asset_id')
        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company
            )
            queryset = AssetAttachment.objects.filter(asset=asset)
        else:
            queryset = AssetAttachment.objects.filter(
                asset__company=self.request.current_company
            )

        queryset = queryset.select_related('asset', 'uploaded_by')

        # الفلترة
        attachment_type = self.request.GET.get('type')
        if attachment_type:
            queryset = queryset.filter(attachment_type=attachment_type)

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(asset__asset_number__icontains=search)
            )

        return queryset.order_by('-uploaded_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.kwargs.get('asset_id')
        asset = None
        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company
            )

        context.update({
            'title': f'مرفقات الأصل {asset.asset_number}' if asset else _('مرفقات الأصول'),
            'asset': asset,
            'can_add': self.request.user.has_perm('assets.change_asset'),
            'attachment_types': AssetAttachment.ATTACHMENT_TYPES,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
            ]
        })

        if asset:
            context['breadcrumbs'].extend([
                {'title': asset.asset_number, 'url': reverse('assets:asset_detail', args=[asset.pk])},
                {'title': _('المرفقات'), 'url': ''},
            ])
        else:
            context['breadcrumbs'].append({'title': _('المرفقات'), 'url': ''})

        return context


class AssetAttachmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, CreateView):
    """إضافة مرفق جديد"""

    model = AssetAttachment
    template_name = 'assets/assets/attachment_form.html'
    permission_required = 'assets.change_asset'
    fields = ['title', 'attachment_type', 'file', 'issue_date', 'expiry_date', 'description']

    def dispatch(self, request, *args, **kwargs):
        self.asset = get_object_or_404(
            Asset,
            pk=self.kwargs['asset_id'],
            company=request.current_company
        )
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'FileInput']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        form.instance.asset = self.asset
        form.instance.uploaded_by = self.request.user

        self.object = form.save()

        messages.success(
            self.request,
            f'✅ تم إضافة المرفق "{self.object.title}" بنجاح'
        )

        return redirect('assets:asset_detail', pk=self.asset.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'إضافة مرفق للأصل {self.asset.asset_number}',
            'asset': self.asset,
            'submit_text': _('رفع المرفق'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.asset.asset_number, 'url': reverse('assets:asset_detail', args=[self.asset.pk])},
                {'title': _('إضافة مرفق'), 'url': ''},
            ]
        })
        return context


class AssetAttachmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, UpdateView):
    """تعديل مرفق"""

    model = AssetAttachment
    template_name = 'assets/assets/attachment_form.html'
    permission_required = 'assets.change_asset'
    fields = ['title', 'attachment_type', 'file', 'issue_date', 'expiry_date', 'description']

    def get_queryset(self):
        return AssetAttachment.objects.filter(
            asset__company=self.request.current_company
        )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # جعل الملف اختياري في التعديل
        form.fields['file'].required = False

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'FileInput']:
                field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()

        messages.success(
            self.request,
            f'✅ تم تحديث المرفق "{self.object.title}" بنجاح'
        )

        return redirect('assets:asset_detail', pk=self.object.asset.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل المرفق: {self.object.title}',
            'asset': self.object.asset,
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.object.asset.asset_number,
                 'url': reverse('assets:asset_detail', args=[self.object.asset.pk])},
                {'title': _('تعديل مرفق'), 'url': ''},
            ]
        })
        return context


class AssetAttachmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف مرفق"""

    model = AssetAttachment
    template_name = 'assets/assets/attachment_confirm_delete.html'
    permission_required = 'assets.change_asset'

    def get_queryset(self):
        return AssetAttachment.objects.filter(
            asset__company=self.request.current_company
        )

    def get_success_url(self):
        return reverse('assets:asset_detail', kwargs={'pk': self.object.asset.pk})

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        title = self.object.title

        # حذف الملف من النظام
        if self.object.file:
            self.object.file.delete()

        messages.success(request, f'✅ تم حذف المرفق "{title}" بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Valuations ====================

class AssetValuationListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة إعادة تقييم الأصول"""

    model = AssetValuation
    template_name = 'assets/assets/valuation_list.html'
    context_object_name = 'valuations'
    permission_required = 'assets.view_asset'
    paginate_by = 50

    def get_queryset(self):
        asset_id = self.kwargs.get('asset_id')
        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company
            )
            queryset = AssetValuation.objects.filter(asset=asset)
        else:
            queryset = AssetValuation.objects.filter(
                asset__company=self.request.current_company
            )

        queryset = queryset.select_related(
            'asset', 'created_by', 'approved_by', 'journal_entry'
        )

        # الفلترة
        is_approved = self.request.GET.get('is_approved')
        if is_approved:
            queryset = queryset.filter(is_approved=is_approved == '1')

        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(valuation_date__gte=date_from)

        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(valuation_date__lte=date_to)

        return queryset.order_by('-valuation_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.kwargs.get('asset_id')
        asset = None
        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company
            )

        # إحصائيات
        valuations = self.get_queryset()
        stats = {
            'total': valuations.count(),
            'approved': valuations.filter(is_approved=True).count(),
            'pending': valuations.filter(is_approved=False).count(),
            'total_increase': valuations.filter(
                is_approved=True,
                difference__gt=0
            ).aggregate(total=Coalesce(Sum('difference'), Decimal('0')))['total'],
            'total_decrease': valuations.filter(
                is_approved=True,
                difference__lt=0
            ).aggregate(total=Coalesce(Sum('difference'), Decimal('0')))['total'],
        }

        context.update({
            'title': f'إعادة تقييم الأصل {asset.asset_number}' if asset else _('إعادة تقييم الأصول'),
            'asset': asset,
            'can_add': self.request.user.has_perm('assets.can_revalue_asset'),
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
            ]
        })

        if asset:
            context['breadcrumbs'].extend([
                {'title': asset.asset_number, 'url': reverse('assets:asset_detail', args=[asset.pk])},
                {'title': _('إعادة التقييم'), 'url': ''},
            ])
        else:
            context['breadcrumbs'].append({'title': _('إعادة التقييم'), 'url': ''})

        return context


class AssetValuationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, CreateView):
    """إنشاء إعادة تقييم"""

    model = AssetValuation
    template_name = 'assets/assets/valuation_form.html'
    permission_required = 'assets.can_revalue_asset'
    fields = [
        'valuation_date', 'new_value', 'reason',
        'valuator_name', 'valuation_report', 'notes'
    ]

    def dispatch(self, request, *args, **kwargs):
        self.asset = get_object_or_404(
            Asset,
            pk=self.kwargs['asset_id'],
            company=request.current_company
        )

        # التحقق من الحالة
        if self.asset.status not in ['active', 'inactive']:
            messages.error(
                request,
                '❌ لا يمكن إعادة تقييم أصل مباع أو مستبعد'
            )
            return redirect('assets:asset_detail', pk=self.asset.pk)

        return super().dispatch(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # القيمة الحالية
        form.fields['new_value'].help_text = f'القيمة الدفترية الحالية: {self.asset.book_value:,.2f}'

        # القيمة الافتراضية
        form.fields['valuation_date'].initial = date.today()

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'FileInput', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 4})

        return form

    @transaction.atomic
    def form_valid(self, form):
        form.instance.asset = self.asset
        form.instance.old_value = self.asset.book_value
        form.instance.created_by = self.request.user

        # حساب الفرق
        form.instance.difference = form.instance.new_value - form.instance.old_value

        try:
            self.object = form.save()

            messages.success(
                self.request,
                f'✅ تم إنشاء طلب إعادة التقييم بنجاح. الفرق: {self.object.difference:+,.2f}'
            )

            return redirect('assets:asset_detail', pk=self.asset.pk)

        except ValidationError as e:
            messages.error(self.request, f'❌ خطأ في التحقق: {e}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ حدث خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'إعادة تقييم الأصل {self.asset.asset_number}',
            'asset': self.asset,
            'submit_text': _('إنشاء طلب التقييم'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.asset.asset_number, 'url': reverse('assets:asset_detail', args=[self.asset.pk])},
                {'title': _('إعادة تقييم'), 'url': ''},
            ]
        })
        return context


class AssetValuationDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل إعادة التقييم"""

    model = AssetValuation
    template_name = 'assets/assets/valuation_detail.html'
    context_object_name = 'valuation'
    permission_required = 'assets.view_asset'

    def get_queryset(self):
        return AssetValuation.objects.filter(
            asset__company=self.request.current_company
        ).select_related(
            'asset', 'created_by', 'approved_by', 'journal_entry'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # نسبة التغيير
        if self.object.old_value > 0:
            change_percentage = (self.object.difference / self.object.old_value) * 100
        else:
            change_percentage = 0

        context.update({
            'title': f'إعادة تقييم: {self.object.asset.asset_number}',
            'can_approve': (
                    self.request.user.has_perm('assets.can_revalue_asset') and
                    not self.object.is_approved
            ),
            'change_percentage': change_percentage,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('الأصول'), 'url': reverse('assets:asset_list')},
                {'title': self.object.asset.asset_number,
                 'url': reverse('assets:asset_detail', args=[self.object.asset.pk])},
                {'title': _('تفاصيل التقييم'), 'url': ''},
            ]
        })
        return context


# ==================== Ajax Views ====================

@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_datatable_ajax(request):
    """Ajax endpoint لجدول الأصول - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر المخصصة
        status = request.GET.get('status', '')
        category = request.GET.get('category', '')
        branch = request.GET.get('branch', '')
        condition = request.GET.get('condition', '')

        # البحث والفلترة
        queryset = Asset.objects.filter(
            company=request.current_company
        ).select_related(
            'category', 'condition', 'cost_center', 'currency'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if category:
            queryset = queryset.filter(category_id=category)

        if branch:
            queryset = queryset.filter(branch_id=branch)

        if condition:
            queryset = queryset.filter(condition_id=condition)

        # البحث العام
        if search_value:
            queryset = queryset.filter(
                Q(asset_number__icontains=search_value) |
                Q(name__icontains=search_value) |
                Q(name_en__icontains=search_value) |
                Q(serial_number__icontains=search_value) |
                Q(barcode__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '1': 'asset_number',
            '2': 'name',
            '3': 'category__code',
            '4': 'purchase_date',
            '5': 'book_value',
            '6': 'status',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-asset_number')
        else:
            queryset = queryset.order_by('-asset_number')

        # العد الإجمالي
        total_records = Asset.objects.filter(company=request.current_company).count()
        filtered_records = queryset.count()

        # الصفحات
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_edit = request.user.has_perm('assets.change_asset')
        can_delete = request.user.has_perm('assets.delete_asset')
        can_view = request.user.has_perm('assets.view_asset')

        for asset in queryset:
            # حالة الأصل
            status_map = {
                'active': '<span class="badge bg-success"><i class="fas fa-check-circle"></i> نشط</span>',
                'inactive': '<span class="badge bg-secondary"><i class="fas fa-pause-circle"></i> غير نشط</span>',
                'under_maintenance': '<span class="badge bg-warning"><i class="fas fa-tools"></i> تحت الصيانة</span>',
                'disposed': '<span class="badge bg-danger"><i class="fas fa-trash"></i> مستبعد</span>',
                'sold': '<span class="badge bg-info"><i class="fas fa-dollar-sign"></i> مباع</span>',
                'lost': '<span class="badge bg-dark"><i class="fas fa-question-circle"></i> مفقود</span>',
                'damaged': '<span class="badge bg-danger"><i class="fas fa-exclamation-triangle"></i> تالف</span>',
            }
            status_badge = status_map.get(asset.status, asset.status)

            # أزرار الإجراءات
            actions = []

            # رابط العرض
            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:asset_detail', args=[asset.pk])}"
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            # رابط التعديل
            if can_edit and asset.status not in ['sold', 'disposed']:
                actions.append(f'''
                    <a href="{reverse('assets:asset_update', args=[asset.pk])}"
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            # رابط الحذف
            if can_delete and asset.status not in ['sold', 'disposed']:
                actions.append(f'''
                    <button type="button" class="btn btn-outline-danger btn-sm"
                            onclick="deleteAsset({asset.pk}, '{asset.asset_number}')"
                            title="حذف" data-bs-toggle="tooltip">
                        <i class="fas fa-trash"></i>
                    </button>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            # Checkbox
            checkbox_html = f'<input type="checkbox" class="asset-checkbox" value="{asset.pk}">'

            data.append([
                checkbox_html,
                f'<a href="{reverse("assets:asset_detail", args=[asset.pk])}">{asset.asset_number}</a>',
                f'''<div>
                    <strong>{asset.name}</strong>
                    {f'<br><small class="text-muted">{asset.name_en}</small>' if asset.name_en else ''}
                </div>''',
                f'''<div>
                    <span class="badge bg-light text-dark">{asset.category.code}</span>
                    <small class="d-block text-muted">{asset.category.name}</small>
                </div>''',
                asset.purchase_date.strftime('%Y-%m-%d'),
                f'''<div class="text-end">
                    <div><strong>{asset.book_value:,.2f}</strong> {asset.currency.code if asset.currency else ''}</div>
                    <small class="text-muted">من أصل {asset.original_cost:,.2f}</small>
                </div>''',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_autocomplete(request):
    """Autocomplete للأصول - محسّن"""

    try:
        term = request.GET.get('term', '').strip()

        if len(term) < 2:
            return JsonResponse([], safe=False)

        # فلترة إضافية
        only_active = request.GET.get('only_active', '1') == '1'
        category_id = request.GET.get('category_id')

        assets = Asset.objects.filter(
            company=request.current_company,
        ).filter(
            Q(asset_number__icontains=term) |
            Q(name__icontains=term) |
            Q(name_en__icontains=term) |
            Q(barcode__icontains=term) |
            Q(serial_number__icontains=term)
        )

        # فلترة الأصول النشطة
        if only_active:
            assets = assets.filter(status='active')

        # فلترة حسب الفئة
        if category_id:
            assets = assets.filter(category_id=category_id)

        assets = assets.select_related('category', 'condition', 'currency')[:20]

        results = []
        for asset in assets:
            results.append({
                'id': asset.id,
                'text': f"{asset.asset_number} - {asset.name}",
                'asset_number': asset.asset_number,
                'name': asset.name,
                'name_en': asset.name_en or '',
                'category': {
                    'id': asset.category.id,
                    'code': asset.category.code,
                    'name': asset.category.name,
                },
                'condition': asset.condition.name if asset.condition else None,
                'book_value': float(asset.book_value),
                'original_cost': float(asset.original_cost),
                'currency': asset.currency.code if asset.currency else '',
                'status': asset.status,
                'purchase_date': asset.purchase_date.strftime('%Y-%m-%d'),
            })

        return JsonResponse(results, safe=False)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_stats_ajax(request):
    """إحصائيات الأصول - للبطاقات الإحصائية"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({'success': False, 'error': 'لا توجد شركة محددة'}, status=400)

    try:
        assets = Asset.objects.filter(company=request.current_company)

        stats = {
            'total_assets': assets.count(),
            'active_assets': assets.filter(status='active').count(),
            'total_cost': float(assets.aggregate(
                total=Coalesce(Sum('original_cost'), Decimal('0'))
            )['total']),
            'total_book_value': float(assets.aggregate(
                total=Coalesce(Sum('book_value'), Decimal('0'))
            )['total']),
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_export(request):
    """تصدير الأصول إلى Excel"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:asset_list')

    try:
        # الفلاتر
        status = request.GET.get('status', '')
        category = request.GET.get('category', '')
        condition = request.GET.get('condition', '')
        search_filter = request.GET.get('search_filter', '')

        # البحث والفلترة
        queryset = Asset.objects.filter(
            company=request.current_company
        ).select_related(
            'category', 'condition', 'cost_center', 'currency',
            'responsible_employee', 'depreciation_method'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if category:
            queryset = queryset.filter(category_id=category)

        if condition:
            queryset = queryset.filter(condition_id=condition)

        if search_filter:
            queryset = queryset.filter(
                Q(asset_number__icontains=search_filter) |
                Q(name__icontains=search_filter) |
                Q(name_en__icontains=search_filter) |
                Q(serial_number__icontains=search_filter) |
                Q(barcode__icontains=search_filter)
            )

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأصول"

        # العناوين
        headers = [
            'رقم الأصل', 'الاسم', 'الاسم بالإنجليزية', 'التصنيف',
            'الحالة', 'تاريخ الشراء', 'التكلفة الأصلية', 'القيمة الدفترية',
            'العملة', 'حالة الأصل', 'طريقة الإهلاك', 'العمر الإنتاجي (شهر)',
            'الموقع الفعلي', 'الرقم التسلسلي', 'الموديل', 'الشركة المصنعة',
            'الباركود', 'المسؤول', 'مركز التكلفة', 'ملاحظات'
        ]

        # كتابة العناوين
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # كتابة البيانات
        for row_num, asset in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=asset.asset_number)
            ws.cell(row=row_num, column=2, value=asset.name)
            ws.cell(row=row_num, column=3, value=asset.name_en or '')
            ws.cell(row=row_num, column=4, value=f"{asset.category.code} - {asset.category.name}")
            ws.cell(row=row_num, column=5, value=asset.condition.name if asset.condition else '')
            ws.cell(row=row_num, column=6, value=asset.purchase_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=7, value=float(asset.original_cost))
            ws.cell(row=row_num, column=8, value=float(asset.book_value))
            ws.cell(row=row_num, column=9, value=asset.currency.code if asset.currency else '')
            ws.cell(row=row_num, column=10, value=asset.get_status_display())
            ws.cell(row=row_num, column=11, value=asset.depreciation_method.name if asset.depreciation_method else '')
            ws.cell(row=row_num, column=12, value=asset.useful_life_months or '')
            ws.cell(row=row_num, column=13, value=asset.physical_location or '')
            ws.cell(row=row_num, column=14, value=asset.serial_number or '')
            ws.cell(row=row_num, column=15, value=asset.model or '')
            ws.cell(row=row_num, column=16, value=asset.manufacturer or '')
            ws.cell(row=row_num, column=17, value=asset.barcode or '')
            ws.cell(row=row_num, column=18, value=asset.responsible_employee.get_full_name() if asset.responsible_employee else '')
            ws.cell(row=row_num, column=19, value=asset.cost_center.name if asset.cost_center else '')
            ws.cell(row=row_num, column=20, value=asset.notes or '')

        # تعديل عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="assets_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'حدث خطأ أثناء تصدير البيانات: {str(e)}')
        return redirect('assets:asset_list')


@login_required
@permission_required_with_message('assets.view_assetcategory')
@require_http_methods(["GET"])
def asset_category_datatable_ajax(request):
    """Ajax endpoint لجدول فئات الأصول - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الحصول على معايير الفلترة الإضافية
        parent_filter = request.GET.get('parent', '')
        level_filter = request.GET.get('level', '')
        is_active_filter = request.GET.get('is_active', '')
        search_filter = request.GET.get('search_filter', '')

        # البحث والفلترة
        queryset = AssetCategory.objects.filter(
            company=request.current_company
        ).select_related('parent').annotate(
            asset_count=Count('assets', filter=Q(assets__status='active')),
            total_value=Coalesce(
                Sum('assets__book_value', filter=Q(assets__status='active')),
                Decimal('0')
            )
        )

        # فلتر الفئة الأب
        if parent_filter == 'null':
            queryset = queryset.filter(parent__isnull=True)
        elif parent_filter:
            queryset = queryset.filter(parent_id=parent_filter)

        # فلتر المستوى
        if level_filter:
            queryset = queryset.filter(level=int(level_filter))

        # فلتر الحالة
        if is_active_filter:
            queryset = queryset.filter(is_active=bool(int(is_active_filter)))

        # البحث النصي
        if search_filter:
            queryset = queryset.filter(
                Q(code__icontains=search_filter) |
                Q(name__icontains=search_filter) |
                Q(name_en__icontains=search_filter)
            )

        if search_value:
            queryset = queryset.filter(
                Q(code__icontains=search_value) |
                Q(name__icontains=search_value) |
                Q(name_en__icontains=search_value)
            )

        queryset = queryset.order_by('code')

        total_records = AssetCategory.objects.filter(company=request.current_company).count()
        filtered_records = queryset.count()

        queryset = queryset[start:start + length]

        data = []
        can_edit = request.user.has_perm('assets.change_assetcategory')
        can_delete = request.user.has_perm('assets.delete_assetcategory')
        can_view = request.user.has_perm('assets.view_assetcategory')

        for category in queryset:
            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:category_detail', args=[category.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            if can_edit:
                actions.append(f'''
                    <a href="{reverse('assets:category_update', args=[category.pk])}" 
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            if can_delete and category.asset_count == 0:
                actions.append(f'''
                    <button type="button" class="btn btn-outline-danger btn-sm" 
                            onclick="deleteCategory({category.pk}, '{category.code}')" 
                            title="حذف" data-bs-toggle="tooltip">
                        <i class="fas fa-trash"></i>
                    </button>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            # تحديد لون شارة المستوى
            level_colors = {
                0: 'linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%)',
                1: 'linear-gradient(135deg, #3b82f6 0%, #2563eb 100%)',
                2: 'linear-gradient(135deg, #10b981 0%, #059669 100%)',
                3: 'linear-gradient(135deg, #f59e0b 0%, #d97706 100%)'
            }
            level_bg = level_colors.get(category.level, 'linear-gradient(135deg, #6c757d 0%, #495057 100%)')

            # شارة الحالة
            status_badge = '<span class="badge bg-success">نشط</span>' if category.is_active else '<span class="badge bg-secondary">غير نشط</span>'

            data.append([
                f'<a href="{reverse("assets:category_detail", args=[category.pk])}">{category.code}</a>',
                f'''<div>
                    <strong>{category.name}</strong>
                    {f'<br><small class="text-muted">{category.name_en}</small>' if category.name_en else ''}
                </div>''',
                category.parent.name if category.parent else '<span class="text-muted">-</span>',
                f'<span class="category-level-badge" style="background: {level_bg}; color: white; padding: 0.35rem 0.65rem; border-radius: 50%; display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; font-weight: 600; font-size: 0.875rem;">{category.level}</span>',
                f'<span class="badge bg-primary">{category.asset_count}</span>',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetcategory')
@require_http_methods(["GET"])
def category_tree_ajax(request):
    """شجرة الفئات - محسّنة"""
    try:
        categories = AssetCategory.objects.filter(
            company=request.current_company,
            is_active=True
        ).select_related('parent').annotate(
            asset_count=Count('assets', filter=Q(assets__status='active'))
        )

        def build_tree(parent=None):
            items = []
            cats = categories.filter(parent=parent)
            for cat in cats:
                item = {
                    'id': cat.id,
                    'text': f"{cat.code} - {cat.name}",
                    'asset_count': cat.asset_count,
                    'state': {
                        'opened': parent is None  # فتح المستوى الأول فقط
                    },
                    'children': build_tree(cat)
                }

                # إضافة أيقونة حسب المستوى
                if cat.asset_count > 0:
                    item['icon'] = 'fas fa-folder-open text-primary'
                else:
                    item['icon'] = 'fas fa-folder text-secondary'

                items.append(item)
            return items

        tree = build_tree()
        return JsonResponse({'success': True, 'tree': tree})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_assetcategory')
@require_http_methods(["GET"])
def category_stats_ajax(request):
    """إحصائيات فئات الأصول - للبطاقات الإحصائية"""
    try:
        company = request.current_company

        # إحصائيات الفئات
        categories = AssetCategory.objects.filter(company=company)

        stats = {
            'total_categories': categories.count(),
            'parent_categories': categories.filter(parent__isnull=True).count(),
            'child_categories': categories.filter(parent__isnull=False).count(),
            'total_assets': Asset.objects.filter(
                company=company,
                category__isnull=False,
                status='active'
            ).count()
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_assetcategory')
@require_http_methods(["GET"])
def category_export(request):
    """تصدير فئات الأصول إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    try:
        company = request.current_company

        # الحصول على معايير الفلترة
        parent_id = request.GET.get('parent', '')
        level = request.GET.get('level', '')
        is_active = request.GET.get('is_active', '')
        search = request.GET.get('search', '')

        # بناء الاستعلام
        queryset = AssetCategory.objects.filter(company=company).select_related('parent').annotate(
            assets_count=Count('assets', filter=Q(assets__status='active'))
        )

        # تطبيق الفلاتر
        if parent_id == 'null':
            queryset = queryset.filter(parent__isnull=True)
        elif parent_id:
            queryset = queryset.filter(parent_id=parent_id)

        if level:
            queryset = queryset.filter(level=int(level))

        if is_active:
            queryset = queryset.filter(is_active=bool(int(is_active)))

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        queryset = queryset.order_by('code')

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "فئات الأصول"

        # تنسيق الرأس
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # عناوين الأعمدة
        headers = [
            'الرمز',
            'الاسم بالعربية',
            'الاسم بالإنجليزية',
            'الفئة الأب',
            'المستوى',
            'عدد الأصول',
            'الحالة',
            'الوصف'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # تعبئة البيانات
        row_num = 2
        for category in queryset:
            ws.cell(row=row_num, column=1, value=category.code)
            ws.cell(row=row_num, column=2, value=category.name)
            ws.cell(row=row_num, column=3, value=category.name_en or '')
            ws.cell(row=row_num, column=4, value=category.parent.name if category.parent else '-')
            ws.cell(row=row_num, column=5, value=category.level)
            ws.cell(row=row_num, column=6, value=category.assets_count)
            ws.cell(row=row_num, column=7, value='نشط' if category.is_active else 'غير نشط')
            ws.cell(row=row_num, column=8, value=category.description or '')
            row_num += 1

        # ضبط عرض الأعمدة
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # حفظ الملف
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'asset_categories_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, f'حدث خطأ أثناء التصدير: {str(e)}')
        return redirect('assets:category_list')


@login_required
@permission_required_with_message('assets.add_asset')
@require_http_methods(["GET"])
def generate_asset_number(request):
    """توليد رقم أصل تلقائي - محسّن"""
    try:
        category_id = request.GET.get('category_id')
        prefix = 'AST'

        if category_id:
            try:
                category = AssetCategory.objects.get(
                    pk=category_id,
                    company=request.current_company
                )
                prefix = category.code
            except AssetCategory.DoesNotExist:
                pass

        # البحث عن آخر رقم
        last_asset = Asset.objects.filter(
            company=request.current_company,
            asset_number__startswith=prefix
        ).order_by('-asset_number').first()

        if last_asset:
            try:
                # استخراج الرقم من نهاية رقم الأصل
                last_number_str = last_asset.asset_number.replace(prefix, '')
                last_number = int(last_number_str)
                new_number = f"{prefix}{last_number + 1:05d}"
            except (ValueError, TypeError):
                new_number = f"{prefix}00001"
        else:
            new_number = f"{prefix}00001"

        return JsonResponse({
            'success': True,
            'asset_number': new_number,
            'prefix': prefix
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_barcode_pdf(request, pk):
    """طباعة باركود الأصل PDF - محسّن"""
    try:
        asset = get_object_or_404(
            Asset,
            pk=pk,
            company=request.current_company
        )

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="barcode_{asset.asset_number}.pdf"'

        # إنشاء PDF
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4

        # العنوان
        p.setFont("Helvetica-Bold", 16)
        p.drawString(200, height - 50, f"Asset Barcode")

        # معلومات الأصل
        p.setFont("Helvetica", 12)
        y_position = height - 100

        info_items = [
            f"Asset Number: {asset.asset_number}",
            f"Name: {asset.name}",
            f"Category: {asset.category.name}",
            f"Purchase Date: {asset.purchase_date.strftime('%Y-%m-%d')}",
            f"Book Value: {asset.book_value:,.2f} {asset.currency.code if asset.currency else ''}",
        ]

        for item in info_items:
            p.drawString(100, y_position, item)
            y_position -= 25

        # توليد الباركود
        if asset.barcode:
            try:
                from barcode import Code128
                from barcode.writer import ImageWriter

                code = Code128(asset.barcode, writer=ImageWriter())
                buffer = BytesIO()
                code.write(buffer)
                buffer.seek(0)

                # إضافة الباركود للـ PDF
                p.drawImage(buffer, 100, y_position - 150, width=400, height=100)
            except Exception as barcode_error:
                p.drawString(100, y_position - 50, f"Error generating barcode: {str(barcode_error)}")

        # تذييل
        p.setFont("Helvetica", 8)
        p.drawString(100, 50, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        p.showPage()
        p.save()

        return response

    except Exception as e:
        messages.error(request, f'❌ خطأ في إنشاء PDF: {str(e)}')
        return redirect('assets:asset_detail', pk=pk)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def asset_qr_code(request, pk):
    """QR Code للأصل - محسّن"""
    try:
        asset = get_object_or_404(
            Asset,
            pk=pk,
            company=request.current_company
        )

        # بيانات QR
        qr_data = json.dumps({
            'asset_number': asset.asset_number,
            'name': asset.name,
            'category': asset.category.code,
            'book_value': float(asset.book_value),
            'url': request.build_absolute_uri(
                reverse('assets:asset_detail', args=[asset.pk])
            )
        }, ensure_ascii=False)

        # إنشاء QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        return HttpResponse(buffer.read(), content_type='image/png')

    except Exception as e:
        return HttpResponse(status=500)


@login_required
@permission_required_with_message('assets.change_asset')
@require_http_methods(["POST"])
def upload_attachment(request, pk):
    """رفع مرفق - محسّن"""
    try:
        asset = get_object_or_404(
            Asset,
            pk=pk,
            company=request.current_company
        )

        file = request.FILES.get('file')
        title = request.POST.get('title', '')
        attachment_type = request.POST.get('attachment_type', 'other')
        description = request.POST.get('description', '')

        if not file:
            return JsonResponse({
                'success': False,
                'error': 'لا يوجد ملف'
            }, status=400)

        if not title:
            title = file.name

        attachment = AssetAttachment.objects.create(
            asset=asset,
            title=title,
            attachment_type=attachment_type,
            file=file,
            description=description,
            uploaded_by=request.user
        )

        return JsonResponse({
            'success': True,
            'id': attachment.id,
            'title': attachment.title,
            'file_name': file.name,
            'file_size': file.size,
            'uploaded_at': attachment.uploaded_at.strftime('%Y-%m-%d %H:%M')
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.change_asset')
@require_http_methods(["POST", "DELETE"])
def delete_attachment(request, pk):
    """حذف مرفق"""
    try:
        attachment = get_object_or_404(
            AssetAttachment,
            pk=pk,
            asset__company=request.current_company
        )

        title = attachment.title

        # حذف الملف
        if attachment.file:
            attachment.file.delete()

        attachment.delete()

        return JsonResponse({
            'success': True,
            'message': f'تم حذف المرفق "{title}" بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.add_asset')
@require_http_methods(["POST"])
def bulk_import_assets(request):
    """استيراد جماعي للأصول - محسّن"""
    try:
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({
                'success': False,
                'error': 'لا يوجد ملف'
            }, status=400)

        # قراءة الملف
        df = pd.read_excel(file)

        # التحقق من الأعمدة المطلوبة
        required_columns = ['asset_number', 'name', 'category_code', 'original_cost', 'purchase_date']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return JsonResponse({
                'success': False,
                'error': f'أعمدة مفقودة: {", ".join(missing_columns)}'
            }, status=400)

        success_count = 0
        error_count = 0
        errors = []

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # البحث عن الفئة
                    category = AssetCategory.objects.get(
                        company=request.current_company,
                        code=row['category_code']
                    )

                    # إنشاء الأصل
                    Asset.objects.create(
                        company=request.current_company,
                        branch=request.current_branch,
                        asset_number=row['asset_number'],
                        name=row['name'],
                        category=category,
                        original_cost=row['original_cost'],
                        purchase_date=pd.to_datetime(row['purchase_date']).date(),
                        depreciation_start_date=pd.to_datetime(row['purchase_date']).date(),
                        currency=request.current_company.base_currency,
                        depreciation_method=category.default_depreciation_method or DepreciationMethod.objects.first(),
                        useful_life_months=row.get('useful_life_months', category.default_useful_life_months or 60),
                        salvage_value=row.get('salvage_value', 0),
                        created_by=request.user
                    )
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"السطر {index + 2}: {str(e)}")

        return JsonResponse({
            'success': True,
            'imported': success_count,
            'errors': error_count,
            'error_details': errors[:10]  # أول 10 أخطاء فقط
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def download_import_template(request):
    """تحميل قالب الاستيراد - محسّن"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets Template"

        # Headers
        headers = [
            'asset_number',
            'name',
            'category_code',
            'original_cost',
            'purchase_date',
            'useful_life_months',
            'salvage_value',
            'serial_number',
            'description'
        ]

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # مثال
        example_row = [
            'AST00001',
            'كمبيوتر محمول Dell',
            'COMP',
            '5000',
            '2024-01-01',
            '60',
            '500',
            'SN123456',
            'كمبيوتر للموظف'
        ]
        ws.append(example_row)

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="assets_template.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f'❌ خطأ في إنشاء القالب: {str(e)}')
        return redirect('assets:asset_list')


@login_required
@permission_required_with_message('assets.change_asset')
@require_http_methods(["POST"])
def bulk_update_status(request):
    """تحديث حالة جماعي"""
    try:
        asset_ids = request.POST.getlist('asset_ids[]')
        status = request.POST.get('status')

        if not asset_ids:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أصول'
            }, status=400)

        updated = Asset.objects.filter(
            id__in=asset_ids,
            company=request.current_company
        ).update(status=status)

        return JsonResponse({
            'success': True,
            'updated': updated,
            'message': f'تم تحديث حالة {updated} أصول بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.change_asset')
@require_http_methods(["POST"])
def bulk_update_location(request):
    """تحديث موقع جماعي"""
    try:
        asset_ids = request.POST.getlist('asset_ids[]')
        location = request.POST.get('location')

        if not asset_ids:
            return JsonResponse({
                'success': False,
                'error': 'لم يتم تحديد أصول'
            }, status=400)

        updated = Asset.objects.filter(
            id__in=asset_ids,
            company=request.current_company
        ).update(physical_location=location)

        return JsonResponse({
            'success': True,
            'updated': updated,
            'message': f'تم تحديث موقع {updated} أصول بنجاح'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== Depreciation Method AJAX ====================

@login_required
@permission_required_with_message('assets.view_depreciationmethod')
@require_http_methods(["GET"])
def depreciation_method_datatable_ajax(request):
    """Ajax endpoint لجدول طرق الإهلاك - Server-side"""

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الحصول على معايير الفلترة الإضافية
        method_type_filter = request.GET.get('method_type', '')
        is_active_filter = request.GET.get('is_active', '')
        search_filter = request.GET.get('search_filter', '')

        # البحث والفلترة
        queryset = DepreciationMethod.objects.all().annotate(
            assets_count=Count('asset', filter=Q(asset__status='active'))
        )

        # فلتر نوع الطريقة
        if method_type_filter:
            queryset = queryset.filter(method_type=method_type_filter)

        # فلتر الحالة
        if is_active_filter:
            queryset = queryset.filter(is_active=bool(int(is_active_filter)))

        # البحث النصي
        if search_filter:
            queryset = queryset.filter(
                Q(code__icontains=search_filter) |
                Q(name__icontains=search_filter) |
                Q(name_en__icontains=search_filter)
            )

        # البحث من DataTable
        if search_value:
            queryset = queryset.filter(
                Q(code__icontains=search_value) |
                Q(name__icontains=search_value) |
                Q(name_en__icontains=search_value)
            )

        # العدد الإجمالي
        records_total = DepreciationMethod.objects.count()
        records_filtered = queryset.count()

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_direction = request.GET.get('order[0][dir]', 'asc')

        if order_column_index:
            order_columns = ['code', 'name', 'method_type', 'rate_percentage', 'assets_count', 'is_active']
            order_column = order_columns[int(order_column_index)]

            if order_direction == 'desc':
                order_column = '-' + order_column

            queryset = queryset.order_by(order_column)
        else:
            queryset = queryset.order_by('code')

        # الـ Pagination
        methods = queryset[start:start + length]

        # بناء البيانات
        data = []
        for method in methods:
            # نوع الطريقة مع badge
            if method.method_type == 'straight_line':
                method_type_html = '<span class="badge bg-primary"><i class="fas fa-minus"></i> {}</span>'.format(method.get_method_type_display())
            elif method.method_type == 'declining_balance':
                method_type_html = '<span class="badge bg-warning"><i class="fas fa-chart-line"></i> {}</span>'.format(method.get_method_type_display())
            elif method.method_type == 'units_of_production':
                method_type_html = '<span class="badge bg-info"><i class="fas fa-industry"></i> {}</span>'.format(method.get_method_type_display())
            else:
                method_type_html = '<span class="badge bg-secondary">{}</span>'.format(method.get_method_type_display())

            # النسبة
            rate_html = '{}%'.format(method.rate_percentage) if method.rate_percentage else '-'

            # عدد الأصول
            assets_count_html = '<span class="badge bg-primary">{}</span>'.format(method.assets_count)

            # الحالة
            if method.is_active:
                status_html = '<span class="badge bg-success"><i class="fas fa-check-circle"></i> نشط</span>'
            else:
                status_html = '<span class="badge bg-secondary"><i class="fas fa-times-circle"></i> غير نشط</span>'

            # الإجراءات
            actions_html = '''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="/assets/depreciation-methods/{}/"
                       class="btn btn-sm btn-outline-info"
                       data-bs-toggle="tooltip"
                       title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
            '''.format(method.pk)

            if request.user.has_perm('assets.change_depreciationmethod'):
                actions_html += '''
                    <a href="/assets/depreciation-methods/{}/update/"
                       class="btn btn-sm btn-outline-primary"
                       data-bs-toggle="tooltip"
                       title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                '''.format(method.pk)

            if request.user.has_perm('assets.delete_depreciationmethod'):
                actions_html += '''
                    <button type="button"
                            class="btn btn-sm btn-outline-danger"
                            onclick="deleteMethod({}, '{}')"
                            data-bs-toggle="tooltip"
                            title="حذف">
                        <i class="fas fa-trash"></i>
                    </button>
                '''.format(method.pk, method.name)

            actions_html += '</div>'

            data.append([
                '<strong>{}</strong>'.format(method.code),
                method.name,
                method_type_html,
                rate_html,
                assets_count_html,
                status_html,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data
        })

    except Exception as e:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.view_depreciationmethod')
@require_http_methods(["GET"])
def depreciation_method_stats_ajax(request):
    """إحصائيات طرق الإهلاك - للبطاقات الإحصائية"""
    try:
        # إحصائيات الطرق
        methods = DepreciationMethod.objects.all()

        stats = {
            'total_methods': methods.count(),
            'active_methods': methods.filter(is_active=True).count(),
            'straight_line_methods': methods.filter(method_type='straight_line').count(),
            'total_assets': Asset.objects.filter(
                depreciation_method__isnull=False,
                status='active'
            ).count()
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_depreciationmethod')
@require_http_methods(["GET"])
def depreciation_method_export(request):
    """تصدير طرق الإهلاك إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    try:
        # الحصول على معايير الفلترة
        method_type = request.GET.get('method_type', '')
        is_active = request.GET.get('is_active', '')
        search = request.GET.get('search', '')

        # بناء الاستعلام
        queryset = DepreciationMethod.objects.all().annotate(
            assets_count=Count('asset', filter=Q(asset__status='active'))
        )

        # تطبيق الفلاتر
        if method_type:
            queryset = queryset.filter(method_type=method_type)

        if is_active:
            queryset = queryset.filter(is_active=bool(int(is_active)))

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        queryset = queryset.order_by('code')

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "طرق الإهلاك"

        # تنسيق الرأس
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # الرؤوس
        headers = ['الرمز', 'الاسم', 'الاسم بالإنجليزية', 'نوع الطريقة', 'نسبة المعدل %', 'عدد الأصول', 'الحالة']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # البيانات
        row_num = 2
        for method in queryset:
            ws.cell(row=row_num, column=1, value=method.code)
            ws.cell(row=row_num, column=2, value=method.name)
            ws.cell(row=row_num, column=3, value=method.name_en or '')
            ws.cell(row=row_num, column=4, value=method.get_method_type_display())
            ws.cell(row=row_num, column=5, value=method.rate_percentage or '')
            ws.cell(row=row_num, column=6, value=method.assets_count)
            ws.cell(row=row_num, column=7, value='نشط' if method.is_active else 'غير نشط')
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="depreciation_methods_{}.xlsx"'.format(datetime.now().strftime("%Y%m%d"))

        return response

    except Exception as e:
        messages.error(request, 'خطأ في التصدير: {}'.format(str(e)))
        return redirect('assets:depreciation_method_list')


# ==================== Asset Condition AJAX ====================

@login_required
@permission_required_with_message('assets.view_assetcondition')
@require_http_methods(["GET"])
def condition_datatable_ajax(request):
    """Ajax endpoint لجدول حالات الأصول - Server-side"""

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الحصول على معايير الفلترة الإضافية
        is_active_filter = request.GET.get('is_active', '')
        search_filter = request.GET.get('search_filter', '')

        # البحث والفلترة
        queryset = AssetCondition.objects.all().annotate(
            assets_count=Count('assets', filter=Q(assets__status='active'))
        )

        # فلتر الحالة
        if is_active_filter:
            queryset = queryset.filter(is_active=bool(int(is_active_filter)))

        # البحث النصي
        if search_filter:
            queryset = queryset.filter(
                Q(name__icontains=search_filter) |
                Q(name_en__icontains=search_filter)
            )

        # البحث من DataTable
        if search_value:
            queryset = queryset.filter(
                Q(name__icontains=search_value) |
                Q(name_en__icontains=search_value)
            )

        # العدد الإجمالي
        records_total = AssetCondition.objects.count()
        records_filtered = queryset.count()

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_direction = request.GET.get('order[0][dir]', 'asc')

        if order_column_index:
            order_columns = ['name', 'name_en', 'color_code', 'assets_count', 'is_active']
            order_column = order_columns[int(order_column_index)]

            if order_direction == 'desc':
                order_column = '-' + order_column

            queryset = queryset.order_by(order_column)
        else:
            queryset = queryset.order_by('name')

        # الـ Pagination
        conditions = queryset[start:start + length]

        # بناء البيانات
        data = []
        for condition in conditions:
            # اللون مع معاينة
            color_html = '<span class="badge" style="background-color: {0}; color: white; padding: 0.5rem 1rem;">{0}</span>'.format(condition.color_code)

            # عدد الأصول
            assets_count_html = '<span class="badge bg-primary">{}</span>'.format(condition.assets_count)

            # الحالة
            if condition.is_active:
                status_html = '<span class="badge bg-success"><i class="fas fa-check-circle"></i> نشط</span>'
            else:
                status_html = '<span class="badge bg-secondary"><i class="fas fa-times-circle"></i> غير نشط</span>'

            # الإجراءات
            actions_html = '''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="/assets/conditions/{}/view/"
                       class="btn btn-sm btn-outline-info"
                       data-bs-toggle="tooltip"
                       title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
            '''.format(condition.pk)

            if request.user.has_perm('assets.change_assetcondition'):
                actions_html += '''
                    <a href="/assets/conditions/{}/update/"
                       class="btn btn-sm btn-outline-primary"
                       data-bs-toggle="tooltip"
                       title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                '''.format(condition.pk)

            if request.user.has_perm('assets.delete_assetcondition'):
                actions_html += '''
                    <button type="button"
                            class="btn btn-sm btn-outline-danger"
                            onclick="deleteCondition({}, '{}')"
                            data-bs-toggle="tooltip"
                            title="حذف">
                        <i class="fas fa-trash"></i>
                    </button>
                '''.format(condition.pk, condition.name)

            actions_html += '</div>'

            data.append([
                '<strong>{}</strong>'.format(condition.name),
                condition.name_en or '-',
                color_html,
                assets_count_html,
                status_html,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data
        })

    except Exception as e:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetcondition')
@require_http_methods(["GET"])
def condition_stats_ajax(request):
    """إحصائيات حالات الأصول - للبطاقات الإحصائية"""
    try:
        # إحصائيات الحالات
        conditions = AssetCondition.objects.all()

        # حساب متوسط قيمة الأصول
        assets_with_condition = Asset.objects.filter(
            condition__isnull=False,
            status='active'
        )
        avg_value = assets_with_condition.aggregate(
            avg=Avg('original_cost')
        )['avg'] or 0

        stats = {
            'total_conditions': conditions.count(),
            'active_conditions': conditions.filter(is_active=True).count(),
            'total_assets': assets_with_condition.count(),
            'avg_value': float(avg_value)
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_assetcondition')
@require_http_methods(["GET"])
def condition_export(request):
    """تصدير حالات الأصول إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    try:
        # الحصول على معايير الفلترة
        is_active = request.GET.get('is_active', '')
        search = request.GET.get('search', '')

        # بناء الاستعلام
        queryset = AssetCondition.objects.all().annotate(
            assets_count=Count('assets', filter=Q(assets__status='active'))
        )

        # تطبيق الفلاتر
        if is_active:
            queryset = queryset.filter(is_active=bool(int(is_active)))

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        queryset = queryset.order_by('name')

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "حالات الأصول"

        # تنسيق الرأس
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # الرؤوس
        headers = ['الاسم', 'الاسم بالإنجليزية', 'رمز اللون', 'عدد الأصول', 'الحالة']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # البيانات
        row_num = 2
        for condition in queryset:
            ws.cell(row=row_num, column=1, value=condition.name)
            ws.cell(row=row_num, column=2, value=condition.name_en or '')
            ws.cell(row=row_num, column=3, value=condition.color_code)
            ws.cell(row=row_num, column=4, value=condition.assets_count)
            ws.cell(row=row_num, column=5, value='نشط' if condition.is_active else 'غير نشط')
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="asset_conditions_{}.xlsx"'.format(datetime.now().strftime("%Y%m%d"))

        return response

    except Exception as e:
        messages.error(request, 'خطأ في التصدير: {}'.format(str(e)))
        return redirect('assets:condition_list')
