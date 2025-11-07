# apps/assets/views/insurance_views.py
"""
Views إدارة التأمين على الأصول - محسّنة وشاملة
- إدارة شركات التأمين
- إدارة البوليصات
- إدارة المطالبات
- تجديد البوليصات
- تقارير التأمين
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    DetailView, TemplateView
)
from django.db.models import (
    Q, Sum, Count, Avg, Max, Min, F,
    DecimalField, Case, When, Value
)
from django.db.models.functions import Coalesce, TruncMonth
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
from django import forms
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
from dateutil.relativedelta import relativedelta

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import (
    InsuranceCompany, AssetInsurance, InsuranceClaim,
    Asset, AssetCategory
)
from ..forms.insurance_forms import (
    InsuranceCompanyForm, AssetInsuranceForm, InsuranceClaimForm
)


# ==================== Insurance Companies ====================

class InsuranceCompanyListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة شركات التأمين - محسّنة"""

    model = InsuranceCompany
    template_name = 'assets/insurance/company_list.html'
    context_object_name = 'companies'
    permission_required = 'assets.view_insurancecompany'
    paginate_by = 50

    def get_queryset(self):
        queryset = InsuranceCompany.objects.filter(
            company=self.request.current_company
        ).annotate(
            active_policies_count=Count(
                'insurance_policies',
                filter=Q(insurance_policies__status='active')
            ),
            total_coverage=Coalesce(
                Sum(
                    'insurance_policies__coverage_amount',
                    filter=Q(insurance_policies__status='active')
                ),
                Decimal('0')
            ),
            total_premiums=Coalesce(
                Sum(
                    'insurance_policies__premium_amount',
                    filter=Q(insurance_policies__status='active')
                ),
                Decimal('0')
            )
        )

        # الفلترة
        search = self.request.GET.get('search')
        has_active = self.request.GET.get('has_active')

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )

        if has_active == '1':
            queryset = queryset.filter(active_policies_count__gt=0)

        # الترتيب
        sort_by = self.request.GET.get('sort', 'name')
        queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات
        companies = InsuranceCompany.objects.filter(
            company=self.request.current_company
        )

        stats = {
            'total_companies': companies.count(),
            'with_active_policies': companies.annotate(
                count=Count(
                    'insurance_policies',
                    filter=Q(insurance_policies__status='active')
                )
            ).filter(count__gt=0).count(),
        }

        context.update({
            'title': _('شركات التأمين'),
            'can_add': self.request.user.has_perm('assets.add_insurancecompany'),
            'can_edit': self.request.user.has_perm('assets.change_insurancecompany'),
            'can_delete': self.request.user.has_perm('assets.delete_insurancecompany'),
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('شركات التأمين'), 'url': ''},
            ]
        })
        return context


class InsuranceCompanyCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء شركة تأمين - محسّن"""

    model = InsuranceCompany
    template_name = 'assets/insurance/company_form.html'
    permission_required = 'assets.add_insurancecompany'
    form_class = InsuranceCompanyForm
    success_url = reverse_lazy('assets:insurance_company_list')

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.created_by = self.request.user
            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء شركة التأمين {self.object.name} بنجاح'
            )

            # Check if "save and add another" was clicked
            if 'save_and_add' in self.request.POST:
                return redirect('assets:insurance_company_create')

            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة شركة تأمين'),
            'submit_text': _('إنشاء الشركة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('شركات التأمين'), 'url': reverse('assets:insurance_company_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class InsuranceCompanyUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل شركة تأمين - محسّن"""

    model = InsuranceCompany
    template_name = 'assets/insurance/company_form.html'
    permission_required = 'assets.change_insurancecompany'
    form_class = InsuranceCompanyForm
    success_url = reverse_lazy('assets:insurance_company_list')

    def get_queryset(self):
        return InsuranceCompany.objects.filter(company=self.request.current_company)

    @transaction.atomic
    def form_valid(self, form):
        try:
            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث شركة التأمين {self.object.name} بنجاح'
            )
            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل شركة التأمين {self.object.name}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('شركات التأمين'), 'url': reverse('assets:insurance_company_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class InsuranceCompanyDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل شركة التأمين - جديد"""

    model = InsuranceCompany
    template_name = 'assets/insurance/company_detail.html'
    context_object_name = 'insurance_company'
    permission_required = 'assets.view_insurancecompany'

    def get_queryset(self):
        return InsuranceCompany.objects.filter(
            company=self.request.current_company
        ).prefetch_related('insurance_policies')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # البوليصات
        policies = self.object.insurance_policies.select_related(
            'asset', 'asset__category'
        ).order_by('-start_date')

        # إحصائيات البوليصات
        policy_stats = policies.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=Q(status='active')),
            total_coverage=Coalesce(
                Sum('coverage_amount', filter=Q(status='active')),
                Decimal('0')
            ),
            total_premiums=Coalesce(
                Sum('premium_amount', filter=Q(status='active')),
                Decimal('0')
            ),
        )

        # المطالبات
        claims = InsuranceClaim.objects.filter(
            insurance__insurance_company=self.object
        ).select_related('insurance__asset').order_by('-filed_date')[:10]

        # إحصائيات المطالبات
        claim_stats = InsuranceClaim.objects.filter(
            insurance__insurance_company=self.object
        ).aggregate(
            total_count=Count('id'),
            approved_count=Count('id', filter=Q(status='approved')),
            paid_count=Count('id', filter=Q(status='paid')),
            total_claimed=Coalesce(Sum('claim_amount'), Decimal('0')),
            total_approved=Coalesce(Sum('approved_amount'), Decimal('0')),
        )

        context.update({
            'title': f'شركة التأمين: {self.object.name}',
            'can_edit': self.request.user.has_perm('assets.change_insurancecompany'),
            'can_delete': self.request.user.has_perm('assets.delete_insurancecompany'),
            'policies': policies[:10],
            'policy_stats': policy_stats,
            'claims': claims,
            'claim_stats': claim_stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('شركات التأمين'), 'url': reverse('assets:insurance_company_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class InsuranceCompanyDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف شركة تأمين - محسّن"""

    model = InsuranceCompany
    template_name = 'assets/insurance/company_confirm_delete.html'
    permission_required = 'assets.delete_insurancecompany'
    success_url = reverse_lazy('assets:insurance_company_list')

    def get_queryset(self):
        return InsuranceCompany.objects.filter(company=self.request.current_company)

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق
        policy_count = self.object.insurance_policies.count()
        if policy_count > 0:
            messages.error(
                request,
                f'❌ لا يمكن حذف شركة تأمين لديها {policy_count} بوليصات'
            )
            return redirect('assets:insurance_company_list')

        company_name = self.object.name
        messages.success(request, f'✅ تم حذف شركة التأمين {company_name} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Insurance ====================

class AssetInsuranceListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة بوليصات التأمين - محسّنة"""

    model = AssetInsurance
    template_name = 'assets/insurance/insurance_list.html'
    context_object_name = 'insurances'
    permission_required = 'assets.view_assetinsurance'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetInsurance.objects.filter(
            company=self.request.current_company
        ).select_related(
            'insurance_company', 'asset', 'asset__category', 'asset__branch'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        insurance_company = self.request.GET.get('insurance_company')
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        coverage_type = self.request.GET.get('coverage_type')
        expiring_soon = self.request.GET.get('expiring_soon')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if insurance_company:
            queryset = queryset.filter(insurance_company_id=insurance_company)

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if coverage_type:
            queryset = queryset.filter(coverage_type=coverage_type)

        if expiring_soon == '1':
            expiry_date = date.today() + timedelta(days=60)
            queryset = queryset.filter(
                status='active',
                end_date__lte=expiry_date,
                end_date__gte=date.today()
            )

        if search:
            queryset = queryset.filter(
                Q(policy_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(insurance_company__name__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-start_date')
        queryset = queryset.order_by(sort_by, '-policy_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # شركات التأمين
        insurance_companies = InsuranceCompany.objects.filter(
            company=self.request.current_company
        ).order_by('name')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        insurances = AssetInsurance.objects.filter(
            company=self.request.current_company
        )

        stats = insurances.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=Q(status='active')),
            expired_count=Count('id', filter=Q(status='expired')),
            total_coverage=Coalesce(
                Sum('coverage_amount', filter=Q(status='active')),
                Decimal('0')
            ),
            total_premiums=Coalesce(
                Sum('premium_amount', filter=Q(status='active')),
                Decimal('0')
            ),
            avg_coverage=Coalesce(
                Avg('coverage_amount', filter=Q(status='active')),
                Decimal('0')
            ),
        )

        # البوليصات المنتهية قريباً (30 يوم)
        expiring_30 = insurances.filter(
            status='active',
            end_date__lte=date.today() + timedelta(days=30),
            end_date__gte=date.today()
        ).count()

        # البوليصات المنتهية قريباً (60 يوم)
        expiring_60 = insurances.filter(
            status='active',
            end_date__lte=date.today() + timedelta(days=60),
            end_date__gte=date.today()
        ).count()

        context.update({
            'title': _('بوليصات التأمين'),
            'can_add': self.request.user.has_perm('assets.add_assetinsurance'),
            'can_edit': self.request.user.has_perm('assets.change_assetinsurance'),
            'can_export': self.request.user.has_perm('assets.view_assetinsurance'),
            'status_choices': AssetInsurance.STATUS_CHOICES,
            'coverage_types': AssetInsurance.COVERAGE_TYPES,
            'insurance_companies': insurance_companies,
            'categories': categories,
            'stats': stats,
            'expiring_30': expiring_30,
            'expiring_60': expiring_60,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': ''},
            ]
        })
        return context


class AssetInsuranceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء بوليصة تأمين - محسّن"""

    model = AssetInsurance
    template_name = 'assets/insurance/insurance_form.html'
    permission_required = 'assets.add_assetinsurance'
    fields = [
        'insurance_company', 'asset', 'coverage_type', 'coverage_description',
        'coverage_amount', 'premium_amount', 'deductible_amount',
        'payment_frequency', 'next_payment_date',
        'start_date', 'end_date', 'renewal_date',
        'status', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['insurance_company'].queryset = InsuranceCompany.objects.filter(
            company=company
        ).order_by('name')

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            status='active'
        ).select_related('category')

        # القيم الافتراضية
        form.fields['start_date'].initial = date.today()
        form.fields['status'].initial = 'active'

        # تعيين widgets للتواريخ بشكل صريح
        date_fields = ['start_date', 'end_date', 'renewal_date', 'next_payment_date']
        for field_name in date_fields:
            if field_name in form.fields:
                form.fields[field_name].widget = forms.DateInput(attrs={
                    'class': 'form-control',
                    'type': 'date'
                })

        # إضافة classes للحقول الأخرى
        for field_name, field in form.fields.items():
            if field_name in date_fields:
                continue  # تم معالجتها بالأعلى
            elif field.widget.__class__.__name__ == 'Select':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-select'})
            elif field.widget.__class__.__name__ == 'Textarea':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control', 'rows': 3})
            elif field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user

            self.object = form.save()

            # تحديث حالة التأمين في الأصل
            asset = self.object.asset
            asset.insurance_status = 'insured'
            asset.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء البوليصة {self.object.policy_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:insurance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة بوليصة تأمين'),
            'submit_text': _('إنشاء البوليصة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': reverse('assets:insurance_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class AssetInsuranceDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل بوليصة التأمين - محسّن"""

    model = AssetInsurance
    template_name = 'assets/insurance/insurance_detail.html'
    context_object_name = 'insurance'
    permission_required = 'assets.view_assetinsurance'

    def get_queryset(self):
        return AssetInsurance.objects.filter(
            company=self.request.current_company
        ).select_related(
            'insurance_company', 'asset', 'asset__category',
            'created_by'
        ).prefetch_related('claims')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # المطالبات
        claims = self.object.claims.select_related('filed_by').order_by('-filed_date')

        # إحصائيات المطالبات
        claim_stats = claims.aggregate(
            total_count=Count('id'),
            approved_count=Count('id', filter=Q(status='approved')),
            paid_count=Count('id', filter=Q(status='paid')),
            rejected_count=Count('id', filter=Q(status='rejected')),
            total_claimed=Coalesce(Sum('claim_amount'), Decimal('0')),
            total_approved=Coalesce(Sum('approved_amount'), Decimal('0')),
        )

        # الأيام المتبقية
        if self.object.end_date:
            days_remaining = (self.object.end_date - date.today()).days
        else:
            days_remaining = None

        # التحذيرات
        warnings = []
        if self.object.is_expiring_soon:
            warnings.append({
                'type': 'warning',
                'icon': 'fa-exclamation-triangle',
                'message': f'البوليصة تنتهي قريباً في {self.object.end_date}'
            })

        if self.object.status == 'expired':
            warnings.append({
                'type': 'danger',
                'icon': 'fa-times-circle',
                'message': 'البوليصة منتهية'
            })

        context.update({
            'title': f'البوليصة {self.object.policy_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assetinsurance') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_assetinsurance') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_activate': (
                    self.request.user.has_perm('assets.change_assetinsurance') and
                    self.object.can_activate()  # ✅ استخدام method من Model
            ),
            'can_add_claim': self.request.user.has_perm('assets.add_insuranceclaim'),
            'can_renew': (
                    self.request.user.has_perm('assets.add_assetinsurance') and
                    self.object.status in ['active', 'expired']
            ),
            'can_pay_premium': (
                    self.request.user.has_perm('assets.change_assetinsurance') and
                    self.object.status in ['draft', 'active']  # ✅ يمكن دفع قسط
            ),
            'claims': claims,
            'claim_stats': claim_stats,
            'days_remaining': days_remaining,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': reverse('assets:insurance_list')},
                {'title': self.object.policy_number, 'url': ''},
            ]
        })
        return context


class AssetInsuranceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل بوليصة تأمين - محسّن"""

    model = AssetInsurance
    template_name = 'assets/insurance/insurance_form.html'
    permission_required = 'assets.change_assetinsurance'
    fields = [
        'insurance_company', 'asset', 'coverage_type', 'coverage_description',
        'coverage_amount', 'premium_amount', 'deductible_amount',
        'payment_frequency', 'next_payment_date',
        'start_date', 'end_date', 'renewal_date',
        'status', 'notes'
    ]

    def get_queryset(self):
        return AssetInsurance.objects.filter(company=self.request.current_company)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['insurance_company'].queryset = InsuranceCompany.objects.filter(
            company=company
        ).order_by('name')

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category')

        # تعيين widgets للتواريخ بشكل صريح
        date_fields = ['start_date', 'end_date', 'renewal_date', 'next_payment_date']
        for field_name in date_fields:
            if field_name in form.fields:
                form.fields[field_name].widget = forms.DateInput(attrs={
                    'class': 'form-control',
                    'type': 'date'
                })

        # إضافة classes للحقول الأخرى
        for field_name, field in form.fields.items():
            if field_name in date_fields:
                continue  # تم معالجتها بالأعلى
            elif field.widget.__class__.__name__ == 'Select':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-select'})
            elif field.widget.__class__.__name__ == 'Textarea':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control', 'rows': 3})
            elif field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            # ✅ التحقق من إمكانية التعديل
            if not self.object.can_edit():
                messages.error(
                    self.request,
                    '❌ لا يمكن تعديل هذه البوليصة. قد تكون منتهية أو ملغاة أو لديها مطالبات مدفوعة'
                )
                return self.form_invalid(form)

            old_status = self.object.status
            self.object = form.save()

            # تحديث حالة التأمين في الأصل
            if old_status != self.object.status:
                asset = self.object.asset
                if self.object.status == 'active':
                    asset.insurance_status = 'insured'
                elif self.object.status in ['expired', 'cancelled']:
                    # التحقق من وجود بوليصات أخرى نشطة
                    other_active = AssetInsurance.objects.filter(
                        asset=asset,
                        status='active'
                    ).exclude(pk=self.object.pk).exists()

                    if not other_active:
                        asset.insurance_status = 'expired'
                asset.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث البوليصة {self.object.policy_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:insurance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل البوليصة {self.object.policy_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': reverse('assets:insurance_list')},
                {'title': self.object.policy_number, 'url': reverse('assets:insurance_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class RenewInsuranceView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """تجديد بوليصة تأمين - جديد"""

    template_name = 'assets/insurance/renew_insurance.html'
    permission_required = 'assets.add_assetinsurance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        insurance_id = self.kwargs.get('pk')
        insurance = get_object_or_404(
            AssetInsurance,
            pk=insurance_id,
            company=self.request.current_company
        )

        # الفترة المقترحة (سنة واحدة)
        suggested_start = insurance.end_date + timedelta(days=1)
        suggested_end = suggested_start + timedelta(days=365)

        context.update({
            'title': f'تجديد البوليصة {insurance.policy_number}',
            'insurance': insurance,
            'suggested_start': suggested_start,
            'suggested_end': suggested_end,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': reverse('assets:insurance_list')},
                {'title': insurance.policy_number, 'url': reverse('assets:insurance_detail', args=[insurance.pk])},
                {'title': _('تجديد'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            insurance_id = kwargs.get('pk')
            old_insurance = get_object_or_404(
                AssetInsurance,
                pk=insurance_id,
                company=request.current_company
            )

            # البيانات الجديدة
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            coverage_amount = Decimal(request.POST.get('coverage_amount', old_insurance.coverage_amount))
            premium_amount = Decimal(request.POST.get('premium_amount', old_insurance.premium_amount))

            # إنشاء البوليصة الجديدة
            new_insurance = AssetInsurance.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                insurance_company=old_insurance.insurance_company,
                asset=old_insurance.asset,
                coverage_type=old_insurance.coverage_type,
                coverage_description=old_insurance.coverage_description,
                coverage_amount=coverage_amount,
                premium_amount=premium_amount,
                deductible_amount=old_insurance.deductible_amount,
                payment_frequency=old_insurance.payment_frequency,
                start_date=start_date,
                end_date=end_date,
                status='active',
                renewed_from=old_insurance,
                created_by=request.user,
            )

            # تحديث البوليصة القديمة
            old_insurance.status = 'expired'
            old_insurance.save()

            messages.success(
                request,
                f'✅ تم تجديد البوليصة بنجاح. رقم البوليصة الجديدة: {new_insurance.policy_number}'
            )

            return redirect('assets:insurance_detail', pk=new_insurance.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في التجديد: {str(e)}')
            return redirect('assets:insurance_detail', pk=insurance_id)


class PayPremiumView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """دفع قسط التأمين - جديد"""

    template_name = 'assets/insurance/pay_premium.html'
    permission_required = 'assets.change_assetinsurance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        insurance_id = self.kwargs.get('pk')
        insurance = get_object_or_404(
            AssetInsurance,
            pk=insurance_id,
            company=self.request.current_company
        )

        context.update({
            'title': f'دفع قسط التأمين - {insurance.policy_number}',
            'insurance': insurance,
            'can_pay': insurance.status in ['draft', 'active'],
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بوليصات التأمين'), 'url': reverse('assets:insurance_list')},
                {'title': insurance.policy_number, 'url': reverse('assets:insurance_detail', args=[insurance.pk])},
                {'title': _('دفع قسط'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            insurance_id = kwargs.get('pk')
            insurance = get_object_or_404(
                AssetInsurance,
                pk=insurance_id,
                company=request.current_company
            )

            # التحقق من الحالة
            if insurance.status not in ['draft', 'active']:
                messages.error(
                    request,
                    f'❌ لا يمكن دفع قسط لبوليصة {insurance.get_status_display()}'
                )
                return redirect('assets:insurance_detail', pk=insurance.pk)

            payment_date = request.POST.get('payment_date')
            if not payment_date:
                payment_date = None

            # ✅ إنشاء قيد دفع القسط
            journal_entry = insurance.create_premium_payment_journal_entry(
                payment_date=payment_date,
                user=request.user
            )

            messages.success(
                request,
                f'✅ تم تسجيل دفع قسط التأمين بمبلغ {insurance.premium_amount:,.2f} '
                f'وإنشاء القيد {journal_entry.number} بنجاح'
            )

            return redirect('assets:insurance_detail', pk=insurance.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:insurance_detail', pk=insurance_id)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في دفع القسط: {str(e)}')
            return redirect('assets:insurance_detail', pk=insurance_id)


# ==================== Insurance Claims ====================

class InsuranceClaimListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة مطالبات التأمين - محسّنة"""

    model = InsuranceClaim
    template_name = 'assets/insurance/claim_list.html'
    context_object_name = 'claims'
    permission_required = 'assets.view_insuranceclaim'
    paginate_by = 50

    def get_queryset(self):
        queryset = InsuranceClaim.objects.filter(
            company=self.request.current_company
        ).select_related(
            'insurance', 'insurance__asset', 'insurance__insurance_company',
            'filed_by', 'reviewed_by'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        claim_type = self.request.GET.get('claim_type')
        insurance = self.request.GET.get('insurance')
        insurance_company = self.request.GET.get('insurance_company')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if claim_type:
            queryset = queryset.filter(claim_type=claim_type)

        if insurance:
            queryset = queryset.filter(insurance_id=insurance)

        if insurance_company:
            queryset = queryset.filter(insurance__insurance_company_id=insurance_company)

        if date_from:
            queryset = queryset.filter(filed_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(filed_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(claim_number__icontains=search) |
                Q(insurance__asset__asset_number__icontains=search) |
                Q(insurance__asset__name__icontains=search) |
                Q(incident_description__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-filed_date')
        queryset = queryset.order_by(sort_by, '-claim_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # شركات التأمين
        insurance_companies = InsuranceCompany.objects.filter(
            company=self.request.current_company
        ).order_by('name')

        # إحصائيات مفصّلة
        claims = InsuranceClaim.objects.filter(
            company=self.request.current_company
        )

        stats = claims.aggregate(
            total_count=Count('id'),
            filed_count=Count('id', filter=Q(status='filed')),
            under_review_count=Count('id', filter=Q(status='under_review')),
            approved_count=Count('id', filter=Q(status='approved')),
            rejected_count=Count('id', filter=Q(status='rejected')),
            paid_count=Count('id', filter=Q(status='paid')),
            total_claimed=Coalesce(Sum('claim_amount'), Decimal('0')),
            total_approved=Coalesce(Sum('approved_amount'), Decimal('0')),
            avg_claim=Coalesce(Avg('claim_amount'), Decimal('0')),
        )

        # نسبة الموافقة
        if stats['total_count'] > 0:
            approval_rate = (stats['approved_count'] + stats['paid_count']) / stats['total_count'] * 100
        else:
            approval_rate = 0

        stats['approval_rate'] = round(approval_rate, 2)

        context.update({
            'title': _('مطالبات التأمين'),
            'can_add': self.request.user.has_perm('assets.add_insuranceclaim'),
            'can_edit': self.request.user.has_perm('assets.change_insuranceclaim'),
            'can_export': self.request.user.has_perm('assets.view_insuranceclaim'),
            'status_choices': InsuranceClaim.STATUS_CHOICES,
            'claim_types': InsuranceClaim.CLAIM_TYPES,
            'insurance_companies': insurance_companies,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('مطالبات التأمين'), 'url': ''},
            ]
        })
        return context


class InsuranceClaimCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء مطالبة تأمين - محسّن"""

    model = InsuranceClaim
    template_name = 'assets/insurance/claim_form.html'
    permission_required = 'assets.add_insuranceclaim'
    fields = [
        'insurance', 'claim_type', 'incident_date', 'incident_time',
        'incident_location', 'incident_description',
        'estimated_damage', 'claim_amount',
        'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        form.fields['insurance'].queryset = AssetInsurance.objects.filter(
            company=self.request.current_company,
            status='active'
        ).select_related('asset', 'insurance_company')

        # القيم الافتراضية
        form.fields['incident_date'].initial = date.today()
        form.fields['claim_type'].initial = 'accident'

        # تعيين widgets بشكل صريح
        if 'incident_date' in form.fields:
            form.fields['incident_date'].widget = forms.DateInput(attrs={
                'class': 'form-control',
                'type': 'date'
            })

        if 'incident_time' in form.fields:
            form.fields['incident_time'].widget = forms.TimeInput(attrs={
                'class': 'form-control',
                'type': 'time'
            })

        # إضافة classes للحقول الأخرى
        for field_name, field in form.fields.items():
            if field_name in ['incident_date', 'incident_time']:
                continue  # تم معالجتها بالأعلى
            elif field.widget.__class__.__name__ == 'Select':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-select'})
            elif field.widget.__class__.__name__ == 'Textarea':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control', 'rows': 4})
            elif field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user
            form.instance.filed_by = self.request.user
            form.instance.filed_date = date.today()
            form.instance.status = 'filed'

            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم تقديم المطالبة {self.object.claim_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:claim_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('تقديم مطالبة تأمين'),
            'submit_text': _('تقديم المطالبة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('مطالبات التأمين'), 'url': reverse('assets:claim_list')},
                {'title': _('تقديم مطالبة'), 'url': ''},
            ]
        })
        return context


class InsuranceClaimDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل مطالبة التأمين - محسّن"""

    model = InsuranceClaim
    template_name = 'assets/insurance/claim_detail.html'
    context_object_name = 'claim'
    permission_required = 'assets.view_insuranceclaim'

    def get_queryset(self):
        return InsuranceClaim.objects.filter(
            company=self.request.current_company
        ).select_related(
            'insurance', 'insurance__asset', 'insurance__insurance_company',
            'filed_by', 'reviewed_by', 'journal_entry', 'created_by'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الفرق بين المطالب والمعتمد
        if self.object.approved_amount:
            difference = self.object.approved_amount - self.object.claim_amount
            difference_pct = (difference / self.object.claim_amount * 100) if self.object.claim_amount > 0 else 0
        else:
            difference = None
            difference_pct = None

        # مدة المعالجة
        if self.object.review_date and self.object.filed_date:
            processing_days = (self.object.review_date - self.object.filed_date).days
        else:
            processing_days = None

        context.update({
            'title': f'المطالبة {self.object.claim_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_insuranceclaim') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_insuranceclaim') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_approve': (
                    self.request.user.has_perm('assets.change_insuranceclaim') and
                    self.object.can_approve()  # ✅ استخدام method من Model
            ),
            'can_pay': (
                    self.request.user.has_perm('assets.change_insuranceclaim') and
                    self.object.can_process_payment()  # ✅ استخدام method من Model
            ),
            'can_cancel': (
                    self.request.user.has_perm('assets.change_insuranceclaim') and
                    self.object.status in ['filed', 'under_review']
            ),
            'difference': difference,
            'difference_pct': difference_pct,
            'processing_days': processing_days,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('مطالبات التأمين'), 'url': reverse('assets:claim_list')},
                {'title': self.object.claim_number, 'url': ''},
            ]
        })
        return context


class InsuranceClaimUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل مطالبة تأمين - محسّن"""

    model = InsuranceClaim
    template_name = 'assets/insurance/claim_form.html'
    permission_required = 'assets.change_insuranceclaim'
    fields = [
        'insurance', 'claim_type', 'incident_date', 'incident_time',
        'incident_location', 'incident_description',
        'estimated_damage', 'claim_amount', 'approved_amount',
        'deductible_applied', 'status', 'rejection_reason', 'notes'
    ]

    def get_queryset(self):
        return InsuranceClaim.objects.filter(
            company=self.request.current_company
        ).exclude(status__in=['paid', 'cancelled'])

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        form.fields['insurance'].queryset = AssetInsurance.objects.filter(
            company=self.request.current_company
        ).select_related('asset', 'insurance_company')

        # تعيين widgets بشكل صريح
        if 'incident_date' in form.fields:
            form.fields['incident_date'].widget = forms.DateInput(attrs={
                'class': 'form-control',
                'type': 'date'
            })

        if 'incident_time' in form.fields:
            form.fields['incident_time'].widget = forms.TimeInput(attrs={
                'class': 'form-control',
                'type': 'time'
            })

        # إضافة classes للحقول الأخرى
        for field_name, field in form.fields.items():
            if field_name in ['incident_date', 'incident_time']:
                continue  # تم معالجتها بالأعلى
            elif field.widget.__class__.__name__ == 'Select':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-select'})
            elif field.widget.__class__.__name__ == 'Textarea':
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control', 'rows': 4})
            elif field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect']:
                if 'class' not in field.widget.attrs:
                    field.widget.attrs.update({'class': 'form-control'})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            # ✅ التحقق من إمكانية التعديل
            if not self.object.can_edit():
                messages.error(
                    self.request,
                    '❌ لا يمكن تعديل هذه المطالبة. قد تكون معتمدة أو مدفوعة أو ملغاة'
                )
                return self.form_invalid(form)

            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث المطالبة {self.object.claim_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:claim_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل المطالبة {self.object.claim_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('مطالبات التأمين'), 'url': reverse('assets:claim_list')},
                {'title': self.object.claim_number, 'url': reverse('assets:claim_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


# ==================== Ajax Views - محسّنة ====================

@login_required
@permission_required_with_message('assets.change_insuranceclaim')
@require_http_methods(["POST"])
def approve_insurance_claim(request, pk):
    """اعتماد مطالبة تأمين - محسّن"""

    try:
        claim = get_object_or_404(
            InsuranceClaim,
            pk=pk,
            company=request.current_company
        )

        # ✅ التحقق من إمكانية الاعتماد
        if not claim.can_approve():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن اعتماد هذه المطالبة. تحقق من حالتها والبوليصة'
            }, status=400)

        approved_amount = Decimal(request.POST.get('approved_amount', 0))
        deductible = Decimal(request.POST.get('deductible_applied', 0))

        if not approved_amount or approved_amount <= 0:
            return JsonResponse({
                'success': False,
                'message': 'يجب تحديد المبلغ المعتمد'
            }, status=400)

        # اعتماد المطالبة
        claim.approve(
            approved_amount=approved_amount,
            deductible_applied=deductible,
            user=request.user
        )

        return JsonResponse({
            'success': True,
            'message': f'تم اعتماد المطالبة {claim.claim_number} بمبلغ {approved_amount:,.2f}',
            'claim_number': claim.claim_number,
            'approved_amount': float(approved_amount),
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في اعتماد المطالبة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_insuranceclaim')
@require_http_methods(["POST"])
def reject_insurance_claim(request, pk):
    """رفض مطالبة تأمين - جديد"""

    try:
        claim = get_object_or_404(
            InsuranceClaim,
            pk=pk,
            company=request.current_company
        )

        if claim.status not in ['filed', 'under_review']:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن رفض هذه المطالبة'
            }, status=400)

        rejection_reason = request.POST.get('rejection_reason', '')

        if not rejection_reason:
            return JsonResponse({
                'success': False,
                'message': 'يجب تحديد سبب الرفض'
            }, status=400)

        # رفض المطالبة
        claim.status = 'rejected'
        claim.rejection_reason = rejection_reason
        claim.reviewed_by = request.user
        claim.review_date = date.today()
        claim.save()

        return JsonResponse({
            'success': True,
            'message': f'تم رفض المطالبة {claim.claim_number}',
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في رفض المطالبة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_insuranceclaim')
@require_http_methods(["POST"])
def process_claim_payment(request, pk):
    """معالجة دفع مطالبة تأمين - محسّن"""

    try:
        claim = get_object_or_404(
            InsuranceClaim,
            pk=pk,
            company=request.current_company
        )

        # ✅ التحقق من إمكانية معالجة الدفع
        if not claim.can_process_payment():
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن معالجة دفع هذه المطالبة. يجب أن تكون معتمدة وليس لديها قيد محاسبي مسبق'
            }, status=400)

        # ✅ معالجة الدفع
        claim.process_payment(user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم معالجة دفع المطالبة {claim.claim_number} بنجاح',
            'journal_entry_number': claim.journal_entry.number if claim.journal_entry else None,
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في معالجة الدفع: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_insuranceclaim')
@require_http_methods(["POST"])
def cancel_insurance_claim(request, pk):
    """إلغاء مطالبة تأمين - جديد"""

    try:
        claim = get_object_or_404(
            InsuranceClaim,
            pk=pk,
            company=request.current_company
        )

        if claim.status not in ['filed', 'under_review']:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن إلغاء هذه المطالبة'
            }, status=400)

        reason = request.POST.get('reason', '')

        claim.status = 'cancelled'
        claim.notes = f"{claim.notes}\nإلغاء: {reason}" if claim.notes else f"إلغاء: {reason}"
        claim.save()

        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء المطالبة {claim.claim_number}',
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إلغاء المطالبة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetinsurance')
@require_http_methods(["GET"])
def insurance_datatable_ajax(request):
    """Ajax endpoint لجدول بوليصات التأمين - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        # إذا كان الطلب للإحصائيات فقط
        if request.GET.get('stats_only'):
            all_insurances = AssetInsurance.objects.filter(company=request.current_company)
            today = date.today()
            expiry_date = today + timedelta(days=60)

            stats = {
                'total': all_insurances.count(),
                'active': all_insurances.filter(status='active').count(),
                'expiring_soon': all_insurances.filter(
                    status='active',
                    end_date__lte=expiry_date,
                    end_date__gte=today
                ).count(),
                'expired': all_insurances.filter(status='expired').count()
            }
            return JsonResponse({'stats': stats})

        # إذا كان الطلب لقائمة شركات التأمين
        if request.GET.get('get_companies'):
            companies = InsuranceCompany.objects.filter(
                company=request.current_company,
                is_active=True
            ).values('id', 'name').order_by('name')
            return JsonResponse({'companies': list(companies)})

        # معاملات DataTables
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        status = request.GET.get('status', '')
        insurance_company = request.GET.get('insurance_company', '')
        asset_filter = request.GET.get('asset', '')
        coverage_type = request.GET.get('coverage_type', '')

        # Query
        queryset = AssetInsurance.objects.filter(
            company=request.current_company
        ).select_related(
            'insurance_company', 'asset', 'asset__category'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if insurance_company:
            queryset = queryset.filter(insurance_company_id=insurance_company)

        if asset_filter:
            queryset = queryset.filter(asset_id=asset_filter)

        if coverage_type:
            queryset = queryset.filter(coverage_type=coverage_type)

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(policy_number__icontains=search_value) |
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value) |
                Q(insurance_company__name__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'policy_number',
            '1': 'asset__asset_number',
            '2': 'insurance_company__name',
            '3': 'coverage_type',
            '4': 'coverage_amount',
            '5': 'end_date',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-policy_number')
        else:
            queryset = queryset.order_by('-start_date', '-policy_number')

        # العد
        total_records = AssetInsurance.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_assetinsurance')
        can_edit = request.user.has_perm('assets.change_assetinsurance')
        can_delete = request.user.has_perm('assets.delete_assetinsurance')

        for insurance in queryset:
            # الحالة
            status_map = {
                'draft': '<span class="badge bg-secondary">مسودة</span>',
                'active': '<span class="badge bg-success">نشط</span>',
                'expired': '<span class="badge bg-danger">منتهي</span>',
                'cancelled': '<span class="badge bg-dark">ملغي</span>',
            }
            status_badge = status_map.get(insurance.status, insurance.status)

            # التحقق من قرب الانتهاء
            if insurance.status == 'active' and insurance.end_date:
                days_remaining = (insurance.end_date - date.today()).days
                if 0 < days_remaining <= 60:
                    status_badge += f' <span class="badge bg-warning text-dark">قرب الانتهاء ({days_remaining} يوم)</span>'

            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:insurance_detail', args=[insurance.pk])}"
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            if can_edit:
                actions.append(f'''
                    <a href="{reverse('assets:insurance_update', args=[insurance.pk])}"
                       class="btn btn-outline-warning btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            # نوع التغطية
            coverage_type_display = insurance.get_coverage_type_display() if hasattr(insurance, 'get_coverage_type_display') else insurance.coverage_type

            data.append([
                # رقم الوثيقة
                f'''<div>
                    <a href="{reverse("assets:insurance_detail", args=[insurance.pk])}" class="text-decoration-none">
                        <strong>{insurance.policy_number}</strong>
                    </a>
                </div>''',
                # الأصل
                f'''<div>
                    <a href="{reverse("assets:asset_detail", args=[insurance.asset.pk])}" class="text-decoration-none">
                        <strong>{insurance.asset.name}</strong>
                    </a>
                    <br><small class="text-muted">{insurance.asset.asset_number}</small>
                </div>''',
                # شركة التأمين
                f'''<div>
                    <strong>{insurance.insurance_company.name}</strong>
                    <br><small class="text-muted">{insurance.insurance_company.phone or '-'}</small>
                </div>''',
                # نوع التغطية
                f'<span class="badge bg-info">{coverage_type_display}</span>',
                # تاريخ البداية
                insurance.start_date.strftime('%Y-%m-%d'),
                # تاريخ الانتهاء
                insurance.end_date.strftime('%Y-%m-%d') if insurance.end_date else '<span class="text-muted">-</span>',
                # القيمة المؤمنة
                f'{insurance.coverage_amount:,.3f}',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in insurance_datatable_ajax: {str(e)}")
        logger.error(traceback.format_exc())

        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_insuranceclaim')
@require_http_methods(["GET"])
def claim_datatable_ajax(request):
    """Ajax endpoint لجدول مطالبات التأمين - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        # إذا كان الطلب للإحصائيات فقط
        if request.GET.get('stats_only'):
            all_claims = InsuranceClaim.objects.filter(company=request.current_company)
            stats = {
                'total': all_claims.count(),
                'pending': all_claims.filter(status='pending').count(),
                'approved': all_claims.filter(status='approved').count(),
                'settled': all_claims.filter(status='settled').count()
            }
            return JsonResponse({'stats': stats})

        # إذا كان الطلب لقائمة الوثائق
        if request.GET.get('get_insurances'):
            insurances = AssetInsurance.objects.filter(
                company=request.current_company,
                status='active'
            ).values('id', 'policy_number').order_by('policy_number')
            return JsonResponse({'insurances': list(insurances)})

        # معاملات DataTables
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        status = request.GET.get('status', '')
        insurance_filter = request.GET.get('insurance', '')
        incident_type = request.GET.get('incident_type', '')
        date_from = request.GET.get('date_from', '')

        # Query
        queryset = InsuranceClaim.objects.filter(
            company=request.current_company
        ).select_related(
            'insurance__asset', 'insurance__insurance_company', 'filed_by'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if insurance_filter:
            queryset = queryset.filter(insurance_id=insurance_filter)

        if incident_type:
            queryset = queryset.filter(incident_type=incident_type)

        if date_from:
            queryset = queryset.filter(incident_date__gte=date_from)

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(claim_number__icontains=search_value) |
                Q(insurance__asset__asset_number__icontains=search_value) |
                Q(insurance__asset__name__icontains=search_value) |
                Q(incident_description__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'claim_number',
            '1': 'filed_date',
            '2': 'claim_type',
            '3': 'insurance__asset__asset_number',
            '4': 'claim_amount',
            '5': 'approved_amount',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-claim_number')
        else:
            queryset = queryset.order_by('-filed_date', '-claim_number')

        # العد
        total_records = InsuranceClaim.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_insuranceclaim')

        for claim in queryset:
            # الحالة
            status_map = {
                'filed': '<span class="badge bg-info"><i class="fas fa-file-upload"></i> مقدم</span>',
                'under_review': '<span class="badge bg-warning"><i class="fas fa-search"></i> قيد المراجعة</span>',
                'approved': '<span class="badge bg-primary"><i class="fas fa-check"></i> معتمد</span>',
                'rejected': '<span class="badge bg-danger"><i class="fas fa-times"></i> مرفوض</span>',
                'paid': '<span class="badge bg-success"><i class="fas fa-money-bill"></i> مدفوع</span>',
                'cancelled': '<span class="badge bg-dark"><i class="fas fa-ban"></i> ملغي</span>',
            }
            status_badge = status_map.get(claim.status, claim.status)

            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:claim_detail', args=[claim.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                f'<a href="{reverse("assets:claim_detail", args=[claim.pk])}">{claim.claim_number}</a>',
                claim.filed_date.strftime('%Y-%m-%d'),
                claim.get_claim_type_display(),
                f'''<div>
                    <strong>{claim.insurance.asset.asset_number}</strong>
                    <br><small class="text-muted">{claim.insurance.asset.name}</small>
                </div>''',
                f'<div class="text-end"><strong>{claim.claim_amount:,.2f}</strong></div>',
                f'<div class="text-end">{claim.approved_amount:,.2f}</div>' if claim.approved_amount else '<span class="text-muted">-</span>',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetinsurance')
@require_http_methods(["GET"])
def insurance_expiring_ajax(request):
    """البوليصات المنتهية قريباً - محسّن"""

    try:
        days = int(request.GET.get('days', 30))
        expiry_date = date.today() + timedelta(days=days)

        insurances = AssetInsurance.objects.filter(
            company=request.current_company,
            status='active',
            end_date__lte=expiry_date,
            end_date__gte=date.today()
        ).select_related(
            'asset', 'asset__category', 'insurance_company'
        ).order_by('end_date')

        results = []
        for insurance in insurances:
            remaining_days = (insurance.end_date - date.today()).days
            results.append({
                'id': insurance.id,
                'policy_number': insurance.policy_number,
                'asset_number': insurance.asset.asset_number,
                'asset_name': insurance.asset.name,
                'category': insurance.asset.category.name,
                'insurance_company': insurance.insurance_company.name,
                'end_date': insurance.end_date.strftime('%Y-%m-%d'),
                'remaining_days': remaining_days,
                'coverage_amount': float(insurance.coverage_amount),
                'premium_amount': float(insurance.premium_amount),
                'url': reverse('assets:insurance_detail', args=[insurance.pk]),
            })

        return JsonResponse({
            'success': True,
            'count': len(results),
            'insurances': results
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


# ==================== Export Functions - جديد ====================

@login_required
@permission_required_with_message('assets.view_assetinsurance')
@require_http_methods(["GET"])
def export_insurance_list_excel(request):
    """تصدير قائمة البوليصات إلى Excel - جديد"""

    try:
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Insurance Policies"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Policy Number', 'Asset Number', 'Asset Name',
            'Insurance Company', 'Coverage Type', 'Coverage Amount',
            'Premium Amount', 'Start Date', 'End Date', 'Status'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # البيانات
        insurances = AssetInsurance.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'insurance_company'
        ).order_by('-start_date')

        row_num = 2
        for insurance in insurances:
            ws.cell(row=row_num, column=1, value=insurance.policy_number)
            ws.cell(row=row_num, column=2, value=insurance.asset.asset_number)
            ws.cell(row=row_num, column=3, value=insurance.asset.name)
            ws.cell(row=row_num, column=4, value=insurance.insurance_company.name)
            ws.cell(row=row_num, column=5, value=insurance.get_coverage_type_display())
            ws.cell(row=row_num, column=6, value=float(insurance.coverage_amount))
            ws.cell(row=row_num, column=7, value=float(insurance.premium_amount))
            ws.cell(row=row_num, column=8, value=insurance.start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=9, value=insurance.end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=10, value=insurance.get_status_display())
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="insurance_policies.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:insurance_list')