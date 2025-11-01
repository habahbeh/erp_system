# apps/assets/views/depreciation_views.py
"""
Views إدارة الإهلاك - محسّنة وشاملة
- احتساب الإهلاك الفردي والجماعي
- جداول الإهلاك المستقبلية
- عكس الإهلاك
- إيقاف/استئناف الإهلاك
- تقارير مفصّلة
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import ListView, DetailView, FormView, TemplateView
from django.db.models import (
    Q, Sum, Count, Avg, Max, Min, F,
    DecimalField, Case, When, Value
)
from django.db.models.functions import Coalesce, TruncMonth
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
from dateutil.relativedelta import relativedelta

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import Asset, AssetDepreciation, AssetCategory
from apps.accounting.models import FiscalYear, AccountingPeriod, JournalEntry


# ==================== List & Detail Views ====================

class AssetDepreciationListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة سجلات الإهلاك - محسّنة"""

    model = AssetDepreciation
    template_name = 'assets/depreciation/depreciation_list.html'
    context_object_name = 'depreciation_records'
    permission_required = 'assets.view_asset'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetDepreciation.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'asset__branch',
            'fiscal_year', 'period', 'created_by', 'journal_entry'
        )

        # الفلترة المتقدمة
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        fiscal_year = self.request.GET.get('fiscal_year')
        period = self.request.GET.get('period')
        branch = self.request.GET.get('branch')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        is_posted = self.request.GET.get('is_posted')
        search = self.request.GET.get('search')

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if fiscal_year:
            queryset = queryset.filter(fiscal_year_id=fiscal_year)

        if period:
            queryset = queryset.filter(period_id=period)

        if branch:
            queryset = queryset.filter(asset__branch_id=branch)

        if date_from:
            queryset = queryset.filter(depreciation_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(depreciation_date__lte=date_to)

        if is_posted:
            queryset = queryset.filter(is_posted=(is_posted == '1'))

        if search:
            queryset = queryset.filter(
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(asset__name_en__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-depreciation_date')
        queryset = queryset.order_by(sort_by, '-id')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # السنوات المالية
        fiscal_years = FiscalYear.objects.filter(
            company=self.request.current_company
        ).order_by('-start_date')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        queryset = self.get_queryset()
        stats = queryset.aggregate(
            total_depreciation=Coalesce(Sum('depreciation_amount'), Decimal('0')),
            count=Count('id'),
            avg_depreciation=Coalesce(Avg('depreciation_amount'), Decimal('0')),
            max_depreciation=Coalesce(Max('depreciation_amount'), Decimal('0')),
            min_depreciation=Coalesce(Min('depreciation_amount'), Decimal('0')),
        )

        # إحصائيات حسب الحالة
        posted_count = queryset.filter(is_posted=True).count()
        unposted_count = queryset.filter(is_posted=False).count()

        stats.update({
            'posted_count': posted_count,
            'unposted_count': unposted_count,
            'posted_percentage': round((posted_count / stats['count'] * 100), 2) if stats['count'] > 0 else 0,
        })

        context.update({
            'title': _('سجلات الإهلاك'),
            'can_calculate': self.request.user.has_perm('assets.can_calculate_depreciation'),
            'can_export': self.request.user.has_perm('assets.view_asset'),
            'fiscal_years': fiscal_years,
            'categories': categories,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('سجلات الإهلاك'), 'url': ''},
            ]
        })
        return context


class AssetDepreciationDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل سجل الإهلاك - محسّن"""

    model = AssetDepreciation
    template_name = 'assets/depreciation/depreciation_detail.html'
    context_object_name = 'depreciation'
    permission_required = 'assets.view_asset'

    def get_queryset(self):
        return AssetDepreciation.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'asset__depreciation_method',
            'fiscal_year', 'period', 'created_by',
            'journal_entry', 'reversal_entry'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # معلومات إضافية
        asset = self.object.asset

        # نسبة الإهلاك في هذا السجل
        if asset.original_cost > 0:
            depreciation_pct_this_record = (
                    self.object.depreciation_amount / asset.original_cost * 100
            )
        else:
            depreciation_pct_this_record = 0

        # الأشهر المتبقية عند هذا السجل
        if self.object.depreciation_date:
            months_passed = (
                    (self.object.depreciation_date.year - asset.depreciation_start_date.year) * 12 +
                    (self.object.depreciation_date.month - asset.depreciation_start_date.month)
            )
            months_remaining = max(0, asset.useful_life_months - months_passed)
        else:
            months_remaining = asset.get_remaining_months()

        # السجلات السابقة واللاحقة
        previous_record = AssetDepreciation.objects.filter(
            asset=asset,
            depreciation_date__lt=self.object.depreciation_date
        ).order_by('-depreciation_date').first()

        next_record = AssetDepreciation.objects.filter(
            asset=asset,
            depreciation_date__gt=self.object.depreciation_date
        ).order_by('depreciation_date').first()

        context.update({
            'title': f'سجل إهلاك - {self.object.asset.asset_number}',
            'can_reverse': (
                    self.request.user.has_perm('assets.can_calculate_depreciation') and
                    self.object.is_posted and
                    not self.object.reversal_entry
            ),
            'can_calculate_next': (
                    self.request.user.has_perm('assets.can_calculate_depreciation') and
                    asset.can_depreciate()  # ✅ استخدام method من Model
            ),
            'depreciation_pct_this_record': depreciation_pct_this_record,
            'months_remaining': months_remaining,
            'previous_record': previous_record,
            'next_record': next_record,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('سجلات الإهلاك'), 'url': reverse('assets:depreciation_list')},
                {'title': f'{self.object.asset.asset_number}', 'url': ''},
            ]
        })
        return context


class AssetDepreciationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """تعديل سجل إهلاك - للسجلات غير المرحّلة فقط"""

    template_name = 'assets/depreciation/depreciation_update.html'
    permission_required = 'assets.change_assetdepreciation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        pk = self.kwargs.get('pk')
        depreciation = get_object_or_404(
            AssetDepreciation,
            pk=pk,
            company=self.request.current_company
        )

        # التحقق من أن السجل غير مرحّل
        if depreciation.is_posted:
            messages.error(self.request, '❌ لا يمكن تعديل سجل إهلاك مرحّل. يجب عكس الإهلاك أولاً.')
            return redirect('assets:depreciation_detail', pk=pk)

        context.update({
            'title': f'تعديل سجل إهلاك - {depreciation.asset.asset_number}',
            'depreciation': depreciation,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('سجلات الإهلاك'), 'url': reverse('assets:depreciation_list')},
                {'title': f'{depreciation.asset.asset_number}', 'url': reverse('assets:depreciation_detail', args=[pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            pk = kwargs.get('pk')
            depreciation = get_object_or_404(
                AssetDepreciation,
                pk=pk,
                company=request.current_company
            )

            # التحقق من أن السجل غير مرحّل
            if depreciation.is_posted:
                messages.error(request, '❌ لا يمكن تعديل سجل إهلاك مرحّل')
                return redirect('assets:depreciation_detail', pk=pk)

            # التحقق من وجود قيد محاسبي
            if depreciation.journal_entry:
                messages.error(request, '❌ لا يمكن تعديل سجل إهلاك مرتبط بقيد محاسبي')
                return redirect('assets:depreciation_detail', pk=pk)

            # تحديث الحقول المسموح بتعديلها
            depreciation_date = request.POST.get('depreciation_date')
            notes = request.POST.get('notes', '')

            if depreciation_date:
                depreciation.depreciation_date = depreciation_date

            depreciation.notes = notes
            depreciation.save()

            messages.success(request, f'✅ تم تعديل سجل الإهلاك بنجاح')
            return redirect('assets:depreciation_detail', pk=pk)

        except Exception as e:
            messages.error(request, f'❌ خطأ في التعديل: {str(e)}')
            return redirect('assets:depreciation_detail', pk=pk)


# ==================== Calculation Views ====================

class CalculateDepreciationView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """احتساب إهلاك أصل واحد - محسّن"""

    template_name = 'assets/depreciation/calculate_depreciation.html'
    permission_required = 'assets.can_calculate_depreciation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        company = self.request.current_company

        # إنشاء الـ form
        from ..forms import SingleAssetDepreciationCalculationForm
        form = SingleAssetDepreciationCalculationForm(company=company)

        asset_id = self.request.GET.get('asset_id') or self.kwargs.get('asset_id')
        asset = None

        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=company
            )
            # تعيين الأصل في الـ form
            form.initial['asset'] = asset

            # معلومات الإهلاك
            depreciation_info = self.get_depreciation_info(asset)
        else:
            depreciation_info = None

        # الأصول المؤهلة للإهلاك
        eligible_assets = Asset.objects.filter(
            company=company,
            status='active',
            depreciation_status='active'
        ).exclude(
            depreciation_method__method_type='units_of_production'
        ).select_related('category')[:20]

        context.update({
            'form': form,
            'title': _('احتساب الإهلاك'),
            'asset': asset,
            'depreciation_info': depreciation_info,
            'eligible_assets': eligible_assets,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('احتساب الإهلاك'), 'url': ''},
            ]
        })
        return context

    def get_depreciation_info(self, asset):
        """حساب معلومات الإهلاك المتوقعة"""
        depreciable_amount = asset.get_depreciable_amount()

        if asset.depreciation_method.method_type == 'straight_line':
            monthly_depreciation = depreciable_amount / asset.useful_life_months
        elif asset.depreciation_method.method_type == 'declining_balance':
            rate = asset.depreciation_method.rate_percentage / 100
            monthly_depreciation = asset.book_value * (rate / 12)
        else:
            monthly_depreciation = Decimal('0')

        # التأكد من عدم تجاوز المبلغ القابل للإهلاك
        remaining = depreciable_amount - asset.accumulated_depreciation
        if monthly_depreciation > remaining:
            monthly_depreciation = remaining

        return {
            'expected_amount': monthly_depreciation,
            'remaining_depreciable': remaining,
            'remaining_months': asset.get_remaining_months(),
            'current_book_value': asset.book_value,
            'expected_book_value': asset.book_value - monthly_depreciation,
        }

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            company = request.current_company

            # إنشاء الـ form مع البيانات المرسلة
            from ..forms import SingleAssetDepreciationCalculationForm
            form = SingleAssetDepreciationCalculationForm(request.POST, company=company)

            if not form.is_valid():
                # عرض أخطاء الـ form
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'❌ {error}')

                # إعادة عرض الـ form مع الأخطاء
                context = self.get_context_data(**kwargs)
                context['form'] = form
                return self.render_to_response(context)

            # الحصول على البيانات من الـ form
            asset = form.cleaned_data['asset']
            calculation_date = form.cleaned_data.get('calculation_date')
            units_used = form.cleaned_data.get('units_used')
            auto_post = form.cleaned_data.get('auto_post', False)

            # ✅ التحقق من إمكانية احتساب الإهلاك
            if not asset.can_depreciate():
                messages.error(
                    request,
                    f'❌ لا يمكن احتساب إهلاك للأصل {asset.asset_number}. '
                    f'تحقق من حالة الأصل وحالة الإهلاك'
                )
                return redirect('assets:asset_detail', pk=asset.pk)

            # ✅ احتساب الإهلاك
            depreciation_record = asset.calculate_monthly_depreciation(user=request.user)

            messages.success(
                request,
                f'✅ تم احتساب إهلاك الأصل {asset.asset_number} بمبلغ {depreciation_record.depreciation_amount:,.2f} بنجاح'
            )

            return redirect('assets:depreciation_detail', pk=depreciation_record.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:calculate_depreciation')

        except PermissionDenied as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:asset_list')

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في احتساب الإهلاك: {str(e)}')
            return redirect('assets:calculate_depreciation')


class BulkDepreciationCalculationView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """احتساب الإهلاك الجماعي - محسّن"""

    template_name = 'assets/depreciation/bulk_calculate.html'
    permission_required = 'assets.can_calculate_depreciation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        company = self.request.current_company

        # إنشاء الـ form
        from ..forms import BulkDepreciationCalculationForm
        form = BulkDepreciationCalculationForm(company=company)

        # الأصول المؤهلة للإهلاك
        eligible_assets_query = Asset.objects.filter(
            company=company,
            status='active',
            depreciation_status='active'
        ).exclude(
            depreciation_method__method_type='units_of_production'
        )

        eligible_count = eligible_assets_query.count()

        # إحصائيات الأصول المؤهلة
        eligible_stats = eligible_assets_query.aggregate(
            total_book_value=Coalesce(Sum('book_value'), Decimal('0')),
            avg_book_value=Coalesce(Avg('book_value'), Decimal('0')),
        )

        # الفئات
        categories = AssetCategory.objects.filter(
            company=company,
            is_active=True
        ).annotate(
            eligible_count=Count(
                'assets',
                filter=Q(
                    assets__status='active',
                    assets__depreciation_status='active'
                )
            )
        ).filter(eligible_count__gt=0).order_by('code')

        # آخر عملية إهلاك جماعي
        last_bulk = AssetDepreciation.objects.filter(
            company=company
        ).order_by('-created_at').first()

        context.update({
            'form': form,
            'title': _('احتساب الإهلاك الجماعي'),
            'eligible_count': eligible_count,
            'eligible_stats': eligible_stats,
            'categories': categories,
            'last_bulk': last_bulk,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('احتساب الإهلاك الجماعي'), 'url': ''},
            ]
        })
        return context

    def post(self, request, *args, **kwargs):
        try:
            category_id = request.POST.get('category_id')
            branch_id = request.POST.get('branch_id')

            # خيارات إضافية
            skip_errors = request.POST.get('skip_errors', 'on') == 'on'
            create_summary = request.POST.get('create_summary', 'on') == 'on'

            # الأصول المؤهلة
            assets = Asset.objects.filter(
                company=request.current_company,
                status='active',
                depreciation_status='active'
            ).exclude(
                depreciation_method__method_type='units_of_production'
            ).select_related('category', 'depreciation_method')

            if category_id:
                assets = assets.filter(category_id=category_id)

            if branch_id:
                assets = assets.filter(branch_id=branch_id)

            # احتساب الإهلاك لكل أصل
            success_count = 0
            error_count = 0
            total_amount = Decimal('0')
            errors = []
            success_assets = []

            import time

            for asset in assets:
                try:
                    # كل عملية إهلاك لها transaction مستقلة
                    depreciation_record = asset.calculate_monthly_depreciation(user=request.user)
                    success_count += 1
                    total_amount += depreciation_record.depreciation_amount
                    success_assets.append({
                        'asset_number': asset.asset_number,
                        'amount': float(depreciation_record.depreciation_amount)
                    })
                    # تأخير بسيط للتأكد من أن القيد السابق تم حفظه بالكامل
                    time.sleep(0.01)  # 10 milliseconds
                except Exception as e:
                    error_count += 1
                    error_msg = f"{asset.asset_number}: {str(e)}"
                    errors.append(error_msg)

                    if not skip_errors:
                        # في حالة عدم تجاوز الأخطاء، نوقف العملية
                        raise Exception(f'فشل في الأصل {asset.asset_number}: {str(e)}')

            # الرسائل النهائية
            if success_count > 0:
                messages.success(
                    request,
                    f'✅ تم احتساب الإهلاك لـ {success_count} أصل بنجاح. '
                    f'إجمالي الإهلاك: {total_amount:,.2f}'
                )

            if error_count > 0:
                error_summary = ', '.join(errors[:5])
                if len(errors) > 5:
                    error_summary += f' ... و {len(errors) - 5} أخطاء أخرى'

                messages.warning(
                    request,
                    f'⚠️ فشل احتساب {error_count} أصل. الأخطاء: {error_summary}'
                )

            # حفظ ملخص في session للعرض
            if create_summary:
                request.session['bulk_depreciation_summary'] = {
                    'success_count': success_count,
                    'error_count': error_count,
                    'total_amount': float(total_amount),
                    'date': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'success_assets': success_assets[:10],  # أول 10 فقط
                    'errors': errors[:10],
                }

            return redirect('assets:depreciation_list')

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في الاحتساب الجماعي: {str(e)}')
            return redirect('assets:bulk_calculate_depreciation')


@login_required
@require_http_methods(["POST", "GET"])
def bulk_depreciation_preview_ajax(request):
    """معاينة الأصول المؤهلة للإهلاك الجماعي عبر AJAX"""
    try:
        company = request.current_company

        # طباعة لأغراض التشخيص
        print(f"🔍 Preview request received - Method: {request.method}")
        print(f"POST data: {request.POST}")
        print(f"GET data: {request.GET}")

        # الحصول على الفلاتر من الطلب (POST أو GET)
        if request.method == 'POST':
            category_id = request.POST.get('category')
            period_year = request.POST.get('period_year')
            period_month = request.POST.get('period_month')
        else:
            category_id = request.GET.get('category')
            period_year = request.GET.get('period_year')
            period_month = request.GET.get('period_month')

        print(f"Filters - Category: {category_id}, Year: {period_year}, Month: {period_month}")

        # بناء الاستعلام للأصول المؤهلة
        assets = Asset.objects.filter(
            company=company,
            status='active',
            depreciation_status='active',
            is_active=True
        ).exclude(
            depreciation_method__method_type='units_of_production'
        ).select_related('category', 'depreciation_method')

        # تطبيق فلتر الفئة إذا وجد
        if category_id:
            assets = assets.filter(category_id=category_id)

        # إعداد بيانات الأصول للعرض
        assets_data = []
        total_depreciation = Decimal('0')
        total_asset_value = Decimal('0')
        total_book_value = Decimal('0')

        print(f"📊 Total assets found: {assets.count()}")

        for asset in assets[:100]:  # حد أقصى 100 أصل للمعاينة
            # حساب الإهلاك المقدر
            estimated_depreciation = Decimal('0')
            if asset.depreciation_method:
                if asset.depreciation_method.method_type == 'straight_line':
                    # القسط الثابت = (القيمة الدفترية - قيمة الإنقاذ) / عدد الأشهر المتبقية
                    remaining_months = asset.useful_life_months or 1
                    estimated_depreciation = (asset.book_value - (asset.salvage_value or Decimal('0'))) / Decimal(remaining_months)
                elif asset.depreciation_method.method_type == 'declining_balance':
                    rate = asset.depreciation_method.rate or Decimal('0')
                    estimated_depreciation = asset.book_value * (rate / Decimal('100')) / Decimal('12')
                elif asset.depreciation_method.method_type == 'double_declining':
                    # معدل مضاعف = (200 / العمر بالسنوات) / 12 شهر
                    if asset.useful_life_months:
                        useful_life_years = asset.useful_life_months / 12
                        rate = (Decimal('200') / Decimal(str(useful_life_years))) / Decimal('12')
                        estimated_depreciation = asset.book_value * (rate / Decimal('100'))

            # التأكد من عدم تجاوز القيمة الدفترية
            if estimated_depreciation > asset.book_value:
                estimated_depreciation = asset.book_value

            # حساب القيمة الدفترية بعد الإهلاك
            new_book_value = asset.book_value - estimated_depreciation

            # تجميع الإحصائيات
            total_depreciation += estimated_depreciation
            total_asset_value += asset.original_cost or Decimal('0')
            total_book_value += asset.book_value or Decimal('0')

            asset_dict = {
                'asset_number': asset.asset_number or '',  # Changed from 'code' to 'asset_number'
                'name': asset.name or '',
                'category': asset.category.name if asset.category else '-',
                'book_value': float(asset.book_value) if asset.book_value else 0.0,  # Added book_value
                'estimated_depreciation': float(estimated_depreciation) if estimated_depreciation else 0.0,  # Changed from depreciation_amount
                'purchase_cost': float(asset.original_cost) if asset.original_cost else 0.0,
                'new_book_value': float(new_book_value) if new_book_value else 0.0
            }
            assets_data.append(asset_dict)

        response_data = {
            'success': True,
            'assets': assets_data,
            'asset_count': assets.count(),
            'total_depreciation': float(total_depreciation) if total_depreciation else 0.0,
            'total_asset_value': float(total_asset_value) if total_asset_value else 0.0,
            'total_book_value': float(total_book_value) if total_book_value else 0.0
        }

        print(f"✅ Response prepared - Assets: {len(assets_data)}, Total Depreciation: {float(total_depreciation)}")
        if len(assets_data) > 0:
            print(f"First asset sample: {assets_data[0]}")

        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ Error in preview: {error_trace}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


# ==================== Additional Actions ====================

class PauseDepreciationView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """إيقاف الإهلاك مؤقتاً - جديد"""

    template_name = 'assets/depreciation/pause_depreciation.html'
    permission_required = 'assets.can_calculate_depreciation'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.kwargs.get('asset_id')
        asset = get_object_or_404(
            Asset,
            pk=asset_id,
            company=self.request.current_company
        )

        context.update({
            'title': f'إيقاف إهلاك الأصل {asset.asset_number}',
            'asset': asset,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': asset.asset_number, 'url': reverse('assets:asset_detail', args=[asset.pk])},
                {'title': _('إيقاف الإهلاك'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            asset_id = kwargs.get('asset_id')
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=request.current_company
            )

            # ✅ التحقق من الحالة الحالية
            if asset.depreciation_status == 'paused':
                messages.warning(request, '⚠️ الإهلاك متوقف مسبقاً')
                return redirect('assets:asset_detail', pk=asset.pk)

            if asset.depreciation_status == 'fully_depreciated':
                messages.error(request, '❌ الأصل مهلك بالكامل، لا يمكن إيقافه')
                return redirect('assets:asset_detail', pk=asset.pk)

            reason = request.POST.get('reason', '')

            # تحديث الحالة
            asset.depreciation_status = 'paused'
            asset.save()

            # تسجيل في الملاحظات
            note = f"تم إيقاف الإهلاك مؤقتاً. السبب: {reason}"
            asset.notes = f"{asset.notes}\n{note}" if asset.notes else note
            asset.save()

            messages.success(request, f'✅ تم إيقاف إهلاك الأصل {asset.asset_number} بنجاح')
            return redirect('assets:asset_detail', pk=asset.pk)

        except Exception as e:
            messages.error(request, f'❌ خطأ: {str(e)}')
            return redirect('assets:asset_detail', pk=asset_id)


class ResumeDepreciationView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """استئناف الإهلاك - جديد"""

    template_name = 'assets/depreciation/resume_depreciation.html'
    permission_required = 'assets.can_calculate_depreciation'

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            asset_id = kwargs.get('asset_id')
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=request.current_company
            )

            if asset.depreciation_status != 'paused':
                messages.warning(request, '⚠️ الإهلاك ليس متوقفاً')
                return redirect('assets:asset_detail', pk=asset.pk)

            # تحديث الحالة
            asset.depreciation_status = 'active'
            asset.save()

            # تسجيل في الملاحظات
            note = f"تم استئناف الإهلاك بتاريخ {timezone.now().strftime('%Y-%m-%d')}"
            asset.notes = f"{asset.notes}\n{note}" if asset.notes else note
            asset.save()

            messages.success(request, f'✅ تم استئناف إهلاك الأصل {asset.asset_number} بنجاح')
            return redirect('assets:asset_detail', pk=asset.pk)

        except Exception as e:
            messages.error(request, f'❌ خطأ: {str(e)}')
            return redirect('assets:asset_detail', pk=asset_id)


# ==================== Ajax Views - محسّنة ====================

@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def depreciation_datatable_ajax(request):
    """Ajax endpoint لجدول سجلات الإهلاك - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر المخصصة
        asset = request.GET.get('asset', '')
        year = request.GET.get('year', '')
        month = request.GET.get('month', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        # Query
        queryset = AssetDepreciation.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'asset__category', 'fiscal_year',
            'period', 'journal_entry', 'created_by'
        )

        # تطبيق الفلاتر
        if asset:
            queryset = queryset.filter(asset_id=asset)

        if year:
            queryset = queryset.filter(period_year=year)

        if month:
            queryset = queryset.filter(period_month=month)

        if date_from:
            queryset = queryset.filter(depreciation_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(depreciation_date__lte=date_to)

        # البحث العام
        if search_value:
            queryset = queryset.filter(
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value) |
                Q(asset__name_en__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'depreciation_date',
            '1': 'asset__asset_number',
            '2': 'asset__name',
            '3': 'depreciation_amount',
            '4': 'accumulated_depreciation_after',
            '5': 'book_value_after',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-id')
        else:
            queryset = queryset.order_by('-depreciation_date', '-id')

        # العد
        total_records = AssetDepreciation.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_asset')
        can_view_journal = request.user.has_perm('accounting.view_journalentry')

        for dep in queryset:
            # الحالة
            if dep.is_posted:
                status_badge = '<span class="badge bg-success"><i class="fas fa-check-circle"></i> مرحّل</span>'
            else:
                status_badge = '<span class="badge bg-warning"><i class="fas fa-clock"></i> غير مرحّل</span>'

            # نسبة الإهلاك
            if dep.asset.original_cost > 0:
                pct = (dep.depreciation_amount / dep.asset.original_cost * 100)
                pct_badge = f'<small class="text-muted">{pct:.2f}%</small>'
            else:
                pct_badge = ''

            # أزرار الإجراءات
            actions = []
            can_edit = request.user.has_perm('assets.change_assetdepreciation')
            can_delete = request.user.has_perm('assets.delete_assetdepreciation')
            can_reverse = request.user.has_perm('assets.can_calculate_depreciation')

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:depreciation_detail', args=[dep.pk])}"
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            if dep.journal_entry and can_view_journal:
                actions.append(f'''
                    <a href="{reverse('accounting:journal_entry_detail', args=[dep.journal_entry.pk])}"
                       class="btn btn-outline-secondary btn-sm" title="القيد" data-bs-toggle="tooltip">
                        <i class="fas fa-file-invoice"></i>
                    </a>
                ''')

            # زر تعديل - للسجلات غير المرحّلة فقط
            if not dep.is_posted and can_edit:
                actions.append(f'''
                    <a href="{reverse('assets:depreciation_update', args=[dep.pk])}"
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            # زر عكس الإهلاك - للسجلات المرحّلة فقط
            if dep.is_posted and can_reverse and not dep.reversal_entry:
                actions.append(f'''
                    <button onclick="reverseDepreciation({dep.pk})"
                       class="btn btn-outline-warning btn-sm" title="عكس" data-bs-toggle="tooltip">
                        <i class="fas fa-undo"></i>
                    </button>
                ''')

            # زر حذف - للسجلات غير المرحّلة فقط
            if not dep.is_posted and can_delete:
                actions.append(f'''
                    <button onclick="deleteDepreciation({dep.pk})"
                       class="btn btn-outline-danger btn-sm" title="حذف" data-bs-toggle="tooltip">
                        <i class="fas fa-trash"></i>
                    </button>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                dep.depreciation_date.strftime('%Y-%m-%d'),
                f'<a href="{reverse("assets:asset_detail", args=[dep.asset.pk])}">{dep.asset.asset_number}</a>',
                f'''<div>
                    <strong>{dep.asset.name}</strong>
                    <br><small class="text-muted">{dep.asset.category.name}</small>
                </div>''',
                f'''<div class="text-end">
                    <strong>{dep.depreciation_amount:,.2f}</strong>
                    {pct_badge}
                </div>''',
                f'<div class="text-end">{dep.accumulated_depreciation_after:,.2f}</div>',
                f'<div class="text-end">{dep.book_value_after:,.2f}</div>',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def depreciation_schedule_ajax(request):
    """جدول الإهلاك المستقبلي - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({'success': False, 'error': 'لا توجد شركة محددة'}, status=400)

    try:
        asset_id = request.GET.get('asset_id')
        months_ahead = int(request.GET.get('months', 24))

        if not asset_id:
            return JsonResponse({'success': False, 'error': 'يجب تحديد الأصل'}, status=400)

        asset = get_object_or_404(
            Asset,
            pk=asset_id,
            company=request.current_company
        )

        # حساب جدول الإهلاك
        schedule = []
        remaining_months = asset.get_remaining_months()
        months_to_calculate = min(remaining_months, months_ahead)

        depreciable_amount = asset.get_depreciable_amount()
        current_accumulated = asset.accumulated_depreciation
        current_book_value = asset.book_value
        current_date = date.today()

        for i in range(months_to_calculate):
            month_date = current_date + relativedelta(months=i)

            # حساب الإهلاك الشهري
            if asset.depreciation_method.method_type == 'straight_line':
                monthly_depreciation = depreciable_amount / asset.useful_life_months
            elif asset.depreciation_method.method_type == 'declining_balance':
                rate = asset.depreciation_method.rate_percentage / 100
                monthly_depreciation = current_book_value * (rate / 12)
            else:
                monthly_depreciation = Decimal('0')

            # التأكد من عدم تجاوز المبلغ القابل للإهلاك
            remaining = depreciable_amount - current_accumulated
            if monthly_depreciation > remaining:
                monthly_depreciation = remaining

            if monthly_depreciation <= 0:
                break

            # التحديث
            current_accumulated += monthly_depreciation
            current_book_value -= monthly_depreciation

            # نسبة الإهلاك
            depreciation_pct = (
                (current_accumulated / asset.original_cost * 100)
                if asset.original_cost > 0 else 0
            )

            schedule.append({
                'month': month_date.strftime('%Y-%m'),
                'month_display': month_date.strftime('%B %Y'),
                'depreciation_amount': float(monthly_depreciation),
                'accumulated_depreciation': float(current_accumulated),
                'book_value': float(current_book_value),
                'depreciation_percentage': round(float(depreciation_pct), 2),
            })

        # ملخص
        summary = {
            'total_future_depreciation': sum(item['depreciation_amount'] for item in schedule),
            'final_book_value': schedule[-1]['book_value'] if schedule else float(asset.book_value),
            'months_calculated': len(schedule),
            'will_be_fully_depreciated': len(schedule) < remaining_months,
        }

        return JsonResponse({
            'success': True,
            'schedule': schedule,
            'summary': summary,
            'asset': {
                'asset_number': asset.asset_number,
                'name': asset.name,
                'original_cost': float(asset.original_cost),
                'current_book_value': float(asset.book_value),
                'current_accumulated': float(asset.accumulated_depreciation),
                'salvage_value': float(asset.salvage_value),
                'remaining_months': remaining_months,
                'depreciation_method': asset.depreciation_method.name,
            }
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'خطأ في حساب جدول الإهلاك: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.can_calculate_depreciation')
@require_http_methods(["POST"])
def calculate_single_depreciation_ajax(request, pk):
    """احتساب إهلاك أصل واحد عبر Ajax - محسّن"""

    try:
        asset = get_object_or_404(
            Asset,
            pk=pk,
            company=request.current_company
        )

        # احتساب الإهلاك
        depreciation_record = asset.calculate_monthly_depreciation(user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم احتساب الإهلاك بنجاح',
            'data': {
                'depreciation_id': depreciation_record.pk,
                'depreciation_amount': float(depreciation_record.depreciation_amount),
                'accumulated_depreciation': float(depreciation_record.accumulated_depreciation_after),
                'book_value': float(depreciation_record.book_value_after),
                'depreciation_date': depreciation_record.depreciation_date.strftime('%Y-%m-%d'),
                'journal_entry_number': (
                    depreciation_record.journal_entry.number
                    if depreciation_record.journal_entry else None
                ),
                'is_posted': depreciation_record.is_posted,
            }
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

    except PermissionDenied as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=403)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في احتساب الإهلاك: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.can_calculate_depreciation')
@require_http_methods(["POST"])
def reverse_depreciation_ajax(request, pk):
    """عكس قيد إهلاك - جديد"""

    try:
        depreciation = get_object_or_404(
            AssetDepreciation,
            pk=pk,
            company=request.current_company
        )

        if not depreciation.is_posted:
            return JsonResponse({
                'success': False,
                'message': 'القيد غير مرحّل'
            }, status=400)

        if depreciation.reversal_entry:
            return JsonResponse({
                'success': False,
                'message': 'القيد معكوس مسبقاً'
            }, status=400)

        # عكس القيد المحاسبي
        from apps.accounting.models import JournalEntry

        reversal_entry = depreciation.journal_entry.reverse_entry(
            user=request.user,
            reversal_date=date.today(),
            reason=f"عكس إهلاك {depreciation.asset.asset_number}"
        )

        # تحديث سجل الإهلاك
        depreciation.reversal_entry = reversal_entry
        depreciation.save()

        # تحديث الأصل
        asset = depreciation.asset
        asset.accumulated_depreciation -= depreciation.depreciation_amount
        asset.book_value += depreciation.depreciation_amount
        asset.save()

        return JsonResponse({
            'success': True,
            'message': 'تم عكس قيد الإهلاك بنجاح',
            'reversal_entry_number': reversal_entry.number,
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في عكس القيد: {str(e)}'
        }, status=500)


# ==================== Export Functions ====================

@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_depreciation_schedule_excel(request):
    """تصدير جدول الإهلاك إلى Excel - جديد"""

    try:
        asset_id = request.GET.get('asset_id')
        months = int(request.GET.get('months', 24))

        if not asset_id:
            messages.error(request, 'يجب تحديد الأصل')
            return redirect('assets:depreciation_list')

        asset = get_object_or_404(
            Asset,
            pk=asset_id,
            company=request.current_company
        )

        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Depreciation Schedule"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # معلومات الأصل
        ws['A1'] = 'Asset Number:'
        ws['B1'] = asset.asset_number
        ws['A2'] = 'Asset Name:'
        ws['B2'] = asset.name
        ws['A3'] = 'Original Cost:'
        ws['B3'] = float(asset.original_cost)
        ws['A4'] = 'Salvage Value:'
        ws['B4'] = float(asset.salvage_value)
        ws['A5'] = 'Current Book Value:'
        ws['B5'] = float(asset.book_value)

        # Headers للجدول
        headers = [
            'Month', 'Depreciation Amount',
            'Accumulated Depreciation', 'Book Value',
            'Depreciation %'
        ]

        row_num = 7
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # حساب البيانات
        depreciable_amount = asset.get_depreciable_amount()
        current_accumulated = asset.accumulated_depreciation
        current_book_value = asset.book_value
        current_date = date.today()

        row_num = 8
        for i in range(min(asset.get_remaining_months(), months)):
            month_date = current_date + relativedelta(months=i)

            if asset.depreciation_method.method_type == 'straight_line':
                monthly_dep = depreciable_amount / asset.useful_life_months
            elif asset.depreciation_method.method_type == 'declining_balance':
                rate = asset.depreciation_method.rate_percentage / 100
                monthly_dep = current_book_value * (rate / 12)
            else:
                monthly_dep = Decimal('0')

            remaining = depreciable_amount - current_accumulated
            if monthly_dep > remaining:
                monthly_dep = remaining

            if monthly_dep <= 0:
                break

            current_accumulated += monthly_dep
            current_book_value -= monthly_dep

            depreciation_pct = (
                (current_accumulated / asset.original_cost * 100)
                if asset.original_cost > 0 else 0
            )

            ws.cell(row=row_num, column=1, value=month_date.strftime('%Y-%m'))
            ws.cell(row=row_num, column=2, value=float(monthly_dep))
            ws.cell(row=row_num, column=3, value=float(current_accumulated))
            ws.cell(row=row_num, column=4, value=float(current_book_value))
            ws.cell(row=row_num, column=5, value=float(depreciation_pct))

            row_num += 1

        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="depreciation_schedule_{asset.asset_number}.xlsx"'
        )

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:depreciation_list')


@login_required
@permission_required_with_message('assets.delete_assetdepreciation')
@require_http_methods(["POST"])
def depreciation_delete_ajax(request, pk):
    """حذف سجل إهلاك - فقط للسجلات غير المرحّلة"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'success': False,
            'message': 'لا توجد شركة محددة'
        }, status=400)

    try:
        # الحصول على سجل الإهلاك
        depreciation = get_object_or_404(
            AssetDepreciation,
            pk=pk,
            company=request.current_company
        )

        # التحقق من أن السجل غير مرحّل
        if depreciation.is_posted:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن حذف سجل إهلاك مرحّل. يجب عكس الإهلاك أولاً.'
            }, status=400)

        # التحقق من عدم وجود قيد محاسبي
        if depreciation.journal_entry:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن حذف سجل الإهلاك لأنه مرتبط بقيد محاسبي.'
            }, status=400)

        # حفظ معلومات للرسالة
        asset_number = depreciation.asset.asset_number
        amount = depreciation.depreciation_amount

        # حذف السجل
        depreciation.delete()

        return JsonResponse({
            'success': True,
            'message': f'تم حذف سجل إهلاك الأصل {asset_number} بمبلغ {amount:,.2f} بنجاح'
        })

    except AssetDepreciation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'سجل الإهلاك غير موجود'
        }, status=404)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في حذف سجل الإهلاك: {str(e)}'
        }, status=500)