# apps/assets/views/base_report_views.py
"""
Base Report Views - القوالب الأساسية للتقارير
"""

from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from decimal import Decimal
from datetime import date, datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.mixins import CompanyMixin


class BaseAssetReportView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """قالب أساسي لتقارير الأصول"""

    permission_required = 'assets.view_asset'
    report_title = 'تقرير الأصول'

    def get_report_filters(self):
        """الحصول على الفلاتر من GET parameters"""
        return {
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'category_id': self.request.GET.get('category_id', ''),
            'status': self.request.GET.get('status', ''),
            'cost_center_id': self.request.GET.get('cost_center_id', ''),
        }

    def get_report_data(self):
        """يجب تنفيذها في الـ subclass"""
        raise NotImplementedError('يجب تنفيذ get_report_data في الـ subclass')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        filters = self.get_report_filters()
        report_data = self.get_report_data()

        context.update({
            'title': _(self.report_title),
            'filters': filters,
            'report_data': report_data,
        })

        return context


class BaseAssetExportView:
    """قالب أساسي لتصدير التقارير إلى Excel"""

    def create_workbook(self, title):
        """إنشاء Workbook جديد"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limit
        return wb, ws

    def style_header(self, ws, row, columns):
        """تنسيق رأس الجدول"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_title(self, ws, row, title, columns_count):
        """تنسيق عنوان التقرير"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns_count)
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color="366092")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_subtitle(self, ws, row, text, columns_count):
        """تنسيق عنوان فرعي"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns_count)
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal="center")

    def add_data_row(self, ws, row, data):
        """إضافة صف بيانات"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def add_total_row(self, ws, row, label, values, label_columns=1):
        """إضافة صف الإجمالي"""
        bold_font = Font(bold=True, size=11)
        fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # دمج خلايا العنوان
        if label_columns > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_columns)

        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = bold_font
        label_cell.fill = fill
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = thin_border

        # القيم
        for col_num, value in enumerate(values, label_columns + 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.font = bold_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def adjust_column_widths(self, ws):
        """ضبط عرض الأعمدة تلقائياً"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_response(self, wb, filename):
        """إنشاء HTTP Response للتحميل"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class BaseAssetReportExportView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, BaseAssetExportView):
    """قالب أساسي لتصدير تقارير الأصول"""

    permission_required = 'assets.view_asset'
    report_title = 'تقرير الأصول'
    filename = 'asset_report.xlsx'

    def get_report_data(self):
        """يجب تنفيذها في الـ subclass"""
        raise NotImplementedError('يجب تنفيذ get_report_data')

    def get_report_columns(self):
        """يجب تنفيذها في الـ subclass"""
        raise NotImplementedError('يجب تنفيذ get_report_columns')

    def format_row_data(self, item):
        """يجب تنفيذها في الـ subclass"""
        raise NotImplementedError('يجب تنفيذ format_row_data')

    def get(self, request, *args, **kwargs):
        """تصدير التقرير"""
        # إنشاء الـ workbook
        wb, ws = self.create_workbook(self.report_title)

        # العنوان الرئيسي
        columns = self.get_report_columns()
        self.style_title(ws, 1, f"{self.report_title} - {request.current_company.name}", len(columns))

        # معلومات التقرير
        today = date.today().strftime('%Y-%m-%d')
        self.style_subtitle(ws, 2, f"تاريخ التقرير: {today}", len(columns))

        # رأس الجدول
        self.style_header(ws, 4, columns)

        # البيانات
        report_data = self.get_report_data()
        row_num = 5

        for item in report_data:
            row_data = self.format_row_data(item)
            self.add_data_row(ws, row_num, row_data)
            row_num += 1

        # ضبط العرض
        self.adjust_column_widths(ws)

        # الاستجابة
        return self.create_response(wb, self.filename)