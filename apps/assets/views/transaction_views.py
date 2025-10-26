# apps/assets/views/transaction_views.py
"""
Views معاملات الأصول - محسّنة وشاملة
- شراء وبيع الأصول
- استبعاد الأصول
- إعادة التقييم
- التحويل بين الفروع
- سير عمل الاعتماد
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    DetailView, FormView, TemplateView
)
from django.db.models import (
    Q, Sum, Count, Avg, Max, Min, F,
    DecimalField, Case, When, Value
)
from django.db.models.functions import Coalesce, TruncMonth, TruncDate
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
from dateutil.relativedelta import relativedelta

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import (
    AssetTransaction, AssetTransfer, Asset, AssetCategory
)
from apps.core.models import BusinessPartner, Branch


# ==================== Asset Transactions ====================

class AssetTransactionListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة معاملات الأصول - محسّنة"""

    model = AssetTransaction
    template_name = 'assets/transactions/transaction_list.html'
    context_object_name = 'transactions'
    permission_required = 'assets.view_assettransaction'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetTransaction.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'asset__branch',
            'business_partner', 'created_by', 'approved_by',
            'journal_entry'
        )

        # الفلترة المتقدمة
        transaction_type = self.request.GET.get('transaction_type')
        status = self.request.GET.get('status')
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        business_partner = self.request.GET.get('business_partner')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)

        if status:
            queryset = queryset.filter(status=status)

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if business_partner:
            queryset = queryset.filter(business_partner_id=business_partner)

        if date_from:
            queryset = queryset.filter(transaction_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(transaction_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(transaction_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(reference_number__icontains=search) |
                Q(description__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-transaction_date')
        queryset = queryset.order_by(sort_by, '-transaction_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # الشركاء
        partners = BusinessPartner.objects.filter(
            company=self.request.current_company
        ).order_by('name')

        # إحصائيات مفصّلة
        transactions = AssetTransaction.objects.filter(
            company=self.request.current_company
        )

        stats = transactions.aggregate(
            total_count=Count('id'),
            draft_count=Count('id', filter=Q(status='draft')),
            approved_count=Count('id', filter=Q(status='approved')),
            completed_count=Count('id', filter=Q(status='completed')),
            cancelled_count=Count('id', filter=Q(status='cancelled')),

            # حسب النوع
            purchase_count=Count('id', filter=Q(transaction_type='purchase')),
            sale_count=Count('id', filter=Q(transaction_type='sale')),
            disposal_count=Count('id', filter=Q(transaction_type='disposal')),
            revaluation_count=Count('id', filter=Q(transaction_type='revaluation')),

            # المبالغ
            total_amount=Coalesce(Sum('amount'), Decimal('0')),
            purchase_amount=Coalesce(
                Sum('amount', filter=Q(transaction_type='purchase')),
                Decimal('0')
            ),
            sale_amount=Coalesce(
                Sum('sale_price', filter=Q(transaction_type='sale')),
                Decimal('0')
            ),
        )

        # معاملات الشهر الحالي
        this_month = transactions.filter(
            transaction_date__year=date.today().year,
            transaction_date__month=date.today().month
        ).count()

        stats['this_month'] = this_month

        context.update({
            'title': _('معاملات الأصول'),
            'can_add': self.request.user.has_perm('assets.add_assettransaction'),
            'can_edit': self.request.user.has_perm('assets.change_assettransaction'),
            'can_delete': self.request.user.has_perm('assets.delete_assettransaction'),
            'can_approve': self.request.user.has_perm('assets.can_approve_transactions'),
            'can_export': self.request.user.has_perm('assets.view_assettransaction'),
            'transaction_types': AssetTransaction.TRANSACTION_TYPES,
            'status_choices': AssetTransaction.STATUS_CHOICES,
            'categories': categories,
            'partners': partners,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': ''},
            ]
        })
        return context


class AssetTransactionCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء معاملة أصل - محسّن"""

    model = AssetTransaction
    template_name = 'assets/transactions/transaction_form.html'
    permission_required = 'assets.add_assettransaction'
    fields = [
        'transaction_date', 'transaction_type', 'asset',
        'amount', 'sale_price', 'payment_method',
        'business_partner', 'reference_number',
        'description', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            status='active'
        ).select_related('category')

        form.fields['business_partner'].queryset = BusinessPartner.objects.filter(
            company=company
        ).order_by('name')
        form.fields['business_partner'].required = False

        # القيم الافتراضية
        form.fields['transaction_date'].initial = date.today()
        form.fields['payment_method'].initial = 'cash'

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user
            form.instance.status = 'draft'

            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء المعاملة {self.object.transaction_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:transaction_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة معاملة أصل'),
            'submit_text': _('إنشاء المعاملة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': reverse('assets:transaction_list')},
                {'title': _('إضافة معاملة'), 'url': ''},
            ]
        })
        return context


class AssetTransactionDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل معاملة الأصل - محسّن"""

    model = AssetTransaction
    template_name = 'assets/transactions/transaction_detail.html'
    context_object_name = 'transaction'
    permission_required = 'assets.view_assettransaction'

    def get_queryset(self):
        return AssetTransaction.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'business_partner',
            'created_by', 'approved_by', 'journal_entry', 'branch'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # حساب الربح/الخسارة للبيع
        gain_loss = None
        if self.object.transaction_type == 'sale':
            if self.object.sale_price and self.object.asset.current_book_value:
                gain_loss = self.object.sale_price - self.object.asset.current_book_value

        # التحذيرات
        warnings = []
        if self.object.status == 'draft':
            warnings.append({
                'type': 'warning',
                'icon': 'fa-exclamation-triangle',
                'message': 'هذه المعاملة في وضع المسودة - يجب اعتمادها'
            })

        if gain_loss and gain_loss < 0:
            warnings.append({
                'type': 'danger',
                'icon': 'fa-arrow-down',
                'message': f'خسارة من البيع: {abs(gain_loss):,.2f}'
            })
        elif gain_loss and gain_loss > 0:
            warnings.append({
                'type': 'success',
                'icon': 'fa-arrow-up',
                'message': f'ربح من البيع: {gain_loss:,.2f}'
            })

        context.update({
            'title': f'المعاملة {self.object.transaction_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assettransaction') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_assettransaction') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_approve': (
                    self.request.user.has_perm('assets.can_approve_transactions') and
                    self.object.status == 'draft'
            ),
            'can_post': (
                    self.request.user.has_perm('assets.can_approve_transactions') and
                    self.object.can_post()  # ✅ استخدام method من Model
            ),
            'can_cancel': (
                    self.request.user.has_perm('assets.change_assettransaction') and
                    self.object.status in ['draft', 'approved']
            ),
            'gain_loss': gain_loss,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': reverse('assets:transaction_list')},
                {'title': self.object.transaction_number, 'url': ''},
            ]
        })
        return context


class AssetTransactionUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل معاملة أصل - محسّن"""

    model = AssetTransaction
    template_name = 'assets/transactions/transaction_form.html'
    permission_required = 'assets.change_assettransaction'
    fields = [
        'transaction_date', 'transaction_type', 'asset',
        'amount', 'sale_price', 'payment_method',
        'business_partner', 'reference_number',
        'description', 'notes'
    ]

    def get_queryset(self):
        return AssetTransaction.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category')

        form.fields['business_partner'].queryset = BusinessPartner.objects.filter(
            company=company
        ).order_by('name')
        form.fields['business_partner'].required = False

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            # ✅ التحقق من إمكانية التعديل
            if not self.object.can_edit():
                messages.error(
                    self.request,
                    '❌ لا يمكن تعديل هذه المعاملة. قد تكون مكتملة أو لديها قيد محاسبي مرحل'
                )
                return self.form_invalid(form)

            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث المعاملة {self.object.transaction_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:transaction_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل المعاملة {self.object.transaction_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': reverse('assets:transaction_list')},
                {'title': self.object.transaction_number,
                 'url': reverse('assets:transaction_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class AssetTransactionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف معاملة أصل - محسّن"""

    model = AssetTransaction
    template_name = 'assets/transactions/transaction_confirm_delete.html'
    permission_required = 'assets.delete_assettransaction'
    success_url = reverse_lazy('assets:transaction_list')

    def get_queryset(self):
        return AssetTransaction.objects.filter(
            company=self.request.current_company,
            status='draft'
        )

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # ✅ استخدام method من Model للتحقق
        if not self.object.can_delete():
            messages.error(request, '❌ لا يمكن حذف هذه المعاملة. قد تكون مكتملة أو لديها قيد محاسبي')
            return redirect('assets:transaction_detail', pk=self.object.pk)

        transaction_number = self.object.transaction_number
        messages.success(request, f'✅ تم حذف المعاملة {transaction_number} بنجاح')

        return super().delete(request, *args, **kwargs)


class ApproveTransactionView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """اعتماد المعاملة - جديد"""

    template_name = 'assets/transactions/approve_transaction.html'
    permission_required = 'assets.can_approve_transactions'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        transaction_id = self.kwargs.get('pk')
        trans = get_object_or_404(
            AssetTransaction,
            pk=transaction_id,
            company=self.request.current_company,
            status='draft'
        )

        context.update({
            'title': f'اعتماد المعاملة {trans.transaction_number}',
            'transaction': trans,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': reverse('assets:transaction_list')},
                {'title': trans.transaction_number, 'url': reverse('assets:transaction_detail', args=[trans.pk])},
                {'title': _('اعتماد'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            transaction_id = kwargs.get('pk')
            trans = get_object_or_404(
                AssetTransaction,
                pk=transaction_id,
                company=request.current_company,
                status='draft'
            )

            approval_notes = request.POST.get('approval_notes', '')

            # اعتماد المعاملة
            trans.status = 'approved'
            trans.approved_by = request.user
            trans.approved_at = timezone.now()
            if approval_notes:
                trans.notes = f"{trans.notes}\nملاحظات الاعتماد: {approval_notes}" if trans.notes else f"ملاحظات الاعتماد: {approval_notes}"
            trans.save()

            messages.success(
                request,
                f'✅ تم اعتماد المعاملة {trans.transaction_number} بنجاح'
            )

            return redirect('assets:transaction_detail', pk=trans.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في الاعتماد: {str(e)}')
            return redirect('assets:transaction_detail', pk=transaction_id)


class CancelTransactionView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """إلغاء المعاملة - جديد"""

    template_name = 'assets/transactions/cancel_transaction.html'
    permission_required = 'assets.change_assettransaction'

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            transaction_id = kwargs.get('pk')
            trans = get_object_or_404(
                AssetTransaction,
                pk=transaction_id,
                company=request.current_company
            )

            if trans.status not in ['draft', 'approved']:
                messages.error(request, 'لا يمكن إلغاء هذه المعاملة')
                return redirect('assets:transaction_detail', pk=trans.pk)

            cancellation_reason = request.POST.get('cancellation_reason', '')

            # إلغاء المعاملة
            trans.status = 'cancelled'
            trans.notes = f"{trans.notes}\nإلغاء: {cancellation_reason}" if trans.notes else f"إلغاء: {cancellation_reason}"
            trans.save()

            messages.success(
                request,
                f'✅ تم إلغاء المعاملة {trans.transaction_number} بنجاح'
            )

            return redirect('assets:transaction_detail', pk=trans.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في الإلغاء: {str(e)}')
            return redirect('assets:transaction_detail', pk=transaction_id)


class PostTransactionView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """ترحيل المعاملة (إنشاء القيد المحاسبي وإكمال العملية) - جديد"""

    template_name = 'assets/transactions/post_transaction.html'
    permission_required = 'assets.can_approve_transactions'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        transaction_id = self.kwargs.get('pk')
        trans = get_object_or_404(
            AssetTransaction,
            pk=transaction_id,
            company=self.request.current_company,
            status='approved'
        )

        context.update({
            'title': f'ترحيل المعاملة {trans.transaction_number}',
            'transaction': trans,
            'can_post': trans.can_post(),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('معاملات الأصول'), 'url': reverse('assets:transaction_list')},
                {'title': trans.transaction_number, 'url': reverse('assets:transaction_detail', args=[trans.pk])},
                {'title': _('ترحيل'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            transaction_id = kwargs.get('pk')
            trans = get_object_or_404(
                AssetTransaction,
                pk=transaction_id,
                company=request.current_company
            )

            # ✅ التحقق من إمكانية الترحيل
            if not trans.can_post():
                messages.error(
                    request,
                    '❌ لا يمكن ترحيل هذه المعاملة. يجب أن تكون معتمدة وليس لديها قيد محاسبي مسبق'
                )
                return redirect('assets:transaction_detail', pk=trans.pk)

            # ✅ ترحيل المعاملة (إنشاء القيد وتحديث الحالة)
            journal_entry = trans.post(user=request.user)

            messages.success(
                request,
                f'✅ تم ترحيل المعاملة {trans.transaction_number} بنجاح وإنشاء القيد {journal_entry.number}'
            )

            return redirect('assets:transaction_detail', pk=trans.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:transaction_detail', pk=transaction_id)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في الترحيل: {str(e)}')
            return redirect('assets:transaction_detail', pk=transaction_id)


# ==================== Specific Transaction Types ====================

class SellAssetView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """بيع أصل - محسّن"""

    template_name = 'assets/transactions/sell_form.html'
    permission_required = 'assets.can_sell_asset'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.request.GET.get('asset_id') or self.kwargs.get('asset_id')
        asset = None

        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company,
                status='active'
            )

        # الشركاء (المشترين)
        buyers = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['customer', 'both']
        ).order_by('name')

        context.update({
            'title': _('بيع أصل'),
            'asset': asset,
            'buyers': buyers,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('بيع أصل'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            asset_id = request.POST.get('asset_id')
            sale_price = Decimal(request.POST.get('sale_price', 0))
            buyer_id = request.POST.get('buyer_id')
            sale_date = request.POST.get('sale_date', date.today())
            payment_method = request.POST.get('payment_method', 'cash')
            reference_number = request.POST.get('reference_number', '')
            description = request.POST.get('description', '')

            if not asset_id or not sale_price or not buyer_id:
                messages.error(request, '❌ يجب إدخال جميع البيانات المطلوبة')
                return redirect('assets:sell_asset')

            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=request.current_company,
                status='active'
            )

            buyer = get_object_or_404(
                BusinessPartner,
                pk=buyer_id,
                company=request.current_company
            )

            # بيع الأصل
            transaction_obj = asset.sell(
                sale_price=sale_price,
                buyer=buyer,
                sale_date=sale_date,
                payment_method=payment_method,
                reference_number=reference_number,
                description=description,
                user=request.user
            )

            # حساب الربح/الخسارة
            gain_loss = sale_price - asset.current_book_value

            if gain_loss > 0:
                message = f'✅ تم بيع الأصل {asset.asset_number} بنجاح بمبلغ {sale_price:,.2f} بربح {gain_loss:,.2f}'
            elif gain_loss < 0:
                message = f'✅ تم بيع الأصل {asset.asset_number} بنجاح بمبلغ {sale_price:,.2f} بخسارة {abs(gain_loss):,.2f}'
            else:
                message = f'✅ تم بيع الأصل {asset.asset_number} بنجاح بمبلغ {sale_price:,.2f}'

            messages.success(request, message)

            return redirect('assets:transaction_detail', pk=transaction_obj.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:sell_asset')

        except PermissionDenied as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:asset_list')

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في بيع الأصل: {str(e)}')
            return redirect('assets:sell_asset')


class DisposeAssetView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """استبعاد أصل - محسّن"""

    template_name = 'assets/transactions/dispose_form.html'
    permission_required = 'assets.can_dispose_asset'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.request.GET.get('asset_id') or self.kwargs.get('asset_id')
        asset = None

        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company,
                status='active'
            )

        context.update({
            'title': _('استبعاد أصل'),
            'asset': asset,
            'disposal_reasons': [
                ('damaged', 'تالف'),
                ('obsolete', 'قديم ومتقادم'),
                ('lost', 'مفقود'),
                ('stolen', 'مسروق'),
                ('end_of_life', 'انتهاء العمر الإنتاجي'),
                ('other', 'أخرى'),
            ],
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('استبعاد أصل'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            asset_id = request.POST.get('asset_id')
            disposal_date = request.POST.get('disposal_date', date.today())
            reason = request.POST.get('reason', '')
            description = request.POST.get('description', '')

            if not asset_id or not reason:
                messages.error(request, '❌ يجب تحديد الأصل وسبب الاستبعاد')
                return redirect('assets:dispose_asset')

            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=request.current_company
            )

            # استبعاد الأصل
            transaction_obj = asset.dispose(
                disposal_date=disposal_date,
                reason=reason,
                description=description,
                user=request.user
            )

            messages.success(
                request,
                f'✅ تم استبعاد الأصل {asset.asset_number} بنجاح'
            )

            return redirect('assets:transaction_detail', pk=transaction_obj.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:dispose_asset')

        except PermissionDenied as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:asset_list')

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في استبعاد الأصل: {str(e)}')
            return redirect('assets:dispose_asset')


class RevalueAssetView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """إعادة تقييم أصل - مكتمل"""

    template_name = 'assets/transactions/revalue_form.html'
    permission_required = 'assets.can_revalue_asset'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.request.GET.get('asset_id') or self.kwargs.get('asset_id')
        asset = None

        if asset_id:
            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=self.request.current_company
            )

        context.update({
            'title': _('إعادة تقييم أصل'),
            'asset': asset,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('إعادة تقييم أصل'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            asset_id = request.POST.get('asset_id')
            new_value = Decimal(request.POST.get('new_value', 0))
            revaluation_date = request.POST.get('revaluation_date', date.today())
            reason = request.POST.get('reason', '')

            if not asset_id or not new_value:
                messages.error(request, '❌ يجب إدخال الأصل والقيمة الجديدة')
                return redirect('assets:revalue_asset')

            asset = get_object_or_404(
                Asset,
                pk=asset_id,
                company=request.current_company
            )

            old_value = asset.current_book_value
            difference = new_value - old_value

            # إنشاء معاملة إعادة التقييم
            transaction_obj = AssetTransaction.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                transaction_type='revaluation',
                transaction_date=revaluation_date,
                asset=asset,
                amount=abs(difference),
                description=f'إعادة تقييم من {old_value:,.2f} إلى {new_value:,.2f}. السبب: {reason}',
                created_by=request.user,
                status='draft'
            )

            messages.success(
                request,
                f'✅ تم إنشاء معاملة إعادة التقييم بنجاح. الفرق: {difference:,.2f}'
            )

            return redirect('assets:transaction_detail', pk=transaction_obj.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:revalue_asset')

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في إعادة التقييم: {str(e)}')
            return redirect('assets:revalue_asset')


# ==================== Asset Transfers ====================

class AssetTransferListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة تحويلات الأصول - محسّنة"""

    model = AssetTransfer
    template_name = 'assets/transfers/transfer_list.html'
    context_object_name = 'transfers'
    permission_required = 'assets.view_assettransfer'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetTransfer.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category',
            'from_branch', 'to_branch',
            'from_cost_center', 'to_cost_center',
            'from_employee', 'to_employee',
            'requested_by', 'approved_by'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        from_branch = self.request.GET.get('from_branch')
        to_branch = self.request.GET.get('to_branch')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if from_branch:
            queryset = queryset.filter(from_branch_id=from_branch)

        if to_branch:
            queryset = queryset.filter(to_branch_id=to_branch)

        if date_from:
            queryset = queryset.filter(transfer_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(transfer_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(transfer_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(reason__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-transfer_date')
        queryset = queryset.order_by(sort_by, '-transfer_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الفروع
        branches = Branch.objects.filter(
            company=self.request.current_company
        ).order_by('code')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        transfers = AssetTransfer.objects.filter(
            company=self.request.current_company
        )

        stats = transfers.aggregate(
            total_count=Count('id'),
            pending_count=Count('id', filter=Q(status='pending')),
            approved_count=Count('id', filter=Q(status='approved')),
            completed_count=Count('id', filter=Q(status='completed')),
            rejected_count=Count('id', filter=Q(status='rejected')),
            cancelled_count=Count('id', filter=Q(status='cancelled')),
        )

        # تحويلات الشهر الحالي
        this_month = transfers.filter(
            transfer_date__year=date.today().year,
            transfer_date__month=date.today().month
        ).count()

        stats['this_month'] = this_month

        context.update({
            'title': _('تحويلات الأصول'),
            'can_add': self.request.user.has_perm('assets.add_assettransfer'),
            'can_edit': self.request.user.has_perm('assets.change_assettransfer'),
            'can_approve': self.request.user.has_perm('assets.can_transfer_asset'),
            'can_export': self.request.user.has_perm('assets.view_assettransfer'),
            'status_choices': AssetTransfer.STATUS_CHOICES,
            'branches': branches,
            'categories': categories,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('تحويلات الأصول'), 'url': ''},
            ]
        })
        return context


class AssetTransferCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء تحويل أصل - محسّن"""

    model = AssetTransfer
    template_name = 'assets/transfers/transfer_form.html'
    permission_required = 'assets.add_assettransfer'
    fields = [
        'transfer_date', 'asset',
        'from_branch', 'from_cost_center', 'from_employee',
        'to_branch', 'to_cost_center', 'to_employee',
        'reason', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            status='active'
        ).select_related('category', 'branch')

        # القيم الافتراضية
        form.fields['transfer_date'].initial = date.today()

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user
            form.instance.requested_by = self.request.user
            form.instance.status = 'pending'

            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء طلب التحويل {self.object.transfer_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:transfer_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('طلب تحويل أصل'),
            'submit_text': _('إنشاء الطلب'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('تحويلات الأصول'), 'url': reverse('assets:transfer_list')},
                {'title': _('طلب تحويل'), 'url': ''},
            ]
        })
        return context


class AssetTransferDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل تحويل الأصل - محسّن"""

    model = AssetTransfer
    template_name = 'assets/transfers/transfer_detail.html'
    context_object_name = 'transfer'
    permission_required = 'assets.view_assettransfer'

    def get_queryset(self):
        return AssetTransfer.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category',
            'from_branch', 'to_branch',
            'from_cost_center', 'to_cost_center',
            'from_employee', 'to_employee',
            'requested_by', 'approved_by', 'delivered_by', 'received_by',
            'created_by'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Timeline
        timeline = []

        if self.object.created_at:
            timeline.append({
                'date': self.object.created_at,
                'user': self.object.requested_by,
                'action': 'طلب التحويل',
                'icon': 'fa-plus-circle',
                'color': 'primary'
            })

        if self.object.approved_at:
            timeline.append({
                'date': self.object.approved_at,
                'user': self.object.approved_by,
                'action': 'اعتماد التحويل',
                'icon': 'fa-check-circle',
                'color': 'success'
            })

        if self.object.received_at:
            timeline.append({
                'date': self.object.received_at,
                'user': self.object.received_by,
                'action': 'استلام الأصل',
                'icon': 'fa-handshake',
                'color': 'info'
            })

        # التحذيرات
        warnings = []
        if self.object.status == 'pending':
            warnings.append({
                'type': 'warning',
                'icon': 'fa-clock',
                'message': 'التحويل في انتظار الاعتماد'
            })

        context.update({
            'title': f'التحويل {self.object.transfer_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assettransfer') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_assettransfer') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_approve': (
                    self.request.user.has_perm('assets.can_transfer_asset') and
                    self.object.can_approve()  # ✅ استخدام method من Model
            ),
            'can_complete': (
                    self.request.user.has_perm('assets.can_transfer_asset') and
                    self.object.can_complete()  # ✅ استخدام method من Model
            ),
            'can_reject': (
                    self.request.user.has_perm('assets.can_transfer_asset') and
                    self.object.can_reject()  # ✅ استخدام method من Model
            ),
            'timeline': sorted(timeline, key=lambda x: x['date']),
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('تحويلات الأصول'), 'url': reverse('assets:transfer_list')},
                {'title': self.object.transfer_number, 'url': ''},
            ]
        })
        return context


class AssetTransferUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل تحويل أصل - محسّن"""

    model = AssetTransfer
    template_name = 'assets/transfers/transfer_form.html'
    permission_required = 'assets.change_assettransfer'
    fields = [
        'transfer_date', 'asset',
        'from_branch', 'from_cost_center', 'from_employee',
        'to_branch', 'to_cost_center', 'to_employee',
        'reason', 'notes'
    ]

    def get_queryset(self):
        return AssetTransfer.objects.filter(
            company=self.request.current_company,
            status='pending'
        )

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category', 'branch')

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            # ✅ التحقق من إمكانية التعديل
            if not self.object.can_edit():
                messages.error(
                    self.request,
                    '❌ لا يمكن تعديل هذا التحويل. قد يكون معتمد أو مكتمل أو تم التسليم/الاستلام'
                )
                return self.form_invalid(form)

            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث طلب التحويل {self.object.transfer_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:transfer_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل التحويل {self.object.transfer_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('تحويلات الأصول'), 'url': reverse('assets:transfer_list')},
                {'title': self.object.transfer_number, 'url': reverse('assets:transfer_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


# ==================== Ajax Views - محسّنة ====================

@login_required
@permission_required_with_message('assets.can_transfer_asset')
@require_http_methods(["POST"])
def approve_transfer(request, pk):
    """اعتماد تحويل أصل - محسّن"""

    try:
        transfer = get_object_or_404(
            AssetTransfer,
            pk=pk,
            company=request.current_company
        )

        approval_notes = request.POST.get('approval_notes', '')

        # ✅ اعتماد التحويل باستخدام model method
        transfer.approve(user=request.user)

        if approval_notes:
            transfer.notes = f"{transfer.notes}\nملاحظات الاعتماد: {approval_notes}" if transfer.notes else f"ملاحظات الاعتماد: {approval_notes}"
            transfer.save(update_fields=['notes', 'updated_at'])

        return JsonResponse({
            'success': True,
            'message': f'تم اعتماد التحويل {transfer.transfer_number} بنجاح'
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في اعتماد التحويل: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.can_transfer_asset')
@require_http_methods(["POST"])
def reject_transfer(request, pk):
    """رفض تحويل أصل - جديد"""

    try:
        transfer = get_object_or_404(
            AssetTransfer,
            pk=pk,
            company=request.current_company
        )

        rejection_reason = request.POST.get('rejection_reason', '')

        # ✅ رفض التحويل باستخدام model method
        transfer.reject(reason=rejection_reason, user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم رفض التحويل {transfer.transfer_number}'
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في رفض التحويل: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.can_transfer_asset')
@require_http_methods(["POST"])
def complete_transfer(request, pk):
    """إكمال تحويل أصل - محسّن"""

    try:
        transfer = get_object_or_404(
            AssetTransfer,
            pk=pk,
            company=request.current_company
        )

        # تحديث بيانات التسليم والاستلام إذا لم تكن موجودة
        if not transfer.delivered_at:
            transfer.delivered_by = request.user
            transfer.delivered_at = timezone.now()

        if not transfer.received_at:
            transfer.received_by = request.user
            transfer.received_at = timezone.now()

        transfer.save(update_fields=['delivered_by', 'delivered_at', 'received_by', 'received_at', 'updated_at'])

        # ✅ إكمال التحويل باستخدام model method
        transfer.complete(user=request.user)

        return JsonResponse({
            'success': True,
            'message': f'تم إكمال التحويل {transfer.transfer_number} بنجاح'
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إكمال التحويل: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assettransaction')
@require_http_methods(["GET"])
def transaction_datatable_ajax(request):
    """Ajax endpoint لجدول المعاملات - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        transaction_type = request.GET.get('transaction_type', '')
        status = request.GET.get('status', '')

        # Query
        queryset = AssetTransaction.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'asset__category', 'business_partner', 'journal_entry'
        )

        # تطبيق الفلاتر
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)

        if status:
            queryset = queryset.filter(status=status)

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(transaction_number__icontains=search_value) |
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value) |
                Q(reference_number__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'transaction_number',
            '1': 'transaction_date',
            '2': 'transaction_type',
            '3': 'asset__asset_number',
            '4': 'asset__name',
            '5': 'amount',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-transaction_number')
        else:
            queryset = queryset.order_by('-transaction_date', '-transaction_number')

        # العد
        total_records = AssetTransaction.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_assettransaction')
        can_edit = request.user.has_perm('assets.change_assettransaction')
        can_delete = request.user.has_perm('assets.delete_assettransaction')

        for trans in queryset:
            # Checkbox
            checkbox = f'<input type="checkbox" class="transaction-checkbox" value="{trans.pk}">'

            # الحالة
            status_map = {
                'draft': '<span class="badge bg-secondary"><i class="fas fa-file"></i> مسودة</span>',
                'approved': '<span class="badge bg-info"><i class="fas fa-check"></i> معتمد</span>',
                'completed': '<span class="badge bg-success"><i class="fas fa-check-double"></i> مكتمل</span>',
                'cancelled': '<span class="badge bg-danger"><i class="fas fa-ban"></i> ملغي</span>',
            }
            status_badge = status_map.get(trans.status, trans.status)

            # النوع
            type_map = {
                'purchase': '<span class="badge bg-success">شراء</span>',
                'sale': '<span class="badge bg-primary">بيع</span>',
                'disposal': '<span class="badge bg-danger">استبعاد</span>',
                'revaluation': '<span class="badge bg-warning text-dark">إعادة تقييم</span>',
                'impairment': '<span class="badge bg-info">هبوط قيمة</span>',
            }
            type_badge = type_map.get(trans.transaction_type, trans.transaction_type)

            # الأصل
            asset_link = f'''
                <a href="{reverse("assets:asset_detail", args=[trans.asset.pk])}">
                    {trans.asset.name}
                </a>
            '''

            # أزرار الإجراءات
            actions = []

            # رابط العرض
            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:transaction_detail', args=[trans.pk])}"
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            # رابط التعديل - فقط للمعاملات المسودة
            if can_edit and trans.status == 'draft':
                actions.append(f'''
                    <a href="{reverse('assets:transaction_update', args=[trans.pk])}"
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            # رابط الحذف - فقط للمعاملات المسودة
            if can_delete and trans.status == 'draft':
                actions.append(f'''
                    <button type="button" class="btn btn-outline-danger btn-sm"
                            onclick="deleteTransaction({trans.pk}, '{trans.transaction_number}')"
                            title="حذف" data-bs-toggle="tooltip">
                        <i class="fas fa-trash"></i>
                    </button>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                checkbox,
                trans.transaction_number,
                trans.transaction_date.strftime('%Y-%m-%d'),
                asset_link,
                type_badge,
                f'<strong>{trans.amount:,.2f}</strong>' if trans.amount else '-',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assettransfer')
@require_http_methods(["GET"])
def transfer_datatable_ajax(request):
    """Ajax endpoint لجدول التحويلات - Server-side"""

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        status = request.GET.get('status', '')
        search_filter = request.GET.get('search_filter', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        # Query
        queryset = AssetTransfer.objects.all().select_related(
            'asset', 'asset__category', 'from_branch', 'to_branch',
            'from_cost_center', 'to_cost_center', 'from_employee', 'to_employee'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if search_filter:
            queryset = queryset.filter(
                Q(transfer_number__icontains=search_filter) |
                Q(asset__code__icontains=search_filter) |
                Q(asset__name__icontains=search_filter) |
                Q(from_branch__name__icontains=search_filter) |
                Q(to_branch__name__icontains=search_filter)
            )

        if date_from:
            queryset = queryset.filter(transfer_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(transfer_date__lte=date_to)

        # البحث من DataTable
        if search_value:
            queryset = queryset.filter(
                Q(transfer_number__icontains=search_value) |
                Q(asset__code__icontains=search_value) |
                Q(asset__name__icontains=search_value)
            )

        # العد
        total_records = AssetTransfer.objects.count()
        filtered_records = queryset.count()

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = ['transfer_number', 'transfer_date', 'asset__code', 'from_branch__name', 'to_branch__name', 'status']

        if order_column_index:
            order_column = order_columns[int(order_column_index)]
            if order_dir == 'desc':
                order_column = '-' + order_column
            queryset = queryset.order_by(order_column)
        else:
            queryset = queryset.order_by('-transfer_date', '-transfer_number')

        # Pagination
        transfers = queryset[start:start + length]

        # إعداد البيانات
        data = []
        for transfer in transfers:
            # من
            from_html = f'''<div class="small">
                <i class="fas fa-building text-muted"></i> {transfer.from_branch.name}
            '''
            if transfer.from_cost_center:
                from_html += f'<br><i class="fas fa-sitemap text-muted"></i> {transfer.from_cost_center.name}'
            if transfer.from_employee:
                from_html += f'<br><i class="fas fa-user text-muted"></i> {transfer.from_employee.get_full_name()}'
            from_html += '</div>'

            # إلى
            to_html = f'''<div class="small">
                <i class="fas fa-building text-muted"></i> {transfer.to_branch.name}
            '''
            if transfer.to_cost_center:
                to_html += f'<br><i class="fas fa-sitemap text-muted"></i> {transfer.to_cost_center.name}'
            if transfer.to_employee:
                to_html += f'<br><i class="fas fa-user text-muted"></i> {transfer.to_employee.get_full_name()}'
            to_html += '</div>'

            # الحالة
            status_map = {
                'pending': '<span class="badge bg-warning"><i class="fas fa-clock"></i> قيد الانتظار</span>',
                'approved': '<span class="badge bg-info"><i class="fas fa-check"></i> معتمد</span>',
                'completed': '<span class="badge bg-success"><i class="fas fa-check-double"></i> مكتمل</span>',
                'rejected': '<span class="badge bg-danger"><i class="fas fa-times"></i> مرفوض</span>',
                'cancelled': '<span class="badge bg-secondary"><i class="fas fa-ban"></i> ملغي</span>',
            }
            status_badge = status_map.get(transfer.status, transfer.status)

            # الإجراءات
            actions_html = f'''
                <div class="btn-group btn-group-sm" role="group">
                    <a href="/assets/transfers/{transfer.pk}/view/"
                       class="btn btn-sm btn-outline-info"
                       data-bs-toggle="tooltip"
                       title="عرض">
                        <i class="fas fa-eye"></i>
                    </a>
            '''

            if transfer.status == 'pending' and request.user.has_perm('assets.change_assettransfer'):
                actions_html += f'''
                    <a href="/assets/transfers/{transfer.pk}/update/"
                       class="btn btn-sm btn-outline-primary"
                       data-bs-toggle="tooltip"
                       title="تعديل">
                        <i class="fas fa-edit"></i>
                    </a>
                    <button type="button"
                            class="btn btn-sm btn-outline-success"
                            onclick="approveTransfer({transfer.pk})"
                            data-bs-toggle="tooltip"
                            title="اعتماد">
                        <i class="fas fa-check"></i>
                    </button>
                    <button type="button"
                            class="btn btn-sm btn-outline-danger"
                            onclick="rejectTransfer({transfer.pk})"
                            data-bs-toggle="tooltip"
                            title="رفض">
                        <i class="fas fa-times"></i>
                    </button>
                '''

            actions_html += '</div>'

            data.append([
                f'<strong>{transfer.transfer_number}</strong>',
                transfer.transfer_date.strftime('%Y-%m-%d'),
                f'<a href="/assets/{transfer.asset.pk}/view/">{transfer.asset.code} - {transfer.asset.name}</a>',
                from_html,
                to_html,
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assettransfer')
@require_http_methods(["GET"])
def transfer_stats_ajax(request):
    """إحصائيات التحويلات - للبطاقات الإحصائية"""
    try:
        transfers = AssetTransfer.objects.all()

        stats = {
            'total_transfers': transfers.count(),
            'pending_transfers': transfers.filter(status='pending').count(),
            'approved_transfers': transfers.filter(status__in=['approved', 'completed']).count(),
            'rejected_transfers': transfers.filter(status='rejected').count(),
        }

        return JsonResponse({'success': True, 'stats': stats})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@permission_required_with_message('assets.view_assettransfer')
@require_http_methods(["GET"])
def transfer_export(request):
    """تصدير التحويلات إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    try:
        # الحصول على معايير الفلترة
        status = request.GET.get('status', '')
        search = request.GET.get('search', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        # بناء الاستعلام
        queryset = AssetTransfer.objects.all().select_related(
            'asset', 'from_branch', 'to_branch',
            'from_cost_center', 'to_cost_center',
            'from_employee', 'to_employee'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if search:
            queryset = queryset.filter(
                Q(transfer_number__icontains=search) |
                Q(asset__code__icontains=search) |
                Q(asset__name__icontains=search)
            )

        if date_from:
            queryset = queryset.filter(transfer_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(transfer_date__lte=date_to)

        queryset = queryset.order_by('-transfer_date')

        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تحويلات الأصول"

        # تنسيق الرأس
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # الرؤوس
        headers = [
            'رقم التحويل', 'التاريخ', 'رقم الأصل', 'اسم الأصل',
            'من الفرع', 'من مركز التكلفة', 'من الموظف',
            'إلى الفرع', 'إلى مركز التكلفة', 'إلى الموظف',
            'الحالة'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # البيانات
        row_num = 2
        status_names = dict(AssetTransfer.STATUS_CHOICES)

        for transfer in queryset:
            ws.cell(row=row_num, column=1, value=transfer.transfer_number)
            ws.cell(row=row_num, column=2, value=transfer.transfer_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=transfer.asset.code)
            ws.cell(row=row_num, column=4, value=transfer.asset.name)
            ws.cell(row=row_num, column=5, value=transfer.from_branch.name)
            ws.cell(row=row_num, column=6, value=transfer.from_cost_center.name if transfer.from_cost_center else '')
            ws.cell(row=row_num, column=7, value=transfer.from_employee.get_full_name() if transfer.from_employee else '')
            ws.cell(row=row_num, column=8, value=transfer.to_branch.name)
            ws.cell(row=row_num, column=9, value=transfer.to_cost_center.name if transfer.to_cost_center else '')
            ws.cell(row=row_num, column=10, value=transfer.to_employee.get_full_name() if transfer.to_employee else '')
            ws.cell(row=row_num, column=11, value=status_names.get(transfer.status, transfer.status))
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ الملف
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="asset_transfers_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    except Exception as e:
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:transfer_list')


@login_required
@permission_required_with_message('assets.can_approve_transactions')
@require_http_methods(["POST"])
def approve_transaction_ajax(request, pk):
    """اعتماد معاملة عبر Ajax - جديد"""

    try:
        trans = get_object_or_404(
            AssetTransaction,
            pk=pk,
            company=request.current_company,
            status='draft'
        )

        approval_notes = request.POST.get('approval_notes', '')

        trans.status = 'approved'
        trans.approved_by = request.user
        trans.approved_at = timezone.now()
        if approval_notes:
            trans.notes = f"{trans.notes}\nملاحظات: {approval_notes}" if trans.notes else f"ملاحظات: {approval_notes}"
        trans.save()

        return JsonResponse({
            'success': True,
            'message': f'تم اعتماد المعاملة {trans.transaction_number} بنجاح'
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في الاعتماد: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_assettransaction')
@require_http_methods(["POST"])
def cancel_transaction_ajax(request, pk):
    """إلغاء معاملة عبر Ajax - جديد"""

    try:
        trans = get_object_or_404(
            AssetTransaction,
            pk=pk,
            company=request.current_company
        )

        if trans.status not in ['draft', 'approved']:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن إلغاء هذه المعاملة'
            }, status=400)

        cancellation_reason = request.POST.get('cancellation_reason', '')

        trans.status = 'cancelled'
        trans.notes = f"{trans.notes}\nإلغاء: {cancellation_reason}" if trans.notes else f"إلغاء: {cancellation_reason}"
        trans.save()

        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء المعاملة {trans.transaction_number} بنجاح'
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في الإلغاء: {str(e)}'
        }, status=500)


# ==================== Export Functions - جديد ====================

@login_required
@permission_required_with_message('assets.view_assettransaction')
@require_http_methods(["GET"])
def export_transactions_excel(request):
    """تصدير المعاملات إلى Excel - جديد"""

    try:
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Transaction Number', 'Date', 'Type',
            'Asset Number', 'Asset Name', 'Amount',
            'Sale Price', 'Business Partner', 'Status'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # البيانات
        transactions = AssetTransaction.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'business_partner'
        ).order_by('-transaction_date')

        row_num = 2
        for trans in transactions:
            ws.cell(row=row_num, column=1, value=trans.transaction_number)
            ws.cell(row=row_num, column=2, value=trans.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=trans.get_transaction_type_display())
            ws.cell(row=row_num, column=4, value=trans.asset.asset_number)
            ws.cell(row=row_num, column=5, value=trans.asset.name)
            ws.cell(row=row_num, column=6, value=float(trans.amount))
            ws.cell(row=row_num, column=7, value=float(trans.sale_price) if trans.sale_price else '')
            ws.cell(row=row_num, column=8, value=trans.business_partner.name if trans.business_partner else '')
            ws.cell(row=row_num, column=9, value=trans.get_status_display())
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:transaction_list')


@login_required
@permission_required_with_message('assets.view_assettransfer')
@require_http_methods(["GET"])
def export_transfers_excel(request):
    """تصدير التحويلات إلى Excel - جديد"""

    try:
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Transfers"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Transfer Number', 'Date', 'Asset Number', 'Asset Name',
            'From Branch', 'To Branch', 'Status', 'Requested By'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # البيانات
        transfers = AssetTransfer.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'from_branch', 'to_branch', 'requested_by'
        ).order_by('-transfer_date')

        row_num = 2
        for transfer in transfers:
            ws.cell(row=row_num, column=1, value=transfer.transfer_number)
            ws.cell(row=row_num, column=2, value=transfer.transfer_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=transfer.asset.asset_number)
            ws.cell(row=row_num, column=4, value=transfer.asset.name)
            ws.cell(row=row_num, column=5, value=transfer.from_branch.name)
            ws.cell(row=row_num, column=6, value=transfer.to_branch.name)
            ws.cell(row=row_num, column=7, value=transfer.get_status_display())
            ws.cell(row=row_num, column=8, value=transfer.requested_by.get_full_name())
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transfers.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:transfer_list')


# ==================== Transaction AJAX Endpoints ====================

@login_required
@require_http_methods(["GET"])
def transaction_stats(request):
    """
    AJAX endpoint for loading transaction statistics
    """
    try:
        transactions = AssetTransaction.objects.filter(
            company=request.current_company
        )

        # Count by transaction type
        sell_count = transactions.filter(transaction_type='sale').count()
        dispose_count = transactions.filter(transaction_type='disposal').count()
        revalue_count = transactions.filter(transaction_type='revaluation').count()

        # Count this month
        today = date.today()
        monthly_transactions = transactions.filter(
            transaction_date__year=today.year,
            transaction_date__month=today.month
        ).count()

        stats = {
            'total_transactions': transactions.count(),
            'monthly_transactions': monthly_transactions,
            'sell_count': sell_count,
            'dispose_count': dispose_count,
            'revalue_count': revalue_count,
        }

        return JsonResponse(stats)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@permission_required('assets.view_assettransaction')
def transaction_export(request):
    """
    Export transactions to Excel
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Asset Transactions"

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'رقم المعاملة', 'التاريخ', 'رقم الأصل', 'اسم الأصل', 'النوع',
            'المبلغ', 'سعر البيع', 'الربح/الخسارة', 'الحالة', 'المستخدم'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Get transactions with filters from request
        transactions = AssetTransaction.objects.filter(
            company=request.current_company
        ).select_related('asset', 'created_by')

        # Apply filters from GET parameters
        transaction_type = request.GET.get('transaction_type')
        status = request.GET.get('status')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

        if transaction_type:
            transactions = transactions.filter(transaction_type=transaction_type)
        if status:
            transactions = transactions.filter(status=status)
        if date_from:
            transactions = transactions.filter(transaction_date__gte=date_from)
        if date_to:
            transactions = transactions.filter(transaction_date__lte=date_to)

        transactions = transactions.order_by('-transaction_date')

        # Transaction type mapping
        type_map = {
            'purchase': 'شراء',
            'sale': 'بيع',
            'disposal': 'استبعاد',
            'revaluation': 'إعادة تقييم',
            'impairment': 'هبوط قيمة'
        }

        # Status mapping
        status_map = {
            'draft': 'مسودة',
            'approved': 'معتمد',
            'completed': 'مكتمل',
            'cancelled': 'ملغي'
        }

        # Data rows
        row_num = 2
        for transaction in transactions:
            gain_loss = (transaction.sale_price or Decimal('0')) - (transaction.amount or Decimal('0'))

            ws.cell(row=row_num, column=1, value=transaction.transaction_number)
            ws.cell(row=row_num, column=2, value=transaction.transaction_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=transaction.asset.asset_number)
            ws.cell(row=row_num, column=4, value=transaction.asset.name)
            ws.cell(row=row_num, column=5, value=type_map.get(transaction.transaction_type, transaction.transaction_type))
            ws.cell(row=row_num, column=6, value=float(transaction.amount or 0))
            ws.cell(row=row_num, column=7, value=float(transaction.sale_price or 0))
            ws.cell(row=row_num, column=8, value=float(gain_loss))
            ws.cell(row=row_num, column=9, value=status_map.get(transaction.status, transaction.status))
            ws.cell(row=row_num, column=10, value=transaction.created_by.get_full_name() if transaction.created_by else '')
            row_num += 1

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save to response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="asset_transactions.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:transaction_list')