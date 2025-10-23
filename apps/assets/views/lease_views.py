# apps/assets/views/lease_views.py
"""
Views إدارة عقود الإيجار - محسّنة وشاملة
- إدارة عقود الإيجار التشغيلي والتمويلي
- إدارة دفعات الإيجار
- إنهاء وتجديد العقود
- تمرين خيار الشراء
- جداول الدفع
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    DetailView, TemplateView
)
from django.db.models import (
    Q, Sum, Count, Avg, Max, Min, F,
    DecimalField, Case, When, Value
)
from django.db.models.functions import Coalesce, TruncMonth
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
from dateutil.relativedelta import relativedelta

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import AssetLease, LeasePayment, Asset, AssetCategory
from apps.core.models import BusinessPartner


# ==================== Asset Leases ====================

class AssetLeaseListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة عقود الإيجار - محسّنة"""

    model = AssetLease
    template_name = 'assets/lease/lease_list.html'
    context_object_name = 'leases'
    permission_required = 'assets.view_assetlease'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetLease.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'lessor', 'branch'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        lease_type = self.request.GET.get('lease_type')
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        lessor = self.request.GET.get('lessor')
        expiring_soon = self.request.GET.get('expiring_soon')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if lease_type:
            queryset = queryset.filter(lease_type=lease_type)

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if lessor:
            queryset = queryset.filter(lessor_id=lessor)

        if expiring_soon == '1':
            expiry_date = date.today() + timedelta(days=90)
            queryset = queryset.filter(
                status='active',
                end_date__lte=expiry_date,
                end_date__gte=date.today()
            )

        if search:
            queryset = queryset.filter(
                Q(lease_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(lessor__name__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-start_date')
        queryset = queryset.order_by(sort_by, '-lease_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الموردين/المؤجرين
        lessors = BusinessPartner.objects.filter(
            company=self.request.current_company,
            partner_type__in=['supplier', 'both']
        ).order_by('name')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        leases = AssetLease.objects.filter(
            company=self.request.current_company
        )

        stats = leases.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=Q(status='active')),
            completed_count=Count('id', filter=Q(status='completed')),
            terminated_count=Count('id', filter=Q(status='terminated')),
            total_monthly=Coalesce(
                Sum('monthly_payment', filter=Q(status='active')),
                Decimal('0')
            ),
            total_deposits=Coalesce(
                Sum('security_deposit', filter=Q(status='active')),
                Decimal('0')
            ),
            avg_monthly=Coalesce(
                Avg('monthly_payment', filter=Q(status='active')),
                Decimal('0')
            ),
        )

        # العقود المنتهية قريباً (90 يوم)
        expiring_90 = leases.filter(
            status='active',
            end_date__lte=date.today() + timedelta(days=90),
            end_date__gte=date.today()
        ).count()

        # الدفعات المتأخرة
        overdue_payments = LeasePayment.objects.filter(
            lease__company=self.request.current_company,
            is_paid=False,
            payment_date__lt=date.today()
        ).count()

        context.update({
            'title': _('عقود الإيجار'),
            'can_add': self.request.user.has_perm('assets.add_assetlease'),
            'can_edit': self.request.user.has_perm('assets.change_assetlease'),
            'can_export': self.request.user.has_perm('assets.view_assetlease'),
            'status_choices': AssetLease.STATUS_CHOICES,
            'lease_types': AssetLease.LEASE_TYPES,
            'lessors': lessors,
            'categories': categories,
            'stats': stats,
            'expiring_90': expiring_90,
            'overdue_payments': overdue_payments,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': ''},
            ]
        })
        return context


class AssetLeaseCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء عقد إيجار - محسّن"""

    model = AssetLease
    template_name = 'assets/lease/lease_form.html'
    permission_required = 'assets.add_assetlease'
    fields = [
        'asset', 'lease_type', 'lessor',
        'start_date', 'end_date', 'payment_frequency',
        'monthly_payment', 'security_deposit',
        'interest_rate', 'residual_value', 'purchase_option_price',
        'status', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            is_leased=True
        ).select_related('category')

        form.fields['lessor'].queryset = BusinessPartner.objects.filter(
            company=company,
            partner_type__in=['supplier', 'both']
        ).order_by('name')

        # القيم الافتراضية
        form.fields['start_date'].initial = date.today()
        form.fields['end_date'].initial = date.today() + relativedelta(years=1)
        form.fields['payment_frequency'].initial = 'monthly'
        form.fields['status'].initial = 'draft'

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user

            self.object = form.save()

            # إنشاء دفعات الإيجار إذا كان العقد نشط
            if self.object.status == 'active':
                self.create_lease_payments()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء عقد الإيجار {self.object.lease_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def create_lease_payments(self):
        """إنشاء دفعات الإيجار المستقبلية - محسّن"""

        current_date = self.object.start_date
        payment_number = 1

        # حساب عدد الدفعات
        months_diff = (
                (self.object.end_date.year - self.object.start_date.year) * 12 +
                (self.object.end_date.month - self.object.start_date.month)
        )

        # تحديد عدد الدفعات حسب التكرار
        if self.object.payment_frequency == 'monthly':
            num_payments = months_diff + 1
        elif self.object.payment_frequency == 'quarterly':
            num_payments = (months_diff // 3) + 1
        elif self.object.payment_frequency == 'semi_annual':
            num_payments = (months_diff // 6) + 1
        elif self.object.payment_frequency == 'annual':
            num_payments = (months_diff // 12) + 1
        else:
            num_payments = months_diff + 1

        # حساب الفائدة للإيجار التمويلي
        if self.object.lease_type == 'finance' and self.object.interest_rate:
            remaining_principal = self.object.total_payments
            monthly_rate = self.object.interest_rate / 100 / 12
        else:
            remaining_principal = Decimal('0')
            monthly_rate = 0

        payments_created = 0

        while current_date <= self.object.end_date and payments_created < num_payments:
            # حساب الأصل والفائدة للإيجار التمويلي
            if self.object.lease_type == 'finance' and self.object.interest_rate:
                interest_amount = remaining_principal * Decimal(str(monthly_rate))
                principal_amount = self.object.monthly_payment - interest_amount
                remaining_principal -= principal_amount

                # التأكد من عدم السالب
                if remaining_principal < 0:
                    principal_amount += remaining_principal
                    remaining_principal = Decimal('0')
            else:
                interest_amount = Decimal('0')
                principal_amount = Decimal('0')

            LeasePayment.objects.create(
                lease=self.object,
                payment_number=payment_number,
                payment_date=current_date,
                amount=self.object.monthly_payment,
                principal_amount=principal_amount,
                interest_amount=interest_amount
            )

            payment_number += 1
            payments_created += 1

            # الانتقال للدفعة التالية
            if self.object.payment_frequency == 'monthly':
                current_date += relativedelta(months=1)
            elif self.object.payment_frequency == 'quarterly':
                current_date += relativedelta(months=3)
            elif self.object.payment_frequency == 'semi_annual':
                current_date += relativedelta(months=6)
            elif self.object.payment_frequency == 'annual':
                current_date += relativedelta(years=1)

    def get_success_url(self):
        return reverse('assets:lease_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة عقد إيجار'),
            'submit_text': _('إنشاء العقد'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class AssetLeaseDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل عقد الإيجار - محسّن"""

    model = AssetLease
    template_name = 'assets/lease/lease_detail.html'
    context_object_name = 'lease'
    permission_required = 'assets.view_assetlease'

    def get_queryset(self):
        return AssetLease.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'lessor', 'created_by', 'branch'
        ).prefetch_related('payments')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الدفعات
        payments = self.object.payments.select_related(
            'journal_entry'
        ).order_by('payment_number')

        # إحصائيات الدفعات
        payment_stats = payments.aggregate(
            total_count=Count('id'),
            paid_count=Count('id', filter=Q(is_paid=True)),
            unpaid_count=Count('id', filter=Q(is_paid=False)),
            overdue_count=Count(
                'id',
                filter=Q(is_paid=False, payment_date__lt=date.today())
            ),
            total_paid=Coalesce(Sum('amount', filter=Q(is_paid=True)), Decimal('0')),
            total_remaining=Coalesce(Sum('amount', filter=Q(is_paid=False)), Decimal('0')),
        )

        # نسبة الإنجاز
        if payment_stats['total_count'] > 0:
            completion_pct = (payment_stats['paid_count'] / payment_stats['total_count']) * 100
        else:
            completion_pct = 0

        # الدفعات المستحقة قريباً (30 يوم)
        upcoming_payments = payments.filter(
            is_paid=False,
            payment_date__gte=date.today(),
            payment_date__lte=date.today() + timedelta(days=30)
        ).order_by('payment_date')[:5]

        # الدفعات المتأخرة
        overdue_payments = payments.filter(
            is_paid=False,
            payment_date__lt=date.today()
        ).order_by('payment_date')[:5]

        # الأيام/الأشهر المتبقية
        if self.object.end_date:
            days_remaining = (self.object.end_date - date.today()).days
            months_remaining = self.object.get_remaining_months()
        else:
            days_remaining = None
            months_remaining = None

        # التحذيرات
        warnings = []
        if overdue_payments.exists():
            warnings.append({
                'type': 'danger',
                'icon': 'fa-exclamation-circle',
                'message': f'يوجد {overdue_payments.count()} دفعات متأخرة'
            })

        if self.object.is_expiring_soon():
            warnings.append({
                'type': 'warning',
                'icon': 'fa-clock',
                'message': f'العقد ينتهي قريباً في {self.object.end_date}'
            })

        context.update({
            'title': f'عقد الإيجار {self.object.lease_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_assetlease') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_activate': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.can_activate()  # ✅ استخدام method من Model
            ),
            'can_terminate': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.can_terminate()  # ✅ استخدام method من Model
            ),
            'can_renew': (
                    self.request.user.has_perm('assets.add_assetlease') and
                    self.object.status in ['active', 'completed']
            ),
            'can_purchase': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.lease_type == 'finance' and
                    self.object.status == 'active' and
                    self.object.purchase_option_price
            ),
            'payments': payments[:12],  # أول 12 دفعة
            'payment_stats': payment_stats,
            'completion_pct': round(completion_pct, 2),
            'upcoming_payments': upcoming_payments,
            'overdue_payments': overdue_payments,
            'days_remaining': days_remaining,
            'months_remaining': months_remaining,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': self.object.lease_number, 'url': ''},
            ]
        })
        return context


class AssetLeaseUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل عقد إيجار - محسّن"""

    model = AssetLease
    template_name = 'assets/lease/lease_form.html'
    permission_required = 'assets.change_assetlease'
    fields = [
        'asset', 'lease_type', 'lessor',
        'start_date', 'end_date', 'payment_frequency',
        'monthly_payment', 'security_deposit',
        'interest_rate', 'residual_value', 'purchase_option_price',
        'status', 'notes'
    ]

    def get_queryset(self):
        return AssetLease.objects.filter(
            company=self.request.current_company
        ).exclude(status__in=['completed', 'cancelled'])

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category')

        form.fields['lessor'].queryset = BusinessPartner.objects.filter(
            company=company,
            partner_type__in=['supplier', 'both']
        ).order_by('name')

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            # ✅ التحقق من إمكانية التعديل
            if not self.object.can_edit():
                messages.error(
                    self.request,
                    '❌ لا يمكن تعديل هذا العقد. قد يكون مكتمل أو منهي أو لديه دفعات مدفوعة'
                )
                return self.form_invalid(form)

            old_status = self.object.status
            self.object = form.save()

            # إنشاء الدفعات إذا تم تفعيل العقد
            if old_status == 'draft' and self.object.status == 'active':
                if not self.object.payments.exists():
                    self.create_lease_payments()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث عقد الإيجار {self.object.lease_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def create_lease_payments(self):
        """إنشاء دفعات الإيجار - نفس الطريقة من Create"""
        current_date = self.object.start_date
        payment_number = 1

        months_diff = (
                (self.object.end_date.year - self.object.start_date.year) * 12 +
                (self.object.end_date.month - self.object.start_date.month)
        )

        if self.object.payment_frequency == 'monthly':
            num_payments = months_diff + 1
        elif self.object.payment_frequency == 'quarterly':
            num_payments = (months_diff // 3) + 1
        elif self.object.payment_frequency == 'semi_annual':
            num_payments = (months_diff // 6) + 1
        elif self.object.payment_frequency == 'annual':
            num_payments = (months_diff // 12) + 1
        else:
            num_payments = months_diff + 1

        if self.object.lease_type == 'finance' and self.object.interest_rate:
            remaining_principal = self.object.total_payments
            monthly_rate = self.object.interest_rate / 100 / 12
        else:
            remaining_principal = Decimal('0')
            monthly_rate = 0

        payments_created = 0

        while current_date <= self.object.end_date and payments_created < num_payments:
            if self.object.lease_type == 'finance' and self.object.interest_rate:
                interest_amount = remaining_principal * Decimal(str(monthly_rate))
                principal_amount = self.object.monthly_payment - interest_amount
                remaining_principal -= principal_amount

                if remaining_principal < 0:
                    principal_amount += remaining_principal
                    remaining_principal = Decimal('0')
            else:
                interest_amount = Decimal('0')
                principal_amount = Decimal('0')

            LeasePayment.objects.create(
                lease=self.object,
                payment_number=payment_number,
                payment_date=current_date,
                amount=self.object.monthly_payment,
                principal_amount=principal_amount,
                interest_amount=interest_amount
            )

            payment_number += 1
            payments_created += 1

            if self.object.payment_frequency == 'monthly':
                current_date += relativedelta(months=1)
            elif self.object.payment_frequency == 'quarterly':
                current_date += relativedelta(months=3)
            elif self.object.payment_frequency == 'semi_annual':
                current_date += relativedelta(months=6)
            elif self.object.payment_frequency == 'annual':
                current_date += relativedelta(years=1)

    def get_success_url(self):
        return reverse('assets:lease_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل عقد الإيجار {self.object.lease_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': self.object.lease_number, 'url': reverse('assets:lease_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class TerminateLeaseView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """إنهاء عقد الإيجار - جديد"""

    template_name = 'assets/lease/terminate_lease.html'
    permission_required = 'assets.change_assetlease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        lease_id = self.kwargs.get('pk')
        lease = get_object_or_404(
            AssetLease,
            pk=lease_id,
            company=self.request.current_company,
            status='active'
        )

        # حساب الدفعات المتبقية
        remaining_payments = lease.payments.filter(is_paid=False)
        remaining_amount = remaining_payments.aggregate(
            total=Coalesce(Sum('amount'), Decimal('0'))
        )['total']

        context.update({
            'title': f'إنهاء عقد الإيجار {lease.lease_number}',
            'lease': lease,
            'remaining_payments_count': remaining_payments.count(),
            'remaining_amount': remaining_amount,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': lease.lease_number, 'url': reverse('assets:lease_detail', args=[lease.pk])},
                {'title': _('إنهاء العقد'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            lease_id = kwargs.get('pk')
            lease = get_object_or_404(
                AssetLease,
                pk=lease_id,
                company=request.current_company
            )

            termination_reason = request.POST.get('termination_reason', '')
            termination_date = request.POST.get('termination_date', date.today())

            # ✅ إنهاء العقد باستخدام model method
            lease.terminate(
                termination_date=termination_date,
                reason=termination_reason,
                user=request.user
            )

            # إلغاء الدفعات المتبقية
            remaining = lease.payments.filter(is_paid=False).update(
                notes='ملغي بسبب إنهاء العقد'
            )

            messages.success(
                request,
                f'✅ تم إنهاء عقد الإيجار {lease.lease_number} بنجاح. تم إلغاء {remaining} دفعات'
            )

            return redirect('assets:lease_detail', pk=lease.pk)

        except ValidationError as e:
            messages.error(request, f'❌ {str(e)}')
            return redirect('assets:lease_detail', pk=lease_id)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في إنهاء العقد: {str(e)}')
            return redirect('assets:lease_detail', pk=lease_id)


class RenewLeaseView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """تجديد عقد الإيجار - جديد"""

    template_name = 'assets/lease/renew_lease.html'
    permission_required = 'assets.add_assetlease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        lease_id = self.kwargs.get('pk')
        lease = get_object_or_404(
            AssetLease,
            pk=lease_id,
            company=self.request.current_company
        )

        # الفترة المقترحة (سنة واحدة)
        suggested_start = lease.end_date + timedelta(days=1)
        suggested_end = suggested_start + relativedelta(years=1)

        context.update({
            'title': f'تجديد عقد الإيجار {lease.lease_number}',
            'lease': lease,
            'suggested_start': suggested_start,
            'suggested_end': suggested_end,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': lease.lease_number, 'url': reverse('assets:lease_detail', args=[lease.pk])},
                {'title': _('تجديد'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            lease_id = kwargs.get('pk')
            old_lease = get_object_or_404(
                AssetLease,
                pk=lease_id,
                company=request.current_company
            )

            # البيانات الجديدة
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            monthly_payment = Decimal(request.POST.get('monthly_payment', old_lease.monthly_payment))

            # إنشاء العقد الجديد
            new_lease = AssetLease.objects.create(
                company=request.current_company,
                branch=request.current_branch,
                asset=old_lease.asset,
                lease_type=old_lease.lease_type,
                lessor=old_lease.lessor,
                start_date=start_date,
                end_date=end_date,
                payment_frequency=old_lease.payment_frequency,
                monthly_payment=monthly_payment,
                security_deposit=old_lease.security_deposit,
                interest_rate=old_lease.interest_rate,
                status='active',
                notes=f'تجديد من العقد {old_lease.lease_number}',
                created_by=request.user,
            )

            # إنشاء الدفعات
            self.create_lease_payments(new_lease)

            # تحديث العقد القديم
            if old_lease.status == 'active':
                old_lease.status = 'completed'
                old_lease.save()

            messages.success(
                request,
                f'✅ تم تجديد العقد بنجاح. رقم العقد الجديد: {new_lease.lease_number}'
            )

            return redirect('assets:lease_detail', pk=new_lease.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في التجديد: {str(e)}')
            return redirect('assets:lease_detail', pk=lease_id)

    def create_lease_payments(self, lease):
        """إنشاء دفعات للعقد الجديد"""
        current_date = lease.start_date
        payment_number = 1

        months_diff = (
                (lease.end_date.year - lease.start_date.year) * 12 +
                (lease.end_date.month - lease.start_date.month)
        )

        if lease.payment_frequency == 'monthly':
            num_payments = months_diff + 1
        elif lease.payment_frequency == 'quarterly':
            num_payments = (months_diff // 3) + 1
        elif lease.payment_frequency == 'semi_annual':
            num_payments = (months_diff // 6) + 1
        elif lease.payment_frequency == 'annual':
            num_payments = (months_diff // 12) + 1
        else:
            num_payments = months_diff + 1

        payments_created = 0

        while current_date <= lease.end_date and payments_created < num_payments:
            LeasePayment.objects.create(
                lease=lease,
                payment_number=payment_number,
                payment_date=current_date,
                amount=lease.monthly_payment,
                principal_amount=Decimal('0'),
                interest_amount=Decimal('0')
            )

            payment_number += 1
            payments_created += 1

            if lease.payment_frequency == 'monthly':
                current_date += relativedelta(months=1)
            elif lease.payment_frequency == 'quarterly':
                current_date += relativedelta(months=3)
            elif lease.payment_frequency == 'semi_annual':
                current_date += relativedelta(months=6)
            elif lease.payment_frequency == 'annual':
                current_date += relativedelta(years=1)


class ExercisePurchaseOptionView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """تمرين خيار الشراء - جديد"""

    template_name = 'assets/lease/exercise_purchase.html'
    permission_required = 'assets.change_assetlease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        lease_id = self.kwargs.get('pk')
        lease = get_object_or_404(
            AssetLease,
            pk=lease_id,
            company=self.request.current_company,
            status='active',
            lease_type='finance'
        )

        if not lease.purchase_option_price:
            messages.error(self.request, 'هذا العقد لا يحتوي على خيار شراء')
            return redirect('assets:lease_detail', pk=lease.pk)

        context.update({
            'title': f'تمرين خيار الشراء - {lease.lease_number}',
            'lease': lease,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('عقود الإيجار'), 'url': reverse('assets:lease_list')},
                {'title': lease.lease_number, 'url': reverse('assets:lease_detail', args=[lease.pk])},
                {'title': _('تمرين خيار الشراء'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            lease_id = kwargs.get('pk')
            lease = get_object_or_404(
                AssetLease,
                pk=lease_id,
                company=request.current_company,
                status='active',
                lease_type='finance'
            )

            purchase_date = request.POST.get('purchase_date', date.today())

            # تحديث الأصل
            asset = lease.asset
            asset.is_leased = False
            asset.original_cost = lease.purchase_option_price
            asset.purchase_date = purchase_date
            asset.save()

            # إكمال العقد
            lease.status = 'completed'
            lease.notes = f"{lease.notes}\nتم تمرين خيار الشراء بتاريخ {purchase_date}" if lease.notes else f"تم تمرين خيار الشراء بتاريخ {purchase_date}"
            lease.save()

            messages.success(
                request,
                f'✅ تم تمرين خيار الشراء بنجاح. الأصل {asset.asset_number} أصبح ملكاً للشركة'
            )

            return redirect('assets:asset_detail', pk=asset.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في تمرين خيار الشراء: {str(e)}')
            return redirect('assets:lease_detail', pk=lease_id)


# ==================== Lease Payments ====================

class LeasePaymentListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة دفعات الإيجار - محسّنة"""

    model = LeasePayment
    template_name = 'assets/lease/payment_list.html'
    context_object_name = 'payments'
    permission_required = 'assets.view_assetlease'
    paginate_by = 50

    def get_queryset(self):
        queryset = LeasePayment.objects.filter(
            lease__company=self.request.current_company
        ).select_related(
            'lease', 'lease__asset', 'lease__asset__category', 'journal_entry'
        )

        # الفلترة المتقدمة
        is_paid = self.request.GET.get('is_paid')
        lease = self.request.GET.get('lease')
        overdue = self.request.GET.get('overdue')
        due_soon = self.request.GET.get('due_soon')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if is_paid:
            queryset = queryset.filter(is_paid=(is_paid == '1'))

        if lease:
            queryset = queryset.filter(lease_id=lease)

        if overdue == '1':
            queryset = queryset.filter(
                is_paid=False,
                payment_date__lt=date.today()
            )

        if due_soon == '1':
            queryset = queryset.filter(
                is_paid=False,
                payment_date__gte=date.today(),
                payment_date__lte=date.today() + timedelta(days=30)
            )

        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(lease__lease_number__icontains=search) |
                Q(lease__asset__asset_number__icontains=search) |
                Q(lease__asset__name__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', 'payment_date')
        queryset = queryset.order_by(sort_by, 'lease', 'payment_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # العقود النشطة
        active_leases = AssetLease.objects.filter(
            company=self.request.current_company,
            status='active'
        ).select_related('asset')

        # إحصائيات مفصّلة
        payments = LeasePayment.objects.filter(
            lease__company=self.request.current_company
        )

        stats = payments.aggregate(
            total_count=Count('id'),
            paid_count=Count('id', filter=Q(is_paid=True)),
            unpaid_count=Count('id', filter=Q(is_paid=False)),
            overdue_count=Count(
                'id',
                filter=Q(is_paid=False, payment_date__lt=date.today())
            ),
            due_soon_count=Count(
                'id',
                filter=Q(
                    is_paid=False,
                    payment_date__gte=date.today(),
                    payment_date__lte=date.today() + timedelta(days=30)
                )
            ),
            total_paid=Coalesce(Sum('amount', filter=Q(is_paid=True)), Decimal('0')),
            total_unpaid=Coalesce(Sum('amount', filter=Q(is_paid=False)), Decimal('0')),
            total_overdue=Coalesce(
                Sum('amount', filter=Q(is_paid=False, payment_date__lt=date.today())),
                Decimal('0')
            ),
        )

        context.update({
            'title': _('دفعات الإيجار'),
            'can_pay': self.request.user.has_perm('assets.change_assetlease'),
            'can_export': self.request.user.has_perm('assets.view_assetlease'),
            'active_leases': active_leases,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('دفعات الإيجار'), 'url': ''},
            ]
        })
        return context


class LeasePaymentDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل دفعة الإيجار - جديد"""

    model = LeasePayment
    template_name = 'assets/lease/payment_detail.html'
    context_object_name = 'payment'
    permission_required = 'assets.view_assetlease'

    def get_queryset(self):
        return LeasePayment.objects.filter(
            lease__company=self.request.current_company
        ).select_related(
            'lease', 'lease__asset', 'lease__lessor',
            'journal_entry', 'paid_by'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الدفعات السابقة واللاحقة
        previous = LeasePayment.objects.filter(
            lease=self.object.lease,
            payment_number__lt=self.object.payment_number
        ).order_by('-payment_number').first()

        next_payment = LeasePayment.objects.filter(
            lease=self.object.lease,
            payment_number__gt=self.object.payment_number
        ).order_by('payment_number').first()

        # الأيام المتأخرة
        if not self.object.is_paid and self.object.payment_date < date.today():
            days_overdue = (date.today() - self.object.payment_date).days
        else:
            days_overdue = None

        context.update({
            'title': f'الدفعة رقم {self.object.payment_number} - {self.object.lease.lease_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.can_edit()  # ✅ استخدام method من Model
            ),
            'can_delete': (
                    self.request.user.has_perm('assets.delete_assetlease') and
                    self.object.can_delete()  # ✅ استخدام method من Model
            ),
            'can_pay': (
                    self.request.user.has_perm('assets.change_assetlease') and
                    self.object.can_pay()  # ✅ استخدام method من Model
            ),
            'previous': previous,
            'next_payment': next_payment,
            'days_overdue': days_overdue,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('دفعات الإيجار'), 'url': reverse('assets:payment_list')},
                {'title': f'دفعة {self.object.payment_number}', 'url': ''},
            ]
        })
        return context


class LeasePaymentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, CreateView):
    """إضافة دفعة إيجار يدوياً - محسّن"""

    model = LeasePayment
    template_name = 'assets/lease/payment_form.html'
    permission_required = 'assets.change_assetlease'
    fields = [
        'lease', 'payment_number', 'payment_date',
        'amount', 'principal_amount', 'interest_amount',
        'notes'
    ]
    success_url = reverse_lazy('assets:payment_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        form.fields['lease'].queryset = AssetLease.objects.filter(
            company=self.request.current_company,
            status='active'
        ).select_related('asset')

        form.fields['payment_date'].initial = date.today()

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            self.object = form.save()

            messages.success(
                self.request,
                f'✅ تم إضافة دفعة الإيجار رقم {self.object.payment_number} بنجاح'
            )

            return redirect(self.success_url)

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة دفعة إيجار'),
            'submit_text': _('إضافة الدفعة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('دفعات الإيجار'), 'url': reverse('assets:payment_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


# ==================== Ajax Views - محسّنة ====================

@login_required
@permission_required_with_message('assets.change_assetlease')
@require_http_methods(["POST"])
def process_lease_payment(request, pk):
    """معالجة دفعة إيجار - محسّن"""

    try:
        payment = get_object_or_404(
            LeasePayment,
            pk=pk,
            lease__company=request.current_company,
            is_paid=False
        )

        paid_date = request.POST.get('paid_date', date.today())
        notes = request.POST.get('notes', '')

        # معالجة الدفع
        payment.process_payment(
            user=request.user,
            paid_date=paid_date,
            notes=notes
        )

        return JsonResponse({
            'success': True,
            'message': f'تم معالجة دفعة الإيجار رقم {payment.payment_number} بنجاح',
            'payment_number': payment.payment_number,
            'amount': float(payment.amount),
            'journal_entry_number': payment.journal_entry.number if payment.journal_entry else None,
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في معالجة الدفع: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_assetlease')
@require_http_methods(["POST"])
def bulk_process_payments(request):
    """معالجة دفعات جماعية - جديد"""

    try:
        payment_ids = request.POST.getlist('payment_ids[]')

        if not payment_ids:
            return JsonResponse({
                'success': False,
                'message': 'لم يتم تحديد دفعات'
            }, status=400)

        success_count = 0
        error_count = 0
        errors = []

        for payment_id in payment_ids:
            try:
                payment = LeasePayment.objects.get(
                    pk=payment_id,
                    lease__company=request.current_company,
                    is_paid=False
                )
                payment.process_payment(user=request.user)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f'الدفعة {payment_id}: {str(e)}')

        return JsonResponse({
            'success': True,
            'message': f'تم معالجة {success_count} دفعات بنجاح',
            'processed': success_count,
            'errors': error_count,
            'error_details': errors[:5],
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في المعالجة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetlease')
@require_http_methods(["GET"])
def lease_datatable_ajax(request):
    """Ajax endpoint لجدول عقود الإيجار - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        status = request.GET.get('status', '')
        lease_type = request.GET.get('lease_type', '')

        # Query
        queryset = AssetLease.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'asset__category', 'lessor'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if lease_type:
            queryset = queryset.filter(lease_type=lease_type)

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(lease_number__icontains=search_value) |
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value) |
                Q(lessor__name__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'lease_number',
            '1': 'asset__asset_number',
            '2': 'lease_type',
            '3': 'lessor__name',
            '4': 'monthly_payment',
            '5': 'start_date',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-lease_number')
        else:
            queryset = queryset.order_by('-start_date', '-lease_number')

        # العد
        total_records = AssetLease.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_assetlease')

        for lease in queryset:
            # الحالة
            status_map = {
                'draft': '<span class="badge bg-secondary"><i class="fas fa-file"></i> مسودة</span>',
                'active': '<span class="badge bg-success"><i class="fas fa-check-circle"></i> نشط</span>',
                'completed': '<span class="badge bg-info"><i class="fas fa-flag-checkered"></i> مكتمل</span>',
                'terminated': '<span class="badge bg-warning"><i class="fas fa-stop-circle"></i> منهي</span>',
                'cancelled': '<span class="badge bg-danger"><i class="fas fa-ban"></i> ملغي</span>',
            }
            status_badge = status_map.get(lease.status, lease.status)

            # التحذيرات
            warnings = []
            if lease.is_expiring_soon():
                days_left = (lease.end_date - date.today()).days
                warnings.append(f'<small class="text-warning"><i class="fas fa-clock"></i> {days_left} يوم</small>')

            # الأشهر المتبقية
            remaining_months = lease.get_remaining_months()

            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:lease_detail', args=[lease.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                f'<a href="{reverse("assets:lease_detail", args=[lease.pk])}">{lease.lease_number}</a>',
                f'''<div>
                    <strong><a href="{reverse("assets:asset_detail", args=[lease.asset.pk])}">{lease.asset.asset_number}</a></strong>
                    <br><small class="text-muted">{lease.asset.name}</small>
                </div>''',
                lease.get_lease_type_display(),
                f'''<div>
                    <strong>{lease.lessor.name}</strong>
                    <br><small class="text-muted">{lease.lessor.phone or ''}</small>
                </div>''',
                f'<div class="text-end"><strong>{lease.monthly_payment:,.2f}</strong></div>',
                f'''<div>
                    {remaining_months} شهر
                    {'<br>' + ' '.join(warnings) if warnings else ''}
                </div>''',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetlease')
@require_http_methods(["GET"])
def payment_datatable_ajax(request):
    """Ajax endpoint لجدول دفعات الإيجار - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        is_paid = request.GET.get('is_paid', '')

        # Query
        queryset = LeasePayment.objects.filter(
            lease__company=request.current_company
        ).select_related(
            'lease__asset', 'lease__lessor', 'journal_entry', 'paid_by'
        )

        # تطبيق الفلاتر
        if is_paid:
            queryset = queryset.filter(is_paid=(is_paid == '1'))

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(lease__lease_number__icontains=search_value) |
                Q(lease__asset__asset_number__icontains=search_value) |
                Q(lease__asset__name__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'asc')

        order_columns = {
            '0': 'lease__lease_number',
            '1': 'payment_number',
            '2': 'payment_date',
            '3': 'lease__asset__asset_number',
            '4': 'amount',
            '5': 'paid_date',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field)
        else:
            queryset = queryset.order_by('payment_date', 'lease', 'payment_number')

        # العد
        total_records = LeasePayment.objects.filter(
            lease__company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_pay = request.user.has_perm('assets.change_assetlease')
        can_view_journal = request.user.has_perm('accounting.view_journalentry')

        for payment in queryset:
            # الحالة
            if payment.is_paid:
                status_badge = '<span class="badge bg-success"><i class="fas fa-check-circle"></i> مدفوع</span>'
            elif payment.payment_date < date.today():
                days_overdue = (date.today() - payment.payment_date).days
                status_badge = f'<span class="badge bg-danger"><i class="fas fa-exclamation-circle"></i> متأخر ({days_overdue} يوم)</span>'
            else:
                days_left = (payment.payment_date - date.today()).days
                status_badge = f'<span class="badge bg-warning"><i class="fas fa-clock"></i> مستحق ({days_left} يوم)</span>'

            # أزرار الإجراءات
            actions = []

            if not payment.is_paid and can_pay:
                actions.append(f'''
                    <button type="button" class="btn btn-outline-success btn-sm" 
                            onclick="processPayment({payment.pk})" title="دفع" data-bs-toggle="tooltip">
                        <i class="fas fa-dollar-sign"></i>
                    </button>
                ''')

            if payment.journal_entry and can_view_journal:
                actions.append(f'''
                    <a href="{reverse('accounting:journal_entry_detail', args=[payment.journal_entry.pk])}" 
                       class="btn btn-outline-info btn-sm" title="القيد" data-bs-toggle="tooltip">
                        <i class="fas fa-file-invoice"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                f'<a href="{reverse("assets:lease_detail", args=[payment.lease.pk])}">{payment.lease.lease_number}</a>',
                f'<strong>قسط {payment.payment_number}</strong>',
                payment.payment_date.strftime('%Y-%m-%d'),
                f'<a href="{reverse("assets:asset_detail", args=[payment.lease.asset.pk])}">{payment.lease.asset.asset_number}</a>',
                f'<div class="text-end"><strong>{payment.amount:,.2f}</strong></div>',
                f'{payment.paid_date.strftime("%Y-%m-%d")}' if payment.paid_date else '<span class="text-muted">-</span>',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetlease')
@require_http_methods(["GET"])
def lease_payment_schedule_ajax(request):
    """جدول دفعات عقد معين - جديد"""

    try:
        lease_id = request.GET.get('lease_id')

        if not lease_id:
            return JsonResponse({
                'success': False,
                'error': 'يجب تحديد العقد'
            }, status=400)

        lease = get_object_or_404(
            AssetLease,
            pk=lease_id,
            company=request.current_company
        )

        payments = lease.payments.order_by('payment_number')

        schedule = []
        for payment in payments:
            schedule.append({
                'payment_number': payment.payment_number,
                'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
                'amount': float(payment.amount),
                'principal_amount': float(payment.principal_amount),
                'interest_amount': float(payment.interest_amount),
                'is_paid': payment.is_paid,
                'paid_date': payment.paid_date.strftime('%Y-%m-%d') if payment.paid_date else None,
            })

        return JsonResponse({
            'success': True,
            'lease_number': lease.lease_number,
            'schedule': schedule,
            'total_payments': len(schedule),
            'total_amount': sum(p['amount'] for p in schedule),
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


# ==================== Export Functions - جديد ====================

@login_required
@permission_required_with_message('assets.view_assetlease')
@require_http_methods(["GET"])
def export_lease_list_excel(request):
    """تصدير قائمة العقود إلى Excel - جديد"""

    try:
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leases"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Lease Number', 'Asset Number', 'Asset Name',
            'Lease Type', 'Lessor', 'Monthly Payment',
            'Start Date', 'End Date', 'Status'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # البيانات
        leases = AssetLease.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'lessor'
        ).order_by('-start_date')

        row_num = 2
        for lease in leases:
            ws.cell(row=row_num, column=1, value=lease.lease_number)
            ws.cell(row=row_num, column=2, value=lease.asset.asset_number)
            ws.cell(row=row_num, column=3, value=lease.asset.name)
            ws.cell(row=row_num, column=4, value=lease.get_lease_type_display())
            ws.cell(row=row_num, column=5, value=lease.lessor.name)
            ws.cell(row=row_num, column=6, value=float(lease.monthly_payment))
            ws.cell(row=row_num, column=7, value=lease.start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=8, value=lease.end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=9, value=lease.get_status_display())
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leases.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:lease_list')