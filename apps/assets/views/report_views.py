# apps/assets/views/report_views.py
"""
Views التقارير
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Sum, Count, Avg, F, Max, Min
from django.utils.translation import gettext_lazy as _
from django.db.models.functions import TruncMonth, TruncYear
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.core.decorators import permission_required_with_message
from ..models import (
    Asset, AssetCategory, AssetDepreciation, AssetMaintenance,
    AssetTransaction, PhysicalCount, AssetInsurance
)


@login_required
@permission_required_with_message('assets.view_asset')
def asset_register_report(request):
    """تقرير سجل الأصول"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    category = request.GET.get('category')
    branch = request.GET.get('branch')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    assets = Asset.objects.filter(
        company=request.current_company
    ).select_related('category', 'branch', 'condition')

    if category:
        assets = assets.filter(category_id=category)

    if branch:
        assets = assets.filter(branch_id=branch)

    if status:
        assets = assets.filter(status=status)

    if date_from:
        assets = assets.filter(purchase_date__gte=date_from)

    if date_to:
        assets = assets.filter(purchase_date__lte=date_to)

    # الإحصائيات
    stats = assets.aggregate(
        total_count=Count('id'),
        total_cost=Sum('original_cost'),
        total_accumulated=Sum('accumulated_depreciation'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation'))
    )

    # حسب الفئة
    by_category = assets.values(
        'category__name'
    ).annotate(
        count=Count('id'),
        total_cost=Sum('original_cost'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation'))
    ).order_by('-total_cost')

    # حسب الحالة
    by_status = assets.values('status').annotate(
        count=Count('id'),
        total_cost=Sum('original_cost')
    ).order_by('-count')

    context = {
        'title': _('تقرير سجل الأصول'),
        'assets': assets.order_by('category', 'asset_number')[:100],
        'stats': stats,
        'by_category': by_category,
        'by_status': by_status,
        'categories': AssetCategory.objects.filter(is_active=True),
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('سجل الأصول'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/asset_register.html', context)


@login_required
@permission_required_with_message('assets.view_asset')
def depreciation_report(request):
    """تقرير الإهلاك"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    category_id = request.GET.get('category', '')

    # البيانات الأساسية
    depreciation_records = AssetDepreciation.objects.filter(
        company=request.current_company
    ).select_related('asset', 'asset__category', 'asset__depreciation_method')

    # تطبيق الفلاتر
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            depreciation_records = depreciation_records.filter(depreciation_date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            depreciation_records = depreciation_records.filter(depreciation_date__lte=date_to_obj)
        except ValueError:
            pass

    if category_id:
        depreciation_records = depreciation_records.filter(asset__category_id=category_id)

    # الإحصائيات العامة
    stats = depreciation_records.aggregate(
        total_depreciation=Sum('depreciation_amount'),
        total_accumulated=Sum('accumulated_depreciation_after'),
        count=Count('id')
    )

    # حساب التكلفة الأصلية والقيمة الدفترية من الأصول المرتبطة
    assets_with_depreciation = Asset.objects.filter(
        company=request.current_company,
        id__in=depreciation_records.values_list('asset_id', flat=True).distinct()
    ).aggregate(
        total_cost=Sum('original_cost'),
        total_book_value=Sum('book_value')
    )

    # دمج الإحصائيات
    totals = {
        'period_depreciation': stats.get('total_depreciation') or Decimal('0'),
        'accumulated_depreciation': stats.get('total_accumulated') or Decimal('0'),
        'original_cost': assets_with_depreciation.get('total_cost') or Decimal('0'),
        'book_value': assets_with_depreciation.get('total_book_value') or Decimal('0'),
        'total_records': stats.get('count') or 0
    }

    # حسب الفئة
    by_category = depreciation_records.values(
        'asset__category__name'
    ).annotate(
        count=Count('id'),
        period_depreciation=Sum('depreciation_amount'),
        accumulated_depreciation=Sum('accumulated_depreciation_after')
    ).order_by('-period_depreciation')

    # إضافة التكلفة والقيمة الدفترية لكل فئة
    for item in by_category:
        category_assets = Asset.objects.filter(
            company=request.current_company,
            category__name=item['asset__category__name']
        ).aggregate(
            total_cost=Sum('original_cost'),
            total_book_value=Sum('book_value')
        )
        item['original_cost'] = category_assets.get('total_cost') or Decimal('0')
        item['book_value'] = category_assets.get('total_book_value') or Decimal('0')

        # حساب نسبة الإهلاك
        if item['original_cost']:
            item['depreciation_percentage'] = (item['accumulated_depreciation'] / item['original_cost']) * 100
        else:
            item['depreciation_percentage'] = 0

    # الفئات للفلتر
    categories = AssetCategory.objects.filter(
        is_active=True
    ).order_by('name')

    # البيانات للمخطط
    category_labels = [item['asset__category__name'] for item in by_category[:10]]
    category_period_data = [float(item['period_depreciation']) for item in by_category[:10]]
    category_accumulated_data = [float(item['accumulated_depreciation']) for item in by_category[:10]]

    context = {
        'title': _('تقرير الإهلاك'),
        'depreciation_records': depreciation_records.order_by('-depreciation_date')[:100],
        'totals': totals,
        'by_category': by_category,
        'categories': categories,
        'date_from': date_from,
        'date_to': date_to,
        'selected_category': category_id,
        'category_labels': json.dumps(category_labels),
        'category_period_data': json.dumps(category_period_data),
        'category_accumulated_data': json.dumps(category_accumulated_data),
        'today': date.today(),
        'now': datetime.now(),
        'user': request.user,
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('تقرير الإهلاك'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/depreciation_report.html', context)


@login_required
@permission_required_with_message('assets.view_asset')
def maintenance_report(request):
    """تقرير الصيانة"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    maintenance_type_id = request.GET.get('maintenance_type', '')
    status_filter = request.GET.get('status', '')

    # البيانات الأساسية
    maintenances = AssetMaintenance.objects.filter(
        company=request.current_company
    ).select_related('asset', 'asset__category', 'maintenance_type', 'asset__branch')

    # تطبيق الفلاتر
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            maintenances = maintenances.filter(scheduled_date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            maintenances = maintenances.filter(scheduled_date__lte=date_to_obj)
        except ValueError:
            pass

    if maintenance_type_id:
        maintenances = maintenances.filter(maintenance_type_id=maintenance_type_id)

    if status_filter:
        maintenances = maintenances.filter(status=status_filter)

    # الإحصائيات العامة
    from django.db.models import Q, Case, When, IntegerField

    stats = maintenances.aggregate(
        total_count=Count('id'),
        completed_count=Count('id', filter=Q(status='completed')),
        in_progress_count=Count('id', filter=Q(status='in_progress')),
        scheduled_count=Count('id', filter=Q(status='scheduled')),
        total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
        avg_cost=Avg(F('labor_cost') + F('parts_cost') + F('other_cost'))
    )

    # دمج الإحصائيات
    totals = {
        'total_count': stats.get('total_count') or 0,
        'completed_count': stats.get('completed_count') or 0,
        'in_progress_count': stats.get('in_progress_count') or 0,
        'scheduled_count': stats.get('scheduled_count') or 0,
        'total_cost': stats.get('total_cost') or Decimal('0'),
        'avg_cost': stats.get('avg_cost') or Decimal('0')
    }

    # حسب النوع
    by_type = maintenances.values(
        'maintenance_type__name'
    ).annotate(
        count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        scheduled=Count('id', filter=Q(status='scheduled')),
        total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
        avg_cost=Avg(F('labor_cost') + F('parts_cost') + F('other_cost'))
    ).order_by('-total_cost')

    # حسب الأصل
    by_asset = maintenances.values(
        'asset__asset_number',
        'asset__name',
        'asset__category__name'
    ).annotate(
        count=Count('id'),
        total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
        last_maintenance=Max('completion_date'),
    ).order_by('-total_cost')[:20]

    # أنواع الصيانة للفلتر
    from ..models import MaintenanceType
    maintenance_types = MaintenanceType.objects.filter(
        is_active=True
    ).order_by('name')

    # البيانات للمخطط - حسب الشهر
    from django.db.models.functions import TruncMonth
    by_month = maintenances.annotate(
        month=TruncMonth('scheduled_date')
    ).values('month').annotate(
        count=Count('id'),
        total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost'))
    ).order_by('month')[:12]  # آخر 12 شهر

    month_labels = [item['month'].strftime('%Y-%m') if item['month'] else '' for item in by_month]
    month_counts = [item['count'] for item in by_month]
    month_costs = [float(item['total_cost'] or 0) for item in by_month]

    # Fetch maintenance records for display
    # Create a fresh queryset to avoid conflicts with aggregations
    maintenance_records = AssetMaintenance.objects.filter(
        company=request.current_company
    ).select_related('asset', 'asset__category', 'maintenance_type', 'asset__branch')

    # Apply same filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            maintenance_records = maintenance_records.filter(scheduled_date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            maintenance_records = maintenance_records.filter(scheduled_date__lte=date_to_obj)
        except ValueError:
            pass

    if maintenance_type_id:
        maintenance_records = maintenance_records.filter(maintenance_type_id=maintenance_type_id)

    if status_filter:
        maintenance_records = maintenance_records.filter(status=status_filter)

    # Order and convert to list to force evaluation
    maintenance_records = list(maintenance_records.order_by('-scheduled_date', '-id')[:100])

    # Debug count
    records_count = len(maintenance_records)

    context = {
        'title': _('تقرير الصيانة'),
        'maintenance_records': maintenance_records,
        'records_count': records_count,  # Add for debugging
        'totals': totals,
        'by_type': by_type,
        'by_asset': by_asset,
        'maintenance_types': maintenance_types,
        'date_from': date_from,
        'date_to': date_to,
        'selected_maintenance_type': maintenance_type_id,
        'selected_status': status_filter,
        'month_labels': json.dumps(month_labels),
        'month_counts': json.dumps(month_counts),
        'month_costs': json.dumps(month_costs),
        'today': date.today(),
        'now': datetime.now(),
        'user': request.user,
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('تقرير الصيانة'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/maintenance_report.html', context)


@login_required
@permission_required_with_message('assets.view_asset')
def asset_movement_report(request):
    """تقرير حركة الأصول"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    transaction_type = request.GET.get('transaction_type')

    transactions = AssetTransaction.objects.filter(
        company=request.current_company
    ).select_related('asset', 'business_partner')

    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)

    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)

    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)

    # الإحصائيات
    stats = transactions.aggregate(
        total_amount=Sum('amount'),
        count=Count('id')
    )

    # حسب النوع
    by_type = transactions.values('transaction_type').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('-count')

    context = {
        'title': _('تقرير حركة الأصول'),
        'transactions': transactions.order_by('-transaction_date')[:100],
        'stats': stats,
        'by_type': by_type,
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('حركة الأصول'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/movement_report.html', context)


@login_required
@permission_required_with_message('assets.view_asset')
def valuation_report(request):
    """تقرير تقييم الأصول"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    category_id = request.GET.get('category', '')
    branch_id = request.GET.get('branch', '')
    status_filter = request.GET.get('status', 'active')

    # البيانات الأساسية
    assets = Asset.objects.filter(
        company=request.current_company
    ).select_related('category', 'branch', 'depreciation_method', 'condition')

    # تطبيق الفلاتر
    if status_filter:
        assets = assets.filter(status=status_filter)

    if category_id:
        assets = assets.filter(category_id=category_id)

    if branch_id:
        assets = assets.filter(branch_id=branch_id)

    # الإحصائيات العامة
    totals = assets.aggregate(
        count=Count('id'),
        total_cost=Sum('original_cost'),
        total_accumulated=Sum('accumulated_depreciation'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation')),
        avg_cost=Avg('original_cost'),
        avg_book_value=Avg(F('original_cost') - F('accumulated_depreciation'))
    )

    # دمج الإحصائيات
    totals = {
        'count': totals.get('count') or 0,
        'total_cost': totals.get('total_cost') or Decimal('0'),
        'total_accumulated': totals.get('total_accumulated') or Decimal('0'),
        'total_book_value': totals.get('total_book_value') or Decimal('0'),
        'avg_cost': totals.get('avg_cost') or Decimal('0'),
        'avg_book_value': totals.get('avg_book_value') or Decimal('0'),
    }

    # حسب الفئة مع النسب المئوية
    by_category = assets.values(
        'category__name'
    ).annotate(
        count=Count('id'),
        total_cost=Sum('original_cost'),
        total_accumulated=Sum('accumulated_depreciation'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation')),
        avg_cost=Avg('original_cost')
    ).order_by('-total_book_value')

    # حساب النسب المئوية
    total_value = totals['total_book_value']
    for cat in by_category:
        if total_value and total_value > 0:
            cat['percentage'] = (float(cat['total_book_value']) / float(total_value)) * 100
        else:
            cat['percentage'] = 0

    # حسب الفرع
    by_branch = assets.values(
        'branch__name'
    ).annotate(
        count=Count('id'),
        total_cost=Sum('original_cost'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation'))
    ).order_by('-total_book_value')

    # حساب النسب المئوية للفروع
    for branch in by_branch:
        if total_value and total_value > 0:
            branch['percentage'] = (float(branch['total_book_value']) / float(total_value)) * 100
        else:
            branch['percentage'] = 0

    # الفئات وال فروع للفلاتر
    categories = AssetCategory.objects.filter(
        company=request.current_company,
        is_active=True
    ).order_by('name')

    from apps.core.models import Branch
    branches = Branch.objects.filter(
        company=request.current_company,
        is_active=True
    ).order_by('name')

    # البيانات للمخططات
    category_labels = [cat['category__name'] for cat in by_category[:8]]
    category_values = [float(cat['total_book_value']) for cat in by_category[:8]]
    category_costs = [float(cat['total_cost']) for cat in by_category[:8]]

    context = {
        'title': _('تقرير تقييم الأصول'),
        'assets': list(assets.order_by('-book_value')[:100]),
        'totals': totals,
        'by_category': by_category,
        'by_branch': by_branch,
        'categories': categories,
        'branches': branches,
        'selected_category': category_id,
        'selected_branch': branch_id,
        'selected_status': status_filter,
        'category_labels': json.dumps(category_labels),
        'category_values': json.dumps(category_values),
        'category_costs': json.dumps(category_costs),
        'today': date.today(),
        'now': datetime.now(),
        'user': request.user,
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('تقرير تقييم الأصول'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/valuation_report.html', context)


@login_required
@permission_required_with_message('assets.can_conduct_physical_count')
def physical_count_report(request):
    """تقرير الجرد الفعلي"""

    if not hasattr(request, 'current_company') or not request.current_company:
        messages.error(request, 'لا توجد شركة محددة')
        return redirect('assets:dashboard')

    # الفلاتر
    cycle_id = request.GET.get('cycle')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    counts = PhysicalCount.objects.filter(
        company=request.current_company
    ).select_related('cycle', 'branch')

    if cycle_id:
        counts = counts.filter(cycle_id=cycle_id)

    if date_from:
        counts = counts.filter(count_date__gte=date_from)

    if date_to:
        counts = counts.filter(count_date__lte=date_to)

    # الإحصائيات
    stats = counts.aggregate(
        total_assets=Sum('total_assets'),
        total_counted=Sum('counted_assets'),
        total_variances=Sum('variance_count'),
        count=Count('id')
    )

    context = {
        'title': _('تقرير الجرد الفعلي'),
        'counts': counts.order_by('-count_date')[:50],
        'stats': stats,
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': reverse('assets:reports')},
            {'title': _('الجرد الفعلي'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/physical_count_report.html', context)


@login_required
@permission_required_with_message('assets.view_asset')
def reports_dashboard(request):
    """لوحة التقارير"""

    context = {
        'title': _('التقارير'),
        'breadcrumbs': [
            {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
            {'title': _('التقارير'), 'url': ''},
        ]
    }

    return render(request, 'assets/reports/dashboard.html', context)


# ==================== Export to Excel ====================

@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_asset_register_excel(request):
    """تصدير سجل الأصول إلى Excel"""

    try:
        from openpyxl.drawing.image import Image as ExcelImage
        import os
        from django.conf import settings

        assets = Asset.objects.filter(
            company=request.current_company
        ).select_related('category', 'branch', 'condition').order_by('category', 'asset_number')

        # إنشاء Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "سجل الأصول"

        # التنسيق
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(bold=True, size=16, color="FFFFFF")
        subtitle_font = Font(bold=True, size=12, color="1F4E78")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_font = Font(size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # إضافة الشعار (إذا كان موجوداً)
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            try:
                img = ExcelImage(logo_path)
                # ضبط حجم الصورة
                img.width = 80
                img.height = 80
                ws.add_image(img, 'A1')
                ws.row_dimensions[1].height = 60
                ws.row_dimensions[2].height = 20
                current_row = 3
            except Exception as e:
                print(f"Error adding logo: {e}")

        # عنوان التقرير والمعلومات
        ws.merge_cells(f'B{current_row}:I{current_row}')
        title_cell = ws[f'B{current_row}']
        title_cell.value = "تقرير سجل الأصول الثابتة"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # معلومات الشركة والفرع
        ws.merge_cells(f'A{current_row}:D{current_row}')
        company_cell = ws[f'A{current_row}']
        company_cell.value = f"الشركة: {request.current_company.name}"
        company_cell.font = subtitle_font
        company_cell.alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'E{current_row}:I{current_row}')
        branch_cell = ws[f'E{current_row}']
        branch_cell.value = f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"
        branch_cell.font = subtitle_font
        branch_cell.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 1

        # تاريخ التقرير والمستخدم
        ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"تاريخ التقرير: {date.today().strftime('%Y-%m-%d')}"
        date_cell.font = Font(size=10, color="666666")
        date_cell.alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells(f'E{current_row}:I{current_row}')
        user_cell = ws[f'E{current_row}']
        user_cell.value = f"المستخدم: {request.user.get_full_name() or request.user.username}"
        user_cell.font = Font(size=10, color="666666")
        user_cell.alignment = Alignment(horizontal='right', vertical='center')
        current_row += 2  # سطر فارغ

        # إحصائيات سريعة
        stats = assets.aggregate(
            total_count=Count('id'),
            total_cost=Sum('original_cost'),
            total_book_value=Sum('book_value')
        )

        ws.merge_cells(f'A{current_row}:C{current_row}')
        stats_cell1 = ws[f'A{current_row}']
        stats_cell1.value = f"عدد الأصول: {stats['total_count'] or 0}"
        stats_cell1.font = Font(bold=True, size=10)
        stats_cell1.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        stats_cell1.alignment = Alignment(horizontal='center', vertical='center')
        stats_cell1.border = border

        ws.merge_cells(f'D{current_row}:F{current_row}')
        stats_cell2 = ws[f'D{current_row}']
        stats_cell2.value = f"إجمالي التكلفة: {stats['total_cost'] or 0:,.2f}"
        stats_cell2.font = Font(bold=True, size=10)
        stats_cell2.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        stats_cell2.alignment = Alignment(horizontal='center', vertical='center')
        stats_cell2.border = border

        ws.merge_cells(f'G{current_row}:I{current_row}')
        stats_cell3 = ws[f'G{current_row}']
        stats_cell3.value = f"القيمة الدفترية: {stats['total_book_value'] or 0:,.2f}"
        stats_cell3.font = Font(bold=True, size=10)
        stats_cell3.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        stats_cell3.alignment = Alignment(horizontal='center', vertical='center')
        stats_cell3.border = border
        current_row += 2  # سطر فارغ

        # عناوين الأعمدة
        headers = [
            'رقم الأصل', 'الاسم', 'الفئة', 'الفرع', 'تاريخ الشراء',
            'التكلفة الأصلية', 'الإهلاك المتراكم', 'القيمة الدفترية', 'الحالة'
        ]

        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[header_row].height = 25
        current_row += 1

        # البيانات
        for asset in assets:
            ws.cell(row=current_row, column=1, value=asset.asset_number).border = border
            ws.cell(row=current_row, column=1).font = data_font
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=current_row, column=2, value=asset.name).border = border
            ws.cell(row=current_row, column=2).font = data_font
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right', vertical='center')

            ws.cell(row=current_row, column=3, value=asset.category.name).border = border
            ws.cell(row=current_row, column=3).font = data_font
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='right', vertical='center')

            ws.cell(row=current_row, column=4, value=asset.branch.name if asset.branch else '').border = border
            ws.cell(row=current_row, column=4).font = data_font
            ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=current_row, column=5, value=asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '').border = border
            ws.cell(row=current_row, column=5).font = data_font
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='center', vertical='center')

            cell6 = ws.cell(row=current_row, column=6, value=float(asset.original_cost))
            cell6.border = border
            cell6.font = data_font
            cell6.number_format = '#,##0.00'
            cell6.alignment = Alignment(horizontal='right', vertical='center')

            cell7 = ws.cell(row=current_row, column=7, value=float(asset.accumulated_depreciation))
            cell7.border = border
            cell7.font = data_font
            cell7.number_format = '#,##0.00'
            cell7.alignment = Alignment(horizontal='right', vertical='center')

            cell8 = ws.cell(row=current_row, column=8, value=float(asset.book_value))
            cell8.border = border
            cell8.font = data_font
            cell8.number_format = '#,##0.00'
            cell8.alignment = Alignment(horizontal='right', vertical='center')

            ws.cell(row=current_row, column=9, value=asset.get_status_display()).border = border
            ws.cell(row=current_row, column=9).font = data_font
            ws.cell(row=current_row, column=9).alignment = Alignment(horizontal='center', vertical='center')

            # تلوين الصفوف بالتناوب
            if current_row % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=current_row, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )

            current_row += 1

        # تعديل عرض الأعمدة
        column_widths = {
            'A': 15,  # رقم الأصل
            'B': 30,  # الاسم
            'C': 20,  # الفئة
            'D': 15,  # الفرع
            'E': 15,  # تاريخ الشراء
            'F': 18,  # التكلفة
            'G': 18,  # الإهلاك
            'H': 18,  # القيمة الدفترية
            'I': 12   # الحالة
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # تجميد الأسطر العلوية
        ws.freeze_panes = f'A{header_row + 1}'

        # الحفظ
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="asset_register_{date.today()}.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:asset_register_report')


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_asset_register_pdf(request):
    """تصدير سجل الأصول إلى PDF"""

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        import os
        from django.conf import settings
        from bidi.algorithm import get_display
        import arabic_reshaper

        # تسجيل خط عربي - Try different font paths
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]

        arabic_font = 'Helvetica'  # Default fallback

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

        assets = Asset.objects.filter(
            company=request.current_company
        ).select_related('category', 'branch', 'condition').order_by('category', 'asset_number')

        # إنشاء PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title="تقرير سجل الأصول الثابتة"
        )

        # عناصر المستند
        elements = []

        # دالة مساعدة لتحويل النص العربي
        def arabic_text(text):
            """Convert Arabic text with proper reshaping and bidi"""
            text_str = str(text)
            if arabic_font == 'ArabicFont':
                try:
                    reshaped_text = arabic_reshaper.reshape(text_str)
                    return get_display(reshaped_text)
                except:
                    return text_str
            return text_str

        # الأنماط
        styles = getSampleStyleSheet()

        # Create custom Arabic style for table content
        arabic_normal_style = ParagraphStyle(
            name='ArabicNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=8,
            alignment=TA_CENTER,
            leading=10
        )

        # نمط العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontName=arabic_font,
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            alignment=TA_CENTER,
            spaceAfter=12,
            fontWeight='BOLD'
        )

        # نمط المعلومات
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=TA_RIGHT,
            rightIndent=0,
            leftIndent=0
        )

        # نمط الإحصائيات
        stats_style = ParagraphStyle(
            'StatsStyle',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=11,
            alignment=TA_CENTER,
            fontWeight='BOLD'
        )

        # إضافة الشعار (إذا كان موجوداً)
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=2*cm, height=2*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.3*cm))
            except:
                pass

        # العنوان
        title = Paragraph(arabic_text("تقرير سجل الأصول الثابتة"), title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))

        # معلومات الشركة والتقرير
        info_data = [
            [
                Paragraph(arabic_text(f"الشركة: {request.current_company.name}"), info_style),
                Paragraph(arabic_text(f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"), info_style),
            ],
            [
                Paragraph(arabic_text(f"تاريخ التقرير: {date.today().strftime('%Y-%m-%d')}"), info_style),
                Paragraph(arabic_text(f"المستخدم: {request.user.get_full_name() or request.user.username}"), info_style),
            ]
        ]

        info_table = Table(info_data, colWidths=[12*cm, 12*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#666666')),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))

        # الإحصائيات
        stats = assets.aggregate(
            total_count=Count('id'),
            total_cost=Sum('original_cost'),
            total_book_value=Sum('book_value')
        )

        stats_data = [[
            Paragraph(arabic_text(f"عدد الأصول: {stats['total_count'] or 0}"), stats_style),
            Paragraph(arabic_text(f"إجمالي التكلفة: {stats['total_cost'] or 0:,.2f}"), stats_style),
            Paragraph(arabic_text(f"القيمة الدفترية: {stats['total_book_value'] or 0:,.2f}"), stats_style),
        ]]

        stats_table = Table(stats_data, colWidths=[8*cm, 8*cm, 8*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.7*cm))

        # جدول البيانات
        data = [[
            Paragraph(arabic_text('رقم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('الاسم'), arabic_normal_style),
            Paragraph(arabic_text('الفئة'), arabic_normal_style),
            Paragraph(arabic_text('الفرع'), arabic_normal_style),
            Paragraph(arabic_text('تاريخ الشراء'), arabic_normal_style),
            Paragraph(arabic_text('التكلفة الأصلية'), arabic_normal_style),
            Paragraph(arabic_text('الإهلاك المتراكم'), arabic_normal_style),
            Paragraph(arabic_text('القيمة الدفترية'), arabic_normal_style),
            Paragraph(arabic_text('الحالة'), arabic_normal_style),
        ]]

        for asset in assets[:100]:  # تحديد 100 أصل لتجنب ملفات PDF كبيرة جداً
            data.append([
                Paragraph(arabic_text(asset.asset_number), arabic_normal_style),
                Paragraph(arabic_text(asset.name[:30]), arabic_normal_style),  # تقصير النص
                Paragraph(arabic_text(asset.category.name[:20]), arabic_normal_style),
                Paragraph(arabic_text(asset.branch.name if asset.branch else '-'), arabic_normal_style),
                Paragraph(arabic_text(asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '-'), arabic_normal_style),
                Paragraph(arabic_text(f"{float(asset.original_cost):,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{float(asset.accumulated_depreciation):,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{float(asset.book_value):,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(asset.get_status_display()), arabic_normal_style),
            ])

        # إنشاء الجدول
        table = Table(data, colWidths=[
            2.5*cm,  # رقم الأصل
            4*cm,    # الاسم
            3*cm,    # الفئة
            2.5*cm,  # الفرع
            2.5*cm,  # تاريخ الشراء
            3*cm,    # التكلفة
            3*cm,    # الإهلاك
            3*cm,    # القيمة الدفترية
            2*cm,    # الحالة
        ])

        # تنسيق الجدول
        table_style = TableStyle([
            # رأس الجدول
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTWEIGHT', (0, 0), (-1, 0), 'BOLD'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),

            # محتوى الجدول
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # رقم الأصل
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),   # الاسم، الفئة، الفرع
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # تاريخ الشراء
            ('ALIGN', (5, 1), (7, -1), 'RIGHT'),   # الأرقام
            ('ALIGN', (8, 1), (8, -1), 'CENTER'),  # الحالة
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),

            # الحدود
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#366092')),

            # الحشو
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ])

        # تلوين الصفوف بالتناوب
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2'))

        table.setStyle(table_style)
        elements.append(table)

        # رسالة إذا كان هناك المزيد من الأصول
        if assets.count() > 100:
            elements.append(Spacer(1, 0.5*cm))
            note = Paragraph(
                arabic_text(f"ملاحظة: يعرض التقرير أول 100 أصل من أصل {assets.count()} أصل. استخدم تصدير Excel للحصول على البيانات الكاملة."),
                info_style
            )
            elements.append(note)

        # بناء المستند
        doc.build(elements)

        # إعادة الملف
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="asset_register_{date.today()}.pdf"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير إلى PDF: {str(e)}')
        return redirect('assets:asset_register_report')


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_depreciation_excel(request):
    """تصدير تقرير الإهلاك إلى Excel - نسخة احترافية"""

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    from django.conf import settings

    try:
        # الفلاتر
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        category_id = request.GET.get('category', '')

        # البيانات الأساسية
        depreciation_records = AssetDepreciation.objects.filter(
            company=request.current_company
        ).select_related('asset', 'asset__category', 'asset__depreciation_method')

        # تطبيق الفلاتر
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                depreciation_records = depreciation_records.filter(depreciation_date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                depreciation_records = depreciation_records.filter(depreciation_date__lte=date_to_obj)
            except ValueError:
                pass

        if category_id:
            depreciation_records = depreciation_records.filter(asset__category_id=category_id)

        depreciation_records = depreciation_records.order_by('-depreciation_date')

        # الإحصائيات
        stats = depreciation_records.aggregate(
            total_depreciation=Sum('depreciation_amount'),
            total_accumulated=Sum('accumulated_depreciation_after'),
            count=Count('id')
        )

        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الإهلاك"

        # التنسيق
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        info_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        # إضافة الشعار
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
            current_row = 6

        # عنوان التقرير
        ws.merge_cells(f'B{current_row}:I{current_row}')
        title_cell = ws[f'B{current_row}']
        title_cell.value = "تقرير الإهلاك"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 2

        # معلومات الشركة والفرع
        ws.merge_cells(f'A{current_row}:D{current_row}')
        company_cell = ws[f'A{current_row}']
        company_cell.value = f"الشركة: {request.current_company.name}"
        company_cell.font = info_font
        company_cell.alignment = Alignment(horizontal='right')

        ws.merge_cells(f'E{current_row}:I{current_row}')
        branch_cell = ws[f'E{current_row}']
        branch_cell.value = f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"
        branch_cell.font = info_font
        branch_cell.alignment = Alignment(horizontal='right')
        current_row += 1

        # تاريخ التقرير والمستخدم
        ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"تاريخ التقرير: {date.today().strftime('%Y-%m-%d')}"
        date_cell.font = info_font
        date_cell.alignment = Alignment(horizontal='right')

        ws.merge_cells(f'E{current_row}:I{current_row}')
        user_cell = ws[f'E{current_row}']
        user_cell.value = f"المستخدم: {request.user.get_full_name()}"
        user_cell.font = info_font
        user_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        # الإحصائيات
        stats_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        stats_font = Font(bold=True, size=11)

        ws.merge_cells(f'A{current_row}:I{current_row}')
        stats_header = ws[f'A{current_row}']
        stats_header.value = "إحصائيات الإهلاك"
        stats_header.font = stats_font
        stats_header.fill = stats_fill
        stats_header.alignment = Alignment(horizontal='center')
        stats_header.border = border
        current_row += 1

        # بيانات الإحصائيات
        ws.merge_cells(f'A{current_row}:C{current_row}')
        count_cell = ws[f'A{current_row}']
        count_cell.value = f"عدد السجلات: {stats['count'] or 0}"
        count_cell.font = info_font
        count_cell.border = border
        count_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(f'D{current_row}:F{current_row}')
        depreciation_cell = ws[f'D{current_row}']
        depreciation_cell.value = f"إهلاك الفترة: {stats['total_depreciation'] or 0:,.2f}"
        depreciation_cell.font = info_font
        depreciation_cell.border = border
        depreciation_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(f'G{current_row}:I{current_row}')
        accumulated_cell = ws[f'G{current_row}']
        accumulated_cell.value = f"الإهلاك المتراكم: {stats['total_accumulated'] or 0:,.2f}"
        accumulated_cell.font = info_font
        accumulated_cell.border = border
        accumulated_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # عناوين الجدول
        headers = [
            'التاريخ', 'رقم الأصل', 'اسم الأصل', 'الفئة',
            'مبلغ الإهلاك', 'الإهلاك المتراكم', 'القيمة الدفترية', 'طريقة الإهلاك', 'الفرع'
        ]

        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[header_row].height = 20
        current_row += 1

        # البيانات
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for idx, record in enumerate(depreciation_records[:1000], 0):  # حد أقصى 1000 سجل
            row = current_row + idx

            # تطبيق لون متناوب
            if idx % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = alt_fill

            # البيانات
            ws.cell(row=row, column=1, value=record.depreciation_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=2, value=record.asset.asset_number).border = border
            ws.cell(row=row, column=3, value=record.asset.name).border = border
            ws.cell(row=row, column=4, value=record.asset.category.name).border = border

            # الأرقام مع التنسيق
            amount_cell = ws.cell(row=row, column=5, value=float(record.depreciation_amount))
            amount_cell.border = border
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = Alignment(horizontal='right')

            accumulated_cell = ws.cell(row=row, column=6, value=float(record.accumulated_depreciation_after))
            accumulated_cell.border = border
            accumulated_cell.number_format = '#,##0.00'
            accumulated_cell.alignment = Alignment(horizontal='right')

            book_value_cell = ws.cell(row=row, column=7, value=float(record.book_value_after))
            book_value_cell.border = border
            book_value_cell.number_format = '#,##0.00'
            book_value_cell.alignment = Alignment(horizontal='right')

            ws.cell(row=row, column=8, value=record.asset.depreciation_method.name if record.asset.depreciation_method else '-').border = border
            ws.cell(row=row, column=9, value=record.asset.branch.name if record.asset.branch else '-').border = border

        # تجميد الصفوف العلوية
        ws.freeze_panes = f'A{header_row + 1}'

        # تعديل عرض الأعمدة
        column_widths = [12, 15, 30, 20, 15, 18, 18, 20, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        # الحفظ
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"depreciation_report_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في تصدير Excel: {str(e)}')
        return redirect('assets:depreciation_report')


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_depreciation_pdf(request):
    """تصدير تقرير الإهلاك إلى PDF - نسخة احترافية"""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    from django.conf import settings

    try:
        # تسجيل الخط العربي
        from bidi.algorithm import get_display
        import arabic_reshaper

        # Try different font paths
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]

        arabic_font = 'Helvetica'  # Default fallback

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

        # دالة لتحويل النص العربي
        def arabic_text(text):
            if arabic_font == 'ArabicFont':
                try:
                    reshaped_text = arabic_reshaper.reshape(str(text))
                    return get_display(reshaped_text)
                except Exception as e:
                    return str(text)
            return str(text)

        # الفلاتر
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        category_id = request.GET.get('category', '')

        # البيانات الأساسية
        depreciation_records = AssetDepreciation.objects.filter(
            company=request.current_company
        ).select_related('asset', 'asset__category', 'asset__depreciation_method')

        # تطبيق الفلاتر
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                depreciation_records = depreciation_records.filter(depreciation_date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                depreciation_records = depreciation_records.filter(depreciation_date__lte=date_to_obj)
            except ValueError:
                pass

        if category_id:
            depreciation_records = depreciation_records.filter(asset__category_id=category_id)

        depreciation_records = depreciation_records.order_by('-depreciation_date')[:100]  # حد أقصى 100 سجل

        # الإحصائيات
        stats = AssetDepreciation.objects.filter(
            company=request.current_company
        ).aggregate(
            total_depreciation=Sum('depreciation_amount'),
            total_accumulated=Sum('accumulated_depreciation_after'),
            count=Count('id')
        )

        # إنشاء ملف PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title=arabic_text("تقرير الإهلاك")
        )

        # العناصر
        elements = []
        styles = getSampleStyleSheet()

        # Create custom Arabic style for table content
        arabic_normal_style = ParagraphStyle(
            name='ArabicNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=9,
            alignment=TA_CENTER,
            leading=12
        )

        # الشعار
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=2*cm, height=2*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.3*cm))
            except:
                pass

        # العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=arabic_font,
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        elements.append(Paragraph(arabic_text("تقرير الإهلاك"), title_style))
        elements.append(Spacer(1, 0.5*cm))

        # معلومات الشركة والتقرير
        info_style = ParagraphStyle(
            'InfoStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        info_data = [[
            Paragraph(arabic_text(f"الشركة: {request.current_company.name}"), info_style),
            Paragraph(arabic_text(f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"), info_style),
            Paragraph(arabic_text(f"التاريخ: {date.today().strftime('%Y-%m-%d')}"), info_style),
        ]]

        info_table = Table(info_data, colWidths=[8*cm, 8*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))

        # الإحصائيات
        stats_style = ParagraphStyle(
            'StatsStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        stats_data = [[
            Paragraph(arabic_text(f"عدد السجلات: {stats['count'] or 0}"), stats_style),
            Paragraph(arabic_text(f"إهلاك الفترة: {stats['total_depreciation'] or 0:,.2f}"), stats_style),
            Paragraph(arabic_text(f"الإهلاك المتراكم: {stats['total_accumulated'] or 0:,.2f}"), stats_style),
        ]]

        stats_table = Table(stats_data, colWidths=[8*cm, 8*cm, 8*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.7*cm))

        # جدول البيانات
        data = [[
            Paragraph(arabic_text('التاريخ'), arabic_normal_style),
            Paragraph(arabic_text('رقم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('اسم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('الفئة'), arabic_normal_style),
            Paragraph(arabic_text('مبلغ الإهلاك'), arabic_normal_style),
            Paragraph(arabic_text('الإهلاك المتراكم'), arabic_normal_style),
            Paragraph(arabic_text('القيمة الدفترية'), arabic_normal_style),
            Paragraph(arabic_text('الطريقة'), arabic_normal_style),
        ]]

        for record in depreciation_records:
            data.append([
                Paragraph(arabic_text(record.depreciation_date.strftime('%Y-%m-%d')), arabic_normal_style),
                Paragraph(arabic_text(record.asset.asset_number), arabic_normal_style),
                Paragraph(arabic_text(record.asset.name), arabic_normal_style),
                Paragraph(arabic_text(record.asset.category.name), arabic_normal_style),
                Paragraph(arabic_text(f"{record.depreciation_amount:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{record.accumulated_depreciation_after:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{record.book_value_after:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(record.asset.depreciation_method.name if record.asset.depreciation_method else '-'), arabic_normal_style),
            ])

        # إنشاء الجدول
        table = Table(data, colWidths=[2.5*cm, 2.5*cm, 5*cm, 3.5*cm, 3*cm, 3.5*cm, 3*cm, 3*cm])

        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])

        # ألوان متناوبة للصفوف
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2'))

        table.setStyle(table_style)
        elements.append(table)

        # ملاحظة إذا كان هناك المزيد من السجلات
        if depreciation_records.count() > 100:
            elements.append(Spacer(1, 0.3*cm))
            note_style = ParagraphStyle(
                'NoteStyle',
                fontName=arabic_font,
                fontSize=9,
                textColor=colors.red,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(
                arabic_text(f"ملاحظة: تم عرض أول 100 سجل من أصل {depreciation_records.count()} سجل"),
                note_style
            ))

        # بناء PDF
        doc.build(elements)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(output.read(), content_type='application/pdf')
        filename = f"depreciation_report_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في تصدير PDF: {str(e)}')
        return redirect('assets:depreciation_report')

@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_maintenance_excel(request):
    """تصدير تقرير الصيانة إلى Excel - نسخة احترافية"""

    from openpyxl.drawing.image import Image as ExcelImage
    import os
    from django.conf import settings

    try:
        # الفلاتر
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        maintenance_type_id = request.GET.get('maintenance_type', '')
        status_filter = request.GET.get('status', '')

        # البيانات الأساسية
        maintenances = AssetMaintenance.objects.filter(
            company=request.current_company
        ).select_related('asset', 'asset__category', 'maintenance_type', 'asset__branch')

        # تطبيق الفلاتر
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                maintenances = maintenances.filter(scheduled_date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                maintenances = maintenances.filter(scheduled_date__lte=date_to_obj)
            except ValueError:
                pass

        if maintenance_type_id:
            maintenances = maintenances.filter(maintenance_type_id=maintenance_type_id)

        if status_filter:
            maintenances = maintenances.filter(status=status_filter)

        maintenances = maintenances.order_by('-scheduled_date')

        # الإحصائيات
        stats = maintenances.aggregate(
            total_count=Count('id'),
            completed_count=Count('id', filter=Q(status='completed')),
            in_progress_count=Count('id', filter=Q(status='in_progress')),
            total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost'))
        )

        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الصيانة"

        # التنسيق
        title_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        info_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        # إضافة الشعار
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            img = ExcelImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
            current_row = 6

        # عنوان التقرير
        ws.merge_cells(f'B{current_row}:J{current_row}')
        title_cell = ws[f'B{current_row}']
        title_cell.value = "تقرير الصيانة"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 2

        # معلومات الشركة والفرع
        ws.merge_cells(f'A{current_row}:E{current_row}')
        company_cell = ws[f'A{current_row}']
        company_cell.value = f"الشركة: {request.current_company.name}"
        company_cell.font = info_font
        company_cell.alignment = Alignment(horizontal='right')

        ws.merge_cells(f'F{current_row}:J{current_row}')
        branch_cell = ws[f'F{current_row}']
        branch_cell.value = f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"
        branch_cell.font = info_font
        branch_cell.alignment = Alignment(horizontal='right')
        current_row += 1

        # تاريخ التقرير والمستخدم
        ws.merge_cells(f'A{current_row}:E{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"تاريخ التقرير: {date.today().strftime('%Y-%m-%d')}"
        date_cell.font = info_font
        date_cell.alignment = Alignment(horizontal='right')

        ws.merge_cells(f'F{current_row}:J{current_row}')
        user_cell = ws[f'F{current_row}']
        user_cell.value = f"المستخدم: {request.user.get_full_name()}"
        user_cell.font = info_font
        user_cell.alignment = Alignment(horizontal='right')
        current_row += 2

        # الإحصائيات
        stats_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        stats_font = Font(bold=True, size=11)

        ws.merge_cells(f'A{current_row}:J{current_row}')
        stats_header = ws[f'A{current_row}']
        stats_header.value = "إحصائيات الصيانة"
        stats_header.font = stats_font
        stats_header.fill = stats_fill
        stats_header.alignment = Alignment(horizontal='center')
        stats_header.border = border
        current_row += 1

        # بيانات الإحصائيات
        ws.merge_cells(f'A{current_row}:C{current_row}')
        count_cell = ws[f'A{current_row}']
        count_cell.value = f"إجمالي الصيانة: {stats['total_count'] or 0}"
        count_cell.font = info_font
        count_cell.border = border
        count_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(f'D{current_row}:F{current_row}')
        completed_cell = ws[f'D{current_row}']
        completed_cell.value = f"مكتملة: {stats['completed_count'] or 0}"
        completed_cell.font = info_font
        completed_cell.border = border
        completed_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells(f'G{current_row}:J{current_row}')
        cost_cell = ws[f'G{current_row}']
        cost_cell.value = f"إجمالي التكلفة: {stats['total_cost'] or 0:,.2f}"
        cost_cell.font = info_font
        cost_cell.border = border
        cost_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # عناوين الجدول
        headers = [
            'التاريخ المجدول', 'رقم الأصل', 'اسم الأصل', 'النوع',
            'تكلفة العمالة', 'تكلفة القطع', 'التكلفة الأخرى', 'إجمالي التكلفة', 'الحالة', 'الفرع'
        ]

        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[header_row].height = 20
        current_row += 1

        # البيانات
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for idx, maintenance in enumerate(maintenances[:1000], 0):  # حد أقصى 1000 سجل
            row = current_row + idx

            # تطبيق لون متناوب
            if idx % 2 == 0:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = alt_fill

            # البيانات
            ws.cell(row=row, column=1, value=maintenance.scheduled_date.strftime('%Y-%m-%d') if maintenance.scheduled_date else '-').border = border
            ws.cell(row=row, column=2, value=maintenance.asset.asset_number if maintenance.asset else '-').border = border
            ws.cell(row=row, column=3, value=maintenance.asset.name if maintenance.asset else '-').border = border
            ws.cell(row=row, column=4, value=maintenance.maintenance_type.name if maintenance.maintenance_type else '-').border = border

            # الأرقام مع التنسيق
            labor_cell = ws.cell(row=row, column=5, value=float(maintenance.labor_cost or 0))
            labor_cell.border = border
            labor_cell.number_format = '#,##0.00'
            labor_cell.alignment = Alignment(horizontal='right')

            parts_cell = ws.cell(row=row, column=6, value=float(maintenance.parts_cost or 0))
            parts_cell.border = border
            parts_cell.number_format = '#,##0.00'
            parts_cell.alignment = Alignment(horizontal='right')

            other_cell = ws.cell(row=row, column=7, value=float(maintenance.other_cost or 0))
            other_cell.border = border
            other_cell.number_format = '#,##0.00'
            other_cell.alignment = Alignment(horizontal='right')

            total_cost = (maintenance.labor_cost or 0) + (maintenance.parts_cost or 0) + (maintenance.other_cost or 0)
            total_cell = ws.cell(row=row, column=8, value=float(total_cost))
            total_cell.border = border
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = Alignment(horizontal='right')

            # الحالة
            status_text = 'مكتملة' if maintenance.status == 'completed' else 'قيد التنفيذ' if maintenance.status == 'in_progress' else 'مجدولة'
            ws.cell(row=row, column=9, value=status_text).border = border
            ws.cell(row=row, column=10, value=maintenance.asset.branch.name if maintenance.asset and maintenance.asset.branch else '-').border = border

        # تجميد الصفوف العلوية
        ws.freeze_panes = f'A{header_row + 1}'

        # تعديل عرض الأعمدة
        column_widths = [15, 15, 30, 20, 15, 15, 15, 18, 15, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        # الحفظ
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"maintenance_report_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في تصدير Excel: {str(e)}')
        return redirect('assets:maintenance_report')


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_maintenance_pdf(request):
    """تصدير تقرير الصيانة إلى PDF - نسخة احترافية"""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    from django.conf import settings
    from bidi.algorithm import get_display
    import arabic_reshaper

    try:
        # Try different font paths
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]

        arabic_font = 'Helvetica'  # Default fallback

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

        # دالة لتحويل النص العربي
        def arabic_text(text):
            """Convert Arabic text with proper reshaping and bidi"""
            text_str = str(text)
            if arabic_font == 'ArabicFont':
                try:
                    reshaped_text = arabic_reshaper.reshape(text_str)
                    return get_display(reshaped_text)
                except Exception as e:
                    return text_str
            return text_str

        # الفلاتر
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        maintenance_type_id = request.GET.get('maintenance_type', '')
        status_filter = request.GET.get('status', '')

        # البيانات الأساسية
        maintenances = AssetMaintenance.objects.filter(
            company=request.current_company
        ).select_related('asset', 'asset__category', 'maintenance_type', 'asset__branch')

        # تطبيق الفلاتر
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                maintenances = maintenances.filter(scheduled_date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                maintenances = maintenances.filter(scheduled_date__lte=date_to_obj)
            except ValueError:
                pass

        if maintenance_type_id:
            maintenances = maintenances.filter(maintenance_type_id=maintenance_type_id)

        if status_filter:
            maintenances = maintenances.filter(status=status_filter)

        maintenances = maintenances.order_by('-scheduled_date')[:100]  # حد أقصى 100 سجل

        # الإحصائيات
        stats = AssetMaintenance.objects.filter(
            company=request.current_company
        ).aggregate(
            total_count=Count('id'),
            completed_count=Count('id', filter=Q(status='completed')),
            total_cost=Sum(F('labor_cost') + F('parts_cost') + F('other_cost'))
        )

        # إنشاء ملف PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title=arabic_text("تقرير الصيانة")
        )

        # العناصر
        elements = []
        styles = getSampleStyleSheet()

        # Create custom Arabic style for table content
        arabic_normal_style = ParagraphStyle(
            name='ArabicNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=9,
            alignment=TA_CENTER,
            leading=12
        )

        # الشعار
        logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=2*cm, height=2*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.3*cm))
            except:
                pass

        # العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=arabic_font,
            fontSize=18,
            textColor=colors.HexColor('#DC3545'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        elements.append(Paragraph(arabic_text("تقرير الصيانة"), title_style))
        elements.append(Spacer(1, 0.5*cm))

        # معلومات الشركة والتقرير
        info_style = ParagraphStyle(
            'InfoStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        info_data = [[
            Paragraph(arabic_text(f"الشركة: {request.current_company.name}"), info_style),
            Paragraph(arabic_text(f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"), info_style),
            Paragraph(arabic_text(f"التاريخ: {date.today().strftime('%Y-%m-%d')}"), info_style),
        ]]

        info_table = Table(info_data, colWidths=[8*cm, 8*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))

        # الإحصائيات
        stats_style = ParagraphStyle(
            'StatsStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        stats_data = [[
            Paragraph(arabic_text(f"إجمالي الصيانة: {stats['total_count'] or 0}"), stats_style),
            Paragraph(arabic_text(f"مكتملة: {stats['completed_count'] or 0}"), stats_style),
            Paragraph(arabic_text(f"إجمالي التكلفة: {stats['total_cost'] or 0:,.2f}"), stats_style),
        ]]

        stats_table = Table(stats_data, colWidths=[8*cm, 8*cm, 8*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.7*cm))

        # جدول البيانات
        data = [[
            Paragraph(arabic_text('التاريخ'), arabic_normal_style),
            Paragraph(arabic_text('رقم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('اسم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('النوع'), arabic_normal_style),
            Paragraph(arabic_text('تكلفة العمالة'), arabic_normal_style),
            Paragraph(arabic_text('تكلفة القطع'), arabic_normal_style),
            Paragraph(arabic_text('إجمالي التكلفة'), arabic_normal_style),
            Paragraph(arabic_text('الحالة'), arabic_normal_style),
        ]]

        for maintenance in maintenances:
            total_cost = (maintenance.labor_cost or 0) + (maintenance.parts_cost or 0) + (maintenance.other_cost or 0)
            status_text = 'مكتملة' if maintenance.status == 'completed' else 'قيد التنفيذ' if maintenance.status == 'in_progress' else 'مجدولة'

            data.append([
                Paragraph(arabic_text(maintenance.scheduled_date.strftime('%Y-%m-%d') if maintenance.scheduled_date else '-'), arabic_normal_style),
                Paragraph(arabic_text(maintenance.asset.asset_number if maintenance.asset else '-'), arabic_normal_style),
                Paragraph(arabic_text(maintenance.asset.name if maintenance.asset else '-'), arabic_normal_style),
                Paragraph(arabic_text(maintenance.maintenance_type.name if maintenance.maintenance_type else '-'), arabic_normal_style),
                Paragraph(arabic_text(f"{maintenance.labor_cost or 0:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{maintenance.parts_cost or 0:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{total_cost:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(status_text), arabic_normal_style),
            ])

        # إنشاء الجدول
        table = Table(data, colWidths=[2.5*cm, 2.5*cm, 5*cm, 3*cm, 3*cm, 3*cm, 3*cm, 2.5*cm])

        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C757D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])

        # ألوان متناوبة للصفوف
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2'))

        table.setStyle(table_style)
        elements.append(table)

        # ملاحظة إذا كان هناك المزيد من السجلات
        if maintenances.count() > 100:
            elements.append(Spacer(1, 0.3*cm))
            note_style = ParagraphStyle(
                'NoteStyle',
                fontName=arabic_font,
                fontSize=9,
                textColor=colors.red,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(
                arabic_text(f"ملاحظة: تم عرض أول 100 سجل من أصل {maintenances.count()} سجل"),
                note_style
            ))

        # بناء PDF
        doc.build(elements)
        output.seek(0)

        # الاستجابة
        response = HttpResponse(output.read(), content_type='application/pdf')
        filename = f"maintenance_report_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في تصدير PDF: {str(e)}')
        return redirect('assets:maintenance_report')


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_valuation_excel(request):
    """تصدير تقرير تقييم الأصول إلى Excel - نسخة احترافية"""

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import os
    from django.conf import settings

    # الفلاتر
    category_id = request.GET.get('category', '')
    branch_id = request.GET.get('branch', '')
    status_filter = request.GET.get('status', 'active')

    # البيانات الأساسية
    assets = Asset.objects.filter(
        company=request.current_company
    ).select_related('category', 'branch', 'depreciation_method')

    # تطبيق الفلاتر
    if status_filter:
        assets = assets.filter(status=status_filter)
    if category_id:
        assets = assets.filter(category_id=category_id)
    if branch_id:
        assets = assets.filter(branch_id=branch_id)

    # الإحصائيات
    stats = assets.aggregate(
        count=Count('id'),
        total_cost=Sum('original_cost'),
        total_accumulated=Sum('accumulated_depreciation'),
        total_book_value=Sum(F('original_cost') - F('accumulated_depreciation'))
    )

    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير التقييم"

    # إضافة الشعار
    logo_path = os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'images', 'logo.png')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width = 80
            img.height = 80
            ws.add_image(img, 'A1')
        except:
            pass

    # العنوان
    ws.merge_cells('A1:J3')
    title_cell = ws['A1']
    title_cell.value = 'تقرير تقييم الأصول الثابتة'
    title_cell.font = Font(size=18, bold=True, color='28A745')  # Green color
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # معلومات الشركة
    row = 4
    ws.merge_cells(f'A{row}:J{row}')
    company_cell = ws[f'A{row}']
    company_cell.value = f"الشركة: {request.current_company.name} | الفرع: {request.current_branch.name if request.current_branch else 'الكل'} | التاريخ: {date.today().strftime('%Y-%m-%d')}"
    company_cell.font = Font(size=11)
    company_cell.alignment = Alignment(horizontal='center')

    # الإحصائيات
    row = 6
    ws.merge_cells(f'A{row}:J{row}')
    stats_cell = ws[f'A{row}']
    stats_cell.value = f"عدد الأصول: {stats['count'] or 0} | التكلفة: {stats['total_cost'] or 0:,.2f} | الإهلاك: {stats['total_accumulated'] or 0:,.2f} | القيمة الدفترية: {stats['total_book_value'] or 0:,.2f}"
    stats_cell.font = Font(size=10, bold=True)
    stats_cell.alignment = Alignment(horizontal='center')
    stats_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # رأس الجدول
    row = 8
    headers = ['رقم الأصل', 'الاسم', 'الفئة', 'الفرع', 'تاريخ الشراء', 'التكلفة الأصلية', 
               'الإهلاك المتراكم', 'القيمة الدفترية', 'العمر (سنة)', 'الحالة']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # Green
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # البيانات
    row = 9
    assets_list = assets.order_by('-book_value')[:1000]  # حد أقصى 1000 أصل
    
    for asset in assets_list:
        status_ar = {
            'active': 'نشط',
            'inactive': 'غير نشط',
            'disposed': 'مستبعد',
            'sold': 'مباع'
        }.get(asset.status, asset.status)

        data = [
            asset.asset_number,
            asset.name,
            asset.category.name if asset.category else '-',
            asset.branch.name if asset.branch else '-',
            asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '-',
            float(asset.original_cost) if asset.original_cost else 0,
            float(asset.accumulated_depreciation) if asset.accumulated_depreciation else 0,
            float(asset.book_value) if asset.book_value else 0,
            asset.age_years if hasattr(asset, 'age_years') and asset.age_years else '-',
            status_ar
        ]

        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            
            # تنسيق الأرقام
            if col in [6, 7, 8]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')
            
            # تلوين الصفوف المتناوبة
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1

    # تعيين عرض الأعمدة
    column_widths = [15, 30, 20, 15, 15, 18, 18, 18, 12, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # حفظ الملف
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=valuation_report_{date.today().strftime("%Y%m%d")}.xlsx'

    return response


@login_required
@permission_required_with_message('assets.view_asset')
@require_http_methods(["GET"])
def export_valuation_pdf(request):
    """تصدير تقرير تقييم الأصول إلى PDF - نسخة احترافية"""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    from django.conf import settings
    from bidi.algorithm import get_display
    import arabic_reshaper

    try:
        # Try different font paths
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'fonts', 'Arial.ttf'),
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]

        arabic_font = 'Helvetica'  # Default fallback

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                    arabic_font = 'ArabicFont'
                    break
                except Exception as e:
                    continue

        # دالة لتحويل النص العربي
        def arabic_text(text):
            """Convert Arabic text with proper reshaping and bidi"""
            text_str = str(text)
            if arabic_font == 'ArabicFont':
                try:
                    reshaped_text = arabic_reshaper.reshape(text_str)
                    return get_display(reshaped_text)
                except Exception as e:
                    return text_str
            return text_str

        # الفلاتر
        category_id = request.GET.get('category', '')
        branch_id = request.GET.get('branch', '')
        status_filter = request.GET.get('status', 'active')

        # البيانات الأساسية
        assets = Asset.objects.filter(
            company=request.current_company
        ).select_related('category', 'branch')

        # تطبيق الفلاتر
        if status_filter:
            assets = assets.filter(status=status_filter)
        if category_id:
            assets = assets.filter(category_id=category_id)
        if branch_id:
            assets = assets.filter(branch_id=branch_id)

        assets = assets.order_by('-book_value')[:100]  # حد أقصى 100 أصل

        # الإحصائيات
        stats = Asset.objects.filter(
            company=request.current_company
        ).aggregate(
            count=Count('id'),
            total_cost=Sum('original_cost'),
            total_accumulated=Sum('accumulated_depreciation'),
            total_book_value=Sum(F('original_cost') - F('accumulated_depreciation'))
        )

        # إنشاء ملف PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title=arabic_text("تقرير تقييم الأصول")
        )

        # العناصر
        elements = []
        styles = getSampleStyleSheet()

        # Create custom Arabic style for table content
        arabic_normal_style = ParagraphStyle(
            name='ArabicNormal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=8,
            alignment=TA_CENTER,
            leading=10
        )

        # الشعار
        logo_path = os.path.join(settings.STATIC_ROOT or (settings.STATICFILES_DIRS[0] if settings.STATICFILES_DIRS else ''), 'images', 'logo.png')
        if not os.path.exists(logo_path):
            logo_path = os.path.join(settings.MEDIA_ROOT, 'logo.png')

        if os.path.exists(logo_path):
            try:
                logo = Image(logo_path, width=2*cm, height=2*cm)
                elements.append(logo)
                elements.append(Spacer(1, 0.3*cm))
            except:
                pass

        # العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=arabic_font,
            fontSize=18,
            textColor=colors.HexColor('#28A745'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        elements.append(Paragraph(arabic_text("تقرير تقييم الأصول"), title_style))
        elements.append(Spacer(1, 0.5*cm))

        # معلومات الشركة والتقرير
        info_style = ParagraphStyle(
            'InfoStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        info_data = [[
            Paragraph(arabic_text(f"الشركة: {request.current_company.name}"), info_style),
            Paragraph(arabic_text(f"الفرع: {request.current_branch.name if request.current_branch else 'الكل'}"), info_style),
            Paragraph(arabic_text(f"التاريخ: {date.today().strftime('%Y-%m-%d')}"), info_style),
        ]]

        info_table = Table(info_data, colWidths=[8*cm, 8*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))

        # الإحصائيات
        stats_style = ParagraphStyle(
            'StatsStyle',
            fontName=arabic_font,
            fontSize=10,
            alignment=TA_CENTER
        )

        stats_data = [[
            Paragraph(arabic_text(f"عدد الأصول: {stats['count'] or 0}"), stats_style),
            Paragraph(arabic_text(f"التكلفة: {stats['total_cost'] or 0:,.2f}"), stats_style),
            Paragraph(arabic_text(f"القيمة الدفترية: {stats['total_book_value'] or 0:,.2f}"), stats_style),
        ]]

        stats_table = Table(stats_data, colWidths=[8*cm, 8*cm, 8*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (0, 0), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.7*cm))

        # جدول البيانات
        data = [[
            Paragraph(arabic_text('رقم الأصل'), arabic_normal_style),
            Paragraph(arabic_text('الاسم'), arabic_normal_style),
            Paragraph(arabic_text('الفئة'), arabic_normal_style),
            Paragraph(arabic_text('الفرع'), arabic_normal_style),
            Paragraph(arabic_text('التكلفة'), arabic_normal_style),
            Paragraph(arabic_text('الإهلاك'), arabic_normal_style),
            Paragraph(arabic_text('القيمة الدفترية'), arabic_normal_style),
            Paragraph(arabic_text('الحالة'), arabic_normal_style),
        ]]

        for asset in assets:
            status_text = {
                'active': 'نشط',
                'inactive': 'غير نشط',
                'disposed': 'مستبعد',
                'sold': 'مباع'
            }.get(asset.status, asset.status)

            data.append([
                Paragraph(arabic_text(asset.asset_number), arabic_normal_style),
                Paragraph(arabic_text(asset.name[:30]), arabic_normal_style),
                Paragraph(arabic_text(asset.category.name if asset.category else '-'), arabic_normal_style),
                Paragraph(arabic_text(asset.branch.name if asset.branch else '-'), arabic_normal_style),
                Paragraph(arabic_text(f"{asset.original_cost:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{asset.accumulated_depreciation:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(f"{asset.book_value:,.2f}"), arabic_normal_style),
                Paragraph(arabic_text(status_text), arabic_normal_style),
            ])

        # إنشاء الجدول
        table = Table(data, colWidths=[2.5*cm, 5*cm, 3*cm, 2.5*cm, 3*cm, 3*cm, 3*cm, 2*cm])

        # تنسيق الجدول
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])

        # ألوان متناوبة للصفوف
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F2F2F2'))

        table.setStyle(table_style)
        elements.append(table)

        # ملاحظة إذا كان هناك المزيد من السجلات
        if assets.count() > 100:
            note_style = ParagraphStyle(
                'NoteStyle',
                fontName=arabic_font,
                fontSize=9,
                textColor=colors.HexColor('#DC3545'),
                alignment=TA_CENTER
            )
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph(
                arabic_text(f"ملاحظة: يعرض هذا التقرير أول 100 أصل. العدد الإجمالي: {assets.count()}"),
                note_style
            ))

        # التذييل
        footer_style = ParagraphStyle(
            'FooterStyle',
            fontName=arabic_font,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(
            arabic_text(f"تم إنشاء هذا التقرير بواسطة: {request.user.get_full_name()} في {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
            footer_style
        ))

        # بناء PDF
        doc.build(elements)

        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=valuation_report_{date.today().strftime("%Y%m%d")}.pdf'

        return response

    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء إنشاء ملف PDF: {str(e)}')
        return redirect('assets:valuation_report')
