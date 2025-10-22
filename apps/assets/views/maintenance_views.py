# apps/assets/views/maintenance_views.py
"""
Views إدارة الصيانة - محسّنة وشاملة
- إدارة أنواع الصيانة
- جدولة الصيانة الدورية
- تنفيذ الصيانة
- تتبع التكاليف
- إدارة الموردين الخارجيين
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView,
    DetailView, TemplateView
)
from django.db.models import (
    Q, Sum, Count, Avg, Max, Min, F,
    DecimalField, Case, When, Value
)
from django.db.models.functions import Coalesce, TruncMonth, TruncDate
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
import json
from datetime import date, timedelta, datetime
from decimal import Decimal
from dateutil.relativedelta import relativedelta
import calendar

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.core.mixins import CompanyMixin, AuditLogMixin
from apps.core.decorators import permission_required_with_message
from ..models import (
    MaintenanceType, MaintenanceSchedule, AssetMaintenance,
    Asset, AssetCategory
)
from apps.core.models import BusinessPartner


# ==================== Maintenance Types ====================

class MaintenanceTypeListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة أنواع الصيانة - محسّنة"""

    model = MaintenanceType
    template_name = 'assets/maintenance/type_list.html'
    context_object_name = 'maintenance_types'
    permission_required = 'assets.view_maintenancetype'
    paginate_by = 50

    def get_queryset(self):
        queryset = MaintenanceType.objects.all().annotate(
            schedule_count=Count('schedules'),
            maintenance_count=Count('maintenances'),
            total_cost=Coalesce(
                Sum('maintenances__labor_cost') +
                Sum('maintenances__parts_cost') +
                Sum('maintenances__other_cost'),
                Decimal('0')
            )
        )

        # الفلترة
        is_active = self.request.GET.get('is_active')
        search = self.request.GET.get('search')

        if is_active:
            queryset = queryset.filter(is_active=(is_active == '1'))

        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(name_en__icontains=search)
            )

        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # إحصائيات
        types = MaintenanceType.objects.all()
        stats = {
            'total': types.count(),
            'active': types.filter(is_active=True).count(),
            'with_schedules': types.annotate(
                count=Count('schedules')
            ).filter(count__gt=0).count(),
        }

        context.update({
            'title': _('أنواع الصيانة'),
            'can_add': self.request.user.has_perm('assets.add_maintenancetype'),
            'can_edit': self.request.user.has_perm('assets.change_maintenancetype'),
            'can_delete': self.request.user.has_perm('assets.delete_maintenancetype'),
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('أنواع الصيانة'), 'url': ''},
            ]
        })
        return context


class MaintenanceTypeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, CreateView):
    """إنشاء نوع صيانة - محسّن"""

    model = MaintenanceType
    template_name = 'assets/maintenance/type_form.html'
    permission_required = 'assets.add_maintenancetype'
    fields = ['code', 'name', 'name_en', 'description', 'is_active']
    success_url = reverse_lazy('assets:maintenance_type_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم إنشاء نوع الصيانة {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة نوع صيانة'),
            'submit_text': _('إنشاء النوع'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('أنواع الصيانة'), 'url': reverse('assets:maintenance_type_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class MaintenanceTypeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, UpdateView):
    """تعديل نوع صيانة - محسّن"""

    model = MaintenanceType
    template_name = 'assets/maintenance/type_form.html'
    permission_required = 'assets.change_maintenancetype'
    fields = ['code', 'name', 'name_en', 'description', 'is_active']
    success_url = reverse_lazy('assets:maintenance_type_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    def form_valid(self, form):
        messages.success(self.request, f'✅ تم تحديث نوع الصيانة {form.instance.name} بنجاح')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل نوع الصيانة {self.object.name}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('أنواع الصيانة'), 'url': reverse('assets:maintenance_type_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class MaintenanceTypeDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """عرض تفاصيل نوع الصيانة - جديد"""

    model = MaintenanceType
    template_name = 'assets/maintenance/type_detail.html'
    context_object_name = 'maintenance_type'
    permission_required = 'assets.view_maintenancetype'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الجدولات
        schedules = self.object.schedules.select_related(
            'asset', 'asset__category'
        ).order_by('-created_at')[:10]

        # الصيانات
        maintenances = self.object.maintenances.select_related(
            'asset', 'asset__category'
        ).order_by('-scheduled_date')[:10]

        # إحصائيات
        stats = maintenances.aggregate(
            total_count=Count('id'),
            completed_count=Count('id', filter=Q(status='completed')),
            total_cost=Coalesce(
                Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
            avg_cost=Coalesce(
                Avg(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
        )

        context.update({
            'title': f'نوع الصيانة: {self.object.name}',
            'can_edit': self.request.user.has_perm('assets.change_maintenancetype'),
            'can_delete': self.request.user.has_perm('assets.delete_maintenancetype'),
            'schedules': schedules,
            'maintenances': maintenances,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('أنواع الصيانة'), 'url': reverse('assets:maintenance_type_list')},
                {'title': self.object.name, 'url': ''},
            ]
        })
        return context


class MaintenanceTypeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف نوع صيانة - محسّن"""

    model = MaintenanceType
    template_name = 'assets/maintenance/type_confirm_delete.html'
    permission_required = 'assets.delete_maintenancetype'
    success_url = reverse_lazy('assets:maintenance_type_list')

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # التحقق
        schedule_count = self.object.schedules.count()
        maintenance_count = self.object.maintenances.count()

        if schedule_count > 0 or maintenance_count > 0:
            messages.error(
                request,
                f'❌ لا يمكن حذف نوع صيانة لديه {schedule_count} جدولات و {maintenance_count} صيانات'
            )
            return redirect('assets:maintenance_type_list')

        type_name = self.object.name
        messages.success(request, f'✅ تم حذف نوع الصيانة {type_name} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Maintenance Schedules ====================

class MaintenanceScheduleListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة جدولة الصيانة - محسّنة"""

    model = MaintenanceSchedule
    template_name = 'assets/maintenance/schedule_list.html'
    context_object_name = 'schedules'
    permission_required = 'assets.view_maintenanceschedule'
    paginate_by = 50

    def get_queryset(self):
        queryset = MaintenanceSchedule.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'asset__branch',
            'maintenance_type', 'assigned_to'
        )

        # الفلترة المتقدمة
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        maintenance_type = self.request.GET.get('maintenance_type')
        frequency = self.request.GET.get('frequency')
        is_active = self.request.GET.get('is_active')
        overdue = self.request.GET.get('overdue')
        due_soon = self.request.GET.get('due_soon')
        search = self.request.GET.get('search')

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if maintenance_type:
            queryset = queryset.filter(maintenance_type_id=maintenance_type)

        if frequency:
            queryset = queryset.filter(frequency=frequency)

        if is_active:
            queryset = queryset.filter(is_active=(is_active == '1'))

        if overdue == '1':
            queryset = queryset.filter(
                is_active=True,
                next_maintenance_date__lt=date.today()
            )

        if due_soon == '1':
            queryset = queryset.filter(
                is_active=True,
                next_maintenance_date__gte=date.today(),
                next_maintenance_date__lte=date.today() + timedelta(days=30)
            )

        if search:
            queryset = queryset.filter(
                Q(schedule_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', 'next_maintenance_date')
        queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # أنواع الصيانة
        maintenance_types = MaintenanceType.objects.filter(is_active=True).order_by('name')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        schedules = MaintenanceSchedule.objects.filter(
            company=self.request.current_company
        )

        stats = schedules.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=Q(is_active=True)),
            inactive_count=Count('id', filter=Q(is_active=False)),
            overdue_count=Count(
                'id',
                filter=Q(is_active=True, next_maintenance_date__lt=date.today())
            ),
            due_soon_count=Count(
                'id',
                filter=Q(
                    is_active=True,
                    next_maintenance_date__gte=date.today(),
                    next_maintenance_date__lte=date.today() + timedelta(days=30)
                )
            ),
            total_estimated=Coalesce(
                Sum('estimated_cost', filter=Q(is_active=True)),
                Decimal('0')
            ),
        )

        context.update({
            'title': _('جدولة الصيانة'),
            'can_add': self.request.user.has_perm('assets.add_maintenanceschedule'),
            'can_edit': self.request.user.has_perm('assets.change_maintenanceschedule'),
            'can_delete': self.request.user.has_perm('assets.delete_maintenanceschedule'),
            'can_export': self.request.user.has_perm('assets.view_maintenanceschedule'),
            'maintenance_types': maintenance_types,
            'categories': categories,
            'frequency_choices': MaintenanceSchedule.FREQUENCY_CHOICES,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('جدولة الصيانة'), 'url': ''},
            ]
        })
        return context


class MaintenanceScheduleCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin,
                                    CreateView):
    """إنشاء جدولة صيانة - محسّن"""

    model = MaintenanceSchedule
    template_name = 'assets/maintenance/schedule_form.html'
    permission_required = 'assets.add_maintenanceschedule'
    fields = [
        'asset', 'maintenance_type', 'frequency', 'custom_days',
        'start_date', 'end_date', 'alert_before_days',
        'assigned_to', 'estimated_cost', 'description', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            status='active'
        ).select_related('category')

        form.fields['maintenance_type'].queryset = MaintenanceType.objects.filter(
            is_active=True
        ).order_by('name')

        # القيم الافتراضية
        form.fields['start_date'].initial = date.today()
        form.fields['alert_before_days'].initial = 7
        form.fields['frequency'].initial = 'monthly'

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user

            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء جدولة الصيانة {self.object.schedule_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:schedule_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة جدولة صيانة'),
            'submit_text': _('إنشاء الجدولة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('جدولة الصيانة'), 'url': reverse('assets:schedule_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class MaintenanceScheduleDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل جدولة الصيانة - محسّن"""

    model = MaintenanceSchedule
    template_name = 'assets/maintenance/schedule_detail.html'
    context_object_name = 'schedule'
    permission_required = 'assets.view_maintenanceschedule'

    def get_queryset(self):
        return MaintenanceSchedule.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'maintenance_type',
            'assigned_to', 'created_by'
        ).prefetch_related('maintenances')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # الصيانات المنفذة
        maintenances = self.object.maintenances.select_related(
            'performed_by', 'external_vendor'
        ).order_by('-scheduled_date')

        # إحصائيات الصيانات
        maintenance_stats = maintenances.aggregate(
            total_count=Count('id'),
            completed_count=Count('id', filter=Q(status='completed')),
            in_progress_count=Count('id', filter=Q(status='in_progress')),
            total_cost=Coalesce(
                Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
            avg_cost=Coalesce(
                Avg(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
        )

        # الأيام المتبقية
        if self.object.next_maintenance_date:
            days_until_next = (self.object.next_maintenance_date - date.today()).days
        else:
            days_until_next = None

        # التحذيرات
        warnings = []
        if self.object.is_overdue():
            warnings.append({
                'type': 'danger',
                'icon': 'fa-exclamation-circle',
                'message': f'الصيانة متأخرة منذ {abs(days_until_next)} يوم'
            })
        elif self.object.is_due_soon():
            warnings.append({
                'type': 'warning',
                'icon': 'fa-clock',
                'message': f'الصيانة مستحقة خلال {days_until_next} يوم'
            })

        context.update({
            'title': f'الجدولة {self.object.schedule_number}',
            'can_edit': self.request.user.has_perm('assets.change_maintenanceschedule'),
            'can_delete': self.request.user.has_perm('assets.delete_maintenanceschedule'),
            'can_add_maintenance': self.request.user.has_perm('assets.add_assetmaintenance'),
            'maintenances': maintenances[:10],
            'maintenance_stats': maintenance_stats,
            'days_until_next': days_until_next,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('جدولة الصيانة'), 'url': reverse('assets:schedule_list')},
                {'title': self.object.schedule_number, 'url': ''},
            ]
        })
        return context


class MaintenanceScheduleUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin,
                                    UpdateView):
    """تعديل جدولة صيانة - محسّن"""

    model = MaintenanceSchedule
    template_name = 'assets/maintenance/schedule_form.html'
    permission_required = 'assets.change_maintenanceschedule'
    fields = [
        'asset', 'maintenance_type', 'frequency', 'custom_days',
        'start_date', 'end_date', 'alert_before_days',
        'assigned_to', 'estimated_cost', 'is_active', 'description', 'notes'
    ]

    def get_queryset(self):
        return MaintenanceSchedule.objects.filter(company=self.request.current_company)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category')

        form.fields['maintenance_type'].queryset = MaintenanceType.objects.filter(
            is_active=True
        ).order_by('name')

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث الجدولة {self.object.schedule_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:schedule_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل الجدولة {self.object.schedule_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('جدولة الصيانة'), 'url': reverse('assets:schedule_list')},
                {'title': self.object.schedule_number, 'url': reverse('assets:schedule_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class MaintenanceScheduleDeleteView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DeleteView):
    """حذف جدولة صيانة - محسّن"""

    model = MaintenanceSchedule
    template_name = 'assets/maintenance/schedule_confirm_delete.html'
    permission_required = 'assets.delete_maintenanceschedule'
    success_url = reverse_lazy('assets:schedule_list')

    def get_queryset(self):
        return MaintenanceSchedule.objects.filter(company=self.request.current_company)

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        schedule_number = self.object.schedule_number
        messages.success(request, f'✅ تم حذف الجدولة {schedule_number} بنجاح')

        return super().delete(request, *args, **kwargs)


# ==================== Asset Maintenance ====================

class AssetMaintenanceListView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, ListView):
    """قائمة صيانة الأصول - محسّنة"""

    model = AssetMaintenance
    template_name = 'assets/maintenance/maintenance_list.html'
    context_object_name = 'maintenances'
    permission_required = 'assets.view_assetmaintenance'
    paginate_by = 50

    def get_queryset(self):
        queryset = AssetMaintenance.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'asset__branch',
            'maintenance_type', 'maintenance_schedule',
            'performed_by', 'external_vendor'
        )

        # الفلترة المتقدمة
        status = self.request.GET.get('status')
        asset = self.request.GET.get('asset')
        category = self.request.GET.get('category')
        maintenance_type = self.request.GET.get('maintenance_type')
        maintenance_category = self.request.GET.get('maintenance_category')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)

        if asset:
            queryset = queryset.filter(asset_id=asset)

        if category:
            queryset = queryset.filter(asset__category_id=category)

        if maintenance_type:
            queryset = queryset.filter(maintenance_type_id=maintenance_type)

        if maintenance_category:
            queryset = queryset.filter(maintenance_category=maintenance_category)

        if date_from:
            queryset = queryset.filter(scheduled_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(scheduled_date__lte=date_to)

        if search:
            queryset = queryset.filter(
                Q(maintenance_number__icontains=search) |
                Q(asset__asset_number__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(description__icontains=search)
            )

        # الترتيب
        sort_by = self.request.GET.get('sort', '-scheduled_date')
        queryset = queryset.order_by(sort_by, '-maintenance_number')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # أنواع الصيانة
        maintenance_types = MaintenanceType.objects.filter(is_active=True).order_by('name')

        # الفئات
        categories = AssetCategory.objects.filter(
            company=self.request.current_company,
            is_active=True
        ).order_by('code')

        # إحصائيات مفصّلة
        maintenances = AssetMaintenance.objects.filter(
            company=self.request.current_company
        )

        stats = maintenances.aggregate(
            total_count=Count('id'),
            scheduled_count=Count('id', filter=Q(status='scheduled')),
            in_progress_count=Count('id', filter=Q(status='in_progress')),
            completed_count=Count('id', filter=Q(status='completed')),
            cancelled_count=Count('id', filter=Q(status='cancelled')),
            total_cost=Coalesce(
                Sum(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
            avg_cost=Coalesce(
                Avg(F('labor_cost') + F('parts_cost') + F('other_cost')),
                Decimal('0')
            ),
        )

        # نسبة الإنجاز
        if stats['total_count'] > 0:
            completion_rate = (stats['completed_count'] / stats['total_count'] * 100)
        else:
            completion_rate = 0

        stats['completion_rate'] = round(completion_rate, 2)

        context.update({
            'title': _('صيانة الأصول'),
            'can_add': self.request.user.has_perm('assets.add_assetmaintenance'),
            'can_edit': self.request.user.has_perm('assets.change_assetmaintenance'),
            'can_export': self.request.user.has_perm('assets.view_assetmaintenance'),
            'status_choices': AssetMaintenance.STATUS_CHOICES,
            'category_choices': AssetMaintenance.MAINTENANCE_CATEGORY,
            'maintenance_types': maintenance_types,
            'categories': categories,
            'stats': stats,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('صيانة الأصول'), 'url': ''},
            ]
        })
        return context


class AssetMaintenanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, CreateView):
    """إنشاء صيانة أصل - محسّن"""

    model = AssetMaintenance
    template_name = 'assets/maintenance/maintenance_form.html'
    permission_required = 'assets.add_assetmaintenance'
    fields = [
        'asset', 'maintenance_type', 'maintenance_category',
        'maintenance_schedule', 'scheduled_date', 'start_date',
        'performed_by', 'external_vendor', 'vendor_invoice_number',
        'labor_cost', 'parts_cost', 'other_cost',
        'is_capital_improvement', 'parts_description',
        'description', 'issues_found', 'recommendations', 'notes'
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company,
            status__in=['active', 'under_maintenance']
        ).select_related('category')

        form.fields['maintenance_type'].queryset = MaintenanceType.objects.filter(
            is_active=True
        ).order_by('name')

        form.fields['maintenance_schedule'].queryset = MaintenanceSchedule.objects.filter(
            company=company,
            is_active=True
        ).select_related('asset')
        form.fields['maintenance_schedule'].required = False

        form.fields['external_vendor'].queryset = BusinessPartner.objects.filter(
            company=company,
            partner_type__in=['supplier', 'both']
        ).order_by('name')
        form.fields['external_vendor'].required = False

        # القيم الافتراضية
        form.fields['scheduled_date'].initial = date.today()
        form.fields['maintenance_category'].initial = 'preventive'

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            form.instance.company = self.request.current_company
            form.instance.branch = self.request.current_branch
            form.instance.created_by = self.request.user
            form.instance.status = 'scheduled'

            self.object = form.save()

            self.log_action('create', self.object)

            messages.success(
                self.request,
                f'✅ تم إنشاء الصيانة {self.object.maintenance_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except ValidationError as e:
            messages.error(self.request, f'❌ {str(e)}')
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:maintenance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': _('إضافة صيانة أصل'),
            'submit_text': _('إنشاء الصيانة'),
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('صيانة الأصول'), 'url': reverse('assets:maintenance_list')},
                {'title': _('إضافة'), 'url': ''},
            ]
        })
        return context


class AssetMaintenanceDetailView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, DetailView):
    """عرض تفاصيل صيانة الأصل - محسّن"""

    model = AssetMaintenance
    template_name = 'assets/maintenance/maintenance_detail.html'
    context_object_name = 'maintenance'
    permission_required = 'assets.view_assetmaintenance'

    def get_queryset(self):
        return AssetMaintenance.objects.filter(
            company=self.request.current_company
        ).select_related(
            'asset', 'asset__category', 'maintenance_type',
            'maintenance_schedule', 'performed_by', 'external_vendor',
            'journal_entry', 'created_by'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # مدة الصيانة
        if self.object.start_date and self.object.completion_date:
            duration_days = (self.object.completion_date - self.object.start_date).days
        else:
            duration_days = None

        # التكلفة الإجمالية
        total_cost = (
                self.object.labor_cost +
                self.object.parts_cost +
                self.object.other_cost
        )

        # التحذيرات
        warnings = []
        if self.object.is_capital_improvement:
            warnings.append({
                'type': 'info',
                'icon': 'fa-info-circle',
                'message': 'هذه صيانة رأسمالية - يجب رسملتها على الأصل'
            })

        context.update({
            'title': f'الصيانة {self.object.maintenance_number}',
            'can_edit': (
                    self.request.user.has_perm('assets.change_assetmaintenance') and
                    self.object.status != 'completed'
            ),
            'can_start': (
                    self.request.user.has_perm('assets.change_assetmaintenance') and
                    self.object.status == 'scheduled'
            ),
            'can_complete': (
                    self.request.user.has_perm('assets.change_assetmaintenance') and
                    self.object.status == 'in_progress'
            ),
            'can_cancel': (
                    self.request.user.has_perm('assets.change_assetmaintenance') and
                    self.object.status in ['scheduled', 'in_progress']
            ),
            'duration_days': duration_days,
            'total_cost': total_cost,
            'warnings': warnings,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('صيانة الأصول'), 'url': reverse('assets:maintenance_list')},
                {'title': self.object.maintenance_number, 'url': ''},
            ]
        })
        return context


class AssetMaintenanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, AuditLogMixin, UpdateView):
    """تعديل صيانة أصل - محسّن"""

    model = AssetMaintenance
    template_name = 'assets/maintenance/maintenance_form.html'
    permission_required = 'assets.change_assetmaintenance'
    fields = [
        'asset', 'maintenance_type', 'maintenance_category',
        'maintenance_schedule', 'scheduled_date', 'start_date', 'completion_date',
        'status', 'performed_by', 'external_vendor', 'vendor_invoice_number',
        'labor_cost', 'parts_cost', 'other_cost',
        'is_capital_improvement', 'parts_description',
        'description', 'issues_found', 'recommendations', 'notes'
    ]

    def get_queryset(self):
        return AssetMaintenance.objects.filter(
            company=self.request.current_company
        ).exclude(status='completed')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        company = self.request.current_company

        form.fields['asset'].queryset = Asset.objects.filter(
            company=company
        ).select_related('category')

        form.fields['maintenance_type'].queryset = MaintenanceType.objects.filter(
            is_active=True
        ).order_by('name')

        form.fields['maintenance_schedule'].queryset = MaintenanceSchedule.objects.filter(
            company=company
        ).select_related('asset')
        form.fields['maintenance_schedule'].required = False

        form.fields['external_vendor'].queryset = BusinessPartner.objects.filter(
            company=company,
            partner_type__in=['supplier', 'both']
        ).order_by('name')
        form.fields['external_vendor'].required = False

        # إضافة classes
        for field_name, field in form.fields.items():
            if field.widget.__class__.__name__ not in ['CheckboxInput', 'RadioSelect', 'Textarea']:
                field.widget.attrs.update({'class': 'form-control'})
            elif field.widget.__class__.__name__ == 'Textarea':
                field.widget.attrs.update({'class': 'form-control', 'rows': 3})

        return form

    @transaction.atomic
    def form_valid(self, form):
        try:
            self.object = form.save()

            self.log_action('update', self.object)

            messages.success(
                self.request,
                f'✅ تم تحديث الصيانة {self.object.maintenance_number} بنجاح'
            )

            return redirect(self.get_success_url())

        except Exception as e:
            messages.error(self.request, f'❌ خطأ: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('assets:maintenance_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': f'تعديل الصيانة {self.object.maintenance_number}',
            'submit_text': _('حفظ التعديلات'),
            'is_update': True,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('صيانة الأصول'), 'url': reverse('assets:maintenance_list')},
                {'title': self.object.maintenance_number,
                 'url': reverse('assets:maintenance_detail', args=[self.object.pk])},
                {'title': _('تعديل'), 'url': ''},
            ]
        })
        return context


class StartMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """بدء الصيانة - جديد"""

    template_name = 'assets/maintenance/start_maintenance.html'
    permission_required = 'assets.change_assetmaintenance'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        maintenance_id = self.kwargs.get('pk')
        maintenance = get_object_or_404(
            AssetMaintenance,
            pk=maintenance_id,
            company=self.request.current_company,
            status='scheduled'
        )

        context.update({
            'title': f'بدء الصيانة {maintenance.maintenance_number}',
            'maintenance': maintenance,
            'breadcrumbs': [
                {'title': _('الأصول الثابتة'), 'url': reverse('assets:dashboard')},
                {'title': _('صيانة الأصول'), 'url': reverse('assets:maintenance_list')},
                {'title': maintenance.maintenance_number,
                 'url': reverse('assets:maintenance_detail', args=[maintenance.pk])},
                {'title': _('بدء الصيانة'), 'url': ''},
            ]
        })
        return context

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            maintenance_id = kwargs.get('pk')
            maintenance = get_object_or_404(
                AssetMaintenance,
                pk=maintenance_id,
                company=request.current_company,
                status='scheduled'
            )

            start_date = request.POST.get('start_date', date.today())

            # بدء الصيانة
            maintenance.status = 'in_progress'
            maintenance.start_date = start_date
            maintenance.save()

            # تحديث حالة الأصل
            asset = maintenance.asset
            asset.status = 'under_maintenance'
            asset.save()

            messages.success(
                request,
                f'✅ تم بدء الصيانة {maintenance.maintenance_number} بنجاح'
            )

            return redirect('assets:maintenance_detail', pk=maintenance.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في بدء الصيانة: {str(e)}')
            return redirect('assets:maintenance_detail', pk=maintenance_id)


class CancelMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, CompanyMixin, TemplateView):
    """إلغاء الصيانة - جديد"""

    template_name = 'assets/maintenance/cancel_maintenance.html'
    permission_required = 'assets.change_assetmaintenance'

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            maintenance_id = kwargs.get('pk')
            maintenance = get_object_or_404(
                AssetMaintenance,
                pk=maintenance_id,
                company=request.current_company
            )

            if maintenance.status not in ['scheduled', 'in_progress']:
                messages.error(request, 'لا يمكن إلغاء هذه الصيانة')
                return redirect('assets:maintenance_detail', pk=maintenance.pk)

            cancellation_reason = request.POST.get('cancellation_reason', '')

            # إلغاء الصيانة
            maintenance.status = 'cancelled'
            maintenance.notes = f"{maintenance.notes}\nإلغاء: {cancellation_reason}" if maintenance.notes else f"إلغاء: {cancellation_reason}"
            maintenance.save()

            # إرجاع حالة الأصل إلى نشط إذا كان تحت الصيانة
            if maintenance.asset.status == 'under_maintenance':
                maintenance.asset.status = 'active'
                maintenance.asset.save()

            messages.success(
                request,
                f'✅ تم إلغاء الصيانة {maintenance.maintenance_number} بنجاح'
            )

            return redirect('assets:maintenance_detail', pk=maintenance.pk)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'❌ خطأ في إلغاء الصيانة: {str(e)}')
            return redirect('assets:maintenance_detail', pk=maintenance_id)


# ==================== Ajax Views - محسّنة ====================

@login_required
@permission_required_with_message('assets.change_assetmaintenance')
@require_http_methods(["POST"])
def complete_maintenance(request, pk):
    """إكمال صيانة أصل - محسّن"""

    try:
        maintenance = get_object_or_404(
            AssetMaintenance,
            pk=pk,
            company=request.current_company,
            status='in_progress'
        )

        completion_date = request.POST.get('completion_date', date.today())
        issues_found = request.POST.get('issues_found', '')
        recommendations = request.POST.get('recommendations', '')

        # إكمال الصيانة
        maintenance.mark_as_completed(
            completion_date=completion_date,
            user=request.user
        )

        # تحديث الحقول الإضافية
        if issues_found:
            maintenance.issues_found = issues_found
        if recommendations:
            maintenance.recommendations = recommendations
        maintenance.save()

        return JsonResponse({
            'success': True,
            'message': f'تم إكمال الصيانة {maintenance.maintenance_number} بنجاح',
            'maintenance_number': maintenance.maintenance_number,
            'total_cost': float(maintenance.total_cost),
        })

    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إكمال الصيانة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_assetmaintenance')
@require_http_methods(["POST"])
def start_maintenance_ajax(request, pk):
    """بدء الصيانة عبر Ajax - جديد"""

    try:
        maintenance = get_object_or_404(
            AssetMaintenance,
            pk=pk,
            company=request.current_company,
            status='scheduled'
        )

        start_date = request.POST.get('start_date', date.today())

        maintenance.status = 'in_progress'
        maintenance.start_date = start_date
        maintenance.save()

        # تحديث حالة الأصل
        asset = maintenance.asset
        asset.status = 'under_maintenance'
        asset.save()

        return JsonResponse({
            'success': True,
            'message': f'تم بدء الصيانة {maintenance.maintenance_number} بنجاح'
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في بدء الصيانة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.change_assetmaintenance')
@require_http_methods(["POST"])
def cancel_maintenance_ajax(request, pk):
    """إلغاء الصيانة عبر Ajax - جديد"""

    try:
        maintenance = get_object_or_404(
            AssetMaintenance,
            pk=pk,
            company=request.current_company
        )

        if maintenance.status not in ['scheduled', 'in_progress']:
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن إلغاء هذه الصيانة'
            }, status=400)

        cancellation_reason = request.POST.get('cancellation_reason', '')

        maintenance.status = 'cancelled'
        maintenance.notes = f"{maintenance.notes}\nإلغاء: {cancellation_reason}" if maintenance.notes else f"إلغاء: {cancellation_reason}"
        maintenance.save()

        # إرجاع حالة الأصل
        if maintenance.asset.status == 'under_maintenance':
            maintenance.asset.status = 'active'
            maintenance.asset.save()

        return JsonResponse({
            'success': True,
            'message': f'تم إلغاء الصيانة {maintenance.maintenance_number} بنجاح'
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': f'خطأ في إلغاء الصيانة: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetmaintenance')
@require_http_methods(["GET"])
def maintenance_datatable_ajax(request):
    """Ajax endpoint لجدول الصيانة - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        status = request.GET.get('status', '')
        maintenance_type = request.GET.get('maintenance_type', '')

        # Query
        queryset = AssetMaintenance.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'asset__category', 'maintenance_type', 'journal_entry'
        )

        # تطبيق الفلاتر
        if status:
            queryset = queryset.filter(status=status)

        if maintenance_type:
            queryset = queryset.filter(maintenance_type_id=maintenance_type)

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(maintenance_number__icontains=search_value) |
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value) |
                Q(description__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'desc')

        order_columns = {
            '0': 'maintenance_number',
            '1': 'scheduled_date',
            '2': 'asset__asset_number',
            '3': 'asset__name',
            '4': 'maintenance_type__name',
            '5': 'labor_cost',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field, '-maintenance_number')
        else:
            queryset = queryset.order_by('-scheduled_date', '-maintenance_number')

        # العد
        total_records = AssetMaintenance.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_assetmaintenance')
        can_edit = request.user.has_perm('assets.change_assetmaintenance')

        for maint in queryset:
            # الحالة
            status_map = {
                'scheduled': '<span class="badge bg-info"><i class="fas fa-calendar"></i> مجدولة</span>',
                'in_progress': '<span class="badge bg-warning"><i class="fas fa-cog fa-spin"></i> جارية</span>',
                'completed': '<span class="badge bg-success"><i class="fas fa-check-circle"></i> مكتملة</span>',
                'cancelled': '<span class="badge bg-danger"><i class="fas fa-ban"></i> ملغاة</span>',
            }
            status_badge = status_map.get(maint.status, maint.status)

            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:maintenance_detail', args=[maint.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            if maint.status != 'completed' and can_edit:
                actions.append(f'''
                    <a href="{reverse('assets:maintenance_update', args=[maint.pk])}" 
                       class="btn btn-outline-primary btn-sm" title="تعديل" data-bs-toggle="tooltip">
                        <i class="fas fa-edit"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                f'<a href="{reverse("assets:maintenance_detail", args=[maint.pk])}">{maint.maintenance_number}</a>',
                maint.scheduled_date.strftime('%Y-%m-%d'),
                f'<a href="{reverse("assets:asset_detail", args=[maint.asset.pk])}">{maint.asset.asset_number}</a>',
                f'''<div>
                    <strong>{maint.asset.name}</strong>
                    <br><small class="text-muted">{maint.asset.category.name}</small>
                </div>''',
                maint.maintenance_type.name,
                f'<div class="text-end"><strong>{maint.total_cost:,.2f}</strong></div>',
                status_badge,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_maintenanceschedule')
@require_http_methods(["GET"])
def schedule_datatable_ajax(request):
    """Ajax endpoint لجدول جدولة الصيانة - محسّن"""

    if not hasattr(request, 'current_company') or not request.current_company:
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'لا توجد شركة محددة'
        })

    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()

        # الفلاتر
        is_active = request.GET.get('is_active', '')

        # Query
        queryset = MaintenanceSchedule.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'asset__category', 'maintenance_type'
        )

        # تطبيق الفلاتر
        if is_active:
            queryset = queryset.filter(is_active=(is_active == '1'))

        # البحث
        if search_value:
            queryset = queryset.filter(
                Q(schedule_number__icontains=search_value) |
                Q(asset__asset_number__icontains=search_value) |
                Q(asset__name__icontains=search_value)
            )

        # الترتيب
        order_column_index = request.GET.get('order[0][column]')
        order_dir = request.GET.get('order[0][dir]', 'asc')

        order_columns = {
            '0': 'schedule_number',
            '1': 'asset__asset_number',
            '2': 'maintenance_type__name',
            '3': 'frequency',
            '4': 'next_maintenance_date',
        }

        if order_column_index and order_column_index in order_columns:
            order_field = order_columns[order_column_index]
            if order_dir == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field)
        else:
            queryset = queryset.order_by('next_maintenance_date')

        # العد
        total_records = MaintenanceSchedule.objects.filter(
            company=request.current_company
        ).count()
        filtered_records = queryset.count()

        # Pagination
        queryset = queryset[start:start + length]

        # إعداد البيانات
        data = []
        can_view = request.user.has_perm('assets.view_maintenanceschedule')

        for schedule in queryset:
            # حالة الجدولة
            if schedule.is_overdue():
                days_overdue = (date.today() - schedule.next_maintenance_date).days
                status = f'<span class="badge bg-danger"><i class="fas fa-exclamation-circle"></i> متأخرة ({days_overdue} يوم)</span>'
            elif schedule.is_due_soon():
                days_left = (schedule.next_maintenance_date - date.today()).days
                status = f'<span class="badge bg-warning"><i class="fas fa-clock"></i> قريبة ({days_left} يوم)</span>'
            elif schedule.is_active:
                status = '<span class="badge bg-success"><i class="fas fa-check-circle"></i> نشطة</span>'
            else:
                status = '<span class="badge bg-secondary"><i class="fas fa-pause-circle"></i> غير نشطة</span>'

            # أزرار الإجراءات
            actions = []

            if can_view:
                actions.append(f'''
                    <a href="{reverse('assets:schedule_detail', args=[schedule.pk])}" 
                       class="btn btn-outline-info btn-sm" title="عرض" data-bs-toggle="tooltip">
                        <i class="fas fa-eye"></i>
                    </a>
                ''')

            actions_html = '<div class="btn-group" role="group">' + ' '.join(actions) + '</div>' if actions else '-'

            data.append([
                f'<a href="{reverse("assets:schedule_detail", args=[schedule.pk])}">{schedule.schedule_number}</a>',
                f'<a href="{reverse("assets:asset_detail", args=[schedule.asset.pk])}">{schedule.asset.asset_number}</a>',
                schedule.maintenance_type.name,
                schedule.get_frequency_display(),
                f'''<div>
                    {schedule.next_maintenance_date.strftime('%Y-%m-%d')}
                    <br><small class="text-muted">بدأت: {schedule.start_date.strftime('%Y-%m-%d')}</small>
                </div>''',
                status,
                actions_html
            ])

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'draw': int(request.GET.get('draw', 1)),
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': f'خطأ في تحميل البيانات: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_maintenanceschedule')
@require_http_methods(["GET"])
def generate_schedule_dates(request):
    """توليد تواريخ الصيانة المستقبلية - محسّن"""

    try:
        schedule_id = request.GET.get('schedule_id')
        months = int(request.GET.get('months', 12))

        if not schedule_id:
            return JsonResponse({
                'success': False,
                'error': 'يجب تحديد الجدولة'
            }, status=400)

        schedule = get_object_or_404(
            MaintenanceSchedule,
            pk=schedule_id,
            company=request.current_company
        )

        dates = []
        current_date = schedule.next_maintenance_date

        for i in range(months):
            if not current_date:
                break

            dates.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'month': current_date.strftime('%B %Y'),
                'day_of_week': current_date.strftime('%A'),
                'estimated_cost': float(schedule.estimated_cost)
            })

            # حساب التاريخ التالي
            if schedule.frequency == 'daily':
                current_date = current_date + timedelta(days=1)
            elif schedule.frequency == 'weekly':
                current_date = current_date + timedelta(weeks=1)
            elif schedule.frequency == 'monthly':
                current_date = current_date + relativedelta(months=1)
            elif schedule.frequency == 'quarterly':
                current_date = current_date + relativedelta(months=3)
            elif schedule.frequency == 'semi_annual':
                current_date = current_date + relativedelta(months=6)
            elif schedule.frequency == 'annual':
                current_date = current_date + relativedelta(years=1)
            elif schedule.frequency == 'custom' and schedule.custom_days:
                current_date = current_date + timedelta(days=schedule.custom_days)
            else:
                break

            # التحقق من تجاوز تاريخ الانتهاء
            if schedule.end_date and current_date > schedule.end_date:
                break

        # حساب الإجماليات
        total_cost = sum(d['estimated_cost'] for d in dates)

        return JsonResponse({
            'success': True,
            'dates': dates,
            'total_dates': len(dates),
            'total_cost': total_cost,
            'schedule': {
                'schedule_number': schedule.schedule_number,
                'asset': schedule.asset.name,
                'maintenance_type': schedule.maintenance_type.name,
                'frequency': schedule.get_frequency_display(),
            }
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'خطأ في توليد التواريخ: {str(e)}'
        }, status=500)


@login_required
@permission_required_with_message('assets.view_assetmaintenance')
@require_http_methods(["GET"])
def maintenance_calendar_ajax(request):
    """تقويم الصيانة - جديد"""

    try:
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))

        # بداية ونهاية الشهر
        start_date = date(year, month, 1)
        _, last_day = calendar.monthrange(year, month)
        end_date = date(year, month, last_day)

        # الصيانات المجدولة
        maintenances = AssetMaintenance.objects.filter(
            company=request.current_company,
            scheduled_date__gte=start_date,
            scheduled_date__lte=end_date
        ).select_related('asset', 'maintenance_type')

        events = []
        for maint in maintenances:
            color = {
                'scheduled': '#17a2b8',
                'in_progress': '#ffc107',
                'completed': '#28a745',
                'cancelled': '#dc3545',
            }.get(maint.status, '#6c757d')

            events.append({
                'id': maint.id,
                'title': f'{maint.asset.asset_number} - {maint.maintenance_type.name}',
                'start': maint.scheduled_date.strftime('%Y-%m-%d'),
                'backgroundColor': color,
                'borderColor': color,
                'url': reverse('assets:maintenance_detail', args=[maint.pk]),
            })

        return JsonResponse({
            'success': True,
            'events': events
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'خطأ في تحميل التقويم: {str(e)}'
        }, status=500)


# ==================== Export Functions - جديد ====================

@login_required
@permission_required_with_message('assets.view_assetmaintenance')
@require_http_methods(["GET"])
def export_maintenance_list_excel(request):
    """تصدير قائمة الصيانة إلى Excel - جديد"""

    try:
        # إنشاء workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance"

        # تنسيق الرأس
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'Maintenance Number', 'Scheduled Date', 'Asset Number',
            'Asset Name', 'Maintenance Type', 'Category',
            'Labor Cost', 'Parts Cost', 'Other Cost', 'Total Cost',
            'Status', 'Completion Date'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # البيانات
        maintenances = AssetMaintenance.objects.filter(
            company=request.current_company
        ).select_related(
            'asset', 'maintenance_type'
        ).order_by('-scheduled_date')

        row_num = 2
        for maint in maintenances:
            ws.cell(row=row_num, column=1, value=maint.maintenance_number)
            ws.cell(row=row_num, column=2, value=maint.scheduled_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=maint.asset.asset_number)
            ws.cell(row=row_num, column=4, value=maint.asset.name)
            ws.cell(row=row_num, column=5, value=maint.maintenance_type.name)
            ws.cell(row=row_num, column=6, value=maint.get_maintenance_category_display())
            ws.cell(row=row_num, column=7, value=float(maint.labor_cost))
            ws.cell(row=row_num, column=8, value=float(maint.parts_cost))
            ws.cell(row=row_num, column=9, value=float(maint.other_cost))
            ws.cell(row=row_num, column=10, value=float(maint.total_cost))
            ws.cell(row=row_num, column=11, value=maint.get_status_display())
            ws.cell(row=row_num, column=12,
                    value=maint.completion_date.strftime('%Y-%m-%d') if maint.completion_date else '')
            row_num += 1

        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # حفظ
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="maintenance.xlsx"'

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        messages.error(request, f'خطأ في التصدير: {str(e)}')
        return redirect('assets:maintenance_list')