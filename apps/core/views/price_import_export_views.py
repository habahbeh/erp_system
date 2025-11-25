# apps/core/views/price_import_export_views.py
"""
Price Import/Export Views
Import and export pricing data (Price Lists, Price List Items, Pricing Rules)
"""

from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Q
from decimal import Decimal
from datetime import datetime

from apps.core.models import (
    PriceList, PriceListItem, PricingRule,
    Item, ItemVariant, ItemCategory, Company
)
from apps.core.utils.import_export import (
    ExcelExporter, CSVExporter, ExcelImporter, CSVImporter,
    BulkImporter, DataValidator, TemplateGenerator
)


class PriceListExportView(LoginRequiredMixin, View):
    """
    Export Price List Items to Excel or CSV
    """

    def get(self, request, *args, **kwargs):
        company = request.current_company
        export_format = request.GET.get('format', 'excel')
        price_list_id = request.GET.get('price_list')

        # Get price list items
        queryset = PriceListItem.objects.filter(
            price_list__company=company
        ).select_related('price_list', 'item', 'variant')

        if price_list_id:
            queryset = queryset.filter(price_list_id=price_list_id)

        queryset = queryset.order_by('price_list__name', 'item__code')

        # Prepare data
        columns = [
            'قائمة الأسعار',
            'رمز الصنف',
            'اسم الصنف',
            'المتغير',
            'السعر',
            'العملة',
            'تاريخ البدء',
            'تاريخ الانتهاء',
            'الحد الأدنى للكمية',
            'نشط'
        ]

        rows = []
        for item in queryset:
            rows.append([
                item.price_list.name,
                item.item.code,
                item.item.name,
                item.variant.code if item.variant else '',
                float(item.price),
                item.currency.code if item.currency else '',
                item.valid_from.strftime('%Y-%m-%d') if item.valid_from else '',
                item.valid_to.strftime('%Y-%m-%d') if item.valid_to else '',
                int(item.min_quantity) if item.min_quantity else 1,
                'نعم' if item.is_active else 'لا'
            ])

        # Generate file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'price_list_items_{timestamp}'

        if export_format == 'csv':
            exporter = CSVExporter(columns)
            exporter.add_rows(rows)
            return exporter.get_response(filename)
        else:
            exporter = ExcelExporter(title='عناصر قوائم الأسعار')
            exporter.add_header(columns)
            exporter.add_rows(rows)
            exporter.freeze_panes(2, 1)
            exporter.add_filter()

            # Add summary
            exporter.add_summary_row(
                'الإجمالي',
                [len(rows), '', '', '', '', '', '', '', '', '']
            )

            return exporter.get_response(filename)


class PriceListImportView(LoginRequiredMixin, TemplateView):
    """
    Import Price List Items from Excel or CSV
    """
    template_name = 'core/pricing/import_prices.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['price_lists'] = PriceList.objects.filter(
            company=self.request.current_company,
            is_active=True
        )
        return context

    def post(self, request, *args, **kwargs):
        company = request.current_company
        uploaded_file = request.FILES.get('import_file')
        file_format = request.POST.get('format', 'excel')
        price_list_id = request.POST.get('price_list')

        if not uploaded_file:
            messages.error(request, 'الرجاء اختيار ملف للاستيراد')
            return redirect('core:price_list_import')

        # Get price list
        try:
            price_list = PriceList.objects.get(id=price_list_id, company=company)
        except PriceList.DoesNotExist:
            messages.error(request, 'قائمة الأسعار غير موجودة')
            return redirect('core:price_list_import')

        # Define expected columns
        expected_columns = [
            'رمز الصنف',
            'السعر',
            'تاريخ البدء',
            'تاريخ الانتهاء',
            'الحد الأدنى للكمية',
            'نشط'
        ]

        # Import based on format
        if file_format == 'csv':
            importer = CSVImporter(uploaded_file, expected_columns)
        else:
            importer = ExcelImporter(uploaded_file, expected_columns)

        # Validate file
        if not importer.validate_file():
            for error in importer.errors:
                messages.error(request, error['message'])
            return redirect('core:price_list_import')

        # Get rows
        rows = importer.get_rows()

        # Process rows
        def process_row(row):
            try:
                # Get item by code
                item_code = DataValidator.validate_required(row.get('رمز الصنف'), 'رمز الصنف')
                item = Item.objects.get(code=item_code, company=company)

                # Validate price
                price = DataValidator.validate_positive_decimal(row.get('السعر'), 'السعر')

                # Validate dates
                valid_from = DataValidator.validate_date(row.get('تاريخ البدء'))
                valid_to = DataValidator.validate_date(row.get('تاريخ الانتهاء'))

                # Validate quantity
                min_quantity = row.get('الحد الأدنى للكمية', 1)
                if min_quantity:
                    min_quantity = int(min_quantity)

                # Validate active
                is_active = DataValidator.validate_boolean(row.get('نشط', True))

                # Create or update price list item
                price_item, created = PriceListItem.objects.update_or_create(
                    price_list=price_list,
                    item=item,
                    variant=None,
                    defaults={
                        'price': price,
                        'valid_from': valid_from,
                        'valid_to': valid_to,
                        'min_quantity': min_quantity,
                        'is_active': is_active,
                        'created_by': request.user if created else None
                    }
                )

                return price_item

            except Item.DoesNotExist:
                raise Exception(f'الصنف برمز "{row.get("رمز الصنف")}" غير موجود')
            except Exception as e:
                raise Exception(str(e))

        # Bulk import
        bulk_importer = BulkImporter(PriceListItem, request.user, company)
        results = bulk_importer.import_rows(rows, process_row)

        # Close importer
        if hasattr(importer, 'close'):
            importer.close()

        # Store results in session
        request.session['import_results'] = results

        messages.success(
            request,
            f'تم استيراد {results["success"]} سعر بنجاح. '
            f'الأخطاء: {results["errors"]}'
        )

        return redirect('core:price_list_import_results')


class PriceListImportResultsView(LoginRequiredMixin, TemplateView):
    """
    Display import results
    """
    template_name = 'core/pricing/import_results.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['results'] = self.request.session.get('import_results', {})
        return context


class PricingRuleExportView(LoginRequiredMixin, View):
    """
    Export Pricing Rules to Excel or CSV
    """

    def get(self, request, *args, **kwargs):
        company = request.current_company
        export_format = request.GET.get('format', 'excel')

        # Get pricing rules
        queryset = PricingRule.objects.filter(
            company=company
        ).select_related('category', 'price_list').order_by('-created_at')

        # Prepare data
        columns = [
            'الرمز',
            'الاسم',
            'النوع',
            'الفئة',
            'قائمة الأسعار',
            'القيمة',
            'نوع القيمة',
            'الأولوية',
            'تاريخ البدء',
            'تاريخ الانتهاء',
            'نشط'
        ]

        rows = []
        for rule in queryset:
            rows.append([
                rule.code,
                rule.name,
                rule.get_rule_type_display(),
                rule.category.name if rule.category else '',
                rule.price_list.name if rule.price_list else '',
                float(rule.value),
                rule.get_value_type_display(),
                rule.priority,
                rule.valid_from.strftime('%Y-%m-%d') if rule.valid_from else '',
                rule.valid_to.strftime('%Y-%m-%d') if rule.valid_to else '',
                'نعم' if rule.is_active else 'لا'
            ])

        # Generate file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'pricing_rules_{timestamp}'

        if export_format == 'csv':
            exporter = CSVExporter(columns)
            exporter.add_rows(rows)
            return exporter.get_response(filename)
        else:
            exporter = ExcelExporter(title='قواعد التسعير')
            exporter.add_header(columns)
            exporter.add_rows(rows)
            exporter.freeze_panes(2, 1)
            exporter.add_filter()
            return exporter.get_response(filename)


class PriceListTemplateDownloadView(LoginRequiredMixin, View):
    """
    Download import template
    """

    def get(self, request, *args, **kwargs):
        template_format = request.GET.get('format', 'excel')

        columns = [
            'رمز الصنف',
            'السعر',
            'تاريخ البدء',
            'تاريخ الانتهاء',
            'الحد الأدنى للكمية',
            'نشط'
        ]

        # Sample data
        sample_data = [
            ['ITEM001', '100.00', '2025-01-01', '2025-12-31', '1', 'نعم'],
            ['ITEM002', '250.50', '2025-01-01', '2025-12-31', '5', 'نعم'],
            ['ITEM003', '75.00', '2025-01-01', '', '1', 'لا'],
        ]

        if template_format == 'csv':
            return TemplateGenerator.generate_csv_template(
                columns,
                sample_data,
                'price_list_template'
            )
        else:
            return TemplateGenerator.generate_excel_template(
                columns,
                sample_data,
                'price_list_template'
            )


class BulkPriceExportView(LoginRequiredMixin, View):
    """
    Export all pricing data (Price Lists, Items, Rules) in one file
    """

    def get(self, request, *args, **kwargs):
        company = request.current_company
        export_format = request.GET.get('format', 'excel')

        if export_format == 'excel':
            return self._export_excel(company)
        else:
            # CSV exports one sheet at a time
            sheet = request.GET.get('sheet', 'price_items')
            return self._export_csv(company, sheet)

    def _export_excel(self, company):
        """Export to Excel with multiple sheets"""
        from openpyxl import Workbook

        workbook = Workbook()

        # Sheet 1: Price List Items
        ws1 = workbook.active
        ws1.title = 'عناصر الأسعار'

        columns1 = ['قائمة الأسعار', 'رمز الصنف', 'اسم الصنف', 'السعر', 'نشط']
        ws1.append(columns1)

        price_items = PriceListItem.objects.filter(
            price_list__company=company
        ).select_related('price_list', 'item')[:1000]  # Limit for performance

        for item in price_items:
            ws1.append([
                item.price_list.name,
                item.item.code,
                item.item.name,
                float(item.price),
                'نعم' if item.is_active else 'لا'
            ])

        # Sheet 2: Price Lists
        ws2 = workbook.create_sheet('قوائم الأسعار')
        columns2 = ['الرمز', 'الاسم', 'العملة', 'نشط']
        ws2.append(columns2)

        price_lists = PriceList.objects.filter(company=company)
        for pl in price_lists:
            ws2.append([
                pl.code,
                pl.name,
                pl.currency.code if pl.currency else '',
                'نعم' if pl.is_active else 'لا'
            ])

        # Sheet 3: Pricing Rules
        ws3 = workbook.create_sheet('قواعد التسعير')
        columns3 = ['الرمز', 'الاسم', 'النوع', 'القيمة', 'نشط']
        ws3.append(columns3)

        rules = PricingRule.objects.filter(company=company)[:1000]
        for rule in rules:
            ws3.append([
                rule.code,
                rule.name,
                rule.get_rule_type_display(),
                float(rule.value),
                'نعم' if rule.is_active else 'لا'
            ])

        # Generate response
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'bulk_pricing_export_{timestamp}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response

    def _export_csv(self, company, sheet):
        """Export specific sheet to CSV"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        if sheet == 'price_lists':
            columns = ['الرمز', 'الاسم', 'العملة', 'نشط']
            exporter = CSVExporter(columns)

            price_lists = PriceList.objects.filter(company=company)
            for pl in price_lists:
                exporter.add_row([
                    pl.code,
                    pl.name,
                    pl.currency.code if pl.currency else '',
                    'نعم' if pl.is_active else 'لا'
                ])

            return exporter.get_response(f'price_lists_{timestamp}')

        elif sheet == 'pricing_rules':
            columns = ['الرمز', 'الاسم', 'النوع', 'القيمة', 'نشط']
            exporter = CSVExporter(columns)

            rules = PricingRule.objects.filter(company=company)
            for rule in rules:
                exporter.add_row([
                    rule.code,
                    rule.name,
                    rule.get_rule_type_display(),
                    float(rule.value),
                    'نعم' if rule.is_active else 'لا'
                ])

            return exporter.get_response(f'pricing_rules_{timestamp}')

        else:  # price_items
            columns = ['قائمة الأسعار', 'رمز الصنف', 'اسم الصنف', 'السعر', 'نشط']
            exporter = CSVExporter(columns)

            price_items = PriceListItem.objects.filter(
                price_list__company=company
            ).select_related('price_list', 'item')[:1000]

            for item in price_items:
                exporter.add_row([
                    item.price_list.name,
                    item.item.code,
                    item.item.name,
                    float(item.price),
                    'نعم' if item.is_active else 'لا'
                ])

            return exporter.get_response(f'price_items_{timestamp}')


class ItemsExportView(LoginRequiredMixin, View):
    """
    Export Items (for price list import reference)
    """

    def get(self, request, *args, **kwargs):
        company = request.current_company
        export_format = request.GET.get('format', 'excel')

        # Get items
        items = Item.objects.filter(
            company=company,
            is_active=True
        ).select_related('category', 'unit')

        # Prepare data
        columns = [
            'الرمز',
            'الاسم',
            'الفئة',
            'الوحدة',
            'الوصف'
        ]

        rows = []
        for item in items:
            rows.append([
                item.code,
                item.name,
                item.category.name if item.category else '',
                item.unit.name if item.unit else '',
                item.description or ''
            ])

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'items_{timestamp}'

        if export_format == 'csv':
            exporter = CSVExporter(columns)
            exporter.add_rows(rows)
            return exporter.get_response(filename)
        else:
            exporter = ExcelExporter(title='الأصناف')
            exporter.add_header(columns)
            exporter.add_rows(rows)
            exporter.freeze_panes(2, 1)
            exporter.add_filter()
            return exporter.get_response(filename)
