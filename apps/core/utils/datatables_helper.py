# apps/core/utils/datatables_helper.py
"""
DataTables Helper Utility
Handles server-side processing for DataTables
"""

from typing import Dict, Any, List, Optional, Tuple
from django.db.models import Q, QuerySet, Model
from django.core.paginator import Paginator
from django.http import JsonResponse


class DataTablesServerSide:
    """
    Helper class for server-side DataTables processing
    """

    def __init__(self, request, queryset: QuerySet, columns: List[Dict[str, Any]]):
        """
        Initialize DataTables server-side processor

        Args:
            request: Django request object with DataTables parameters
            queryset: Base queryset to filter
            columns: List of column configurations
                Format: [
                    {
                        'name': 'field_name',
                        'searchable': True,
                        'orderable': True,
                        'search_fields': ['field1', 'field2']  # Optional
                    }
                ]
        """
        self.request = request
        self.queryset = queryset
        self.columns = columns

        # Extract DataTables parameters from request
        self.draw = int(request.GET.get('draw', 1))
        self.start = int(request.GET.get('start', 0))
        self.length = int(request.GET.get('length', 10))
        self.search_value = request.GET.get('search[value]', '')
        self.order_column_index = int(request.GET.get('order[0][column]', 0))
        self.order_direction = request.GET.get('order[0][dir]', 'asc')

    def get_filtered_queryset(self) -> QuerySet:
        """
        Apply search filters to queryset

        Returns:
            Filtered queryset
        """
        queryset = self.queryset

        if self.search_value:
            # Build Q object for OR search across searchable fields
            search_query = Q()

            for column in self.columns:
                if column.get('searchable', True):
                    search_fields = column.get('search_fields', [column['name']])

                    for field in search_fields:
                        # Handle related fields (e.g., 'category__name')
                        if '__' in field or field.endswith('__icontains'):
                            search_query |= Q(**{f'{field}__icontains': self.search_value})
                        else:
                            search_query |= Q(**{f'{field}__icontains': self.search_value})

            queryset = queryset.filter(search_query)

        return queryset

    def get_ordered_queryset(self, queryset: QuerySet) -> QuerySet:
        """
        Apply ordering to queryset

        Args:
            queryset: Queryset to order

        Returns:
            Ordered queryset
        """
        if 0 <= self.order_column_index < len(self.columns):
            column = self.columns[self.order_column_index]

            if column.get('orderable', True):
                order_field = column['name']

                # Add '-' prefix for descending order
                if self.order_direction == 'desc':
                    order_field = f'-{order_field}'

                queryset = queryset.order_by(order_field)

        return queryset

    def paginate_queryset(self, queryset: QuerySet) -> Tuple[List, int]:
        """
        Paginate queryset

        Args:
            queryset: Queryset to paginate

        Returns:
            Tuple of (paginated_list, total_count)
        """
        total = queryset.count()

        # Handle special case of -1 (show all)
        if self.length == -1:
            return list(queryset), total

        # Calculate page number
        page_num = (self.start // self.length) + 1

        paginator = Paginator(queryset, self.length)
        page = paginator.get_page(page_num)

        return list(page), total

    def get_response_data(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Build DataTables response

        Args:
            data: List of row data dictionaries

        Returns:
            DataTables-compatible response dictionary
        """
        # Get filtered queryset
        filtered_queryset = self.get_filtered_queryset()

        # Get ordered queryset
        ordered_queryset = self.get_ordered_queryset(filtered_queryset)

        # Get total counts
        total_records = self.queryset.count()
        filtered_records = filtered_queryset.count()

        return {
            'draw': self.draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        }

    def process(self, row_callback) -> JsonResponse:
        """
        Process DataTables request and return JSON response

        Args:
            row_callback: Function to convert model instance to row data
                Signature: row_callback(instance) -> dict

        Returns:
            JsonResponse with DataTables data
        """
        # Get filtered and ordered queryset
        filtered_queryset = self.get_filtered_queryset()
        ordered_queryset = self.get_ordered_queryset(filtered_queryset)

        # Paginate
        paginated_items, _ = self.paginate_queryset(ordered_queryset)

        # Convert to row data
        data = [row_callback(item) for item in paginated_items]

        # Build response
        response_data = self.get_response_data(data)

        return JsonResponse(response_data, safe=False)


class DataTablesColumnBuilder:
    """
    Helper class for building DataTables column configurations
    """

    @staticmethod
    def text_column(
        name: str,
        label: str,
        searchable: bool = True,
        orderable: bool = True,
        search_fields: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Create a text column configuration

        Args:
            name: Field name
            label: Column label
            searchable: Whether column is searchable
            orderable: Whether column is orderable
            search_fields: List of fields to search in

        Returns:
            Column configuration dictionary
        """
        return {
            'name': name,
            'label': label,
            'searchable': searchable,
            'orderable': orderable,
            'search_fields': search_fields or [name],
            'type': 'text'
        }

    @staticmethod
    def number_column(
        name: str,
        label: str,
        orderable: bool = True
    ) -> Dict[str, Any]:
        """
        Create a number column configuration

        Args:
            name: Field name
            label: Column label
            orderable: Whether column is orderable

        Returns:
            Column configuration dictionary
        """
        return {
            'name': name,
            'label': label,
            'searchable': False,
            'orderable': orderable,
            'type': 'number'
        }

    @staticmethod
    def date_column(
        name: str,
        label: str,
        orderable: bool = True
    ) -> Dict[str, Any]:
        """
        Create a date column configuration

        Args:
            name: Field name
            label: Column label
            orderable: Whether column is orderable

        Returns:
            Column configuration dictionary
        """
        return {
            'name': name,
            'label': label,
            'searchable': False,
            'orderable': orderable,
            'type': 'date'
        }

    @staticmethod
    def boolean_column(
        name: str,
        label: str,
        orderable: bool = True
    ) -> Dict[str, Any]:
        """
        Create a boolean column configuration

        Args:
            name: Field name
            label: Column label
            orderable: Whether column is orderable

        Returns:
            Column configuration dictionary
        """
        return {
            'name': name,
            'label': label,
            'searchable': False,
            'orderable': orderable,
            'type': 'boolean'
        }

    @staticmethod
    def actions_column(label: str = 'الإجراءات') -> Dict[str, Any]:
        """
        Create an actions column configuration

        Args:
            label: Column label

        Returns:
            Column configuration dictionary
        """
        return {
            'name': 'actions',
            'label': label,
            'searchable': False,
            'orderable': False,
            'type': 'actions'
        }


class DataTablesExporter:
    """
    Helper class for exporting DataTables data
    """

    @staticmethod
    def to_excel(queryset: QuerySet, columns: List[Dict], filename: str):
        """
        Export queryset to Excel

        Args:
            queryset: Data to export
            columns: Column configurations
            filename: Output filename

        Returns:
            HttpResponse with Excel file
        """
        try:
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, Alignment, PatternFill

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            # Write headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for idx, column in enumerate(columns, start=1):
                if column.get('type') != 'actions':
                    cell = ws.cell(row=1, column=idx)
                    cell.value = column['label']
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

            # Write data
            for row_idx, item in enumerate(queryset, start=2):
                col_idx = 1
                for column in columns:
                    if column.get('type') != 'actions':
                        field_name = column['name']
                        value = getattr(item, field_name, '')

                        # Handle related fields
                        if '__' in field_name:
                            parts = field_name.split('__')
                            value = item
                            for part in parts:
                                value = getattr(value, part, '')
                                if value is None:
                                    value = ''
                                    break

                        ws.cell(row=row_idx, column=col_idx, value=str(value))
                        col_idx += 1

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)

            return response

        except ImportError:
            from django.http import HttpResponse
            return HttpResponse("openpyxl library is required for Excel export", status=500)

    @staticmethod
    def to_csv(queryset: QuerySet, columns: List[Dict], filename: str):
        """
        Export queryset to CSV

        Args:
            queryset: Data to export
            columns: Column configurations
            filename: Output filename

        Returns:
            HttpResponse with CSV file
        """
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)

        # Write headers
        headers = [col['label'] for col in columns if col.get('type') != 'actions']
        writer.writerow(headers)

        # Write data
        for item in queryset:
            row = []
            for column in columns:
                if column.get('type') != 'actions':
                    field_name = column['name']
                    value = getattr(item, field_name, '')

                    # Handle related fields
                    if '__' in field_name:
                        parts = field_name.split('__')
                        value = item
                        for part in parts:
                            value = getattr(value, part, '')
                            if value is None:
                                value = ''
                                break

                    row.append(str(value))

            writer.writerow(row)

        return response
