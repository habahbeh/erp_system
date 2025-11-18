# apps/core/utils/uom_import_export.py
"""
Bulk Import/Export System for UoM Conversions

⭐ NEW Week 2 Day 4

Features:
- Export conversions to Excel
- Import conversions from Excel
- Validation during import
- Error reporting
- Template generation
"""

from decimal import Decimal, InvalidOperation
from typing import List, Dict, Tuple, Optional
import io

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.translation import gettext_lazy as _

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from apps.core.models import Company, UoMGroup, UnitOfMeasure, UoMConversion


class UoMConversionExporter:
    """
    Export UoM Conversions to Excel

    ⭐ NEW Week 2 Day 4

    Features:
    - Export all conversions for a company
    - Export conversions for a specific group
    - Formatted Excel with headers and styling
    - Multiple sheets (by group)
    """

    def __init__(self, company: Company):
        """
        Initialize exporter

        Args:
            company: Company instance
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        self.company = company

    def export_all_conversions(self) -> Workbook:
        """
        Export all conversions for the company to Excel

        Returns:
            Workbook: Excel workbook with all conversions

        Structure:
        - One sheet per UoM Group
        - Summary sheet
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Get all groups
        groups = UoMGroup.objects.filter(
            company=self.company,
            is_active=True
        ).order_by('name')

        if not groups.exists():
            # Create empty template
            ws = wb.create_sheet("Template")
            self._create_template_sheet(ws)
            return wb

        # Create summary sheet
        summary_ws = wb.create_sheet("ملخص - Summary", 0)
        self._create_summary_sheet(summary_ws, groups)

        # Create sheet for each group
        for group in groups:
            ws = wb.create_sheet(f"{group.name}")
            self._export_group_conversions(ws, group)

        return wb

    def export_group_conversions(self, group: UoMGroup) -> Workbook:
        """
        Export conversions for a specific group

        Args:
            group: UoMGroup instance

        Returns:
            Workbook: Excel workbook with group conversions
        """
        wb = Workbook()
        ws = wb.active
        ws.title = group.name

        self._export_group_conversions(ws, group)

        return wb

    def _create_summary_sheet(self, ws, groups):
        """Create summary sheet with statistics"""

        # Title
        ws['A1'] = 'ملخص التحويلات - Conversions Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # Company info
        ws['A3'] = 'الشركة - Company:'
        ws['B3'] = self.company.name
        ws['A3'].font = Font(bold=True)

        # Headers
        headers = ['المجموعة - Group', 'عدد الوحدات - Units', 'عدد التحويلات - Conversions',
                   'الوحدة الأساسية - Base Unit', 'نشط - Active']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        row = 6
        for group in groups:
            unit_count = group.get_unit_count()
            conv_count = group.get_all_conversions().count()
            base_unit = group.base_uom.name if group.base_uom else 'غير محدد - Not Set'

            ws.cell(row=row, column=1).value = group.name
            ws.cell(row=row, column=2).value = unit_count
            ws.cell(row=row, column=3).value = conv_count
            ws.cell(row=row, column=4).value = base_unit
            ws.cell(row=row, column=5).value = 'نعم - Yes' if group.is_active else 'لا - No'

            row += 1

        # Auto-width columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _export_group_conversions(self, ws, group: UoMGroup):
        """Export conversions for a specific group to worksheet"""

        # Title
        ws['A1'] = f'تحويلات {group.name} - {group.name} Conversions'
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')

        # Group info
        ws['A2'] = f'الرمز - Code: {group.code}'
        ws['A3'] = f'الوحدة الأساسية - Base Unit: {group.base_uom.name if group.base_uom else "غير محدد"}'

        # Headers
        headers = [
            'من وحدة - From UoM',
            'رمز الوحدة - Code',
            'معامل التحويل - Factor',
            'الصيغة - Formula',
            'نوع - Type',
            'ملاحظات - Notes'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')

        # Get conversions
        conversions = UoMConversion.objects.filter(
            company=self.company,
            from_uom__uom_group=group,
            item__isnull=True,
            variant__isnull=True,
            is_active=True
        ).select_related('from_uom').order_by('from_uom__name')

        # Data
        row = 6
        for conv in conversions:
            ws.cell(row=row, column=1).value = conv.from_uom.name
            ws.cell(row=row, column=2).value = conv.from_uom.code
            ws.cell(row=row, column=3).value = float(conv.conversion_factor)
            ws.cell(row=row, column=4).value = conv.formula_expression or ''
            ws.cell(row=row, column=5).value = 'عام - Global'
            ws.cell(row=row, column=6).value = conv.notes or ''

            row += 1

        # Auto-width columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Freeze panes
        ws.freeze_panes = 'A6'

    def _create_template_sheet(self, ws):
        """Create empty template sheet for import"""

        # Title
        ws['A1'] = 'قالب استيراد التحويلات - Conversions Import Template'
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')

        # Instructions
        ws['A2'] = 'تعليمات - Instructions:'
        ws['A2'].font = Font(bold=True)

        instructions = [
            '1. رمز المجموعة (Group Code) يجب أن يكون موجوداً مسبقاً',
            '2. رمز الوحدة (From UoM Code) يجب أن يكون موجوداً مسبقاً',
            '3. معامل التحويل (Factor) يجب أن يكون رقماً موجباً',
            '4. اتركوا العمود "نوع" فارغاً (سيكون "عام" افتراضياً)',
        ]

        for i, instruction in enumerate(instructions, 3):
            ws[f'A{i}'] = instruction

        # Headers
        headers = [
            'رمز المجموعة - Group Code*',
            'رمز الوحدة - From UoM Code*',
            'معامل التحويل - Factor*',
            'الصيغة - Formula',
            'نوع - Type',
            'ملاحظات - Notes'
        ]

        row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Example data
        example_row = row + 1
        ws.cell(row=example_row, column=1).value = 'WEIGHT'
        ws.cell(row=example_row, column=2).value = 'KG'
        ws.cell(row=example_row, column=3).value = 1000
        ws.cell(row=example_row, column=4).value = '1 كيلو = 1000 جرام'
        ws.cell(row=example_row, column=5).value = 'عام'
        ws.cell(row=example_row, column=6).value = 'مثال توضيحي'

        # Auto-width
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def save_to_file(self, wb: Workbook, filepath: str):
        """Save workbook to file"""
        wb.save(filepath)

    def save_to_bytes(self, wb: Workbook) -> bytes:
        """Save workbook to bytes (for HTTP response)"""
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class UoMConversionImporter:
    """
    Import UoM Conversions from Excel

    ⭐ NEW Week 2 Day 4

    Features:
    - Import conversions from Excel
    - Validate data before saving
    - Error reporting with line numbers
    - Rollback on error
    - Skip duplicates option
    """

    def __init__(self, company: Company):
        """
        Initialize importer

        Args:
            company: Company instance
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel import")

        self.company = company
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
        self.success_count = 0
        self.skip_count = 0

    def import_from_file(self, filepath: str, skip_duplicates: bool = True) -> Dict:
        """
        Import conversions from Excel file

        Args:
            filepath: Path to Excel file
            skip_duplicates: Skip existing conversions if True

        Returns:
            Dict with results: {
                'success': bool,
                'created': int,
                'skipped': int,
                'errors': List[Dict],
                'warnings': List[Dict]
            }
        """
        wb = load_workbook(filepath)
        return self._import_from_workbook(wb, skip_duplicates)

    def import_from_bytes(self, data: bytes, skip_duplicates: bool = True) -> Dict:
        """
        Import conversions from Excel bytes (for file upload)

        Args:
            data: Excel file bytes
            skip_duplicates: Skip existing conversions if True

        Returns:
            Dict with results
        """
        buffer = io.BytesIO(data)
        wb = load_workbook(buffer)
        return self._import_from_workbook(wb, skip_duplicates)

    def _import_from_workbook(self, wb: Workbook, skip_duplicates: bool) -> Dict:
        """Import conversions from workbook"""

        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.skip_count = 0

        # Process all sheets except summary
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ملخص - Summary', 'Template', 'قالب']:
                continue

            ws = wb[sheet_name]
            self._import_from_sheet(ws, skip_duplicates)

        return {
            'success': len(self.errors) == 0,
            'created': self.success_count,
            'skipped': self.skip_count,
            'errors': self.errors,
            'warnings': self.warnings
        }

    def _import_from_sheet(self, ws, skip_duplicates: bool):
        """Import conversions from a single sheet"""

        # Find header row (contains "From UoM" or "رمز الوحدة")
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and ('Group Code' in str(cell_value) or 'رمز المجموعة' in str(cell_value)):
                header_row = row
                break

        if header_row is None:
            self.errors.append({
                'sheet': ws.title,
                'error': 'لم يتم العثور على صف الرؤوس (Headers not found)'
            })
            return

        # Read data rows
        for row_num in range(header_row + 1, ws.max_row + 1):
            # Skip empty rows
            if not ws.cell(row=row_num, column=1).value:
                continue

            # Skip example rows
            first_cell = str(ws.cell(row=row_num, column=1).value).strip()
            if first_cell.startswith('مثال') or first_cell.startswith('Example'):
                continue

            self._import_row(ws, row_num, skip_duplicates)

    def _import_row(self, ws, row_num: int, skip_duplicates: bool):
        """Import a single row"""

        try:
            # Extract data
            group_code = str(ws.cell(row=row_num, column=1).value or '').strip().upper()
            from_uom_code = str(ws.cell(row=row_num, column=2).value or '').strip().upper()
            factor_value = ws.cell(row=row_num, column=3).value
            formula = ws.cell(row=row_num, column=4).value or ''
            notes = ws.cell(row=row_num, column=6).value or ''

            # Validate required fields
            if not group_code:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': 'رمز المجموعة مطلوب (Group Code required)'
                })
                return

            if not from_uom_code:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': 'رمز الوحدة مطلوب (From UoM Code required)'
                })
                return

            if factor_value is None:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': 'معامل التحويل مطلوب (Factor required)'
                })
                return

            # Convert factor to Decimal
            try:
                factor = Decimal(str(factor_value))
            except (InvalidOperation, ValueError):
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': f'معامل تحويل غير صحيح: {factor_value} (Invalid factor)'
                })
                return

            # Find group
            try:
                group = UoMGroup.objects.get(
                    company=self.company,
                    code=group_code,
                    is_active=True
                )
            except UoMGroup.DoesNotExist:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': f'المجموعة غير موجودة: {group_code} (Group not found)'
                })
                return

            # Find from_uom
            try:
                from_uom = UnitOfMeasure.objects.get(
                    company=self.company,
                    code=from_uom_code,
                    is_active=True
                )
            except UnitOfMeasure.DoesNotExist:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': f'الوحدة غير موجودة: {from_uom_code} (Unit not found)'
                })
                return

            # Check if from_uom belongs to group
            if from_uom.uom_group_id != group.id:
                self.errors.append({
                    'sheet': ws.title,
                    'row': row_num,
                    'error': f'الوحدة {from_uom_code} لا تنتمي للمجموعة {group_code}'
                })
                return

            # Check for duplicate
            existing = UoMConversion.objects.filter(
                company=self.company,
                from_uom=from_uom,
                item__isnull=True,
                variant__isnull=True
            ).first()

            if existing:
                if skip_duplicates:
                    self.skip_count += 1
                    self.warnings.append({
                        'sheet': ws.title,
                        'row': row_num,
                        'warning': f'تحويل موجود مسبقاً: {from_uom_code} (Skipped duplicate)'
                    })
                    return
                else:
                    # Update existing
                    existing.conversion_factor = factor
                    existing.formula_expression = formula
                    existing.notes = notes
                    existing.save()
                    self.success_count += 1
                    return

            # Create conversion
            with transaction.atomic():
                conversion = UoMConversion(
                    company=self.company,
                    from_uom=from_uom,
                    conversion_factor=factor,
                    formula_expression=formula,
                    notes=notes,
                    is_active=True
                )

                # Run validation
                conversion.full_clean()
                conversion.save()

                self.success_count += 1

        except ValidationError as e:
            self.errors.append({
                'sheet': ws.title,
                'row': row_num,
                'error': str(e)
            })
        except Exception as e:
            self.errors.append({
                'sheet': ws.title,
                'row': row_num,
                'error': f'خطأ غير متوقع: {str(e)} (Unexpected error)'
            })


def export_conversions_to_excel(company: Company, group: Optional[UoMGroup] = None) -> bytes:
    """
    Helper function to export conversions to Excel

    Args:
        company: Company instance
        group: Optional UoMGroup - export specific group only

    Returns:
        bytes: Excel file content

    Example:
        data = export_conversions_to_excel(company)
        # Save to file or return as HTTP response
    """
    exporter = UoMConversionExporter(company)

    if group:
        wb = exporter.export_group_conversions(group)
    else:
        wb = exporter.export_all_conversions()

    return exporter.save_to_bytes(wb)


def import_conversions_from_excel(
    company: Company,
    file_data: bytes,
    skip_duplicates: bool = True
) -> Dict:
    """
    Helper function to import conversions from Excel

    Args:
        company: Company instance
        file_data: Excel file bytes
        skip_duplicates: Skip existing conversions

    Returns:
        Dict with results

    Example:
        with open('conversions.xlsx', 'rb') as f:
            result = import_conversions_from_excel(company, f.read())
            print(f"Created: {result['created']}, Errors: {len(result['errors'])}")
    """
    importer = UoMConversionImporter(company)
    return importer.import_from_bytes(file_data, skip_duplicates)
