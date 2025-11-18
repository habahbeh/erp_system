# apps/core/utils/import_export.py
"""
Import/Export Utilities
Comprehensive system for importing and exporting data
Supports Excel (XLSX) and CSV formats
"""

import csv
import io
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import models, transaction
from django.core.exceptions import ValidationError


class ExcelExporter:
    """
    Export data to Excel format with formatting
    """

    def __init__(self, title: str = "تصدير البيانات"):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = title[:31]  # Excel sheet name limit
        self.current_row = 1

        # Styles
        self.header_font = Font(name='Cairo', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.data_font = Font(name='Cairo', size=11)
        self.data_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_header(self, columns: List[str], rtl: bool = True):
        """Add header row with styling"""
        for col_num, column_name in enumerate(columns, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = column_name
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Set RTL direction
        if rtl:
            self.worksheet.sheet_view.rightToLeft = True

        # Auto-adjust column widths
        for col_num in range(1, len(columns) + 1):
            self.worksheet.column_dimensions[get_column_letter(col_num)].width = 20

        self.current_row += 1

    def add_row(self, data: List[Any]):
        """Add data row"""
        for col_num, value in enumerate(data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = self._format_value(value)
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.border

        self.current_row += 1

    def add_rows(self, rows: List[List[Any]]):
        """Add multiple data rows"""
        for row in rows:
            self.add_row(row)

    def _format_value(self, value: Any) -> Any:
        """Format value for Excel"""
        if isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, bool):
            return 'نعم' if value else 'لا'
        elif value is None:
            return ''
        return str(value)

    def add_summary_row(self, label: str, values: List[Any], label_col: int = 1):
        """Add a summary row (totals, etc.)"""
        self.current_row += 1  # Empty row

        cell = self.worksheet.cell(row=self.current_row, column=label_col)
        cell.value = label
        cell.font = Font(name='Cairo', size=11, bold=True)
        cell.alignment = self.data_alignment
        cell.border = self.border

        for col_num, value in enumerate(values, label_col + 1):
            cell = self.worksheet.cell(row=self.current_row, column=col_num)
            cell.value = self._format_value(value)
            cell.font = Font(name='Cairo', size=11, bold=True)
            cell.alignment = self.data_alignment
            cell.border = self.border

        self.current_row += 1

    def freeze_panes(self, row: int = 2, col: int = 1):
        """Freeze header row"""
        self.worksheet.freeze_panes = self.worksheet.cell(row=row, column=col)

    def add_filter(self, start_row: int = 1, end_row: int = None, end_col: int = None):
        """Add auto-filter to header row"""
        if end_row is None:
            end_row = self.current_row - 1
        if end_col is None:
            end_col = self.worksheet.max_column

        self.worksheet.auto_filter.ref = f'A{start_row}:{get_column_letter(end_col)}{end_row}'

    def get_response(self, filename: str) -> HttpResponse:
        """Generate HTTP response for download"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        # Save workbook to response
        self.workbook.save(response)
        return response

    def save_to_file(self, filepath: str):
        """Save workbook to file"""
        self.workbook.save(filepath)


class CSVExporter:
    """
    Export data to CSV format
    """

    def __init__(self, columns: List[str]):
        self.columns = columns
        self.rows = []

    def add_row(self, data: List[Any]):
        """Add data row"""
        formatted_row = [self._format_value(v) for v in data]
        self.rows.append(formatted_row)

    def add_rows(self, rows: List[List[Any]]):
        """Add multiple rows"""
        for row in rows:
            self.add_row(row)

    def _format_value(self, value: Any) -> str:
        """Format value for CSV"""
        if isinstance(value, Decimal):
            return str(float(value))
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, bool):
            return 'نعم' if value else 'لا'
        elif value is None:
            return ''
        return str(value)

    def get_response(self, filename: str) -> HttpResponse:
        """Generate HTTP response for download"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        # Add BOM for Excel to recognize UTF-8
        response.write('\ufeff')

        writer = csv.writer(response)
        writer.writerow(self.columns)
        writer.writerows(self.rows)

        return response


class ExcelImporter:
    """
    Import data from Excel format with validation
    """

    def __init__(self, file, expected_columns: List[str], rtl: bool = True):
        self.file = file
        self.expected_columns = expected_columns
        self.rtl = rtl
        self.workbook = None
        self.worksheet = None
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        self.skipped_count = 0

    def validate_file(self) -> bool:
        """Validate Excel file structure"""
        try:
            self.workbook = openpyxl.load_workbook(self.file, read_only=True)
            self.worksheet = self.workbook.active

            # Validate headers
            headers = [cell.value for cell in self.worksheet[1]]
            headers = [str(h).strip() if h else '' for h in headers]

            # Remove empty trailing columns
            while headers and not headers[-1]:
                headers.pop()

            if headers != self.expected_columns:
                self.errors.append({
                    'type': 'header_mismatch',
                    'message': f'تنسيق الملف غير صحيح. الأعمدة المتوقعة: {", ".join(self.expected_columns)}'
                })
                return False

            return True

        except Exception as e:
            self.errors.append({
                'type': 'file_error',
                'message': f'خطأ في قراءة الملف: {str(e)}'
            })
            return False

    def get_rows(self) -> List[Dict[str, Any]]:
        """Get all data rows as dictionaries"""
        if not self.worksheet:
            return []

        rows = []
        headers = [cell.value for cell in self.worksheet[1]]

        for row_num, row in enumerate(self.worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            row_data = {}
            for col_num, (header, value) in enumerate(zip(headers, row), start=1):
                if header:
                    row_data[header] = value
                    row_data['_row_number'] = row_num
                    row_data['_col_number'] = col_num

            rows.append(row_data)

        return rows

    def validate_row(self, row: Dict[str, Any], validators: Dict[str, callable]) -> Tuple[bool, List[str]]:
        """Validate a single row using provided validators"""
        errors = []

        for field, validator in validators.items():
            if field in row:
                try:
                    validator(row[field])
                except ValidationError as e:
                    errors.append(f'{field}: {e.message}')
                except Exception as e:
                    errors.append(f'{field}: {str(e)}')

        return len(errors) == 0, errors

    def close(self):
        """Close workbook"""
        if self.workbook:
            self.workbook.close()


class CSVImporter:
    """
    Import data from CSV format with validation
    """

    def __init__(self, file, expected_columns: List[str]):
        self.file = file
        self.expected_columns = expected_columns
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        self.skipped_count = 0

    def validate_file(self) -> bool:
        """Validate CSV file structure"""
        try:
            # Decode file
            self.file.seek(0)
            content = self.file.read().decode('utf-8-sig')
            self.file.seek(0)

            # Read first line (headers)
            reader = csv.reader(io.StringIO(content))
            headers = next(reader)
            headers = [h.strip() for h in headers]

            if headers != self.expected_columns:
                self.errors.append({
                    'type': 'header_mismatch',
                    'message': f'تنسيق الملف غير صحيح. الأعمدة المتوقعة: {", ".join(self.expected_columns)}'
                })
                return False

            return True

        except Exception as e:
            self.errors.append({
                'type': 'file_error',
                'message': f'خطأ في قراءة الملف: {str(e)}'
            })
            return False

    def get_rows(self) -> List[Dict[str, Any]]:
        """Get all data rows as dictionaries"""
        self.file.seek(0)
        content = self.file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))

        rows = []
        for row_num, row in enumerate(reader, start=2):
            if not any(row.values()):  # Skip empty rows
                continue

            row['_row_number'] = row_num
            rows.append(row)

        return rows


class BulkImporter:
    """
    Bulk import with transaction support and progress tracking
    """

    def __init__(self, model: models.Model, user, company):
        self.model = model
        self.user = user
        self.company = company
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.warnings = []

    @transaction.atomic
    def import_rows(
        self,
        rows: List[Dict[str, Any]],
        row_processor: callable,
        batch_size: int = 100
    ) -> Dict[str, Any]:
        """
        Import rows in batches with transaction support

        Args:
            rows: List of row dictionaries
            row_processor: Function to process each row (returns model instance or None)
            batch_size: Number of rows to process in each batch

        Returns:
            Dictionary with import results
        """
        total_rows = len(rows)
        processed = 0

        for i in range(0, total_rows, batch_size):
            batch = rows[i:i + batch_size]

            for row in batch:
                try:
                    instance = row_processor(row)
                    if instance:
                        self.success_count += 1
                    else:
                        self.skipped_count += 1

                except ValidationError as e:
                    self.error_count += 1
                    self.errors.append({
                        'row': row.get('_row_number', '?'),
                        'message': str(e)
                    })

                except Exception as e:
                    self.error_count += 1
                    self.errors.append({
                        'row': row.get('_row_number', '?'),
                        'message': f'خطأ غير متوقع: {str(e)}'
                    })

                processed += 1

        return {
            'total': total_rows,
            'success': self.success_count,
            'errors': self.error_count,
            'skipped': getattr(self, 'skipped_count', 0),
            'error_details': self.errors,
            'warnings': self.warnings
        }


class DataValidator:
    """
    Common validation functions for import
    """

    @staticmethod
    def validate_required(value: Any, field_name: str = 'الحقل'):
        """Validate required field"""
        if value is None or str(value).strip() == '':
            raise ValidationError(f'{field_name} مطلوب')
        return value

    @staticmethod
    def validate_decimal(value: Any, field_name: str = 'القيمة') -> Decimal:
        """Validate and convert to Decimal"""
        if value is None or str(value).strip() == '':
            return Decimal('0')

        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            raise ValidationError(f'{field_name} يجب أن يكون رقم صحيح')

    @staticmethod
    def validate_positive_decimal(value: Any, field_name: str = 'القيمة') -> Decimal:
        """Validate positive decimal"""
        decimal_value = DataValidator.validate_decimal(value, field_name)
        if decimal_value < 0:
            raise ValidationError(f'{field_name} يجب أن يكون رقم موجب')
        return decimal_value

    @staticmethod
    def validate_boolean(value: Any) -> bool:
        """Validate and convert to boolean"""
        if isinstance(value, bool):
            return value

        str_value = str(value).strip().lower()
        if str_value in ['1', 'true', 'yes', 'نعم', 'صح']:
            return True
        elif str_value in ['0', 'false', 'no', 'لا', 'خطأ', '']:
            return False
        else:
            raise ValidationError('قيمة منطقية غير صحيحة')

    @staticmethod
    def validate_date(value: Any, field_name: str = 'التاريخ'):
        """Validate and convert to date"""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        try:
            # Try to parse string date
            from dateutil import parser
            return parser.parse(str(value)).date()
        except:
            raise ValidationError(f'{field_name} تنسيق التاريخ غير صحيح')

    @staticmethod
    def validate_choice(value: Any, choices: List[Any], field_name: str = 'الخيار'):
        """Validate value is in choices"""
        if value not in choices:
            raise ValidationError(f'{field_name} يجب أن يكون أحد: {", ".join(map(str, choices))}')
        return value

    @staticmethod
    def validate_max_length(value: Any, max_length: int, field_name: str = 'النص'):
        """Validate string max length"""
        str_value = str(value) if value else ''
        if len(str_value) > max_length:
            raise ValidationError(f'{field_name} يجب ألا يتجاوز {max_length} حرف')
        return str_value


class TemplateGenerator:
    """
    Generate import templates
    """

    @staticmethod
    def generate_excel_template(
        columns: List[str],
        sample_data: List[List[Any]] = None,
        filename: str = 'template'
    ) -> HttpResponse:
        """Generate Excel template with sample data"""
        exporter = ExcelExporter(title='قالب الاستيراد')
        exporter.add_header(columns)

        if sample_data:
            exporter.add_rows(sample_data)

        exporter.freeze_panes(2, 1)

        return exporter.get_response(filename)

    @staticmethod
    def generate_csv_template(
        columns: List[str],
        sample_data: List[List[Any]] = None,
        filename: str = 'template'
    ) -> HttpResponse:
        """Generate CSV template with sample data"""
        exporter = CSVExporter(columns)

        if sample_data:
            exporter.add_rows(sample_data)

        return exporter.get_response(filename)
